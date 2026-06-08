from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_CONFIG: dict[str, Any] = {
    "locations": {
        "RC Richmond": {"short_code": "RVA"},
        "RC Virginia Beach": {"short_code": "VB"},
    },
    "public_min_guest_count": 1,
    "public_name_aliases": {
        "Bar 1 Bar 1": "Bar",
        "Bar Server": "Bar",
        "BarPatio Bartender Patio": "Patio",
    },
    "public_exclude_name_contains": ["Banquet", "Takeout", "Server Server"],
}

METRICS = [
    ("gross_sales", "Gross Sales"),
    ("guest_count", "Guest Count"),
    ("check_average", "Check Average"),
    ("wine_sales", "Wine Sales"),
    ("rate_of_sale_by_guest_count", "Rate of Sale by Guest Count"),
    ("average_ticket_time", "Average Ticket Time"),
]


@dataclass(frozen=True)
class MetricRecord:
    source_file: str
    report_date: date
    location: str
    raw_user_name: str
    display_name: str
    is_location_total: bool
    gross_sales: float
    guest_count: float
    check_average: float
    wine_sales: float
    wine_pct: float
    rate_of_sale_by_guest_count: float
    average_ticket_time_seconds: float


def load_config(path: Path) -> dict[str, Any]:
    config = json.loads(json.dumps(DEFAULT_CONFIG))
    if path.exists():
        user_config = json.loads(path.read_text(encoding="utf-8"))
        for key, value in user_config.items():
            if isinstance(value, dict) and isinstance(config.get(key), dict):
                config[key].update(value)
            else:
                config[key] = value
    return config


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except TypeError:
        pass
    return str(value).strip() == ""


def to_float(value: Any) -> float:
    if is_blank(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).replace(",", "").replace("$", "").strip())


def ticket_time_to_seconds(value: Any) -> float:
    if is_blank(value):
        return 0.0
    if isinstance(value, time):
        return (
            value.hour * 3600
            + value.minute * 60
            + value.second
            + value.microsecond / 1_000_000
        )
    if isinstance(value, datetime):
        t = value.time()
        return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1_000_000
    if isinstance(value, (int, float)):
        number = float(value)
        return number * 86400 if 0 <= number < 1 else number

    text = str(value).strip()
    match = re.match(r"^(\d+):(\d{2}):(\d{2})(?:\.(\d+))?$", text)
    if not match:
        return 0.0
    hours, minutes, seconds, fraction = match.groups()
    micros = float(f"0.{fraction}") if fraction else 0.0
    return int(hours) * 3600 + int(minutes) * 60 + int(seconds) + micros


def parse_date_text(text: str) -> date:
    text = text.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Could not parse report date: {text}")


def parse_report_date(df: pd.DataFrame, path: Path) -> date:
    for value in df.to_numpy().ravel():
        if isinstance(value, str) and "Date(s):" in value:
            match = re.search(r"Date\(s\):\s*([^\n]+)", value)
            if match:
                date_text = match.group(1).split("-")[-1].strip()
                return parse_date_text(date_text)
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", path.name)
    if match:
        month, day, year = map(int, match.groups())
        return date(year, month, day) - timedelta(days=1)
    raise ValueError(f"Could not find report date in {path.name}")


def find_header_row(df: pd.DataFrame) -> tuple[int, int, list[int]]:
    for row_index, row in df.iterrows():
        values = ["" if is_blank(v) else str(v).strip() for v in row.tolist()]
        if "Store" in values and "Gross Sales" in values:
            location_col = values.index("Store")
            block_starts = [idx for idx, value in enumerate(values) if value == "Gross Sales"]
            if block_starts:
                return row_index, location_col, block_starts
    raise ValueError("Could not find the metric header row.")


def display_name_for(raw_name: str, config: dict[str, Any]) -> str:
    aliases = config.get("public_name_aliases", {})
    return aliases.get(raw_name, raw_name)


def parse_daily_report(path: Path, config: dict[str, Any]) -> list[MetricRecord]:
    df = pd.ExcelFile(path, engine="xlrd").parse("Report(All)", header=None)
    report_date = parse_report_date(df, path)
    header_row, location_col, block_starts = find_header_row(df)
    location_names = set(config["locations"])
    records: list[MetricRecord] = []

    for _, row in df.iloc[header_row + 1 :].iterrows():
        location_value = row.get(location_col)
        if is_blank(location_value):
            continue

        location = str(location_value).strip()
        if location not in location_names:
            continue

        raw_user_name = "" if is_blank(row.get(1)) else str(row.get(1)).strip()

        for start_col in block_starts:
            values = [row.get(start_col + offset) for offset in range(len(METRICS))]
            if all(is_blank(value) for value in values):
                continue

            gross_sales = to_float(values[0])
            guest_count = to_float(values[1])
            wine_sales = to_float(values[3])
            rate = to_float(values[4])
            ticket_seconds = ticket_time_to_seconds(values[5])

            if raw_user_name and gross_sales == 0 and guest_count == 0 and wine_sales == 0:
                continue

            check_average = gross_sales / guest_count if guest_count else 0.0
            wine_pct = wine_sales / gross_sales if gross_sales else 0.0
            records.append(
                MetricRecord(
                    source_file=path.name,
                    report_date=report_date,
                    location=location,
                    raw_user_name=raw_user_name,
                    display_name=display_name_for(raw_user_name, config),
                    is_location_total=raw_user_name == "",
                    gross_sales=gross_sales,
                    guest_count=guest_count,
                    check_average=check_average,
                    wine_sales=wine_sales,
                    wine_pct=wine_pct,
                    rate_of_sale_by_guest_count=rate,
                    average_ticket_time_seconds=ticket_seconds,
                )
            )
            break

    return records


def read_all_reports(input_dir: Path, config: dict[str, Any]) -> list[MetricRecord]:
    paths = sorted(input_dir.glob("Daily Report*.xls"))
    if not paths:
        raise FileNotFoundError(f"No daily .xls reports found in {input_dir}")

    records: list[MetricRecord] = []
    for path in paths:
        records.extend(parse_daily_report(path, config))
    return records


def week_period_for(day: date) -> tuple[date, date]:
    week_start = day - timedelta(days=day.weekday())
    return week_start, week_start + timedelta(days=6)


def selected_public_dates(
    records: Iterable[MetricRecord], week_start: str | None, week_end: str | None
) -> tuple[date, date]:
    all_dates = sorted({record.report_date for record in records})
    if not all_dates:
        raise ValueError("No report dates were found.")

    if week_start and week_end:
        return date.fromisoformat(week_start), date.fromisoformat(week_end)
    if week_start:
        start = date.fromisoformat(week_start)
        return start, start + timedelta(days=6)
    if week_end:
        end = date.fromisoformat(week_end)
        return end - timedelta(days=6), end

    end = max(all_dates)
    return end - timedelta(days=6), end


def aggregate_records(
    records: Iterable[MetricRecord], key_fields: tuple[str, ...], extra: dict[str, Any] | None = None
) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], dict[str, Any]] = {}

    for record in records:
        key = tuple(getattr(record, field) for field in key_fields)
        if key not in groups:
            groups[key] = {field: getattr(record, field) for field in key_fields}
            if extra:
                groups[key].update(extra)
            groups[key].update(
                {
                    "gross_sales": 0.0,
                    "guest_count": 0.0,
                    "wine_sales": 0.0,
                    "rate_weighted_sum": 0.0,
                    "rate_weight": 0.0,
                    "ticket_weighted_sum": 0.0,
                    "ticket_weight": 0.0,
                    "active_dates": set(),
                    "source_files": set(),
                }
            )

        group = groups[key]
        group["gross_sales"] += record.gross_sales
        group["guest_count"] += record.guest_count
        group["wine_sales"] += record.wine_sales
        if record.guest_count > 0:
            group["rate_weighted_sum"] += (
                record.rate_of_sale_by_guest_count * record.guest_count
            )
            group["rate_weight"] += record.guest_count
            group["ticket_weighted_sum"] += (
                record.average_ticket_time_seconds * record.guest_count
            )
            group["ticket_weight"] += record.guest_count
        if record.guest_count > 0 or record.gross_sales > 0:
            group["active_dates"].add(record.report_date)
        group["source_files"].add(record.source_file)

    rollups: list[dict[str, Any]] = []
    for group in groups.values():
        gross_sales = group["gross_sales"]
        guest_count = group["guest_count"]
        wine_sales = group["wine_sales"]
        rollup = dict(group)
        rollup["check_average"] = gross_sales / guest_count if guest_count else 0.0
        rollup["wine_pct"] = wine_sales / gross_sales if gross_sales else 0.0
        rollup["rate_of_sale_by_guest_count"] = (
            group["rate_weighted_sum"] / group["rate_weight"] if group["rate_weight"] else 0.0
        )
        rollup["average_ticket_time_seconds"] = (
            group["ticket_weighted_sum"] / group["ticket_weight"]
            if group["ticket_weight"]
            else 0.0
        )
        rollup["active_days"] = len(group["active_dates"])
        rollup["source_days"] = len(group["active_dates"])
        rollup["source_files"] = ", ".join(sorted(group["source_files"]))
        for helper in (
            "rate_weighted_sum",
            "rate_weight",
            "ticket_weighted_sum",
            "ticket_weight",
            "active_dates",
        ):
            rollup.pop(helper, None)
        rollups.append(rollup)

    return rollups


def public_excluded(row: dict[str, Any], config: dict[str, Any]) -> bool:
    if row["guest_count"] < float(config.get("public_min_guest_count", 1)):
        return True
    haystack = f"{row.get('raw_user_name', '')} {row.get('display_name', '')}".casefold()
    for pattern in config.get("public_exclude_name_contains", []):
        if str(pattern).casefold() in haystack:
            return True
    return False


def format_date_range(start: date, end: date) -> str:
    if start == end:
        return start.strftime("%m/%d/%Y")
    return f"{start:%m/%d/%Y} - {end:%m/%d/%Y}"


def duration_fraction(seconds: float) -> float:
    return seconds / 86400 if seconds else 0.0


def style_public_sheet(ws, last_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="E1E1E1")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 78
    widths = {"A": 18, "B": 28, "C": 13, "D": 13, "E": 15, "F": 12, "G": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for hidden_col in ("C", "D", "F"):
        ws.column_dimensions[hidden_col].hidden = True

    ws["A1"].font = Font(name="Calibri", size=12, color="68696C")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=7):
        for cell in row:
            cell.font = Font(name="Calibri", size=12, bold=cell.row == 3)
            cell.border = border
        row[2].number_format = "$#,##0.00"
        row[3].number_format = "#,##0"
        row[4].number_format = "$#,##0.00"
        row[5].number_format = "$#,##0.00"
        row[6].number_format = "0.00%"

    if last_row >= 2:
        ws.auto_filter.ref = f"A2:G{last_row}"
        ws.freeze_panes = "A3"


def apply_public_metric_highlights(ws, last_row: int) -> None:
    first_server_row = 4
    if last_row < first_server_row:
        return

    light_green = PatternFill("solid", fgColor="92D050")
    light_yellow = PatternFill("solid", fgColor="FFFF00")
    light_red = PatternFill("solid", fgColor="F4CCCC")

    rows = list(range(first_server_row, last_row + 1))
    check_values = [
        (row, ws.cell(row=row, column=5).value or 0)
        for row in rows
    ]
    wine_values = []
    for row in rows:
        gross_sales = ws.cell(row=row, column=3).value or 0
        wine_sales = ws.cell(row=row, column=6).value or 0
        wine_pct = wine_sales / gross_sales if gross_sales else 0
        wine_values.append((row, wine_pct))

    top_check_row = max(check_values, key=lambda item: item[1])[0]
    top_wine_row = max(wine_values, key=lambda item: item[1])[0]
    bottom_check_rows = [
        row for row, _ in sorted(check_values, key=lambda item: (item[1], item[0]))[:3]
    ]

    for row in bottom_check_rows:
        ws.cell(row=row, column=5).fill = light_red
    ws.cell(row=top_check_row, column=5).fill = light_green
    ws.cell(row=top_wine_row, column=7).fill = light_yellow


def write_report_definition(ws, location: str, start: date, end: date) -> None:
    bullet = "\u25cf"
    filter_text = (
        "FILTERS\n"
        f"  {bullet}  Date(s):  {format_date_range(start, end)}\n"
        f"  {bullet}  Location Type:  Store\n"
        f"  {bullet}  Locations:  {location}"
    )
    rows = [
        ["Daily Report - TM (Auto-Run)"],
        [filter_text],
        [],
        ["Column", "Definition"],
        ["Gross Sales", "KPI: GrossAmount"],
        ["Guest Count", "KPI: GuestCount"],
        ["Check Average", "KPI: GrossGuestAvg"],
        ["Wine Sales", "KPI: MenuPrice\nMenuCategory: Banquet, Wine - Glass, Wine Bottle, Wine Glass"],
        ["Rate of Sale by Guest Count", "KPI: RateOfSaleGuestCount"],
        ["Average Ticket Time", "KPI: AvgTicketTime"],
    ]
    for row in rows:
        ws.append(row)

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 76
    ws.row_dimensions[2].height = 78
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["B8"].alignment = Alignment(wrap_text=True)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor="E1E1E1")
        cell.font = Font(bold=True)


def write_public_workbook(
    location: str,
    selected_records: list[MetricRecord],
    output_dir: Path,
    config: dict[str, Any],
) -> Path:
    actual_start = min(record.report_date for record in selected_records)
    actual_end = max(record.report_date for record in selected_records)
    short_code = config["locations"][location]["short_code"]

    location_records = [record for record in selected_records if record.location == location]
    totals = [record for record in location_records if record.is_location_total]
    server_records = [record for record in location_records if not record.is_location_total]

    total_rows = aggregate_records(totals, ("location", "display_name", "raw_user_name"))
    public_rows = aggregate_records(server_records, ("location", "display_name", "raw_user_name"))
    public_rows = [row for row in public_rows if not public_excluded(row, config)]
    public_rows.sort(key=lambda row: (-row["check_average"], row["display_name"]))

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "Report(All)"
    ws.merge_cells("A1:F1")

    bullet = "\u25cf"
    ws["A1"] = (
        "FILTERS\n"
        f"  {bullet}  Date(s):  {format_date_range(actual_start, actual_end)}\n"
        f"  {bullet}  Location Type:  Store\n"
        f"  {bullet}  Locations:  {location}"
    )
    headers = ["Store", "", "Gross Sales", "Guest Count", "Check Average", "Wine Sales", "Wine %"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_index, value=header)

    rows_to_write = []
    if total_rows:
        rows_to_write.append(total_rows[0])
    rows_to_write.extend(public_rows)

    for row_index, row in enumerate(rows_to_write, start=3):
        ws.cell(row=row_index, column=1, value=location)
        ws.cell(row=row_index, column=2, value=row["display_name"])
        ws.cell(row=row_index, column=3, value=row["gross_sales"])
        ws.cell(row=row_index, column=4, value=row["guest_count"])
        ws.cell(row=row_index, column=5, value=row["check_average"])
        ws.cell(row=row_index, column=6, value=row["wine_sales"])
        ws.cell(row=row_index, column=7, value=f"=F{row_index}/C{row_index}")

    last_row = max(2, len(rows_to_write) + 2)
    style_public_sheet(ws, last_row)
    apply_public_metric_highlights(ws, last_row)
    write_report_definition(wb.create_sheet("Report Definition"), location, actual_start, actual_end)

    output_path = output_dir / f"Check_Wine_{short_code}{actual_end:%m%d%y}.xlsx"
    wb.save(output_path)
    return output_path


def write_table_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    table_name: str,
    widths: dict[str, float] | None = None,
) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(size=15, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="7A1E1E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_index, value=header)
        cell.fill = PatternFill("solid", fgColor="E1E1E1")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=4):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            header = headers[col_index - 1]
            if "Date" in header or header in {"Week Start", "Week End", "Latest Week End"}:
                cell.number_format = "m/d/yyyy"
            elif "Wine %" in header:
                cell.number_format = "0.00%"
            elif "Rate of Sale" in header:
                cell.number_format = "0.00"
            elif "Ticket Time" in header:
                cell.number_format = "[h]:mm:ss"
            elif header in {"Gross Sales", "Wine Sales", "Total Gross Sales", "Total Wine Sales", "Check Average", "Latest Check Average", "Weighted Check Average"}:
                cell.number_format = "$#,##0.00"
            elif "Guest" in header or "Days" in header or "Weeks" in header:
                cell.number_format = "#,##0"

    if rows:
        ref = f"A3:{get_column_letter(len(headers))}{len(rows) + 3}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    default_widths = {
        "A": 14,
        "B": 14,
        "C": 20,
        "D": 24,
        "E": 24,
        "F": 18,
        "G": 12,
        "H": 16,
        "I": 14,
        "J": 12,
        "K": 20,
        "L": 18,
        "M": 12,
        "N": 12,
        "O": 18,
        "P": 18,
    }
    default_widths.update(widths or {})
    for col, width in default_widths.items():
        ws.column_dimensions[col].width = width


def weekly_rollups(records: list[MetricRecord]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    grouped: dict[tuple[date, date], list[MetricRecord]] = defaultdict(list)
    for record in records:
        grouped[week_period_for(record.report_date)].append(record)

    server_rows: list[dict[str, Any]] = []
    location_rows: list[dict[str, Any]] = []
    for (week_start, week_end), period_records in sorted(grouped.items()):
        servers = [record for record in period_records if not record.is_location_total]
        locations = [record for record in period_records if record.is_location_total]
        server_rows.extend(
            aggregate_records(
                servers,
                ("location", "raw_user_name", "display_name"),
                {"week_start": week_start, "week_end": week_end},
            )
        )
        location_rows.extend(
            aggregate_records(
                locations,
                ("location",),
                {"week_start": week_start, "week_end": week_end},
            )
        )
    return server_rows, location_rows


def trend_summary_rows(weekly_server_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in weekly_server_rows:
        groups[(row["location"], row["raw_user_name"], row["display_name"])].append(row)

    summary: list[dict[str, Any]] = []
    for (location, raw_name, display_name), rows in groups.items():
        rows.sort(key=lambda row: row["week_end"])
        total_gross = sum(row["gross_sales"] for row in rows)
        total_guests = sum(row["guest_count"] for row in rows)
        total_wine = sum(row["wine_sales"] for row in rows)
        rate_weighted = sum(row["rate_of_sale_by_guest_count"] * row["guest_count"] for row in rows)
        ticket_weighted = sum(row["average_ticket_time_seconds"] * row["guest_count"] for row in rows)
        latest = rows[-1]
        summary.append(
            {
                "location": location,
                "raw_user_name": raw_name,
                "display_name": display_name,
                "weeks_tracked": len(rows),
                "total_gross_sales": total_gross,
                "total_guest_count": total_guests,
                "weighted_check_average": total_gross / total_guests if total_guests else 0.0,
                "total_wine_sales": total_wine,
                "overall_wine_pct": total_wine / total_gross if total_gross else 0.0,
                "weighted_rate": rate_weighted / total_guests if total_guests else 0.0,
                "weighted_ticket_time_seconds": ticket_weighted / total_guests if total_guests else 0.0,
                "latest_week_end": latest["week_end"],
                "latest_check_average": latest["check_average"],
                "latest_wine_pct": latest["wine_pct"],
                "latest_rate": latest["rate_of_sale_by_guest_count"],
                "latest_ticket_time_seconds": latest["average_ticket_time_seconds"],
            }
        )

    summary.sort(key=lambda row: (row["location"], -row["weighted_check_average"], row["display_name"]))
    return summary


def write_master_workbook(
    records: list[MetricRecord],
    output_path: Path,
    config: dict[str, Any],
    source_dir: Path,
    public_start: date,
    public_end: date,
) -> Path:
    weekly_server_rows, weekly_location_rows = weekly_rollups(records)
    trend_rows = trend_summary_rows(weekly_server_rows)

    wb = Workbook()
    wb.remove(wb.active)

    server_headers = [
        "Week Start",
        "Week End",
        "Location",
        "Raw Server",
        "Display Name",
        "Gross Sales",
        "Guest Count",
        "Check Average",
        "Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
        "Active Days",
        "Source Days",
    ]
    server_data = [
        [
            row["week_start"],
            row["week_end"],
            row["location"],
            row["raw_user_name"],
            row["display_name"],
            row["gross_sales"],
            row["guest_count"],
            row["check_average"],
            row["wine_sales"],
            row["wine_pct"],
            row["rate_of_sale_by_guest_count"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["active_days"],
            row["source_days"],
        ]
        for row in sorted(
            weekly_server_rows,
            key=lambda item: (item["week_end"], item["location"], item["display_name"]),
        )
    ]
    write_table_sheet(wb, "Weekly Server Metrics", server_headers, server_data, "WeeklyServerMetrics")

    location_headers = [
        "Week Start",
        "Week End",
        "Location",
        "Gross Sales",
        "Guest Count",
        "Check Average",
        "Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
        "Active Days",
        "Source Days",
    ]
    location_data = [
        [
            row["week_start"],
            row["week_end"],
            row["location"],
            row["gross_sales"],
            row["guest_count"],
            row["check_average"],
            row["wine_sales"],
            row["wine_pct"],
            row["rate_of_sale_by_guest_count"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["active_days"],
            row["source_days"],
        ]
        for row in sorted(weekly_location_rows, key=lambda item: (item["week_end"], item["location"]))
    ]
    write_table_sheet(wb, "Weekly Location Metrics", location_headers, location_data, "WeeklyLocationMetrics")

    daily_server_headers = [
        "Report Date",
        "Location",
        "Raw Server",
        "Display Name",
        "Source File",
        "Gross Sales",
        "Guest Count",
        "Check Average",
        "Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
    ]
    daily_server_data = [
        [
            record.report_date,
            record.location,
            record.raw_user_name,
            record.display_name,
            record.source_file,
            record.gross_sales,
            record.guest_count,
            record.check_average,
            record.wine_sales,
            record.wine_pct,
            record.rate_of_sale_by_guest_count,
            duration_fraction(record.average_ticket_time_seconds),
        ]
        for record in sorted(
            (item for item in records if not item.is_location_total),
            key=lambda item: (item.report_date, item.location, item.display_name),
        )
    ]
    write_table_sheet(wb, "Daily Server Detail", daily_server_headers, daily_server_data, "DailyServerDetail")

    daily_location_data = [
        [
            record.report_date,
            record.location,
            record.source_file,
            record.gross_sales,
            record.guest_count,
            record.check_average,
            record.wine_sales,
            record.wine_pct,
            record.rate_of_sale_by_guest_count,
            duration_fraction(record.average_ticket_time_seconds),
        ]
        for record in sorted(
            (item for item in records if item.is_location_total),
            key=lambda item: (item.report_date, item.location),
        )
    ]
    write_table_sheet(
        wb,
        "Daily Location Detail",
        [
            "Report Date",
            "Location",
            "Source File",
            "Gross Sales",
            "Guest Count",
            "Check Average",
            "Wine Sales",
            "Wine %",
            "Rate of Sale by Guest Count",
            "Average Ticket Time",
        ],
        daily_location_data,
        "DailyLocationDetail",
    )

    trend_headers = [
        "Location",
        "Raw Server",
        "Display Name",
        "Weeks Tracked",
        "Total Gross Sales",
        "Total Guest Count",
        "Weighted Check Average",
        "Total Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
        "Latest Week End",
        "Latest Check Average",
        "Latest Wine %",
        "Latest Rate of Sale by Guest Count",
        "Latest Ticket Time",
    ]
    trend_data = [
        [
            row["location"],
            row["raw_user_name"],
            row["display_name"],
            row["weeks_tracked"],
            row["total_gross_sales"],
            row["total_guest_count"],
            row["weighted_check_average"],
            row["total_wine_sales"],
            row["overall_wine_pct"],
            row["weighted_rate"],
            duration_fraction(row["weighted_ticket_time_seconds"]),
            row["latest_week_end"],
            row["latest_check_average"],
            row["latest_wine_pct"],
            row["latest_rate"],
            duration_fraction(row["latest_ticket_time_seconds"]),
        ]
        for row in trend_rows
    ]
    write_table_sheet(
        wb,
        "Server Trend Summary",
        trend_headers,
        trend_data,
        "ServerTrendSummary",
        widths={"A": 22, "B": 28, "C": 28, "D": 14, "J": 20, "K": 18, "O": 24, "P": 18},
    )

    notes = wb.create_sheet("Run Notes", 0)
    notes.sheet_view.showGridLines = False
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 90
    notes["A1"] = "Red Onion Server Master"
    notes["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="7A1E1E")
    notes.merge_cells("A1:B1")
    note_rows = [
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source Folder", str(source_dir)),
        ("Raw Reports Read", len({record.source_file for record in records})),
        ("Date Coverage", format_date_range(min(r.report_date for r in records), max(r.report_date for r in records))),
        ("Public Snapshot Dates", format_date_range(public_start, public_end)),
        ("Public Exclude Patterns", ", ".join(config.get("public_exclude_name_contains", []))),
        ("Metric Rule", "Check average and wine percent are recalculated from rolled-up sales, guests, and wine sales."),
        ("Metric Rule", "Rate of sale and ticket time are guest-weighted averages."),
    ]
    for row_index, (label, value) in enumerate(note_rows, start=3):
        notes.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        notes.cell(row=row_index, column=2, value=value)
        notes.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def run(args: argparse.Namespace) -> list[Path]:
    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    config_path = Path(args.config).resolve()
    config = load_config(config_path)
    records = read_all_reports(input_dir, config)
    public_start, public_end = selected_public_dates(records, args.week_start, args.week_end)
    selected_records = [
        record for record in records if public_start <= record.report_date <= public_end
    ]
    if not selected_records:
        raise ValueError(f"No records found between {public_start} and {public_end}.")
    actual_public_start = min(record.report_date for record in selected_records)
    actual_public_end = max(record.report_date for record in selected_records)

    generated: list[Path] = []
    for location in config["locations"]:
        location_records = [record for record in selected_records if record.location == location]
        if not location_records:
            continue
        generated.append(write_public_workbook(location, selected_records, output_dir, config))

    master_path = output_dir / "Red_Onion_Server_Master.xlsx"
    generated.append(
        write_master_workbook(
            records,
            master_path,
            config,
            input_dir,
            actual_public_start,
            actual_public_end,
        )
    )
    return generated


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build Red Onion weekly metric workbooks.")
    parser.add_argument("--input-dir", default=".", help="Folder containing raw daily .xls reports.")
    parser.add_argument("--output-dir", default="outputs", help="Folder for generated workbooks.")
    parser.add_argument("--config", default="red_onion_config.json", help="Path to config JSON.")
    parser.add_argument("--week-start", help="Optional public snapshot start date, YYYY-MM-DD.")
    parser.add_argument("--week-end", help="Optional public snapshot end date, YYYY-MM-DD.")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    generated = run(args)
    print("Generated:")
    for path in generated:
        print(f"  {path}")


if __name__ == "__main__":
    main()
