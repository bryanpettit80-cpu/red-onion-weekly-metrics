from __future__ import annotations

import json
from argparse import Namespace
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
import pytest

import red_onion_weekly_metrics as metrics
from red_onion_integrity import read_json_manifest


def reconciled_day(report_date: date, source_file: str) -> list[metrics.MetricRecord]:
    records: list[metrics.MetricRecord] = []
    for location in metrics.DEFAULT_CONFIG["locations"]:
        values = {
            "source_file": source_file,
            "report_date": report_date,
            "location": location,
            "gross_sales": 100.0,
            "guest_count": 10.0,
            "check_average": 10.0,
            "wine_sales": 5.0,
            "wine_pct": 0.05,
            "rate_of_sale_by_guest_count": 0.20,
            "average_ticket_time_seconds": 4800.0,
            "source_format": ".xls",
            "parser_engine": "synthetic-test",
            "report_date_source": "embedded-filter",
        }
        records.append(
            metrics.MetricRecord(
                **values,
                raw_user_name="",
                display_name=f"{location} Total",
                is_location_total=True,
            )
        )
        records.append(
            metrics.MetricRecord(
                **values,
                raw_user_name="entity-01",
                display_name="entity-01",
                is_location_total=False,
            )
        )
    return records


def rebuild_args(tmp_path: Path) -> Namespace:
    return Namespace(
        input_dir=str(tmp_path / "01 Daily Reports - Drop Here"),
        output_dir=str(tmp_path / "02 Finished Reports"),
        archive_dir=str(tmp_path / "03 Archive"),
        config=str(tmp_path / "missing-config.json"),
        integrity_anchor_dir=str(
            tmp_path.parent / f"trusted-anchor-{tmp_path.name}"
        ),
        week_start=None,
        week_end=None,
        migrate_history_from=[],
        migrate_history_only=False,
        rebuild_from_history=False,
        initialize_integrity_baseline=False,
    )


def test_history_rebuild_ignores_active_intake_and_commits_one_manifest(
    tmp_path: Path,
    monkeypatch,
) -> None:
    args = rebuild_args(tmp_path)
    input_dir = Path(args.input_dir)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    input_dir.mkdir(parents=True)
    poison = input_dir / "Daily Report poison.xls"
    poison.write_bytes(b"must-not-be-read-or-moved")

    week_end = date(2026, 7, 19)
    raw_week = (
        archive_dir
        / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
        / f"week-ending-{week_end.isoformat()}"
    )
    raw_week.mkdir(parents=True)
    report_records: dict[str, list[metrics.MetricRecord]] = {}
    for offset in range(metrics.OPERATING_WEEK_DAYS):
        report_date = week_end - timedelta(
            days=metrics.OPERATING_WEEK_DAYS - 1 - offset
        )
        filename = f"Daily Report {report_date.isoformat()}.xls"
        (raw_week / filename).write_bytes(f"history-{report_date}".encode())
        report_records[filename] = reconciled_day(report_date, filename)
    partial_date = date(2026, 7, 21)
    partial_week = (
        archive_dir
        / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
        / "week-ending-2026-07-26"
    )
    partial_week.mkdir(parents=True)
    partial_filename = f"Daily Report {partial_date.isoformat()}.xls"
    (partial_week / partial_filename).write_bytes(b"later-partial-history")
    report_records[partial_filename] = reconciled_day(
        partial_date, partial_filename
    )

    def fake_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        if path.name == poison.name:
            raise AssertionError("history rebuild inspected the active drop folder")
        return report_records[path.name]

    def fake_public(
        location,
        records,
        output_path,
        config,
        start,
        end,
    ):
        result = output_path / f"public-{location}.xlsx"
        result.write_bytes(b"validated-public-placeholder")
        return result

    master_record_dates: list[date] = []
    real_master_writer = metrics.write_master_workbook

    def capture_master(records, *writer_args, **writer_kwargs):
        master_record_dates.extend(record.report_date for record in records)
        return real_master_writer(records, *writer_args, **writer_kwargs)

    monkeypatch.setattr(metrics, "parse_daily_report", fake_parse)
    monkeypatch.setattr(metrics, "write_public_workbook", fake_public)
    monkeypatch.setattr(metrics, "write_master_workbook", capture_master)

    baseline_args = Namespace(**vars(args))
    baseline_args.initialize_integrity_baseline = True
    metrics.run(baseline_args)

    args.rebuild_from_history = True
    generated = metrics.run(args)

    assert poison.read_bytes() == b"must-not-be-read-or-moved"
    assert output_dir / "Red_Onion_Server_Master.xlsx" in generated
    workbook = load_workbook(output_dir / "Red_Onion_Server_Master.xlsx")
    try:
        assert "Recent Movement Signals" in workbook.sheetnames
        assert "Rising & Falling Stars" not in workbook.sheetnames
    finally:
        workbook.close()

    manifest_path = metrics.latest_integrity_manifest_path(archive_dir)
    assert manifest_path is not None
    payload = read_json_manifest(manifest_path)
    assert payload["kind"] == "history-rebuild"
    assert payload["details"]["rebuild_from_history"] is True
    assert payload["details"]["active_input_inventory"] == []
    assert payload["details"]["archived_destinations"] == []
    assert payload["details"]["history_week_end"] == week_end.isoformat()
    assert max(master_record_dates) == week_end
    assert partial_date not in master_record_dates
    attempt_payloads = [
        json.loads(path.read_text(encoding="utf-8"))
        for path in (
            archive_dir / metrics.RUN_ATTEMPT_FOLDER
        ).glob("*.json")
    ]
    rebuild_attempt = next(
        item for item in attempt_payloads if item["operation"] == "history-rebuild"
    )
    assert rebuild_attempt["readiness"]["workbook"] == "Ready"
    assert rebuild_attempt["readiness"]["distribution"] == "Ready"
    assert (
        rebuild_attempt["readiness"]["recovery"]
        == "ExternalCheckRequired"
    )
    status = (output_dir / metrics.LAST_RUN_STATUS_FILE).read_text(encoding="utf-8")
    assert "Workbook: Ready" in status
    assert "Local publication: Ready" in status


def test_history_rebuild_requires_history_for_every_configured_location(
    tmp_path: Path,
    monkeypatch,
) -> None:
    args = rebuild_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    week_end = date(2026, 7, 19)
    raw_week = (
        archive_dir
        / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
        / f"week-ending-{week_end.isoformat()}"
    )
    raw_week.mkdir(parents=True)
    report_records: dict[str, list[metrics.MetricRecord]] = {}
    for offset in range(metrics.OPERATING_WEEK_DAYS):
        report_date = week_end - timedelta(
            days=metrics.OPERATING_WEEK_DAYS - 1 - offset
        )
        filename = f"Daily Report {report_date.isoformat()}.xls"
        (raw_week / filename).write_bytes(f"history-{report_date}".encode())
        report_records[filename] = [
            record
            for record in reconciled_day(report_date, filename)
            if record.location == "RC Richmond"
        ]

    monkeypatch.setattr(
        metrics,
        "parse_daily_report",
        lambda path, config: report_records[path.name],
    )
    baseline_args = Namespace(**vars(args))
    baseline_args.initialize_integrity_baseline = True
    metrics.run(baseline_args)
    args.rebuild_from_history = True

    with pytest.raises(ValueError, match="Missing all history for: RC Virginia Beach"):
        metrics.run(args)

    assert not Path(args.output_dir).exists()
