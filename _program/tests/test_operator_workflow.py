from __future__ import annotations

from argparse import Namespace
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

import red_onion_weekly_metrics as metrics


@pytest.mark.parametrize(
    ("hidden_columns", "expected_bounds"),
    [
        (frozenset(), (1, 12)),
        (frozenset({1, 2}), (3, 14)),
        (frozenset({12, 13}), (1, 14)),
    ],
)
def test_management_menu_preserves_column_widths_and_visibility(
    hidden_columns: frozenset[int],
    expected_bounds: tuple[int, int],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data Quality"
    for index, width in enumerate((11, 96, 13, 22, 48, 69, 12, 14, 18, 50, 9, 10, 7, 16), start=1):
        dimension = worksheet.column_dimensions[
            metrics.get_column_letter(index)
        ]
        dimension.width = width
        dimension.hidden = index in hidden_columns
    before = {
        column: (
            worksheet.column_dimensions[column].width,
            worksheet.column_dimensions[column].hidden,
        )
        for column in (
            metrics.get_column_letter(index)
            for index in range(1, 15)
        )
    }

    metrics.add_management_navigation(worksheet)

    after = {
        column: (
            worksheet.column_dimensions[column].width,
            worksheet.column_dimensions[column].hidden,
        )
        for column in before
    }
    assert after == before
    assert metrics.management_menu_bounds(worksheet) == expected_bounds


def make_record(
    day: date,
    location: str = "RC Richmond",
    raw_user_name: str = "Server One",
    display_name: str | None = None,
    is_location_total: bool = False,
    gross_sales: float = 100.0,
    guest_count: float = 10.0,
    wine_sales: float = 20.0,
    rate: float = 1.0,
    ticket_seconds: float = 120.0,
    source_file: str | None = None,
) -> metrics.MetricRecord:
    if is_location_total and raw_user_name == "Server One":
        raw_user_name = ""
    if display_name is None:
        display_name = raw_user_name
    return metrics.MetricRecord(
        source_file=source_file or f"Daily Report {day.isoformat()}.xls",
        report_date=day,
        location=location,
        raw_user_name=raw_user_name,
        display_name=display_name,
        is_location_total=is_location_total,
        gross_sales=gross_sales,
        guest_count=guest_count,
        check_average=gross_sales / guest_count if guest_count else 0.0,
        wine_sales=wine_sales,
        wine_pct=wine_sales / gross_sales if gross_sales else 0.0,
        rate_of_sale_by_guest_count=rate,
        average_ticket_time_seconds=ticket_seconds,
    )


def weekly_row(
    week_end: date,
    *,
    location: str = "RC Richmond",
    display_name: str = "Server One",
    gross_sales: float = 1000.0,
    guest_count: float = 50.0,
    wine_sales: float = 100.0,
    rate: float = 2.0,
    ticket_seconds: float = 1200.0,
    active_days: int = 3,
) -> dict:
    return {
        "week_start": week_end - timedelta(days=metrics.OPERATING_WEEK_DAYS - 1),
        "week_end": week_end,
        "location": location,
        "raw_user_name": display_name,
        "display_name": display_name,
        "gross_sales": gross_sales,
        "guest_count": guest_count,
        "check_average": gross_sales / guest_count if guest_count else 0.0,
        "wine_sales": wine_sales,
        "wine_pct": wine_sales / gross_sales if gross_sales else 0.0,
        "rate_of_sale_by_guest_count": rate,
        "average_ticket_time_seconds": ticket_seconds,
        "rate_available": True,
        "ticket_time_available": True,
        "check_count": guest_count,
        "check_count_available": True,
        "active_days": active_days,
        "source_days": active_days,
        "source_files": "synthetic.xls",
    }


def ranked_row(week_end: date, *, ranks: tuple[int, int, int, int], **kwargs) -> dict:
    row = weekly_row(week_end, **kwargs)
    row.update(
        {
            "check_average_rank": ranks[0],
            "wine_pct_rank": ranks[1],
            "rate_rank": ranks[2],
            "ticket_time_rank": ranks[3],
        }
    )
    return row


def full_week_records(
    week_end: date,
    *,
    location: str,
    server: str = "Server One",
    weekly_gross: float = 1200.0,
    weekly_guests: float = 60.0,
    weekly_wine: float = 120.0,
    rate: float = 0.20,
    ticket_seconds: float = 80 * 60,
) -> list[metrics.MetricRecord]:
    records: list[metrics.MetricRecord] = []
    for offset in range(metrics.OPERATING_WEEK_DAYS):
        day = week_end - timedelta(days=metrics.OPERATING_WEEK_DAYS - 1 - offset)
        source_file = f"{location}-{day.isoformat()}.xls"
        records.append(
            make_record(
                day,
                location,
                is_location_total=True,
                gross_sales=weekly_gross / metrics.OPERATING_WEEK_DAYS,
                guest_count=weekly_guests / metrics.OPERATING_WEEK_DAYS,
                wine_sales=weekly_wine / metrics.OPERATING_WEEK_DAYS,
                rate=rate,
                ticket_seconds=ticket_seconds,
                source_file=source_file,
            )
        )
        records.append(
            make_record(
                day,
                location,
                raw_user_name=server,
                gross_sales=weekly_gross / metrics.OPERATING_WEEK_DAYS,
                guest_count=weekly_guests / metrics.OPERATING_WEEK_DAYS,
                wine_sales=weekly_wine / metrics.OPERATING_WEEK_DAYS,
                rate=rate,
                ticket_seconds=ticket_seconds,
                source_file=source_file,
            )
        )
    return records


def args_for(tmp_path: Path) -> Namespace:
    return Namespace(
        input_dir=str(tmp_path / "Daily Reports"),
        output_dir=str(tmp_path / "Output"),
        archive_dir=str(tmp_path / "Archive - Old Files"),
        config=str(tmp_path / "missing-config.json"),
        week_start=None,
        week_end=None,
    )


def initialize_integrity(args: Namespace) -> None:
    baseline_args = Namespace(**vars(args), initialize_integrity_baseline=True)
    metrics.run(baseline_args)


def test_zero_active_files_stop_with_clear_message(tmp_path: Path) -> None:
    args = args_for(tmp_path)
    Path(args.input_dir).mkdir(parents=True)

    with pytest.raises(FileNotFoundError, match=r"No active daily reports \(.xls or .xlsx\)"):
        metrics.run(args)


def test_no_data_export_stops_with_file_specific_message(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    args = args_for(tmp_path)
    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True)
    source = input_dir / "Daily Report - TM - 07-12-2026.xls"
    source.write_text("no data", encoding="utf-8")

    class NoDataWorkbook:
        sheet_names = ["No Data Available"]

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc_value, traceback):
            return False

    monkeypatch.setattr(metrics.pd, "ExcelFile", lambda *args, **kwargs: NoDataWorkbook())
    initialize_integrity(args)

    with pytest.raises(ValueError) as exc_info:
        metrics.run(args)

    message = str(exc_info.value)
    assert source.name in message
    assert "No Data Available" in message
    assert "No workbooks were created and no source files were moved" in message
    assert source.exists()


def test_mixed_active_weeks_stop_with_exact_file_names(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    args = args_for(tmp_path)
    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True)
    first = input_dir / "Daily Report - TM - 06-07-2026.xls"
    second = input_dir / "Daily Report - TM - 06-14-2026.xls"
    first.write_text("first", encoding="utf-8")
    second.write_text("second", encoding="utf-8")

    def fake_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        if path.name == first.name:
            return [make_record(date(2026, 6, 7))]
        return [make_record(date(2026, 6, 14))]

    monkeypatch.setattr(metrics, "parse_daily_report", fake_parse)
    initialize_integrity(args)

    with pytest.raises(ValueError) as exc_info:
        metrics.run(args)

    message = str(exc_info.value)
    assert "more than one Tuesday-Sunday operating week" in message
    assert first.name in message
    assert second.name in message
    assert first.exists()
    assert second.exists()


def test_successful_run_archives_active_files_and_keeps_master_history(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    args = args_for(tmp_path)
    input_dir = Path(args.input_dir)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    input_dir.mkdir(parents=True)
    archived_source = (
        archive_dir
        / "processed-daily-reports"
        / "week-ending-2026-06-07"
        / "Daily Report archived.xls"
    )
    archived_source.parent.mkdir(parents=True)
    active_source = input_dir / "Daily Report active.xls"
    archived_source.write_text("archived", encoding="utf-8")
    active_source.write_text("active", encoding="utf-8")
    master_records: list[metrics.MetricRecord] = []
    real_master_writer = metrics.write_master_workbook

    def fake_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        if path.name == archived_source.name:
            day = date(2026, 6, 7)
            location = "RC Virginia Beach"
        else:
            day = date(2026, 6, 14)
            location = "RC Richmond"
        return [
            make_record(day, location, is_location_total=True),
            make_record(day, location),
        ]

    def fake_public(location, records, output_path, config, start, end):
        output_file = output_path / f"public-{location}.xlsx"
        output_file.parent.mkdir(parents=True, exist_ok=True)
        output_file.write_text("public", encoding="utf-8")
        return output_file

    def fake_master(records, output_path, config, input_path, start, end):
        master_records.extend(records)
        return real_master_writer(
            records, output_path, config, input_path, start, end
        )

    monkeypatch.setattr(metrics, "parse_daily_report", fake_parse)
    monkeypatch.setattr(metrics, "write_public_workbook", fake_public)
    monkeypatch.setattr(metrics, "write_master_workbook", fake_master)
    initialize_integrity(args)

    generated = metrics.run(args)

    assert not active_source.exists()
    assert (
        archive_dir
        / "processed-daily-reports"
        / "week-ending-2026-06-14"
        / active_source.name
    ).exists()
    assert {record.report_date for record in master_records} == {
        date(2026, 6, 7),
        date(2026, 6, 14),
    }
    assert output_dir / "Red_Onion_Server_Master.xlsx" in generated


def test_failed_run_leaves_active_files_in_place(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    args = args_for(tmp_path)
    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True)
    active_source = input_dir / "Daily Report active.xls"
    active_source.write_text("active", encoding="utf-8")

    def fake_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        day = date(2026, 6, 14)
        return [
            make_record(day, is_location_total=True),
            make_record(day),
        ]

    def fake_public(location, records, output_path, config, start, end):
        output_path.mkdir(parents=True, exist_ok=True)
        output_file = output_path / f"public-{location}.xlsx"
        output_file.write_text("public", encoding="utf-8")
        return output_file

    def failing_master(records, output_path, config, input_path, start, end):
        raise RuntimeError("workbook failed")

    monkeypatch.setattr(metrics, "parse_daily_report", fake_parse)
    monkeypatch.setattr(metrics, "write_public_workbook", fake_public)
    monkeypatch.setattr(metrics, "write_master_workbook", failing_master)
    initialize_integrity(args)

    with pytest.raises(RuntimeError, match="workbook failed"):
        metrics.run(args)

    assert active_source.exists()


def test_duplicate_archive_handling_preserves_different_same_name_files(tmp_path: Path) -> None:
    source_dir = tmp_path / "Daily Reports"
    archive_root = tmp_path / "Archive - Old Files"
    source_dir.mkdir()
    source = source_dir / "Daily Report active.xls"
    source.write_text("new content", encoding="utf-8")
    existing = (
        archive_root
        / "processed-daily-reports"
        / "week-ending-2026-06-14"
        / source.name
    )
    existing.parent.mkdir(parents=True)
    existing.write_text("old content", encoding="utf-8")

    archived = metrics.archive_processed_files([source], archive_root, date(2026, 6, 14))

    assert not source.exists()
    assert existing.read_text(encoding="utf-8") == "old content"
    assert archived == [existing.with_name("Daily Report active (1).xls")]
    assert archived[0].read_text(encoding="utf-8") == "new content"


def test_duplicate_archive_handling_removes_identical_active_copy(tmp_path: Path) -> None:
    source_dir = tmp_path / "Daily Reports"
    archive_root = tmp_path / "Archive - Old Files"
    source_dir.mkdir()
    source = source_dir / "Daily Report active.xls"
    source.write_text("same content", encoding="utf-8")
    existing = (
        archive_root
        / "processed-daily-reports"
        / "week-ending-2026-06-14"
        / source.name
    )
    existing.parent.mkdir(parents=True)
    existing.write_text("same content", encoding="utf-8")

    archived = metrics.archive_processed_files([source], archive_root, date(2026, 6, 14))

    assert not source.exists()
    assert archived == [existing]
    assert existing.read_text(encoding="utf-8") == "same content"


def test_parser_defaults_point_to_operator_root_folders() -> None:
    args = metrics.build_parser().parse_args([])

    assert Path(args.input_dir).name == "01 Daily Reports - Drop Here"
    assert Path(args.output_dir).name == "02 Finished Reports"
    assert Path(args.archive_dir).name == "03 Archive"
    assert Path(args.input_dir).parent == metrics.PROJECT_ROOT
    assert Path(args.config).parent == metrics.PROGRAM_DIR


def test_launchers_route_to_named_operator_workspace() -> None:
    root_launcher = (metrics.REPOSITORY_ROOT / "Run Weekly Snapshot.cmd").read_text(
        encoding="utf-8"
    )
    powershell_runner = (metrics.PROGRAM_DIR / "Run-WeeklySnapshot.ps1").read_text(
        encoding="utf-8"
    )

    assert "Red Onion Weekly Metrics Automation" in root_launcher
    assert "-OperationsRoot" in root_launcher
    assert "01 Daily Reports - Drop Here" in powershell_runner
    assert "02 Finished Reports" in powershell_runner
    assert "03 Archive" in powershell_runner
    assert "$ReportExitCode = $LASTEXITCODE" in powershell_runner
    assert "exit $ReportExitCode" in powershell_runner


def test_star_scoring_uses_only_person_action_metrics(tmp_path: Path) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    prior_week = date(2026, 6, 7)
    latest_week = date(2026, 6, 14)
    ranked_rows = [
        ranked_row(
            prior_week,
            display_name="Alex Rising",
            gross_sales=1000,
            guest_count=50,
            wine_sales=80,
            rate=2.0,
            ticket_seconds=25 * 60,
            ranks=(5, 5, 5, 5),
        ),
        ranked_row(
            latest_week,
            display_name="Alex Rising",
            gross_sales=1400,
            guest_count=50,
            wine_sales=168,
            rate=1.5,
            ticket_seconds=18 * 60,
            ranks=(1, 1, 1, 1),
        ),
        ranked_row(
            prior_week,
            display_name="Jordan Falling",
            gross_sales=1500,
            guest_count=50,
            wine_sales=180,
            rate=1.5,
            ticket_seconds=18 * 60,
            ranks=(1, 1, 1, 1),
        ),
        ranked_row(
            latest_week,
            display_name="Jordan Falling",
            gross_sales=1000,
            guest_count=50,
            wine_sales=50,
            rate=2.0,
            ticket_seconds=25 * 60,
            ranks=(6, 6, 6, 6),
        ),
    ]

    stars = metrics.server_star_rows(metrics.server_week_trend_rows(ranked_rows), config)
    by_name = {row["display_name"]: row for row in stars}

    rising = by_name["Alex Rising"]
    assert rising["category"] == "Rising Star"
    assert rising["composite_score"] == 4
    assert rising["average_rank_movement"] == 4
    assert "Sales / Guest +$8.00" in rising["why"]
    assert "Wine +4.0 pts" in rising["why"]
    assert "Rate" not in rising["why"]
    assert "Ticket" not in rising["why"]

    falling = by_name["Jordan Falling"]
    assert falling["category"] == "Falling Star"
    assert falling["composite_score"] == -4
    assert falling["average_rank_movement"] == -5


def test_star_rows_exclude_low_volume_admin_takeout_and_placeholders(tmp_path: Path) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    prior_week = date(2026, 6, 7)
    latest_week = date(2026, 6, 14)
    ranked_rows: list[dict] = []
    for name, guest_count, active_days in (
        ("AGM Manager", 50, 3),
        ("Takeout", 50, 3),
        ("1 Server Server", 50, 3),
        ("Low Signal Person", 5, 1),
        ("Real Person", 50, 3),
    ):
        ranked_rows.extend(
            [
                ranked_row(
                    prior_week,
                    display_name=name,
                    gross_sales=1000,
                    guest_count=guest_count,
                    wine_sales=100,
                    active_days=active_days,
                    ranks=(5, 5, 5, 5),
                ),
                ranked_row(
                    latest_week,
                    display_name=name,
                    gross_sales=1500,
                    guest_count=guest_count,
                    wine_sales=180,
                    rate=1.0,
                    ticket_seconds=10 * 60,
                    active_days=active_days,
                    ranks=(1, 1, 1, 1),
                ),
            ]
        )

    stars = metrics.server_star_rows(metrics.server_week_trend_rows(ranked_rows), config)

    assert [row["display_name"] for row in stars] == ["Real Person"]


def test_management_scoring_uses_peer_reference_and_ignores_rank(tmp_path: Path) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    for family in ("management_score_thresholds", "management_peer_score_thresholds"):
        config[family] = {
            "check_average": {"neutral": 2.5, "strong": 5.0, "lower_is_better": False},
            "wine_pct": {"neutral": 0.005, "strong": 0.01, "lower_is_better": False},
            "rate_of_sale_by_guest_count": {"neutral": 0.005, "strong": 0.01, "lower_is_better": True},
            "average_ticket_time_seconds": {"neutral": 150.0, "strong": 300.0, "lower_is_better": True},
        }
    weeks = [date(2026, 5, 17) + timedelta(days=7 * index) for index in range(6)]
    weekly_servers: list[dict] = []
    weekly_locations: list[dict] = []
    for index, week_end in enumerate(weeks):
        improved = index >= 4
        weekly_servers.append(
            weekly_row(
                week_end,
                display_name="Alex Rising",
                gross_sales=3500 if improved else 1000,
                guest_count=100,
                wine_sales=700 if improved else 10,
                rate=0.10 if improved else 0.40,
                ticket_seconds=50 * 60 if improved else 120 * 60,
                active_days=6,
            )
        )
        for peer_index in range(6):
            weekly_servers.append(
                weekly_row(
                    week_end,
                    display_name=f"Peer {peer_index}",
                    gross_sales=2000,
                    guest_count=100,
                    wine_sales=160,
                    rate=0.25,
                    ticket_seconds=5000,
                    active_days=6,
                )
            )
        weekly_locations.append(
            weekly_row(
                week_end,
                display_name="",
                gross_sales=14000,
                guest_count=700,
                wine_sales=1120,
                rate=0.25,
                ticket_seconds=5000,
                active_days=6,
            )
        )
    ranked = [
        {
            **row,
            "check_average_rank": 999,
            "wine_pct_rank": 999,
            "rate_rank": 999,
            "ticket_time_rank": 999,
        }
        for row in weekly_servers
    ]

    rows = metrics.management_server_rows(
        weekly_servers,
        weekly_locations,
        ranked,
        {"RC Richmond": {"check_average": -999999.0}},
        config,
    )

    result = next(row for row in rows if row["display_name"] == "Alex Rising")
    assert result["prominent"] is True
    assert result["momentum"] == "Upward"
    assert result["performance_level"] == "Above Peer Reference"
    assert result["rank_modifier"] == 0
    assert result["average_rank_movement"] is None
    assert result["peer_cohort_size"] == 6
    assert set(result["benchmark_sources"].values()) == {
        "Same-store prior-four-week median"
    }
    assert result["action"] == "Context Review"


def test_management_confidence_fails_closed_and_excludes_service_areas(
    tmp_path: Path,
) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    weeks = [date(2026, 6, 7), date(2026, 6, 14), date(2026, 6, 21)]
    servers: list[dict] = []
    ranked: list[dict] = []
    for name, guests, days in (("Bar", 100, 6), ("Patio", 40, 4), ("Low Sample", 24, 6), ("Real Person", 25, 3)):
        for index, week_end in enumerate(weeks):
            row = weekly_row(
                week_end,
                display_name=name,
                gross_sales=1000 + index * 300,
                guest_count=guests,
                wine_sales=100 + index * 40,
                rate=0.20 - index * 0.01,
                ticket_seconds=(80 - index * 5) * 60,
                active_days=days,
            )
            servers.append(row)
            ranked.append(
                {**row, "check_average_rank": 5 - index, "wine_pct_rank": 5 - index,
                 "rate_rank": 5 - index, "ticket_time_rank": 5 - index}
            )
    locations = [
        weekly_row(week_end, display_name="", gross_sales=5000, guest_count=250, active_days=6)
        for week_end in weeks
    ]

    rows = metrics.management_server_rows(servers, locations, ranked, {}, config)
    by_name = {row["display_name"]: row for row in rows}

    assert "Bar" not in by_name
    assert "Patio" not in by_name
    assert by_name["Low Sample"]["prominent"] is False
    assert by_name["Low Sample"]["action"] == "Monitor"
    assert by_name["Low Sample"]["confidence"] == "Limited Volume"
    assert by_name["Real Person"]["prominent"] is False
    assert by_name["Real Person"]["performance_level"] == "Reference Unavailable"
    assert by_name["Real Person"]["confidence"] == "Reference Unavailable"


def test_management_baseline_excludes_short_weeks_and_target_takes_precedence(
    tmp_path: Path,
) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    rows = [
        weekly_row(date(2026, 6, 7), gross_sales=100, guest_count=5, active_days=2),
        weekly_row(date(2026, 6, 14), gross_sales=1000, guest_count=50, active_days=6),
        weekly_row(date(2026, 6, 21), gross_sales=1200, guest_count=60, active_days=6),
        weekly_row(date(2026, 6, 28), gross_sales=1500, guest_count=60, active_days=6),
    ]
    full_by_location, _ = metrics.full_week_ends_by_location(rows)
    targets = {"RC Richmond": {"check_average": 30.0}}

    result = metrics.management_entity_rows(
        rows, "location", targets, config, full_by_location
    )[0]

    assert result["baseline_weeks"] == 2
    assert result["baseline"]["gross_sales"] == 1100
    assert result["baseline"]["guest_count"] == 55
    assert result["benchmark_values"]["check_average"] == 30.0
    assert result["benchmark_sources"]["check_average"] == "Target"
    assert result["benchmark_sources"]["gross_sales"] == "2-week baseline"


def test_action_tracking_carries_manual_fields_and_moves_cleared_items_to_history() -> None:
    signal = {
        "Entity Key": "server|richmond|server one|coaching",
        "Priority": "High",
        "Location": "RC Richmond",
        "Person / Area": "Server One",
        "Action": "Coach Now",
        "Signal": "Falling / Below Benchmark",
        "Why It Matters": "Watch: check average",
        "Recommended Next Step": "Coach check building.",
        "Performance Level": "Below Benchmark",
        "Momentum": "Falling",
        "Confidence": "High",
        "Last Seen": date(2026, 6, 21),
    }
    current, history = metrics.merge_management_actions([signal], {})
    assert not history
    current[0]["Status"] = "In Progress"
    current[0]["Owner"] = "Pat Manager"
    current[0]["Due Date"] = date(2026, 6, 30)
    current[0]["Context Notes"] = "Review Friday"
    current[0]["Review Disposition"] = "Coaching Accepted"
    current[0]["Reviewed By"] = "Pat Manager"
    current[0]["Review Date"] = date(2026, 6, 23)
    next_signal = {**signal, "Last Seen": date(2026, 6, 28)}

    carried, history = metrics.merge_management_actions(
        [next_signal], {"active_actions": current, "action_history": []}
    )

    assert carried[0]["Action ID"] == current[0]["Action ID"]
    assert carried[0]["Status"] == "In Progress"
    assert carried[0]["Owner"] == "Pat Manager"
    assert carried[0]["Due Date"] == date(2026, 6, 30)
    assert carried[0]["Context Notes"] == "Review Friday"
    assert carried[0]["Review Disposition"] == "Coaching Accepted"
    assert carried[0]["Reviewed By"] == "Pat Manager"
    assert carried[0]["Review Date"] == date(2026, 6, 23)
    assert carried[0]["Weeks Open"] == 2
    cleared, history = metrics.merge_management_actions(
        [], {"active_actions": carried, "action_history": []}
    )
    assert cleared == []
    assert history[0]["Signal State"] == "Cleared"


def test_new_evidence_resets_a_recurring_action_to_review_needed() -> None:
    signal = {
        "Entity Key": "server|richmond|server one|coaching",
        "Evidence ID": "EVIDENCE-ONE",
        "Priority": "Medium",
        "Location": "RC Richmond",
        "Person / Area": "Server One",
        "Action": "Coaching Prompt",
        "Signal": "Downward / Below Peer Reference",
        "Why It Matters": "Watch: check average",
        "Recommended Next Step": "Review comparable work context.",
        "Peer Comparison": "Below Peer Reference",
        "Recent Movement": "Downward",
        "Evidence Status": "Stable",
        "Last Seen": date(2026, 6, 21),
    }
    current, _ = metrics.merge_management_actions([signal], {})
    current[0].update(
        {
            "Status": "In Progress",
            "Owner": "Pat Manager",
            "Review Disposition": "Coaching Accepted",
            "Reviewed By": "Pat Manager",
            "Review Date": date(2026, 6, 23),
        }
    )
    changed_signal = {
        **signal,
        "Evidence ID": "EVIDENCE-TWO",
        "Last Seen": date(2026, 6, 28),
    }

    carried, _ = metrics.merge_management_actions(
        [changed_signal],
        {"active_actions": current, "action_history": []},
    )

    assert carried[0]["Action ID"] == current[0]["Action ID"]
    assert carried[0]["Status"] == "Review Needed"
    assert carried[0]["Owner"] == "Pat Manager"
    assert carried[0]["Review Disposition"] == "Pending Review"
    assert carried[0]["Reviewed By"] == ""
    assert carried[0]["Review Date"] is None


def test_action_episode_id_uses_stable_sha256_identifier() -> None:
    action_id = metrics.action_episode_id(
        "server|richmond|server one|coaching", date(2026, 6, 21)
    )

    assert action_id == "56AFCB611719"
    assert len(action_id) == 12


def test_store_week_trend_deltas() -> None:
    prior_week = weekly_row(
        date(2026, 6, 7),
        gross_sales=1000,
        guest_count=50,
        wine_sales=100,
        rate=2.0,
        ticket_seconds=25 * 60,
        active_days=6,
    )
    latest_week = weekly_row(
        date(2026, 6, 14),
        gross_sales=1500,
        guest_count=60,
        wine_sales=180,
        rate=1.5,
        ticket_seconds=20 * 60,
        active_days=6,
    )

    trend_rows = metrics.weekly_metric_trend_rows([prior_week, latest_week], ("location",))
    latest = trend_rows[-1]

    assert latest["gross_sales_change"] == 500
    assert latest["guest_count_change"] == 10
    assert latest["check_average_change"] == 5
    assert latest["wine_pct_change"] == pytest.approx(0.02)
    assert latest["rate_change"] == -0.5
    assert latest["ticket_time_change_minutes"] == -5
    assert latest["trend_note"] == "Improving"


def test_all_stores_group_aggregation_and_trend_deltas() -> None:
    records = [
        make_record(
            date(2026, 6, 7),
            "RC Richmond",
            is_location_total=True,
            gross_sales=1000,
            guest_count=50,
            wine_sales=100,
            rate=2,
            ticket_seconds=600,
            source_file="richmond-prior.xls",
        ),
        make_record(
            date(2026, 6, 7),
            "RC Virginia Beach",
            is_location_total=True,
            gross_sales=500,
            guest_count=25,
            wine_sales=50,
            rate=4,
            ticket_seconds=1200,
            source_file="vb-prior.xls",
        ),
        make_record(
            date(2026, 6, 14),
            "RC Richmond",
            is_location_total=True,
            gross_sales=1200,
            guest_count=60,
            wine_sales=180,
            rate=1.5,
            ticket_seconds=600,
            source_file="richmond-latest.xls",
        ),
        make_record(
            date(2026, 6, 14),
            "RC Virginia Beach",
            is_location_total=True,
            gross_sales=900,
            guest_count=30,
            wine_sales=90,
            rate=3,
            ticket_seconds=900,
            source_file="vb-latest.xls",
        ),
    ]

    group_rows = metrics.group_weekly_rows(records)
    prior = group_rows[0]
    assert prior["group"] == "All Stores"
    assert prior["gross_sales"] == 1500
    assert prior["guest_count"] == 75
    assert prior["check_average"] == 20
    assert prior["wine_pct"] == pytest.approx(0.1)
    assert prior["rate_of_sale_by_guest_count"] == pytest.approx(
        75 / (50 / 2 + 25 / 4)
    )
    assert prior["ticket_time_available"] is False
    assert prior["average_ticket_time_seconds"] == 0
    assert prior["active_days"] == 1

    trend_rows = metrics.weekly_metric_trend_rows(group_rows, ("group",))
    latest = trend_rows[-1]
    assert latest["gross_sales_change"] == 600
    assert latest["guest_count_change"] == 15
    assert latest["check_average_change"] == pytest.approx((2100 / 90) - 20)
    assert latest["wine_pct_change"] == pytest.approx((270 / 2100) - 0.1)


def test_master_workbook_contains_star_store_group_and_dashboard_sections(tmp_path: Path) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    records: list[metrics.MetricRecord] = []
    for week_end, gross, wine, rate, ticket_seconds in (
        (date(2026, 6, 7), 1000, 100, 0.20, 80 * 60),
        (date(2026, 6, 14), 1100, 110, 0.20, 80 * 60),
        (date(2026, 6, 21), 600, 20, 0.30, 95 * 60),
    ):
        for location in ("RC Richmond", "RC Virginia Beach"):
            records.extend(
                full_week_records(
                    week_end,
                    location=location,
                    weekly_gross=gross,
                    weekly_guests=60,
                    weekly_wine=wine,
                    rate=rate,
                    ticket_seconds=ticket_seconds,
                    server="Alex Rising",
                )
            )

    output_path = tmp_path / "Red_Onion_Server_Master.xlsx"
    metrics.write_master_workbook(
        records,
        output_path,
        config,
        tmp_path / "Daily Reports",
        date(2026, 6, 16),
        date(2026, 6, 21),
    )

    wb = load_workbook(output_path)
    assert wb.sheetnames[: len(metrics.VISIBLE_MANAGEMENT_SHEETS)] == (
        metrics.VISIBLE_MANAGEMENT_SHEETS
    )
    assert wb["_Dashboard Chart Data"].sheet_state == "veryHidden"
    assert wb["Weekly Server Metrics"].sheet_state == "veryHidden"
    assert wb["Store Week Trends"].sheet_state == "veryHidden"
    assert wb.active.title == "How to Use"
    assert wb["How to Use"].sheet_state == "visible"
    assert wb["Dashboard"].sheet_state == "visible"
    assert wb["Action Board"].sheet_state == "visible"

    guide = wb["How to Use"]
    guide_values = {
        value
        for row in guide.iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    }
    assert set(metrics.HOW_TO_USE_SECTION_HEADINGS).issubset(guide_values)
    assert all(
        cell.protection.locked is not False for cell in guide._cells.values()
    )
    assert not guide._charts
    assert not guide._images
    assert guide.freeze_panes == "A3"
    assert guide.sheet_view.zoomScale == 85
    assert guide.sheet_properties.tabColor.rgb[-6:] == "FFD966"
    assert all(
        guide.column_dimensions[metrics.get_column_letter(column)].width == 13
        for column in range(1, 13)
    )
    assert guide.row_dimensions[1].height == 32
    assert guide.row_dimensions[2].height == 24
    assert guide.row_dimensions[3].height == 24
    assert guide.row_dimensions[4].height == 54
    assert guide.row_dimensions[30].height == 42
    assert guide.row_dimensions[58].height == 24
    assert guide.row_dimensions[59].height == 72
    assert guide["A44"].value == metrics.HOW_TO_USE_SECTION_HEADINGS[6]
    assert metrics.MANAGEMENT_METHODOLOGY_VERSION in "\n".join(guide_values)
    for phrase in (
        "never be the sole or determinative basis",
        "Do not infer protected characteristics from names.",
        "ExternalCheckRequired",
        "Rate of Sale is opportunities divided by qualifying sales",
        "Ticket Time is weighted only by complete Check Count coverage",
    ):
        assert phrase in "\n".join(guide_values)
    for field in (
        "D — Status",
        "E — Owner",
        "F — Due Date",
        "N — Context Notes",
        "U — Review Disposition",
        "V — Reviewed By",
        "W — Review Date",
    ):
        assert field in guide_values
    for choice in (*metrics.ACTION_STATUS_CHOICES, *metrics.REVIEW_DISPOSITION_CHOICES):
        assert choice in "\n".join(guide_values)
    assert guide["A45"].value == "Sheet - click to open"
    workbook_map_rows = range(
        46, 46 + len(metrics.VISIBLE_MANAGEMENT_SHEETS)
    )
    assert [
        guide.cell(row=row, column=1).value for row in workbook_map_rows
    ] == list(metrics.VISIBLE_MANAGEMENT_SHEETS)
    assert [
        guide.cell(row=row, column=1).hyperlink.target
        for row in workbook_map_rows
    ] == [
        f"#'{sheet_name}'!A1"
        for sheet_name in metrics.VISIBLE_MANAGEMENT_SHEETS
    ]
    assert all(
        guide.cell(row=row, column=1).font.underline == "single"
        and guide.cell(row=row, column=1).protection.locked is True
        for row in workbook_map_rows
    )

    for sheet_name in metrics.VISIBLE_MANAGEMENT_SHEETS:
        worksheet = wb[sheet_name]
        start_column, end_column = metrics.management_menu_bounds(worksheet)
        expected_merge = (
            f"{metrics.get_column_letter(start_column)}2:"
            f"{metrics.get_column_letter(end_column)}2"
        )
        menu_cell = worksheet.cell(row=2, column=start_column)
        assert worksheet.row_dimensions[2].height == 24
        assert {
            str(merged)
            for merged in worksheet.merged_cells.ranges
            if merged.min_row <= 2 <= merged.max_row
        } == {expected_merge}
        assert menu_cell.value == metrics.MANAGEMENT_MENU_LABEL
        assert menu_cell.hyperlink.target == metrics.MANAGEMENT_MENU_TARGET
        assert menu_cell.font.underline == "single"
        assert menu_cell.font.bold is True
        assert menu_cell.font.size == 9
        assert metrics.workbook_color_suffix(menu_cell.font.color) == "7A1E1E"
        assert menu_cell.protection.locked is True
        assert menu_cell.alignment.horizontal == "left"
        assert menu_cell.alignment.vertical == "center"
        assert menu_cell.alignment.shrink_to_fit is True
        assert menu_cell.alignment.indent == 1
        assert (
            metrics.workbook_color_suffix(menu_cell.fill.fgColor)
            == "F2F2F2"
        )
        assert getattr(menu_cell.border.left, "style", None) is None
        assert getattr(menu_cell.border.right, "style", None) is None
        assert menu_cell.border.top.style == "thin"
        assert menu_cell.border.bottom.style == "thin"
        assert (
            worksheet.column_dimensions[
                metrics.get_column_letter(start_column)
            ].hidden
            is not True
        )
        assert (
            worksheet.column_dimensions[
                metrics.get_column_letter(end_column)
            ].hidden
            is not True
        )
        assert (
            worksheet.page_setup.fitToWidth
            == metrics.MANAGEMENT_PRINT_WIDTHS[sheet_name]
        )
        if sheet_name == "Data Quality":
            assert (
                str(worksheet.page_setup.paperSize)
                == str(worksheet.PAPERSIZE_TABLOID)
            )
        assert worksheet.page_setup.orientation == "landscape"
        assert worksheet.print_area
        assert worksheet.print_title_rows
        assert any(
            merged.min_row == merged.max_row == 1
            and merged.min_col == 1
            and merged.max_col >= 12
            for merged in worksheet.merged_cells.ranges
        )
        frozen_at = worksheet.freeze_panes
        assert frozen_at is not None
        assert int("".join(character for character in str(frozen_at) if character.isdigit())) >= 3
    assert metrics.management_menu_bounds(wb["Action Board"]) == (3, 21)
    assert metrics.management_menu_bounds(wb["Action History"]) == (3, 14)
    assert metrics.management_menu_bounds(wb["Evidence Detail"]) == (1, 14)
    assert "Action Focus" not in wb.sheetnames
    assert "Server Scorecard" not in wb.sheetnames
    assert "Recent Movement Signals" not in wb.sheetnames

    dashboard_values = {
        value
        for row in wb["Dashboard"].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    }
    assert "TOP THREE REVIEW ITEMS" in dashboard_values
    assert "RECOGNITION REVIEW" in dashboard_values
    assert "STORE SNAPSHOT" in dashboard_values
    assert any("READY FOR HUMAN REVIEW" in value for value in dashboard_values)
    assert len(wb["Dashboard"]._charts) == 0
    assert len(wb["Store & Group Scorecards"]._charts) == 2
    assert wb["Store & Group Scorecards"]._charts[0].legend.position == "b"
    assert wb["Store & Group Scorecards"]._charts[0].anchor._from.row == 3
    assert wb["Store & Group Scorecards"]._charts[1].anchor._from.row == 16

    action_values = {
        value
        for row in wb["Action Board"].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    }
    assert "Owner" in action_values
    assert "Due Date" in action_values
    assert "Status" in action_values
    assert "Recommended Next Step" in action_values
    assert len(wb["Action Board"].data_validations.dataValidation) == 3
    assert len(wb["Action Board"].conditional_formatting) >= 3
    assert wb["Action Board"]["C3"].value.startswith("Single action queue:")
    assert "C3:W3" in {
        str(merged) for merged in wb["Action Board"].merged_cells.ranges
    }
    assert wb["Action Board"].row_dimensions[3].height == 36
    assert all(
        wb["Action Board"].column_dimensions[column].hidden is True
        for column in ("A", "B", "J", "M", "O", "P", "Q", "R", "T")
    )
    assert wb["Action Board"].row_dimensions[4].height == 30
    assert wb["Action Board"].row_dimensions[5].height == 60
    assert wb["Action Board"]["L5"].alignment.wrap_text is True
    assert wb["Action Board"]["N5"].alignment.wrap_text is True
    assert wb["Action Board"]["F5"].number_format == "m/d/yyyy"
    assert wb["Action Board"]["M5"].number_format == "m/d/yyyy"
    assert wb["Action Board"]["Q5"].number_format == "m/d/yyyy"
    assert wb["Action Board"]["W5"].number_format == "m/d/yyyy"
    assert wb["Action History"].row_dimensions[4].height == 30
    evidence = wb["Evidence Detail"]
    assert evidence.row_dimensions[3].height == 54
    assert "exact raw Evidence Sources and Metric Evidence columns are hidden" in (
        evidence["A3"].value
    )
    assert evidence.row_dimensions[4].height == 42
    assert evidence.column_dimensions["D"].width == 46
    assert evidence.column_dimensions["L"].hidden is True
    assert evidence.column_dimensions["M"].hidden is True
    assert all(
        evidence.row_dimensions[row].height in {75, 90}
        for row in range(5, evidence.max_row + 1)
    )
    evidence_headers = {
        evidence.cell(row=4, column=column).value: column
        for column in range(1, evidence.max_column + 1)
    }
    for row in range(5, evidence.max_row + 1):
        assert str(
            evidence.cell(
                row=row,
                column=evidence_headers["Evidence Sources"],
            ).value
        ).startswith("[")
        assert str(
            evidence.cell(
                row=row,
                column=evidence_headers["Metric Evidence"],
            ).value
        ).startswith("{")
    run_notes = wb["Run Notes"]
    run_note_rows = {
        run_notes.cell(row=row, column=1).value: row
        for row in range(1, run_notes.max_row + 1)
    }
    for label in (
        "8-Week Direction",
        "Peer Comparison",
        "Signal Scoring",
        "Metric Caveat",
    ):
        assert run_notes.row_dimensions[run_note_rows[label]].height == 60
        assert run_notes.cell(
            row=run_note_rows[label],
            column=2,
        ).alignment.wrap_text is True
    assert "Check Count Coverage" in run_note_rows
    assert wb["Management Setup"].row_dimensions[5].height == 30
    assert wb["Management Setup"]["C5"].alignment.wrap_text is True
    assert all(
        wb["Store & Group Scorecards"].row_dimensions[row].height == 30
        for row in range(18, 30)
    )
    assert "A18:G23" in {
        str(merged)
        for merged in wb["Store & Group Scorecards"].merged_cells.ranges
    }
    assert wb["Store & Group Scorecards"]["A18"].alignment.wrap_text is True
    assert wb["Dashboard"].row_dimensions[13].height == 66
    assert wb["Dashboard"].row_dimensions[14].height == 66
    assert wb["Dashboard"].row_dimensions[15].height == 66
    assert wb["Dashboard"].row_dimensions[19].height == 60
    assert wb["Dashboard"].row_dimensions[20].height == 60
    assert wb["Dashboard"]["K19"].alignment.wrap_text is True
    assert wb["Dashboard"]["K20"].alignment.wrap_text is True
    assert wb["Dashboard"].row_dimensions[9].height == 32
    trends = wb["Team Trends"]
    assert trends.row_dimensions[1].height == 42
    assert trends.row_dimensions[3].height == 30
    assert trends.row_dimensions[4].height == 54
    assert trends.freeze_panes == "E4"
    assert trends.page_setup.fitToWidth == 2
    assert trends.max_row == 5
    assert "TeamTrends" in trends.tables
    assert [trends.cell(row=row, column=1).value for row in (4, 5)] == [
        "RC Richmond",
        "RC Virginia Beach",
    ]
    assert [trends.cell(row=row, column=2).value for row in (4, 5)] == [
        "Alex Rising",
        "Alex Rising",
    ]
    assert trends["C4"].value.date() == date(2026, 6, 21)
    assert trends["D4"].value.date() == date(2026, 6, 14)
    assert trends["F4"].value == pytest.approx(10.0)
    assert trends["G4"].value == pytest.approx(10.0 - (1100 / 60))
    assert trends["H4"].value == pytest.approx(20 / 600)
    assert trends["I4"].value == pytest.approx((20 / 600) - (110 / 1100))
    assert trends["J4"].value == "Watch"
    assert trends["N4"].value == "Reference Unavailable"
    assert trends["O4"].value == "Monitor"
    assert wb["Data Quality"].row_dimensions[3].height == 48
    assert "$A$1:$L$" in str(wb["Data Quality"].print_area)
    assert len(wb["Store & Group Scorecards"].row_breaks.brk) == 3
    assert [
        page_break.id
        for page_break in wb["Store & Group Scorecards"].row_breaks.brk
    ] == [16, 29, 42]
    for sheet_name, widths in {
        "Team Trends": {
            "C": 15,
            "D": 17,
            "F": 17,
            "I": 17,
            "J": 20,
            "K": 23,
            "L": 22,
            "P": 42,
            "Q": 52,
        },
        "Data Quality": {"D": 13, "F": 69},
        "Management Setup": {"D": 25, "E": 48},
        "Run Notes": {"A": 30, "B": 96},
        "Evidence Detail": {"D": 46},
    }.items():
        for column, expected_width in widths.items():
            assert (
                wb[sheet_name].column_dimensions[column].width
                == expected_width
            )
    assert "ManagementTargets" in wb["Management Setup"].tables
    assert "Alex Rising" in {
        cell.value for row in trends.iter_rows() for cell in row
    }
    assert "Reference Unavailable" in {
        cell.value for row in trends.iter_rows() for cell in row
    }


def test_incomplete_latest_week_hides_team_trends_and_server_actions(tmp_path: Path) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    records: list[metrics.MetricRecord] = []
    for week_end in (date(2026, 6, 14), date(2026, 6, 21), date(2026, 6, 28)):
        for location in ("RC Richmond", "RC Virginia Beach"):
            records.extend(
                full_week_records(
                    week_end,
                    location=location,
                    weekly_gross=1200,
                    weekly_guests=60,
                    weekly_wine=120,
                    server="Alex Server",
                )
            )
    records = [
        record
        for record in records
        if not (
            record.report_date == date(2026, 6, 27)
            and record.source_file.endswith("2026-06-27.xls")
        )
    ]
    output_path = tmp_path / "Red_Onion_Server_Master.xlsx"

    metrics.write_master_workbook(
        records,
        output_path,
        config,
        tmp_path / "Daily Reports",
        date(2026, 6, 23),
        date(2026, 6, 28),
    )

    workbook = load_workbook(output_path, data_only=False)
    assert workbook["Team Trends"].max_row == 3
    assert "Action Focus" not in workbook.sheetnames
    assert "Server Scorecard" not in workbook.sheetnames
    assert "Recent Movement Signals" not in workbook.sheetnames
    dashboard_values = {
        cell.value for row in workbook["Dashboard"].iter_rows() for cell in row
    }
    assert "Alex Server" not in dashboard_values
    assert any(
        "paused" in value.casefold()
        for value in dashboard_values
        if isinstance(value, str)
    )
    workbook.close()


def test_master_regeneration_preserves_targets_and_manual_action_fields(tmp_path: Path) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    records: list[metrics.MetricRecord] = []
    for week_end, gross, wine in (
        (date(2026, 6, 7), 1000, 100),
        (date(2026, 6, 14), 1100, 110),
        (date(2026, 6, 21), 600, 20),
    ):
        for location in ("RC Richmond", "RC Virginia Beach"):
            records.extend(
                full_week_records(
                    week_end,
                    location=location,
                    weekly_gross=gross,
                    weekly_guests=60,
                    weekly_wine=wine,
                    rate=0.30 if week_end == date(2026, 6, 21) else 0.20,
                    ticket_seconds=(95 if week_end == date(2026, 6, 21) else 80) * 60,
                    server="Alex Rising",
                )
            )
    output_path = tmp_path / "Red_Onion_Server_Master.xlsx"
    metrics.write_master_workbook(
        records, output_path, config, tmp_path / "Daily Reports",
        date(2026, 6, 16), date(2026, 6, 21),
    )
    wb = load_workbook(output_path)
    setup = wb["Management Setup"]
    setup["D7"] = 28.0
    setup["A21"] = "Pat Manager"
    setup["B21"] = "Yes"
    actions = wb["Action Board"]
    assert actions.max_row >= 5
    actions["D5"] = "In Progress"
    actions["E5"] = "Pat Manager"
    actions["F5"] = date(2026, 6, 30)
    actions["N5"] = "Follow up Friday"
    actions["U5"] = "Coaching Accepted"
    actions["V5"] = "Pat Manager"
    actions["W5"] = date(2026, 6, 23)
    action_id = actions["A5"].value
    wb.save(output_path)
    wb.close()

    metrics.write_master_workbook(
        records, output_path, config, tmp_path / "Daily Reports",
        date(2026, 6, 16), date(2026, 6, 21),
    )

    regenerated = load_workbook(output_path, data_only=False)
    assert regenerated["Management Setup"]["D7"].value == 28.0
    roster = regenerated["Management Setup"].tables[metrics.OWNER_ROSTER_TABLE_NAME]
    assert roster.ref == "A20:B70"
    assert regenerated["Management Setup"]["A21"].value == "Pat Manager"
    assert regenerated["Management Setup"]["B21"].value == "Yes"
    action_rows = metrics.records_from_sheet(regenerated["Action Board"], "Action ID")
    carried = next(row for row in action_rows if row["Action ID"] == action_id)
    assert carried["Status"] == "In Progress"
    assert carried["Owner"] == "Pat Manager"
    assert carried["Due Date"].date() == date(2026, 6, 30)
    assert carried["Context Notes"] == "Follow up Friday"
    assert carried["Review Disposition"] == "Coaching Accepted"
    assert carried["Reviewed By"] == "Pat Manager"
    assert carried["Review Date"].date() == date(2026, 6, 23)
    regenerated.close()


def test_unreadable_existing_master_fails_without_overwriting(tmp_path: Path) -> None:
    output_path = tmp_path / "Red_Onion_Server_Master.xlsx"
    output_path.write_text("not an xlsx", encoding="utf-8")
    original = output_path.read_bytes()

    with pytest.raises(RuntimeError, match="Could not read the existing master workbook"):
        metrics.write_master_workbook(
            [make_record(date(2026, 6, 7))],
            output_path,
            metrics.load_config(tmp_path / "missing-config.json"),
            tmp_path,
            date(2026, 6, 3),
            date(2026, 6, 7),
        )

    assert output_path.read_bytes() == original
