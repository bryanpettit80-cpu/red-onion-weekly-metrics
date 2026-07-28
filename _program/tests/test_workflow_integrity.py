from __future__ import annotations

from argparse import Namespace
from dataclasses import replace
from datetime import date, timedelta
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pytest

import red_onion_integrity as integrity
import red_onion_weekly_metrics as metrics


ACTIVE_DAY = date(2026, 7, 21)  # Tuesday
ACTIVE_WEEK_END = date(2026, 7, 26)  # Sunday


def workflow_args(tmp_path: Path, *, initialize_baseline: bool = False) -> Namespace:
    return Namespace(
        input_dir=str(tmp_path / "01 Daily Reports - Drop Here"),
        output_dir=str(tmp_path / "02 Finished Reports"),
        archive_dir=str(tmp_path / "03 Archive"),
        config=str(tmp_path / "missing-config.json"),
        week_start=None,
        week_end=None,
        migrate_history_from=[],
        migrate_history_only=False,
        initialize_integrity_baseline=initialize_baseline,
    )


def minimal_records(day: date, source_file: str) -> list[metrics.MetricRecord]:
    common = {
        "source_file": source_file,
        "report_date": day,
        "location": "RC Richmond",
        "gross_sales": 1000.0,
        "guest_count": 50.0,
        "check_average": 20.0,
        "wine_sales": 100.0,
        "wine_pct": 0.10,
        "rate_of_sale_by_guest_count": 0.20,
        "average_ticket_time_seconds": 600.0,
    }
    return [
        metrics.MetricRecord(
            raw_user_name="",
            display_name="",
            is_location_total=True,
            **common,
        ),
        metrics.MetricRecord(
            raw_user_name="Alex Server",
            display_name="Alex Server",
            is_location_total=False,
            **common,
        ),
    ]


@pytest.fixture(scope="module")
def valid_master_template(tmp_path_factory: pytest.TempPathFactory) -> Path:
    root = tmp_path_factory.mktemp("workflow-integrity-master")
    path = root / "Red_Onion_Server_Master.xlsx"
    metrics.write_master_workbook(
        minimal_records(ACTIVE_DAY, "Daily Report template.xlsx"),
        path,
        metrics.load_config(root / "missing-config.json"),
        root / "source",
        ACTIVE_DAY,
        ACTIVE_DAY,
    )
    assert metrics.verify_existing_management_workbook_integrity(path)
    return path


def manifest_paths(archive_dir: Path) -> list[Path]:
    root = metrics.integrity_manifest_dir(archive_dir)
    return sorted(root.glob("*.json"), key=lambda path: path.name.casefold())


def manifest_by_kind(archive_dir: Path, kind: str) -> Path:
    matches = [
        path
        for path in manifest_paths(archive_dir)
        if integrity.read_json_manifest(path).get("kind") == kind
    ]
    assert len(matches) == 1
    return matches[0]


def install_synthetic_parser(monkeypatch: pytest.MonkeyPatch) -> None:
    def parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        return minimal_records(ACTIVE_DAY, path.name)

    monkeypatch.setattr(metrics, "parse_daily_report", parse)


def downgrade_master_to_pre_consolidation_layout(path: Path) -> str:
    """Model the manifest-bound workbook before action/trend consolidation."""

    workbook = load_workbook(path, data_only=False)
    try:
        # A historical workbook cannot also carry the redesign management layer.
        # Removing these sheets keeps the compatibility fixture on the legacy
        # routing path and prevents redesign-only validations from leaking into
        # the historical protection contracts.
        for sheet_name in metrics.VISIBLE_MANAGEMENT_SHEETS:
            if sheet_name in workbook.sheetnames:
                workbook.remove(workbook[sheet_name])
        if "Team Trends" in workbook.sheetnames:
            current_actions = metrics.records_from_sheet(
                workbook["Action Board"],
                "Action ID",
            )
            workbook["Team Trends"].title = "Server Scorecard"
            workbook["Server Scorecard"]["A1"] = "Server Scorecard"
            metrics.write_action_focus_sheet(workbook, current_actions)
            metrics.write_rising_falling_sheet(workbook, [])
        else:
            missing = [
                name
                for name in metrics.PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
                if name not in workbook.sheetnames
            ]
            assert not missing

        guide = workbook["How to Use"]
        for merged in list(guide.merged_cells.ranges):
            if merged.min_row <= 57 and merged.max_row >= 46:
                guide.unmerge_cells(str(merged))
        for row in range(46, 58):
            for column in range(1, 13):
                cell = guide.cell(row=row, column=column)
                cell.value = None
                cell.hyperlink = None

        workbook_map = {
            "How to Use": (
                "start-here workflow, interpretation, controls, and limitations",
                "No - locked",
            ),
            "Dashboard": ("aggregate weekly brief", "No - locked"),
            "Action Focus": (
                "prioritized current prompts",
                "No - locked",
            ),
            "Action Board": (
                "current review and task-tracking queue",
                "Yes - seven blue fields only",
            ),
            "Server Scorecard": (
                "full server performance review",
                "No - locked",
            ),
            "Store & Group Scorecards": (
                "operational context only",
                "No - locked",
            ),
            "Recent Movement Signals": (
                "prominent recent movement signals",
                "No - locked",
            ),
            "Evidence Detail": (
                "evidence weeks, cohort, thresholds, and stability",
                "No - locked",
            ),
            "Action History": (
                "preserved closed and dismissed items",
                "No - locked",
            ),
            "Data Quality": (
                "completeness, reconciliation, and provenance",
                "No - locked",
            ),
            "Management Setup": (
                "custodian-maintained targets and owner roster",
                "Yes - blue cells only",
            ),
            "Run Notes": (
                "release, integrity, methodology, and assumptions",
                "No - locked",
            ),
        }
        for row, (_, target) in enumerate(
            metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
            start=46,
        ):
            guide.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=2,
            )
            guide.merge_cells(
                start_row=row,
                start_column=3,
                end_row=row,
                end_column=10,
            )
            guide.merge_cells(
                start_row=row,
                start_column=11,
                end_row=row,
                end_column=12,
            )
            sheet_cell = guide.cell(row=row, column=1, value=target)
            sheet_cell.hyperlink = f"#'{target}'!A1"
            sheet_cell.font = Font(
                bold=True,
                color="7A1E1E",
                underline="single",
            )
            purpose, editable = workbook_map[target]
            guide.cell(row=row, column=3, value=purpose)
            guide.cell(row=row, column=11, value=editable)
            for column in (1, 3, 11):
                guide.cell(row=row, column=column).alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                )
            guide.row_dimensions[row].height = 42 if target in {
                "How to Use",
                "Server Scorecard",
                "Evidence Detail",
                "Data Quality",
                "Management Setup",
                "Run Notes",
            } else 36

        visible_sheets = set(
            metrics.PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
        )
        for worksheet in workbook.worksheets:
            if worksheet.title in visible_sheets:
                worksheet.sheet_state = "visible"
                row_two_merges = [
                    str(merged)
                    for merged in worksheet.merged_cells.ranges
                    if merged.min_row <= 2 <= merged.max_row
                ]
                for merged_range in row_two_merges:
                    worksheet.unmerge_cells(merged_range)
                metrics.add_management_navigation(worksheet)
                metrics.configure_management_print_layout(worksheet)
            else:
                worksheet.sheet_state = "veryHidden"
            metrics.protect_worksheet(worksheet, "test-pass")
        workbook._sheets = [
            *[
                workbook[name]
                for name in metrics.PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
            ],
            *[
                worksheet
                for worksheet in workbook.worksheets
                if worksheet.title not in visible_sheets
            ],
        ]
        workbook.active = 0
        workbook.save(path)
    finally:
        workbook.close()
    return metrics.stamp_generated_content_digest(path)


def downgrade_master_to_pre_protection_contract(path: Path) -> str:
    """Model the protected owner-roster workbook deployed before PR #10."""

    current = load_workbook(path, data_only=False, read_only=True)
    try:
        needs_layout_downgrade = "Team Trends" in current.sheetnames
    finally:
        current.close()
    if needs_layout_downgrade:
        downgrade_master_to_pre_consolidation_layout(path)
    workbook = load_workbook(path, data_only=False)
    if "How to Use" in workbook.sheetnames:
        workbook.remove(workbook["How to Use"])
    run_notes = workbook["Run Notes"]
    marker_row = next(
        (
            row
            for row in range(1, run_notes.max_row + 1)
            if run_notes.cell(row=row, column=1).value == "Protection Contract"
        ),
        None,
    )
    if marker_row is not None:
        run_notes.delete_rows(marker_row, 1)
    for worksheet in workbook.worksheets:
        worksheet.protection.objects = False
        worksheet.protection.scenarios = False
    for sheet_name in ("Management Setup", "Action Board"):
        for validation in workbook[sheet_name].data_validations.dataValidation:
            validation.showErrorMessage = False
            validation.errorStyle = None
            validation.errorTitle = None
            validation.error = None
    workbook.save(path)
    workbook.close()
    return metrics.stamp_generated_content_digest(path)


def replace_menu_with_legacy_tabs(
    workbook,
    *,
    previous_v2: bool,
) -> None:
    """Replace the current merged menu with the exact prior multi-link bar."""

    for sheet_name in metrics.PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS:
        worksheet = workbook[sheet_name]
        row_two_merges = [
            str(merged)
            for merged in worksheet.merged_cells.ranges
            if merged.min_row <= 2 <= merged.max_row
        ]
        for merged_range in row_two_merges:
            worksheet.unmerge_cells(merged_range)
        navigation_columns = metrics.management_navigation_columns(
            worksheet,
            metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
        )
        for column, (label, target) in zip(
            navigation_columns,
            metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
            strict=True,
        ):
            cell = worksheet.cell(row=2, column=column, value=label)
            cell.hyperlink = f"#'{target}'!A1"
            is_current = target == sheet_name
            cell.fill = PatternFill(
                "solid",
                fgColor=(
                    "7A1E1E"
                    if previous_v2 and is_current
                    else "F2F2F2"
                ),
            )
            cell.font = Font(
                color=(
                    "FFFFFF"
                    if previous_v2 and is_current
                    else "7A1E1E"
                ),
                bold=True,
                underline=(
                    None
                    if previous_v2
                    else ("single" if is_current else None)
                ),
                size=9,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                shrink_to_fit=True,
            )
            cell.border = Border(
                left=(
                    Side(style="thin", color="B7B7B7")
                    if previous_v2
                    else Side()
                ),
                right=(
                    Side(style="thin", color="B7B7B7")
                    if previous_v2
                    else Side()
                ),
                top=Side(style="thin", color="B7B7B7"),
                bottom=Side(style="thin", color="B7B7B7"),
            )


def downgrade_master_to_v2_usability_contract(path: Path) -> str:
    """Model the manifest-bound v0.3.2 workbook deployed before v3."""

    downgrade_master_to_pre_consolidation_layout(path)
    workbook = load_workbook(path, data_only=False)
    try:
        guide = workbook["How to Use"]
        for cell in guide._cells.values():
            if not isinstance(cell.value, str):
                continue
            cell.value = cell.value.replace(
                metrics.MANAGEMENT_METHODOLOGY_VERSION,
                metrics.PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION,
            )
            if "Rate of Sale is opportunities divided by qualifying sales" in cell.value:
                cell.value = (
                    "Rate of Sale currently assumes lower is better. Ticket Time is "
                    "guest-weighted because ticket/check count is unavailable. These "
                    "are action-driving business assumptions pending source-owner "
                    "confirmation."
                )
        replace_menu_with_legacy_tabs(workbook, previous_v2=True)
        workbook.save(path)
    finally:
        workbook.close()
    return metrics.stamp_generated_content_digest(path)


def downgrade_master_to_legacy_v3_navigation(path: Path) -> str:
    """Model the manifest-bound v3 workbook emitted before the single menu."""

    downgrade_master_to_pre_consolidation_layout(path)
    workbook = load_workbook(path, data_only=False)
    try:
        replace_menu_with_legacy_tabs(workbook, previous_v2=False)
        workbook.save(path)
    finally:
        workbook.close()
    return metrics.stamp_generated_content_digest(path)


def downgrade_master_to_v1_action_focus(path: Path) -> str:
    """Model the protected 20-column Action Focus workbook from v0.2.x."""

    downgrade_master_to_pre_consolidation_layout(path)
    workbook = load_workbook(path, data_only=False)
    try:
        if "How to Use" in workbook.sheetnames:
            workbook.remove(workbook["How to Use"])
        workbook["Recent Movement Signals"].title = "Rising & Falling Stars"
        action_board = workbook["Action Board"]
        table = action_board.tables["ActionBoardTable"]
        min_col, min_row, _, max_row = metrics.range_boundaries(table.ref)
        assert min_col == 1
        first_data_row = min_row + 1
        for row in range(first_data_row, max_row + 1):
            if action_board.cell(row=row, column=4).value == "Review Needed":
                action_board.cell(row=row, column=4).value = "Open"
        action_board.delete_cols(
            len(metrics.LEGACY_ACTION_HEADERS_V1) + 1,
            len(metrics.ACTION_HEADERS) - len(metrics.LEGACY_ACTION_HEADERS_V1),
        )
        for column, header in enumerate(metrics.LEGACY_ACTION_HEADERS_V1, start=1):
            action_board.cell(row=min_row, column=column).value = header
        table.ref = (
            f"A{min_row}:"
            f"{metrics.get_column_letter(len(metrics.LEGACY_ACTION_HEADERS_V1))}{max_row}"
        )
        table.tableColumns = table.tableColumns[
            : len(metrics.LEGACY_ACTION_HEADERS_V1)
        ]
        for table_column, header in zip(
            table.tableColumns,
            metrics.LEGACY_ACTION_HEADERS_V1,
            strict=True,
        ):
            table_column.name = header
        if table.autoFilter is not None:
            table.autoFilter.ref = table.ref

        status_validation = next(
            item
            for item in action_board.data_validations.dataValidation
            if item.formula1 == f'"{",".join(metrics.ACTION_STATUS_CHOICES)}"'
        )
        owner_validation = next(
            item
            for item in action_board.data_validations.dataValidation
            if item.formula1 == f"={metrics.OWNER_ROSTER_DEFINED_NAME}"
        )
        status_validation.formula1 = (
            f'"{",".join(metrics.LEGACY_ACTION_STATUS_CHOICES)}"'
        )
        status_validation.sqref = f"D{first_data_row}:D{max_row}"
        owner_validation.sqref = f"E{first_data_row}:E{max_row}"
        action_board.data_validations.dataValidation = [
            status_validation,
            owner_validation,
        ]

        action_history = workbook["Action History"]
        history_table = action_history.tables.get("ActionHistoryTable")
        history_header_row = 4
        history_max_row = action_history.max_row
        if history_table is not None:
            _, history_header_row, _, history_max_row = metrics.range_boundaries(
                history_table.ref
            )
        action_history.delete_cols(
            len(metrics.LEGACY_ACTION_HEADERS_V1) + 1,
            len(metrics.ACTION_HEADERS) - len(metrics.LEGACY_ACTION_HEADERS_V1),
        )
        for column, header in enumerate(metrics.LEGACY_ACTION_HEADERS_V1, start=1):
            action_history.cell(row=history_header_row, column=column).value = header
        if history_table is not None:
            history_table.ref = (
                f"A{history_header_row}:"
                f"{metrics.get_column_letter(len(metrics.LEGACY_ACTION_HEADERS_V1))}"
                f"{history_max_row}"
            )
            history_table.tableColumns = history_table.tableColumns[
                : len(metrics.LEGACY_ACTION_HEADERS_V1)
            ]
            for table_column, header in zip(
                history_table.tableColumns,
                metrics.LEGACY_ACTION_HEADERS_V1,
                strict=True,
            ):
                table_column.name = header
            if history_table.autoFilter is not None:
                history_table.autoFilter.ref = history_table.ref
        workbook.save(path)
    finally:
        workbook.close()
    return metrics.stamp_generated_content_digest(path)


def downgrade_master_to_pre_guide_v031(path: Path) -> str:
    """Model the protected v0.3.1 workbook immediately before the guide."""

    downgrade_master_to_pre_consolidation_layout(path)
    workbook = load_workbook(path, data_only=False)
    try:
        workbook.remove(workbook["How to Use"])
        workbook.active = workbook.sheetnames.index("Dashboard")
        workbook.save(path)
    finally:
        workbook.close()
    return metrics.stamp_generated_content_digest(path)


def downgrade_master_to_previous_action_schema(path: Path) -> str:
    """Model the protected pre-Action-Focus master currently deployed in production."""

    downgrade_master_to_v1_action_focus(path)
    workbook = load_workbook(path, data_only=False)
    try:
        workbook.remove(workbook["Action Focus"])
        workbook.remove(workbook["Evidence Detail"])
        workbook.save(path)
    finally:
        workbook.close()
    return metrics.stamp_generated_content_digest(path)


def test_explicit_baseline_succeeds_without_active_input_and_is_idempotent(
    tmp_path: Path,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    archive_dir = Path(args.archive_dir)

    first = metrics.run(args)

    assert len(first) == 1
    baseline = first[0]
    assert baseline.parent == metrics.integrity_manifest_dir(archive_dir)
    payload = integrity.read_json_manifest(baseline)
    assert payload["schema_version"] == 1
    assert payload["kind"] == "integrity-baseline"
    assert payload["raw_inventory"] == []
    assert payload["derived_archive_inventory"] == []
    assert payload["published_output_inventory"] == []
    assert payload["master_generated_content_sha256"] is None
    assert "previous_manifest" not in payload
    assert [entry.path for entry in integrity.verify_manifest_chain(baseline, baseline.parent)] == [
        baseline.resolve()
    ]
    anchor = metrics.integrity_anchor_path(archive_dir)
    assert anchor.is_file()
    anchor_payload = integrity.read_json_manifest(anchor, root=anchor.parent)
    assert anchor_payload["manifest_path"] == baseline.name
    assert anchor_payload["manifest_sha256"] == integrity.canonical_json_sha256(payload)
    assert not Path(args.input_dir).exists()

    second = metrics.run(args)

    assert second == [baseline]
    assert manifest_paths(archive_dir) == [baseline]
    assert integrity.read_json_manifest(anchor, root=anchor.parent) == anchor_payload


def test_replacement_path_rebind_verifies_backed_up_head(
    tmp_path: Path,
) -> None:
    original_root = tmp_path / "original-operator-root"
    original_args = workflow_args(
        original_root, initialize_baseline=True
    )
    [original_manifest] = metrics.run(original_args)
    original_archive = Path(original_args.archive_dir)
    source_anchor = metrics.integrity_anchor_path(original_archive)
    source_payload = integrity.read_json_manifest(
        source_anchor, root=source_anchor.parent
    )

    replacement_root = tmp_path / "replacement-operator-root"
    shutil.copytree(original_root, replacement_root)
    replacement_archive = replacement_root / "03 Archive"
    replacement_output = replacement_root / "02 Finished Reports"
    replacement_anchor_dir = tmp_path / "replacement-machine-anchors"

    rebound, receipt = metrics.rebind_restored_integrity_anchor(
        replacement_archive,
        replacement_output,
        source_anchor,
        replacement_anchor_dir,
    )

    replacement_manifest = (
        metrics.integrity_manifest_dir(replacement_archive)
        / original_manifest.name
    )
    assert metrics.verify_integrity_anchor(
        replacement_archive, replacement_anchor_dir
    ) == (
        replacement_manifest.resolve(),
        source_payload["manifest_sha256"],
    )
    rebound_payload = integrity.read_json_manifest(
        rebound, root=rebound.parent
    )
    assert (
        rebound_payload["archive_identity_sha256"]
        != source_payload["archive_identity_sha256"]
    )
    receipt_payload = integrity.read_json_manifest(
        receipt, root=receipt.parent
    )
    assert receipt_payload["contract"] == (
        "IntegrityAnchorRestoreRebindReceiptV1"
    )
    assert receipt_payload["source_anchor_sha256"] == metrics.sha256_file(
        source_anchor
    )
    assert receipt_payload["manifest_sha256"] == source_payload[
        "manifest_sha256"
    ]


def test_rewritten_manifest_head_and_raw_data_cannot_replace_trusted_history(
    tmp_path: Path,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    raw_root = metrics.canonical_daily_archive_dir(archive_dir)
    raw = raw_root / "week-ending-2026-07-19" / "Daily Report evidence.xlsx"
    raw.parent.mkdir(parents=True)
    raw.write_bytes(b"original raw evidence")
    [baseline] = metrics.run(args)
    anchor = metrics.integrity_anchor_path(archive_dir)
    pinned = integrity.read_json_manifest(anchor, root=anchor.parent)

    raw.write_bytes(b"rewritten raw evidence")
    rewritten = integrity.read_json_manifest(baseline)
    rewritten["raw_inventory"] = [integrity.fingerprint_file(raw_root, raw).to_dict()]
    integrity.write_json_manifest_atomic(
        baseline,
        rewritten,
        root=metrics.integrity_manifest_dir(archive_dir),
    )

    for allow_initialize in (False, True):
        with pytest.raises(
            integrity.IntegrityError,
            match="does not match the machine-local trusted anchor",
        ):
            metrics.ensure_integrity_preflight(
                archive_dir,
                output_dir,
                Path(args.config),
                metrics.load_config(Path(args.config)),
                allow_initialize=allow_initialize,
            )

    assert raw.read_bytes() == b"rewritten raw evidence"
    assert manifest_paths(archive_dir) == [baseline]
    assert integrity.read_json_manifest(anchor, root=anchor.parent) == pinned


def test_deleted_established_chain_cannot_be_explicitly_reinitialized(
    tmp_path: Path,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    archive_dir = Path(args.archive_dir)
    [baseline] = metrics.run(args)
    anchor = metrics.integrity_anchor_path(archive_dir)
    pinned = integrity.read_json_manifest(anchor, root=anchor.parent)
    baseline.unlink()

    with pytest.raises(
        integrity.IntegrityError,
        match="pinned by the machine-local integrity anchor is missing",
    ):
        metrics.run(args)

    assert manifest_paths(archive_dir) == []
    assert integrity.read_json_manifest(anchor, root=anchor.parent) == pinned


def test_existing_verified_chain_requires_explicit_one_time_anchor_adoption(
    tmp_path: Path,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    existing = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="integrity-baseline",
        run_id="pre-anchor-deployment",
        previous_manifest=None,
    )

    with pytest.raises(
        integrity.IntegrityError,
        match="no machine-local trusted-head anchor",
    ):
        metrics.ensure_integrity_preflight(
            archive_dir, output_dir, config_path, config
        )

    adopted, _, adopted_sha256 = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
        allow_initialize=True,
    )

    assert adopted == existing
    assert metrics.verify_integrity_anchor(archive_dir) == (
        existing.resolve(),
        adopted_sha256,
    )


def test_adopted_pre_contract_master_is_regenerated_with_strict_protection(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    downgrade_master_to_pre_consolidation_layout(master)
    workbook = load_workbook(master, data_only=False)
    try:
        roster = workbook["Management Setup"]
        roster["A21"] = "Avery Manager"
        roster["B21"] = "Yes"
        workbook["Recent Movement Signals"].title = "Rising & Falling Stars"
        workbook.save(master)
    finally:
        workbook.close()
    legacy_digest = downgrade_master_to_pre_protection_contract(master)
    genesis = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="integrity-baseline",
        run_id="pre-protection-genesis",
        previous_manifest=None,
    )
    genesis_sha256 = integrity.canonical_json_sha256(
        integrity.read_json_manifest(genesis)
    )
    existing = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="pre-protection-contract",
        previous_manifest=genesis,
        expected_previous_sha256=genesis_sha256,
    )
    assert integrity.read_json_manifest(existing)[
        "master_generated_content_sha256"
    ] == legacy_digest
    before_adoption = master.read_bytes()

    with pytest.raises(
        integrity.IntegrityError,
        match="predates the current protection contract",
    ):
        metrics.verify_existing_management_workbook_integrity(master)

    adopted, adopted_payload, adopted_sha256 = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
        allow_initialize=True,
    )

    assert adopted == existing
    assert metrics.verify_integrity_anchor(archive_dir) == (
        existing.resolve(),
        adopted_sha256,
    )
    assert adopted_payload["_legacy_master_upgrade_pending"] is True
    assert master.read_bytes() == before_adoption
    assert len(
        integrity.verify_manifest_chain(existing, metrics.integrity_manifest_dir(archive_dir))
    ) == 2

    args.migrate_history_from = [str(tmp_path / "legacy-history")]
    args.migrate_history_only = True
    pinned_anchor = integrity.read_json_manifest(
        metrics.integrity_anchor_path(archive_dir),
        root=metrics.integrity_anchor_path(archive_dir).parent,
    )
    with pytest.raises(
        integrity.IntegrityError,
        match="History-only migration is blocked",
    ):
        metrics.run(args)
    assert master.read_bytes() == before_adoption
    assert manifest_paths(archive_dir) == [genesis, existing]
    assert integrity.read_json_manifest(
        metrics.integrity_anchor_path(archive_dir),
        root=metrics.integrity_anchor_path(archive_dir).parent,
    ) == pinned_anchor
    args.migrate_history_from = []
    args.migrate_history_only = False

    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True)
    (input_dir / "Daily Report - TM - 07-21-2026.xlsx").write_bytes(
        b"captured active report"
    )
    install_synthetic_parser(monkeypatch)

    generated = metrics.run(args)

    assert master in generated
    assert metrics.validate_management_workbook(master)
    latest = metrics.latest_integrity_manifest_path(archive_dir)
    assert latest is not None and latest not in {genesis, existing}
    assert len(
        integrity.verify_manifest_chain(latest, metrics.integrity_manifest_dir(archive_dir))
    ) == 3
    assert metrics.verify_integrity_anchor(archive_dir)[0] == latest.resolve()
    upgraded = load_workbook(master, data_only=False)
    try:
        assert all(
            worksheet.protection.objects is True
            and worksheet.protection.scenarios is True
            for worksheet in upgraded.worksheets
        )
        assert any(
            upgraded["Run Notes"].cell(row=row, column=1).value
            == "Protection Contract"
            for row in range(1, upgraded["Run Notes"].max_row + 1)
        )
        assert metrics.owner_roster_from_sheet(upgraded["Management Setup"])[0] == {
            "Owner Name": "Avery Manager",
            "Active": "Yes",
        }
    finally:
        upgraded.close()


def test_manifest_pinned_v2_usability_master_is_regenerated_as_v3(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    workbook = load_workbook(master, data_only=False)
    try:
        workbook["Management Setup"]["A21"] = "Avery Manager"
        workbook["Management Setup"]["B21"] = "Yes"
        workbook.save(master)
    finally:
        workbook.close()
    v2_digest = downgrade_master_to_v2_usability_contract(master)
    existing = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="v2-usability-contract",
        previous_manifest=None,
    )
    assert integrity.read_json_manifest(existing)[
        "master_generated_content_sha256"
    ] == v2_digest
    before_upgrade = master.read_bytes()

    with pytest.raises(
        integrity.IntegrityError,
        match="predates the current usability contract",
    ):
        metrics.verify_existing_management_workbook_integrity(
            master,
            expected_digest=v2_digest,
        )

    adopted, adopted_payload, adopted_sha256 = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
        allow_initialize=True,
    )
    assert adopted == existing
    assert adopted_payload["_legacy_master_upgrade_pending"] is True
    assert metrics.verify_integrity_anchor(archive_dir) == (
        existing.resolve(),
        adopted_sha256,
    )
    assert master.read_bytes() == before_upgrade

    _, anchored_payload, _ = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
    )
    assert anchored_payload["_legacy_master_upgrade_pending"] is True

    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True)
    (input_dir / "Daily Report - TM - 07-21-2026.xlsx").write_bytes(
        b"captured active report"
    )
    install_synthetic_parser(monkeypatch)
    generated = metrics.run(args)

    assert master in generated
    assert metrics.validate_management_workbook(master)
    upgraded = load_workbook(master, data_only=False)
    try:
        guide_text = "\n".join(
            str(cell.value)
            for cell in upgraded["How to Use"]._cells.values()
            if cell.value not in (None, "")
        )
        assert metrics.MANAGEMENT_METHODOLOGY_VERSION in guide_text
        assert metrics.PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION not in guide_text
        assert metrics.owner_roster_from_sheet(upgraded["Management Setup"])[0] == {
            "Owner Name": "Avery Manager",
            "Active": "Yes",
        }
    finally:
        upgraded.close()


def test_manifest_pinned_legacy_v3_navigation_is_regenerated_with_menu(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    workbook = load_workbook(master, data_only=False)
    try:
        workbook["Management Setup"]["A21"] = "Avery Manager"
        workbook["Management Setup"]["B21"] = "Yes"
        seeded_action = {
            "Action ID": "NAV-UPGRADE-STATE",
            "Entity Key": "server|rc richmond|alex server",
            "Priority": "Medium",
            "Status": "Open",
            "Owner": "Avery Manager",
            "Due Date": None,
            "Location": "RC Richmond",
            "Person / Area": "Alex Server",
            "Action": "Review operating context",
            "Signal": "Context Review",
            "Why It Matters": "Test protected editable-state carry-forward.",
            "Recommended Next Step": "Review with the manager",
            "Last Seen": ACTIVE_DAY,
            "Context Notes": "Preserve this protected manager note.",
            "Peer Comparison": "Reference Unavailable",
            "Recent Movement": "Not Evaluated",
            "First Seen": ACTIVE_DAY,
            "Weeks Open": 1,
            "Evidence Status": "Developing",
            "Signal State": "Current",
            "Review Disposition": "Pending Review",
            "Reviewed By": "",
            "Review Date": None,
        }
        metrics.write_action_tracking_sheet(
            workbook,
            "Action Board",
            [seeded_action],
            editable=True,
        )
        action_board = workbook["Action Board"]
        metrics.add_management_navigation(action_board)
        metrics.configure_management_print_layout(action_board)
        metrics.protect_worksheet(action_board, "test-pass")
        workbook._sheets = [
            *[
                workbook[name]
                for name in metrics.VISIBLE_MANAGEMENT_SHEETS
                if name in workbook.sheetnames
            ],
            *[
                worksheet
                for worksheet in workbook.worksheets
                if worksheet.title not in metrics.VISIBLE_MANAGEMENT_SHEETS
            ],
        ]
        workbook.save(master)
    finally:
        workbook.close()
    legacy_digest = downgrade_master_to_legacy_v3_navigation(master)
    existing = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="legacy-v3-navigation",
        previous_manifest=None,
    )
    assert integrity.read_json_manifest(existing)[
        "master_generated_content_sha256"
    ] == legacy_digest

    with pytest.raises(
        integrity.IntegrityError,
        match="predates the current navigation contract",
    ):
        metrics.verify_existing_management_workbook_integrity(
            master,
            expected_digest=legacy_digest,
        )

    adopted, adopted_payload, adopted_sha256 = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
        allow_initialize=True,
    )
    assert adopted == existing
    assert adopted_payload["_legacy_master_upgrade_pending"] is True
    assert metrics.verify_integrity_anchor(archive_dir) == (
        existing.resolve(),
        adopted_sha256,
    )

    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True)
    (input_dir / "Daily Report - TM - 07-21-2026.xlsx").write_bytes(
        b"captured active report"
    )
    install_synthetic_parser(monkeypatch)
    generated = metrics.run(args)

    assert master in generated
    assert metrics.validate_management_workbook(master)
    upgraded = load_workbook(master, data_only=False)
    try:
        assert metrics.owner_roster_from_sheet(
            upgraded["Management Setup"]
        )[0] == {
            "Owner Name": "Avery Manager",
            "Active": "Yes",
        }
        carried_actions = [
            *metrics.records_from_sheet(
                upgraded["Action Board"],
                "Action ID",
            ),
            *metrics.records_from_sheet(
                upgraded["Action History"],
                "Action ID",
            ),
        ]
        carried = next(
            row
            for row in carried_actions
            if row["Action ID"] == "NAV-UPGRADE-STATE"
        )
        assert (
            carried["Context Notes"]
            == "Preserve this protected manager note."
        )
        assert carried["Owner"] == "Avery Manager"
        for sheet_name in metrics.VISIBLE_MANAGEMENT_SHEETS:
            worksheet = upgraded[sheet_name]
            start_column, end_column = metrics.management_menu_bounds(worksheet)
            assert {
                str(merged)
                for merged in worksheet.merged_cells.ranges
                if merged.min_row <= 2 <= merged.max_row
            } == {
                (
                    f"{metrics.get_column_letter(start_column)}2:"
                    f"{metrics.get_column_letter(end_column)}2"
                )
            }
            assert (
                worksheet.cell(row=2, column=start_column).hyperlink.target
                == metrics.PREVIEW_MENU_TARGET
            )
    finally:
        upgraded.close()


def test_manifest_pinned_pre_consolidation_layout_requires_upgrade_and_regenerates(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    legacy_digest = downgrade_master_to_pre_consolidation_layout(master)

    legacy = load_workbook(master, data_only=False)
    try:
        assert [
            worksheet.title
            for worksheet in legacy.worksheets
            if worksheet.sheet_state == "visible"
        ] == metrics.PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
        assert "Team Trends" not in legacy.sheetnames
        assert all(
            name in legacy.sheetnames
            for name in (
                "Action Focus",
                "Server Scorecard",
                "Recent Movement Signals",
            )
        )
        metrics.require_management_menu_contract(
            legacy,
            visible_sheets=(
                metrics.PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
            ),
            navigation_links=metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
        )
    finally:
        legacy.close()

    existing = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="pre-consolidation-layout",
        previous_manifest=None,
    )
    assert integrity.read_json_manifest(existing)[
        "master_generated_content_sha256"
    ] == legacy_digest

    with pytest.raises(
        integrity.IntegrityError,
        match="predates the current action and trend layout",
    ):
        metrics.verify_existing_management_workbook_integrity(
            master,
            expected_digest=legacy_digest,
        )
    with pytest.raises(
        integrity.IntegrityError,
        match="predates the current action and trend layout",
    ):
        metrics.verify_existing_management_workbook_integrity(
            master,
            allow_legacy_protection_upgrade=True,
        )
    assert (
        metrics.verify_existing_management_workbook_integrity(
            master,
            expected_digest=legacy_digest,
            allow_legacy_protection_upgrade=True,
        )
        == legacy_digest
    )

    adopted, adopted_payload, adopted_sha256 = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
        allow_initialize=True,
    )
    assert adopted == existing
    assert adopted_payload["_legacy_master_upgrade_pending"] is True
    assert metrics.verify_integrity_anchor(archive_dir) == (
        existing.resolve(),
        adopted_sha256,
    )

    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True)
    (input_dir / "Daily Report - TM - 07-21-2026.xlsx").write_bytes(
        b"captured active report"
    )
    install_synthetic_parser(monkeypatch)

    generated = metrics.run(args)

    assert master in generated
    assert metrics.validate_management_workbook(master)
    upgraded = load_workbook(master, data_only=False)
    try:
        assert [
            worksheet.title
            for worksheet in upgraded.worksheets
            if worksheet.sheet_state == "visible"
        ] == metrics.VISIBLE_MANAGEMENT_SHEETS
        assert "Team Trends" in upgraded.sheetnames
        assert all(
            name not in upgraded.sheetnames
            for name in (
                "Action Focus",
                "Server Scorecard",
                "Recent Movement Signals",
            )
        )
    finally:
        upgraded.close()


def test_manifest_pinned_legacy_v3_navigation_rejects_style_tampering(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    master = tmp_path / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    downgrade_master_to_legacy_v3_navigation(master)
    workbook = load_workbook(master, data_only=False)
    try:
        dashboard = workbook["Dashboard"]
        current_column = next(
            column
            for column, (_, target) in zip(
                metrics.management_navigation_columns(
                    dashboard,
                    metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
                ),
                metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
                strict=True,
            )
            if target == "Dashboard"
        )
        dashboard.cell(row=2, column=current_column).font = Font(
            color="7A1E1E",
            bold=True,
            underline=None,
            size=9,
        )
        workbook.save(master)
    finally:
        workbook.close()
    tampered_digest = metrics.stamp_generated_content_digest(master)

    with pytest.raises(
        integrity.IntegrityError,
        match="navigation style does not match the contract",
    ):
        metrics.verify_existing_management_workbook_integrity(
            master,
            expected_digest=tampered_digest,
            allow_legacy_protection_upgrade=True,
        )


def test_legacy_v3_navigation_with_missing_sheet_fails_closed(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    master = tmp_path / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    legacy_digest = downgrade_master_to_legacy_v3_navigation(master)
    workbook = load_workbook(master, data_only=False)
    try:
        workbook.remove(workbook["Dashboard"])
        assert (
            metrics.workbook_uses_legacy_multi_link_navigation(workbook)
            is False
        )
        workbook.save(master)
    finally:
        workbook.close()

    with pytest.raises(
        integrity.IntegrityError,
        match="missing required sheets",
    ):
        metrics.verify_existing_management_workbook_integrity(
            master,
            expected_digest=legacy_digest,
            allow_legacy_protection_upgrade=True,
        )


def test_manifest_pinned_v2_usability_upgrade_rejects_navigation_tampering(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    downgrade_master_to_v2_usability_contract(master)
    workbook = load_workbook(master, data_only=False)
    try:
        dashboard = workbook["Dashboard"]
        current_column = next(
            column
            for column, (_, target) in zip(
                metrics.management_navigation_columns(
                    dashboard,
                    metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
                ),
                metrics.PRE_CONSOLIDATION_NAVIGATION_LINKS,
                strict=True,
            )
            if target == "Dashboard"
        )
        dashboard.cell(row=2, column=current_column).fill = PatternFill(
            "solid",
            fgColor="F2F2F2",
        )
        workbook.save(master)
    finally:
        workbook.close()
    metrics.stamp_generated_content_digest(master)
    metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="tampered-v2-usability-contract",
        previous_manifest=None,
    )

    with pytest.raises(
        integrity.IntegrityError,
        match="navigation style does not match the contract",
    ):
        metrics.ensure_integrity_preflight(
            archive_dir,
            output_dir,
            config_path,
            config,
            allow_initialize=True,
        )
    assert not metrics.integrity_anchor_exists(archive_dir)


def test_v2_usability_upgrade_combines_history_migration_and_rebuild_atomically(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    workbook = load_workbook(master, data_only=False)
    try:
        workbook["Management Setup"]["A21"] = "Avery Manager"
        workbook["Management Setup"]["B21"] = "Yes"
        workbook.save(master)
    finally:
        workbook.close()
    downgrade_master_to_v2_usability_contract(master)

    canonical_week_end = date(2026, 7, 19)
    canonical_week = (
        archive_dir
        / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
        / f"week-ending-{canonical_week_end.isoformat()}"
    )
    canonical_week.mkdir(parents=True)
    staged_week_end = date(2026, 7, 12)
    migration_source = tmp_path / "approved-history"
    migration_source.mkdir()
    report_dates: dict[str, date] = {}
    for week_end, destination in (
        (canonical_week_end, canonical_week),
        (staged_week_end, migration_source),
    ):
        for offset in range(metrics.OPERATING_WEEK_DAYS):
            report_date = week_end - timedelta(
                days=metrics.OPERATING_WEEK_DAYS - 1 - offset
            )
            filename = f"Daily Report - TM - {report_date.isoformat()}.xls"
            (destination / filename).write_bytes(
                f"protected-history-{report_date.isoformat()}".encode()
            )
            report_dates[filename] = report_date

    def parse_complete_day(
        path: Path,
        _config: dict,
    ) -> list[metrics.MetricRecord]:
        report_date = report_dates[path.name]
        richmond = minimal_records(report_date, path.name)
        virginia_beach = [
            replace(record, location="RC Virginia Beach")
            for record in richmond
        ]
        return [*richmond, *virginia_beach]

    monkeypatch.setattr(metrics, "parse_daily_report", parse_complete_day)
    existing = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="v2-before-combined-rebuild",
        previous_manifest=None,
    )
    _, adopted_payload, _ = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
        allow_initialize=True,
    )
    assert adopted_payload["_legacy_master_upgrade_pending"] is True
    before_master = master.read_bytes()
    before_anchor = integrity.read_json_manifest(
        metrics.integrity_anchor_path(archive_dir),
        root=metrics.integrity_anchor_path(archive_dir).parent,
    )

    args.migrate_history_from = [str(migration_source)]
    args.migrate_history_only = True
    with pytest.raises(
        integrity.IntegrityError,
        match="History-only migration is blocked",
    ):
        metrics.run(args)
    assert master.read_bytes() == before_master
    assert manifest_paths(archive_dir) == [existing]
    assert integrity.read_json_manifest(
        metrics.integrity_anchor_path(archive_dir),
        root=metrics.integrity_anchor_path(archive_dir).parent,
    ) == before_anchor

    active_input = Path(args.input_dir) / "Daily Report active.xls"
    active_input.parent.mkdir(parents=True)
    active_input.write_bytes(b"must-not-be-read-or-moved")
    args.migrate_history_only = False
    args.rebuild_from_history = True
    generated = metrics.run(args)

    assert master in generated
    assert active_input.read_bytes() == b"must-not-be-read-or-moved"
    assert len(
        list(
            (
                archive_dir / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
            ).rglob("*.xls")
        )
    ) == 12
    latest = metrics.latest_integrity_manifest_path(archive_dir)
    assert latest is not None and latest != existing
    latest_payload = integrity.read_json_manifest(latest)
    assert latest_payload["kind"] == "history-rebuild"
    assert latest_payload["details"]["rebuild_from_history"] is True
    assert len(
        integrity.verify_manifest_chain(
            latest,
            metrics.integrity_manifest_dir(archive_dir),
        )
    ) == 2
    assert metrics.verify_integrity_anchor(archive_dir)[0] == latest.resolve()
    assert metrics.validate_management_workbook(master)
    upgraded = load_workbook(master, data_only=False)
    try:
        assert metrics.owner_roster_from_sheet(upgraded["Management Setup"])[0] == {
            "Owner Name": "Avery Manager",
            "Active": "Yes",
        }
    finally:
        upgraded.close()


def test_protected_v1_action_focus_master_accepts_only_legacy_status_contract(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    master = tmp_path / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    legacy_digest = downgrade_master_to_v1_action_focus(master)

    assert metrics.verify_existing_management_workbook_integrity(
        master,
        expected_digest=legacy_digest,
    ) == legacy_digest

    workbook = load_workbook(master, data_only=False)
    try:
        formulas = {
            validation.formula1
            for validation in workbook[
                "Action Board"
            ].data_validations.dataValidation
        }
        assert (
            f'"{",".join(metrics.LEGACY_ACTION_STATUS_CHOICES)}"'
            in formulas
        )
        assert f'"{",".join(metrics.ACTION_STATUS_CHOICES)}"' not in formulas
    finally:
        workbook.close()


def test_protected_v031_pre_guide_master_uses_exact_compatibility_contract(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    master = tmp_path / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    pre_guide_digest = downgrade_master_to_pre_guide_v031(master)

    assert metrics.validate_pre_guide_management_workbook(
        master, pre_guide_digest
    ) == pre_guide_digest
    assert metrics.verify_existing_management_workbook_integrity(
        master,
        expected_digest=pre_guide_digest,
    ) == pre_guide_digest


def test_evidence_source_accepts_manifest_pinned_pre_guide_v031_master(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_dir = tmp_path / "02 Finished Reports"
    archive_dir = tmp_path / "03 Archive"
    output_dir.mkdir()
    archive_dir.mkdir()
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    pre_guide_digest = downgrade_master_to_pre_guide_v031(master)
    workbook = load_workbook(master, data_only=False)
    try:
        expected_evidence = {
            str(row["Action ID"]): (
                row["Evidence Sources"],
                row["Metric Evidence"],
            )
            for row in metrics.records_from_sheet(
                workbook["Evidence Detail"], "Evidence ID"
            )
            if row.get("Action ID")
        }
    finally:
        workbook.close()
    manifest = archive_dir / "manifest.json"
    manifest.write_text("{}", encoding="utf-8")
    manifest_sha256 = "a" * 64
    manifest_payload = {
        "run_id": "pre-guide-run",
        "created_at_utc": "2026-07-23T12:00:00+00:00",
        "master_generated_content_sha256": pre_guide_digest,
        "provenance": {
            "git": {"commit": "b" * 40},
            "effective_config_sha256": "c" * 64,
        },
    }
    monkeypatch.setattr(
        metrics,
        "latest_integrity_manifest_path",
        lambda archive: manifest,
    )
    monkeypatch.setattr(
        metrics,
        "verify_integrity_anchor",
        lambda archive, anchor: (manifest.resolve(), manifest_sha256),
    )
    monkeypatch.setattr(
        metrics,
        "verify_integrity_state",
        lambda archive, output, selected_manifest: (
            manifest_payload,
            manifest_sha256,
        ),
    )

    source, rows, workbook_path = metrics.verified_evidence_source(
        Namespace(
            output_dir=str(output_dir),
            archive_dir=str(archive_dir),
            integrity_anchor_dir=str(tmp_path / "anchors"),
        )
    )

    assert workbook_path == master
    assert source["workbook_generated_content_sha256"] == pre_guide_digest
    actual_evidence = {
        str(row["Action ID"]): (
            row["Evidence Sources"],
            row["Metric Evidence"],
        )
        for row in rows
    }
    assert actual_evidence
    assert actual_evidence == {
        action_id: expected_evidence[action_id]
        for action_id in actual_evidence
    }


def test_v2_evidence_export_rejects_legacy_action_schema(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    master = tmp_path / "legacy-v1-action-focus.xlsx"
    shutil.copy2(valid_master_template, master)
    legacy_digest = downgrade_master_to_v1_action_focus(master)

    with pytest.raises(
        integrity.IntegrityError,
        match="legacy workbook evidence cannot be relabeled or promoted as V2",
    ):
        metrics.validate_v2_management_evidence_workbook(master, legacy_digest)


@pytest.mark.parametrize(
    ("tamper_kind", "message"),
    [
        ("missing", "missing required sheets"),
        ("hidden", "expected 'visible'"),
        ("reordered", "approved order"),
        ("unlocked", "editable-cell protection"),
        ("navigation_target", "invalid preview navigation target"),
        ("navigation_style", "preview navigation style"),
        ("menu_merge", "preview workbook-menu merge"),
        ("review_header", "missing required review columns"),
        ("visible_label", "label check_average as Sales / Guest"),
        ("quality_guidance", "missing required guidance"),
        ("title_merge", "title band must extend through column L"),
    ],
)
def test_current_workbook_rejects_redesign_usability_contract_tampering(
    tmp_path: Path,
    valid_master_template: Path,
    tamper_kind: str,
    message: str,
) -> None:
    master = tmp_path / f"tampered-{tamper_kind}.xlsx"
    shutil.copy2(valid_master_template, master)
    workbook = load_workbook(master, data_only=False)
    try:
        review = workbook["Weekly Review"]
        if tamper_kind == "missing":
            workbook.remove(review)
        elif tamper_kind == "hidden":
            review.sheet_state = "hidden"
        elif tamper_kind == "reordered":
            workbook._sheets = [
                workbook["Follow-up Queue"],
                *[
                    worksheet
                    for worksheet in workbook.worksheets
                    if worksheet.title != "Follow-up Queue"
                ],
            ]
        elif tamper_kind == "unlocked":
            review["A4"].protection = metrics.Protection(locked=False)
        elif tamper_kind == "navigation_target":
            review["A2"].hyperlink = "#'Follow-up Queue'!A1"
        elif tamper_kind == "navigation_style":
            review["A2"].fill = metrics.PatternFill("solid", fgColor="00FF00")
        elif tamper_kind == "menu_merge":
            review.unmerge_cells("A2:L2")
        elif tamper_kind == "review_header":
            header_row = next(
                (
                    row
                    for row in range(1, review.max_row + 1)
                    if review.cell(row=row, column=1).value == "Review Level"
                ),
                None,
            )
            assert header_row is not None
            review.cell(row=header_row, column=6, value="Metric")
        elif tamper_kind == "visible_label":
            review["A4"] = "Check Average is the primary metric."
        elif tamper_kind == "quality_guidance":
            quality = workbook["Data Quality & Audit"]
            for cell in quality._cells.values():
                if isinstance(cell.value, str) and (
                    "Sales / Guest is gross sales divided by guests" in cell.value
                ):
                    cell.value = "Metric guidance removed."
                    break
        else:
            review.unmerge_cells("A1:M1")
            review.merge_cells("A1:K1")
        workbook.save(master)
    finally:
        workbook.close()
    tampered_digest = metrics.stamp_generated_content_digest(master)

    with pytest.raises(integrity.IntegrityError, match=message):
        metrics.validate_management_workbook(master, tampered_digest)


@pytest.mark.parametrize(
    "tamper_kind",
    [
        "review_text",
        "navigation_target",
        "navigation_height",
        "menu_merge",
        "quality_text",
        "review_width",
    ],
)
def test_generated_content_digest_covers_workbook_usability_contract(
    tmp_path: Path,
    valid_master_template: Path,
    tamper_kind: str,
) -> None:
    master = tmp_path / f"digest-{tamper_kind}.xlsx"
    shutil.copy2(valid_master_template, master)
    original_digest = metrics.workbook_generated_content_sha256(master)
    workbook = load_workbook(master, data_only=False)
    try:
        if tamper_kind == "review_text":
            workbook["Weekly Review"]["A4"] = "Changed scope"
        elif tamper_kind == "navigation_target":
            workbook["Weekly Review"]["A2"].hyperlink = "#'Follow-up Queue'!A1"
        elif tamper_kind == "navigation_height":
            workbook["Follow-up Queue"].row_dimensions[2].height = 30
        elif tamper_kind == "menu_merge":
            workbook["Weekly Review"].unmerge_cells("A2:L2")
        elif tamper_kind == "quality_text":
            workbook["Data Quality & Audit"]["B14"] = "Changed audit guidance"
        else:
            workbook["Weekly Review"].column_dimensions["M"].width = 54
        workbook.save(master)
    finally:
        workbook.close()

    assert metrics.workbook_generated_content_sha256(master) != original_digest


def test_protected_previous_action_master_accepts_legacy_status_contract(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    master = tmp_path / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    legacy_digest = downgrade_master_to_previous_action_schema(master)

    assert metrics.verify_existing_management_workbook_integrity(
        master,
        expected_digest=legacy_digest,
    ) == legacy_digest


def test_current_action_board_rejects_legacy_status_validation(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    master = tmp_path / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    workbook = load_workbook(master, data_only=False)
    try:
        status_validation = next(
            item
            for item in workbook[
                "Action Board"
            ].data_validations.dataValidation
            if item.formula1 == f'"{",".join(metrics.ACTION_STATUS_CHOICES)}"'
        )
        status_validation.formula1 = (
            f'"{",".join(metrics.LEGACY_ACTION_STATUS_CHOICES)}"'
        )
        workbook.save(master)
    finally:
        workbook.close()

    with pytest.raises(
        integrity.IntegrityError,
        match="Action Board status list validation is missing or duplicated",
    ):
        metrics.validate_management_workbook(master)


def test_pre_contract_adoption_rejects_a_manifest_digest_mismatch(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    recorded_digest = downgrade_master_to_pre_protection_contract(master)

    workbook = load_workbook(master, data_only=False)
    try:
        workbook["Dashboard"]["A1"] = "Unrecorded generated heading"
        workbook.save(master)
    finally:
        workbook.close()
    state = metrics.build_integrity_state(archive_dir, output_dir)
    state["master_generated_content_sha256"] = recorded_digest
    metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="mismatched-pre-contract-master",
        previous_manifest=None,
        integrity_state=state,
    )

    with pytest.raises(
        integrity.IntegrityError,
        match="Master workbook generated-content verification failed",
    ):
        metrics.ensure_integrity_preflight(
            archive_dir,
            output_dir,
            config_path,
            config,
            allow_initialize=True,
        )

    assert not metrics.integrity_anchor_exists(archive_dir)


@pytest.mark.parametrize(
    ("tamper_kind", "message"),
    [
        ("protection", "pre-approved protection contract"),
        ("validation_range", "approved legacy shape"),
    ],
)
def test_pre_contract_adoption_rejects_a_self_stamped_near_legacy_shape(
    tmp_path: Path,
    valid_master_template: Path,
    tamper_kind: str,
    message: str,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    downgrade_master_to_pre_protection_contract(master)
    workbook = load_workbook(master, data_only=False)
    try:
        if tamper_kind == "protection":
            workbook["Dashboard"].protection.objects = True
        else:
            validation = next(
                item
                for item in workbook[
                    "Management Setup"
                ].data_validations.dataValidation
                if item.formula1 == '"Yes,No"'
            )
            validation.sqref = "Z1"
        workbook.save(master)
    finally:
        workbook.close()
    metrics.stamp_generated_content_digest(master)
    metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="near-legacy-protection-shape",
        previous_manifest=None,
    )

    with pytest.raises(
        integrity.IntegrityError,
        match=message,
    ):
        metrics.ensure_integrity_preflight(
            archive_dir,
            output_dir,
            config_path,
            config,
            allow_initialize=True,
        )

    assert not metrics.integrity_anchor_exists(archive_dir)


def test_explicit_adoption_rejects_an_unknown_protection_contract(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)
    workbook = load_workbook(master, data_only=False)
    try:
        run_notes = workbook["Run Notes"]
        marker_row = next(
            row
            for row in range(1, run_notes.max_row + 1)
            if run_notes.cell(row=row, column=1).value
            == metrics.WORKBOOK_PROTECTION_CONTRACT_LABEL
        )
        run_notes.cell(
            row=marker_row,
            column=2,
            value="objects-scenarios-stop-validation-v999",
        )
        workbook.save(master)
    finally:
        workbook.close()
    metrics.stamp_generated_content_digest(master)
    metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="unknown-protection-contract",
        previous_manifest=None,
    )

    with pytest.raises(
        integrity.IntegrityError,
        match="unsupported protection contract",
    ):
        metrics.ensure_integrity_preflight(
            archive_dir,
            output_dir,
            config_path,
            config,
            allow_initialize=True,
        )

    assert not metrics.integrity_anchor_exists(archive_dir)


def test_anchor_advance_failure_preserves_prior_trusted_head(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    archive_dir = Path(args.archive_dir)
    output_dir = Path(args.output_dir)
    config_path = Path(args.config)
    config = metrics.load_config(config_path)
    [baseline] = metrics.run(args)
    _, baseline_sha256 = metrics.verify_integrity_anchor(archive_dir)
    anchor = metrics.integrity_anchor_path(archive_dir)
    pinned = integrity.read_json_manifest(anchor, root=anchor.parent)
    successor = metrics.write_integrity_manifest(
        archive_dir=archive_dir,
        output_dir=output_dir,
        config_path=config_path,
        config=config,
        kind="weekly-run",
        run_id="failed-anchor-advance",
        previous_manifest=baseline,
        expected_previous_sha256=baseline_sha256,
    )
    _, successor_sha256 = metrics.verify_integrity_state(
        archive_dir, output_dir, successor
    )

    def fail_before_atomic_replace(*args, **kwargs):
        raise OSError("simulated local anchor write failure")

    monkeypatch.setattr(
        metrics, "write_json_manifest_atomic", fail_before_atomic_replace
    )
    with pytest.raises(OSError, match="simulated local anchor write failure"):
        metrics.advance_integrity_anchor(
            archive_dir,
            baseline,
            baseline_sha256,
            successor,
            successor_sha256,
        )

    assert integrity.read_json_manifest(anchor, root=anchor.parent) == pinned


def test_successful_staged_run_chains_manifest_snapshots_outputs_and_deletes_source_last(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    archive_dir = Path(args.archive_dir)
    input_dir.mkdir(parents=True)
    source = input_dir / "Daily Report - TM - 07-21-2026.xlsx"
    source_bytes = b"unaltered-red-onion-source"
    source.write_bytes(source_bytes)
    install_synthetic_parser(monkeypatch)

    def write_valid_master(
        records: list[metrics.MetricRecord],
        output_path: Path,
        config: dict,
        source_dir: Path,
        public_start: date,
        public_end: date,
    ) -> Path:
        shutil.copy2(valid_master_template, output_path)
        return output_path

    monkeypatch.setattr(metrics, "write_master_workbook", write_valid_master)
    metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    original_delete = metrics.quarantine_and_delete_captured_inputs
    deletion_observation: dict[str, object] = {}

    def assert_delete_is_last(
        captures: list[metrics.CapturedActiveInput],
        copies: list[metrics.VerifiedArchiveCopy],
        managed_input_dir: Path,
        run_id: str,
    ) -> None:
        captures = list(captures)
        copies = list(copies)
        weekly_manifest = manifest_by_kind(archive_dir, "weekly-run")
        weekly_payload = integrity.read_json_manifest(weekly_manifest)
        snapshot = (
            metrics.generated_workbook_archive_dir(archive_dir)
            / f"week-ending-{ACTIVE_WEEK_END.isoformat()}"
            / weekly_payload["run_id"]
        )
        assert source.exists()
        assert copies == [
            metrics.VerifiedArchiveCopy(
                source=source.resolve(),
                destination=(
                    metrics.canonical_daily_archive_dir(archive_dir)
                    / f"week-ending-{ACTIVE_WEEK_END.isoformat()}"
                    / source.name
                ).resolve(),
                created=True,
                sha256=integrity.sha256_file(source),
            )
        ]
        assert copies[0].destination.read_bytes() == source_bytes
        assert (snapshot / "published" / "Red_Onion_Server_Master.xlsx").is_file()
        assert (output_dir / "Red_Onion_Server_Master.xlsx").is_file()
        assert len(list((snapshot / "published").glob("*.xlsx"))) == 2
        assert len(list(output_dir.glob("*.xlsx"))) == 2
        deletion_observation.update(
            source_existed=True,
            manifest=weekly_manifest,
            snapshot=snapshot,
        )
        assert captures[0].fingerprint.sha256 == copies[0].sha256
        original_delete(captures, copies, managed_input_dir, run_id)

    monkeypatch.setattr(
        metrics, "quarantine_and_delete_captured_inputs", assert_delete_is_last
    )

    generated = metrics.run(args)

    assert deletion_observation["source_existed"] is True
    assert not source.exists()
    assert {path.name for path in generated} == {
        f"Check_Wine_RVA{ACTIVE_WEEK_END:%m%d%y}.xlsx",
        "Red_Onion_Server_Master.xlsx",
    }
    raw_root = metrics.canonical_daily_archive_dir(archive_dir)
    archived = raw_root / f"week-ending-{ACTIVE_WEEK_END.isoformat()}" / source.name
    assert archived.read_bytes() == source_bytes
    expected_raw = integrity.fingerprint_file(raw_root, archived).to_dict()

    baseline = manifest_by_kind(archive_dir, "integrity-baseline")
    weekly = manifest_by_kind(archive_dir, "weekly-run")
    chain = integrity.verify_manifest_chain(weekly, weekly.parent)
    assert [entry.path for entry in chain] == [weekly.resolve(), baseline.resolve()]
    trusted_manifest, trusted_sha256 = metrics.verify_integrity_anchor(archive_dir)
    assert trusted_manifest == weekly.resolve()
    assert trusted_sha256 == chain[0].sha256
    weekly_payload = dict(chain[0].payload)
    assert weekly_payload["raw_inventory"] == [expected_raw]
    assert weekly_payload["details"]["archived_destinations"] == [
        f"week-ending-{ACTIVE_WEEK_END.isoformat()}/{source.name}"
    ]
    assert weekly_payload["details"]["active_input_inventory"] == [
        {
            "path": source.name,
            "size": len(source_bytes),
            "sha256": integrity.sha256_file(archived),
        }
    ]
    snapshot = Path(deletion_observation["snapshot"])
    expected_derived = integrity.build_raw_inventory(
        metrics.generated_workbook_archive_dir(archive_dir)
    )
    assert weekly_payload["derived_archive_inventory"] == [
        item.to_dict() for item in expected_derived
    ]
    assert {path.name for path in (snapshot / "published").glob("*.xlsx")} == {
        path.name for path in generated
    }


def test_raw_archive_tampering_fails_preflight_before_writers_or_source_deletion(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    baseline_args = workflow_args(tmp_path, initialize_baseline=True)
    archive_dir = Path(baseline_args.archive_dir)
    raw = (
        metrics.canonical_daily_archive_dir(archive_dir)
        / "week-ending-2026-07-19"
        / "Daily Report archived.xlsx"
    )
    raw.parent.mkdir(parents=True)
    raw.write_bytes(b"original")
    metrics.run(baseline_args)
    baseline = manifest_by_kind(archive_dir, "integrity-baseline")
    raw.write_bytes(b"tampered")  # Same size; only the SHA-256 changes.

    args = workflow_args(tmp_path)
    active = Path(args.input_dir) / "Daily Report active.xlsx"
    active.parent.mkdir(parents=True)
    active_bytes = b"active-source"
    active.write_bytes(active_bytes)
    calls: list[str] = []

    def should_not_run(*args, **kwargs):
        calls.append("unexpected")
        raise AssertionError("Parsing, writing, and deletion must not run after preflight failure")

    monkeypatch.setattr(metrics, "parse_daily_report", should_not_run)
    monkeypatch.setattr(metrics, "write_public_workbook", should_not_run)
    monkeypatch.setattr(metrics, "write_master_workbook", should_not_run)
    monkeypatch.setattr(metrics, "delete_verified_active_sources", should_not_run)

    with pytest.raises(
        integrity.IntegrityError,
        match=(
            r"Canonical raw archive verification failed\.[\s\S]*"
            r"expected\(size=8, sha256=.*\); actual\(size=8, sha256=.*\)"
        ),
    ):
        metrics.run(args)

    assert calls == []
    assert active.read_bytes() == active_bytes
    assert raw.read_bytes() == b"tampered"
    assert manifest_paths(archive_dir) == [baseline]
    assert not metrics.generated_workbook_archive_dir(archive_dir).exists()
    assert not Path(args.output_dir).exists()


def test_master_preflight_allows_management_inputs_but_rejects_generated_tampering(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    output_dir = Path(args.output_dir)
    archive_dir = Path(args.archive_dir)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    shutil.copy2(valid_master_template, master)

    baseline_args = workflow_args(tmp_path, initialize_baseline=True)
    [baseline] = metrics.run(baseline_args)

    workbook = load_workbook(master, data_only=False)
    workbook["Management Setup"]["B6"] = 1234.0
    workbook["Management Setup"]["A21"] = "Pat Manager"
    workbook["Management Setup"]["B21"] = "Yes"
    workbook.active = workbook.sheetnames.index("Action Board")
    workbook.save(master)
    workbook.close()

    same_manifest, _, _ = metrics.ensure_integrity_preflight(
        archive_dir,
        output_dir,
        Path(args.config),
        metrics.load_config(Path(args.config)),
    )
    assert same_manifest == baseline
    assert manifest_paths(archive_dir) == [baseline]
    assert metrics.verify_existing_management_workbook_integrity(master)

    workbook = load_workbook(master, data_only=False)
    workbook["Dashboard"]["A1"] = "TAMPERED GENERATED TITLE"
    workbook.save(master)
    workbook.close()

    active = Path(args.input_dir) / "Daily Report active.xlsx"
    active.parent.mkdir(parents=True)
    active_bytes = b"active-source"
    active.write_bytes(active_bytes)
    calls: list[str] = []

    def should_not_run(*args, **kwargs):
        calls.append("unexpected")
        raise AssertionError("Parsing, writing, and deletion must not run after preflight failure")

    monkeypatch.setattr(metrics, "parse_daily_report", should_not_run)
    monkeypatch.setattr(metrics, "write_public_workbook", should_not_run)
    monkeypatch.setattr(metrics, "write_master_workbook", should_not_run)
    monkeypatch.setattr(metrics, "delete_verified_active_sources", should_not_run)

    with pytest.raises(
        integrity.IntegrityError,
        match="Master workbook generated-content verification failed",
    ):
        metrics.run(args)

    assert calls == []
    assert active.read_bytes() == active_bytes
    assert manifest_paths(archive_dir) == [baseline]
    assert not metrics.generated_workbook_archive_dir(archive_dir).exists()
    workbook = load_workbook(master, data_only=False)
    assert workbook["Management Setup"]["B6"].value == 1234.0
    assert workbook["Management Setup"]["A21"].value == "Pat Manager"
    workbook.close()


def test_writer_failure_preserves_prior_outputs_and_active_source_without_weekly_artifacts(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    output_dir = Path(args.output_dir)
    archive_dir = Path(args.archive_dir)
    output_dir.mkdir(parents=True)
    master = output_dir / "Red_Onion_Server_Master.xlsx"
    prior_public = output_dir / f"Check_Wine_RVA{ACTIVE_WEEK_END:%m%d%y}.xlsx"
    shutil.copy2(valid_master_template, master)
    prior_public.write_bytes(b"prior-public-output")
    prior_master_bytes = master.read_bytes()
    prior_public_bytes = prior_public.read_bytes()
    [baseline] = metrics.run(workflow_args(tmp_path, initialize_baseline=True))

    active = Path(args.input_dir) / "Daily Report active.xlsx"
    active.parent.mkdir(parents=True)
    active_bytes = b"active-source"
    active.write_bytes(active_bytes)
    install_synthetic_parser(monkeypatch)

    def write_staged_public(
        location: str,
        selected_records: list[metrics.MetricRecord],
        output_path: Path,
        config: dict,
        public_start: date,
        public_end: date,
    ) -> Path:
        staged = output_path / prior_public.name
        staged.write_bytes(b"new-staged-public")
        return staged

    def fail_master_writer(*args, **kwargs):
        raise RuntimeError("simulated master writer failure")

    deletion_calls: list[object] = []
    monkeypatch.setattr(metrics, "write_public_workbook", write_staged_public)
    monkeypatch.setattr(metrics, "write_master_workbook", fail_master_writer)
    monkeypatch.setattr(
        metrics,
        "delete_verified_active_sources",
        lambda copies: deletion_calls.append(list(copies)),
    )

    with pytest.raises(RuntimeError, match="simulated master writer failure"):
        metrics.run(args)

    assert deletion_calls == []
    assert active.read_bytes() == active_bytes
    assert master.read_bytes() == prior_master_bytes
    assert prior_public.read_bytes() == prior_public_bytes
    assert manifest_paths(archive_dir) == [baseline]
    assert not metrics.generated_workbook_archive_dir(archive_dir).exists()
    assert not metrics.canonical_daily_archive_dir(archive_dir).exists()
    assert list(output_dir.glob(".weekly-run-*")) == []
