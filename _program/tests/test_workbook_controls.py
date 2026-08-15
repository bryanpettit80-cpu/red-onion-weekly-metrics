from __future__ import annotations

from base64 import b64decode
from copy import deepcopy
from datetime import date
from pathlib import Path
import shutil
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, BarChart3D, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout
from openpyxl.chart.plotarea import DataTable
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, GradientFill, PatternFill, Protection
from openpyxl.utils import range_boundaries
from openpyxl.utils.datetime import CALENDAR_MAC_1904
from openpyxl.utils.protection import hash_password
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.scenario import InputCells, Scenario, ScenarioList
from openpyxl.worksheet.table import TableFormula
from openpyxl.worksheet.views import SheetView
from openpyxl.workbook.defined_name import DefinedName
import pytest

import red_onion_weekly_metrics as metrics


def action_row() -> dict[str, object]:
    row = {header: None for header in metrics.ACTION_HEADERS}
    row.update(
        {
            "Action ID": "A1B2C3D4E5F6",
            "Entity Key": "server|rc richmond|server one|coaching",
            "Priority": "High",
            "Status": "Open",
            "Owner": "Avery Manager",
            "Location": "RC Richmond",
            "Person / Area": "Server One",
            "Action Type": "Coaching",
            "Key Evidence": "Ticket time is above the benchmark.",
            "Recommended Next Step": "Review the next three shifts.",
            "First Seen": date(2026, 7, 12),
            "Last Seen": date(2026, 7, 19),
            "Active This Run": "Yes",
        }
    )
    return row


def build_controlled_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    roster = [
        {"Owner Name": "Avery Manager", "Active": "Yes"},
        {"Owner Name": "Former Manager", "Active": "No"},
    ]
    metrics.write_management_setup_sheet(
        workbook,
        {},
        roster,
        metrics.DEFAULT_CONFIG,
    )
    metrics.write_owner_validation_sheet(workbook, roster)
    metrics.write_action_tracking_sheet(
        workbook,
        "Action Board",
        [action_row()],
        editable=True,
    )
    run_notes = workbook.create_sheet("Run Notes")
    run_notes["A10"] = metrics.WORKBOOK_PROTECTION_CONTRACT_LABEL
    run_notes["B10"] = metrics.WORKBOOK_PROTECTION_CONTRACT
    run_notes["A11"] = metrics.RUN_NOTES_DIGEST_LABEL
    run_notes["B11"] = "Pending test digest"
    technical = workbook.create_sheet("_Technical")
    technical["A1"] = "Generated detail"
    technical["A2"] = 1
    technical["A3"] = 2
    technical.merge_cells("D1:E1")
    technical["D1"] = "Merged heading"
    technical.conditional_formatting.add(
        "A2:A3",
        CellIsRule(
            operator="greaterThan",
            formula=["1"],
            fill=PatternFill(fill_type="solid", fgColor="FFFF00"),
        ),
    )
    chart = BarChart()
    chart.title = "Generated chart"
    chart.add_data(Reference(technical, min_col=1, min_row=2, max_row=3))
    technical.add_chart(chart, "G1")
    metrics.finalize_management_workbook(workbook)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)
    workbook.close()


def unlocked_cells(worksheet) -> set[str]:
    return {
        cell.coordinate
        for row in worksheet.iter_rows()
        for cell in row
        if not cell.protection.locked
    }


def add_image_overlay(path: Path) -> None:
    """Add a valid OOXML picture anchor without requiring Pillow in the test runtime."""
    xdr_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    drawing_rel_ns = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    content_type_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    with ZipFile(path) as source:
        source_parts = {item.filename: source.read(item) for item in source.infolist()}
        drawing_part = next(
            name
            for name in source_parts
            if name.startswith("xl/drawings/drawing") and name.endswith(".xml")
        )
        relationship_part = (
            "xl/drawings/_rels/"
            f"{drawing_part.rsplit('/', 1)[-1]}.rels"
        )

    drawing = ElementTree.fromstring(source_parts[drawing_part])
    overlay = ElementTree.fromstring(
        f"""
        <xdr:oneCellAnchor
            xmlns:xdr="{xdr_ns}"
            xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
            xmlns:r="{drawing_rel_ns}">
          <xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
          <xdr:ext cx="7600000" cy="1500000" />
          <xdr:pic>
            <xdr:nvPicPr><xdr:cNvPr id="999" name="Dashboard Overlay" /><xdr:cNvPicPr /></xdr:nvPicPr>
            <xdr:blipFill><a:blip r:embed="rId999" /><a:stretch><a:fillRect /></a:stretch></xdr:blipFill>
            <xdr:spPr><a:prstGeom prst="rect"><a:avLst /></a:prstGeom></xdr:spPr>
          </xdr:pic>
          <xdr:clientData />
        </xdr:oneCellAnchor>
        """
    )
    drawing.append(overlay)
    source_parts[drawing_part] = ElementTree.tostring(
        drawing,
        encoding="utf-8",
        xml_declaration=False,
    )

    relationships = ElementTree.fromstring(source_parts[relationship_part])
    ElementTree.SubElement(
        relationships,
        f"{{{package_rel_ns}}}Relationship",
        {
            "Id": "rId999",
            "Type": f"{drawing_rel_ns}/image",
            "Target": "../media/dashboard-overlay.png",
        },
    )
    source_parts[relationship_part] = ElementTree.tostring(
        relationships,
        encoding="utf-8",
        xml_declaration=False,
    )

    content_types = ElementTree.fromstring(source_parts["[Content_Types].xml"])
    if not any(
        node.attrib.get("Extension", "").casefold() == "png"
        for node in content_types
    ):
        ElementTree.SubElement(
            content_types,
            f"{{{content_type_ns}}}Default",
            {"Extension": "png", "ContentType": "image/png"},
        )
    source_parts["[Content_Types].xml"] = ElementTree.tostring(
        content_types,
        encoding="utf-8",
        xml_declaration=False,
    )
    source_parts["xl/media/dashboard-overlay.png"] = b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR4nGP4z8DwHwAFgAI/"
        "sQ2eAAAAAElFTkSuQmCC"
    )

    replacement = path.with_name(f".{path.name}.overlay.tmp")
    with ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target:
        for name, content in source_parts.items():
            target.writestr(name, content)
    replacement.replace(path)


def mutate_first_xml_part(
    path: Path,
    prefix: str,
    mutation,
) -> None:
    with ZipFile(path) as source:
        source_parts = {item.filename: source.read(item) for item in source.infolist()}
    part = next(
        name for name in sorted(source_parts) if name.startswith(prefix) and name.endswith(".xml")
    )
    root = ElementTree.fromstring(source_parts[part])
    mutation(root)
    source_parts[part] = ElementTree.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    )
    replacement = path.with_name(f".{path.name}.xml-rewrite")
    with ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target:
        for name, content in source_parts.items():
            target.writestr(name, content)
    replacement.replace(path)


def add_legacy_vml_drawing_part(path: Path) -> None:
    with ZipFile(path) as source:
        source_parts = {item.filename: source.read(item) for item in source.infolist()}
    source_parts["xl/drawings/vmlDrawing999.vml"] = (
        b'<xml xmlns:v="urn:schemas-microsoft-com:vml"><v:shape id="overlay" /></xml>'
    )
    replacement = path.with_name(f".{path.name}.vml.tmp")
    with ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target:
        for name, content in source_parts.items():
            target.writestr(name, content)
    replacement.replace(path)


def add_external_non_xl_relationship(path: Path) -> None:
    with ZipFile(path) as source:
        source_parts = {item.filename: source.read(item) for item in source.infolist()}
    source_parts["customXml/_rels/item1.xml.rels"] = b"""<?xml version="1.0"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="urn:red-onion:test" Target="https://example.com/data" TargetMode="External" />
</Relationships>
"""
    replacement = path.with_name(f".{path.name}.external-rel.tmp")
    with ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target:
        for name, content in source_parts.items():
            target.writestr(name, content)
    replacement.replace(path)


def add_unsupported_active_part(path: Path) -> None:
    with ZipFile(path) as source:
        source_parts = {item.filename: source.read(item) for item in source.infolist()}
    source_parts["xl/connections.xml"] = b"""<?xml version="1.0"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" />
"""
    replacement = path.with_name(f".{path.name}.active-part.tmp")
    with ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target:
        for name, content in source_parts.items():
            target.writestr(name, content)
    replacement.replace(path)


def test_owner_roster_supports_more_than_twenty_people_and_never_shrinks(
    tmp_path: Path,
) -> None:
    owners = [
        {"Owner Name": f"Manager {index:02d}", "Active": "Yes"}
        for index in range(1, 26)
    ]
    workbook = Workbook()
    metrics.write_management_setup_sheet(
        workbook,
        {},
        owners,
        metrics.DEFAULT_CONFIG,
        roster_capacity=65,
    )
    setup = workbook["Management Setup"]
    assert setup.tables[metrics.OWNER_ROSTER_TABLE_NAME].ref == "A20:B85"
    assert [setup.cell(row=row, column=1).value for row in range(21, 46)] == [
        entry["Owner Name"] for entry in owners
    ]
    roster = metrics.owner_roster_from_sheet(setup)
    roster_capacity = metrics.owner_roster_capacity_from_sheet(setup)
    workbook.close()

    assert len(roster) == 25
    assert roster_capacity == 65

    regenerated = Workbook()
    metrics.write_management_setup_sheet(
        regenerated,
        {},
        roster[:2],
        metrics.DEFAULT_CONFIG,
        roster_capacity,
    )
    assert (
        regenerated["Management Setup"].tables[metrics.OWNER_ROSTER_TABLE_NAME].ref
        == "A20:B85"
    )
    regenerated.close()


def test_owner_roster_has_at_least_fifty_editable_rows() -> None:
    workbook = Workbook()
    metrics.write_management_setup_sheet(
        workbook,
        {},
        [{"Owner Name": "Avery Manager", "Active": "Yes"}],
        metrics.DEFAULT_CONFIG,
    )

    table = workbook["Management Setup"].tables[metrics.OWNER_ROSTER_TABLE_NAME]
    _, header_row, _, last_row = range_boundaries(table.ref)
    assert last_row - header_row == metrics.OWNER_ROSTER_MIN_EDIT_ROWS == 50
    workbook.close()


@pytest.mark.parametrize(
    ("owners", "message"),
    [
        (
            [
                {"Owner Name": "Pat Manager", "Active": "Yes"},
                {"Owner Name": " pat manager ", "Active": "No"},
            ],
            "duplicate name",
        ),
        (
            [{"Owner Name": "Pat Manager", "Active": "Maybe"}],
            "Active values must be Yes or No",
        ),
    ],
)
def test_owner_roster_rejects_duplicates_and_invalid_active_values(
    owners: list[dict[str, str]],
    message: str,
) -> None:
    workbook = Workbook()
    with pytest.raises(ValueError, match=message):
        metrics.write_management_setup_sheet(
            workbook,
            {},
            owners,
            metrics.DEFAULT_CONFIG,
        )
    workbook.close()


def test_active_owner_defined_name_and_action_validation_are_live() -> None:
    workbook = Workbook()
    roster = [
        {"Owner Name": "Avery Manager", "Active": "Yes"},
        {"Owner Name": "Former Manager", "Active": "No"},
    ]
    metrics.write_management_setup_sheet(
        workbook,
        {},
        roster,
        metrics.DEFAULT_CONFIG,
    )
    metrics.write_owner_validation_sheet(workbook, roster)
    metrics.write_action_tracking_sheet(
        workbook,
        "Action Board",
        [action_row()],
        editable=True,
    )

    assert metrics.active_owner_names(roster) == ["Avery Manager"]
    helper = workbook[metrics.OWNER_VALIDATION_SHEET]
    assert helper.sheet_state == "veryHidden"
    assert helper["A2"].value == (
        '=IF(\'Management Setup\'!$B$21="Yes",'
        '\'Management Setup\'!$A$21,"")'
    )
    assert helper["A3"].value == (
        '=IF(\'Management Setup\'!$B$22="Yes",'
        '\'Management Setup\'!$A$22,"")'
    )
    defined_name = workbook.defined_names[metrics.OWNER_ROSTER_DEFINED_NAME]
    assert defined_name.attr_text == "'_Validation Lists'!$A$2:$A$51"

    action_validations = workbook["Action Board"].data_validations.dataValidation
    status_validation = next(
        item
        for item in action_validations
        if item.formula1 == f'"{",".join(metrics.ACTION_STATUS_CHOICES)}"'
    )
    owner_validation = next(
        item
        for item in action_validations
        if item.formula1 == f"={metrics.OWNER_ROSTER_DEFINED_NAME}"
    )
    disposition_validation = next(
        item
        for item in action_validations
        if item.formula1
        == f'"{",".join(metrics.REVIEW_DISPOSITION_CHOICES)}"'
    )
    assert str(owner_validation.sqref) == "E5 V5"
    assert str(disposition_validation.sqref) == "U5"
    active_validation = workbook["Management Setup"].data_validations.dataValidation[0]
    assert active_validation.formula1 == '"Yes,No"'
    assert str(active_validation.sqref) == "B21:B70"
    for validation in (
        active_validation,
        status_validation,
        owner_validation,
        disposition_validation,
    ):
        assert validation.showErrorMessage is True
        assert validation.errorStyle == "stop"
        assert validation.errorTitle
        assert validation.error
    assert active_validation.allowBlank is True
    assert status_validation.allowBlank is False
    assert owner_validation.allowBlank is True
    assert disposition_validation.allowBlank is False
    workbook.close()


def test_workbook_protection_uses_exact_unlocked_cell_allowlist(
    tmp_path: Path,
) -> None:
    path = tmp_path / "protected.xlsx"
    build_controlled_workbook(path)

    workbook = load_workbook(path, data_only=False)
    target_rows = range(6, 9)
    expected_setup = {
        f"{column}{row}"
        for row in target_rows
        for column in "BCDEFG"
    }
    expected_setup.update(
        f"{column}{row}"
        for row in range(21, 71)
        for column in "AB"
    )
    assert unlocked_cells(workbook["Management Setup"]) == expected_setup
    assert unlocked_cells(workbook["Action Board"]) == {
        "D5",
        "E5",
        "F5",
        "N5",
        "U5",
        "V5",
        "W5",
    }
    assert unlocked_cells(workbook[metrics.OWNER_VALIDATION_SHEET]) == set()
    assert unlocked_cells(workbook["_Technical"]) == set()
    workbook.close()


def test_all_sheets_and_workbook_structure_are_password_protected(
    tmp_path: Path,
) -> None:
    path = tmp_path / "protected.xlsx"
    build_controlled_workbook(path)

    workbook = load_workbook(path, data_only=False)
    assert workbook.security.lockStructure is True
    assert workbook.security.workbookPassword == hash_password(
        metrics.WORKBOOK_OPERATOR_PASSWORD
    )
    assert workbook[metrics.OWNER_VALIDATION_SHEET].sheet_state == "veryHidden"
    assert workbook["_Technical"].sheet_state == "veryHidden"
    for worksheet in workbook.worksheets:
        assert worksheet.protection.sheet is True
        assert worksheet.protection.objects is True
        assert worksheet.protection.scenarios is True
        assert worksheet.protection.password == hash_password(
            metrics.WORKBOOK_OPERATOR_PASSWORD
        )
    workbook.close()


def test_strict_validation_rejects_a_restamped_validation_range_move(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "moved-validation-range.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    try:
        validation = next(
            item
            for item in workbook["Management Setup"].data_validations.dataValidation
            if item.formula1 == '"Yes,No"'
        )
        validation.sqref = "Z1"
        workbook.save(path)
    finally:
        workbook.close()
    tampered_digest = metrics.stamp_generated_content_digest(path)
    monkeypatch.setattr(
        metrics,
        "VISIBLE_MANAGEMENT_SHEETS",
        ["Management Setup", "Action Board", "Run Notes"],
    )

    with pytest.raises(
        metrics.IntegrityError,
        match="must block invalid pasted or typed values",
    ):
        metrics.validate_management_workbook(path, tampered_digest)


def test_strict_validation_rejects_a_duplicate_contract_marker(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "duplicate-contract-marker.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    try:
        workbook["Run Notes"].append(
            [metrics.WORKBOOK_PROTECTION_CONTRACT_LABEL, "attacker-contract"]
        )
        workbook.save(path)
    finally:
        workbook.close()
    tampered_digest = metrics.stamp_generated_content_digest(path)
    monkeypatch.setattr(
        metrics,
        "VISIBLE_MANAGEMENT_SHEETS",
        ["Management Setup", "Action Board", "Run Notes"],
    )

    with pytest.raises(metrics.IntegrityError, match="duplicated protection contract"):
        metrics.validate_management_workbook(path, tampered_digest)


def test_image_overlay_is_rejected_before_semantic_digest(tmp_path: Path) -> None:
    path = tmp_path / "image-overlay.xlsx"
    build_controlled_workbook(path)

    assert metrics.workbook_generated_content_sha256(path)
    add_image_overlay(path)

    with pytest.raises(
        metrics.IntegrityError,
        match="Images and embedded objects are not allowed",
    ):
        metrics.workbook_generated_content_sha256(path)


def test_legacy_vml_drawing_is_rejected_before_semantic_digest(
    tmp_path: Path,
) -> None:
    path = tmp_path / "legacy-vml-overlay.xlsx"
    build_controlled_workbook(path)
    add_legacy_vml_drawing_part(path)

    with pytest.raises(metrics.IntegrityError, match="Legacy VML"):
        metrics.workbook_generated_content_sha256(path)


@pytest.mark.parametrize(
    ("sheet_name", "coordinate", "invalid_value", "message"),
    [
        (
            "Management Setup",
            "B21",
            "Maybe",
            "Owner Roster Active values must be Yes or No",
        ),
        (
            "Action Board",
            "D5",
            "Maybe",
            "Action Board Status values must be one of",
        ),
    ],
)
def test_pasted_invalid_list_values_stop_before_workbook_generation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    sheet_name: str,
    coordinate: str,
    invalid_value: str,
    message: str,
) -> None:
    path = tmp_path / "invalid-pasted-value.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook[sheet_name][coordinate] = invalid_value
    workbook.save(path)
    workbook.close()

    monkeypatch.setattr(
        metrics,
        "verify_existing_management_workbook_integrity",
        lambda _path, **_kwargs: None,
    )

    def unexpected_generation(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("Workbook generation started before pasted values were validated")

    monkeypatch.setattr(metrics, "_write_master_workbook_base", unexpected_generation)
    with pytest.raises(ValueError, match=message):
        metrics.write_master_workbook(
            [],
            path,
            metrics.DEFAULT_CONFIG,
            tmp_path,
            date(2026, 7, 14),
            date(2026, 7, 19),
        )


def test_generated_content_digest_ignores_only_approved_management_edits(
    tmp_path: Path,
) -> None:
    path = tmp_path / "controlled.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    setup = workbook["Management Setup"]
    setup["B6"] = 1234.0
    setup["A21"] = "New Manager"
    setup["B21"] = "No"
    actions = workbook["Action Board"]
    actions["D5"] = "In Progress"
    actions["E5"] = "New Manager"
    actions["F5"] = date(2026, 7, 31)
    actions["N5"] = "Follow up Friday"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) == original_digest


@pytest.mark.parametrize(
    ("tamper_kind", "changes_substantive_digest"),
    [("style", True), ("internal_hyperlink", True)],
)
def test_editable_cells_exclude_only_scalar_value(
    tmp_path: Path,
    tamper_kind: str,
    changes_substantive_digest: bool,
) -> None:
    path = tmp_path / f"editable-metadata-{tamper_kind}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    cell = workbook["Management Setup"]["A21"]
    if tamper_kind == "style":
        cell.font = Font(name="Arial", bold=True, color="FFFF0000")
    else:
        cell.hyperlink = "#'Action Board'!A5"
    workbook.save(path)
    workbook.close()

    assert (
        metrics.workbook_generated_content_sha256(path) != original_digest
    ) is changes_substantive_digest


def test_editable_cell_formula_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "editable-formula.xlsx"
    build_controlled_workbook(path)

    workbook = load_workbook(path, data_only=False)
    workbook["Management Setup"]["B6"] = "=1+1"
    workbook.save(path)
    workbook.close()

    with pytest.raises(metrics.IntegrityError, match="Formulas are not allowed"):
        metrics.workbook_generated_content_sha256(path)


def test_external_hyperlink_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "external-hyperlink.xlsx"
    build_controlled_workbook(path)

    workbook = load_workbook(path, data_only=False)
    workbook["Management Setup"]["A21"].hyperlink = "https://example.com/unsafe"
    workbook.save(path)
    workbook.close()

    with pytest.raises(metrics.IntegrityError, match="External hyperlinks are not allowed"):
        metrics.workbook_generated_content_sha256(path)


def test_external_relationship_outside_xl_package_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "external-package-relationship.xlsx"
    build_controlled_workbook(path)
    add_external_non_xl_relationship(path)

    with pytest.raises(metrics.IntegrityError, match="External relationships"):
        metrics.workbook_generated_content_sha256(path)


def test_unsupported_active_workbook_part_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "active-workbook-part.xlsx"
    build_controlled_workbook(path)
    add_unsupported_active_part(path)

    with pytest.raises(metrics.IntegrityError, match="active workbook parts"):
        metrics.workbook_generated_content_sha256(path)


def test_internal_hyperlink_target_and_location_serializations_are_equivalent() -> None:
    target_form = Hyperlink(ref="A1", target="#'Action Board'!A5")
    location_form = Hyperlink(ref="A1", location="'Action Board'!A5")

    assert metrics.workbook_internal_hyperlink_destination(
        target_form,
        sheet_name="How to Use",
        coordinate="A1",
    ) == metrics.workbook_internal_hyperlink_destination(
        location_form,
        sheet_name="How to Use",
        coordinate="A1",
    )


def test_internal_hyperlink_tooltip_and_display_are_substantive(tmp_path: Path) -> None:
    path = tmp_path / "internal-hyperlink-display.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    cell = workbook["Management Setup"]["A21"]
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        target="#'Action Board'!A5",
        tooltip="Open action",
        display="Action link",
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["Management Setup"]["A21"].hyperlink.tooltip = "Changed tooltip"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_generated_error_and_text_cells_with_same_value_have_distinct_digest(
    tmp_path: Path,
) -> None:
    path = tmp_path / "generated-data-type.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A2"].value = "#N/A"
    workbook["_Technical"]["A2"].data_type = "s"
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A2"].data_type = "e"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_rich_text_cells_are_rejected_before_inline_format_can_be_hidden(
    tmp_path: Path,
) -> None:
    path = tmp_path / "rich-text.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A1"] = CellRichText(
        [
            "Revenue ",
            TextBlock(InlineFont(color="000000"), "123"),
        ]
    )
    workbook.save(path)
    workbook.close()

    with pytest.raises(metrics.IntegrityError, match="Rich-text cells are not allowed"):
        metrics.workbook_generated_content_sha256(path)


def test_external_formula_reference_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "external-formula.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A2"] = "='[Other.xlsx]Sheet1'!A1"
    workbook.save(path)
    workbook.close()

    with pytest.raises(metrics.IntegrityError, match="External formula references"):
        metrics.workbook_generated_content_sha256(path)


@pytest.mark.parametrize(
    ("tamper_kind", "changes_substantive_digest"),
    [
        ("cell_style", True),
        ("blank_cell_style", True),
        ("row_layout", True),
        ("column_layout", True),
        ("cell_number_format", True),
        ("generated_data_type", True),
        ("quote_prefix", True),
        ("pivot_button", True),
        ("chart", True),
        ("chart_series_color", True),
        ("chart_anchor", True),
        ("conditional_formatting", True),
        ("conditional_formatting_range", True),
        ("table_style", True),
        ("validation_prompt", True),
        ("validation_disable_prompts", True),
        ("sheet_zero_height", True),
        ("sheet_show_formulas", True),
        ("merged_cells", True),
        ("filter", True),
        ("protection_password", True),
    ],
)
def test_substantive_digest_separates_schema_from_presentation_metadata(
    tmp_path: Path,
    tamper_kind: str,
    changes_substantive_digest: bool,
) -> None:
    path = tmp_path / f"object-state-{tamper_kind}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    actions = workbook["Action Board"]
    technical = workbook["_Technical"]
    if tamper_kind == "cell_style":
        actions["G5"].font = Font(name="Arial", bold=True)
    elif tamper_kind == "blank_cell_style":
        technical["B10"].fill = PatternFill(fill_type="solid", fgColor="000000")
    elif tamper_kind == "cell_number_format":
        actions["G5"].number_format = ";;;"
    elif tamper_kind == "generated_data_type":
        technical["A2"].value = "#N/A"
        technical["A2"].data_type = "e"
    elif tamper_kind == "quote_prefix":
        technical["A2"].quotePrefix = True
    elif tamper_kind == "pivot_button":
        technical["A2"].pivotButton = True
    elif tamper_kind == "row_layout":
        actions.row_dimensions[5].height = 99
    elif tamper_kind == "column_layout":
        actions.column_dimensions["G"].width = 3
    elif tamper_kind == "chart":
        technical._charts[0].title = "Changed chart"
    elif tamper_kind == "chart_series_color":
        technical._charts[0].series[0].graphicalProperties.solidFill = "00FF00"
    elif tamper_kind == "chart_anchor":
        technical._charts[0].anchor = "ZZ1000"
    elif tamper_kind == "conditional_formatting":
        conditional_format = list(technical.conditional_formatting)[0]
        conditional_format.rules[0].formula = ["2"]
    elif tamper_kind == "conditional_formatting_range":
        conditional_format = list(technical.conditional_formatting)[0]
        [rule] = conditional_format.rules
        technical.conditional_formatting._cf_rules.pop(conditional_format)
        technical.conditional_formatting.add("A3", rule)
    elif tamper_kind == "table_style":
        table_style = actions.tables["ActionBoardTable"].tableStyleInfo
        table_style.showRowStripes = not bool(table_style.showRowStripes)
    elif tamper_kind == "validation_prompt":
        validation = actions.data_validations.dataValidation[0]
        validation.showInputMessage = True
        validation.promptTitle = "Changed prompt"
        validation.prompt = "Changed instructions"
    elif tamper_kind == "validation_disable_prompts":
        actions.data_validations.disablePrompts = True
    elif tamper_kind == "sheet_zero_height":
        technical.sheet_format.zeroHeight = True
    elif tamper_kind == "sheet_show_formulas":
        technical.sheet_view.showFormulas = True
    elif tamper_kind == "merged_cells":
        technical.unmerge_cells("D1:E1")
    elif tamper_kind == "filter":
        actions.auto_filter.ref = "A4:N5"
    else:
        actions.protection.set_password("different-password")
    workbook.save(path)
    workbook.close()

    assert (
        metrics.workbook_generated_content_sha256(path) != original_digest
    ) is changes_substantive_digest


def test_substantive_digest_keeps_auto_filter_criteria_with_same_range(
    tmp_path: Path,
) -> None:
    path = tmp_path / "auto-filter-criteria.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].auto_filter.ref = "A1:A3"
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].auto_filter.add_filter_column(0, ["Generated detail"])
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    ("plain", "excel_escaped"),
    [
        ("$#,##0.00", r"\$#,##0.00"),
        ("yyyy-mm-dd", r"yyyy\-mm\-dd"),
        ("$0.00", r"\$0.00"),
        ("+$#,##0.00;-$#,##0.00", r"\+\$#,##0.00;\-\$#,##0.00"),
        ("$#,##0", r"\$#,##0"),
    ],
)
def test_substantive_digest_normalizes_only_literal_number_format_escapes(
    tmp_path: Path,
    plain: str,
    excel_escaped: str,
) -> None:
    path = tmp_path / "number-format-escape.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A2"].number_format = plain
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A2"].number_format = excel_escaped
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) == original_digest


def test_substantive_digest_keeps_material_date_number_format_change(
    tmp_path: Path,
) -> None:
    path = tmp_path / "number-format-material.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A2"].number_format = "m/d/yyyy"
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A2"].number_format = "mm-dd-yy"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_explicit_row_height_changes(
    tmp_path: Path,
) -> None:
    path = tmp_path / "row-height-change.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].row_dimensions[2].height = 22.0
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].row_dimensions[2].height = 22.1
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_normalizes_empty_chart_layout(tmp_path: Path) -> None:
    path = tmp_path / "empty-chart-layout.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]._charts[0].layout = Layout()
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) == original_digest


def test_finalization_materializes_excel_row_height_serialization() -> None:
    workbook = Workbook()
    workbook.active.title = "Run Notes"
    workbook.create_sheet("_Technical")
    workbook["_Technical"].row_dimensions[2].height = 22.0

    metrics.finalize_management_workbook(workbook)

    assert workbook["_Technical"].row_dimensions[2].height == 21.95
    workbook.close()


def test_substantive_digest_keeps_one_pixel_column_width_change(
    tmp_path: Path,
) -> None:
    path = tmp_path / "column-width-change.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].column_dimensions["A"].width = 30.0
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].column_dimensions["A"].width = 30.08
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_normalizes_safe_excel_materialized_defaults(
    tmp_path: Path,
) -> None:
    path = tmp_path / "excel-materialized-defaults.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    technical["A1"].font = Font(name=None, bold=True, color="00112233")
    technical["A1"].hyperlink = Hyperlink(
        ref="A1", target="#'Action Board'!A5"
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    technical["A1"].font = Font(name="Calibri", bold=True, color="FF112233")
    technical["A1"].hyperlink.display = technical["A1"].value
    validation = next(
        item
        for item in workbook["Action Board"].data_validations.dataValidation
        if item.formula1 == f"={metrics.OWNER_ROSTER_DEFINED_NAME}"
    )
    validation.formula1 = metrics.OWNER_ROSTER_DEFINED_NAME
    validation.errorStyle = None
    table_style = workbook["Action Board"].tables["ActionBoardTable"].tableStyleInfo
    table_style.showFirstColumn = False
    table_style.showLastColumn = False
    table_style.showColumnStripes = False
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) == original_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "axis_delete",
        "vary_colors",
        "title_overlay",
        "legend_overlay",
        "axis_auto",
        "label_alignment",
        "flat_labels",
        "cross_between",
    ),
)
def test_substantive_digest_keeps_visible_chart_defaults(
    tmp_path: Path,
    mutation: str,
) -> None:
    path = tmp_path / f"chart-default-{mutation}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    chart = workbook["_Technical"]._charts[0]
    if mutation == "axis_delete":
        chart.x_axis.delete = True
    elif mutation == "vary_colors":
        chart.varyColors = True
    elif mutation == "title_overlay":
        chart.title.overlay = True
    elif mutation == "legend_overlay":
        chart.legend.overlay = True
    elif mutation == "axis_auto":
        chart.x_axis.auto = True
    elif mutation == "label_alignment":
        chart.x_axis.lblAlgn = "r"
    elif mutation == "flat_labels":
        chart.x_axis.noMultiLvlLbl = True
    else:
        chart.y_axis.crossBetween = "between"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_chart_data_labels(tmp_path: Path) -> None:
    path = tmp_path / "chart-data-labels.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]._charts[0].dLbls = DataLabelList(showVal=True)
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_chart_formula_normalization_preserves_dollar_in_sheet_name(
    tmp_path: Path,
) -> None:
    path = tmp_path / "chart-dollar-sheet-name.xlsx"
    workbook = Workbook()
    source_with_dollar = workbook.active
    source_with_dollar.title = "Cash$"
    source_without_dollar = workbook.create_sheet("Cash")
    host = workbook.create_sheet("Dashboard")
    for row_index, (with_dollar, without_dollar) in enumerate(
        ((1, 100), (2, 200)),
        start=1,
    ):
        source_with_dollar.cell(row=row_index, column=1, value=with_dollar)
        source_without_dollar.cell(row=row_index, column=1, value=without_dollar)
    chart = BarChart()
    chart.add_data(Reference(source_with_dollar, min_col=1, min_row=1, max_row=2))
    host.add_chart(chart, "A1")
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["Dashboard"]._charts[0].series[0].val.numRef.f = "'Cash'!$A$1:$A$2"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_stop_if_true_rule_priority(tmp_path: Path) -> None:
    path = tmp_path / "conditional-format-priority.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    first = next(iter(technical.conditional_formatting)).rules[0]
    first.stopIfTrue = True
    technical.conditional_formatting.add(
        "A2:A3",
        CellIsRule(
            operator="greaterThan",
            formula=["2"],
            stopIfTrue=True,
            fill=PatternFill(fill_type="solid", fgColor="FF0000"),
        ),
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    rules = next(iter(workbook["_Technical"].conditional_formatting)).rules
    rules[0].priority, rules[1].priority = rules[1].priority, rules[0].priority
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_table_formula_array_semantics(tmp_path: Path) -> None:
    path = tmp_path / "table-formula-array.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    table = workbook["Action Board"].tables["ActionBoardTable"]
    table.tableColumns[0].calculatedColumnFormula = TableFormula(
        array=False,
        attr_text="[@[Priority]]",
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    table = workbook["Action Board"].tables["ActionBoardTable"]
    table.tableColumns[0].calculatedColumnFormula.array = True
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    ("attribute", "value"),
    (
        ("headerRowCellStyle", "Good"),
        ("dataCellStyle", "Neutral"),
        ("totalsRowCellStyle", "Total"),
        ("insertRow", True),
    ),
)
def test_substantive_digest_keeps_table_level_semantics(
    tmp_path: Path,
    attribute: str,
    value: object,
) -> None:
    path = tmp_path / f"table-{attribute}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    table = workbook["Action Board"].tables["ActionBoardTable"]
    setattr(table, attribute, value)
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "freeze_panes",
        "grid_lines",
        "right_to_left",
        "row_col_headers",
        "zoom_scale",
        "view_mode",
        "custom_grid_color",
        "window_protection",
    ),
)
def test_substantive_digest_keeps_meaningful_sheet_view_settings(
    tmp_path: Path,
    mutation: str,
) -> None:
    path = tmp_path / f"sheet-view-{mutation}.xlsx"
    build_controlled_workbook(path)
    if mutation == "custom_grid_color":
        workbook = load_workbook(path, data_only=False)
        workbook["_Technical"].sheet_view.defaultGridColor = False
        workbook["_Technical"].sheet_view.colorId = 10
        workbook.save(path)
        workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    if mutation == "freeze_panes":
        technical.freeze_panes = "B3"
    elif mutation == "grid_lines":
        technical.sheet_view.showGridLines = False
    elif mutation == "right_to_left":
        technical.sheet_view.rightToLeft = True
    elif mutation == "row_col_headers":
        technical.sheet_view.showRowColHeaders = False
    elif mutation == "zoom_scale":
        technical.sheet_view.zoomScale = 40
    elif mutation == "view_mode":
        technical.sheet_view.view = "pageBreakPreview"
    elif mutation == "custom_grid_color":
        technical.sheet_view.colorId = 11
    else:
        technical.sheet_view.windowProtection = True
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_supports_gradient_fill_semantics(tmp_path: Path) -> None:
    path = tmp_path / "gradient-fill.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A1"].fill = GradientFill(
        type="linear",
        degree=45,
        stop=("FFFFFF", "000000"),
    )
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_excel_materialized_row_heights(tmp_path: Path) -> None:
    path = tmp_path / "excel-materialized-row-height.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook.create_sheet("Dashboard")
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["Dashboard"].row_dimensions[18].height = (
        metrics.EXCEL_AUTOFIT_ROW_HEIGHTS["Dashboard"][18]
    )
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_number_format_normalization_preserves_quoted_backslash(tmp_path: Path) -> None:
    path = tmp_path / "quoted-number-format.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A1"].number_format = r'0"\$"'
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A1"].number_format = '0"$"'
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_number_format_normalization_tracks_escaped_quote_literal(
    tmp_path: Path,
) -> None:
    path = tmp_path / "escaped-quote-number-format.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A1"].number_format = r'0"say \"hi \$"'
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["A1"].number_format = r'0"say \"hi $"'
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    "mutation",
    (
        "page_setup",
        "page_margin",
        "print_options",
        "header_footer",
        "row_break",
        "tab_color",
    ),
)
def test_substantive_digest_keeps_print_and_sheet_properties(
    tmp_path: Path,
    mutation: str,
) -> None:
    path = tmp_path / f"print-property-{mutation}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    if mutation == "page_setup":
        technical.page_setup.orientation = "landscape"
    elif mutation == "page_margin":
        technical.page_margins.left = 2.0
    elif mutation == "print_options":
        technical.print_options.gridLines = True
    elif mutation == "header_footer":
        technical.oddHeader.center.text = "Generated heading"
    elif mutation == "row_break":
        technical.row_breaks.append(Break(id=2))
    else:
        technical.sheet_properties.tabColor = "FF0000"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    "mutation",
    ("print_area", "title_rows", "title_columns"),
)
def test_substantive_digest_keeps_worksheet_print_scope(
    tmp_path: Path,
    mutation: str,
) -> None:
    path = tmp_path / f"print-scope-{mutation}.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    technical.print_area = "A1:A2"
    technical.print_title_rows = "1:1"
    technical.print_title_cols = "A:A"
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    if mutation == "print_area":
        technical.print_area = "A1:B3"
    elif mutation == "title_rows":
        technical.print_title_rows = "1:2"
    else:
        technical.print_title_cols = "A:B"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_workbook_view_and_epoch(tmp_path: Path) -> None:
    path = tmp_path / "workbook-view-epoch.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook.views[0].showSheetTabs = False
    workbook.save(path)
    workbook.close()
    view_digest = metrics.workbook_generated_content_sha256(path)
    assert view_digest != original_digest

    workbook = load_workbook(path, data_only=False)
    workbook.epoch = CALENDAR_MAC_1904
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != view_digest


def test_substantive_digest_keeps_multiple_sheet_views(tmp_path: Path) -> None:
    path = tmp_path / "multiple-sheet-views.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].views.sheetView.append(
        SheetView(workbookViewId=1, showGridLines=True)
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].views.sheetView[1].showGridLines = False
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_grouped_tab_selection(tmp_path: Path) -> None:
    path = tmp_path / "grouped-tab-selection.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["Action Board"].sheet_view.tabSelected = True
    workbook["_Technical"].sheet_view.tabSelected = True
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].sheet_view.tabSelected = False
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_sheet_local_defined_names(tmp_path: Path) -> None:
    path = tmp_path / "local-defined-name.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    technical["B3"] = "=LocalRate"
    technical.defined_names.add(
        DefinedName("LocalRate", attr_text="'_Technical'!$A$1")
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].defined_names["LocalRate"].attr_text = (
        "'_Technical'!$A$2"
    )
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_array_formula_scope(tmp_path: Path) -> None:
    path = tmp_path / "array-formula-scope.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["C1"] = ArrayFormula(
        ref="C1:C2",
        text="=A2:A3*2",
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["C1"].value.ref = "C1:C3"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_data_table_formula_digest_is_deterministic_and_complete(
    tmp_path: Path,
) -> None:
    path = tmp_path / "data-table-formula.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["C1"] = DataTableFormula(
        ref="C1:C2",
        r1="A2",
    )
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)
    assert metrics.workbook_generated_content_sha256(path) == original_digest

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]["C1"].value.r1 = "A3"
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_what_if_scenarios_are_rejected(tmp_path: Path) -> None:
    path = tmp_path / "what-if-scenario.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"].scenarios = ScenarioList(
        scenario=[
            Scenario(
                name="Tamper",
                inputCells=[InputCells(r="A2", val="999")],
            )
        ]
    )
    workbook.save(path)
    workbook.close()

    with pytest.raises(metrics.IntegrityError, match="What-if scenarios"):
        metrics.workbook_generated_content_sha256(path)


def test_substantive_digest_keeps_validation_prompt_position(tmp_path: Path) -> None:
    path = tmp_path / "validation-prompt-position.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["Action Board"].data_validations.xWindow = 10
    workbook["Action Board"].data_validations.yWindow = 20
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["Action Board"].data_validations.xWindow = 500
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_conditional_format_pivot_scope(
    tmp_path: Path,
) -> None:
    path = tmp_path / "conditional-format-pivot.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    conditional_format = next(iter(workbook["_Technical"].conditional_formatting))
    conditional_format.pivot = True
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_table_totals_row_default(tmp_path: Path) -> None:
    path = tmp_path / "table-totals-default.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["Action Board"].tables["ActionBoardTable"].totalsRowShown = False
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_chart_space_style(tmp_path: Path) -> None:
    path = tmp_path / "chart-space-style.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    def add_style(root) -> None:
        namespace = str(root.tag).split("}", 1)[0].lstrip("{")
        root.insert(0, ElementTree.Element(f"{{{namespace}}}style", {"val": "11"}))

    mutate_first_xml_part(path, "xl/charts/chart", add_style)

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_normalizes_excel_chart_space_defaults(
    tmp_path: Path,
) -> None:
    path = tmp_path / "chart-space-excel-defaults.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    def add_defaults(root) -> None:
        namespace = str(root.tag).split("}", 1)[0].lstrip("{")
        root.insert(
            0,
            ElementTree.Element(
                f"{{{namespace}}}roundedCorners", {"val": "1"}
            ),
        )
        root.insert(
            1,
            ElementTree.Element(f"{{{namespace}}}style", {"val": "2"}),
        )

    mutate_first_xml_part(path, "xl/charts/chart", add_defaults)

    assert metrics.workbook_generated_content_sha256(path) == original_digest


def test_substantive_digest_normalizes_duplicate_differential_styles(
    tmp_path: Path,
) -> None:
    path = tmp_path / "duplicate-differential-style.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    def duplicate_style(root) -> None:
        differentials = next(
            node for node in root if str(node.tag).endswith("}dxfs")
        )
        differentials.append(deepcopy(list(differentials)[0]))
        differentials.attrib["count"] = str(len(differentials))

    mutate_first_xml_part(path, "xl/styles.xml", duplicate_style)

    assert metrics.workbook_generated_content_sha256(path) == original_digest


def test_substantive_digest_keeps_chart_title_whitespace(tmp_path: Path) -> None:
    path = tmp_path / "chart-title-whitespace.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]._charts[0].title = " Generated chart "
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_chart_plot_data_table(tmp_path: Path) -> None:
    path = tmp_path / "chart-plot-data-table.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]._charts[0].plot_area.dTable = DataTable(
        showHorzBorder=True
    )
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_chart_3d_view(tmp_path: Path) -> None:
    path = tmp_path / "chart-3d-view.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    chart = BarChart3D()
    chart.add_data(Reference(technical, min_col=1, min_row=2, max_row=3))
    chart.view3D.rotX = 10
    technical.add_chart(chart, "G20")
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]._charts[1].view3D.rotX = 60
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_chart_anchor_client_data(tmp_path: Path) -> None:
    path = tmp_path / "chart-anchor-client-data.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]._charts[0].anchor.clientData.fPrintsWithSheet = False
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_substantive_digest_keeps_chart_frame_visibility(tmp_path: Path) -> None:
    path = tmp_path / "chart-frame-visibility.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    def hide_frame(root) -> None:
        nonvisual = next(
            node for node in root.iter() if str(node.tag).endswith("}cNvPr")
        )
        nonvisual.attrib["hidden"] = "1"

    mutate_first_xml_part(path, "xl/drawings/drawing", hide_frame)

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_raw_chart_digest_preserves_attribute_namespace(tmp_path: Path) -> None:
    path = tmp_path / "chart-attribute-namespace.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["_Technical"]._charts[0].title = " Revenue "
    workbook.save(path)
    workbook.close()
    original_digest = metrics.workbook_generated_content_sha256(path)

    def replace_space_namespace(root) -> None:
        xml_space = "{http://www.w3.org/XML/1998/namespace}space"
        text_node = next(node for node in root.iter() if xml_space in node.attrib)
        value = text_node.attrib.pop(xml_space)
        text_node.attrib["{urn:red-onion-test}space"] = value

    mutate_first_xml_part(path, "xl/charts/chart", replace_space_namespace)

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    "part_prefix",
    ("xl/worksheets/sheet", "xl/tables/table"),
)
def test_unsupported_worksheet_and_table_extensions_are_rejected(
    tmp_path: Path,
    part_prefix: str,
) -> None:
    path = tmp_path / f"unsupported-extension-{part_prefix.rsplit('/', 1)[-1]}.xlsx"
    build_controlled_workbook(path)

    def add_extension(root) -> None:
        namespace = str(root.tag).split("}", 1)[0].lstrip("{")
        extension_list = ElementTree.SubElement(root, f"{{{namespace}}}extLst")
        extension = ElementTree.SubElement(
            extension_list,
            f"{{{namespace}}}ext",
            {"uri": "{RED-ONION-UNSUPPORTED-TEST}"},
        )
        ElementTree.SubElement(extension, "{urn:red-onion-test}feature")

    mutate_first_xml_part(path, part_prefix, add_extension)

    with pytest.raises(metrics.IntegrityError, match="Unsupported"):
        metrics.workbook_generated_content_sha256(path)


@pytest.mark.parametrize(
    ("part_prefix", "element_name"),
    (
        ("xl/workbook.xml", "customWorkbookViews"),
        ("xl/workbook.xml", "fileRecoveryPr"),
        ("xl/workbook.xml", "smartTagTypes"),
        ("xl/workbook.xml", "webPublishingObjects"),
        ("xl/worksheets/sheet", "customSheetViews"),
        ("xl/worksheets/sheet", "ignoredErrors"),
        ("xl/worksheets/sheet", "webPublishItems"),
    ),
)
def test_unsupported_workbook_and_worksheet_features_are_rejected(
    tmp_path: Path,
    part_prefix: str,
    element_name: str,
) -> None:
    path = tmp_path / f"unsupported-{element_name}.xlsx"
    build_controlled_workbook(path)

    def add_feature(root) -> None:
        namespace = str(root.tag).split("}", 1)[0].lstrip("{")
        ElementTree.SubElement(root, f"{{{namespace}}}{element_name}")

    mutate_first_xml_part(path, part_prefix, add_feature)

    with pytest.raises(metrics.IntegrityError, match=element_name):
        metrics.workbook_generated_content_sha256(path)


def test_unsupported_namespaced_layout_attribute_is_rejected(
    tmp_path: Path,
) -> None:
    path = tmp_path / "unsupported-namespaced-layout-attribute.xlsx"
    build_controlled_workbook(path)

    def change_descent(root) -> None:
        namespace = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
        row = next(node for node in root.iter() if str(node.tag).endswith("}row"))
        row.attrib[f"{{{namespace}}}dyDescent"] = "0.50"

    mutate_first_xml_part(path, "xl/worksheets/sheet", change_descent)

    with pytest.raises(metrics.IntegrityError, match="namespaced XML attribute"):
        metrics.workbook_generated_content_sha256(path)


@pytest.mark.parametrize(
    ("part_prefix", "local_name", "replacement_namespace"),
    (
        (
            "xl/workbook.xml",
            "workbookPr",
            "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac",
        ),
        (
            "xl/styles.xml",
            "font",
            "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
        ),
    ),
)
def test_extension_namespaces_cannot_replace_core_elements(
    tmp_path: Path,
    part_prefix: str,
    local_name: str,
    replacement_namespace: str,
) -> None:
    path = tmp_path / f"wrong-namespace-{local_name}.xlsx"
    build_controlled_workbook(path)

    def replace_namespace(root) -> None:
        node = next(
            node
            for node in root.iter()
            if str(node.tag).rsplit("}", 1)[-1] == local_name
        )
        node.tag = f"{{{replacement_namespace}}}{local_name}"

    mutate_first_xml_part(path, part_prefix, replace_namespace)

    with pytest.raises(metrics.IntegrityError, match="Unsupported XML namespace"):
        metrics.workbook_generated_content_sha256(path)


def test_duplicate_conditional_format_priorities_are_rejected(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-conditional-format-priority.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    technical = workbook["_Technical"]
    technical.conditional_formatting.add(
        "A2:A3",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            stopIfTrue=True,
            fill=PatternFill(fill_type="solid", fgColor="FF0000"),
        ),
    )
    rules = next(iter(technical.conditional_formatting)).rules
    rules[0].priority = 1
    rules[1].priority = 1
    workbook.save(path)
    workbook.close()

    with pytest.raises(metrics.IntegrityError, match="unique priorities"):
        metrics.workbook_generated_content_sha256(path)


def test_substantive_digest_keeps_theme_color_semantics(tmp_path: Path) -> None:
    path = tmp_path / "theme-color.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    assert workbook.loaded_theme is not None
    theme_root = ElementTree.fromstring(workbook.loaded_theme)
    dark_one = next(
        node for node in theme_root.iter() if str(node.tag).endswith("}dk1")
    )
    [dark_color] = list(dark_one)
    original_color = dark_color.attrib.get("lastClr")
    dark_color.attrib["lastClr"] = (
        "FFFFFF" if original_color != "FFFFFF" else "000000"
    )
    workbook.loaded_theme = ElementTree.tostring(theme_root, encoding="utf-8")
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


def test_cell_comment_vml_drawing_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "comment-vml.xlsx"
    build_controlled_workbook(path)
    workbook = load_workbook(path, data_only=False)
    workbook["Action Board"]["G5"].comment = Comment(
        "Changed generated metadata", "Test"
    )
    workbook.save(path)
    workbook.close()

    with pytest.raises(metrics.IntegrityError, match="Legacy VML"):
        metrics.workbook_generated_content_sha256(path)


@pytest.mark.parametrize(
    "tamper_kind",
    [
        "generated_value",
        "generated_formula",
        "cell_lock",
        "sheet_protection",
        "workbook_structure",
        "defined_name",
    ],
)
def test_generated_content_digest_detects_generated_and_protection_tampering(
    tmp_path: Path,
    tamper_kind: str,
) -> None:
    path = tmp_path / f"controlled-{tamper_kind}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    if tamper_kind == "generated_value":
        workbook["Action Board"]["G5"] = "Tampered location"
    elif tamper_kind == "generated_formula":
        workbook[metrics.OWNER_VALIDATION_SHEET]["A2"] = "=1"
    elif tamper_kind == "cell_lock":
        workbook["Action Board"]["G5"].protection = Protection(locked=False)
    elif tamper_kind == "sheet_protection":
        workbook["Action Board"].protection.sheet = False
    elif tamper_kind == "workbook_structure":
        workbook.security.lockStructure = False
    else:
        workbook.defined_names[metrics.OWNER_ROSTER_DEFINED_NAME].attr_text = (
            "'_Validation Lists'!$A$2:$A$3"
        )
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    ("attribute", "drift_value", "unsafe"),
    [
        ("calcMode", None, False),
        ("fullCalcOnLoad", None, False),
        ("forceFullCalc", None, False),
        ("calcMode", "manual", True),
        ("fullCalcOnLoad", False, True),
        ("forceFullCalc", False, True),
    ],
)
def test_calculation_metadata_defaults_warn_but_explicit_unsafe_values_fail(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    attribute: str,
    drift_value: object,
    unsafe: bool,
) -> None:
    path = tmp_path / f"calculation-{attribute}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.stamp_generated_content_digest(path)

    workbook = load_workbook(path, data_only=False)
    setattr(workbook.calculation, attribute, drift_value)
    workbook.save(path)
    workbook.close()

    actual_digest = metrics.workbook_generated_content_sha256(path)
    assert (actual_digest != original_digest) is unsafe

    # This compact fixture deliberately contains only the two visible sheets
    # involved in workbook controls. Narrow the expected visible set so the
    # validation reaches the calculation guard being exercised here.
    monkeypatch.setattr(
        metrics,
        "VISIBLE_MANAGEMENT_SHEETS",
        ["Management Setup", "Action Board", "Run Notes"],
    )
    if unsafe:
        with pytest.raises(metrics.IntegrityError, match="unsafe calculation settings"):
            metrics.validate_management_workbook(path, original_digest)
    else:
        assert metrics.validate_management_workbook(path, original_digest) == original_digest


@pytest.mark.parametrize(
    ("attribute", "value"),
    [
        ("refMode", "R1C1"),
        ("iterate", True),
        ("iterateCount", 12),
        ("iterateDelta", 0.25),
        ("fullPrecision", False),
        ("calcOnSave", False),
        ("concurrentCalc", False),
        ("concurrentManualCount", 2),
    ],
)
def test_substantive_digest_keeps_other_calculation_semantics(
    tmp_path: Path,
    attribute: str,
    value: object,
) -> None:
    path = tmp_path / f"calculation-semantic-{attribute}.xlsx"
    build_controlled_workbook(path)
    original_digest = metrics.workbook_generated_content_sha256(path)

    workbook = load_workbook(path, data_only=False)
    setattr(workbook.calculation, attribute, value)
    workbook.save(path)
    workbook.close()

    assert metrics.workbook_generated_content_sha256(path) != original_digest


@pytest.mark.parametrize(
    "expected_scheme",
    [
        metrics.WORKBOOK_DIGEST_SCHEME,
        metrics.PREVIOUS_WORKBOOK_DIGEST_SCHEME,
        metrics.LEGACY_WORKBOOK_DIGEST_SCHEME,
    ],
)
def test_manifest_expected_digest_scheme_rejects_present_blank_stamp(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    expected_scheme: str,
) -> None:
    digest = "a" * 64
    monkeypatch.setattr(
        metrics, "workbook_generated_content_sha256", lambda _path: digest
    )
    monkeypatch.setattr(metrics, "workbook_metadata_sha256", lambda _path: digest)

    with pytest.raises(
        metrics.IntegrityError,
        match="digest scheme does not match the manifest-recorded scheme",
    ):
        metrics.verify_recorded_workbook_digest(
            tmp_path / "present-blank-scheme.xlsx",
            stamped_digest=digest,
            stamped_scheme="",
            expected_digest=digest,
            expected_scheme=expected_scheme,
        )


@pytest.mark.parametrize(
    ("stamped_scheme", "expected_scheme"),
    [
        (metrics.WORKBOOK_DIGEST_SCHEME, metrics.WORKBOOK_DIGEST_SCHEME),
        (
            metrics.LEGACY_WORKBOOK_DIGEST_SCHEME,
            metrics.LEGACY_WORKBOOK_DIGEST_SCHEME,
        ),
        (None, metrics.LEGACY_WORKBOOK_DIGEST_SCHEME),
    ],
)
def test_manifest_expected_digest_scheme_preserves_supported_stamps(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    stamped_scheme: str | None,
    expected_scheme: str,
) -> None:
    digest = "b" * 64
    monkeypatch.setattr(
        metrics, "workbook_generated_content_sha256", lambda _path: digest
    )
    monkeypatch.setattr(metrics, "workbook_metadata_sha256", lambda _path: digest)

    assert (
        metrics.verify_recorded_workbook_digest(
            tmp_path / "supported-scheme.xlsx",
            stamped_digest=digest,
            stamped_scheme=stamped_scheme,
            expected_digest=digest,
            expected_scheme=expected_scheme,
        )
        == digest
    )


def test_legacy_v2_metadata_drift_uses_pinned_reference_but_rejects_value_change(
    tmp_path: Path,
) -> None:
    reference = tmp_path / "reference.xlsx"
    live = tmp_path / "live.xlsx"
    build_controlled_workbook(reference)
    legacy_digest = metrics.workbook_metadata_sha256(reference)
    workbook = load_workbook(reference, data_only=False)
    workbook["Run Notes"]["B11"] = legacy_digest
    workbook.save(reference)
    workbook.close()
    assert metrics.workbook_metadata_sha256(reference) == legacy_digest
    shutil.copy2(reference, live)

    workbook = load_workbook(live, data_only=False)
    workbook.calculation.calcMode = None
    workbook.save(live)
    workbook.close()

    assert metrics.workbook_metadata_sha256(live) != legacy_digest
    with pytest.warns(UserWarning, match="metadata differs"):
        assert metrics.verify_recorded_workbook_digest(
            live,
            stamped_digest=legacy_digest,
            stamped_scheme=metrics.LEGACY_WORKBOOK_DIGEST_SCHEME,
            expected_digest=legacy_digest,
            expected_scheme=metrics.LEGACY_WORKBOOK_DIGEST_SCHEME,
            legacy_reference_path=reference,
        ) == legacy_digest

    workbook = load_workbook(live, data_only=False)
    workbook["Action Board"]["G5"] = "Changed generated location"
    workbook.save(live)
    workbook.close()
    with pytest.raises(metrics.IntegrityError, match="substantive-content verification"):
        metrics.verify_recorded_workbook_digest(
            live,
            stamped_digest=legacy_digest,
            stamped_scheme=metrics.LEGACY_WORKBOOK_DIGEST_SCHEME,
            expected_digest=legacy_digest,
            expected_scheme=metrics.LEGACY_WORKBOOK_DIGEST_SCHEME,
            legacy_reference_path=reference,
        )


def test_manifest_pinned_legacy_master_requires_one_exact_inventory_entry(
    tmp_path: Path,
) -> None:
    archive_dir = tmp_path / "03 Archive"
    generated_root = archive_dir / metrics.GENERATED_WORKBOOK_ARCHIVE_FOLDER
    run_id = "run-123"
    reference = (
        generated_root
        / "week-ending-2026-08-02"
        / run_id
        / "published"
        / "Red_Onion_Server_Master.xlsx"
    )
    reference.parent.mkdir(parents=True)
    build_controlled_workbook(reference)
    fingerprint = metrics.fingerprint_file(generated_root, reference).to_dict()
    payload = {
        "run_id": run_id,
        "derived_archive_inventory": [fingerprint],
    }

    assert metrics.manifest_pinned_master_snapshot(archive_dir, payload) == reference
    payload["derived_archive_inventory"].append(dict(fingerprint))
    with pytest.raises(metrics.IntegrityError, match="exactly one archived published master"):
        metrics.manifest_pinned_master_snapshot(archive_dir, payload)
