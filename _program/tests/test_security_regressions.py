from __future__ import annotations

from argparse import Namespace
from dataclasses import replace
from datetime import date, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula
import pytest

import red_onion_weekly_metrics as metrics


def initialize_integrity(args: Namespace) -> None:
    values = vars(args).copy()
    values.update(
        initialize_integrity_baseline=True,
        migrate_history_from=[],
        migrate_history_only=False,
    )
    metrics.run(Namespace(**values))


def make_record(
    day: date,
    *,
    source_file: str | None = None,
    raw_user_name: str = "Server One",
    display_name: str | None = None,
    gross_sales: float = 100.0,
) -> metrics.MetricRecord:
    display_name = raw_user_name if display_name is None else display_name
    return metrics.MetricRecord(
        source_file=source_file or f"Daily Report {day.isoformat()}.xlsx",
        report_date=day,
        location="RC Richmond",
        raw_user_name=raw_user_name,
        display_name=display_name,
        is_location_total=False,
        gross_sales=gross_sales,
        guest_count=10.0,
        check_average=gross_sales / 10.0,
        wine_sales=20.0,
        wine_pct=20.0 / gross_sales,
        rate_of_sale_by_guest_count=0.2,
        average_ticket_time_seconds=1200.0,
    )


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r"])
def test_excel_safe_text_neutralizes_formula_prefixes_idempotently(prefix: str) -> None:
    value = f"{prefix}SUM(1,1)"

    safe = metrics.excel_safe_text(value)

    assert safe == f"'{value}"
    assert metrics.excel_safe_text(safe) == safe
    assert metrics.excel_safe_text(123) == 123


def test_report_names_are_sanitized_before_workbook_generation(tmp_path: Path) -> None:
    malicious_name = '=HYPERLINK("https://example.invalid","click")'
    source_path = tmp_path / "Daily Report malicious.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report(All)"
    worksheet.append(["Date(s): 07/18/2026"])
    worksheet.append(
        [
            "Store",
            "User",
            "Gross Sales",
            "Guest Count",
            "Check Average",
            "Wine Sales",
            "Rate of Sale by Guest Count",
            "Average Ticket Time",
        ]
    )
    worksheet.append(
        ["RC Richmond", malicious_name, 100.0, 10.0, 10.0, 20.0, 0.2, time(0, 20)]
    )
    worksheet["B3"].data_type = "s"
    workbook.save(source_path)
    workbook.close()

    config = metrics.load_config(tmp_path / "missing-config.json")
    records = metrics.parse_daily_report(source_path, config)
    assert records[0].raw_user_name == malicious_name
    assert records[0].display_name == malicious_name

    output_dir = tmp_path / "output"
    output_dir.mkdir()
    output_path = metrics.write_public_workbook(
        "RC Richmond",
        records,
        output_dir,
        config,
        date(2026, 7, 18),
        date(2026, 7, 18),
    )
    generated = load_workbook(output_path, data_only=False)
    name_cell = generated["Report(All)"]["B3"]
    assert name_cell.value == f"'{malicious_name}"
    assert name_cell.data_type == "s"
    generated.close()


def test_formula_escaping_does_not_hide_distinct_report_conflicts(tmp_path: Path) -> None:
    config = metrics.load_config(tmp_path / "missing-config.json")
    formula_name = "=Alex"
    literal_name = "'=Alex"
    first_path = tmp_path / "Daily Report active.xlsx"
    second_path = tmp_path / "Daily Report archived.xlsx"

    for path, server_name in (
        (first_path, formula_name),
        (second_path, literal_name),
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report(All)"
        worksheet.append(["Date(s): 07/18/2026"])
        worksheet.append(
            [
                "Store",
                "User",
                "Gross Sales",
                "Guest Count",
                "Check Average",
                "Wine Sales",
                "Rate of Sale by Guest Count",
                "Average Ticket Time",
            ]
        )
        worksheet.append(
            ["RC Richmond", server_name, 100.0, 10.0, 10.0, 20.0, 0.2, time(0, 20)]
        )
        worksheet["B3"].data_type = "s"
        workbook.save(path)
        workbook.close()

    records_by_path = {
        first_path: metrics.parse_daily_report(first_path, config),
        second_path: metrics.parse_daily_report(second_path, config),
    }
    assert records_by_path[first_path][0].raw_user_name == formula_name
    assert records_by_path[second_path][0].raw_user_name == literal_name

    with pytest.raises(ValueError, match="Conflicting daily reports"):
        metrics.resolve_report_duplicates(records_by_path)


def test_generic_workbook_tables_escape_untrusted_identity_columns() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    worksheet = metrics.write_table_sheet(
        workbook,
        "Identity Detail",
        ["Raw Server", "Display Name", "Source File", "Trusted Note"],
        [["=Raw", "+Display", "@source.xlsx", "=trusted formula-like note"]],
        "IdentityDetail",
    )

    for coordinate in ("A4", "B4", "C4"):
        assert worksheet[coordinate].value.startswith("'")
        assert worksheet[coordinate].data_type == "s"
    assert worksheet["D4"].value == "=trusted formula-like note"
    assert worksheet["D4"].data_type == "f"
    workbook.close()


def test_data_quality_sheet_escapes_source_file_names() -> None:
    workbook = Workbook()
    report_day = date(2026, 7, 18)

    metrics.write_data_quality_sheet(
        workbook,
        [make_record(report_day, source_file="=source.xlsx")],
        [],
        report_day,
        report_day,
    )

    source_cells = [
        cell
        for row in workbook["Data Quality"].iter_rows()
        for cell in row
        if isinstance(cell.value, str) and "source.xlsx" in cell.value
    ]
    assert len(source_cells) == 2
    assert all(cell.value.startswith("'") for cell in source_cells)
    assert all(cell.data_type == "s" for cell in source_cells)
    workbook.close()


def test_carried_forward_management_text_is_sanitized(tmp_path: Path) -> None:
    malicious = "=HYPERLINK(\"https://example.invalid\",\"open\")"
    existing_path = tmp_path / "Red_Onion_Server_Master.xlsx"
    existing = Workbook()
    setup = existing.active
    setup.title = "Management Setup"
    setup["J5"] = "Owner List"
    setup["J6"] = malicious
    actions = existing.create_sheet("Action Board")
    for column, header in enumerate(metrics.ACTION_HEADERS, start=1):
        actions.cell(row=4, column=column, value=header)
    actions["A5"] = "A1B2C3D4E5F6"
    actions["B5"] = "server|rc richmond|server one|coaching"
    actions["D5"] = "Open"
    actions["E5"] = malicious
    actions["H5"] = malicious
    actions["N5"] = ArrayFormula("N5", malicious)
    existing.save(existing_path)
    existing.close()

    state = metrics.read_management_state(existing_path)
    assert state["owners"] == [f"'{malicious}"]
    assert state["owner_roster"] == [
        {"Owner Name": f"'{malicious}", "Active": "Yes"}
    ]
    assert state["active_actions"][0]["Owner"] == f"'{malicious}"
    assert state["active_actions"][0]["Person / Area"] == f"'{malicious}"
    assert state["active_actions"][0]["Context Notes"] == f"'{malicious}"

    regenerated = Workbook()
    metrics.write_management_setup_sheet(
        regenerated,
        {},
        state["owner_roster"],
        metrics.DEFAULT_CONFIG,
        state["owner_roster_capacity"],
    )
    metrics.write_action_tracking_sheet(
        regenerated, "Action Board", state["active_actions"], editable=True
    )
    setup = regenerated["Management Setup"]
    assert metrics.OWNER_ROSTER_TABLE_NAME in setup.tables
    assert setup["A21"].value == f"'{malicious}"
    assert setup["A21"].data_type == "s"
    assert setup["B21"].value == "Yes"
    for coordinate in ("E5", "H5", "N5"):
        assert regenerated["Action Board"][coordinate].data_type == "s"
    regenerated.close()


@pytest.mark.parametrize(
    ("week_start", "week_end"),
    [
        ("2026-07-19", "2026-07-14"),
        ("2026-07-14", "2026-07-21"),
        ("2026-07-13", "2026-07-18"),
    ],
)
def test_public_date_selection_rejects_non_weekly_ranges(
    week_start: str, week_end: str
) -> None:
    records = [make_record(date(2026, 7, 18))]

    with pytest.raises(ValueError, match="Tuesday-Sunday operating week"):
        metrics.selected_public_dates(records, week_start, week_end)


@pytest.mark.parametrize(
    ("week_start", "week_end"),
    [
        ("9999-12-31", None),
        (None, "0001-01-01"),
        ("9999-12-31", "9999-12-31"),
    ],
)
def test_public_date_selection_rejects_unrepresentable_week_bounds(
    week_start: str | None, week_end: str | None
) -> None:
    records = [make_record(date(2026, 7, 18))]

    with pytest.raises(ValueError, match="Tuesday-Sunday operating week"):
        metrics.selected_public_dates(records, week_start, week_end)


def test_one_sided_public_dates_stay_within_the_same_operating_week() -> None:
    records = [make_record(date(2026, 7, 18))]

    assert metrics.selected_public_dates(records, "2026-07-16", None) == (
        date(2026, 7, 16),
        date(2026, 7, 19),
    )
    assert metrics.selected_public_dates(records, None, "2026-07-17") == (
        date(2026, 7, 14),
        date(2026, 7, 17),
    )


def test_data_quality_coverage_rejects_extreme_ranges_without_iteration() -> None:
    records = [make_record(date.min), make_record(date.max)]

    with pytest.raises(ValueError, match="Data Quality date coverage"):
        metrics.validated_data_quality_coverage(
            records, date(2026, 7, 14), date(2026, 7, 19)
        )


def test_duplicate_signature_distinguishes_unavailable_metrics_from_real_zero() -> None:
    available = make_record(date(2026, 7, 18))
    unavailable = replace(
        available,
        rate_available=False,
        ticket_time_available=False,
    )

    assert metrics.semantic_report_signature([available]) != (
        metrics.semantic_report_signature([unavailable])
    )


def test_run_rejects_excessive_history_span_before_writes_or_moves(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "01 Daily Reports - Drop Here"
    archive_dir = tmp_path / "03 Archive"
    output_dir = tmp_path / "02 Finished Reports"
    input_dir.mkdir()
    migration_source = tmp_path / "legacy-history"
    migration_source.mkdir()
    historical_source = migration_source / "Daily Report historical.xlsx"
    active_source = input_dir / "Daily Report active.xlsx"
    historical_source.write_text("historical", encoding="utf-8")
    active_source.write_text("active", encoding="utf-8")
    active_day = date(9000, 1, 1)
    while not metrics.is_operating_day(active_day):
        active_day += metrics.timedelta(days=1)
    archived_day = date(1000, 1, 1)

    def fake_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        day = active_day if path.name == active_source.name else archived_day
        person = make_record(day, source_file=path.name)
        total = replace(
            person,
            raw_user_name="",
            display_name="",
            is_location_total=True,
        )
        return [total, person]

    def unexpected_writer(*args, **kwargs):
        pytest.fail("workbook writer must not run for an excessive date range")

    monkeypatch.setattr(metrics, "parse_daily_report", fake_parse)
    monkeypatch.setattr(metrics, "write_public_workbook", unexpected_writer)
    monkeypatch.setattr(metrics, "write_master_workbook", unexpected_writer)
    args = Namespace(
        input_dir=str(input_dir),
        output_dir=str(output_dir),
        archive_dir=str(archive_dir),
        config=str(tmp_path / "missing-config.json"),
        week_start=None,
        week_end=None,
        migrate_history_from=[str(migration_source)],
        migrate_history_only=False,
    )

    initialize_integrity(args)
    with pytest.raises(ValueError, match="Data Quality date coverage"):
        metrics.run(args)

    assert active_source.exists()
    assert historical_source.exists()
    assert not metrics.canonical_daily_archive_dir(archive_dir).exists()
    assert not output_dir.exists()


def test_active_and_archive_identical_reports_are_used_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "01 Daily Reports - Drop Here"
    archive_dir = tmp_path / "03 Archive"
    output_dir = tmp_path / "02 Finished Reports"
    input_dir.mkdir()
    archived_source = (
        archive_dir
        / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
        / "week-ending-2026-07-19"
        / "Daily Report archived.xlsx"
    )
    archived_source.parent.mkdir(parents=True)
    active_source = input_dir / "Daily Report active.xlsx"
    archived_source.write_text("archived copy", encoding="utf-8")
    active_source.write_text("active copy", encoding="utf-8")
    master_records: list[metrics.MetricRecord] = []
    real_master_writer = metrics.write_master_workbook

    def fake_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        person = make_record(date(2026, 7, 18), source_file=path.name)
        total = replace(
            person,
            raw_user_name="",
            display_name="",
            is_location_total=True,
        )
        return [total, person]

    def fake_public(location, records, output_path, config, start, end):
        output_path.mkdir(parents=True, exist_ok=True)
        result = output_path / "public.xlsx"
        result.write_text("public", encoding="utf-8")
        return result

    def fake_master(records, output_path, config, source_dir, start, end):
        master_records.extend(records)
        return real_master_writer(
            records, output_path, config, source_dir, start, end
        )

    monkeypatch.setattr(metrics, "parse_daily_report", fake_parse)
    monkeypatch.setattr(metrics, "write_public_workbook", fake_public)
    monkeypatch.setattr(metrics, "write_master_workbook", fake_master)
    args = Namespace(
        input_dir=str(input_dir),
        output_dir=str(output_dir),
        archive_dir=str(archive_dir),
        config=str(tmp_path / "missing-config.json"),
        week_start=None,
        week_end=None,
        migrate_history_from=[],
        migrate_history_only=False,
    )

    initialize_integrity(args)
    metrics.run(args)

    assert len(master_records) == 2
    assert {record.report_date for record in master_records} == {date(2026, 7, 18)}
    assert not active_source.exists()


def test_active_and_archive_conflict_stops_before_writes_or_moves(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    input_dir = tmp_path / "01 Daily Reports - Drop Here"
    archive_dir = tmp_path / "03 Archive"
    output_dir = tmp_path / "02 Finished Reports"
    input_dir.mkdir()
    archived_source = (
        archive_dir
        / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
        / "week-ending-2026-07-19"
        / "Daily Report archived.xlsx"
    )
    archived_source.parent.mkdir(parents=True)
    active_source = input_dir / "Daily Report active.xlsx"
    archived_source.write_text("archived", encoding="utf-8")
    active_source.write_text("active", encoding="utf-8")

    def fake_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        gross_sales = 125.0 if path.name == active_source.name else 100.0
        return [
            make_record(
                date(2026, 7, 18), source_file=path.name, gross_sales=gross_sales
            )
        ]

    monkeypatch.setattr(metrics, "parse_daily_report", fake_parse)
    args = Namespace(
        input_dir=str(input_dir),
        output_dir=str(output_dir),
        archive_dir=str(archive_dir),
        config=str(tmp_path / "missing-config.json"),
        week_start=None,
        week_end=None,
        migrate_history_from=[],
        migrate_history_only=False,
    )

    initialize_integrity(args)
    with pytest.raises(ValueError, match="Conflicting daily reports"):
        metrics.run(args)

    assert active_source.exists()
    assert not output_dir.exists()


def test_dashboard_quality_warnings_do_not_overlap_action_tables() -> None:
    report_day = date(2026, 7, 18)
    warning_rows = [
        {
            "week_end": report_day - metrics.timedelta(days=index * 7),
            "location": f"Store {index}",
            "source_days": metrics.OPERATING_WEEK_DAYS - 1,
            "check_average": 10.0,
        }
        for index in range(6)
    ]
    action_rows = [
        {
            "action": "Coach",
            "priority": "High",
            "location": "RC Richmond",
            "subject": "Server One",
            "impact": "High",
            "evidence": "Evidence",
            "recommended_follow_up": "Follow up",
        }
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)

    metrics.write_dashboard_sheet(
        workbook,
        [make_record(report_day)],
        warning_rows,
        [],
        [],
        action_rows,
        [],
        [],
        metrics.DEFAULT_CONFIG,
        Path("reports"),
        report_day,
        report_day,
    )

    dashboard = workbook["Dashboard"]
    value_rows = {
        cell.value: cell.row
        for row in dashboard.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert {f"Store {index}" for index in range(6)} <= set(value_rows)
    assert value_rows["Coach First"] > max(
        value_rows[f"Store {index}"] for index in range(6)
    )
    workbook.close()
