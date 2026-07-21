from __future__ import annotations

from argparse import Namespace
from datetime import date
import os
from pathlib import Path
import shutil
import subprocess
import threading

from openpyxl import load_workbook
import pytest

import red_onion_integrity as integrity
import red_onion_weekly_metrics as metrics


ACTIVE_DAY = date(2026, 7, 21)  # Tuesday
ACTIVE_WEEK_END = date(2026, 7, 26)  # Sunday
ACTIVE_NAME = "Daily Report - TM - 07-21-2026.xlsx"


def workflow_args(tmp_path: Path, *, initialize_baseline: bool = False) -> Namespace:
    return Namespace(
        input_dir=str(tmp_path / "01 Daily Reports - Drop Here"),
        output_dir=str(tmp_path / "02 Finished Reports"),
        archive_dir=str(tmp_path / "03 Archive"),
        config=str(tmp_path / "missing-config.json"),
        week_start=None,
        week_end=None,
        migrate_history_from=[],
        migrate_history_only=False,
        initialize_integrity_baseline=initialize_baseline,
    )


def minimal_records(day: date, source_file: str) -> list[metrics.MetricRecord]:
    common = {
        "source_file": source_file,
        "report_date": day,
        "location": "RC Richmond",
        "gross_sales": 1000.0,
        "guest_count": 50.0,
        "check_average": 20.0,
        "wine_sales": 100.0,
        "wine_pct": 0.10,
        "rate_of_sale_by_guest_count": 0.20,
        "average_ticket_time_seconds": 600.0,
    }
    return [
        metrics.MetricRecord(
            raw_user_name="",
            display_name="",
            is_location_total=True,
            **common,
        ),
        metrics.MetricRecord(
            raw_user_name="Alex Server",
            display_name="Alex Server",
            is_location_total=False,
            **common,
        ),
    ]


@pytest.fixture(scope="module")
def valid_master_template(tmp_path_factory: pytest.TempPathFactory) -> Path:
    root = tmp_path_factory.mktemp("transaction-race-master")
    path = root / "Red_Onion_Server_Master.xlsx"
    metrics.write_master_workbook(
        minimal_records(ACTIVE_DAY, ACTIVE_NAME),
        path,
        metrics.load_config(root / "missing-config.json"),
        root / "source",
        ACTIVE_DAY,
        ACTIVE_DAY,
    )

    # Give the fixture one real, unlocked manager-notes cell so the publication
    # race test exercises the same carry-forward surface an operator edits.
    action = {header: None for header in metrics.ACTION_HEADERS}
    action.update(
        {
            "Action ID": "RACE-ACTION",
            "Priority": "Medium",
            "Status": "Open",
            "Person / Area": "Alex Server",
            "Manager Notes": "",
        }
    )
    workbook = load_workbook(path, data_only=False)
    metrics.write_action_tracking_sheet(
        workbook, "Action Board", [action], editable=True
    )
    metrics.finalize_management_workbook(workbook)
    workbook.save(path)
    workbook.close()
    metrics.stamp_generated_content_digest(path)
    metrics.validate_management_workbook(path)

    workbook = load_workbook(path, data_only=False)
    assert workbook["Action Board"]["N5"].protection.locked is False
    workbook.close()
    return path


def active_source(args: Namespace, content: bytes = b"source-version-A") -> Path:
    source = Path(args.input_dir) / ACTIVE_NAME
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_bytes(content)
    return source


def manifest_paths(archive_dir: Path) -> list[Path]:
    root = metrics.integrity_manifest_dir(archive_dir)
    return sorted(root.glob("*.json"), key=lambda path: path.name.casefold())


def install_lightweight_workflow(
    monkeypatch: pytest.MonkeyPatch,
    valid_master_template: Path,
    *,
    parser=None,
    master_writer=None,
) -> None:
    if parser is None:
        parser = lambda path, config: minimal_records(ACTIVE_DAY, path.name)

    def write_public(
        location: str,
        selected_records: list[metrics.MetricRecord],
        output_dir: Path,
        config: dict,
        public_start: date,
        public_end: date,
    ) -> Path:
        path = output_dir / f"Check_Wine_RVA{public_end:%m%d%y}.xlsx"
        path.write_bytes(b"synthetic-public-workbook")
        return path

    if master_writer is None:
        def master_writer(
            records: list[metrics.MetricRecord],
            output_path: Path,
            config: dict,
            source_dir: Path,
            public_start: date,
            public_end: date,
        ) -> Path:
            # A prior master is copied into staging by run(). For a first run,
            # seed the staging path with a fully validated template.
            if not output_path.exists():
                shutil.copy2(valid_master_template, output_path)
            return output_path

    monkeypatch.setattr(metrics, "parse_daily_report", parser)
    monkeypatch.setattr(metrics, "write_public_workbook", write_public)
    monkeypatch.setattr(metrics, "write_master_workbook", master_writer)


def test_missing_manifest_with_managed_state_fails_closed_instead_of_rebaselining(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    output_dir = Path(args.output_dir)
    archive_dir = Path(args.archive_dir)
    output_dir.mkdir(parents=True)
    shutil.copy2(valid_master_template, output_dir / valid_master_template.name)
    raw = metrics.canonical_daily_archive_dir(archive_dir) / "evidence.bin"
    raw.parent.mkdir(parents=True)
    raw.write_bytes(b"recorded-raw")
    [baseline] = metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    baseline.unlink()
    source = active_source(args)
    install_lightweight_workflow(monkeypatch, valid_master_template)

    with pytest.raises(integrity.IntegrityError):
        metrics.run(args)

    assert source.read_bytes() == b"source-version-A"
    assert raw.read_bytes() == b"recorded-raw"
    assert manifest_paths(archive_dir) == []


def test_active_input_is_version_pinned_before_parsing_and_rejected_if_replaced(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    [baseline] = metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    source = active_source(args, b"source-version-A")
    parse_calls: list[bytes] = []

    def replace_while_parsing(path: Path, config: dict) -> list[metrics.MetricRecord]:
        parse_calls.append(path.read_bytes())
        source.write_bytes(b"source-version-B")  # Same length, different SHA-256.
        return minimal_records(ACTIVE_DAY, path.name)

    install_lightweight_workflow(
        monkeypatch, valid_master_template, parser=replace_while_parsing
    )

    with pytest.raises(integrity.IntegrityError):
        metrics.run(args)

    assert parse_calls == [b"source-version-A"]
    assert source.read_bytes() == b"source-version-B"
    assert manifest_paths(archive_dir) == [baseline]
    assert not metrics.generated_workbook_archive_dir(archive_dir).exists()


def test_predecessor_manifest_tampered_after_preflight_is_not_blessed_by_new_manifest(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    [baseline] = metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    source = active_source(args)
    tampered = False

    def tamper_predecessor(path: Path, config: dict) -> list[metrics.MetricRecord]:
        nonlocal tampered
        if not tampered:
            payload = integrity.read_json_manifest(baseline)
            payload["tampered_mid_run"] = True
            integrity.write_json_manifest_atomic(
                baseline,
                payload,
                root=metrics.integrity_manifest_dir(archive_dir),
            )
            tampered = True
        return minimal_records(ACTIVE_DAY, path.name)

    install_lightweight_workflow(
        monkeypatch, valid_master_template, parser=tamper_predecessor
    )

    with pytest.raises(integrity.IntegrityError):
        metrics.run(args)

    assert source.exists()
    assert manifest_paths(archive_dir) == [baseline]
    assert integrity.read_json_manifest(baseline)["tampered_mid_run"] is True


def test_raw_archive_tampered_after_preflight_is_not_blessed_by_new_manifest(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    raw = metrics.canonical_daily_archive_dir(archive_dir) / "evidence.bin"
    raw.parent.mkdir(parents=True)
    raw.write_bytes(b"archive-version-A")
    [baseline] = metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    source = active_source(args)
    tampered = False

    def tamper_archive(path: Path, config: dict) -> list[metrics.MetricRecord]:
        nonlocal tampered
        if not tampered:
            raw.write_bytes(b"archive-version-B")  # Same length, different SHA-256.
            tampered = True
        return minimal_records(ACTIVE_DAY, path.name)

    install_lightweight_workflow(
        monkeypatch, valid_master_template, parser=tamper_archive
    )

    with pytest.raises(integrity.IntegrityError):
        metrics.run(args)

    assert source.exists()
    assert raw.read_bytes() == b"archive-version-B"
    assert manifest_paths(archive_dir) == [baseline]


def test_second_run_is_rejected_while_first_run_holds_the_workflow_lock(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    active_source(args)
    first_entered_parser = threading.Event()
    release_first = threading.Event()

    def blocking_parser(path: Path, config: dict) -> list[metrics.MetricRecord]:
        if threading.current_thread().name == "race-first-run":
            first_entered_parser.set()
            if not release_first.wait(timeout=15):
                raise AssertionError("Timed out waiting to release the first run")
        return minimal_records(ACTIVE_DAY, path.name)

    install_lightweight_workflow(
        monkeypatch, valid_master_template, parser=blocking_parser
    )
    first_results: list[list[Path]] = []
    first_errors: list[BaseException] = []

    def run_first() -> None:
        try:
            first_results.append(metrics.run(args))
        except BaseException as exc:  # Preserve the exact thread failure for assertions.
            first_errors.append(exc)

    first = threading.Thread(target=run_first, name="race-first-run", daemon=True)
    first.start()
    assert first_entered_parser.wait(timeout=15), "The first run never reached its parser"

    second_error: BaseException | None = None
    try:
        metrics.run(args)
    except BaseException as exc:
        second_error = exc
    finally:
        release_first.set()
        first.join(timeout=20)

    assert not first.is_alive(), "The first run did not terminate"
    assert isinstance(second_error, integrity.IntegrityError)
    assert any(
        phrase in str(second_error).casefold()
        for phrase in ("already running", "another run", "run lock", "in progress")
    )
    assert first_errors == []
    assert len(first_results) == 1


def test_input_symlink_is_rejected_before_any_external_target_is_parsed(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    outside = tmp_path / "outside-input" / ACTIVE_NAME
    outside.parent.mkdir(parents=True)
    outside.write_bytes(b"outside-source")
    link = Path(args.input_dir) / ACTIVE_NAME
    link.parent.mkdir(parents=True)
    try:
        link.symlink_to(outside)
    except OSError:
        pytest.skip("File symlinks are unavailable on this host")

    parse_calls: list[Path] = []

    def record_parse(path: Path, config: dict) -> list[metrics.MetricRecord]:
        parse_calls.append(path)
        return minimal_records(ACTIVE_DAY, path.name)

    install_lightweight_workflow(
        monkeypatch, valid_master_template, parser=record_parse
    )

    with pytest.raises(integrity.IntegrityError):
        metrics.run(args)

    assert parse_calls == []
    assert link.is_symlink()
    assert outside.read_bytes() == b"outside-source"


def test_output_symlink_is_rejected_without_reading_or_replacing_external_master(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    outside = tmp_path / "outside-output" / valid_master_template.name
    outside.parent.mkdir(parents=True)
    shutil.copy2(valid_master_template, outside)
    before = outside.read_bytes()
    output_link = Path(args.output_dir) / valid_master_template.name
    output_link.parent.mkdir(parents=True)
    try:
        output_link.symlink_to(outside)
    except OSError:
        pytest.skip("File symlinks are unavailable on this host")

    with pytest.raises(integrity.IntegrityError):
        metrics.run(args)

    assert output_link.is_symlink()
    assert outside.read_bytes() == before
    assert manifest_paths(Path(args.archive_dir)) == []


def test_broken_output_symlink_is_rejected_before_baseline_creation(
    tmp_path: Path,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    missing_target = (
        tmp_path / "outside-output" / "Red_Onion_Server_Master.xlsx"
    )
    output_link = Path(args.output_dir) / missing_target.name
    output_link.parent.mkdir(parents=True)
    try:
        output_link.symlink_to(missing_target)
    except OSError:
        pytest.skip("File symlinks are unavailable on this host")

    with pytest.raises(integrity.IntegrityError, match="link|reparse"):
        metrics.run(args)

    assert output_link.is_symlink()
    assert not missing_target.exists()
    assert manifest_paths(Path(args.archive_dir)) == []


def test_recorded_master_replaced_by_symlink_is_rejected_before_read(
    tmp_path: Path,
    valid_master_template: Path,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    live_master = Path(args.output_dir) / valid_master_template.name
    live_master.parent.mkdir(parents=True)
    shutil.copy2(valid_master_template, live_master)
    metrics.run(args)
    baseline_manifests = manifest_paths(Path(args.archive_dir))
    assert len(baseline_manifests) == 1

    outside = tmp_path / "outside-output" / valid_master_template.name
    outside.parent.mkdir(parents=True)
    shutil.copy2(valid_master_template, outside)
    before = outside.read_bytes()
    live_master.unlink()
    try:
        live_master.symlink_to(outside)
    except OSError:
        if os.path.lexists(live_master):
            live_master.unlink()
        shutil.copy2(valid_master_template, live_master)
        pytest.skip("File symlinks are unavailable on this host")

    with pytest.raises(integrity.IntegrityError, match="link|reparse"):
        metrics.run(args)

    assert live_master.is_symlink()
    assert outside.read_bytes() == before
    assert manifest_paths(Path(args.archive_dir)) == baseline_manifests


def test_manager_edit_after_staging_is_preserved_when_publication_aborts(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True)
    live_master = output_dir / valid_master_template.name
    shutil.copy2(valid_master_template, live_master)
    metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    active_source(args)
    late_note = "Manager note saved while the weekly run was staging"
    edit_injected = False

    def write_staged_master_then_edit_live_copy(
        records: list[metrics.MetricRecord],
        output_path: Path,
        config: dict,
        source_dir: Path,
        public_start: date,
        public_end: date,
    ) -> Path:
        nonlocal edit_injected
        assert output_path.is_file(), "run() should copy the live master into staging first"
        workbook = load_workbook(live_master, data_only=False)
        workbook["Action Board"]["N5"] = late_note
        workbook.save(live_master)
        workbook.close()
        edit_injected = True
        return output_path

    install_lightweight_workflow(
        monkeypatch,
        valid_master_template,
        master_writer=write_staged_master_then_edit_live_copy,
    )

    with pytest.raises(integrity.IntegrityError, match="newer edit was preserved"):
        metrics.run(args)

    assert edit_injected is True
    workbook = load_workbook(live_master, data_only=False)
    assert workbook["Action Board"]["N5"].value == late_note
    workbook.close()


def test_migration_source_replaced_after_plan_construction_aborts_without_commit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    archive_dir = tmp_path / "03 Archive"
    baseline_args = workflow_args(tmp_path, initialize_baseline=True)
    [baseline] = metrics.run(baseline_args)
    source_root = tmp_path / "legacy-history"
    source_root.mkdir()
    source = source_root / "Daily Report historical.xlsx"
    source.write_bytes(b"migration-source-A")

    monkeypatch.setattr(
        metrics,
        "parse_daily_report",
        lambda path, config: minimal_records(date(2026, 7, 14), path.name),
    )
    original_build = metrics.build_history_migration_plan

    def build_then_replace(
        source_dirs: list[Path],
        supplied_archive_dir: Path,
        config: dict,
        **kwargs,
    ) -> metrics.HistoryMigrationPlan:
        plan = original_build(source_dirs, supplied_archive_dir, config, **kwargs)
        source.write_bytes(b"migration-source-B")  # Same size, different SHA-256.
        return plan

    monkeypatch.setattr(metrics, "build_history_migration_plan", build_then_replace)
    args = workflow_args(tmp_path)
    args.migrate_history_from = [str(source_root)]
    args.migrate_history_only = True

    with pytest.raises(
        integrity.IntegrityError,
        match="History migration source changed during the run",
    ):
        metrics.run(args)

    assert source.read_bytes() == b"migration-source-B"
    assert manifest_paths(archive_dir) == [baseline]
    raw_root = archive_dir / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
    assert not raw_root.exists() or not any(path.is_file() for path in raw_root.rglob("*"))


def test_publication_conflict_after_displacement_preserves_file_and_recovery_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_dir = tmp_path / "02 Finished Reports"
    archive_dir = tmp_path / "03 Archive"
    stage_dir = tmp_path / "stage"
    output_dir.mkdir()
    archive_dir.mkdir()
    stage_dir.mkdir()
    name = "Check_Wine_RVA072626.xlsx"
    staged = stage_dir / name
    final = output_dir / name
    staged.write_bytes(b"new generated output")
    final.write_bytes(b"original manager output")
    original_hash = metrics.sha256_file(final)
    run_id = "publication-race"
    conflicting_bytes = b"Dropbox conflict after displacement"
    original_copy = metrics.verified_copy_file

    def inject_conflict_after_displacement(
        source: Path,
        destination: Path,
        *,
        replace: bool = False,
    ) -> str:
        if source.absolute() == staged.absolute() and destination.absolute() == final.absolute():
            assert not final.exists()
            final.write_bytes(conflicting_bytes)
        return original_copy(source, destination, replace=replace)

    monkeypatch.setattr(metrics, "verified_copy_file", inject_conflict_after_displacement)

    with pytest.raises(integrity.IntegrityError, match="retained the recovery snapshot"):
        metrics.snapshot_and_publish_outputs(
            staged_paths=[staged],
            output_dir=output_dir,
            archive_dir=archive_dir,
            week_end=ACTIVE_WEEK_END,
            run_id=run_id,
            expected_existing_hashes={name.casefold(): original_hash},
        )

    snapshot = (
        archive_dir
        / metrics.GENERATED_WORKBOOK_ARCHIVE_FOLDER
        / f"week-ending-{ACTIVE_WEEK_END.isoformat()}"
        / run_id
    )
    assert final.read_bytes() == conflicting_bytes
    assert (snapshot / "replaced" / name).read_bytes() == b"original manager output"
    assert (snapshot / "published" / name).read_bytes() == b"new generated output"


def test_raw_archive_replacement_before_forced_rollback_is_preserved(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    [baseline] = metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    source = active_source(args)
    install_lightweight_workflow(monkeypatch, valid_master_template)
    replacement = b"Dropbox replacement in raw archive"

    def replace_raw_then_fail(**kwargs):
        archived = (
            metrics.canonical_daily_archive_dir(archive_dir)
            / f"week-ending-{ACTIVE_WEEK_END.isoformat()}"
            / ACTIVE_NAME
        )
        assert archived.is_file(), "The pinned raw copy must exist before publication"
        archived.write_bytes(replacement)
        raise RuntimeError("forced publication failure")

    monkeypatch.setattr(metrics, "snapshot_and_publish_outputs", replace_raw_then_fail)

    with pytest.raises(
        integrity.IntegrityError,
        match="preserved a changed raw-archive replacement",
    ):
        metrics.run(args)

    archived = (
        metrics.canonical_daily_archive_dir(archive_dir)
        / f"week-ending-{ACTIVE_WEEK_END.isoformat()}"
        / ACTIVE_NAME
    )
    assert archived.read_bytes() == replacement
    assert source.read_bytes() == b"source-version-A"
    assert manifest_paths(archive_dir) == [baseline]


def test_archived_history_is_parsed_only_from_manifest_pinned_bytes(
    tmp_path: Path,
    valid_master_template: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    args = workflow_args(tmp_path)
    archive_dir = Path(args.archive_dir)
    raw = (
        metrics.canonical_daily_archive_dir(archive_dir)
        / "week-ending-2026-07-19"
        / "Daily Report historical.xlsx"
    )
    raw.parent.mkdir(parents=True)
    original = b"manifest-recorded-history"
    injected = b"temporarily-manipulated!!"
    assert len(original) == len(injected)
    raw.write_bytes(original)
    metrics.run(workflow_args(tmp_path, initialize_baseline=True))
    active_source(args)
    observed_history_bytes: list[bytes] = []

    def parse_with_restore_attack(
        path: Path, config: dict
    ) -> list[metrics.MetricRecord]:
        if path.name == raw.name:
            if path.absolute() == raw.absolute():
                raw.write_bytes(injected)
                observed_history_bytes.append(path.read_bytes())
                raw.write_bytes(original)
            else:
                observed_history_bytes.append(path.read_bytes())
            return minimal_records(date(2026, 7, 14), path.name)
        return minimal_records(ACTIVE_DAY, path.name)

    install_lightweight_workflow(
        monkeypatch, valid_master_template, parser=parse_with_restore_attack
    )

    metrics.run(args)

    assert observed_history_bytes == [original]
    assert raw.read_bytes() == original


def test_rollback_created_file_does_not_delete_replacement_after_hash(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    raw_root = tmp_path / "raw"
    candidate = raw_root / "week-ending-2026-07-26" / "Daily Report current.xlsx"
    candidate.parent.mkdir(parents=True)
    candidate.write_bytes(b"run-owned-version")
    expected_hash = metrics.sha256_file(candidate)
    replacement = b"Dropbox replacement after comparison"
    original_sha256 = metrics.sha256_file
    injected = False

    def inject_after_comparison(path: Path) -> str:
        nonlocal injected
        digest = original_sha256(path)
        supplied = Path(path)
        if not injected and supplied.parent == candidate.parent:
            candidate.write_bytes(replacement)
            injected = True
        return digest

    monkeypatch.setattr(metrics, "sha256_file", inject_after_comparison)

    conflicts = metrics.rollback_created_files(
        [candidate], raw_root, expected_hashes={candidate: expected_hash}
    )

    assert injected is True
    assert conflicts == []
    assert candidate.read_bytes() == replacement


def test_output_rollback_does_not_overwrite_replacement_after_hash(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_dir = tmp_path / "02 Finished Reports"
    snapshot_dir = tmp_path / "snapshot"
    output_dir.mkdir()
    snapshot_dir.mkdir()
    final = output_dir / "Red_Onion_Server_Master.xlsx"
    backup = snapshot_dir / final.name
    final.write_bytes(b"run-staged-output")
    backup.write_bytes(b"pre-run-manager-output")
    staged_hash = metrics.sha256_file(final)
    original_hash = metrics.sha256_file(backup)
    replacement = b"manager edit after rollback comparison"
    original_sha256 = metrics.sha256_file
    injected = False

    def inject_after_comparison(path: Path) -> str:
        nonlocal injected
        digest = original_sha256(path)
        supplied = Path(path)
        if not injected and supplied.parent == output_dir:
            final.write_bytes(replacement)
            injected = True
        return digest

    monkeypatch.setattr(metrics, "sha256_file", inject_after_comparison)

    conflicts = metrics.rollback_published_outputs(
        {
            final: metrics.OutputRollback(
                backup=backup,
                original_sha256=original_hash,
                backup_sha256=original_hash,
            )
        },
        {final: staged_hash},
    )

    assert injected is True
    assert conflicts == [final]
    assert final.read_bytes() == replacement


def test_legacy_archive_helper_rejects_symlink_without_deleting_target(
    tmp_path: Path,
) -> None:
    source_dir = tmp_path / "input"
    archive_dir = tmp_path / "archive"
    archive_dir.mkdir()
    outside = tmp_path / "outside" / "Daily Report external.xlsx"
    source_dir.mkdir()
    outside.parent.mkdir()
    outside.write_bytes(b"external raw report")
    link = source_dir / "Daily Report submitted.xlsx"
    try:
        link.symlink_to(outside)
    except OSError:
        pytest.skip("File symlinks are unavailable on this host")

    with pytest.raises(integrity.IntegrityError, match="link|reparse"):
        metrics.archive_processed_files([link], archive_dir, ACTIVE_WEEK_END)

    assert link.is_symlink()
    assert outside.read_bytes() == b"external raw report"
    assert not archive_dir.exists() or not any(
        path.is_file() for path in archive_dir.rglob("*")
    )


def test_legacy_archive_helper_rejects_linked_source_directory(
    tmp_path: Path,
) -> None:
    outside = tmp_path / "outside-source"
    outside.mkdir()
    target = outside / "Daily Report external.xlsx"
    target.write_bytes(b"external raw report")
    linked_dir = tmp_path / "linked-input"
    try:
        linked_dir.symlink_to(outside, target_is_directory=True)
    except OSError:
        if os.name != "nt":
            pytest.skip("Directory links are unavailable on this host")
        junction = subprocess.run(
            ["cmd.exe", "/c", "mklink", "/J", str(linked_dir), str(outside)],
            capture_output=True,
            text=True,
            check=False,
        )
        if junction.returncode != 0:
            pytest.skip("Directory symlinks or junctions are unavailable on this host")

    archive_dir = tmp_path / "archive"
    archive_dir.mkdir()
    submitted = linked_dir / target.name
    with pytest.raises(integrity.IntegrityError, match="link|reparse"):
        metrics.archive_processed_files([submitted], archive_dir, ACTIVE_WEEK_END)

    assert target.read_bytes() == b"external raw report"
    assert not archive_dir.exists() or not any(
        path.is_file() for path in archive_dir.rglob("*")
    )


def test_legacy_migration_helper_refuses_manifest_managed_archive(
    tmp_path: Path,
) -> None:
    args = workflow_args(tmp_path, initialize_baseline=True)
    archive_dir = Path(args.archive_dir)
    metrics.run(args)
    source_dir = tmp_path / "legacy-history"
    source_dir.mkdir()
    source = source_dir / "Daily Report historical.xlsx"
    source.write_bytes(b"legacy report bytes")

    with pytest.raises(
        integrity.IntegrityError,
        match="outside the locked integrity transaction",
    ):
        metrics.migrate_history_files(
            [source_dir], archive_dir, metrics.DEFAULT_CONFIG
        )

    assert source.read_bytes() == b"legacy report bytes"
    assert metrics.archived_daily_report_paths(archive_dir) == []


def test_raw_archive_directory_link_rejects_external_writes_where_supported(
    tmp_path: Path,
) -> None:
    archive_dir = tmp_path / "03 Archive"
    outside = tmp_path / "outside-archive-target"
    input_dir = tmp_path / "01 Daily Reports - Drop Here"
    archive_dir.mkdir()
    outside.mkdir()
    input_dir.mkdir()
    raw_link = archive_dir / metrics.CANONICAL_DAILY_ARCHIVE_FOLDER
    try:
        raw_link.symlink_to(outside, target_is_directory=True)
    except OSError:
        if os.name != "nt":
            pytest.skip("Directory symlinks are unavailable on this host")
        junction = subprocess.run(
            ["cmd.exe", "/c", "mklink", "/J", str(raw_link), str(outside)],
            capture_output=True,
            text=True,
            check=False,
        )
        if junction.returncode != 0:
            pytest.skip("Directory symlinks or junctions are unavailable on this host")
    source = input_dir / ACTIVE_NAME
    source.write_bytes(b"pinned source")
    captures = metrics.capture_active_inputs([source], input_dir)

    with pytest.raises(integrity.IntegrityError, match="link|junction"):
        metrics.copy_captured_active_files_verified(
            captures,
            archive_dir,
            ACTIVE_WEEK_END,
        )

    assert list(outside.iterdir()) == []
