from __future__ import annotations

import json
from argparse import Namespace
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

import red_onion_weekly_metrics as metrics
from red_onion_runtime import write_json_atomic


def sample_signal() -> dict:
    return {
        "Entity Key": "server|rc richmond|alex|coaching",
        "Priority": "Medium",
        "Status": "Review Needed",
        "Owner": "Pat Manager",
        "Due Date": date(2026, 7, 30),
        "Location": "RC Richmond",
        "Person / Area": "Alex",
        "Action": "Coaching Prompt",
        "Signal": "Downward / Below Peer Reference / 8-week Downward",
        "Why It Matters": "Watch: check average",
        "Recommended Next Step": "Review recent shifts.",
        "Peer Comparison": "Below Peer Reference",
        "Recent Movement": "Downward",
        "Evidence Status": "Stable",
        "Comparator Type": "Same-store prior-four-week median",
        "Peer Cohort Size": 8,
        "Peer Cohort Weeks": 4,
        "Threshold Version": "2026.07-v3",
        "Recurring Drivers": "check_average",
        "Stability Result": "Stable under every active-day removal",
        "Last Seen": date(2026, 7, 19),
        "_evidence_week_ends": ["2026-07-12", "2026-07-19"],
        "_source_evidence": [
            {
                "source_file": "Daily Report.xlsx",
                "sha256": "a" * 64,
                "format": ".xlsx",
                "parser_engine": "openpyxl",
                "report_date_source": "Workbook Date(s) field",
                "report_date": "2026-07-19",
            }
        ],
        "_metric_evidence": {
            "recent_metric_scores": {"check_average": -2},
            "guest_count": 42,
        },
    }


def current_action() -> dict:
    action = metrics.enrich_management_signal(sample_signal())
    action.update(
        {
            "Action ID": "ABC123",
            "First Seen": date(2026, 7, 19),
            "Weeks Open": 1,
            "Context Notes": "",
            "Signal State": "Current",
            "Review Disposition": "Pending Review",
            "Reviewed By": "",
            "Review Date": None,
        }
    )
    return action


def test_stable_codes_and_evidence_id_are_deterministic() -> None:
    first = metrics.enrich_management_signal(sample_signal())
    second = metrics.enrich_management_signal(sample_signal())

    assert first["Action Code"] == "COACHING_PROMPT"
    assert first["Reason Code"] == "SERVER_TWO_WEEK_DOWNWARD_BELOW_PEER_STABLE"
    assert first["Evidence ID"] == second["Evidence ID"]
    assert first["Evidence Week Ends"] == "2026-07-12, 2026-07-19"
    sources = json.loads(first["Evidence Sources"])
    assert sources[0]["sha256"] == "a" * 64


def test_management_signal_matches_reviewed_semantic_golden() -> None:
    result = metrics.enrich_management_signal(sample_signal())
    fixture = Path(__file__).parent / "fixtures" / "management_signal_golden.json"
    expected = json.loads(fixture.read_text(encoding="utf-8"))

    assert {field: result[field] for field in expected} == expected


def test_context_only_change_updates_audit_evidence_without_resetting_review() -> None:
    def signal(ticket_seconds: float) -> dict:
        raw = sample_signal()
        raw["_metric_evidence"] = {
            "decision_metric_fields": ["check_average", "wine_pct"],
            "recent_metric_scores": {
                "check_average": -2,
                "wine_pct": -1,
            },
            "context_only_metrics": {
                "average_ticket_time_seconds": ticket_seconds,
                "ticket_time_available": True,
                "ticket_time_weight_basis": "Check Count",
            },
        }
        return metrics.enrich_management_signal(raw)

    prior = signal(100.0)
    prior.update(
        {
            "Action ID": "ABC123",
            "Status": "In Progress",
            "First Seen": date(2026, 7, 12),
            "Weeks Open": 2,
            "Context Notes": "Comparable shifts reviewed.",
            "Review Disposition": "Coaching Accepted",
            "Reviewed By": "Pat Manager",
            "Review Date": date(2026, 7, 20),
            "Signal State": "Current",
        }
    )
    changed_context = signal(200.0)
    assert changed_context["Evidence ID"] != prior["Evidence ID"]

    current, _ = metrics.merge_management_actions(
        [changed_context],
        {"active_actions": [prior], "action_history": []},
    )

    assert current[0]["Evidence ID"] == changed_context["Evidence ID"]
    assert current[0]["Status"] == "In Progress"
    assert current[0]["Review Disposition"] == "Coaching Accepted"
    assert current[0]["Reviewed By"] == "Pat Manager"
    assert current[0]["Review Date"] == date(2026, 7, 20)


def test_action_board_is_single_queue_and_evidence_links_to_it() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    action = current_action()
    metrics.write_action_tracking_sheet(
        wb, "Action Board", [action], editable=True
    )
    metrics.write_evidence_detail_sheet(wb, [action], [])

    board = wb["Action Board"]
    evidence = wb["Evidence Detail"]
    assert "Action Focus" not in wb.sheetnames
    assert board["C3"].value.startswith("Single action queue:")
    assert "C3:W3" in {str(merged) for merged in board.merged_cells.ranges}
    assert board["C5"].value == "Medium"
    assert all(
        board.column_dimensions[column].hidden is True
        for column in ("A", "B", "J", "M", "O", "P", "Q", "R", "T")
    )
    assert evidence["A5"].value == action["Evidence ID"]
    assert evidence["A5"].hyperlink.target == "#'Action Board'!C5"
    assert evidence.column_dimensions["L"].hidden is True
    assert evidence.column_dimensions["M"].hidden is True
    assert evidence.row_dimensions[5].height in {75, 90}
    assert evidence["L5"].value == action["Evidence Sources"]
    assert evidence["M5"].value == action["Metric Evidence"]
    assert json.loads(evidence["L5"].value)[0]["parser_engine"] == "openpyxl"
    assert "protected values remain unchanged" in evidence["A3"].value
    assert all(
        cell.protection.locked is not False
        for cell in evidence._cells.values()
    )


def test_verified_evidence_uses_live_editable_action_fields(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    action = current_action()
    metrics.write_action_tracking_sheet(
        wb, "Action Board", [action], editable=True
    )
    metrics.write_evidence_detail_sheet(wb, [action], [])
    board = wb["Action Board"]
    board["D5"] = "Blocked"
    board["E5"] = "Current Manager"
    board["F5"] = date(2026, 8, 6)
    board["N5"] = "Waiting for comparable shift context."
    board["U5"] = "Coaching Accepted"
    board["V5"] = "Reviewing Manager"
    board["W5"] = date(2026, 8, 1)
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    workbook_path = output_dir / "Red_Onion_Server_Master.xlsx"
    wb.save(workbook_path)
    wb.close()

    archive_dir = tmp_path / "archive"
    archive_dir.mkdir()
    manifest_path = archive_dir / "manifest.json"
    manifest_path.write_text("{}", encoding="utf-8")
    manifest_sha256 = "b" * 64
    manifest_payload = {
        "run_id": "run-1",
        "created_at_utc": "2026-07-23T12:00:00+00:00",
        "master_generated_content_sha256": "c" * 64,
        "provenance": {
            "git": {"commit": "d" * 40},
            "effective_config_sha256": "e" * 64,
        },
    }
    monkeypatch.setattr(
        metrics,
        "latest_integrity_manifest_path",
        lambda archive: manifest_path,
    )
    monkeypatch.setattr(
        metrics,
        "verify_integrity_anchor",
        lambda archive, anchor: (manifest_path.resolve(), manifest_sha256),
    )
    monkeypatch.setattr(
        metrics,
        "verify_integrity_state",
        lambda archive, output, manifest: (
            manifest_payload,
            manifest_sha256,
        ),
    )
    monkeypatch.setattr(
        metrics, "managed_master_workbook_path", lambda output: workbook_path
    )
    monkeypatch.setattr(
        metrics,
        "validate_v2_management_evidence_workbook",
        lambda workbook, expected_digest: None,
    )
    args = Namespace(
        output_dir=str(output_dir),
        archive_dir=str(archive_dir),
        integrity_anchor_dir=str(tmp_path / "anchors"),
    )

    _, rows, _ = metrics.verified_evidence_source(args)

    assert rows[0]["Status"] == "Blocked"
    assert rows[0]["Owner"] == "Current Manager"
    assert metrics.as_date(rows[0]["Due Date"]) == date(2026, 8, 6)
    assert rows[0]["Action Code"] == action["Action Code"]
    assert rows[0]["Evidence ID"] == action["Evidence ID"]
    assert rows[0]["Review Disposition"] == "Coaching Accepted"
    assert rows[0]["Reviewed By"] == "Reviewing Manager"


def test_completed_disposition_requires_reviewer_and_review_date() -> None:
    action = current_action()
    action["Status"] = "Open"
    action["Review Disposition"] = "Coaching Accepted"

    with pytest.raises(ValueError, match="Reviewed By, and Review Date"):
        metrics.validate_action_board_records([action])


def test_completed_disposition_rejects_reviewer_outside_active_roster() -> None:
    action = current_action()
    action.update(
        {
            "Status": "Open",
            "Review Disposition": "Coaching Accepted",
            "Reviewed By": "Pasted Reviewer",
            "Review Date": date(2026, 7, 23),
        }
    )

    with pytest.raises(ValueError, match="active person from the Owner Roster"):
        metrics.validate_action_board_records(
            [action],
            allowed_reviewers=["Authorized Manager"],
        )


def sample_package() -> metrics.ManagementEvidencePackageV2:
    record = metrics.EvidenceRecordV2(
        action_id="ABC123",
        evidence_id="EVIDENCE1",
        action_code="COACHING_PROMPT",
        reason_code="SERVER_TWO_WEEK_DOWNWARD_BELOW_PEER_STABLE",
        location="RC Richmond",
        person_or_area="Alex",
        priority="Medium",
        status="In Progress",
        owner="Pat Manager",
        due_date="2026-07-30",
        recommended_next_step="Review recent shifts.",
        why_it_matters="Watch: check average",
        evidence_week_ends="2026-07-12, 2026-07-19",
        evidence_sources=(
            {
                "source_file": "Daily Report.xlsx",
                "sha256": "a" * 64,
            },
        ),
        metric_evidence={"guest_count": 42},
        methodology_version=metrics.MANAGEMENT_METHODOLOGY_VERSION,
        comparator_type="Same-store prior-four-week median",
        peer_cohort_size=8,
        peer_cohort_weeks=4,
        threshold_version="2026.07-v3",
        evidence_status="Stable",
        recurring_drivers="check_average",
        stability_result="Stable under every active-day removal",
        review_disposition="Coaching Accepted",
        reviewed_by="Authorized Manager",
        review_date="2026-07-23",
    )
    return metrics.ManagementEvidencePackageV2(
        source={
            "manifest_path": "manifest.json",
            "manifest_sha256": "b" * 64,
            "manifest_run_id": "run-1",
            "manifest_created_at_utc": "2026-07-23T12:00:00+00:00",
            "workbook_file": "Red_Onion_Server_Master.xlsx",
            "workbook_generated_content_sha256": "c" * 64,
            "generator_commit": "d" * 40,
            "effective_config_sha256": "e" * 64,
            "methodology_version": metrics.MANAGEMENT_METHODOLOGY_VERSION,
        },
        records=(record,),
        retention_delete_after="2027-07-23",
    )


def sample_legacy_package() -> metrics.ManagementEvidencePackageV1:
    record = metrics.EvidenceRecord(
        action_id="ABC123",
        evidence_id="EVIDENCE1",
        action_code="COACH_NOW",
        reason_code="SERVER_FALLING_BELOW_BENCHMARK",
        location="RC Richmond",
        person_or_area="Alex",
        priority="High",
        status="Open",
        owner="Pat Manager",
        due_date="2026-07-30",
        recommended_next_step="Review recent shifts.",
        why_it_matters="Watch: check average",
        evidence_week_ends="2026-07-12, 2026-07-19",
        evidence_sources=(),
        metric_evidence={"guest_count": 42},
        methodology_version="2026.07-v1",
    )
    return metrics.ManagementEvidencePackageV1(
        source=sample_package().source,
        records=(record,),
        retention_delete_after="2027-07-23",
    )


def test_candidate_requires_exact_approval_and_promotes_identical_bytes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    package = sample_package()
    monkeypatch.setattr(
        metrics, "build_management_evidence_package", lambda args: package
    )
    candidate = tmp_path / "review" / "candidate.json"
    args = Namespace(archive_dir=str(tmp_path / "archive"))
    paths = metrics.stage_management_evidence(args, candidate)
    candidate_bytes = candidate.read_bytes()
    payload = json.loads(candidate_bytes)
    assert payload["retention"]["days"] == 365
    assert payload["distribution"]["automatic_upload"] is False
    assert payload["distribution"]["automatic_send"] is False
    assert payload["records"][0]["person_or_area"] == "Alex"

    template = json.loads(paths[2].read_text(encoding="utf-8"))
    template.update(
        {
            "approved_by": "Authorized Manager",
            "approved_at_utc": "2026-07-23T13:00:00+00:00",
            "purpose": "Approved AI management coaching review",
        }
    )
    approval = tmp_path / "review" / "approval.json"
    write_json_atomic(approval, template)
    (tmp_path / "archive").mkdir()
    monkeypatch.setattr(
        metrics,
        "verified_evidence_source",
        lambda args: (package.source, [], tmp_path / "master.xlsx"),
    )

    promoted = metrics.promote_approved_management_evidence(
        args, candidate, approval
    )

    assert promoted[0].read_bytes() == candidate_bytes
    receipt = json.loads(promoted[1].read_text(encoding="utf-8"))
    assert receipt["approved_by"] == "Authorized Manager"
    assert receipt["automatic_upload"] is False
    assert receipt["automatic_send"] is False


def test_approval_hash_mismatch_is_rejected_before_archive_creation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    package = sample_package()
    monkeypatch.setattr(
        metrics, "build_management_evidence_package", lambda args: package
    )
    args = Namespace(archive_dir=str(tmp_path / "archive"))
    candidate = tmp_path / "candidate.json"
    paths = metrics.stage_management_evidence(args, candidate)
    approval = json.loads(paths[2].read_text(encoding="utf-8"))
    approval.update(
        {
            "candidate_sha256": "0" * 64,
            "approved_by": "Authorized Manager",
            "approved_at_utc": "2026-07-23T13:00:00+00:00",
            "purpose": "Review",
        }
    )
    approval_path = tmp_path / "approval.json"
    write_json_atomic(approval_path, approval)

    with pytest.raises(metrics.IntegrityError, match="does not match"):
        metrics.promote_approved_management_evidence(
            args, candidate, approval_path
        )

    assert not (tmp_path / "archive").exists()


def test_legacy_v1_candidate_cannot_be_promoted_as_v2(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    legacy_package = sample_legacy_package()
    monkeypatch.setattr(
        metrics,
        "build_management_evidence_package",
        lambda args: legacy_package,
    )
    args = Namespace(archive_dir=str(tmp_path / "archive"))
    candidate = tmp_path / "legacy-candidate.json"
    paths = metrics.stage_management_evidence(args, candidate)
    approval = json.loads(paths[2].read_text(encoding="utf-8"))
    approval.update(
        {
            "approved_by": "Authorized Manager",
            "approved_at_utc": "2026-07-23T13:00:00+00:00",
            "purpose": "Review",
        }
    )
    approval_path = tmp_path / "legacy-approval.json"
    write_json_atomic(approval_path, approval)

    with pytest.raises(metrics.IntegrityError, match="approved evidence contract"):
        metrics.promote_approved_management_evidence(
            args, candidate, approval_path
        )

    assert not (tmp_path / "archive").exists()
