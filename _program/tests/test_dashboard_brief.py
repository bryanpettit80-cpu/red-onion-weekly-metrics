from __future__ import annotations

from datetime import date, timedelta

from openpyxl import Workbook

import red_onion_weekly_metrics as metrics


METRIC_VALUES = {
    "gross_sales": 12000.0,
    "guest_count": 500.0,
    "check_average": 24.0,
    "wine_pct": 0.12,
    "rate_of_sale_by_guest_count": 0.18,
    "average_ticket_time_seconds": 72 * 60.0,
}


def record_for(day: date) -> metrics.MetricRecord:
    return metrics.MetricRecord(
        source_file=f"daily-{day.isoformat()}.xlsx",
        report_date=day,
        location="RC Richmond",
        raw_user_name="",
        display_name="",
        is_location_total=True,
        gross_sales=100.0,
        guest_count=10.0,
        check_average=10.0,
        wine_sales=10.0,
        wine_pct=0.10,
        rate_of_sale_by_guest_count=0.20,
        average_ticket_time_seconds=3600.0,
    )


def weekly_location_row(
    week_end: date, location: str, source_days: int = metrics.OPERATING_WEEK_DAYS
) -> dict:
    return {
        "week_start": week_end - timedelta(days=metrics.OPERATING_WEEK_DAYS - 1),
        "week_end": week_end,
        "location": location,
        **METRIC_VALUES,
        "active_days": source_days,
        "source_days": source_days,
    }


def management_item(entity: str, *, group: bool = False) -> dict:
    benchmark = {**METRIC_VALUES, "gross_sales": 11000.0, "guest_count": 480.0}
    return {
        "entity": "All Stores" if group else entity,
        "latest": dict(METRIC_VALUES),
        "prior": dict(METRIC_VALUES),
        "benchmark_values": benchmark,
        "benchmark_sources": {field: "4-week baseline" for field in METRIC_VALUES},
        "prior_changes": {field: 0.0 for field in METRIC_VALUES},
        "benchmark_changes": {
            field: METRIC_VALUES[field] - benchmark[field] for field in METRIC_VALUES
        },
        "priority": "Monitor",
        "status": "On Track / Mixed",
        "recommended_focus": "Continue monitoring the current service plan.",
    }


def action(
    entity_key: str,
    person: str,
    priority: str,
    *,
    owner: str = "Pat Manager",
) -> dict:
    return {
        "Entity Key": entity_key,
        "Priority": priority,
        "Status": "Open",
        "Location": "RC Richmond",
        "Person / Area": person,
        "Action": "Coach Now",
        "Signal": "Traffic opportunity",
        "Why It Matters": "Guest traffic is below baseline | supporting detail",
        "Recommended Next Step": "Review the next two shifts",
        "Owner": owner,
        "Due Date": date(2026, 7, 15),
    }


def dashboard_inputs(source_days: int = metrics.OPERATING_WEEK_DAYS):
    week_end = date(2026, 7, 12)
    records = [
        record_for(week_end - timedelta(days=offset))
        for offset in range(metrics.OPERATING_WEEK_DAYS)
    ]
    if source_days < metrics.OPERATING_WEEK_DAYS:
        records = [record for record in records if record.report_date != date(2026, 7, 11)]
    weekly_locations = [
        weekly_location_row(week_end, location, source_days)
        for location in ("RC Richmond", "RC Virginia Beach")
    ]
    stores = [management_item(location) for location in ("RC Richmond", "RC Virginia Beach")]
    groups = [management_item("All Stores", group=True)]
    actions = [
        action("server|alex", "Alex", "High"),
        action("server|alex", "Alex duplicate", "Medium"),
        action("server|blair", "Blair", "Medium"),
        action("store|richmond", "RC Richmond", "Review"),
        action("server|casey", "Casey", "Recognize"),
    ]
    return records, weekly_locations, stores, groups, actions


def test_dashboard_is_one_screen_and_deduplicates_actions() -> None:
    wb = Workbook()
    records, weekly_locations, stores, groups, actions = dashboard_inputs()

    metrics.write_management_dashboard_sheet(
        wb, records, weekly_locations, [], [], stores, groups, actions
    )

    ws = wb["Dashboard"]
    values = {
        value
        for row in ws.iter_rows(min_row=1, max_row=24, min_col=1, max_col=12, values_only=True)
        for value in row
        if isinstance(value, str)
    }
    assert {"Sales vs Benchmark", "Guests vs Benchmark", "Review Items"}.issubset(values)
    assert {"TOP THREE REVIEW ITEMS", "STORE SNAPSHOT", "RECOGNITION REVIEW"}.issubset(values)
    assert ws["A7"].value == "+9.1%"
    assert ws["E7"].value == "+4.2%"
    assert ws["I7"].value == 3
    assert ws["C13"].value == "Alex | RC Richmond"
    assert ws["C14"].value == "Blair | RC Richmond"
    assert ws["C15"].value == "RC Richmond"
    assert "Alex duplicate" not in values
    assert "Casey" in ws["A23"].value
    assert len(ws._charts) == 0
    assert ws.max_row == 24
    assert "$A$1:$L$24" in str(ws.print_area)
    assert ws.freeze_panes == "A5"
    assert ws.sheet_view.zoomScale == 90
    assert ws.row_dimensions[9].height == 32
    for row in (19, 20):
        assert ws.cell(row=row, column=6).alignment.horizontal == "center"
        assert ws.cell(row=row, column=7).alignment.horizontal == "center"
        assert ws.cell(row=row, column=6).border.right.style == "thin"


def test_dashboard_action_dedup_keeps_same_topic_for_distinct_stores() -> None:
    richmond = action("store|rc richmond|traffic-watch", "RC Richmond", "Review")
    richmond.update(
        {
            "Location": "RC Richmond",
            "Action": "Store Review",
            "Signal": "Traffic Watch",
        }
    )
    virginia_beach = action(
        "store|rc virginia beach|traffic-watch",
        "RC Virginia Beach",
        "Review",
    )
    virginia_beach.update(
        {
            "Location": "RC Virginia Beach",
            "Action": "Store Review",
            "Signal": "Traffic Watch",
        }
    )

    selected = metrics.deduplicated_dashboard_actions(
        [richmond, dict(richmond), virginia_beach],
        {"Review"},
    )

    assert [row["Entity Key"] for row in selected] == [
        "store|rc richmond|traffic-watch",
        "store|rc virginia beach|traffic-watch",
    ]


def test_dashboard_does_not_display_zero_denominator_per_check_values() -> None:
    wb = Workbook()
    records, weekly_locations, stores, groups, actions = dashboard_inputs()
    for store in stores:
        store["latest"].update(
            {
                "check_count": 0.0,
                "check_count_available": True,
                "per_check_available": False,
                "sales_per_check": 0.0,
                "guests_per_check": 0.0,
            }
        )

    metrics.write_management_dashboard_sheet(
        wb, records, weekly_locations, [], [], stores, groups, actions
    )

    dashboard = wb["Dashboard"]
    for row in (19, 20):
        assert dashboard.cell(row=row, column=7).value == 0
        assert dashboard.cell(row=row, column=9).value == "n/a"
        assert dashboard.cell(row=row, column=10).value == "n/a"


def test_incomplete_week_pauses_comparisons_actions_and_recognition() -> None:
    wb = Workbook()
    records, weekly_locations, stores, groups, actions = dashboard_inputs(source_days=5)

    metrics.write_management_dashboard_sheet(
        wb, records, weekly_locations, [], [], stores, groups, actions
    )

    ws = wb["Dashboard"]
    assert ws["A7"].value == "PAUSED"
    assert ws["E7"].value == "PAUSED"
    assert ws["I7"].value == "PAUSED"
    assert "Jul 11" in ws["A4"].value
    assert "Actions are paused" in ws["A13"].value
    assert "Recognition is paused" in ws["A23"].value
    assert ws["C19"].value == "Preliminary"
    assert ws["C20"].value == "Preliminary"
    assert len(ws._charts) == 0


def test_missing_configured_location_pauses_dashboard_without_stale_metrics() -> None:
    wb = Workbook()
    records, _weekly_locations, stores, groups, actions = dashboard_inputs()
    latest_week_end = date(2026, 7, 12)
    weekly_locations = [
        weekly_location_row(latest_week_end, "RC Richmond"),
        weekly_location_row(latest_week_end - timedelta(days=7), "RC Virginia Beach"),
    ]

    metrics.write_management_dashboard_sheet(
        wb,
        records,
        weekly_locations,
        [],
        [],
        stores,
        groups,
        actions,
        metrics.DEFAULT_CONFIG,
    )

    dashboard = wb["Dashboard"]
    assert dashboard["A7"].value == "PAUSED"
    assert dashboard["E7"].value == "PAUSED"
    assert dashboard["I7"].value == "PAUSED"
    assert "PRELIMINARY" in dashboard["A4"].value
    assert "RC Virginia Beach data" in dashboard["A4"].value
    assert dashboard["A19"].value == "RC Richmond"
    assert dashboard["A20"].value == "RC Virginia Beach"
    assert dashboard["C20"].value == "Missing"
    assert dashboard["E20"].value is None
    assert dashboard["F20"].value is None
    assert dashboard["G20"].value is None

    metrics.write_management_data_quality_sheet(
        wb, weekly_locations, metrics.DEFAULT_CONFIG
    )
    data_quality = wb["Data Quality"]
    assert "incomplete" in data_quality["A3"].value.lower()
    assert data_quality["B7"].value == "RC Virginia Beach"
    assert data_quality["G7"].value == "Missing"


def test_data_quality_heatmap_shows_latest_16_weeks_and_gaps() -> None:
    workbook = Workbook()
    first_week_end = date(2026, 3, 22)
    week_ends = [
        first_week_end + timedelta(days=7 * offset) for offset in range(18)
    ]
    partial_week_end = week_ends[-2]
    location_missing_week_end = week_ends[-3]
    global_missing_week_end = week_ends[-4]
    weekly_locations = []
    for week_end in week_ends:
        for location in ("RC Richmond", "RC Virginia Beach"):
            if week_end == global_missing_week_end:
                continue
            if (
                week_end == location_missing_week_end
                and location == "RC Virginia Beach"
            ):
                continue
            source_days = (
                3
                if week_end == partial_week_end and location == "RC Richmond"
                else metrics.OPERATING_WEEK_DAYS
            )
            weekly_locations.append(
                weekly_location_row(week_end, location, source_days)
            )

    metrics.write_management_data_quality_sheet(
        workbook,
        weekly_locations,
        metrics.DEFAULT_CONFIG,
    )

    quality = workbook["Data Quality"]
    assert quality["A10"].value == "Source Completeness - Latest 16 Weeks"
    assert "Latest week appears first" in quality["A11"].value
    assert quality["A12"].value == "Week Ending"
    assert quality["B12"].value == "RC Richmond"
    assert quality["C12"].value == "RC Virginia Beach"
    assert quality["A13"].value == week_ends[-1]
    assert quality["A28"].value == week_ends[2]
    displayed_dates = {
        quality.cell(row=row, column=1).value for row in range(13, 29)
    }
    assert week_ends[0] not in displayed_dates
    assert week_ends[1] not in displayed_dates

    status_rows = {
        quality.cell(row=row, column=1).value: row for row in range(13, 29)
    }
    partial_row = status_rows[partial_week_end]
    location_missing_row = status_rows[location_missing_week_end]
    global_missing_row = status_rows[global_missing_week_end]
    assert quality.cell(row=partial_row, column=2).value == "Partial"
    assert quality.cell(row=partial_row, column=3).value == "Complete"
    assert quality.cell(row=location_missing_row, column=2).value == "Complete"
    assert quality.cell(row=location_missing_row, column=3).value == "Missing"
    assert quality.cell(row=global_missing_row, column=2).value == "Missing"
    assert quality.cell(row=global_missing_row, column=3).value == "Missing"
    assert quality.cell(row=partial_row, column=2).fill.fgColor.rgb[-6:] == "FFF2CC"
    assert (
        quality.cell(row=location_missing_row, column=3).fill.fgColor.rgb[-6:]
        == "F4CCCC"
    )
    assert quality.cell(row=13, column=2).fill.fgColor.rgb[-6:] == "D9EAD3"
    assert quality["A30"].value == "Historical Exceptions"


def test_data_quality_heatmap_handles_empty_history() -> None:
    workbook = Workbook()

    metrics.write_management_data_quality_sheet(
        workbook,
        [],
        metrics.DEFAULT_CONFIG,
    )

    quality = workbook["Data Quality"]
    assert quality["A10"].value == "Source Completeness - No Weekly History"
    assert quality["A13"].value == (
        "No weekly location history is available for the completeness view."
    )
    assert quality["A15"].value == "Historical Exceptions"


def test_data_quality_heatmap_scales_past_default_location_count() -> None:
    workbook = Workbook()
    locations = [f"RC Location {index:02d}" for index in range(1, 13)]
    config = {
        **metrics.DEFAULT_CONFIG,
        "locations": {
            location: {"short_code": f"L{index:02d}"}
            for index, location in enumerate(locations, start=1)
        },
    }
    week_end = date(2026, 7, 19)
    weekly_locations = [
        weekly_location_row(week_end, location, metrics.OPERATING_WEEK_DAYS)
        for location in locations
    ]

    metrics.write_management_data_quality_sheet(
        workbook,
        weekly_locations,
        config,
    )

    quality = workbook["Data Quality"]
    assert quality["B17"].value == locations[-1]
    assert quality["A20"].value == "Source Completeness - Latest 1 Week"
    assert quality["M22"].value == locations[-1]
    assert quality["M23"].value == "Complete"
    assert "A20:M20" in {str(merged) for merged in quality.merged_cells.ranges}
    assert "$A$1:$M$" in str(quality.print_area)


def test_scorecard_charts_use_latest_eight_complete_weeks() -> None:
    wb = Workbook()
    first_week_end = date(2026, 5, 10)
    complete_week_ends = [first_week_end + timedelta(days=7 * offset) for offset in range(9)]
    partial_week_end = complete_week_ends[-1] + timedelta(days=7)
    weekly_locations = [
        weekly_location_row(week_end, location)
        for week_end in complete_week_ends
        for location in ("RC Richmond", "RC Virginia Beach")
    ]
    weekly_locations.extend(
        weekly_location_row(partial_week_end, location, 5)
        for location in ("RC Richmond", "RC Virginia Beach")
    )
    weekly_groups = [
        {"week_end": week_end, "guest_count": 900 + index * 10}
        for index, week_end in enumerate([*complete_week_ends, partial_week_end])
    ]

    metrics.write_store_group_scorecards_sheet(
        wb,
        [
            management_item("All Stores", group=True),
            management_item("RC Richmond"),
            management_item("RC Virginia Beach"),
        ],
        metrics.DEFAULT_CONFIG,
        weekly_locations,
        weekly_groups,
    )

    scorecards = wb["Store & Group Scorecards"]
    chart_data = wb["_Dashboard Chart Data"]
    assert len(scorecards._charts) == 2
    assert chart_data.sheet_state == "veryHidden"
    assert [chart_data.cell(row=row, column=1).value for row in range(2, 10)] == [
        week_end.strftime("%m/%d") for week_end in complete_week_ends[-8:]
    ]
    assert chart_data["A10"].value is None
    assert partial_week_end.strftime("%m/%d") not in {
        chart_data.cell(row=row, column=1).value for row in range(2, 12)
    }
    assert [series.tx.v for series in scorecards._charts[0].series] == [
        "Richmond",
        "Virginia Beach",
    ]
    assert scorecards._charts[0].series[0].cat.strRef.f.endswith("$A$2:$A$9")
    assert scorecards._charts[1].series[0].cat.strRef.f.endswith("$F$2:$F$9")
