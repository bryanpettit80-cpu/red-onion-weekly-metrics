from __future__ import annotations

from copy import deepcopy
from datetime import date, timedelta
from pathlib import Path
from statistics import stdev

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.protection import hash_password
import pytest

import red_onion_weekly_metrics as metrics


def weekly_server_row(
    week_end: date,
    name: str,
    *,
    spg: float = 100.0,
    wine_pct: float = 0.10,
    guests: float = 50.0,
    active_days: int = 3,
    location: str = "RC Richmond",
) -> dict[str, object]:
    gross_sales = spg * guests
    return {
        "week_start": week_end - timedelta(days=5),
        "week_end": week_end,
        "location": location,
        "raw_user_name": name,
        "display_name": name,
        "gross_sales": gross_sales,
        "guest_count": guests,
        "check_average": spg,
        "wine_sales": gross_sales * wine_pct,
        "wine_pct": wine_pct,
        "active_days": active_days,
        "source_days": active_days,
    }


def weekly_location_row(
    week_end: date, *, location: str = "RC Richmond"
) -> dict[str, object]:
    return {
        "week_start": week_end - timedelta(days=5),
        "week_end": week_end,
        "location": location,
        "gross_sales": 100_000.0,
        "guest_count": 900.0,
        "check_average": 100_000.0 / 900.0,
        "wine_sales": 10_000.0,
        "wine_pct": 0.10,
        "active_days": 6,
        "source_days": 6,
    }


def eight_week_dataset() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    first = date(2026, 6, 21)
    week_ends = [first + timedelta(days=7 * offset) for offset in range(8)]
    server_rows: list[dict[str, object]] = []
    for week_end in week_ends:
        server_rows.append(
            weekly_server_row(
                week_end,
                "Taylor Guest",
                spg=112.0,
                wine_pct=0.12,
                guests=60.0,
            )
        )
        for peer_number in range(1, 6):
            server_rows.append(
                weekly_server_row(week_end, f"Peer {peer_number}")
            )
    return server_rows, [weekly_location_row(week_end) for week_end in week_ends]


def one_location_config() -> dict[str, object]:
    config = deepcopy(metrics.load_config(Path("missing-config.json")))
    config["locations"] = {"RC Richmond": {"short_code": "RVA"}}
    return config


def row_for(model: dict[str, object], name: str) -> dict[str, object]:
    return next(
        row
        for row in model["summary_rows"]  # type: ignore[index]
        if row["display_name"] == name
    )


@pytest.mark.parametrize("number", ["4040", "5050", "7070", "8080"])
def test_shared_pos_number_recognizes_observed_leading_identities(number: str) -> None:
    row = {"raw_user_name": f"{number} Server Server2", "display_name": "Alias"}
    assert metrics.shared_pos_number(row) == number


def test_shared_pos_number_does_not_reclassify_an_embedded_number() -> None:
    assert (
        metrics.shared_pos_number(
            {"raw_user_name": "Dining Room 5050", "display_name": "Dining Room"}
        )
        is None
    )


@pytest.mark.parametrize("name", ["Barbara Smith", "Bart Jones"])
def test_weekly_area_bar_pattern_does_not_match_person_name_prefixes(
    name: str,
) -> None:
    row = weekly_server_row(date(2026, 8, 9), name)

    assert metrics.weekly_area_for_row(row, one_location_config()) == "Dining Room"


def test_weekly_area_patterns_match_explicit_labels_and_fallback_totals() -> None:
    week_end = date(2026, 8, 9)
    rows = [
        weekly_server_row(week_end, "Main Bar", spg=60.0, guests=10.0),
        weekly_server_row(week_end, "Barbara Smith", spg=100.0, guests=20.0),
        weekly_server_row(week_end, "Bart Jones", spg=80.0, guests=30.0),
    ]
    config = one_location_config()

    model = metrics.build_shared_area_trends_model(
        rows,
        [weekly_location_row(week_end)],
        config,
    )

    assert metrics.weekly_area_for_row(rows[0], config) == "Bar"
    for location in ("All Stores", "RC Richmond"):
        area_rows = {
            row["area"]: row
            for row in model["area_rows"]
            if row["location"] == location
        }
        assert area_rows["Bar"]["latest_guest_count"] == pytest.approx(10.0)
        assert area_rows["Bar"]["latest_check_average"] == pytest.approx(60.0)
        assert area_rows["Dining Room"]["latest_guest_count"] == pytest.approx(50.0)
        assert area_rows["Dining Room"]["latest_gross_sales"] == pytest.approx(
            4_400.0
        )
        assert area_rows["Dining Room"]["latest_check_average"] == pytest.approx(
            88.0
        )


def test_weekly_area_uses_a_configured_literal_phrase() -> None:
    config = one_location_config()
    config["weekly_area_name_patterns"]["Bar"] = ["Cocktail Station"]  # type: ignore[index]

    assert (
        metrics.weekly_area_for_row(
            weekly_server_row(date(2026, 8, 9), "Cocktail Station 1"),
            config,
        )
        == "Bar"
    )


@pytest.mark.parametrize(
    ("spg_gap", "wine_gap", "expected"),
    [
        (11.5, 0.0, "Strong"),
        (0.0, 0.041, "Strong"),
        (-11.5, 0.0, "Below"),
        (0.0, -0.041, "Below"),
        (5.0, 0.01, "Near Peer"),
        (5.0, -0.01, "Mixed"),
        (None, None, "Not Qualified"),
    ],
)
def test_weekly_band_boundaries(
    spg_gap: float | None, wine_gap: float | None, expected: str
) -> None:
    assert (
        metrics.performance_consistency_weekly_band(
            spg_gap,
            wine_gap,
            spg_boundary=11.5,
            wine_boundary=0.041,
        )
        == expected
    )


def test_model_uses_capped_weighted_peer_gaps_and_sample_sd() -> None:
    server_rows, location_rows = eight_week_dataset()
    config = one_location_config()

    model = metrics.build_performance_consistency_model(
        server_rows, location_rows, config
    )
    target = row_for(model, "Taylor Guest")

    assert model["week_ends"] == tuple(
        date(2026, 6, 21) + timedelta(days=7 * offset) for offset in range(8)
    )
    assert target["qualified_weeks"] == 8
    assert target["qualified_guests"] == 480
    assert target["average_spg_gap"] == pytest.approx(12.0)
    assert target["average_wine_gap"] == pytest.approx(0.02)
    assert target["weekly_spg_gap_sd"] == pytest.approx(stdev([12.0] * 8))
    assert target["weekly_wine_gap_sd"] == pytest.approx(stdev([0.02] * 8))
    assert target["recent_spg_delta"] == pytest.approx(0.0)
    assert target["recent_wine_delta"] == pytest.approx(0.0)
    assert target["performance_level"] == "Strong"
    assert target["consistency"] == "High"
    assert target["confidence"] == "High"
    assert target["overall_read"] == "Consistently Strong"
    overall = model["overall_store"]
    assert overall["gross_sales"] == pytest.approx(100_000.0)
    assert overall["check_average"] == pytest.approx(100_000.0 / 900.0)
    assert overall["baseline_week_count"] == 4
    assert overall["gross_sales_change_pct"] == pytest.approx(0.0)


def test_models_require_every_configured_location_for_a_complete_shared_week() -> None:
    server_rows, location_rows = eight_week_dataset()
    config = metrics.load_config(Path("missing-config.json"))

    full_by_location, shared_weeks = metrics.full_week_ends_by_location(
        location_rows,
        config["locations"],
    )
    performance = metrics.build_performance_consistency_model(
        server_rows,
        location_rows,
        config,
    )
    shared_areas = metrics.build_shared_area_trends_model(
        server_rows,
        location_rows,
        config,
    )

    assert set(full_by_location) == {"RC Richmond", "RC Virginia Beach"}
    assert full_by_location["RC Virginia Beach"] == set()
    assert shared_weeks == set()
    assert performance["week_ends"] == ()
    assert performance["overall_store"] is None
    assert shared_areas["week_ends"] == ()
    assert shared_areas["latest_complete_week_end"] is None


def test_peer_pool_excludes_people_not_on_the_latest_complete_roster() -> None:
    first = date(2026, 6, 21)
    week_ends = [first + timedelta(days=7 * offset) for offset in range(8)]
    server_rows: list[dict[str, object]] = []
    for week_end in week_ends:
        server_rows.append(weekly_server_row(week_end, "Taylor Guest"))
        for peer_number in range(1, 5):
            server_rows.append(weekly_server_row(week_end, f"Current Peer {peer_number}"))
        if week_end != week_ends[-1]:
            server_rows.append(weekly_server_row(week_end, "Former Peer"))
    location_rows = [weekly_location_row(week_end) for week_end in week_ends]

    model = metrics.build_performance_consistency_model(
        server_rows,
        location_rows,
        one_location_config(),
    )
    target = row_for(model, "Taylor Guest")

    assert target["qualified_weeks"] == 0
    assert target["confidence"] == "Insufficient"
    assert target["overall_read"] == "Insufficient Data"
    assert "Former Peer" not in {
        row["display_name"] for row in model["summary_rows"]  # type: ignore[index]
    }


def test_shared_pos_row_has_no_scorecard_and_does_not_shift_peer_median() -> None:
    first = date(2026, 6, 21)
    week_ends = [first + timedelta(days=7 * offset) for offset in range(8)]
    server_rows: list[dict[str, object]] = []
    for week_end in week_ends:
        server_rows.append(weekly_server_row(week_end, "Taylor Guest", spg=100.0))
        for peer_number, peer_spg in enumerate((80, 90, 100, 110, 120), start=1):
            server_rows.append(
                weekly_server_row(
                    week_end,
                    f"Peer {peer_number}",
                    spg=float(peer_spg),
                )
            )
        # This high-value shared identity would move the six-row median from
        # $100 to $105 if it leaked into the current-person peer pool.
        server_rows.append(weekly_server_row(week_end, "5050 Bar", spg=1_000.0))
    location_rows = [weekly_location_row(week_end) for week_end in week_ends]

    model = metrics.build_performance_consistency_model(
        server_rows,
        location_rows,
        one_location_config(),
    )

    assert "5050 Bar" not in {
        row["display_name"] for row in model["summary_rows"]  # type: ignore[index]
    }
    assert "5050 Bar" not in {
        row["display_name"] for row in model["detail_rows"]  # type: ignore[index]
    }
    taylor_details = [
        row
        for row in model["detail_rows"]  # type: ignore[index]
        if row["display_name"] == "Taylor Guest"
    ]
    assert len(taylor_details) == 8
    assert {row["distinct_peer_count"] for row in taylor_details} == {5}
    assert {row["peer_spg"] for row in taylor_details} == {100.0}
    assert row_for(model, "Taylor Guest")["average_spg_gap"] == pytest.approx(0.0)


def test_historical_shared_pos_alias_is_absent_from_row_lookup_and_peer_pool() -> None:
    first = date(2026, 6, 21)
    week_ends = [first + timedelta(days=7 * offset) for offset in range(8)]
    server_rows: list[dict[str, object]] = []
    for week_end in week_ends:
        server_rows.append(weekly_server_row(week_end, "Taylor Guest", spg=100.0))
        for peer_number, peer_spg in enumerate((80, 90, 110, 120), start=1):
            server_rows.append(
                weekly_server_row(
                    week_end,
                    f"Peer {peer_number}",
                    spg=float(peer_spg),
                )
            )
        alias = weekly_server_row(week_end, "Alex Guest", spg=100.0)
        if week_end == week_ends[0]:
            alias["display_name"] = "5050 Bar"
            alias["check_average"] = 1_000.0
            alias["gross_sales"] = 50_000.0
        server_rows.append(alias)
    location_rows = [weekly_location_row(week_end) for week_end in week_ends]
    config = one_location_config()
    config["management_peer_reference"][  # type: ignore[index]
        "min_distinct_peers_per_week"
    ] = 4

    model = metrics.build_performance_consistency_model(
        server_rows,
        location_rows,
        config,
    )

    alex_details = [
        row
        for row in model["detail_rows"]  # type: ignore[index]
        if row["raw_user_name"] == "Alex Guest"
    ]
    assert len(alex_details) == 7
    assert week_ends[0] not in {row["week_end"] for row in alex_details}
    first_taylor = next(
        row
        for row in model["detail_rows"]  # type: ignore[index]
        if row["display_name"] == "Taylor Guest" and row["week_end"] == week_ends[0]
    )
    assert first_taylor["distinct_peer_count"] == 4
    assert first_taylor["peer_spg"] == pytest.approx(100.0)
    assert first_taylor["spg_gap"] == pytest.approx(0.0)


def test_confidence_gates_include_the_exact_six_200_and_four_150_boundaries() -> None:
    first = date(2026, 6, 21)
    week_ends = [first + timedelta(days=7 * offset) for offset in range(8)]
    high_guests = (34, 34, 33, 33, 33, 33, 24, 24)
    provisional_guests = (38, 38, 37, 37, 24, 24, 24, 24)
    server_rows: list[dict[str, object]] = []
    high_gaps = (0, 2, 4, 6, 8, 10, 0, 0)
    for offset, week_end in enumerate(week_ends):
        server_rows.append(
            weekly_server_row(
                week_end,
                "High Boundary",
                spg=100 + high_gaps[offset],
                guests=high_guests[offset],
            )
        )
        server_rows.append(
            weekly_server_row(
                week_end,
                "Provisional Boundary",
                guests=provisional_guests[offset],
            )
        )
        for peer_number in range(1, 6):
            server_rows.append(weekly_server_row(week_end, f"Peer {peer_number}"))
    location_rows = [weekly_location_row(week_end) for week_end in week_ends]

    model = metrics.build_performance_consistency_model(
        server_rows,
        location_rows,
        one_location_config(),
    )
    high = row_for(model, "High Boundary")
    provisional = row_for(model, "Provisional Boundary")

    assert (high["qualified_weeks"], high["qualified_guests"]) == (6, 200)
    assert high["confidence"] == "High"
    assert high["weekly_spg_gap_sd"] == pytest.approx(stdev(high_gaps[:6]))
    assert (provisional["qualified_weeks"], provisional["qualified_guests"]) == (
        4,
        150,
    )
    assert provisional["confidence"] == "Provisional"


def test_shared_numbers_and_areas_use_complete_week_weighted_rollups() -> None:
    first = date(2026, 7, 26)
    week_ends = [first + timedelta(days=7 * offset) for offset in range(3)]
    rows: list[dict[str, object]] = []
    for offset, week_end in enumerate(week_ends):
        rows.extend(
            [
                weekly_server_row(
                    week_end,
                    "5050 Server Server2",
                    spg=80 + offset * 5,
                    guests=20,
                ),
                weekly_server_row(week_end, "Bar 1 Bar 1", spg=50, guests=10),
                weekly_server_row(
                    week_end, "Bar Server", spg=100, guests=30
                ),
                weekly_server_row(
                    week_end, "BarPatio Bartender Patio", spg=70, guests=20
                ),
                weekly_server_row(
                    week_end, "Banquet Banquet 1", spg=140, guests=50
                ),
                weekly_server_row(
                    week_end, "Taylor Dining", spg=90, guests=40
                ),
                weekly_server_row(
                    week_end, "Wine Dinner Special", spg=160, guests=25
                ),
                weekly_server_row(week_end, "Takeout", spg=60, guests=100),
            ]
        )
    location_rows = [weekly_location_row(week_end) for week_end in week_ends]
    config = one_location_config()

    model = metrics.build_shared_area_trends_model(rows, location_rows, config)

    assert model["week_ends"] == tuple(week_ends)
    shared = model["shared_rows"][0]
    assert shared["shared_number"] == "5050"
    assert shared["latest_check_average"] == pytest.approx(90.0)
    assert shared["wow_check_average"] == pytest.approx(5.0)
    all_store = {
        row["area"]: row
        for row in model["area_rows"]
        if row["location"] == "All Stores"
    }
    assert all_store["Bar"]["latest_check_average"] == pytest.approx(87.5)
    assert all_store["Patio"]["latest_check_average"] == pytest.approx(70.0)
    assert all_store["Dining Room"]["latest_check_average"] == pytest.approx(90.0)
    assert all_store["Banquets"]["latest_check_average"] == pytest.approx(140.0)
    assert all_store["Wine Dinners"]["latest_check_average"] == pytest.approx(160.0)
    assert all_store["Dining Room"]["latest_guest_count"] == pytest.approx(40.0)


def test_shared_number_can_feed_a_configured_area_without_becoming_a_person() -> None:
    week_ends = [date(2026, 8, 2), date(2026, 8, 9)]
    rows = [
        weekly_server_row(week_end, "7070 Server Server3", spg=150, guests=20)
        for week_end in week_ends
    ]
    config = one_location_config()
    config["weekly_shared_number_areas"] = {"7070": "Wine Dinners"}
    model = metrics.build_shared_area_trends_model(
        rows,
        [weekly_location_row(week_end) for week_end in week_ends],
        config,
    )

    assert model["shared_rows"][0]["shared_number"] == "7070"
    wine = next(
        row
        for row in model["area_rows"]
        if row["location"] == "All Stores" and row["area"] == "Wine Dinners"
    )
    assert wine["latest_check_average"] == pytest.approx(150.0)
    assert metrics.dashboard_excluded(rows[-1], config) is True


def test_area_wow_does_not_bridge_a_missing_complete_week() -> None:
    week_ends = [date(2026, 7, 26), date(2026, 8, 2), date(2026, 8, 9)]
    rows = [
        weekly_server_row(week_ends[0], "Bar Server", spg=70),
        weekly_server_row(week_ends[2], "Bar Server", spg=90),
    ]
    model = metrics.build_shared_area_trends_model(
        rows,
        [weekly_location_row(week_end) for week_end in week_ends],
        one_location_config(),
    )
    bar = next(
        row
        for row in model["area_rows"]
        if row["location"] == "All Stores" and row["area"] == "Bar"
    )
    assert bar["latest_check_average"] == pytest.approx(90.0)
    assert bar["wow_check_average"] is None


def test_generated_sheets_preserve_navigation_protection_and_audit_boundary(
    tmp_path: Path,
) -> None:
    server_rows, location_rows = eight_week_dataset()
    config = one_location_config()
    model = metrics.build_performance_consistency_model(
        server_rows, location_rows, config
    )
    for week_end in model["week_ends"]:
        server_rows.append(
            weekly_server_row(week_end, "5050 Server Server2", spg=85, guests=40)
        )
    shared_area_model = metrics.build_shared_area_trends_model(
        server_rows, location_rows, config
    )
    latest = model["latest_complete_week_end"]
    readiness = metrics.LatestWeekReadiness(
        latest_week_end=latest,
        latest_location_rows=(location_rows[-1],),
        configured_locations=("RC Richmond",),
        location_gaps=(),
        expected_dates=tuple(latest - timedelta(days=offset) for offset in range(6)),
        received_dates=frozenset(latest - timedelta(days=offset) for offset in range(6)),
        missing_dates=(),
        ready=True,
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    metrics.write_performance_consistency_sheets(
        workbook, model, config, readiness
    )
    metrics.write_shared_area_trends_sheet(workbook, shared_area_model)
    metrics.finalize_management_workbook(workbook)
    path = tmp_path / "performance-consistency.xlsx"
    workbook.save(path)
    workbook.close()

    reloaded = load_workbook(path, data_only=False)
    try:
        for name in (
            "Performance Dashboard",
            "Server Scorecards",
            "Weekly Performance",
            "Shared & Area Trends",
            "Methodology",
        ):
            sheet = reloaded[name]
            assert sheet.sheet_state == "visible"
            assert sheet.protection.sheet is True
            assert sheet.protection.objects is True
            assert sheet.protection.scenarios is True
            assert sheet.protection.password == hash_password(
                metrics.WORKBOOK_OPERATOR_PASSWORD
            )
            assert sheet.sheet_view.topLeftCell == "A1"
            assert all(cell.protection.locked is not False for cell in sheet._cells.values())
            assert all(cell.comment is None for cell in sheet._cells.values())
            start_column, _ = metrics.management_menu_bounds(sheet)
            menu = sheet.cell(row=2, column=start_column)
            assert menu.value == metrics.MANAGEMENT_MENU_LABEL
            assert menu.hyperlink.target == metrics.MANAGEMENT_MENU_TARGET

        assert reloaded["_Consistency Calc"].sheet_state == "veryHidden"
        assert reloaded["_Consistency Calc"].protection.sheet is True
        assert reloaded["Server Scorecards"].tables["ServerScorecardsTable"].ref == "A6:P12"
        assert reloaded["Weekly Performance"].tables["WeeklyPerformanceTable"].ref == "A6:O12"
        assert "SharedPosTrendTable" in reloaded["Shared & Area Trends"].tables
        assert "AreaCheckAverageTrendTable" in reloaded["Shared & Area Trends"].tables
        shared_area_sheet = reloaded["Shared & Area Trends"]
        expected_shared_area_widths = {
            "A": 20.7109375,
            "B": 19.140625,
            "C": 30.85546875,
            "D": 18.140625,
            "E": 18.5703125,
            "F": 18.5703125,
            "G": 18.140625,
            "H": 18.140625,
            "I": 18.5703125,
            "J": 18.5703125,
            "K": 18.5703125,
            "L": 15.85546875,
            "M": 17.42578125,
            "N": 21.140625,
            "O": 17.42578125,
            "P": 21.140625,
            "Q": 13.0,
        }
        assert metrics.SHARED_AREA_TRENDS_COLUMN_WIDTHS == expected_shared_area_widths
        for column, width in expected_shared_area_widths.items():
            dimension = shared_area_sheet.column_dimensions[column]
            assert dimension.width == pytest.approx(width)
            assert dimension.bestFit is False
        for table_name in ("SharedPosTrendTable", "AreaCheckAverageTrendTable"):
            _, header_row, _, _ = range_boundaries(
                shared_area_sheet.tables[table_name].ref
            )
            assert shared_area_sheet.row_dimensions[header_row].height == 30
        assert len(reloaded["Performance Dashboard"]._charts) == 1
        assert reloaded["Performance Dashboard"]["A6"].value.startswith(
            "OVERALL STORE PERFORMANCE"
        )
        assert reloaded["Performance Dashboard"]["A8"].value == pytest.approx(
            100_000.0
        )
        assert reloaded.security.workbookPassword == hash_password(
            metrics.WORKBOOK_OPERATOR_PASSWORD
        )
        methodology_text = "\n".join(
            str(cell.value)
            for cell in reloaded["Methodology"]._cells.values()
            if cell.value not in (None, "")
        )
        operator_text = "\n".join(
            str(cell.value)
            for sheet_name in (
                "Performance Dashboard",
                "Server Scorecards",
                "Weekly Performance",
                "Shared & Area Trends",
                "Methodology",
            )
            for cell in reloaded[sheet_name]._cells.values()
            if cell.value not in (None, "")
        )
        assert metrics.PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION in methodology_text
        assert "fixed validated snapshot" not in methodology_text.casefold()
        assert "locks were removed" not in methodology_text.casefold()
        assert "unhiding" not in methodology_text.casefold()
        assert "Action Board" not in operator_text
        assert "See Data Quality" not in operator_text
        assert "Management Center Current Actions" in operator_text
        assert all(
            reloaded["Methodology"].row_dimensions[row].height >= 50
            for row in range(6, 13)
        )
    finally:
        reloaded.close()
