from __future__ import annotations

import argparse
import filecmp
import hashlib
import json
import math
import os
import re
import secrets
import shutil
import stat
import sys
import tempfile
import uuid
import warnings
from collections import defaultdict
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from statistics import median, stdev
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableStyleInfo

from red_onion_integrity import (
    FileFingerprint,
    IntegrityError,
    build_raw_inventory,
    canonical_json_sha256,
    collect_provenance,
    fingerprint_file,
    read_json_manifest,
    sha256_file,
    verify_manifest_chain,
    verify_raw_inventory,
    write_chained_manifest_atomic,
    write_json_manifest_atomic,
)
from red_onion_config import DEFAULT_CONFIG, load_config
from red_onion_runtime import (
    RunAttemptRecorder,
    RunReadiness,
    RunStage,
    safe_message,
    write_json_atomic,
)
from red_onion_fairness import (
    CandidatePolarity,
    MetricBand,
    PeerComparison,
    PeerObservation,
    PromptAction,
    RecentMovement,
    WeeklyCandidateSignal,
    assess_common_store_shock,
    classify_candidate,
    evaluate_two_week_persistence,
    leave_one_day_stability,
    leave_one_out_same_store_peer_reference,
    score_metric,
)


METRICS = [
    ("gross_sales", "Gross Sales"),
    ("guest_count", "Guest Count"),
    ("check_average", "Check Average"),
    ("wine_sales", "Wine Sales"),
    ("rate_of_sale_by_guest_count", "Rate of Sale by Guest Count"),
    ("average_ticket_time", "Average Ticket Time"),
]

MANAGEMENT_METRICS: tuple[tuple[str, str, str], ...] = (
    ("gross_sales", "Gross Sales", "currency"),
    ("guest_count", "Guest Count", "count"),
    ("check_average", "Sales / Guest", "currency"),
    ("wine_pct", "Wine %", "percent"),
    ("rate_of_sale_by_guest_count", "Rate of Sale", "rate"),
    ("average_ticket_time_seconds", "Ticket Time", "duration"),
)

STORE_GROUP_DISPLAY_METRICS: tuple[tuple[str, str, str], ...] = (
    ("gross_sales", "Gross Sales", "currency"),
    ("guest_count", "Guests", "count"),
    ("check_count", "Checks", "count"),
    ("check_average", "Sales / Guest", "currency"),
    ("sales_per_check", "Sales / Check", "currency"),
    ("guests_per_check", "Guests / Check", "ratio"),
    ("wine_pct", "Wine %", "percent"),
    ("rate_of_sale_by_guest_count", "Rate of Sale", "rate"),
    ("average_ticket_time_seconds", "Ticket Time", "duration"),
)

TARGET_FIELDS: tuple[tuple[str, str], ...] = (
    ("gross_sales", "Weekly Sales Target"),
    ("guest_count", "Weekly Guest Target"),
    ("check_average", "Check Average Target"),
    ("wine_pct", "Wine % Target"),
    ("rate_of_sale_by_guest_count", "Rate Target"),
    ("average_ticket_time_seconds", "Ticket Time Target (Min)"),
)

SERVER_PERSON_ACTION_FIELDS: tuple[str, ...] = (
    "check_average",
    "wine_pct",
)

SERVER_CONTEXT_FIELDS: tuple[str, ...] = (
    "rate_of_sale_by_guest_count",
    "average_ticket_time_seconds",
)

# Retain the established name for compatibility with callers and tests while
# making its decision scope explicit. Rate of Sale and Ticket Time remain
# visible operational context but cannot create person-level signals.
SERVER_TREND_FIELDS: tuple[str, ...] = SERVER_PERSON_ACTION_FIELDS

ACTION_HEADERS = [
    "Action ID",
    "Entity Key",
    "Priority",
    "Status",
    "Owner",
    "Due Date",
    "Location",
    "Person / Area",
    "Action",
    "Signal",
    "Why It Matters",
    "Recommended Next Step",
    "Last Seen",
    "Context Notes",
    "Peer Comparison",
    "Recent Movement",
    "First Seen",
    "Weeks Open",
    "Evidence Status",
    "Signal State",
    "Review Disposition",
    "Reviewed By",
    "Review Date",
]
LEGACY_ACTION_HEADERS_V1 = [
    "Action ID",
    "Entity Key",
    "Priority",
    "Status",
    "Owner",
    "Due Date",
    "Location",
    "Person / Area",
    "Action",
    "Signal",
    "Why It Matters",
    "Recommended Next Step",
    "Last Seen",
    "Manager Notes",
    "Performance Level",
    "Momentum",
    "First Seen",
    "Weeks Open",
    "Confidence",
    "Signal State",
]

LEGACY_ACTION_STATUS_CHOICES: tuple[str, ...] = (
    "Open",
    "In Progress",
    "Blocked",
    "Complete",
    "Dismissed",
)
ACTION_STATUS_CHOICES: tuple[str, ...] = (
    "Review Needed",
    *LEGACY_ACTION_STATUS_CHOICES,
)
REVIEW_DISPOSITION_CHOICES: tuple[str, ...] = (
    "Pending Review",
    "Coaching Accepted",
    "Recognition Accepted",
    "Context Explains",
    "Data Issue",
    "Monitor",
)
MANAGEMENT_METHODOLOGY_VERSION = "2026.07-v3"
PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION = "2026.07-v2"
PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION = "2026.08-v1"
PERFORMANCE_CONSISTENCY_WINDOW_WEEKS = 8
PERFORMANCE_CONSISTENCY_BLOCK_WEEKS = 4
PERFORMANCE_CONSISTENCY_GUEST_WEIGHT_CAP = 50.0
PERFORMANCE_CONSISTENCY_HIGH_MIN_WEEKS = 6
PERFORMANCE_CONSISTENCY_HIGH_MIN_GUESTS = 200.0
PERFORMANCE_CONSISTENCY_PROVISIONAL_MIN_WEEKS = 4
PERFORMANCE_CONSISTENCY_PROVISIONAL_MIN_GUESTS = 150.0
PERFORMANCE_CONSISTENCY_RECENT_SPG_THRESHOLD = 5.0
PERFORMANCE_CONSISTENCY_RECENT_WINE_THRESHOLD = 0.015
WEEKLY_SHARED_AREA_WINDOW_WEEKS = 8
WEEKLY_AREA_ORDER = ("Bar", "Patio", "Dining Room", "Banquets", "Wine Dinners")
PERFORMANCE_CONSISTENCY_COLORS: dict[str, tuple[str, str]] = {
    "Consistently Strong": ("E8F3E8", "1F5B23"),
    "Consistent / Near Peer": ("E3F0EF", "2F6F6D"),
    "Strong / Variable": ("FFF4CC", "B77900"),
    "Mixed / Monitor": ("FFF4CC", "B77900"),
    "Inconsistent": ("FFF4CC", "B77900"),
    "Consistently Below": ("F8E8E9", "8B1E23"),
    "Below / Variable": ("F8E8E9", "8B1E23"),
    "Insufficient Data": ("E9E9E9", "5C5C5C"),
    "High": ("E8F3E8", "1F5B23"),
    "Moderate": ("FFF4CC", "B77900"),
    "Low": ("F8E8E9", "8B1E23"),
    "Insufficient": ("E9E9E9", "5C5C5C"),
    "Strong": ("E8F3E8", "1F5B23"),
    "Near Peer": ("E8F0FA", "3E6FA8"),
    "Mixed": ("FFF4CC", "B77900"),
    "Below": ("F8E8E9", "8B1E23"),
    "Not Qualified": ("E9E9E9", "5C5C5C"),
    "—": ("F5F5F5", "7F7F7F"),
}
PREVIOUS_RATE_OF_SALE_GUIDANCE = (
    "Rate of Sale currently assumes lower is better."
)
PREVIOUS_TICKET_TIME_GUIDANCE = "Ticket Time is guest-weighted"
MANAGEMENT_SIGNAL_DISCLAIMER = (
    "Rule-based observational coaching signal—not a statistical, causal, or "
    "employment decision. Verify comparable work context and source accuracy."
)
EVIDENCE_DETAIL_HEADERS: tuple[str, ...] = (
    "Evidence ID",
    "Action ID",
    "Action Code",
    "Reason Code",
    "Priority",
    "Status",
    "Owner",
    "Due Date",
    "Location",
    "Person / Area",
    "Evidence Week Ends",
    "Evidence Sources",
    "Metric Evidence",
    "Comparator Type",
    "Peer Cohort Size",
    "Peer Cohort Weeks",
    "Threshold Version",
    "Evidence Status",
    "Recurring Drivers",
    "Stability Result",
    "Review Disposition",
    "Reviewed By",
    "Review Date",
    "Methodology Version",
    "Last Seen",
)
ACTION_EVIDENCE_FIELDS: tuple[str, ...] = (
    "Evidence ID",
    "Action Code",
    "Reason Code",
    "Evidence Week Ends",
    "Evidence Sources",
    "Metric Evidence",
    "Comparator Type",
    "Peer Cohort Size",
    "Peer Cohort Weeks",
    "Threshold Version",
    "Evidence Status",
    "Recurring Drivers",
    "Stability Result",
    "Methodology Version",
)

PRE_ACTION_FOCUS_VISIBLE_MANAGEMENT_SHEETS = [
    "Dashboard",
    "Action Board",
    "Server Scorecard",
    "Store & Group Scorecards",
    "Rising & Falling Stars",
    "Action History",
    "Data Quality",
    "Management Setup",
    "Run Notes",
]
LEGACY_V1_VISIBLE_MANAGEMENT_SHEETS = [
    "Dashboard",
    "Action Focus",
    "Action Board",
    "Server Scorecard",
    "Store & Group Scorecards",
    "Rising & Falling Stars",
    "Evidence Detail",
    "Action History",
    "Data Quality",
    "Management Setup",
    "Run Notes",
]
PRE_GUIDE_VISIBLE_MANAGEMENT_SHEETS = [
    "Dashboard",
    "Action Focus",
    "Action Board",
    "Server Scorecard",
    "Store & Group Scorecards",
    "Recent Movement Signals",
    "Evidence Detail",
    "Action History",
    "Data Quality",
    "Management Setup",
    "Run Notes",
]
PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS = [
    "How to Use",
    *PRE_GUIDE_VISIBLE_MANAGEMENT_SHEETS,
]
PRE_PERFORMANCE_VISIBLE_MANAGEMENT_SHEETS = [
    "How to Use",
    "Dashboard",
    "Action Board",
    "Team Trends",
    "Store & Group Scorecards",
    "Evidence Detail",
    "Action History",
    "Data Quality",
    "Management Setup",
    "Run Notes",
]
PRE_SIMPLIFICATION_VISIBLE_MANAGEMENT_SHEETS = [
    "How to Use",
    "Performance Dashboard",
    "Server Scorecards",
    "Weekly Performance",
    "Shared & Area Trends",
    "Methodology",
    "Dashboard",
    "Action Board",
    "Team Trends",
    "Store & Group Scorecards",
    "Evidence Detail",
    "Action History",
    "Data Quality",
    "Management Setup",
    "Run Notes",
]
VISIBLE_MANAGEMENT_SHEETS = [
    "How to Use",
    "Performance Dashboard",
    "Server Scorecards",
    "Weekly Performance",
    "Shared & Area Trends",
    "Methodology",
    "Management Center",
]
PRE_CONSOLIDATION_NAVIGATION_LINKS: tuple[tuple[str, str], ...] = (
    ("Guide", "How to Use"),
    ("Dashboard", "Dashboard"),
    ("Focus", "Action Focus"),
    ("Actions", "Action Board"),
    ("Servers", "Server Scorecard"),
    ("Stores", "Store & Group Scorecards"),
    ("Movement", "Recent Movement Signals"),
    ("Evidence", "Evidence Detail"),
    ("History", "Action History"),
    ("Quality", "Data Quality"),
    ("Setup", "Management Setup"),
    ("Run", "Run Notes"),
)
PRE_PERFORMANCE_NAVIGATION_LINKS: tuple[tuple[str, str], ...] = (
    ("Guide", "How to Use"),
    ("Dashboard", "Dashboard"),
    ("Actions", "Action Board"),
    ("Trends", "Team Trends"),
    ("Stores", "Store & Group Scorecards"),
    ("Evidence", "Evidence Detail"),
    ("History", "Action History"),
    ("Quality", "Data Quality"),
    ("Setup", "Management Setup"),
    ("Run", "Run Notes"),
)
PRE_SIMPLIFICATION_NAVIGATION_LINKS: tuple[tuple[str, str], ...] = (
    ("Guide", "How to Use"),
    ("Performance", "Performance Dashboard"),
    ("Scorecards", "Server Scorecards"),
    ("Weekly", "Weekly Performance"),
    ("Areas", "Shared & Area Trends"),
    ("Method", "Methodology"),
    ("Dashboard", "Dashboard"),
    ("Actions", "Action Board"),
    ("Trends", "Team Trends"),
    ("Stores", "Store & Group Scorecards"),
    ("Evidence", "Evidence Detail"),
    ("History", "Action History"),
    ("Quality", "Data Quality"),
    ("Setup", "Management Setup"),
    ("Run", "Run Notes"),
)
MANAGEMENT_NAVIGATION_LINKS: tuple[tuple[str, str], ...] = (
    ("Guide", "How to Use"),
    ("Performance", "Performance Dashboard"),
    ("Scorecards", "Server Scorecards"),
    ("Weekly", "Weekly Performance"),
    ("Areas", "Shared & Area Trends"),
    ("Method", "Methodology"),
    ("Manage", "Management Center"),
)
MANAGEMENT_MENU_LABEL = "Workbook Menu - click to choose a sheet on How to Use"
MANAGEMENT_MENU_TARGET = "#'How to Use'!A44"
MANAGEMENT_MENU_VISIBLE_COLUMN_COUNT = 12
MANAGEMENT_PRINT_WIDTHS: dict[str, int] = {
    "How to Use": 1,
    "Performance Dashboard": 1,
    "Server Scorecards": 1,
    "Weekly Performance": 1,
    "Shared & Area Trends": 1,
    "Methodology": 1,
    "Management Center": 1,
}
SHARED_AREA_TRENDS_COLUMN_WIDTHS: dict[str, float] = {
    "A": 20.7109375,
    "B": 19.140625,
    "C": 30.85546875,
    "D": 18.140625,
    "E": 18.5703125,
    "F": 18.5703125,
    "G": 18.140625,
    "H": 18.140625,
    "I": 18.5703125,
    "J": 18.5703125,
    "K": 18.5703125,
    "L": 15.85546875,
    "M": 17.42578125,
    "N": 21.140625,
    "O": 17.42578125,
    "P": 21.140625,
    "Q": 13.0,
}
PRE_SIMPLIFICATION_HOW_TO_USE_SECTION_HEADINGS: tuple[str, ...] = (
    "SCOPE AND PROHIBITED USE",
    "SIX-STEP WEEKLY WORKFLOW",
    "REQUIRED EVIDENCE REVIEW",
    "SIGNAL INTERPRETATION",
    "ACTION BOARD EDIT SURFACE",
    "DISPOSITION MEANINGS",
    "WORKBOOK MAP",
    "LIMITS, ASSUMPTIONS, AND ESCALATION",
)
HOW_TO_USE_SECTION_HEADINGS: tuple[str, ...] = (
    "SCOPE AND PROHIBITED USE",
    "SIX-STEP WEEKLY WORKFLOW",
    "REQUIRED EVIDENCE REVIEW",
    "SIGNAL INTERPRETATION",
    "MANAGEMENT CENTER EDIT SURFACE",
    "DISPOSITION MEANINGS",
    "WORKBOOK MAP",
    "LIMITS, ASSUMPTIONS, AND ESCALATION",
)

OPERATING_WEEK_START_WEEKDAY = 1  # Tuesday; date.weekday() uses Monday=0.
OPERATING_WEEK_END_WEEKDAY = 6  # Sunday.
OPERATING_WEEK_DAYS = 6
OPERATING_WEEK_LABEL = "Tuesday-Sunday"
DATA_QUALITY_COMPLETENESS_WEEKS = 16
FILENAME_DATE_BUSINESS_DATE_OFFSET_DAYS = 1
DAILY_REPORT_PREFIX = "daily report"
DAILY_REPORT_EXTENSIONS = frozenset({".xls", ".xlsx"})
DAILY_REPORT_FORMAT_LABEL = ".xls or .xlsx"
CANONICAL_DAILY_ARCHIVE_FOLDER = "processed-daily-reports"
INTEGRITY_MANIFEST_FOLDER = "run-manifests"
RUN_ATTEMPT_FOLDER = "run-attempts"
LAST_RUN_STATUS_FILE = "LAST RUN STATUS.txt"
INTEGRITY_ANCHOR_ENVIRONMENT_VARIABLE = "RED_ONION_INTEGRITY_ANCHOR_DIR"
INTEGRITY_ANCHOR_SCHEMA_VERSION = 1
GENERATED_WORKBOOK_ARCHIVE_FOLDER = "generated-workbooks"
OWNER_ROSTER_TABLE_NAME = "OwnerRoster"
OWNER_ROSTER_DEFINED_NAME = "ActiveOwnerChoices"
OWNER_VALIDATION_SHEET = "_Validation Lists"
MANAGEMENT_CENTER_SHEET = "Management Center"
OWNER_ROSTER_HEADERS = ("Owner Name", "Active")
OWNER_ROSTER_MIN_EDIT_ROWS = 50
OWNER_ROSTER_SPARE_ROWS = 10
OWNER_ROSTER_MAX_ROWS = 200
OWNER_ROSTER_VISIBLE_EDIT_ROWS = 15
LEGACY_WORKBOOK_DIGEST_SCHEME = "red-onion-generated-content-v2"
PREVIOUS_WORKBOOK_DIGEST_SCHEME = "red-onion-substantive-content-v3"
WORKBOOK_DIGEST_SCHEME = "red-onion-substantive-content-v4"
RUN_NOTES_DIGEST_LABEL = "Generated Content SHA-256"
WORKBOOK_PROTECTION_CONTRACT_LABEL = "Protection Contract"
WORKBOOK_PROTECTION_CONTRACT = "objects-scenarios-stop-validation-v1"
WORKBOOK_OPERATOR_PASSWORD = "redonion"
WORKBOOK_OPERATOR_UNLOCK_LABEL = "Operator Unlock"
EXCEL_DATE_NUMBER_FORMAT = "mm-dd-yy"
EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
UNTRUSTED_WORKBOOK_TEXT_HEADERS = frozenset(
    {"Raw Server", "Display Name", "Source File", "Server", "Person / Area"}
)
MAX_DATA_QUALITY_DATE_ROWS = 10_000
PROGRAM_DIR = Path(__file__).resolve().parent
REPOSITORY_ROOT = PROGRAM_DIR.parent
PROJECT_ROOT = (
    REPOSITORY_ROOT.parent
    if REPOSITORY_ROOT.name == "Red Onion Weekly Metrics Automation"
    else REPOSITORY_ROOT
)
DEFAULT_INPUT_DIR = PROJECT_ROOT / "01 Daily Reports - Drop Here"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "02 Finished Reports"
DEFAULT_ARCHIVE_DIR = PROJECT_ROOT / "03 Archive"
DEFAULT_CONFIG_PATH = PROGRAM_DIR / "red_onion_config.json"


@dataclass(frozen=True)
class MetricRecord:
    source_file: str
    report_date: date
    location: str
    raw_user_name: str
    display_name: str
    is_location_total: bool
    gross_sales: float
    guest_count: float
    check_average: float
    wine_sales: float
    wine_pct: float
    rate_of_sale_by_guest_count: float
    average_ticket_time_seconds: float
    source_sha256: str = ""
    source_format: str = ""
    parser_engine: str = ""
    report_date_source: str = "Unknown"
    rate_available: bool = True
    ticket_time_available: bool = True
    check_count: float = 0.0
    check_count_available: bool = False


@dataclass(frozen=True)
class ReportResolution:
    records_by_path: dict[Path, list[MetricRecord]]
    duplicate_paths: tuple[Path, ...]
    business_dates: tuple[date, ...]


@dataclass(frozen=True)
class HistoryMigrationPlan:
    copy_pairs: tuple[tuple[Path, Path], ...]
    effective_records_by_path: dict[Path, list[MetricRecord]]
    duplicate_paths: tuple[Path, ...]
    business_dates: tuple[date, ...]
    captured_sources: tuple[CapturedActiveInput, ...] = ()


@dataclass(frozen=True)
class HistoryMigrationResult:
    copied_paths: tuple[Path, ...]
    duplicate_files_ignored: int
    business_dates_considered: int


@dataclass(frozen=True)
class VerifiedArchiveCopy:
    source: Path
    destination: Path
    created: bool
    sha256: str


@dataclass(frozen=True)
class CapturedActiveInput:
    """One immutable, in-memory snapshot of an operator drop-folder file."""

    source: Path
    fingerprint: FileFingerprint
    content: bytes


@dataclass(frozen=True)
class OutputRollback:
    """Evidence needed to restore a replaced output without losing a newer edit."""

    backup: Path | None
    original_sha256: str | None
    backup_sha256: str | None
    displaced: Path | None = None


@dataclass(frozen=True)
class LatestWeekReadiness:
    latest_week_end: date | None
    latest_location_rows: tuple[dict[str, Any], ...]
    configured_locations: tuple[str, ...]
    location_gaps: tuple[str, ...]
    expected_dates: tuple[date, ...]
    received_dates: frozenset[date]
    missing_dates: tuple[date, ...]
    ready: bool

    @property
    def missing_parts(self) -> tuple[str, ...]:
        return (
            *(f"{report_date:%b} {report_date.day}" for report_date in self.missing_dates),
            *(f"{location} data" for location in self.location_gaps),
        )

    @property
    def missing_text(self) -> str:
        return ", ".join(self.missing_parts)


@dataclass(frozen=True)
class EvidenceRecord:
    action_id: str
    evidence_id: str
    action_code: str
    reason_code: str
    location: str
    person_or_area: str
    priority: str
    status: str
    owner: str
    due_date: str | None
    recommended_next_step: str
    why_it_matters: str
    evidence_week_ends: str
    evidence_sources: tuple[dict[str, Any], ...]
    metric_evidence: dict[str, Any]
    methodology_version: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "action_id": self.action_id,
            "evidence_id": self.evidence_id,
            "action_code": self.action_code,
            "reason_code": self.reason_code,
            "location": self.location,
            "person_or_area": self.person_or_area,
            "priority": self.priority,
            "status": self.status,
            "owner": self.owner,
            "due_date": self.due_date,
            "recommended_next_step": self.recommended_next_step,
            "why_it_matters": self.why_it_matters,
            "evidence_week_ends": self.evidence_week_ends,
            "evidence_sources": list(self.evidence_sources),
            "metric_evidence": self.metric_evidence,
            "methodology_version": self.methodology_version,
        }


@dataclass(frozen=True)
class ManagementEvidencePackageV1:
    source: dict[str, Any]
    records: tuple[EvidenceRecord, ...]
    retention_delete_after: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "contract": "ManagementEvidencePackageV1",
            "schema_version": 1,
            "classification": "Restricted Employee Performance Information",
            "permitted_use": (
                "Management-approved analysis of identifiable action and coaching evidence."
            ),
            "retention": {
                "days": 365,
                "delete_after": self.retention_delete_after,
                "automatic_deletion": False,
            },
            "distribution": {
                "mode": "Manual local export after exact-fingerprint approval",
                "automatic_upload": False,
                "automatic_send": False,
            },
            "source": self.source,
            "records": [record.to_dict() for record in self.records],
        }


@dataclass(frozen=True)
class EvidenceRecordV2(EvidenceRecord):
    comparator_type: str = "Same-store prior-four-week median"
    peer_cohort_size: int = 0
    peer_cohort_weeks: int = 0
    threshold_version: str = MANAGEMENT_METHODOLOGY_VERSION
    evidence_status: str = ""
    recurring_drivers: str = ""
    stability_result: str = ""
    review_disposition: str = "Pending Review"
    reviewed_by: str = ""
    review_date: str | None = None

    def to_dict(self) -> dict[str, Any]:
        payload = super().to_dict()
        payload.update(
            {
                "comparator_type": self.comparator_type,
                "peer_cohort_size": self.peer_cohort_size,
                "peer_cohort_weeks": self.peer_cohort_weeks,
                "threshold_version": self.threshold_version,
                "evidence_status": self.evidence_status,
                "recurring_drivers": self.recurring_drivers,
                "stability_result": self.stability_result,
                "review_disposition": self.review_disposition,
                "reviewed_by": self.reviewed_by,
                "review_date": self.review_date,
            }
        )
        return payload


@dataclass(frozen=True)
class ManagementEvidencePackageV2:
    source: dict[str, Any]
    records: tuple[EvidenceRecordV2, ...]
    retention_delete_after: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "contract": "ManagementEvidencePackageV2",
            "schema_version": 2,
            "classification": "Restricted Employee Performance Information",
            "permitted_use": (
                "Human-reviewed observational coaching and recognition prompts only."
            ),
            "prohibited_use": (
                "Not the sole or determinative basis for pay, scheduling, discipline, "
                "promotion, or termination."
            ),
            "retention": {
                "days": 365,
                "delete_after": self.retention_delete_after,
                "automatic_deletion": False,
            },
            "distribution": {
                "mode": "Manual local export after exact-fingerprint approval",
                "automatic_upload": False,
                "automatic_send": False,
            },
            "source": self.source,
            "records": [record.to_dict() for record in self.records],
        }


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except TypeError:
        pass
    return str(value).strip() == ""


def to_float(value: Any) -> float:
    if is_blank(value):
        return 0.0
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        number = float(str(value).replace(",", "").replace("$", "").strip())
    if not math.isfinite(number):
        raise ValueError(f"Expected a finite number, received {value!r}.")
    return number


def ticket_time_to_seconds(value: Any) -> float:
    if is_blank(value):
        return 0.0
    if isinstance(value, time):
        return (
            value.hour * 3600
            + value.minute * 60
            + value.second
            + value.microsecond / 1_000_000
        )
    if isinstance(value, datetime):
        t = value.time()
        return t.hour * 3600 + t.minute * 60 + t.second + t.microsecond / 1_000_000
    if isinstance(value, (int, float)):
        number = float(value)
        return number * 86400 if 0 <= number < 1 else number

    text = str(value).strip()
    match = re.match(r"^(\d+):(\d{2}):(\d{2})(?:\.(\d+))?$", text)
    if not match:
        return 0.0
    hours, minutes, seconds, fraction = match.groups()
    micros = float(f"0.{fraction}") if fraction else 0.0
    return int(hours) * 3600 + int(minutes) * 60 + int(seconds) + micros


def optional_float(value: Any) -> tuple[float, bool]:
    """Return a finite numeric value plus whether it was actually available."""

    if is_blank(value):
        return 0.0, False
    try:
        return to_float(value), True
    except (TypeError, ValueError, OverflowError):
        return 0.0, False


def optional_nonnegative_float(value: Any) -> tuple[float, bool]:
    """Return an available finite value only when it is non-negative."""

    number, available = optional_float(value)
    return (number, True) if available and number >= 0 else (0.0, False)


def optional_check_count(value: Any) -> tuple[float, bool]:
    """Parse a non-negative whole-number Check Count."""

    number, available = optional_nonnegative_float(value)
    if not available or not float(number).is_integer():
        return 0.0, False
    return number, True


def optional_ticket_time_seconds(value: Any) -> tuple[float, bool]:
    """Parse ticket time without turning malformed or non-finite input into a benefit."""

    if is_blank(value):
        return 0.0, False
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)) or float(value) < 0:
            return 0.0, False
    if isinstance(value, str):
        match = re.match(
            r"^(\d+):(\d{2}):(\d{2})(?:\.(\d+))?$", value.strip()
        )
        if not match:
            return 0.0, False
        _, minutes, seconds, _ = match.groups()
        if int(minutes) >= 60 or int(seconds) >= 60:
            return 0.0, False
    try:
        seconds = ticket_time_to_seconds(value)
    except (TypeError, ValueError, OverflowError):
        return 0.0, False
    return (
        (seconds, True)
        if math.isfinite(seconds) and seconds >= 0
        else (0.0, False)
    )


def parse_date_text(text: str) -> date:
    text = text.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Could not parse report date: {text}")


def parse_report_date_with_source(df: pd.DataFrame, path: Path) -> tuple[date, str]:
    for value in df.to_numpy().ravel():
        if isinstance(value, str) and "Date(s):" in value:
            match = re.search(r"Date\(s\):\s*([^\n]+)", value)
            if match:
                date_text = match.group(1).split("-")[-1].strip()
                return parse_date_text(date_text), "Workbook Date(s) field"
    match = re.search(r"(\d{2})-(\d{2})-(\d{4})", path.name)
    if match:
        month, day, year = map(int, match.groups())
        return (
            date(year, month, day)
            - timedelta(days=FILENAME_DATE_BUSINESS_DATE_OFFSET_DAYS),
            "Filename fallback minus one business-date day",
        )
    raise ValueError(f"Could not find report date in {path.name}")


def parse_report_date(df: pd.DataFrame, path: Path) -> date:
    return parse_report_date_with_source(df, path)[0]


def find_header_row(df: pd.DataFrame) -> tuple[int, int, list[int]]:
    for row_index, row in df.iterrows():
        values = ["" if is_blank(v) else str(v).strip() for v in row.tolist()]
        if "Store" in values and "Gross Sales" in values:
            location_col = values.index("Store")
            block_starts = [idx for idx, value in enumerate(values) if value == "Gross Sales"]
            if block_starts:
                return row_index, location_col, block_starts
    raise ValueError("Could not find the metric header row.")


def display_name_for(raw_name: str, config: dict[str, Any]) -> str:
    aliases = config.get("public_name_aliases", {})
    return aliases.get(raw_name, raw_name)


def excel_safe_text(value: Any) -> Any:
    """Neutralize strings that spreadsheet applications may interpret as formulas."""
    if isinstance(value, str) and value.startswith(EXCEL_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def excel_safe_cell_value(cell) -> Any:
    """Return a non-executable value, including for openpyxl formula objects."""
    value = cell.value
    if cell.data_type != "f":
        return excel_safe_text(value)
    formula_text = value if isinstance(value, str) else getattr(value, "text", None)
    if not isinstance(formula_text, str) or not formula_text:
        return "[Blocked spreadsheet formula]"
    return excel_safe_text(formula_text)


def is_numbered_server_placeholder(name: Any) -> bool:
    text = "" if is_blank(name) else str(name).strip()
    pattern = r"\d+\s+Server\d*(?:\s+Server\d*)*"
    return re.fullmatch(pattern, text, re.IGNORECASE) is not None


def daily_report_excel_engine(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return "xlrd"
    if suffix == ".xlsx":
        return "openpyxl"
    raise ValueError(
        f"Unsupported daily report format for {path.name}. Use {DAILY_REPORT_FORMAT_LABEL}."
    )


def parse_daily_report(path: Path, config: dict[str, Any]) -> list[MetricRecord]:
    parser_engine = daily_report_excel_engine(path)
    source_sha256 = sha256_file(path)
    with pd.ExcelFile(path, engine=parser_engine) as workbook:
        if "Report(All)" not in workbook.sheet_names:
            if "No Data Available" in workbook.sheet_names:
                raise ValueError(
                    "the source export contains a 'No Data Available' worksheet instead of report "
                    "data. Replace it with a complete Daily Report export, then rerun. No "
                    "workbooks were created and no source files were moved."
                )
            available_sheets = ", ".join(workbook.sheet_names) or "none"
            raise ValueError(
                "the required 'Report(All)' worksheet is missing "
                f"(worksheets found: {available_sheets}). Replace the file with a complete Red Onion "
                "Daily Report export, then rerun."
            )
        df = workbook.parse("Report(All)", header=None)
    report_date, report_date_source = parse_report_date_with_source(df, path)
    header_row, location_col, block_starts = find_header_row(df)
    header_values = [
        "" if is_blank(value) else str(value).strip()
        for value in df.iloc[header_row].tolist()
    ]
    location_names = set(config["locations"])
    records: list[MetricRecord] = []

    for _, row in df.iloc[header_row + 1 :].iterrows():
        location_value = row.get(location_col)
        if is_blank(location_value):
            continue

        location = str(location_value).strip()
        if location not in location_names:
            continue

        source_raw_user_name = "" if is_blank(row.get(1)) else str(row.get(1)).strip()
        raw_user_name = source_raw_user_name
        display_name = display_name_for(source_raw_user_name, config)

        for start_col in block_starts:
            values = [row.get(start_col + offset) for offset in range(len(METRICS))]
            check_count_col = start_col + len(METRICS)
            has_check_count = (
                check_count_col < len(header_values)
                and header_values[check_count_col] == "Check Count"
            )
            check_count_value = row.get(check_count_col) if has_check_count else None
            if all(is_blank(value) for value in values) and is_blank(check_count_value):
                continue

            gross_sales = to_float(values[0])
            guest_count = to_float(values[1])
            wine_sales = to_float(values[3])
            rate, rate_available = optional_nonnegative_float(values[4])
            ticket_seconds, ticket_time_available = optional_ticket_time_seconds(
                values[5]
            )
            check_count, check_count_available = (
                optional_check_count(check_count_value)
                if has_check_count
                else (0.0, False)
            )

            if (
                raw_user_name
                and gross_sales == 0
                and guest_count == 0
                and wine_sales == 0
                and not (check_count_available and check_count > 0)
            ):
                continue

            check_average = gross_sales / guest_count if guest_count else 0.0
            wine_pct = wine_sales / gross_sales if gross_sales else 0.0
            records.append(
                MetricRecord(
                    source_file=excel_safe_text(path.name),
                    report_date=report_date,
                    location=location,
                    raw_user_name=raw_user_name,
                    display_name=display_name,
                    is_location_total=source_raw_user_name == "",
                    gross_sales=gross_sales,
                    guest_count=guest_count,
                    check_average=check_average,
                    wine_sales=wine_sales,
                    wine_pct=wine_pct,
                    rate_of_sale_by_guest_count=rate,
                    average_ticket_time_seconds=ticket_seconds,
                    source_sha256=source_sha256,
                    source_format=path.suffix.lower(),
                    parser_engine=parser_engine,
                    report_date_source=report_date_source,
                    rate_available=(
                        guest_count <= 0 or (rate_available and rate > 0)
                    ),
                    ticket_time_available=(
                        (guest_count <= 0 and check_count <= 0)
                        or ticket_time_available
                    ),
                    check_count=check_count,
                    check_count_available=check_count_available,
                )
            )
            break

    return records


def is_daily_report_name(name: str) -> bool:
    path = Path(name)
    return (
        not path.name.startswith("~$")
        and path.name.casefold().startswith(DAILY_REPORT_PREFIX)
        and path.suffix.lower() in DAILY_REPORT_EXTENSIONS
    )


def is_daily_report_path(path: Path) -> bool:
    return path.is_file() and is_daily_report_name(path.name)


def find_daily_report_paths(root: Path, *, recursive: bool) -> list[Path]:
    if not root.exists() or not root.is_dir():
        return []
    candidates = root.rglob("*") if recursive else root.iterdir()
    return sorted(
        (path for path in candidates if is_daily_report_path(path)),
        key=lambda path: str(path).casefold(),
    )


def daily_report_paths(input_dir: Path) -> list[Path]:
    return find_daily_report_paths(input_dir, recursive=False)


def path_is_link_or_reparse(path: Path) -> bool:
    """Return whether a directory entry redirects through a link/reparse point."""

    try:
        metadata = os.lstat(path)
    except FileNotFoundError:
        return False
    return stat.S_ISLNK(metadata.st_mode) or bool(
        getattr(metadata, "st_file_attributes", 0) & 0x400
    )


def managed_subdirectory(
    root: Path,
    *parts: str,
    purpose: str,
    create: bool,
) -> Path:
    """Traverse/create managed directories without accepting links or junctions."""

    current = root.resolve()
    for part in parts:
        if not part or Path(part).name != part:
            raise IntegrityError(f"Unsafe managed {purpose} directory component: {part!r}")
        candidate = current / part
        if os.path.lexists(candidate):
            if path_is_link_or_reparse(candidate):
                raise IntegrityError(
                    f"Refusing {purpose} directory link or Windows junction: {candidate}"
                )
            if not candidate.is_dir():
                raise IntegrityError(f"Managed {purpose} path is not a directory: {candidate}")
        elif create:
            candidate.mkdir()
        current = candidate
    return current


def validate_managed_tree_no_reparse(root: Path, *, purpose: str) -> None:
    """Reject every link/reparse entry below a managed inventory root."""

    if not os.path.lexists(root):
        return
    if path_is_link_or_reparse(root) or not root.is_dir():
        raise IntegrityError(f"Managed {purpose} root is unsafe: {root}")
    for current, directory_names, file_names in os.walk(root, followlinks=False):
        current_path = Path(current)
        for name in [*directory_names, *file_names]:
            candidate = current_path / name
            if path_is_link_or_reparse(candidate):
                raise IntegrityError(
                    f"Refusing {purpose} link or Windows reparse point: {candidate}"
                )


def managed_direct_child(
    root: Path,
    supplied: Path,
    *,
    purpose: str,
    require_file: bool,
) -> Path:
    """Return a direct child without ever resolving a managed leaf through a link."""

    root = root.resolve()
    supplied = Path(supplied)
    try:
        parent = supplied.parent.resolve()
    except OSError as exc:
        raise IntegrityError(f"Could not validate {purpose}: {supplied}") from exc
    if parent != root or supplied.name in {"", ".", ".."}:
        raise IntegrityError(f"Refusing {purpose} outside its managed folder: {supplied}")
    candidate = root / supplied.name
    if os.path.lexists(candidate) and path_is_link_or_reparse(candidate):
        raise IntegrityError(
            f"Refusing {purpose} link or Windows reparse point: {candidate.name}. "
            "Replace it with a normal file inside the managed folder."
        )
    if require_file and not candidate.is_file():
        raise IntegrityError(f"Managed {purpose} file is missing: {candidate}")
    if os.path.lexists(candidate) and not candidate.is_file():
        raise IntegrityError(f"Managed {purpose} entry is not a regular file: {candidate}")
    return candidate


def managed_master_workbook_path(output_dir: Path) -> Path:
    """Return the master leaf only when it is a normal managed output entry."""

    return managed_direct_child(
        output_dir,
        output_dir / "Red_Onion_Server_Master.xlsx",
        purpose="master workbook",
        require_file=False,
    )


def _file_identity(metadata: os.stat_result) -> tuple[int, int, int | None, int | None]:
    return (
        metadata.st_size,
        metadata.st_mtime_ns,
        getattr(metadata, "st_dev", None),
        getattr(metadata, "st_ino", None),
    )


def capture_regular_file(source: Path, *, fingerprint_path: str) -> CapturedActiveInput:
    """Capture one regular file while proving its directory entry stayed stable."""

    before = os.lstat(source)
    if path_is_link_or_reparse(source) or not stat.S_ISREG(before.st_mode):
        raise IntegrityError(f"Managed source is not a normal file: {source}")
    digest = hashlib.sha256()
    chunks: list[bytes] = []
    with source.open("rb") as handle:
        opened = os.fstat(handle.fileno())
        if _file_identity(opened) != _file_identity(before):
            raise IntegrityError(f"Managed source changed while opening: {source.name}")
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
            chunks.append(chunk)
    after = os.lstat(source)
    if path_is_link_or_reparse(source) or _file_identity(after) != _file_identity(before):
        raise IntegrityError(
            f"Managed source changed while it was captured: {source.name}. Rerun after "
            "Dropbox has finished syncing."
        )
    content = b"".join(chunks)
    return CapturedActiveInput(
        source,
        FileFingerprint(
            path=fingerprint_path,
            size=len(content),
            sha256=digest.hexdigest(),
        ),
        content,
    )


def managed_recursive_file(roots: Iterable[Path], supplied: Path, *, purpose: str) -> Path:
    """Validate a recursive managed source without resolving through reparse points."""

    candidate = supplied.absolute()
    matching_root: Path | None = None
    relative: Path | None = None
    for root in roots:
        resolved_root = root.resolve()
        try:
            relative = candidate.relative_to(resolved_root)
            matching_root = resolved_root
            break
        except ValueError:
            continue
    if matching_root is None or relative is None or not relative.parts:
        raise IntegrityError(f"Refusing {purpose} outside its declared source folder: {supplied}")
    cursor = matching_root
    for part in relative.parts:
        cursor = cursor / part
        if os.path.lexists(cursor) and path_is_link_or_reparse(cursor):
            raise IntegrityError(
                f"Refusing {purpose} link or Windows reparse point: {cursor}"
            )
    if not candidate.is_file():
        raise IntegrityError(f"Managed {purpose} file is missing: {candidate}")
    return candidate


def regular_file_without_reparse_ancestors(supplied: Path, *, purpose: str) -> Path:
    """Validate an arbitrary lexical path without following a linked ancestor."""

    candidate = Path(supplied).absolute()
    cursor = Path(candidate.anchor)
    for part in candidate.parts[1:]:
        cursor = cursor / part
        if os.path.lexists(cursor) and path_is_link_or_reparse(cursor):
            raise IntegrityError(
                f"Refusing {purpose} link or Windows reparse point: {cursor}"
            )
    if not candidate.is_file():
        raise IntegrityError(f"Managed {purpose} file is missing: {candidate}")
    return candidate


def capture_migration_inputs(
    source_paths: Iterable[Path], source_roots: Iterable[Path]
) -> list[CapturedActiveInput]:
    roots = [root.resolve() for root in source_roots]
    captures: list[CapturedActiveInput] = []
    for supplied in source_paths:
        source = managed_recursive_file(roots, supplied, purpose="history migration source")
        captures.append(
            capture_regular_file(source, fingerprint_path=source.name)
        )
    return captures


def capture_active_inputs(
    source_paths: Iterable[Path], input_dir: Path
) -> list[CapturedActiveInput]:
    """Pin the exact bytes used for parsing, archiving, manifesting, and cleanup."""

    captures: list[CapturedActiveInput] = []
    seen_names: set[str] = set()
    for supplied in source_paths:
        source = managed_direct_child(
            input_dir, supplied, purpose="active input", require_file=True
        )
        key = source.name.casefold()
        if key in seen_names:
            raise IntegrityError(f"Duplicate active input name: {source.name}")
        seen_names.add(key)

        captures.append(capture_regular_file(source, fingerprint_path=source.name))
    return captures


def verify_captured_active_inputs(
    captures: Iterable[CapturedActiveInput], input_dir: Path
) -> None:
    """Require every live pathname to remain the exact captured input version."""

    for capture in captures:
        source = managed_direct_child(
            input_dir, capture.source, purpose="active input", require_file=True
        )
        if source.stat().st_size != capture.fingerprint.size:
            raise IntegrityError(
                f"Active input was replaced during the run: {source.name}. "
                "No active source files were deleted."
            )
        if sha256_file(source) != capture.fingerprint.sha256:
            raise IntegrityError(
                f"Active input was replaced during the run: {source.name}. "
                "No active source files were deleted."
            )


def verify_captured_migration_inputs(
    captures: Iterable[CapturedActiveInput],
) -> None:
    """Stop a migration when its external source changed after plan construction."""

    for capture in captures:
        source = capture.source.absolute()
        if (
            path_is_link_or_reparse(source)
            or not source.is_file()
            or source.stat().st_size != capture.fingerprint.size
            or sha256_file(source) != capture.fingerprint.sha256
        ):
            raise IntegrityError(
                f"History migration source changed during the run: {source.name}. "
                "No migration files were committed; rebuild the plan and rerun."
            )


def canonical_daily_archive_dir(archive_dir: Path) -> Path:
    return managed_subdirectory(
        archive_dir,
        CANONICAL_DAILY_ARCHIVE_FOLDER,
        purpose="canonical raw archive",
        create=False,
    )


def archived_daily_report_paths(archive_dir: Path) -> list[Path]:
    return find_daily_report_paths(canonical_daily_archive_dir(archive_dir), recursive=True)


def capture_archived_report_inputs(
    archive_dir: Path,
    *,
    expected_inventory: Iterable[FileFingerprint] | None = None,
) -> Iterable[CapturedActiveInput]:
    """Yield manifest-pinned history captures so parsing retains one payload at a time."""

    raw_root = canonical_daily_archive_dir(archive_dir)
    if expected_inventory is None:
        for supplied in archived_daily_report_paths(archive_dir):
            source = managed_recursive_file(
                [raw_root], supplied, purpose="canonical archived daily report"
            )
            relative = source.absolute().relative_to(raw_root.resolve()).as_posix()
            yield capture_regular_file(source, fingerprint_path=relative)
        return

    expected_reports = sorted(
        (
            fingerprint
            for fingerprint in expected_inventory
            if is_daily_report_name(Path(fingerprint.path).name)
        ),
        key=lambda fingerprint: fingerprint.path.casefold(),
    )
    for expected in expected_reports:
        supplied = raw_root.joinpath(*expected.path.split("/"))
        source = managed_recursive_file(
            [raw_root], supplied, purpose="manifest-recorded archived daily report"
        )
        capture = capture_regular_file(source, fingerprint_path=expected.path)
        if capture.fingerprint != expected:
            raise IntegrityError(
                f"Archived daily report changed before it could be pinned: {source.name}. "
                "No outputs or archive files were changed."
            )
        yield capture
        del capture


def read_captured_reports_by_path(
    captures: Iterable[CapturedActiveInput], config: dict[str, Any]
) -> dict[Path, list[MetricRecord]]:
    """Parse staged captures one at a time, releasing each byte payload after use."""

    records_by_path: dict[Path, list[MetricRecord]] = {}
    with tempfile.TemporaryDirectory(prefix=".archived-report-read-") as stage_name:
        stage_root = Path(stage_name)
        for index, capture in enumerate(captures):
            staged_dir = stage_root / f"{index:06d}"
            staged_dir.mkdir()
            staged_path = staged_dir / capture.source.name
            written_hash = verified_write_bytes(capture.content, staged_path)
            if written_hash != capture.fingerprint.sha256:
                raise IntegrityError(
                    f"Archived report staging verification failed for {capture.source.name}."
                )
            try:
                records = parse_daily_report(staged_path, config)
            except Exception as exc:
                raise ValueError(
                    f"Could not process {capture.source.name}: {exc}"
                ) from exc
            if not records:
                raise ValueError(f"No metric rows were found in {capture.source}")
            records_by_path[capture.source] = records
            del capture
    return records_by_path


def read_reports_by_path(paths: Iterable[Path], config: dict[str, Any]) -> dict[Path, list[MetricRecord]]:
    records_by_path: dict[Path, list[MetricRecord]] = {}
    for path in paths:
        try:
            records = parse_daily_report(path, config)
        except Exception as exc:
            raise ValueError(f"Could not process {path.name}: {exc}") from exc
        if not records:
            raise ValueError(f"No metric rows were found in {path}")
        records_by_path[path] = records
    return records_by_path


def report_date_for_records(path: Path, records: list[MetricRecord]) -> date:
    report_dates = sorted({record.report_date for record in records})
    if len(report_dates) != 1:
        raise ValueError(f"{path.name} contains multiple report dates: {report_dates}")
    return report_dates[0]


def semantic_report_signature(
    records: list[MetricRecord],
    *,
    include_check_count: bool = True,
) -> tuple[tuple[Any, ...], ...]:
    def normalized_number(value: float) -> float:
        number = round(float(value), 6)
        return 0.0 if number == 0 else number

    def record_signature(record: MetricRecord) -> tuple[Any, ...]:
        signature: tuple[Any, ...] = (
            record.location.strip(),
            record.raw_user_name.strip(),
            bool(record.is_location_total),
            normalized_number(record.gross_sales),
            normalized_number(record.guest_count),
            normalized_number(record.wine_sales),
            normalized_number(record.rate_of_sale_by_guest_count),
            normalized_number(record.average_ticket_time_seconds),
            bool(record.rate_available),
            bool(record.ticket_time_available),
        )
        if include_check_count:
            signature += (
                normalized_number(record.check_count),
                bool(record.check_count_available),
            )
        return signature

    signature_records = (
        records
        if include_check_count
        else [
            record
            for record in records
            if not (
                record.check_count_available
                and record.check_count > 0
                and record.gross_sales == 0
                and record.guest_count == 0
                and record.wine_sales == 0
            )
        ]
    )
    return tuple(sorted(record_signature(record) for record in signature_records))


def resolve_report_duplicates(
    records_by_path: dict[Path, list[MetricRecord]],
) -> ReportResolution:
    reports_by_date: dict[date, list[tuple[Path, list[MetricRecord]]]] = defaultdict(list)
    for path, records in records_by_path.items():
        reports_by_date[report_date_for_records(path, records)].append((path, records))

    resolved: dict[Path, list[MetricRecord]] = {}
    duplicates: list[Path] = []
    for report_date, reports in sorted(reports_by_date.items()):
        reports.sort(key=lambda item: str(item[0]).casefold())
        core_signatures = {
            semantic_report_signature(records, include_check_count=False)
            for _, records in reports
        }
        conflicting_paths: list[Path] = []
        if len(core_signatures) == 1:
            # A newer export can contain the same daily facts plus Check Count.
            # Prefer the most complete copy instead of treating additive
            # context as a conflict. Conflicting values at the same coverage
            # level still fail closed.
            richest_coverage = max(
                sum(record.check_count_available for record in records)
                for _, records in reports
            )
            richest_reports = [
                (path, records)
                for path, records in reports
                if sum(record.check_count_available for record in records)
                == richest_coverage
            ]
            selected_path, selected_records = richest_reports[0]
            selected_signature = semantic_report_signature(selected_records)
            conflicting_paths = [
                path
                for path, records in richest_reports[1:]
                if semantic_report_signature(records) != selected_signature
            ]
        else:
            selected_path, selected_records = reports[0]
            selected_core_signature = semantic_report_signature(
                selected_records, include_check_count=False
            )
            conflicting_paths = [
                path
                for path, records in reports[1:]
                if semantic_report_signature(records, include_check_count=False)
                != selected_core_signature
            ]
        if conflicting_paths:
            paths = [selected_path, *conflicting_paths]
            path_lines = "\n".join(f"  - {path}" for path in paths)
            raise ValueError(
                "Conflicting daily reports were found for business date "
                f"{report_date.isoformat()}. These files contain different metric data:\n"
                f"{path_lines}\n"
                "Keep only the correct source or reconcile the reports before rerunning. "
                "No history files were copied, no workbooks were created, and no active "
                "source files were moved."
            )
        resolved[selected_path] = selected_records
        duplicates.extend(path for path, _ in reports if path != selected_path)

    return ReportResolution(
        records_by_path=resolved,
        duplicate_paths=tuple(duplicates),
        business_dates=tuple(sorted(reports_by_date)),
    )


def flatten_report_records(records_by_path: dict[Path, list[MetricRecord]]) -> list[MetricRecord]:
    records: list[MetricRecord] = []
    for path_records in records_by_path.values():
        records.extend(path_records)
    return records


def read_all_reports(input_dir: Path, config: dict[str, Any]) -> list[MetricRecord]:
    paths = daily_report_paths(input_dir)
    if not paths:
        raise FileNotFoundError(
            f"No daily reports ({DAILY_REPORT_FORMAT_LABEL}) found in {input_dir}"
        )

    resolution = resolve_report_duplicates(read_reports_by_path(paths, config))
    return flatten_report_records(resolution.records_by_path)


def is_operating_day(day: date) -> bool:
    return OPERATING_WEEK_START_WEEKDAY <= day.weekday() <= OPERATING_WEEK_END_WEEKDAY


def week_period_for(day: date) -> tuple[date, date]:
    days_since_start = (day.weekday() - OPERATING_WEEK_START_WEEKDAY) % 7
    week_start = day - timedelta(days=days_since_start)
    return week_start, week_start + timedelta(days=OPERATING_WEEK_DAYS - 1)


def active_week_for_paths(records_by_path: dict[Path, list[MetricRecord]]) -> tuple[date, date]:
    groups: dict[tuple[date, date], list[Path]] = defaultdict(list)
    for path, path_records in records_by_path.items():
        report_dates = sorted({record.report_date for record in path_records})
        if len(report_dates) != 1:
            raise ValueError(f"{path.name} contains multiple report dates: {report_dates}")
        groups[week_period_for(report_dates[0])].append(path)

    if not groups:
        raise ValueError("No active daily report files were parsed.")

    if len(groups) > 1:
        lines = [
            "The daily-report drop folder contains files from more than one Tuesday-Sunday operating week.",
            "Move the extra files out of 01 Daily Reports - Drop Here and rerun:",
        ]
        for (_, week_end), paths in sorted(groups.items()):
            lines.append(f"  week-ending-{week_end.isoformat()}:")
            for path in sorted(paths):
                lines.append(f"    {path.name}")
        raise ValueError("\n".join(lines))

    return next(iter(groups))


def selected_public_dates(
    records: Iterable[MetricRecord], week_start: str | None, week_end: str | None
) -> tuple[date, date]:
    all_dates = sorted({record.report_date for record in records})
    if not all_dates:
        raise ValueError("No report dates were found.")
    operating_dates = [report_date for report_date in all_dates if is_operating_day(report_date)]
    if not operating_dates:
        raise ValueError("No Tuesday-Sunday operating report dates were found.")

    try:
        if week_start and week_end:
            start, end = date.fromisoformat(week_start), date.fromisoformat(week_end)
        elif week_start:
            start = date.fromisoformat(week_start)
            _, end = week_period_for(start)
        elif week_end:
            end = date.fromisoformat(week_end)
            start, _ = week_period_for(end)
        else:
            start, end = week_period_for(max(operating_dates))

        invalid_range = (
            start > end
            or not is_operating_day(start)
            or not is_operating_day(end)
            or week_period_for(start) != week_period_for(end)
        )
    except OverflowError as exc:
        raise ValueError(
            "Public snapshot dates must be ordered operating days within one "
            f"{OPERATING_WEEK_LABEL} operating week; received "
            f"{week_start or 'automatic'} through {week_end or 'automatic'}."
        ) from exc
    if invalid_range:
        raise ValueError(
            "Public snapshot dates must be ordered operating days within one "
            f"{OPERATING_WEEK_LABEL} operating week; received "
            f"{start.isoformat()} through {end.isoformat()}."
        )
    return start, end


def operating_day_count(start: date, end: date) -> int:
    if start > end:
        return 0
    total_days = (end - start).days + 1
    full_weeks, remainder = divmod(total_days, 7)
    return full_weeks * OPERATING_WEEK_DAYS + sum(
        is_operating_day(start + timedelta(days=offset)) for offset in range(remainder)
    )


def validated_data_quality_coverage(
    records: Iterable[MetricRecord], public_start: date, public_end: date
) -> tuple[date, date, int]:
    report_dates = [record.report_date for record in records]
    if not report_dates:
        raise ValueError("No report dates were found for Data Quality coverage.")
    coverage_start = min(min(report_dates), public_start)
    coverage_end = max(max(report_dates), public_end)
    row_count = operating_day_count(coverage_start, coverage_end)
    if row_count > MAX_DATA_QUALITY_DATE_ROWS:
        raise ValueError(
            "Data Quality date coverage from "
            f"{coverage_start.isoformat()} through {coverage_end.isoformat()} would create "
            f"{row_count:,} operating-day rows, above the safe limit of "
            f"{MAX_DATA_QUALITY_DATE_ROWS:,}. Reconcile the outlier report date and rerun. "
            "No history files were copied, no workbooks were created, and no active "
            "source files were moved."
        )
    return coverage_start, coverage_end, row_count


def validate_daily_location_reconciliation(
    records: Iterable[MetricRecord],
    config: dict[str, Any],
) -> None:
    """Require daily person rows to reconcile to each configured location total."""

    grouped: dict[tuple[date, str], list[MetricRecord]] = defaultdict(list)
    for record in records:
        if is_operating_day(record.report_date):
            grouped[(record.report_date, record.location)].append(record)
    failures: list[str] = []
    for (report_date, location), rows in sorted(grouped.items()):
        if location not in config.get("locations", {}):
            continue
        totals = [row for row in rows if row.is_location_total]
        people = [row for row in rows if not row.is_location_total]
        if len(totals) != 1:
            failures.append(
                f"{report_date.isoformat()} {location}: expected one location total, "
                f"found {len(totals)}"
            )
            continue
        total = totals[0]
        if any(row.check_count_available for row in rows):
            missing_check_rows = [
                row.raw_user_name for row in rows if not row.check_count_available
            ]
            if missing_check_rows:
                preview = ", ".join(missing_check_rows[:3])
                failures.append(
                    f"{report_date.isoformat()} {location}: Check Count is only "
                    f"partially available (missing for {preview})"
                )
            else:
                person_checks = sum(row.check_count for row in people)
                if abs(person_checks - total.check_count) > 1e-6:
                    failures.append(
                        f"{report_date.isoformat()} {location}: check count person rows "
                        f"{person_checks:.4f} != location total {total.check_count:.4f}"
                    )
        comparisons = (
            (
                "guest count",
                sum(row.guest_count for row in people),
                total.guest_count,
                1e-6,
            ),
            (
                "gross sales",
                sum(row.gross_sales for row in people),
                total.gross_sales,
                0.01,
            ),
            (
                "wine sales",
                sum(row.wine_sales for row in people),
                total.wine_sales,
                0.01,
            ),
        )
        for label, person_value, total_value, tolerance in comparisons:
            if abs(person_value - total_value) > tolerance:
                failures.append(
                    f"{report_date.isoformat()} {location}: {label} person rows "
                    f"{person_value:.4f} != location total {total_value:.4f}"
                )
    if failures:
        detail = "; ".join(failures[:5])
        raise ValueError(
            "Daily reconciliation failed. No history files were copied, no workbooks "
            f"were created, and no active source files were moved. {detail}"
        )


def archive_destination_for(
    source_path: Path,
    archive_dir: Path,
    reserved_destinations: set[Path] | None = None,
) -> tuple[Path, bool]:
    reserved_destinations = reserved_destinations or set()
    destination = archive_dir / source_path.name
    if not destination.exists() and destination not in reserved_destinations:
        return destination, False

    if destination.exists() and filecmp.cmp(source_path, destination, shallow=False):
        return destination, True

    counter = 1
    while True:
        candidate = archive_dir / f"{source_path.stem} ({counter}){source_path.suffix}"
        if not candidate.exists() and candidate not in reserved_destinations:
            return candidate, False
        if candidate.exists() and filecmp.cmp(source_path, candidate, shallow=False):
            return candidate, True
        counter += 1


def verified_copy_file(source: Path, destination: Path, *, replace: bool = False) -> str:
    """Copy one file through a verified same-folder temporary file."""
    source = source.absolute()
    destination = destination.absolute()
    if not source.is_file():
        raise FileNotFoundError(f"Source file was not found: {source}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    if os.path.lexists(destination) and not replace:
        raise FileExistsError(f"Copy destination already exists: {destination}")

    expected_hash = sha256_file(source)
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{destination.name}.",
            suffix=".copying",
            dir=str(destination.parent),
            delete=False,
        ) as handle:
            temp_path = Path(handle.name)
        shutil.copy2(source, temp_path)
        copied_hash = sha256_file(temp_path)
        current_source_hash = sha256_file(source)
        if copied_hash != expected_hash or current_source_hash != expected_hash:
            raise IntegrityError(
                f"File changed or failed verification while copying {source.name}; "
                "no active source files were deleted."
            )
        if os.path.lexists(destination) and not replace:
            raise FileExistsError(f"Copy destination appeared during the run: {destination}")
        if replace:
            os.replace(temp_path, destination)
        elif os.name == "nt":
            try:
                os.rename(temp_path, destination)
            except OSError as exc:
                if os.path.lexists(destination):
                    raise FileExistsError(
                        f"Copy destination appeared during the run: {destination}"
                    ) from exc
                raise
        else:
            os.link(temp_path, destination)
            temp_path.unlink()
        temp_path = None
        return expected_hash
    finally:
        if temp_path is not None and temp_path.exists():
            temp_path.unlink()


def copy_processed_files_verified(
    source_paths: Iterable[Path],
    archive_root: Path,
    week_end: date,
    *,
    anchor_dir: Path | None = None,
) -> list[VerifiedArchiveCopy]:
    """Pin and archive normal source entries without deleting them."""

    reject_legacy_raw_mutation_if_protected(
        archive_root, "legacy processed-file archiving", anchor_dir=anchor_dir
    )
    captures = [
        capture_regular_file(
            regular_file_without_reparse_ancestors(
                source_path, purpose="legacy archive source"
            ),
            fingerprint_path=Path(source_path).name,
        )
        for source_path in source_paths
    ]
    raw_root = managed_subdirectory(
        archive_root,
        CANONICAL_DAILY_ARCHIVE_FOLDER,
        purpose="canonical raw archive",
        create=True,
    )
    archive_dir = managed_subdirectory(
        raw_root,
        f"week-ending-{week_end.isoformat()}",
        purpose="raw archive week",
        create=True,
    )
    reserved: set[Path] = set()
    plans: list[tuple[CapturedActiveInput, Path, bool]] = []
    for capture in captures:
        destination, already_archived = archive_destination_for_capture(
            capture, archive_dir, reserved
        )
        plans.append((capture, destination.absolute(), already_archived))
        reserved.add(destination)

    completed: list[VerifiedArchiveCopy] = []
    try:
        for capture, destination, already_archived in plans:
            if already_archived:
                if sha256_file(destination) != capture.fingerprint.sha256:
                    raise IntegrityError(
                        f"Existing archive copy no longer matches {capture.source.name}; "
                        "no active source files were deleted."
                    )
                created = False
            else:
                copied_hash = verified_write_bytes(capture.content, destination)
                if copied_hash != capture.fingerprint.sha256:
                    raise IntegrityError(
                        f"Archive verification failed for {capture.source.name}."
                    )
                created = True
            completed.append(
                VerifiedArchiveCopy(
                    capture.source,
                    destination,
                    created,
                    capture.fingerprint.sha256,
                )
            )
    except Exception:
        created = [copy for copy in completed if copy.created]
        conflicts = rollback_created_files(
            [copy.destination for copy in created],
            archive_dir,
            expected_hashes={copy.destination: copy.sha256 for copy in created},
        )
        if conflicts:
            raise IntegrityError(
                "Archive rollback preserved a file that changed during the failed copy: "
                f"{conflicts[0]}"
            )
        raise
    return completed


def verified_write_bytes(content: bytes, destination: Path, *, replace: bool = False) -> str:
    """Write pinned bytes atomically and verify the exact persisted content."""

    destination = destination.absolute()
    destination.parent.mkdir(parents=True, exist_ok=True)
    if os.path.lexists(destination) and not replace:
        raise FileExistsError(f"Write destination already exists: {destination}")
    expected_hash = hashlib.sha256(content).hexdigest()
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{destination.name}.",
            suffix=".writing",
            dir=str(destination.parent),
            delete=False,
        ) as handle:
            temp_path = Path(handle.name)
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        if sha256_file(temp_path) != expected_hash:
            raise IntegrityError(f"Pinned-byte write verification failed for {destination.name}.")
        if os.path.lexists(destination) and not replace:
            raise FileExistsError(f"Write destination appeared during the run: {destination}")
        if replace:
            os.replace(temp_path, destination)
        elif os.name == "nt":
            try:
                os.rename(temp_path, destination)
            except OSError as exc:
                if os.path.lexists(destination):
                    raise FileExistsError(
                        f"Write destination appeared during the run: {destination}"
                    ) from exc
                raise
        else:
            os.link(temp_path, destination)
            temp_path.unlink()
        temp_path = None
        if sha256_file(destination) != expected_hash:
            raise IntegrityError(f"Pinned-byte write changed after publication: {destination.name}.")
        return expected_hash
    finally:
        if temp_path is not None and temp_path.exists():
            temp_path.unlink()


def archive_destination_for_capture(
    capture: CapturedActiveInput,
    archive_dir: Path,
    reserved_destinations: set[Path] | None = None,
) -> tuple[Path, bool]:
    """Choose a collision-safe archive name using the pinned input fingerprint."""

    reserved_destinations = reserved_destinations or set()
    counter = 0
    while True:
        suffix = "" if counter == 0 else f" ({counter})"
        candidate = archive_dir / (
            f"{capture.source.stem}{suffix}{capture.source.suffix}"
        )
        if candidate in reserved_destinations:
            counter += 1
            continue
        if not os.path.lexists(candidate):
            return candidate, False
        managed_direct_child(
            archive_dir, candidate, purpose="raw archive destination", require_file=True
        )
        if (
            candidate.stat().st_size == capture.fingerprint.size
            and sha256_file(candidate) == capture.fingerprint.sha256
        ):
            return candidate, True
        counter += 1


def copy_captured_active_files_verified(
    captures: Iterable[CapturedActiveInput], archive_root: Path, week_end: date
) -> list[VerifiedArchiveCopy]:
    """Archive the exact bytes that were parsed, never a later pathname version."""

    raw_root = managed_subdirectory(
        archive_root,
        CANONICAL_DAILY_ARCHIVE_FOLDER,
        purpose="canonical raw archive",
        create=True,
    )
    archive_dir = managed_subdirectory(
        raw_root,
        f"week-ending-{week_end.isoformat()}",
        purpose="raw archive week",
        create=True,
    )
    reserved: set[Path] = set()
    plans: list[tuple[CapturedActiveInput, Path, bool]] = []
    for capture in captures:
        destination, already_archived = archive_destination_for_capture(
            capture, archive_dir, reserved
        )
        plans.append((capture, destination, already_archived))
        reserved.add(destination)

    completed: list[VerifiedArchiveCopy] = []
    try:
        for capture, destination, already_archived in plans:
            if already_archived:
                created = False
            else:
                copied_hash = verified_write_bytes(capture.content, destination)
                if copied_hash != capture.fingerprint.sha256:
                    raise IntegrityError(
                        f"Pinned archive verification failed for {capture.source.name}."
                    )
                created = True
            completed.append(
                VerifiedArchiveCopy(
                    capture.source,
                    destination,
                    created,
                    capture.fingerprint.sha256,
                )
            )
    except Exception:
        created = [copy for copy in completed if copy.created]
        conflicts = rollback_created_files(
            [copy.destination for copy in created],
            archive_dir,
            expected_hashes={copy.destination: copy.sha256 for copy in created},
        )
        if conflicts:
            raise IntegrityError(
                "Pinned archive rollback preserved a changed replacement: "
                f"{conflicts[0]}"
            )
        raise
    return completed


def quarantine_and_delete_captured_inputs(
    captures: Iterable[CapturedActiveInput],
    copies: Iterable[VerifiedArchiveCopy],
    input_dir: Path,
    run_id: str,
) -> None:
    """Atomically move verified inputs aside before unlinking their exact entries."""

    captures_by_name = {capture.source.name.casefold(): capture for capture in captures}
    for copy in copies:
        capture = captures_by_name[copy.source.name.casefold()]
        if sha256_file(copy.destination) != capture.fingerprint.sha256:
            raise IntegrityError(
                f"Archive copy changed before cleanup: {copy.destination.name}. "
                "The active source was not deleted."
            )
        verify_captured_active_inputs([capture], input_dir)
        source = managed_direct_child(
            input_dir, capture.source, purpose="active input", require_file=True
        )
        quarantine = input_dir.resolve() / (
            f".{source.name}.{run_id}.{uuid.uuid4().hex}.processed"
        )
        if os.path.lexists(quarantine):
            raise IntegrityError(f"Unexpected cleanup quarantine collision: {quarantine.name}")
        os.replace(source, quarantine)
        try:
            quarantine_is_expected = (
                not path_is_link_or_reparse(quarantine)
                and quarantine.is_file()
                and quarantine.stat().st_size == capture.fingerprint.size
                and sha256_file(quarantine) == capture.fingerprint.sha256
            )
            if not quarantine_is_expected:
                if not os.path.lexists(source):
                    os.replace(quarantine, source)
                raise IntegrityError(
                    f"Active input changed at cleanup time: {source.name}. "
                    "The replacement was not deleted."
                )
            quarantine.unlink()
        except OSError as exc:
            raise IntegrityError(
                f"The run was committed, but active source cleanup failed for "
                f"{source.name}: {exc}. The archive and generated outputs are intact."
            ) from exc


def delete_verified_active_sources(copies: Iterable[VerifiedArchiveCopy]) -> None:
    """Quarantine exact source entries before deletion; never unlink a resolved target."""

    copies = list(copies)
    for copy in copies:
        source = regular_file_without_reparse_ancestors(
            copy.source, purpose="legacy active source"
        )
        if sha256_file(source) != copy.sha256:
            raise IntegrityError(
                f"Active source changed before cleanup: {source.name}. "
                "The verified outputs and archive remain, but the source was not deleted."
            )
        destination = regular_file_without_reparse_ancestors(
            copy.destination, purpose="legacy archive destination"
        )
        if sha256_file(destination) != copy.sha256:
            raise IntegrityError(
                f"Archive copy changed before cleanup: {destination.name}. "
                "The active source was not deleted."
            )
    for copy in copies:
        source = regular_file_without_reparse_ancestors(
            copy.source, purpose="legacy active source"
        )
        quarantine = source.parent / (
            f".{source.name}.{uuid.uuid4().hex}.processed"
        )
        if os.path.lexists(quarantine):
            raise IntegrityError(
                f"Unexpected active-source quarantine collision: {quarantine.name}"
            )
        try:
            os.replace(source, quarantine)
            if (
                path_is_link_or_reparse(quarantine)
                or not quarantine.is_file()
                or sha256_file(quarantine) != copy.sha256
            ):
                if not os.path.lexists(source):
                    os.replace(quarantine, source)
                raise IntegrityError(
                    f"Active source changed at cleanup time: {source.name}. "
                    "The replacement was not deleted."
                )
            quarantine.unlink()
        except OSError as exc:
            raise IntegrityError(
                f"The run was committed, but active source cleanup failed for "
                f"{source.name}: {exc}. The archive and generated outputs are intact; "
                "remove the duplicate active file after confirming it is closed."
            ) from exc


def archive_processed_files(
    source_paths: Iterable[Path],
    archive_root: Path,
    week_end: date,
    *,
    anchor_dir: Path | None = None,
) -> list[Path]:
    copies = copy_processed_files_verified(
        source_paths, archive_root, week_end, anchor_dir=anchor_dir
    )
    delete_verified_active_sources(copies)
    return [copy.destination for copy in copies]


def migration_daily_report_paths(source_dirs: Iterable[Path]) -> list[Path]:
    paths: dict[Path, None] = {}
    for source_dir in source_dirs:
        source_dir = source_dir.resolve()
        if not source_dir.exists() or not source_dir.is_dir():
            raise FileNotFoundError(f"History migration source folder was not found: {source_dir}")
        source_paths = find_daily_report_paths(source_dir, recursive=True)
        if not source_paths:
            raise FileNotFoundError(
                f"No daily reports ({DAILY_REPORT_FORMAT_LABEL}) were found under history "
                f"migration source: {source_dir}"
            )
        for path in source_paths:
            safe_path = managed_recursive_file(
                [source_dir], path, purpose="history migration source"
            )
            paths[safe_path] = None
    return sorted(paths, key=lambda path: str(path).casefold())


def build_history_migration_plan(
    source_dirs: Iterable[Path],
    archive_root: Path,
    config: dict[str, Any],
    *,
    expected_raw_inventory: Iterable[FileFingerprint] | None = None,
) -> HistoryMigrationPlan:
    source_dirs = [Path(path).resolve() for path in source_dirs]
    source_paths = migration_daily_report_paths(source_dirs)
    captured_sources = capture_migration_inputs(source_paths, source_dirs)
    source_records_by_path = read_captured_reports_by_path(captured_sources, config)
    captures_by_path = {
        capture.source.absolute(): capture for capture in captured_sources
    }
    canonical_captures = capture_archived_report_inputs(
        archive_root, expected_inventory=expected_raw_inventory
    )
    canonical_records_by_path = read_captured_reports_by_path(
        canonical_captures, config
    )

    combined_records_by_path = dict(canonical_records_by_path)
    combined_records_by_path.update(source_records_by_path)
    resolution = resolve_report_duplicates(combined_records_by_path)

    # Canonical history is the provenance boundary for dates that have already
    # been archived. Duplicate resolution may prefer a migration source with
    # richer optional context, but that source is intentionally not copied for
    # an existing date. Keep the validated canonical report as the effective
    # input so an unarchived staging file cannot influence rebuilt outputs.
    canonical_resolution = resolve_report_duplicates(canonical_records_by_path)
    effective_records_by_path = dict(resolution.records_by_path)
    effective_paths_by_date = {
        report_date_for_records(path, records): path
        for path, records in effective_records_by_path.items()
    }
    for canonical_path, canonical_records in canonical_resolution.records_by_path.items():
        canonical_date = report_date_for_records(canonical_path, canonical_records)
        effective_path = effective_paths_by_date[canonical_date]
        effective_records_by_path.pop(effective_path)
        effective_records_by_path[canonical_path] = canonical_records
        effective_paths_by_date[canonical_date] = canonical_path

    duplicate_paths = tuple(
        sorted(
            (
                path
                for path in combined_records_by_path
                if path not in effective_records_by_path
            ),
            key=lambda path: (
                report_date_for_records(path, combined_records_by_path[path]),
                str(path).casefold(),
            ),
        )
    )

    canonical_dates = {
        report_date_for_records(path, records)
        for path, records in canonical_records_by_path.items()
    }
    source_paths_by_date: dict[date, list[Path]] = defaultdict(list)
    for path, records in source_records_by_path.items():
        source_paths_by_date[report_date_for_records(path, records)].append(path)

    copy_pairs: list[tuple[Path, Path]] = []
    reserved_destinations: set[Path] = set()
    for report_date, paths in sorted(source_paths_by_date.items()):
        if report_date in canonical_dates:
            continue
        source_path = min(
            paths,
            key=lambda path: (path.suffix.lower() != ".xlsx", str(path).casefold()),
        )
        _, week_end = week_period_for(report_date)
        destination_dir = managed_subdirectory(
            canonical_daily_archive_dir(archive_root),
            f"week-ending-{week_end.isoformat()}",
            purpose="raw archive week",
            create=False,
        )
        destination, already_present = archive_destination_for_capture(
            captures_by_path[source_path.absolute()],
            destination_dir,
            reserved_destinations,
        )
        if not already_present:
            copy_pairs.append((source_path, destination))
            reserved_destinations.add(destination)

    return HistoryMigrationPlan(
        copy_pairs=tuple(copy_pairs),
        effective_records_by_path=effective_records_by_path,
        duplicate_paths=duplicate_paths,
        business_dates=resolution.business_dates,
        captured_sources=tuple(captured_sources),
    )


def apply_history_migration_plan(plan: HistoryMigrationPlan) -> tuple[Path, ...]:
    destinations = [destination for _, destination in plan.copy_pairs]
    if len(set(destinations)) != len(destinations):
        raise RuntimeError(
            "History migration preflight selected the same destination more than once. "
            "No files were copied."
        )
    conflicts = [destination for destination in destinations if os.path.lexists(destination)]
    if conflicts:
        raise RuntimeError(
            f"History migration destination appeared after preflight: {conflicts[0]}. "
            "No files were copied; rerun the migration."
        )

    captures_by_path = {
        capture.source.absolute(): capture for capture in plan.captured_sources
    }
    copied_paths: list[Path] = []
    copied_hashes: dict[Path, str] = {}
    try:
        for source_path, destination in plan.copy_pairs:
            raw_root = destination.parent.parent
            safe_raw_root = managed_subdirectory(
                raw_root.parent,
                raw_root.name,
                purpose="canonical raw archive",
                create=True,
            )
            safe_week = managed_subdirectory(
                safe_raw_root,
                destination.parent.name,
                purpose="raw archive week",
                create=True,
            )
            if safe_week / destination.name != destination.absolute():
                raise IntegrityError(
                    f"History migration destination escaped the raw archive: {destination}"
                )
            capture = captures_by_path.get(source_path.absolute())
            if capture is None:
                copied_hash = verified_copy_file(source_path, destination)
            else:
                copied_hash = verified_write_bytes(capture.content, destination)
                if copied_hash != capture.fingerprint.sha256:
                    raise IntegrityError(
                        f"Pinned history migration copy failed for {source_path.name}."
                    )
            copied_paths.append(destination)
            copied_hashes[destination.absolute()] = copied_hash
    except Exception:
        conflicts = rollback_created_files(
            copied_paths,
            destinations[0].parent.parent if destinations else Path.cwd(),
            expected_hashes=copied_hashes,
        )
        if conflicts:
            raise IntegrityError(
                "History migration rollback preserved a changed replacement: "
                f"{conflicts[0]}"
            )
        raise
    return tuple(copied_paths)


def history_migration_expected_hashes(
    plan: HistoryMigrationPlan, copied_paths: Iterable[Path]
) -> dict[Path, str]:
    captures = {
        capture.source.absolute(): capture.fingerprint.sha256
        for capture in plan.captured_sources
    }
    source_for_destination = {
        destination.absolute(): source.absolute()
        for source, destination in plan.copy_pairs
    }
    expected: dict[Path, str] = {}
    for supplied in copied_paths:
        destination = supplied.absolute()
        source = source_for_destination[destination]
        expected[destination] = captures.get(source) or sha256_file(source)
    return expected


def migrate_history_files(
    source_dirs: Iterable[Path],
    archive_root: Path,
    config: dict[str, Any],
    *,
    anchor_dir: Path | None = None,
) -> HistoryMigrationResult:
    reject_legacy_raw_mutation_if_protected(
        archive_root, "legacy history migration", anchor_dir=anchor_dir
    )
    archive_root = archive_root.resolve()
    archive_root.mkdir(parents=True, exist_ok=True)
    plan = build_history_migration_plan(source_dirs, archive_root, config)
    copied_paths = apply_history_migration_plan(plan)
    return HistoryMigrationResult(
        copied_paths=copied_paths,
        duplicate_files_ignored=len(plan.duplicate_paths),
        business_dates_considered=len(plan.business_dates),
    )


def aggregate_records(
    records: Iterable[MetricRecord], key_fields: tuple[str, ...], extra: dict[str, Any] | None = None
) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], dict[str, Any]] = {}

    for record in records:
        key = tuple(getattr(record, field) for field in key_fields)
        if key not in groups:
            groups[key] = {field: getattr(record, field) for field in key_fields}
            if extra:
                groups[key].update(extra)
            groups[key].update(
                {
                    "gross_sales": 0.0,
                    "guest_count": 0.0,
                    "wine_sales": 0.0,
                    "check_count": 0.0,
                    "rate_qualifying_sales": 0.0,
                    "rate_opportunities": 0.0,
                    "ticket_weighted_sum": 0.0,
                    "ticket_weight": 0.0,
                    "rate_available": True,
                    "ticket_time_available": True,
                    "check_count_available": True,
                    "has_active_rows": False,
                    "active_dates": set(),
                    "daily_records": [],
                    "source_files": set(),
                    "source_evidence": set(),
                }
            )

        group = groups[key]
        group["gross_sales"] += record.gross_sales
        group["guest_count"] += record.guest_count
        group["wine_sales"] += record.wine_sales
        group["check_count"] += (
            record.check_count if record.check_count_available else 0.0
        )
        # Check Count is operational context only. A check-only row must not
        # create an active day or make a person eligible for review.
        active_row = record.guest_count > 0 or record.gross_sales > 0
        if active_row:
            group["has_active_rows"] = True
            group["check_count_available"] = (
                group["check_count_available"] and record.check_count_available
            )
        if record.guest_count > 0:
            group["rate_available"] = group["rate_available"] and record.rate_available
            if record.rate_available and record.rate_of_sale_by_guest_count > 0:
                group["rate_qualifying_sales"] += (
                    record.guest_count / record.rate_of_sale_by_guest_count
                )
                group["rate_opportunities"] += record.guest_count
        if record.check_count > 0 and record.check_count_available:
            group["ticket_time_available"] = (
                group["ticket_time_available"] and record.ticket_time_available
            )
        if (
            record.check_count > 0
            and record.check_count_available
            and record.ticket_time_available
        ):
            group["ticket_weighted_sum"] += (
                record.average_ticket_time_seconds * record.check_count
            )
            group["ticket_weight"] += record.check_count
        if active_row:
            group["active_dates"].add(record.report_date)
        group["daily_records"].append(
            {
                "report_date": record.report_date,
                "gross_sales": record.gross_sales,
                "guest_count": record.guest_count,
                "wine_sales": record.wine_sales,
                "rate_of_sale_by_guest_count": record.rate_of_sale_by_guest_count,
                "average_ticket_time_seconds": record.average_ticket_time_seconds,
                "rate_available": record.rate_available,
                "ticket_time_available": record.ticket_time_available,
                "check_count": record.check_count,
                "check_count_available": record.check_count_available,
            }
        )
        group["source_files"].add(record.source_file)
        group["source_evidence"].add(
            (
                record.source_file,
                record.source_sha256,
                record.source_format,
                record.parser_engine,
                record.report_date_source,
                record.report_date.isoformat(),
            )
        )

    rollups: list[dict[str, Any]] = []
    for group in groups.values():
        gross_sales = group["gross_sales"]
        guest_count = group["guest_count"]
        wine_sales = group["wine_sales"]
        rollup = dict(group)
        rollup["check_average"] = gross_sales / guest_count if guest_count else 0.0
        rollup["wine_pct"] = wine_sales / gross_sales if gross_sales else 0.0
        rollup["check_count_available"] = bool(
            group["has_active_rows"] and group["check_count_available"]
        )
        rollup["per_check_available"] = bool(
            rollup["check_count_available"] and group["check_count"] > 0
        )
        rollup["sales_per_check"] = (
            gross_sales / group["check_count"]
            if rollup["per_check_available"]
            else 0.0
        )
        rollup["guests_per_check"] = (
            guest_count / group["check_count"]
            if rollup["per_check_available"]
            else 0.0
        )
        rollup["rate_available"] = bool(
            group["rate_available"]
            and group["rate_opportunities"] > 0
            and group["rate_qualifying_sales"] > 0
        )
        rollup["rate_of_sale_by_guest_count"] = (
            group["rate_opportunities"] / group["rate_qualifying_sales"]
            if rollup["rate_available"]
            else 0.0
        )
        rollup["ticket_time_available"] = bool(
            rollup["check_count_available"]
            and group["ticket_time_available"]
            and group["ticket_weight"] > 0
        )
        rollup["average_ticket_time_seconds"] = (
            group["ticket_weighted_sum"] / group["ticket_weight"]
            if rollup["ticket_time_available"]
            else 0.0
        )
        rollup["ticket_time_weight_basis"] = (
            "Check Count"
            if rollup["ticket_time_available"]
            else "Unavailable (Check Count missing or incomplete)"
        )
        rollup["active_days"] = len(group["active_dates"])
        rollup["source_days"] = len(group["active_dates"])
        rollup["source_files"] = ", ".join(sorted(group["source_files"]))
        rollup["daily_records"] = sorted(
            group["daily_records"], key=lambda item: item["report_date"]
        )
        rollup["source_evidence"] = [
            {
                "source_file": item[0],
                "sha256": item[1] or None,
                "format": item[2] or None,
                "parser_engine": item[3] or None,
                "report_date_source": item[4],
                "report_date": item[5],
            }
            for item in sorted(group["source_evidence"])
        ]
        for helper in (
            "rate_qualifying_sales",
            "rate_opportunities",
            "ticket_weighted_sum",
            "ticket_weight",
            "has_active_rows",
            "active_dates",
        ):
            rollup.pop(helper, None)
        rollups.append(rollup)

    return rollups


def row_name_haystack(row: dict[str, Any]) -> str:
    return f"{row.get('raw_user_name', '')} {row.get('display_name', '')}".casefold()


def row_matches_name_patterns(row: dict[str, Any], patterns: Iterable[Any]) -> bool:
    haystack = row_name_haystack(row)
    return any(str(pattern).casefold() in haystack for pattern in patterns)


def row_matches_weekly_area_patterns(
    row: dict[str, Any], patterns: Iterable[Any]
) -> bool:
    """Match configured area labels as literal whole tokens or phrases."""

    names = (
        "" if is_blank(row.get(field)) else str(row.get(field)).casefold()
        for field in ("raw_user_name", "display_name")
    )
    normalized_names = tuple(names)
    for pattern in patterns:
        phrase = str(pattern).strip().casefold()
        if not phrase:
            continue
        token_phrase = re.compile(rf"(?<!\w){re.escape(phrase)}(?!\w)")
        if any(token_phrase.search(name) for name in normalized_names):
            return True
    return False


def dashboard_exclusion_patterns(config: dict[str, Any]) -> list[Any]:
    patterns: list[Any] = []
    for key in ("dashboard_exclude_name_contains", "public_exclude_name_contains"):
        for pattern in config.get(key, []):
            if pattern not in patterns:
                patterns.append(pattern)
    return patterns


def dashboard_trend_eligible(row: dict[str, Any], config: dict[str, Any]) -> bool:
    min_guest_count = float(config.get("dashboard_min_guest_count_for_trends", 25))
    min_active_days = int(config.get("dashboard_min_active_days_for_trends", 3))
    return row.get("guest_count", 0) >= min_guest_count and row.get("active_days", 0) >= min_active_days


def public_excluded(row: dict[str, Any], config: dict[str, Any]) -> bool:
    if row["guest_count"] < float(config.get("public_min_guest_count", 1)):
        return True
    if is_numbered_server_placeholder(row.get("raw_user_name")) or is_numbered_server_placeholder(
        row.get("display_name")
    ):
        return True
    return row_matches_name_patterns(row, config.get("public_exclude_name_contains", []))


def dashboard_excluded(row: dict[str, Any], config: dict[str, Any]) -> bool:
    if is_numbered_server_placeholder(row.get("raw_user_name")) or is_numbered_server_placeholder(
        row.get("display_name")
    ):
        return True
    exact_names = {
        str(value).strip().casefold()
        for value in config.get("dashboard_exclude_exact_names", [])
    }
    if str(row.get("display_name", "")).strip().casefold() in exact_names:
        return True
    return row_matches_name_patterns(row, dashboard_exclusion_patterns(config))


def format_date_range(start: date, end: date) -> str:
    if start == end:
        return start.strftime("%m/%d/%Y")
    return f"{start:%m/%d/%Y} - {end:%m/%d/%Y}"


def duration_fraction(seconds: float | None) -> float | None:
    if seconds is None:
        return None
    return seconds / 86400 if seconds else 0.0


def style_public_sheet(ws, last_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="E1E1E1")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 78
    widths = {"A": 18, "B": 28, "C": 13, "D": 13, "E": 15, "F": 12, "G": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for hidden_col in ("C", "D", "F"):
        ws.column_dimensions[hidden_col].hidden = True

    ws["A1"].font = Font(name="Calibri", size=12, color="68696C")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=7):
        for cell in row:
            cell.font = Font(name="Calibri", size=12, bold=cell.row == 3)
            cell.border = border
        row[2].number_format = "$#,##0.00"
        row[3].number_format = "#,##0"
        row[4].number_format = "$#,##0.00"
        row[5].number_format = "$#,##0.00"
        row[6].number_format = "0.00%"

    if last_row >= 2:
        ws.auto_filter.ref = f"A2:G{last_row}"
        ws.freeze_panes = "A3"


def apply_public_metric_highlights(ws, last_row: int) -> None:
    first_server_row = 4
    if last_row < first_server_row:
        return

    light_green = PatternFill("solid", fgColor="92D050")
    light_yellow = PatternFill("solid", fgColor="FFFF00")
    light_red = PatternFill("solid", fgColor="F4CCCC")

    rows = list(range(first_server_row, last_row + 1))
    check_values = [
        (row, ws.cell(row=row, column=5).value or 0)
        for row in rows
    ]
    wine_values = []
    for row in rows:
        gross_sales = ws.cell(row=row, column=3).value or 0
        wine_sales = ws.cell(row=row, column=6).value or 0
        wine_pct = wine_sales / gross_sales if gross_sales else 0
        wine_values.append((row, wine_pct))

    top_check_row = max(check_values, key=lambda item: item[1])[0]
    top_wine_row = max(wine_values, key=lambda item: item[1])[0]
    bottom_check_rows = [
        row for row, _ in sorted(check_values, key=lambda item: (item[1], item[0]))[:3]
    ]

    for row in bottom_check_rows:
        ws.cell(row=row, column=5).fill = light_red
    ws.cell(row=top_check_row, column=5).fill = light_green
    ws.cell(row=top_wine_row, column=7).fill = light_yellow


def write_report_definition(ws, location: str, start: date, end: date) -> None:
    bullet = "\u25cf"
    filter_text = (
        "FILTERS\n"
        f"  {bullet}  Date(s):  {format_date_range(start, end)}\n"
        f"  {bullet}  Location Type:  Store\n"
        f"  {bullet}  Locations:  {location}"
    )
    rows = [
        ["Daily Report - TM (Auto-Run)"],
        [filter_text],
        [],
        ["Column", "Definition"],
        ["Gross Sales", "KPI: GrossAmount"],
        ["Guest Count", "KPI: GuestCount"],
        ["Check Average", "KPI: GrossGuestAvg"],
        ["Wine Sales", "KPI: MenuPrice\nMenuCategory: Banquet, Wine - Glass, Wine Bottle, Wine Glass"],
        ["Rate of Sale by Guest Count", "KPI: RateOfSaleGuestCount"],
        ["Average Ticket Time", "KPI: AvgTicketTime"],
    ]
    for row in rows:
        ws.append(row)

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 76
    ws.row_dimensions[2].height = 78
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["B8"].alignment = Alignment(wrap_text=True)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor="E1E1E1")
        cell.font = Font(bold=True)


def write_public_workbook(
    location: str,
    selected_records: list[MetricRecord],
    output_dir: Path,
    config: dict[str, Any],
    public_start: date,
    public_end: date,
) -> Path:
    actual_start = min(record.report_date for record in selected_records)
    actual_end = max(record.report_date for record in selected_records)
    short_code = config["locations"][location]["short_code"]

    location_records = [record for record in selected_records if record.location == location]
    totals = [record for record in location_records if record.is_location_total]
    server_records = [record for record in location_records if not record.is_location_total]

    total_rows = aggregate_records(totals, ("location", "display_name", "raw_user_name"))
    public_rows = aggregate_records(server_records, ("location", "display_name", "raw_user_name"))
    public_rows = [row for row in public_rows if not public_excluded(row, config)]
    public_rows.sort(key=lambda row: (-row["check_average"], row["display_name"]))

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "Report(All)"
    ws.merge_cells("A1:F1")

    bullet = "\u25cf"
    ws["A1"] = (
        "FILTERS\n"
        f"  {bullet}  Week:  {format_date_range(public_start, public_end)} ({OPERATING_WEEK_LABEL})\n"
        f"  {bullet}  Source Date(s):  {format_date_range(actual_start, actual_end)}\n"
        f"  {bullet}  Location Type:  Store\n"
        f"  {bullet}  Locations:  {location}"
    )
    headers = ["Store", "", "Gross Sales", "Guest Count", "Check Average", "Wine Sales", "Wine %"]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_index, value=header)

    rows_to_write = []
    if total_rows:
        rows_to_write.append(total_rows[0])
    rows_to_write.extend(public_rows)

    for row_index, row in enumerate(rows_to_write, start=3):
        ws.cell(row=row_index, column=1, value=location)
        ws.cell(row=row_index, column=2, value=excel_safe_text(row["display_name"]))
        ws.cell(row=row_index, column=3, value=row["gross_sales"])
        ws.cell(row=row_index, column=4, value=row["guest_count"])
        ws.cell(row=row_index, column=5, value=row["check_average"])
        ws.cell(row=row_index, column=6, value=row["wine_sales"])
        ws.cell(row=row_index, column=7, value=row["wine_pct"])

    last_row = max(2, len(rows_to_write) + 2)
    style_public_sheet(ws, last_row)
    apply_public_metric_highlights(ws, last_row)
    write_report_definition(wb.create_sheet("Report Definition"), location, public_start, public_end)

    output_path = output_dir / f"Check_Wine_{short_code}{public_end:%m%d%y}.xlsx"
    wb.save(output_path)
    return output_path


def write_table_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    table_name: str,
    widths: dict[str, float] | None = None,
    index: int | None = None,
):
    ws = wb.create_sheet(title, index=index)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(size=15, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="7A1E1E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_index, value=header)
        cell.fill = PatternFill("solid", fgColor="E1E1E1")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=4):
        for col_index, value in enumerate(row, start=1):
            header = headers[col_index - 1]
            if header in UNTRUSTED_WORKBOOK_TEXT_HEADERS:
                value = excel_safe_text(value)
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if "Date" in header or header in {"Week Start", "Week End", "Latest Week End"}:
                cell.number_format = EXCEL_DATE_NUMBER_FORMAT
            elif "Composite Score" in header or "Rank Movement" in header:
                cell.number_format = "0.0"
            elif "Rank" in header:
                cell.number_format = "#,##0"
            elif "Ticket Time Change (Min)" in header:
                cell.number_format = "0.0"
            elif "Check Average Change" in header:
                cell.number_format = "$#,##0.00"
            elif "Wine % Change" in header:
                cell.number_format = "0.00%"
            elif "Rate Change" in header:
                cell.number_format = "0.00"
            elif "Wine %" in header:
                cell.number_format = "0.00%"
            elif "Rate of Sale" in header:
                cell.number_format = "0.00"
            elif "Ticket Time" in header:
                cell.number_format = "[h]:mm:ss"
            elif (
                "Gross Sales" in header
                or "Wine Sales" in header
                or ("Check Average" in header and "Rank" not in header)
            ):
                cell.number_format = "$#,##0.00"
            elif "Guest" in header or "Days" in header or "Weeks" in header:
                cell.number_format = "#,##0"

    if rows:
        ref = f"A3:{get_column_letter(len(headers))}{len(rows) + 3}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    default_widths = {
        "A": 14,
        "B": 14,
        "C": 20,
        "D": 24,
        "E": 24,
        "F": 18,
        "G": 12,
        "H": 16,
        "I": 14,
        "J": 12,
        "K": 20,
        "L": 18,
        "M": 12,
        "N": 12,
        "O": 18,
        "P": 18,
    }
    for index in range(1, len(headers) + 1):
        col = get_column_letter(index)
        default_widths.setdefault(col, min(max(len(headers[index - 1]) + 2, 12), 24))
    default_widths.update(widths or {})
    for col, width in default_widths.items():
        ws.column_dimensions[col].width = width
    return ws


def weekly_rollups(records: list[MetricRecord]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    grouped: dict[tuple[date, date], list[MetricRecord]] = defaultdict(list)
    for record in records:
        if not is_operating_day(record.report_date):
            continue
        grouped[week_period_for(record.report_date)].append(record)

    server_rows: list[dict[str, Any]] = []
    location_rows: list[dict[str, Any]] = []
    for (week_start, week_end), period_records in sorted(grouped.items()):
        servers = [record for record in period_records if not record.is_location_total]
        locations = [record for record in period_records if record.is_location_total]
        server_rows.extend(
            aggregate_records(
                servers,
                ("location", "raw_user_name", "display_name"),
                {"week_start": week_start, "week_end": week_end},
            )
        )
        location_rows.extend(
            aggregate_records(
                locations,
                ("location",),
                {"week_start": week_start, "week_end": week_end},
            )
        )
    return server_rows, location_rows


def assign_rank(
    rows: list[dict[str, Any]],
    value_field: str,
    rank_field: str,
    *,
    higher_is_better: bool,
    min_guest_count: int,
    availability_field: str | None = None,
) -> None:
    eligible = [
        row
        for row in rows
        if row.get("guest_count", 0) >= min_guest_count
        and row.get(value_field) is not None
        and (
            availability_field is None
            or bool(row.get(availability_field, False))
        )
    ]
    eligible.sort(
        key=lambda row: row[value_field],
        reverse=higher_is_better,
    )

    previous_value: float | None = None
    previous_rank = 0
    for index, row in enumerate(eligible, start=1):
        value = row[value_field]
        rank = previous_rank if previous_value == value else index
        row[rank_field] = rank
        previous_rank = rank
        previous_value = value

    for row in rows:
        row.setdefault(rank_field, None)


def weekly_server_rank_rows(
    weekly_server_rows: list[dict[str, Any]],
    min_guest_count: int,
) -> list[dict[str, Any]]:
    groups: dict[tuple[date, date, str], list[dict[str, Any]]] = defaultdict(list)
    for row in weekly_server_rows:
        ranked = dict(row)
        groups[(ranked["week_start"], ranked["week_end"], ranked["location"])].append(ranked)

    ranked_rows: list[dict[str, Any]] = []
    for group_rows in groups.values():
        assign_rank(
            group_rows,
            "check_average",
            "check_average_rank",
            higher_is_better=True,
            min_guest_count=min_guest_count,
        )
        assign_rank(
            group_rows,
            "wine_pct",
            "wine_pct_rank",
            higher_is_better=True,
            min_guest_count=min_guest_count,
        )
        assign_rank(
            group_rows,
            "rate_of_sale_by_guest_count",
            "rate_rank",
            higher_is_better=False,
            min_guest_count=min_guest_count,
            availability_field="rate_available",
        )
        assign_rank(
            group_rows,
            "average_ticket_time_seconds",
            "ticket_time_rank",
            higher_is_better=False,
            min_guest_count=min_guest_count,
            availability_field="ticket_time_available",
        )
        ranked_rows.extend(group_rows)

    ranked_rows.sort(
        key=lambda row: (
            row["week_end"],
            row["location"],
            row.get("check_average_rank") or 9999,
            row["display_name"],
        )
    )
    return ranked_rows


def trend_note(
    check_change: float | None,
    wine_change: float | None,
    rate_change: float | None,
    ticket_change_minutes: float | None,
) -> str:
    if check_change is None:
        return "No prior week"

    improved = 0
    declined = 0
    # Person-level trend labels are limited to the two validated action
    # measures. Rate of Sale and Ticket Time remain adjacent context only.
    for value, higher_is_better in (
        (check_change, True),
        (wine_change, True),
    ):
        if value is None or abs(value) < 0.000001:
            continue
        if (higher_is_better and value > 0) or ((not higher_is_better) and value < 0):
            improved += 1
        else:
            declined += 1

    if improved >= 2 and improved > declined:
        return "Improving"
    if declined >= 2 and declined > improved:
        return "Watch"
    return "Mixed"


def metric_delta(current: dict[str, Any], previous: dict[str, Any] | None, field: str) -> float | None:
    if previous is None:
        return None
    return current[field] - previous[field]


def rank_movement(current_rank: Any, previous_rank: Any) -> float | None:
    if current_rank is None or previous_rank is None:
        return None
    return float(previous_rank) - float(current_rank)


def star_score_component(
    change: float | None, *, lower_is_better: bool, strong_threshold: float
) -> int:
    if change is None or abs(change) < 0.000001:
        return 0
    directional_change = -change if lower_is_better else change
    if directional_change >= strong_threshold:
        return 2
    if directional_change > 0:
        return 1
    if directional_change <= -strong_threshold:
        return -2
    return -1


def format_change(change: float | None, kind: str) -> str:
    if change is None:
        return ""
    sign = "+" if change > 0 else "-" if change < 0 else ""
    value = abs(change)
    if kind == "currency":
        return f"{sign}${value:,.2f}"
    if kind == "pct_points":
        return f"{sign}{value * 100:.1f} pts"
    if kind == "minutes":
        return f"{sign}{value:.1f} min"
    if kind == "rank":
        return f"{sign}{value:.1f}"
    return f"{sign}{value:.2f}"


def average_rank_movement(row: dict[str, Any]) -> float | None:
    movements = [
        row.get("check_average_rank_movement"),
        row.get("wine_pct_rank_movement"),
    ]
    available = [movement for movement in movements if movement is not None]
    if not available:
        return None
    return sum(available) / len(available)


def server_week_trend_rows(ranked_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in ranked_rows:
        groups[(row["location"], row["raw_user_name"], row["display_name"])].append(row)

    trend_rows: list[dict[str, Any]] = []
    for (_, _, _), rows in groups.items():
        rows.sort(key=lambda row: row["week_end"])
        previous: dict[str, Any] | None = None
        for row in rows:
            check_change = metric_delta(row, previous, "check_average")
            wine_change = metric_delta(row, previous, "wine_pct")
            rate_change = (
                metric_delta(row, previous, "rate_of_sale_by_guest_count")
                if previous
                and row.get("rate_available")
                and previous.get("rate_available")
                else None
            )
            ticket_change_minutes = (
                (row["average_ticket_time_seconds"] - previous["average_ticket_time_seconds"]) / 60
                if previous
                and row.get("ticket_time_available")
                and previous.get("ticket_time_available")
                else None
            )
            rank_fields = (
                ("check_average_rank", "check_average_rank_movement"),
                ("wine_pct_rank", "wine_pct_rank_movement"),
                ("rate_rank", "rate_rank_movement"),
                ("ticket_time_rank", "ticket_time_rank_movement"),
            )
            rank_values: dict[str, Any] = {}
            for rank_field, movement_field in rank_fields:
                prior_field = f"prior_{rank_field}"
                prior_rank = previous.get(rank_field) if previous else None
                rank_values[prior_field] = prior_rank
                rank_values[movement_field] = rank_movement(row.get(rank_field), prior_rank)
            trend_rows.append(
                {
                    **row,
                    "check_average_change": check_change,
                    "wine_pct_change": wine_change,
                    "rate_change": rate_change,
                    "ticket_time_change_minutes": ticket_change_minutes,
                    **rank_values,
                    "trend_note": trend_note(
                        check_change,
                        wine_change,
                        rate_change,
                        ticket_change_minutes,
                    ),
                }
            )
            previous = row

    trend_rows.sort(
        key=lambda row: (
            row["week_end"],
            row["location"],
            row.get("check_average_rank") or 9999,
            row["display_name"],
        )
    )
    return trend_rows


def star_classification(score: float) -> str:
    if score >= 2:
        return "Rising Star"
    if score <= -2:
        return "Falling Star"
    return "Stable"


def star_why(row: dict[str, Any], components: list[tuple[str, int, str]]) -> str:
    drivers = [
        label
        for label, score, _ in sorted(
            components,
            key=lambda item: (abs(item[1]), item[2]),
            reverse=True,
        )
        if score
    ]
    if not drivers:
        return "Minimal week-over-week movement"
    return "; ".join(drivers[:4])


def server_star_rows(
    server_trend_detail_rows: list[dict[str, Any]], config: dict[str, Any]
) -> list[dict[str, Any]]:
    rows_with_prior = [
        row
        for row in server_trend_detail_rows
        if row.get("check_average_change") is not None
    ]
    if not rows_with_prior:
        return []

    latest_week_end = max(row["week_end"] for row in rows_with_prior)
    star_rows: list[dict[str, Any]] = []
    for row in rows_with_prior:
        if row["week_end"] != latest_week_end:
            continue
        if dashboard_excluded(row, config) or not dashboard_trend_eligible(row, config):
            continue

        metric_components = [
            (
                "Sales / Guest",
                star_score_component(
                    row.get("check_average_change"),
                    lower_is_better=False,
                    strong_threshold=5.0,
                ),
                f"Sales / Guest {format_change(row.get('check_average_change'), 'currency')}",
                "check",
            ),
            (
                "Wine %",
                star_score_component(
                    row.get("wine_pct_change"),
                    lower_is_better=False,
                    strong_threshold=0.01,
                ),
                f"Wine {format_change(row.get('wine_pct_change'), 'pct_points')}",
                "wine",
            ),
        ]
        avg_rank_move = average_rank_movement(row)
        score = sum(component[1] for component in metric_components)
        components = [(label, component_score, sort_key) for _, component_score, label, sort_key in metric_components]
        star_rows.append(
            {
                "category": star_classification(score),
                "composite_score": score,
                "week_start": row["week_start"],
                "week_end": row["week_end"],
                "location": row["location"],
                "raw_user_name": row["raw_user_name"],
                "display_name": row["display_name"],
                "guest_count": row["guest_count"],
                "active_days": row["active_days"],
                "check_average": row["check_average"],
                "check_average_change": row["check_average_change"],
                "wine_pct": row["wine_pct"],
                "wine_pct_change": row["wine_pct_change"],
                "rate_of_sale_by_guest_count": row["rate_of_sale_by_guest_count"],
                "rate_change": row["rate_change"],
                "average_ticket_time_seconds": row["average_ticket_time_seconds"],
                "ticket_time_change_minutes": row["ticket_time_change_minutes"],
                "average_rank_movement": avg_rank_move,
                "why": star_why(row, components),
            }
        )

    category_order = {"Rising Star": 0, "Falling Star": 1, "Stable": 2}
    star_rows.sort(
        key=lambda row: (
            category_order[row["category"]],
            -row["composite_score"] if row["category"] == "Rising Star" else row["composite_score"],
            row["location"],
            row["display_name"],
        )
    )
    return star_rows


def compact_money(value: float | None, *, signed: bool = False) -> str:
    if value is None:
        return ""
    sign = ""
    if signed:
        sign = "+" if value > 0 else "-" if value < 0 else ""
    return f"{sign}${abs(value):,.0f}"


def compact_count(value: float | None, *, signed: bool = False) -> str:
    if value is None:
        return ""
    sign = "+" if signed and value > 0 else ""
    return f"{sign}{value:,.0f}"


def compact_pct_points(value: float | None) -> str:
    if value is None:
        return ""
    sign = "+" if value > 0 else "-" if value < 0 else ""
    return f"{sign}{abs(value) * 100:.1f} pts"


def compact_minutes(value: float | None) -> str:
    if value is None:
        return ""
    sign = "+" if value > 0 else "-" if value < 0 else ""
    return f"{sign}{abs(value):.1f} min"


def star_action(row: dict[str, Any]) -> tuple[str, str, str]:
    if row["category"] == "Falling Star":
        priority = "High" if row["composite_score"] <= -4 else "Medium"
        return (
            priority,
            "Coach",
            "What comparable-shift context explains the Sales / Guest and Wine % "
            "movement, and what one coaching step is supported?",
        )
    if row["category"] == "Rising Star":
        priority = "Recognize" if row["composite_score"] >= 4 else "Share"
        return (
            priority,
            "Recognize",
            "What supported practice should be recognized and shared after context review?",
        )
    return (
        "Monitor",
        "Monitor",
        "What should be watched in the next qualified week before acting?",
    )


def store_status(row: dict[str, Any]) -> tuple[str, str, str]:
    gross_down = (row.get("gross_sales_change") or 0) < 0
    guests_down = (row.get("guest_count_change") or 0) < 0
    check_down = (row.get("check_average_change") or 0) < 0
    wine_down = (row.get("wine_pct_change") or 0) < 0
    negative_count = sum([gross_down, guests_down, check_down, wine_down])

    if gross_down and guests_down:
        return (
            "High",
            "Traffic Watch",
            "What changed in reservations, events, operating hours, staffing, or local demand?",
        )
    if check_down and wine_down:
        return (
            "Medium",
            "Upsell Watch",
            "Did guest mix, menu availability, training, or shift mix explain the "
            "Sales / Guest and Wine % change?",
        )
    if negative_count >= 2:
        return (
            "Medium",
            "Mixed Watch",
            "Which validated movement changed, what context explains it, and what "
            "one follow-up should be assigned?",
        )
    return (
        "Monitor",
        "Stable / Mixed",
        "What, if anything, needs follow-up before the next weekly review?",
    )


def trend_evidence(row: dict[str, Any]) -> str:
    return (
        f"Sales {compact_money(row.get('gross_sales_change'), signed=True)}; "
        f"guests {compact_count(row.get('guest_count_change'), signed=True)}; "
        f"check {compact_money(row.get('check_average_change'), signed=True)}; "
        f"wine {compact_pct_points(row.get('wine_pct_change'))}; "
        f"ticket {compact_minutes(row.get('ticket_time_change_minutes'))}"
    )


def build_action_board_rows(
    star_rows: list[dict[str, Any]],
    store_trend_summary_rows: list[dict[str, Any]],
    group_trend_summary_rows: list[dict[str, Any]],
    weekly_location_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    sort_order = {
        "High": 0,
        "Medium": 1,
        "Recognize": 2,
        "Share": 3,
        "Monitor": 4,
        "Review": 5,
    }

    for star in sorted(
        (row for row in star_rows if row["category"] == "Falling Star"),
        key=lambda row: (row["composite_score"], row["location"], row["display_name"]),
    )[:6]:
        priority, action, follow_up = star_action(star)
        rows.append(
            {
                "priority": priority,
                "action": action,
                "location": star["location"],
                "subject": star["display_name"],
                "signal": "Falling Star",
                "impact": f"Score {star['composite_score']:+.0f}",
                "evidence": star["why"],
                "recommended_follow_up": follow_up,
                "week_end": star["week_end"],
                "guest_count": star["guest_count"],
                "active_days": star["active_days"],
                "_sort": sort_order[priority],
            }
        )

    for star in sorted(
        (row for row in star_rows if row["category"] == "Rising Star"),
        key=lambda row: (-row["composite_score"], row["location"], row["display_name"]),
    )[:6]:
        priority, action, follow_up = star_action(star)
        rows.append(
            {
                "priority": priority,
                "action": action,
                "location": star["location"],
                "subject": star["display_name"],
                "signal": "Rising Star",
                "impact": f"Score {star['composite_score']:+.0f}",
                "evidence": star["why"],
                "recommended_follow_up": follow_up,
                "week_end": star["week_end"],
                "guest_count": star["guest_count"],
                "active_days": star["active_days"],
                "_sort": sort_order[priority],
            }
        )

    for store in store_trend_summary_rows:
        priority, signal, follow_up = store_status(store)
        rows.append(
            {
                "priority": priority,
                "action": "Store Review",
                "location": store["location"],
                "subject": store["location"],
                "signal": signal,
                "impact": trend_evidence(store),
                "evidence": f"Latest week {store['latest_week_end']:%m/%d/%Y}",
                "recommended_follow_up": follow_up,
                "week_end": store["latest_week_end"],
                "guest_count": store["latest_guest_count"],
                "active_days": None,
                "_sort": sort_order[priority] + 0.5,
            }
        )

    for group in group_trend_summary_rows:
        priority, signal, follow_up = store_status(group)
        rows.append(
            {
                "priority": priority,
                "action": "Group Review",
                "location": group.get("group", "All Stores"),
                "subject": group.get("group", "All Stores"),
                "signal": signal,
                "impact": trend_evidence(group),
                "evidence": f"Latest week {group['latest_week_end']:%m/%d/%Y}",
                "recommended_follow_up": follow_up,
                "week_end": group["latest_week_end"],
                "guest_count": group["latest_guest_count"],
                "active_days": None,
                "_sort": sort_order[priority] + 0.25,
            }
        )

    for status, location, week_end, detail in data_quality_warning_rows(weekly_location_rows):
        if status == "OK":
            continue
        rows.append(
            {
                "priority": "Review",
                "action": "Data Quality",
                "location": location,
                "subject": location,
                "signal": status,
                "impact": detail,
                "evidence": "Short weeks can distort trends and star movement.",
                "recommended_follow_up": "Confirm whether missing source days are expected before coaching from that week.",
                "week_end": week_end,
                "guest_count": None,
                "active_days": None,
                "_sort": sort_order["Review"],
            }
        )

    rows.sort(key=lambda row: (row["_sort"], row["action"], row["location"], row["subject"]))
    return rows


def weekly_metric_trend_rows(
    weekly_rows: list[dict[str, Any]], identity_fields: tuple[str, ...]
) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in weekly_rows:
        groups[tuple(row[field] for field in identity_fields)].append(row)

    trend_rows: list[dict[str, Any]] = []
    for rows in groups.values():
        rows.sort(key=lambda row: row["week_end"])
        previous: dict[str, Any] | None = None
        for row in rows:
            check_change = metric_delta(row, previous, "check_average")
            wine_change = metric_delta(row, previous, "wine_pct")
            rate_change = (
                metric_delta(row, previous, "rate_of_sale_by_guest_count")
                if previous
                and row.get("rate_available")
                and previous.get("rate_available")
                else None
            )
            ticket_change_minutes = (
                (row["average_ticket_time_seconds"] - previous["average_ticket_time_seconds"]) / 60
                if previous
                and row.get("ticket_time_available")
                and previous.get("ticket_time_available")
                else None
            )
            trend_rows.append(
                {
                    **row,
                    "prior_week_end": previous["week_end"] if previous else None,
                    "gross_sales_change": metric_delta(row, previous, "gross_sales"),
                    "guest_count_change": metric_delta(row, previous, "guest_count"),
                    "check_average_change": check_change,
                    "wine_sales_change": metric_delta(row, previous, "wine_sales"),
                    "wine_pct_change": wine_change,
                    "rate_change": rate_change,
                    "ticket_time_change_minutes": ticket_change_minutes,
                    "trend_note": trend_note(
                        check_change,
                        wine_change,
                        rate_change,
                        ticket_change_minutes,
                    ),
                }
            )
            previous = row

    trend_rows.sort(key=lambda row: (row["week_end"],) + tuple(row[field] for field in identity_fields))
    return trend_rows


def group_weekly_rows(records: list[MetricRecord]) -> list[dict[str, Any]]:
    grouped: dict[tuple[date, date], list[MetricRecord]] = defaultdict(list)
    for record in records:
        if record.is_location_total and is_operating_day(record.report_date):
            grouped[week_period_for(record.report_date)].append(record)

    rows: list[dict[str, Any]] = []
    for (week_start, week_end), period_records in sorted(grouped.items()):
        rows.extend(
            aggregate_records(
                period_records,
                (),
                {"week_start": week_start, "week_end": week_end, "group": "All Stores"},
            )
        )
    return rows


def weekly_metric_summary_rows(
    weekly_rows: list[dict[str, Any]], identity_fields: tuple[str, ...]
) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in weekly_rows:
        groups[tuple(row[field] for field in identity_fields)].append(row)

    summary: list[dict[str, Any]] = []
    for key, rows in groups.items():
        rows.sort(key=lambda row: row["week_end"])
        total_gross = sum(row["gross_sales"] for row in rows)
        total_guests = sum(row["guest_count"] for row in rows)
        total_wine = sum(row["wine_sales"] for row in rows)
        operational_context = aggregate_weekly_rows(rows) or {}
        latest = rows[-1]
        prior = rows[-2] if len(rows) > 1 else None
        overall_rate_available = bool(
            operational_context.get("rate_available", False)
        )
        overall_ticket_available = bool(
            operational_context.get("ticket_time_available", False)
        )
        latest_rate_available = bool(latest.get("rate_available", False))
        prior_rate_available = bool(prior and prior.get("rate_available", False))
        latest_ticket_available = bool(
            latest.get("ticket_time_available", False)
        )
        prior_ticket_available = bool(
            prior and prior.get("ticket_time_available", False)
        )
        row = {field: value for field, value in zip(identity_fields, key)}
        row.update(
            {
                "weeks_tracked": len(rows),
                "active_days_tracked": sum(item["active_days"] for item in rows),
                "source_days_tracked": sum(item["source_days"] for item in rows),
                "total_gross_sales": total_gross,
                "total_guest_count": total_guests,
                "weighted_check_average": total_gross / total_guests if total_guests else 0.0,
                "total_wine_sales": total_wine,
                "overall_wine_pct": total_wine / total_gross if total_gross else 0.0,
                "weighted_rate": (
                    operational_context.get("rate_of_sale_by_guest_count")
                    if overall_rate_available
                    else None
                ),
                "weighted_ticket_time_seconds": (
                    operational_context.get("average_ticket_time_seconds")
                    if overall_ticket_available
                    else None
                ),
                "rate_available": overall_rate_available,
                "ticket_time_available": overall_ticket_available,
                "ticket_time_weight_basis": operational_context.get(
                    "ticket_time_weight_basis",
                    "Unavailable (Check Count missing or incomplete)",
                ),
                "first_week_end": rows[0]["week_end"],
                "latest_week_end": latest["week_end"],
                "prior_week_end": prior["week_end"] if prior else None,
                "latest_gross_sales": latest["gross_sales"],
                "prior_gross_sales": prior["gross_sales"] if prior else None,
                "gross_sales_change": metric_delta(latest, prior, "gross_sales"),
                "latest_guest_count": latest["guest_count"],
                "prior_guest_count": prior["guest_count"] if prior else None,
                "guest_count_change": metric_delta(latest, prior, "guest_count"),
                "latest_check_average": latest["check_average"],
                "prior_check_average": prior["check_average"] if prior else None,
                "check_average_change": metric_delta(latest, prior, "check_average"),
                "latest_wine_pct": latest["wine_pct"],
                "prior_wine_pct": prior["wine_pct"] if prior else None,
                "wine_pct_change": metric_delta(latest, prior, "wine_pct"),
                "latest_rate": (
                    latest["rate_of_sale_by_guest_count"]
                    if latest_rate_available
                    else None
                ),
                "prior_rate": (
                    prior["rate_of_sale_by_guest_count"]
                    if prior_rate_available
                    else None
                ),
                "rate_change": (
                    metric_delta(latest, prior, "rate_of_sale_by_guest_count")
                    if latest_rate_available and prior_rate_available
                    else None
                ),
                "latest_ticket_time_seconds": (
                    latest["average_ticket_time_seconds"]
                    if latest_ticket_available
                    else None
                ),
                "prior_ticket_time_seconds": (
                    prior["average_ticket_time_seconds"]
                    if prior_ticket_available
                    else None
                ),
                "ticket_time_change_minutes": (
                    (latest["average_ticket_time_seconds"] - prior["average_ticket_time_seconds"]) / 60
                    if latest_ticket_available and prior_ticket_available
                    else None
                ),
            }
        )
        summary.append(row)

    summary.sort(key=lambda row: tuple(row[field] for field in identity_fields))
    return summary


def trend_summary_rows(weekly_server_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in weekly_server_rows:
        groups[(row["location"], row["raw_user_name"], row["display_name"])].append(row)

    summary: list[dict[str, Any]] = []
    for (location, raw_name, display_name), rows in groups.items():
        rows.sort(key=lambda row: row["week_end"])
        total_gross = sum(row["gross_sales"] for row in rows)
        total_guests = sum(row["guest_count"] for row in rows)
        total_wine = sum(row["wine_sales"] for row in rows)
        operational_context = aggregate_weekly_rows(rows) or {}
        latest = rows[-1]
        prior = rows[-2] if len(rows) > 1 else None
        overall_rate_available = bool(
            operational_context.get("rate_available", False)
        )
        overall_ticket_available = bool(
            operational_context.get("ticket_time_available", False)
        )
        latest_rate_available = bool(latest.get("rate_available", False))
        prior_rate_available = bool(prior and prior.get("rate_available", False))
        latest_ticket_available = bool(
            latest.get("ticket_time_available", False)
        )
        prior_ticket_available = bool(
            prior and prior.get("ticket_time_available", False)
        )
        summary.append(
            {
                "location": location,
                "raw_user_name": raw_name,
                "display_name": display_name,
                "weeks_tracked": len(rows),
                "active_days_tracked": sum(row["active_days"] for row in rows),
                "total_gross_sales": total_gross,
                "total_guest_count": total_guests,
                "weighted_check_average": total_gross / total_guests if total_guests else 0.0,
                "total_wine_sales": total_wine,
                "overall_wine_pct": total_wine / total_gross if total_gross else 0.0,
                "weighted_rate": (
                    operational_context.get("rate_of_sale_by_guest_count")
                    if overall_rate_available
                    else None
                ),
                "weighted_ticket_time_seconds": (
                    operational_context.get("average_ticket_time_seconds")
                    if overall_ticket_available
                    else None
                ),
                "rate_available": overall_rate_available,
                "ticket_time_available": overall_ticket_available,
                "ticket_time_weight_basis": operational_context.get(
                    "ticket_time_weight_basis",
                    "Unavailable (Check Count missing or incomplete)",
                ),
                "first_week_end": rows[0]["week_end"],
                "latest_week_end": latest["week_end"],
                "prior_week_end": prior["week_end"] if prior else None,
                "latest_check_average": latest["check_average"],
                "prior_check_average": prior["check_average"] if prior else None,
                "check_average_change": (
                    latest["check_average"] - prior["check_average"] if prior else None
                ),
                "latest_wine_pct": latest["wine_pct"],
                "prior_wine_pct": prior["wine_pct"] if prior else None,
                "wine_pct_change": latest["wine_pct"] - prior["wine_pct"] if prior else None,
                "latest_rate": (
                    latest["rate_of_sale_by_guest_count"]
                    if latest_rate_available
                    else None
                ),
                "prior_rate": (
                    prior["rate_of_sale_by_guest_count"]
                    if prior_rate_available
                    else None
                ),
                "rate_change": (
                    latest["rate_of_sale_by_guest_count"] - prior["rate_of_sale_by_guest_count"]
                    if latest_rate_available and prior_rate_available
                    else None
                ),
                "latest_ticket_time_seconds": (
                    latest["average_ticket_time_seconds"]
                    if latest_ticket_available
                    else None
                ),
                "prior_ticket_time_seconds": (
                    prior["average_ticket_time_seconds"]
                    if prior_ticket_available
                    else None
                ),
                "ticket_time_change_minutes": (
                    (latest["average_ticket_time_seconds"] - prior["average_ticket_time_seconds"]) / 60
                    if latest_ticket_available and prior_ticket_available
                    else None
                ),
            }
        )

    summary.sort(key=lambda row: (row["location"], -row["weighted_check_average"], row["display_name"]))
    return summary


def style_title(ws, title: str, end_col: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(size=16, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="7A1E1E")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)


def style_section_header(ws, row: int, start_col: int, end_col: int, label: str) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor="E1E1E1")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=row, column=start_col, value=label)


def soft_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def priority_fill(value: Any) -> PatternFill | None:
    text = str(value or "").casefold()
    if text in {
        "high", "coach", "coach now", "falling", "declining", "falling star",
        "below benchmark", "traffic watch", "incomplete week",
        "coaching prompt", "downward", "below peer reference",
    }:
        return soft_fill("F4CCCC")
    if text in {
        "medium", "coach fundamentals", "protect performance", "reinforce improvement",
        "store review", "group review", "upsell watch", "service watch", "mixed watch",
        "building history", "low sample", "context review", "sensitive",
        "limited history", "limited volume", "reference unavailable",
    }:
        return soft_fill("FFF2CC")
    if text in {
        "recognize", "recognize & replicate", "share", "rising", "improving",
        "rising star", "above benchmark", "recognition prompt", "upward",
        "above peer reference",
    }:
        return soft_fill("D9EAD3")
    if text in {
        "review", "review needed", "data quality", "short week", "eligible",
    }:
        return soft_fill("D9EAF7")
    if text in {
        "monitor", "stable", "mixed", "stable / mixed", "on track", "not scored",
        "not evaluated", "no prior week", "within peer range",
    }:
        return soft_fill("EDEDED")
    return None


def apply_dashboard_number_formats(ws, row_start: int, row_end: int) -> None:
    for row in range(row_start, row_end + 1):
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row=row_start - 1, column=col).value or "")
            cell = ws.cell(row=row, column=col)
            if "Date" in header or "Week End" in header:
                cell.number_format = EXCEL_DATE_NUMBER_FORMAT
            elif "Composite Score" in header or "Rank Movement" in header:
                cell.number_format = "0.0"
            elif "Wine %" in header:
                cell.number_format = "0.00%"
            elif "Rate" in header:
                cell.number_format = "0.00"
            elif "Ticket Time Change (Min)" in header:
                cell.number_format = "0.0"
            elif "Ticket Time" in header:
                cell.number_format = "[h]:mm:ss"
            elif "Gross Sales" in header or "Check Average" in header or "Wine Sales" in header:
                cell.number_format = "$#,##0.00"
            elif "Guest" in header or "Rank" in header:
                cell.number_format = "#,##0"


def data_quality_warning_rows(weekly_location_rows: list[dict[str, Any]]) -> list[list[Any]]:
    short_rows = [
        row
        for row in sorted(weekly_location_rows, key=lambda item: (item["week_end"], item["location"]))
        if row["source_days"] < OPERATING_WEEK_DAYS
    ]
    if not short_rows:
        return [["OK", "All tracked store weeks include a full Tuesday-Sunday source set.", None, None]]
    return [
        [
            "Short Week",
            row["location"],
            row["week_end"],
            f"{row['source_days']} of {OPERATING_WEEK_DAYS} source days",
        ]
        for row in short_rows[:6]
    ]


def write_dashboard_table(
    ws,
    title: str,
    row: int,
    col: int,
    headers: list[str],
    rows: list[list[Any]],
) -> int:
    end_col = col + len(headers) - 1
    style_section_header(ws, row, col, end_col, title)
    for offset, header in enumerate(headers):
        cell = ws.cell(row=row + 1, column=col + offset, value=header)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_offset, values in enumerate(rows, start=2):
        for col_offset, value in enumerate(values):
            cell = ws.cell(row=row + row_offset, column=col + col_offset, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if rows:
        apply_dashboard_number_formats(ws, row + 2, row + len(rows) + 1)
    return row + len(rows) + 3


def write_dashboard_sheet(
    wb: Workbook,
    records: list[MetricRecord],
    weekly_location_rows: list[dict[str, Any]],
    ranked_rows: list[dict[str, Any]],
    star_rows: list[dict[str, Any]],
    action_rows: list[dict[str, Any]],
    store_trend_summary_rows: list[dict[str, Any]],
    group_trend_summary_rows: list[dict[str, Any]],
    config: dict[str, Any],
    source_dir: Path,
    public_start: date,
    public_end: date,
) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    style_title(ws, "Red Onion Weekly Performance Dashboard", 13)
    ws.freeze_panes = "A4"
    for col, width in {
        "A": 15,
        "B": 18,
        "C": 20,
        "D": 12,
        "E": 30,
        "F": 34,
        "G": 16,
        "H": 15,
        "I": 18,
        "J": 18,
        "K": 18,
        "L": 34,
        "M": 30,
    }.items():
        ws.column_dimensions[col].width = width

    dates = sorted({record.report_date for record in records})
    latest_week_end = max((row["week_end"] for row in weekly_location_rows), default=None)
    latest_location_rows = [
        row for row in weekly_location_rows if latest_week_end and row["week_end"] == latest_week_end
    ]

    summary_rows = [
        ["Latest Week", public_end],
        ["Reports Read", len({record.source_file for record in records})],
        ["Date Coverage", format_date_range(min(dates), max(dates))],
        ["Weeks Tracked", len({row["week_end"] for row in weekly_location_rows})],
    ]
    snapshot_end = write_dashboard_table(ws, "Snapshot", 3, 1, ["Item", "Value"], summary_rows)

    group_rows = []
    for row in group_trend_summary_rows:
        _, status, focus = store_status(row)
        group_rows.append(
            [
                row.get("group", "All Stores"),
                status,
                compact_money(row["latest_gross_sales"]),
                compact_money(row["gross_sales_change"], signed=True),
                compact_count(row["guest_count_change"], signed=True),
                compact_money(row["check_average_change"], signed=True),
                compact_pct_points(row["wine_pct_change"]),
                compact_minutes(row["ticket_time_change_minutes"]),
                focus,
            ]
        )
    group_end = write_dashboard_table(
        ws,
        "All-Stores Group Pulse",
        3,
        4,
        [
            "Group",
            "Status",
            "Sales",
            "Sales Change",
            "Guest Change",
            "Check Change",
            "Wine Change",
            "Ticket Change",
            "Focus",
        ],
        group_rows,
    )

    data_quality_end = write_dashboard_table(
        ws,
        "Data Quality",
        max(snapshot_end, group_end) + 1,
        1,
        ["Status", "Location / Note", "Week End", "Detail"],
        data_quality_warning_rows(weekly_location_rows),
    )

    coach_rows = [
        [
            excel_safe_text(row["priority"]),
            excel_safe_text(row["location"]),
            excel_safe_text(row["subject"]),
            excel_safe_text(row["impact"]),
            excel_safe_text(row["evidence"]),
            excel_safe_text(row["recommended_follow_up"]),
        ]
        for row in action_rows
        if row["action"] == "Coach"
    ][:5]
    recognize_rows = [
        [
            excel_safe_text(row["priority"]),
            excel_safe_text(row["location"]),
            excel_safe_text(row["subject"]),
            excel_safe_text(row["impact"]),
            excel_safe_text(row["evidence"]),
            excel_safe_text(row["recommended_follow_up"]),
        ]
        for row in action_rows
        if row["action"] == "Recognize"
    ][:5]
    if not coach_rows:
        coach_rows = [["", "", "No coach-now items", "", "", ""]]
    if not recognize_rows:
        recognize_rows = [["", "", "No recognition items", "", "", ""]]
    action_tables_row = max(max(snapshot_end, group_end) + 7, data_quality_end)
    coach_end = write_dashboard_table(
        ws,
        "Coach First",
        action_tables_row,
        1,
        ["Priority", "Location", "Server", "Impact", "Evidence", "Recommended Follow-Up"],
        coach_rows,
    )
    recognize_end = write_dashboard_table(
        ws,
        "Recognize / Replicate",
        action_tables_row,
        8,
        ["Priority", "Location", "Server", "Impact", "Evidence", "Recommended Follow-Up"],
        recognize_rows,
    )

    store_rows = []
    for row in store_trend_summary_rows:
        _, status, focus = store_status(row)
        store_rows.append(
            [
                row["location"],
                status,
                compact_money(row["gross_sales_change"], signed=True),
                compact_count(row["guest_count_change"], signed=True),
                compact_money(row["check_average_change"], signed=True),
                compact_pct_points(row["wine_pct_change"]),
                compact_minutes(row["ticket_time_change_minutes"]),
                focus,
            ]
        )
    store_end = write_dashboard_table(
        ws,
        "Store Action Pulse",
        max(coach_end, recognize_end) + 1,
        1,
        [
            "Location",
            "Status",
            "Sales Change",
            "Guest Count",
            "Check Change",
            "Wine Change",
            "Ticket Change",
            "Recommended Focus",
        ],
        store_rows,
    )

    latest_ranked = [
        row
        for row in ranked_rows
        if latest_week_end
        and row["week_end"] == latest_week_end
        and not dashboard_excluded(row, config)
        and dashboard_trend_eligible(row, config)
    ]
    leaderboard_sections = [
        (
            "Highest Check Average",
            lambda rows: sorted(
                rows,
                key=lambda row: (row.get("check_average_rank") or 9999, row["display_name"]),
            )[:3],
            "Check Average",
            "check_average",
            "$#,##0.00",
        ),
        (
            "Highest Wine %",
            lambda rows: sorted(
                rows,
                key=lambda row: (row.get("wine_pct_rank") or 9999, row["display_name"]),
            )[:3],
            "Wine %",
            "wine_pct",
            "0.00%",
        ),
        (
            "Best Rate of Sale by Guest Count",
            lambda rows: sorted(
                rows,
                key=lambda row: (row.get("rate_rank") or 9999, row["display_name"]),
            )[:3],
            "Rate of Sale by Guest Count",
            "rate_of_sale_by_guest_count",
            "0.00",
        ),
        (
            "Fastest Ticket Time",
            lambda rows: sorted(
                rows,
                key=lambda row: (row.get("ticket_time_rank") or 9999, row["display_name"]),
            )[:3],
            "Ticket Time",
            "average_ticket_time_seconds",
            "[h]:mm:ss",
        ),
    ]

    table_row = store_end + 1
    for title, selector, value_label, value_field, number_format in leaderboard_sections:
        style_section_header(ws, table_row, 1, 6, title)
        for col, header in enumerate(["Location", "Rank", "Server", "Guest Count", value_label, "Active Days"], start=1):
            cell = ws.cell(row=table_row + 1, column=col, value=header)
            cell.fill = PatternFill("solid", fgColor="F3F4F6")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        write_row = table_row + 2
        for location in sorted({row["location"] for row in latest_ranked}):
            rows = [
                row
                for row in latest_ranked
                if row["location"] == location and row.get("guest_count", 0) > 0
            ]
            for visible_rank, selected in enumerate(selector(rows), start=1):
                if value_field == "average_ticket_time_seconds":
                    value = duration_fraction(selected[value_field])
                else:
                    value = selected[value_field]
                ws.cell(row=write_row, column=1, value=location)
                ws.cell(row=write_row, column=2, value=visible_rank)
                ws.cell(
                    row=write_row,
                    column=3,
                    value=excel_safe_text(selected["display_name"]),
                )
                ws.cell(row=write_row, column=4, value=selected["guest_count"])
                ws.cell(row=write_row, column=5, value=value)
                ws.cell(row=write_row, column=6, value=selected["active_days"])
                ws.cell(row=write_row, column=5).number_format = number_format
                write_row += 1
        table_row = write_row + 2

    for row_index in range(4, ws.max_row + 1):
        for col_index in range(1, min(ws.max_column, 13) + 1):
            cell = ws.cell(row=row_index, column=col_index)
            fill = priority_fill(cell.value)
            if fill:
                cell.fill = fill
        if row_index <= store_end:
            ws.row_dimensions[row_index].height = 42 if row_index >= max(snapshot_end, group_end) + 9 else 32
        else:
            ws.row_dimensions[row_index].height = 18

    if latest_location_rows:
        chart_ws = wb.create_sheet("_Dashboard Chart Data")
        chart_ws.sheet_state = "hidden"
        chart_anchor_row = 10
        helper_row = chart_anchor_row
        helper_location_col = 1
        helper_value_col = 2
        chart_ws.cell(row=helper_row, column=helper_location_col, value="Location")
        chart_ws.cell(row=helper_row, column=helper_value_col, value="Check Average")
        for index, row in enumerate(sorted(latest_location_rows, key=lambda item: item["location"]), start=1):
            chart_ws.cell(row=helper_row + index, column=helper_location_col, value=row["location"])
            chart_ws.cell(row=helper_row + index, column=helper_value_col, value=row["check_average"])
            chart_ws.cell(row=helper_row + index, column=helper_value_col).number_format = "$#,##0.00"
        chart = BarChart()
        chart.type = "col"
        chart.title = "Latest Check Average By Location"
        chart.y_axis.title = "Check Average"
        chart.x_axis.title = "Location"
        data = Reference(chart_ws, min_col=helper_value_col, min_row=helper_row, max_row=helper_row + len(latest_location_rows))
        cats = Reference(chart_ws, min_col=helper_location_col, min_row=helper_row + 1, max_row=helper_row + len(latest_location_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.height = 7
        chart.width = 13
        ws.add_chart(chart, f"H{max(store_end + 2, table_row - 12)}")



def write_data_quality_sheet(
    wb: Workbook,
    records: list[MetricRecord],
    weekly_location_rows: list[dict[str, Any]],
    public_start: date,
    public_end: date,
) -> None:
    source_groups: dict[str, list[MetricRecord]] = defaultdict(list)
    for record in records:
        source_groups[record.source_file].append(record)

    source_headers = [
        "Source File",
        "Report Date",
        "Locations Found",
        "Server Rows",
        "Location Total Rows",
        "Status",
    ]
    source_rows: list[list[Any]] = []
    for source_file, source_records in sorted(source_groups.items()):
        locations = sorted({record.location for record in source_records})
        source_rows.append(
            [
                excel_safe_text(source_file),
                min(record.report_date for record in source_records),
                ", ".join(locations),
                sum(1 for record in source_records if not record.is_location_total),
                sum(1 for record in source_records if record.is_location_total),
                "OK" if len(locations) >= 2 else "Review",
            ]
        )

    ws = wb.create_sheet("Data Quality")
    style_title(ws, "Data Quality Checks", len(source_headers))
    ws.freeze_panes = "A4"
    for col_index, header in enumerate(source_headers, start=1):
        cell = ws.cell(row=3, column=col_index, value=header)
        cell.fill = PatternFill("solid", fgColor="E1E1E1")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(source_rows, start=4):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if col_index == 2:
                cell.number_format = EXCEL_DATE_NUMBER_FORMAT
    if source_rows:
        ref = f"A3:{get_column_letter(len(source_headers))}{len(source_rows) + 3}"
        table = Table(displayName="DataQualitySources", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)

    for col, width in {"A": 56, "B": 14, "C": 30, "D": 14, "E": 18, "F": 14}.items():
        ws.column_dimensions[col].width = width

    all_dates = sorted({record.report_date for record in records})
    check_row = len(source_rows) + 6
    style_section_header(ws, check_row, 1, 4, "Date Coverage")
    date_headers = ["Date", "Status", "Locations With Totals", "Source Files"]
    for col_index, header in enumerate(date_headers, start=1):
        cell = ws.cell(row=check_row + 1, column=col_index, value=header)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.font = Font(bold=True)
    expected_dates = []
    if all_dates:
        current, coverage_end, _ = validated_data_quality_coverage(
            records, public_start, public_end
        )
        while current <= coverage_end:
            if is_operating_day(current):
                expected_dates.append(current)
            current += timedelta(days=1)

    location_totals_by_date: dict[date, set[str]] = defaultdict(set)
    source_files_by_date: dict[date, set[str]] = defaultdict(set)
    for record in records:
        if record.is_location_total:
            location_totals_by_date[record.report_date].add(record.location)
        source_files_by_date[record.report_date].add(record.source_file)

    for offset, report_date in enumerate(expected_dates, start=2):
        row_index = check_row + offset
        status = "OK" if report_date in all_dates else "Missing"
        ws.cell(row=row_index, column=1, value=report_date).number_format = (
            EXCEL_DATE_NUMBER_FORMAT
        )
        ws.cell(row=row_index, column=2, value=status)
        ws.cell(
            row=row_index,
            column=3,
            value=", ".join(sorted(location_totals_by_date.get(report_date, set()))),
        )
        ws.cell(
            row=row_index,
            column=4,
            value=excel_safe_text(
                ", ".join(sorted(source_files_by_date.get(report_date, set())))
            ),
        )
        if status == "Missing":
            ws.cell(row=row_index, column=2).fill = PatternFill("solid", fgColor="F4CCCC")

    location_check_row = check_row + len(expected_dates) + 5
    style_section_header(ws, location_check_row, 1, 5, "Weekly Location Coverage")
    loc_headers = ["Week End", "Location", "Active Days", "Source Days", "Status"]
    for col_index, header in enumerate(loc_headers, start=1):
        cell = ws.cell(row=location_check_row + 1, column=col_index, value=header)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.font = Font(bold=True)
    for offset, row in enumerate(sorted(weekly_location_rows, key=lambda item: (item["week_end"], item["location"])), start=2):
        row_index = location_check_row + offset
        status = "OK" if row["source_days"] >= OPERATING_WEEK_DAYS else "Short Week"
        ws.cell(row=row_index, column=1, value=row["week_end"]).number_format = (
            EXCEL_DATE_NUMBER_FORMAT
        )
        ws.cell(row=row_index, column=2, value=row["location"])
        ws.cell(row=row_index, column=3, value=row["active_days"])
        ws.cell(row=row_index, column=4, value=row["source_days"])
        ws.cell(row=row_index, column=5, value=status)
        if status != "OK":
            ws.cell(row=row_index, column=5).fill = PatternFill("solid", fgColor="FFF2CC")


def metric_week_trend_headers(entity_label: str) -> list[str]:
    return [
        "Week Start",
        "Week End",
        entity_label,
        "Gross Sales",
        "Gross Sales Change",
        "Guest Count",
        "Guest Count Change",
        "Check Average",
        "Check Average Change",
        "Wine Sales",
        "Wine Sales Change",
        "Wine %",
        "Wine % Change",
        "Rate of Sale by Guest Count",
        "Rate Change",
        "Average Ticket Time",
        "Ticket Time Change (Min)",
        "Active Days",
        "Source Days",
        "Trend Note",
    ]


def metric_week_trend_data(rows: list[dict[str, Any]], entity_field: str) -> list[list[Any]]:
    return [
        [
            row["week_start"],
            row["week_end"],
            row[entity_field],
            row["gross_sales"],
            row["gross_sales_change"],
            row["guest_count"],
            row["guest_count_change"],
            row["check_average"],
            row["check_average_change"],
            row["wine_sales"],
            row["wine_sales_change"],
            row["wine_pct"],
            row["wine_pct_change"],
            row["rate_of_sale_by_guest_count"],
            row["rate_change"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["ticket_time_change_minutes"],
            row["active_days"],
            row["source_days"],
            row["trend_note"],
        ]
        for row in rows
    ]


def metric_summary_headers(entity_label: str) -> list[str]:
    return [
        entity_label,
        "Weeks Tracked",
        "Active Days Tracked",
        "Source Days Tracked",
        "Total Gross Sales",
        "Total Guest Count",
        "Weighted Check Average",
        "Total Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
        "First Week End",
        "Latest Week End",
        "Prior Week End",
        "Latest Gross Sales",
        "Prior Gross Sales",
        "Gross Sales Change",
        "Latest Guest Count",
        "Prior Guest Count",
        "Guest Count Change",
        "Latest Check Average",
        "Prior Check Average",
        "Check Average Change",
        "Latest Wine %",
        "Prior Wine %",
        "Wine % Change",
        "Latest Rate of Sale by Guest Count",
        "Prior Rate of Sale by Guest Count",
        "Rate Change",
        "Latest Ticket Time",
        "Prior Ticket Time",
        "Ticket Time Change (Min)",
    ]


def metric_summary_data(rows: list[dict[str, Any]], entity_field: str) -> list[list[Any]]:
    return [
        [
            row[entity_field],
            row["weeks_tracked"],
            row["active_days_tracked"],
            row["source_days_tracked"],
            row["total_gross_sales"],
            row["total_guest_count"],
            row["weighted_check_average"],
            row["total_wine_sales"],
            row["overall_wine_pct"],
            row["weighted_rate"],
            duration_fraction(row["weighted_ticket_time_seconds"]),
            row["first_week_end"],
            row["latest_week_end"],
            row["prior_week_end"],
            row["latest_gross_sales"],
            row["prior_gross_sales"],
            row["gross_sales_change"],
            row["latest_guest_count"],
            row["prior_guest_count"],
            row["guest_count_change"],
            row["latest_check_average"],
            row["prior_check_average"],
            row["check_average_change"],
            row["latest_wine_pct"],
            row["prior_wine_pct"],
            row["wine_pct_change"],
            row["latest_rate"],
            row["prior_rate"],
            row["rate_change"],
            duration_fraction(row["latest_ticket_time_seconds"]),
            duration_fraction(row["prior_ticket_time_seconds"])
            if row["prior_ticket_time_seconds"] is not None
            else None,
            row["ticket_time_change_minutes"],
        ]
        for row in rows
    ]


def write_action_board_sheet(wb: Workbook, action_rows: list[dict[str, Any]]) -> None:
    headers = [
        "Priority",
        "Action",
        "Location",
        "Person / Area",
        "Signal",
        "Score / Impact",
        "Evidence",
        "Recommended Follow-Up",
        "Week End",
        "Guest Count",
        "Active Days",
    ]
    data = [
        [
            excel_safe_text(row["priority"]),
            excel_safe_text(row["action"]),
            excel_safe_text(row["location"]),
            excel_safe_text(row["subject"]),
            excel_safe_text(row["signal"]),
            excel_safe_text(row["impact"]),
            excel_safe_text(row["evidence"]),
            excel_safe_text(row["recommended_follow_up"]),
            row["week_end"],
            row["guest_count"],
            row["active_days"],
        ]
        for row in action_rows
    ]
    ws = write_table_sheet(
        wb,
        "Action Board",
        headers,
        data,
        "ActionBoard",
        widths={
            "A": 13,
            "B": 16,
            "C": 20,
            "D": 24,
            "E": 18,
            "F": 32,
            "G": 44,
            "H": 56,
            "I": 14,
            "J": 12,
            "K": 12,
        },
        index=1,
    )
    ws.sheet_properties.tabColor = "7A1E1E"
    for row in range(4, len(data) + 4):
        priority = ws.cell(row=row, column=1).value
        action = ws.cell(row=row, column=2).value
        signal = ws.cell(row=row, column=5).value
        fill = priority_fill(priority) or priority_fill(action) or priority_fill(signal)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        for col in (6, 7, 8):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 42


def _write_master_workbook_base(
    records: list[MetricRecord],
    output_path: Path,
    config: dict[str, Any],
    source_dir: Path,
    public_start: date,
    public_end: date,
) -> Path:
    weekly_server_rows, weekly_location_rows = weekly_rollups(records)
    rank_min_guest_count = int(
        config.get("master_min_guest_count_for_rankings", config.get("public_min_guest_count", 1))
    )
    ranked_rows = weekly_server_rank_rows(weekly_server_rows, rank_min_guest_count)
    server_trend_detail_rows = server_week_trend_rows(ranked_rows)
    server_star_detail_rows = server_star_rows(server_trend_detail_rows, config)
    trend_rows = trend_summary_rows(weekly_server_rows)
    store_week_trend_detail_rows = weekly_metric_trend_rows(weekly_location_rows, ("location",))
    store_trend_summary = weekly_metric_summary_rows(weekly_location_rows, ("location",))
    weekly_group_rows = group_weekly_rows(records)
    group_week_trend_detail_rows = weekly_metric_trend_rows(weekly_group_rows, ("group",))
    group_trend_summary = weekly_metric_summary_rows(weekly_group_rows, ("group",))
    action_board_rows = build_action_board_rows(
        server_star_detail_rows,
        store_trend_summary,
        group_trend_summary,
        weekly_location_rows,
    )

    wb = Workbook()
    wb.remove(wb.active)
    write_dashboard_sheet(
        wb,
        records,
        weekly_location_rows,
        ranked_rows,
        server_star_detail_rows,
        action_board_rows,
        store_trend_summary,
        group_trend_summary,
        config,
        source_dir,
        public_start,
        public_end,
    )
    write_action_board_sheet(wb, action_board_rows)

    star_headers = [
        "Category",
        "Suggested Action",
        "Priority",
        "Composite Score",
        "Week Start",
        "Week End",
        "Location",
        "Raw Server",
        "Display Name",
        "Guest Count",
        "Active Days",
        "Check Average",
        "Check Average Change",
        "Wine %",
        "Wine % Change",
        "Rate of Sale by Guest Count",
        "Rate Change",
        "Average Ticket Time",
        "Ticket Time Change (Min)",
        "Average Rank Movement",
        "Why",
    ]
    star_data = [
        [
            row["category"],
            star_action(row)[1],
            star_action(row)[0],
            row["composite_score"],
            row["week_start"],
            row["week_end"],
            row["location"],
            row["raw_user_name"],
            row["display_name"],
            row["guest_count"],
            row["active_days"],
            row["check_average"],
            row["check_average_change"],
            row["wine_pct"],
            row["wine_pct_change"],
            row["rate_of_sale_by_guest_count"],
            row["rate_change"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["ticket_time_change_minutes"],
            row["average_rank_movement"],
            row["why"],
        ]
        for row in server_star_detail_rows
    ]
    star_ws = write_table_sheet(
        wb,
        "Rising & Falling Stars",
        star_headers,
        star_data,
        "RisingFallingStars",
        widths={
            "A": 16,
            "B": 16,
            "C": 14,
            "D": 16,
            "E": 14,
            "F": 14,
            "G": 22,
            "H": 28,
            "I": 28,
            "P": 24,
            "U": 54,
        },
    )
    for row_index in range(4, len(star_data) + 4):
        fill = (
            priority_fill(star_ws.cell(row=row_index, column=1).value)
            or priority_fill(star_ws.cell(row=row_index, column=2).value)
            or priority_fill(star_ws.cell(row=row_index, column=3).value)
        )
        if fill:
            for col_index in range(1, len(star_headers) + 1):
                star_ws.cell(row=row_index, column=col_index).fill = fill
        star_ws.cell(row=row_index, column=len(star_headers)).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        star_ws.row_dimensions[row_index].height = 32

    server_headers = [
        "Week Start",
        "Week End",
        "Location",
        "Raw Server",
        "Display Name",
        "Gross Sales",
        "Guest Count",
        "Check Average",
        "Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
        "Active Days",
        "Source Days",
    ]
    server_data = [
        [
            row["week_start"],
            row["week_end"],
            row["location"],
            row["raw_user_name"],
            row["display_name"],
            row["gross_sales"],
            row["guest_count"],
            row["check_average"],
            row["wine_sales"],
            row["wine_pct"],
            row["rate_of_sale_by_guest_count"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["active_days"],
            row["source_days"],
        ]
        for row in sorted(
            weekly_server_rows,
            key=lambda item: (item["week_end"], item["location"], item["display_name"]),
        )
    ]
    write_table_sheet(wb, "Weekly Server Metrics", server_headers, server_data, "WeeklyServerMetrics")

    ranking_headers = [
        "Week Start",
        "Week End",
        "Location",
        "Raw Server",
        "Display Name",
        "Guest Count",
        "Check Average",
        "Check Average Rank",
        "Wine %",
        "Wine % Rank",
        "Rate of Sale by Guest Count",
        "Rate Rank",
        "Average Ticket Time",
        "Ticket Time Rank",
        "Active Days",
    ]
    ranking_data = [
        [
            row["week_start"],
            row["week_end"],
            row["location"],
            row["raw_user_name"],
            row["display_name"],
            row["guest_count"],
            row["check_average"],
            row["check_average_rank"],
            row["wine_pct"],
            row["wine_pct_rank"],
            row["rate_of_sale_by_guest_count"],
            row["rate_rank"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["ticket_time_rank"],
            row["active_days"],
        ]
        for row in ranked_rows
    ]
    write_table_sheet(
        wb,
        "Weekly Server Rankings",
        ranking_headers,
        ranking_data,
        "WeeklyServerRankings",
        widths={"A": 14, "B": 14, "C": 22, "D": 28, "E": 28, "K": 22, "M": 18},
    )

    trend_detail_headers = [
        "Week Start",
        "Week End",
        "Location",
        "Raw Server",
        "Display Name",
        "Guest Count",
        "Check Average",
        "Check Average Rank",
        "Prior Check Average Rank",
        "Check Average Rank Movement",
        "Check Average Change",
        "Wine %",
        "Wine % Rank",
        "Prior Wine % Rank",
        "Wine % Rank Movement",
        "Wine % Change",
        "Rate of Sale by Guest Count",
        "Rate Rank",
        "Prior Rate Rank",
        "Rate Rank Movement",
        "Rate Change",
        "Average Ticket Time",
        "Ticket Time Rank",
        "Prior Ticket Time Rank",
        "Ticket Time Rank Movement",
        "Ticket Time Change (Min)",
        "Active Days",
        "Trend Note",
    ]
    trend_detail_data = [
        [
            row["week_start"],
            row["week_end"],
            row["location"],
            row["raw_user_name"],
            row["display_name"],
            row["guest_count"],
            row["check_average"],
            row["check_average_rank"],
            row["prior_check_average_rank"],
            row["check_average_rank_movement"],
            row["check_average_change"],
            row["wine_pct"],
            row["wine_pct_rank"],
            row["prior_wine_pct_rank"],
            row["wine_pct_rank_movement"],
            row["wine_pct_change"],
            row["rate_of_sale_by_guest_count"],
            row["rate_rank"],
            row["prior_rate_rank"],
            row["rate_rank_movement"],
            row["rate_change"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["ticket_time_rank"],
            row["prior_ticket_time_rank"],
            row["ticket_time_rank_movement"],
            row["ticket_time_change_minutes"],
            row["active_days"],
            row["trend_note"],
        ]
        for row in server_trend_detail_rows
    ]
    write_table_sheet(
        wb,
        "Server Week-over-Week Detail",
        trend_detail_headers,
        trend_detail_data,
        "ServerWeekTrends",
        widths={
            "A": 14,
            "B": 14,
            "C": 22,
            "D": 28,
            "E": 28,
            "I": 22,
            "J": 24,
            "M": 18,
            "N": 20,
            "O": 22,
            "R": 20,
            "S": 20,
            "T": 22,
            "W": 22,
            "X": 24,
            "Y": 22,
        },
    )

    location_headers = [
        "Week Start",
        "Week End",
        "Location",
        "Gross Sales",
        "Guest Count",
        "Check Average",
        "Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
        "Active Days",
        "Source Days",
    ]
    location_data = [
        [
            row["week_start"],
            row["week_end"],
            row["location"],
            row["gross_sales"],
            row["guest_count"],
            row["check_average"],
            row["wine_sales"],
            row["wine_pct"],
            row["rate_of_sale_by_guest_count"],
            duration_fraction(row["average_ticket_time_seconds"]),
            row["active_days"],
            row["source_days"],
        ]
        for row in sorted(weekly_location_rows, key=lambda item: (item["week_end"], item["location"]))
    ]
    write_table_sheet(wb, "Weekly Location Metrics", location_headers, location_data, "WeeklyLocationMetrics")

    trend_tab_widths = {
        "A": 14,
        "B": 14,
        "C": 24,
        "E": 20,
        "G": 20,
        "I": 22,
        "K": 20,
        "M": 16,
        "N": 24,
        "Q": 22,
        "T": 16,
    }
    write_table_sheet(
        wb,
        "Store Week Trends",
        metric_week_trend_headers("Location"),
        metric_week_trend_data(store_week_trend_detail_rows, "location"),
        "StoreWeekTrends",
        widths=trend_tab_widths,
    )
    write_table_sheet(
        wb,
        "Store Trend Summary",
        metric_summary_headers("Location"),
        metric_summary_data(store_trend_summary, "location"),
        "StoreTrendSummary",
        widths={
            "A": 24,
            "J": 24,
            "K": 18,
            "Q": 20,
            "W": 22,
            "AA": 24,
            "AC": 16,
            "AF": 22,
        },
    )
    write_table_sheet(
        wb,
        "Group Week Trends",
        metric_week_trend_headers("Group"),
        metric_week_trend_data(group_week_trend_detail_rows, "group"),
        "GroupWeekTrends",
        widths=trend_tab_widths,
    )
    write_table_sheet(
        wb,
        "Group Trend Summary",
        metric_summary_headers("Group"),
        metric_summary_data(group_trend_summary, "group"),
        "GroupTrendSummary",
        widths={
            "A": 18,
            "J": 24,
            "K": 18,
            "Q": 20,
            "W": 22,
            "AA": 24,
            "AC": 16,
            "AF": 22,
        },
    )

    daily_server_headers = [
        "Report Date",
        "Location",
        "Raw Server",
        "Display Name",
        "Source File",
        "Gross Sales",
        "Guest Count",
        "Check Average",
        "Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
    ]
    daily_server_data = [
        [
            record.report_date,
            record.location,
            record.raw_user_name,
            record.display_name,
            record.source_file,
            record.gross_sales,
            record.guest_count,
            record.check_average,
            record.wine_sales,
            record.wine_pct,
            record.rate_of_sale_by_guest_count,
            duration_fraction(record.average_ticket_time_seconds),
        ]
        for record in sorted(
            (item for item in records if not item.is_location_total),
            key=lambda item: (item.report_date, item.location, item.display_name),
        )
    ]
    write_table_sheet(wb, "Daily Server Detail", daily_server_headers, daily_server_data, "DailyServerDetail")

    daily_location_data = [
        [
            record.report_date,
            record.location,
            record.source_file,
            record.gross_sales,
            record.guest_count,
            record.check_average,
            record.wine_sales,
            record.wine_pct,
            record.rate_of_sale_by_guest_count,
            duration_fraction(record.average_ticket_time_seconds),
        ]
        for record in sorted(
            (item for item in records if item.is_location_total),
            key=lambda item: (item.report_date, item.location),
        )
    ]
    write_table_sheet(
        wb,
        "Daily Location Detail",
        [
            "Report Date",
            "Location",
            "Source File",
            "Gross Sales",
            "Guest Count",
            "Check Average",
            "Wine Sales",
            "Wine %",
            "Rate of Sale by Guest Count",
            "Average Ticket Time",
        ],
        daily_location_data,
        "DailyLocationDetail",
    )

    trend_headers = [
        "Location",
        "Raw Server",
        "Display Name",
        "Weeks Tracked",
        "Active Days Tracked",
        "Total Gross Sales",
        "Total Guest Count",
        "Weighted Check Average",
        "Total Wine Sales",
        "Wine %",
        "Rate of Sale by Guest Count",
        "Average Ticket Time",
        "First Week End",
        "Latest Week End",
        "Prior Week End",
        "Latest Check Average",
        "Prior Check Average",
        "Check Average Change",
        "Latest Wine %",
        "Prior Wine %",
        "Wine % Change",
        "Latest Rate of Sale by Guest Count",
        "Prior Rate of Sale by Guest Count",
        "Rate Change",
        "Latest Ticket Time",
        "Prior Ticket Time",
        "Ticket Time Change (Min)",
    ]
    trend_data = [
        [
            row["location"],
            row["raw_user_name"],
            row["display_name"],
            row["weeks_tracked"],
            row["active_days_tracked"],
            row["total_gross_sales"],
            row["total_guest_count"],
            row["weighted_check_average"],
            row["total_wine_sales"],
            row["overall_wine_pct"],
            row["weighted_rate"],
            duration_fraction(row["weighted_ticket_time_seconds"]),
            row["first_week_end"],
            row["latest_week_end"],
            row["prior_week_end"],
            row["latest_check_average"],
            row["prior_check_average"],
            row["check_average_change"],
            row["latest_wine_pct"],
            row["prior_wine_pct"],
            row["wine_pct_change"],
            row["latest_rate"],
            row["prior_rate"],
            row["rate_change"],
            duration_fraction(row["latest_ticket_time_seconds"]),
            duration_fraction(row["prior_ticket_time_seconds"])
            if row["prior_ticket_time_seconds"] is not None
            else None,
            row["ticket_time_change_minutes"],
        ]
        for row in trend_rows
    ]
    write_table_sheet(
        wb,
        "Server Trend Summary",
        trend_headers,
        trend_data,
        "ServerTrendSummary",
        widths={
            "A": 22,
            "B": 28,
            "C": 28,
            "D": 14,
            "E": 18,
            "K": 22,
            "L": 18,
            "R": 20,
            "U": 16,
            "V": 24,
            "X": 14,
            "AA": 22,
        },
    )

    write_data_quality_sheet(wb, records, weekly_location_rows, public_start, public_end)

    notes = wb.create_sheet("Run Notes", 2)
    notes.sheet_view.showGridLines = False
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 90
    notes["A1"] = "Red Onion Server Master"
    notes["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    notes["A1"].fill = PatternFill("solid", fgColor="7A1E1E")
    notes.merge_cells("A1:B1")
    note_rows = [
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source Folder", str(source_dir)),
        ("Operating Week", f"{OPERATING_WEEK_LABEL}; Mondays are closed and excluded from weekly rollups."),
        (
            "Report Date Source",
            "Uses each raw workbook's Date(s) value; filename date fallback is one day after the business date.",
        ),
        ("Raw Reports Read", len({record.source_file for record in records})),
        ("Date Coverage", format_date_range(min(r.report_date for r in records), max(r.report_date for r in records))),
        ("Public Snapshot Dates", format_date_range(public_start, public_end)),
        ("Public Exclude Patterns", ", ".join(config.get("public_exclude_name_contains", []))),
        ("Dashboard Exclude Patterns", ", ".join(config.get("dashboard_exclude_name_contains", []))),
        ("Ranking Minimum Guest Count", rank_min_guest_count),
        (
            "Dashboard Trend Eligibility",
            "Server trend lists include rows with guest count >= "
            f"{config.get('dashboard_min_guest_count_for_trends', 25)} or active days >= "
            f"{config.get('dashboard_min_active_days_for_trends', 3)}.",
        ),
        (
            "Dashboard",
            "Dashboard highlights coach-first items, recognition opportunities, store action pulses, group pulse, and current metric leaders.",
        ),
        (
            "Action Board",
            "Prioritized follow-up list for coaching, recognition, store review, group review, and data-quality checks.",
        ),
        (
            "Rising/Falling Stars",
            "Person trend labels use Sales / Guest and Wine % only. Rate of Sale and Ticket Time remain context only.",
        ),
        (
            "Trend Tabs",
            "Server, store, and group trend tabs show week-over-week changes plus latest/prior/total-period summaries.",
        ),
        ("Metric Rule", "Sales / Guest and Wine % are recalculated from rolled-up sales, guests, and wine sales."),
        ("Metric Rule", "Rate of Sale uses opportunities divided by inferred qualifying sales. Ticket Time is Check Count-weighted only when Check Count coverage is complete."),
    ]
    for row_index, (label, value) in enumerate(note_rows, start=3):
        notes.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        notes.cell(row=row_index, column=2, value=value)
        notes.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return datetime.fromisoformat(value.strip()).date()
        except ValueError:
            return None
    return None


def safe_pct_delta(current: float | None, comparison: float | None) -> float | None:
    if current is None or comparison in (None, 0):
        return None
    return (current - comparison) / comparison


def management_threshold(
    config: dict[str, Any],
    field: str,
    *,
    comparator: str = "movement",
) -> dict[str, Any]:
    family = (
        "management_peer_score_thresholds"
        if comparator == "peer"
        else "management_score_thresholds"
    )
    defaults = DEFAULT_CONFIG[family][field]
    configured = config.get(family, {}).get(field, {})
    return {**defaults, **configured}


def directional_score(change: float | None, threshold: dict[str, Any]) -> int:
    if change is None:
        return 0
    directional = -change if threshold.get("lower_is_better") else change
    neutral = float(threshold["neutral"])
    strong = float(threshold["strong"])
    if directional >= strong:
        return 2
    if directional >= neutral:
        return 1
    if directional <= -strong:
        return -2
    if directional <= -neutral:
        return -1
    return 0


def directional_variance(current: float, benchmark: float, threshold: dict[str, Any]) -> float:
    variance = current - benchmark
    return -variance if threshold.get("lower_is_better") else variance


def operational_context_rollup(rows: Iterable[dict[str, Any]]) -> dict[str, Any]:
    selected = list(rows)
    active_rows = [
        row
        for row in selected
        if (
            float(row.get("guest_count", 0) or 0) > 0
            or float(row.get("gross_sales", 0) or 0) > 0
        )
    ]
    check_count_available = bool(active_rows) and all(
        bool(row.get("check_count_available", False)) for row in active_rows
    )
    check_count = sum(
        float(row.get("check_count", 0) or 0)
        for row in selected
        if bool(row.get("check_count_available", False))
    )
    per_check_available = bool(check_count_available and check_count > 0)

    rate_rows = [
        row
        for row in selected
        if float(row.get("guest_count", 0) or 0) > 0
    ]
    rate_available = bool(rate_rows) and all(
        bool(row.get("rate_available", True))
        and float(row.get("rate_of_sale_by_guest_count", 0) or 0) > 0
        for row in rate_rows
    )
    rate_opportunities = sum(
        float(row.get("guest_count", 0) or 0) for row in rate_rows
    )
    rate_qualifying_sales = sum(
        float(row.get("guest_count", 0) or 0)
        / float(row.get("rate_of_sale_by_guest_count", 0) or 0)
        for row in rate_rows
        if bool(row.get("rate_available", True))
        and float(row.get("rate_of_sale_by_guest_count", 0) or 0) > 0
    )
    rate_available = bool(
        rate_available and rate_opportunities > 0 and rate_qualifying_sales > 0
    )

    ticket_rows = [
        row
        for row in selected
        if (
            bool(row.get("check_count_available", False))
            and float(row.get("check_count", 0) or 0) > 0
        )
    ]
    ticket_time_available = bool(
        check_count_available
        and ticket_rows
        and all(bool(row.get("ticket_time_available", True)) for row in ticket_rows)
    )
    ticket_weight = sum(
        float(row.get("check_count", 0) or 0) for row in ticket_rows
    )
    ticket_weighted = sum(
        float(row.get("average_ticket_time_seconds", 0) or 0)
        * float(row.get("check_count", 0) or 0)
        for row in ticket_rows
        if bool(row.get("ticket_time_available", True))
    )
    return {
        "check_count": check_count,
        "check_count_available": check_count_available,
        "per_check_available": per_check_available,
        "sales_per_check": 0.0,
        "guests_per_check": 0.0,
        "rate_of_sale_by_guest_count": (
            rate_opportunities / rate_qualifying_sales if rate_available else 0.0
        ),
        "rate_available": rate_available,
        "average_ticket_time_seconds": (
            ticket_weighted / ticket_weight
            if ticket_time_available and ticket_weight
            else 0.0
        ),
        "ticket_time_available": ticket_time_available,
        "ticket_time_weight_basis": (
            "Check Count"
            if ticket_time_available
            else "Unavailable (Check Count missing or incomplete)"
        ),
    }


def aggregate_weekly_rows(rows: Iterable[dict[str, Any]]) -> dict[str, Any] | None:
    selected = list(rows)
    if not selected:
        return None
    gross_sales = sum(float(row.get("gross_sales", 0) or 0) for row in selected)
    guest_count = sum(float(row.get("guest_count", 0) or 0) for row in selected)
    wine_sales = sum(float(row.get("wine_sales", 0) or 0) for row in selected)
    context = operational_context_rollup(selected)
    if context["per_check_available"]:
        context["sales_per_check"] = gross_sales / context["check_count"]
        context["guests_per_check"] = guest_count / context["check_count"]
    return {
        "gross_sales": gross_sales,
        "guest_count": guest_count,
        "check_average": gross_sales / guest_count if guest_count else 0.0,
        "wine_sales": wine_sales,
        "wine_pct": wine_sales / gross_sales if gross_sales else 0.0,
        **context,
    }


def shared_pos_number(row: dict[str, Any]) -> str | None:
    """Return a leading four-digit shared POS identity, when one is present."""

    for field in ("raw_user_name", "display_name"):
        value = row.get(field)
        text = "" if is_blank(value) else str(value).strip()
        match = re.match(r"^(\d{4})\b", text)
        if match:
            return match.group(1)
    return None


def weekly_area_for_row(
    row: dict[str, Any], config: dict[str, Any]
) -> str | None:
    """Map a weekly POS row into one non-person area without guessing shared IDs."""

    number = shared_pos_number(row)
    if number is not None:
        mapped_area = config.get("weekly_shared_number_areas", {}).get(number)
        return str(mapped_area) if mapped_area in WEEKLY_AREA_ORDER else None
    patterns = config.get(
        "weekly_area_name_patterns",
        DEFAULT_CONFIG.get("weekly_area_name_patterns", {}),
    )
    # Specific multi-word concepts precede broader matches such as Bar.
    for area in ("Wine Dinners", "Patio", "Banquets", "Bar", "Dining Room"):
        if row_matches_weekly_area_patterns(row, patterns.get(area, [])):
            return area
    # Remaining dashboard-eligible named people are the dining-room population.
    # Managers, takeout, generic placeholders, and other excluded identities are
    # intentionally not swept into this residual bucket.
    if dashboard_excluded(row, config):
        return None
    return "Dining Room"


def weekly_trend_summary(
    weekly_metrics: dict[date, dict[str, Any]],
    week_ends: tuple[date, ...],
) -> dict[str, Any]:
    latest_week = week_ends[-1] if week_ends else None
    prior_week = week_ends[-2] if len(week_ends) >= 2 else None
    latest = weekly_metrics.get(latest_week) if latest_week else None
    prior = weekly_metrics.get(prior_week) if prior_week else None
    window = aggregate_weekly_rows(weekly_metrics.values())
    return {
        "weekly_metrics": weekly_metrics,
        "available_weeks": len(weekly_metrics),
        "latest_gross_sales": latest["gross_sales"] if latest else None,
        "latest_guest_count": latest["guest_count"] if latest else None,
        "latest_check_average": latest["check_average"] if latest else None,
        "latest_wine_pct": latest["wine_pct"] if latest else None,
        "wow_check_average": (
            float(latest["check_average"]) - float(prior["check_average"])
            if latest is not None and prior is not None
            else None
        ),
        "window_check_average": window["check_average"] if window else None,
    }


def build_shared_area_trends_model(
    weekly_server_rows: list[dict[str, Any]],
    weekly_location_rows: list[dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, Any]:
    """Build high-level shared-number and operating-area weekly trend views."""

    _, global_full = full_week_ends_by_location(
        weekly_location_rows,
        config.get("locations", {}),
    )
    week_ends = tuple(sorted(global_full)[-WEEKLY_SHARED_AREA_WINDOW_WEEKS:])
    if not week_ends:
        return {
            "week_ends": (),
            "latest_complete_week_end": None,
            "shared_rows": [],
            "area_rows": [],
        }

    selected = set(week_ends)
    shared_buckets: dict[tuple[str, str, date], list[dict[str, Any]]] = defaultdict(list)
    area_buckets: dict[tuple[str, str, date], list[dict[str, Any]]] = defaultdict(list)
    shared_labels: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in weekly_server_rows:
        week_end = row.get("week_end")
        if week_end not in selected:
            continue
        location = str(row.get("location") or "")
        number = shared_pos_number(row)
        if number is not None:
            shared_buckets[(location, number, week_end)].append(row)
            shared_labels[(location, number)].add(
                str(row.get("raw_user_name") or row.get("display_name") or number)
            )
        area = weekly_area_for_row(row, config)
        if area is None:
            continue
        area_buckets[(location, area, week_end)].append(row)
        area_buckets[("All Stores", area, week_end)].append(row)

    shared_rows: list[dict[str, Any]] = []
    for location, number in sorted(shared_labels):
        weekly_metrics = {
            week_end: aggregate_weekly_rows(
                shared_buckets[(location, number, week_end)]
            )
            for week_end in week_ends
            if shared_buckets.get((location, number, week_end))
        }
        normalized = {
            week_end: value
            for week_end, value in weekly_metrics.items()
            if value is not None
        }
        shared_rows.append(
            {
                "location": location,
                "shared_number": number,
                "pos_label": ", ".join(sorted(shared_labels[(location, number)])),
                **weekly_trend_summary(normalized, week_ends),
            }
        )

    configured_locations = [str(location) for location in config.get("locations", {})]
    location_order = ["All Stores", *configured_locations]
    area_rows: list[dict[str, Any]] = []
    for location in location_order:
        for area in WEEKLY_AREA_ORDER:
            weekly_metrics = {
                week_end: aggregate_weekly_rows(area_buckets[(location, area, week_end)])
                for week_end in week_ends
                if area_buckets.get((location, area, week_end))
            }
            normalized = {
                week_end: value
                for week_end, value in weekly_metrics.items()
                if value is not None
            }
            mapping_status = (
                "Awaiting POS mapping"
                if area == "Wine Dinners" and not normalized
                else "Named-server fallback"
                if area == "Dining Room"
                else "Configured name match"
            )
            area_rows.append(
                {
                    "location": location,
                    "area": area,
                    "mapping_status": mapping_status,
                    **weekly_trend_summary(normalized, week_ends),
                }
            )
    return {
        "week_ends": week_ends,
        "latest_complete_week_end": week_ends[-1],
        "shared_rows": shared_rows,
        "area_rows": area_rows,
    }


def overall_store_performance(
    weekly_location_rows: list[dict[str, Any]],
    week_ends: Iterable[date],
) -> dict[str, Any] | None:
    selected_weeks = tuple(sorted(set(week_ends)))
    if not selected_weeks:
        return None
    weekly = {
        week_end: aggregate_weekly_rows(
            row for row in weekly_location_rows if row.get("week_end") == week_end
        )
        for week_end in selected_weeks
    }
    latest = weekly.get(selected_weeks[-1])
    if latest is None:
        return None
    baseline_weeks = selected_weeks[:-1][-PERFORMANCE_CONSISTENCY_BLOCK_WEEKS:]
    baseline_rows = [weekly[week_end] for week_end in baseline_weeks if weekly.get(week_end)]
    baseline = aggregate_weekly_rows(baseline_rows)
    baseline_count = len(baseline_rows)
    baseline_sales = (
        sum(float(row["gross_sales"]) for row in baseline_rows) / baseline_count
        if baseline_count
        else None
    )
    baseline_guests = (
        sum(float(row["guest_count"]) for row in baseline_rows) / baseline_count
        if baseline_count
        else None
    )

    def pct_change(current: float, comparison: float | None) -> float | None:
        return current / comparison - 1 if comparison else None

    return {
        "week_end": selected_weeks[-1],
        "complete_location_count": len(
            {
                str(row.get("location") or "")
                for row in weekly_location_rows
                if row.get("week_end") == selected_weeks[-1]
                and int(row.get("source_days", 0) or 0) >= OPERATING_WEEK_DAYS
            }
        ),
        "baseline_week_count": baseline_count,
        "gross_sales": float(latest["gross_sales"]),
        "gross_sales_change_pct": pct_change(
            float(latest["gross_sales"]), baseline_sales
        ),
        "guest_count": float(latest["guest_count"]),
        "guest_count_change_pct": pct_change(
            float(latest["guest_count"]), baseline_guests
        ),
        "check_average": float(latest["check_average"]),
        "check_average_change": (
            float(latest["check_average"]) - float(baseline["check_average"])
            if baseline
            else None
        ),
        "wine_pct": float(latest["wine_pct"]),
        "wine_pct_change": (
            float(latest["wine_pct"]) - float(baseline["wine_pct"])
            if baseline
            else None
        ),
    }


def management_trend_classification(
    metric_scores: dict[str, int],
    rank_modifier: int,
    *,
    improving_label: str,
    declining_label: str,
) -> tuple[str, int]:
    composite_score = sum(metric_scores.values()) + rank_modifier
    positive_count = sum(score > 0 for score in metric_scores.values())
    negative_count = sum(score < 0 for score in metric_scores.values())
    if composite_score >= 3 and positive_count >= 2:
        return improving_label, composite_score
    if composite_score <= -3 and negative_count >= 2:
        return declining_label, composite_score
    return "Stable", composite_score


def management_rank_block_movement(
    recent_rows: list[dict[str, Any]],
    comparison_rows: list[dict[str, Any]],
    rank_lookup: dict[tuple[date, str, str], dict[str, Any]],
) -> tuple[float | None, int]:
    movements: list[float] = []
    for rank_field in (
        "check_average_rank",
        "wine_pct_rank",
    ):
        recent_ranks = [
            rank_lookup[(row["week_end"], row["location"], row["raw_user_name"])].get(rank_field)
            for row in recent_rows
            if (row["week_end"], row["location"], row["raw_user_name"]) in rank_lookup
            and rank_lookup[(row["week_end"], row["location"], row["raw_user_name"])].get(rank_field)
            is not None
        ]
        comparison_ranks = [
            rank_lookup[(row["week_end"], row["location"], row["raw_user_name"])].get(rank_field)
            for row in comparison_rows
            if (row["week_end"], row["location"], row["raw_user_name"]) in rank_lookup
            and rank_lookup[(row["week_end"], row["location"], row["raw_user_name"])].get(rank_field)
            is not None
        ]
        if recent_ranks and comparison_ranks:
            movements.append(
                sum(float(value) for value in comparison_ranks) / len(comparison_ranks)
                - sum(float(value) for value in recent_ranks) / len(recent_ranks)
            )
    average_movement = sum(movements) / len(movements) if movements else None
    modifier = (
        1
        if average_movement is not None and average_movement >= 3
        else -1
        if average_movement is not None and average_movement <= -3
        else 0
    )
    return average_movement, modifier


def server_trend_context(recent_momentum: str, long_term_direction: str) -> str:
    contexts = {
        ("Rising", "Improving"): "The recent gain continues a sustained longer-term improvement.",
        ("Rising", "Declining"): "This is a recent rebound against a softer longer-term trend.",
        ("Falling", "Improving"): "This is recent softening within a stronger longer-term trend.",
        ("Falling", "Declining"): "The recent decline continues a sustained longer-term decline.",
        ("Stable", "Improving"): "Recent results are stable while the longer-term direction is improving.",
        ("Stable", "Declining"): "Recent results are stable while the longer-term direction is declining.",
    }
    return contexts.get((recent_momentum, long_term_direction), "")


def full_week_ends_by_location(
    weekly_location_rows: list[dict[str, Any]],
    required_locations: Iterable[str] | None = None,
) -> tuple[dict[str, set[date]], set[date]]:
    observed_locations = {str(row["location"]) for row in weekly_location_rows}
    configured_locations = {
        str(location) for location in (required_locations or ()) if str(location)
    }
    locations = sorted(configured_locations or observed_locations)
    by_location = {
        location: {
            row["week_end"]
            for row in weekly_location_rows
            if row["location"] == location and row.get("source_days", 0) >= OPERATING_WEEK_DAYS
        }
        for location in locations
    }
    if not locations:
        return by_location, set()
    global_full = set.intersection(*(by_location[location] for location in locations))
    return by_location, global_full


def performance_consistency_weighted_average(
    rows: Iterable[dict[str, Any]], field: str
) -> float | None:
    selected = [
        row
        for row in rows
        if row.get(field) is not None and float(row.get("guest_weight", 0) or 0) > 0
    ]
    total_weight = sum(float(row["guest_weight"]) for row in selected)
    if not selected or total_weight <= 0:
        return None
    return sum(float(row[field]) * float(row["guest_weight"]) for row in selected) / total_weight


def performance_consistency_level(
    spg_gap: float | None,
    wine_gap: float | None,
    *,
    spg_boundary: float,
    wine_boundary: float,
) -> str:
    """Classify peer-adjusted selling outcomes without collapsing the two axes."""

    if spg_gap is None or wine_gap is None:
        return "Not Evaluated"
    if (
        (spg_gap >= spg_boundary and wine_gap >= 0)
        or (wine_gap >= wine_boundary and spg_gap >= 0)
    ):
        return "Strong"
    if (
        (spg_gap <= -spg_boundary and wine_gap <= 0)
        or (wine_gap <= -wine_boundary and spg_gap <= 0)
    ):
        return "Below"
    return "On Track / Mixed"


def performance_consistency_weekly_band(
    spg_gap: float | None,
    wine_gap: float | None,
    *,
    spg_boundary: float,
    wine_boundary: float,
) -> str:
    level = performance_consistency_level(
        spg_gap,
        wine_gap,
        spg_boundary=spg_boundary,
        wine_boundary=wine_boundary,
    )
    if level in {"Strong", "Below"}:
        return level
    if level == "Not Evaluated":
        return "Not Qualified"
    return "Near Peer" if float(spg_gap) * float(wine_gap) >= 0 else "Mixed"


def build_performance_consistency_model(
    weekly_server_rows: list[dict[str, Any]],
    weekly_location_rows: list[dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, Any]:
    """Build the protected eight-week performance/consistency review layer.

    The model intentionally remains separate from the gated Action Board signal.
    It uses only complete shared weeks, a current-roster snapshot, same-store
    leave-one-person-out weekly medians, and explicit sample/confidence gates.
    """

    _, global_full = full_week_ends_by_location(
        weekly_location_rows,
        config.get("locations", {}),
    )
    week_ends = sorted(global_full)[-PERFORMANCE_CONSISTENCY_WINDOW_WEEKS:]
    if not week_ends:
        return {
            "methodology_version": PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION,
            "week_ends": (),
            "complete_week_count": 0,
            "latest_complete_week_end": None,
            "summary_rows": [],
            "detail_rows": [],
            "category_counts": {
                "Consistently Strong": 0,
                "Consistent / Near Peer": 0,
                "Variable / Inconsistent": 0,
                "Below / Review": 0,
                "Insufficient Data": 0,
            },
            "store_consistency": [],
            "overall_store": None,
        }

    latest_complete_week_end = week_ends[-1]
    spg_threshold = management_threshold(config, "check_average")
    wine_threshold = management_threshold(config, "wine_pct")
    spg_boundary = float(spg_threshold["neutral"])
    wine_boundary = float(wine_threshold["neutral"])
    spg_moderate_max = float(spg_threshold["strong"])
    wine_moderate_max = float(wine_threshold["strong"])
    min_peers = int(
        config.get("management_peer_reference", {}).get(
            "min_distinct_peers_per_week",
            DEFAULT_CONFIG["management_peer_reference"]["min_distinct_peers_per_week"],
        )
    )

    current_rows = [
        row
        for row in weekly_server_rows
        if row["week_end"] == latest_complete_week_end
        and shared_pos_number(row) is None
        and not dashboard_excluded(row, config)
    ]
    roster = {
        (str(row["location"]), str(row["raw_user_name"])): row
        for row in current_rows
    }
    selected_weeks = set(week_ends)
    row_lookup = {
        (row["week_end"], str(row["location"]), str(row["raw_user_name"])): row
        for row in weekly_server_rows
        if row["week_end"] in selected_weeks
        and shared_pos_number(row) is None
        and not dashboard_excluded(row, config)
    }
    qualified_peer_pool: dict[tuple[date, str], list[dict[str, Any]]] = defaultdict(list)
    for row in weekly_server_rows:
        identity = (str(row["location"]), str(row["raw_user_name"]))
        if (
            row["week_end"] in selected_weeks
            and identity in roster
            and shared_pos_number(row) is None
            and not dashboard_excluded(row, config)
            and dashboard_trend_eligible(row, config)
        ):
            qualified_peer_pool[(row["week_end"], str(row["location"]))].append(row)

    detail_rows: list[dict[str, Any]] = []
    detail_by_server: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    weekly_bands_by_server: dict[tuple[str, str], dict[date, str]] = {
        key: {week_end: "—" for week_end in week_ends} for key in roster
    }
    recent_week_ends = set(week_ends[-PERFORMANCE_CONSISTENCY_BLOCK_WEEKS:])

    for location, raw_user_name in sorted(roster):
        roster_row = roster[(location, raw_user_name)]
        for week_end in week_ends:
            row = row_lookup.get((week_end, location, raw_user_name))
            if row is None:
                continue
            base_sample = dashboard_trend_eligible(row, config)
            peers = (
                [
                    peer
                    for peer in qualified_peer_pool.get((week_end, location), [])
                    if str(peer["raw_user_name"]) != raw_user_name
                ]
                if base_sample
                else []
            )
            distinct_peer_count = len(
                {str(peer["raw_user_name"]) for peer in peers}
            )
            qualified = base_sample and distinct_peer_count >= min_peers
            peer_spg = (
                float(median(float(peer["check_average"]) for peer in peers))
                if qualified
                else None
            )
            peer_wine = (
                float(median(float(peer["wine_pct"]) for peer in peers))
                if qualified
                else None
            )
            spg_gap = (
                float(row["check_average"]) - peer_spg
                if peer_spg is not None
                else None
            )
            wine_gap = (
                float(row["wine_pct"]) - peer_wine
                if peer_wine is not None
                else None
            )
            weekly_band = performance_consistency_weekly_band(
                spg_gap,
                wine_gap,
                spg_boundary=spg_boundary,
                wine_boundary=wine_boundary,
            )
            guest_count = float(row.get("guest_count", 0) or 0)
            detail = {
                "week_end": week_end,
                "location": location,
                "raw_user_name": raw_user_name,
                "display_name": str(roster_row.get("display_name") or raw_user_name),
                "gross_sales": float(row.get("gross_sales", 0) or 0),
                "guest_count": guest_count,
                "check_average": float(row.get("check_average", 0) or 0),
                "wine_sales": float(row.get("wine_sales", 0) or 0),
                "wine_pct": float(row.get("wine_pct", 0) or 0),
                "active_days": int(row.get("active_days", 0) or 0),
                "base_sample": base_sample,
                "distinct_peer_count": distinct_peer_count if base_sample else 0,
                "peer_spg": peer_spg,
                "peer_wine_pct": peer_wine,
                "spg_gap": spg_gap,
                "wine_gap": wine_gap,
                "guest_weight": min(
                    guest_count, PERFORMANCE_CONSISTENCY_GUEST_WEIGHT_CAP
                )
                if qualified
                else 0.0,
                "qualified": qualified,
                "qualified_gross_sales": (
                    float(row.get("gross_sales", 0) or 0) if qualified else None
                ),
                "qualified_guests": guest_count if qualified else None,
                "qualified_wine_sales": (
                    float(row.get("wine_sales", 0) or 0) if qualified else None
                ),
                "period": "Recent 4" if week_end in recent_week_ends else "Prior 4",
                "weekly_band": weekly_band,
            }
            detail_rows.append(detail)
            detail_by_server[(location, raw_user_name)].append(detail)
            weekly_bands_by_server[(location, raw_user_name)][week_end] = weekly_band

    summary_rows: list[dict[str, Any]] = []
    for key, current in roster.items():
        location, raw_user_name = key
        qualified_rows = [row for row in detail_by_server[key] if row["qualified"]]
        qualified_weeks = len(qualified_rows)
        qualified_guests = sum(float(row["guest_count"]) for row in qualified_rows)
        qualified_sales = sum(float(row["gross_sales"]) for row in qualified_rows)
        qualified_wine_sales = sum(float(row["wine_sales"]) for row in qualified_rows)
        eight_week_spg = (
            qualified_sales / qualified_guests if qualified_guests else None
        )
        eight_week_wine_pct = (
            qualified_wine_sales / qualified_sales if qualified_sales else None
        )
        average_spg_gap = performance_consistency_weighted_average(
            qualified_rows, "spg_gap"
        )
        average_wine_gap = performance_consistency_weighted_average(
            qualified_rows, "wine_gap"
        )
        spg_values = [float(row["spg_gap"]) for row in qualified_rows]
        wine_values = [float(row["wine_gap"]) for row in qualified_rows]
        weekly_spg_gap_sd = float(stdev(spg_values)) if len(spg_values) >= 2 else None
        weekly_wine_gap_sd = float(stdev(wine_values)) if len(wine_values) >= 2 else None

        performance_level = performance_consistency_level(
            average_spg_gap,
            average_wine_gap,
            spg_boundary=spg_boundary,
            wine_boundary=wine_boundary,
        )
        confidence = (
            "High"
            if qualified_weeks >= PERFORMANCE_CONSISTENCY_HIGH_MIN_WEEKS
            and qualified_guests >= PERFORMANCE_CONSISTENCY_HIGH_MIN_GUESTS
            else "Provisional"
            if qualified_weeks >= PERFORMANCE_CONSISTENCY_PROVISIONAL_MIN_WEEKS
            and qualified_guests >= PERFORMANCE_CONSISTENCY_PROVISIONAL_MIN_GUESTS
            else "Insufficient"
        )
        if confidence == "Insufficient":
            consistency = "Insufficient"
        elif (
            weekly_spg_gap_sd is not None
            and weekly_wine_gap_sd is not None
            and weekly_spg_gap_sd <= spg_boundary
            and weekly_wine_gap_sd <= wine_boundary
        ):
            consistency = "High"
        elif (
            weekly_spg_gap_sd is not None
            and weekly_wine_gap_sd is not None
            and weekly_spg_gap_sd <= spg_moderate_max
            and weekly_wine_gap_sd <= wine_moderate_max
        ):
            consistency = "Moderate"
        else:
            consistency = "Low"

        if confidence == "Insufficient":
            overall_read = "Insufficient Data"
        elif performance_level == "Strong":
            overall_read = (
                "Consistently Strong" if consistency == "High" else "Strong / Variable"
            )
        elif performance_level == "Below":
            overall_read = (
                "Consistently Below" if consistency == "High" else "Below / Variable"
            )
        elif consistency == "High":
            overall_read = "Consistent / Near Peer"
        elif consistency == "Moderate":
            overall_read = "Mixed / Monitor"
        else:
            overall_read = "Inconsistent"

        recent_rows = [row for row in qualified_rows if row["period"] == "Recent 4"]
        prior_rows = [row for row in qualified_rows if row["period"] == "Prior 4"]
        recent_spg = performance_consistency_weighted_average(recent_rows, "spg_gap")
        prior_spg = performance_consistency_weighted_average(prior_rows, "spg_gap")
        recent_wine = performance_consistency_weighted_average(recent_rows, "wine_gap")
        prior_wine = performance_consistency_weighted_average(prior_rows, "wine_gap")
        recent_spg_delta = (
            recent_spg - prior_spg
            if recent_spg is not None and prior_spg is not None
            else None
        )
        recent_wine_delta = (
            recent_wine - prior_wine
            if recent_wine is not None and prior_wine is not None
            else None
        )
        performance_score = (
            float(average_spg_gap) / spg_boundary
            + float(average_wine_gap) / wine_boundary
            if average_spg_gap is not None and average_wine_gap is not None
            else float("-inf")
        )
        summary_rows.append(
            {
                "location": location,
                "raw_user_name": raw_user_name,
                "display_name": str(current.get("display_name") or raw_user_name),
                "overall_read": overall_read,
                "performance_level": performance_level,
                "consistency": consistency,
                "confidence": confidence,
                "eight_week_spg": eight_week_spg,
                "average_spg_gap": average_spg_gap,
                "eight_week_wine_pct": eight_week_wine_pct,
                "average_wine_gap": average_wine_gap,
                "weekly_spg_gap_sd": weekly_spg_gap_sd,
                "weekly_wine_gap_sd": weekly_wine_gap_sd,
                "qualified_weeks": qualified_weeks,
                "qualified_guests": qualified_guests,
                "recent_spg_delta": recent_spg_delta,
                "recent_wine_delta": recent_wine_delta,
                "performance_score": performance_score,
                "weekly_bands": weekly_bands_by_server[key],
            }
        )

    category_order = {
        "Consistently Strong": 0,
        "Consistent / Near Peer": 1,
        "Strong / Variable": 2,
        "Mixed / Monitor": 3,
        "Inconsistent": 4,
        "Consistently Below": 5,
        "Below / Variable": 6,
        "Insufficient Data": 7,
    }
    summary_rows.sort(
        key=lambda row: (
            category_order.get(str(row["overall_read"]), 99),
            -float(row["performance_score"]),
            str(row["location"]),
            str(row["display_name"]),
        )
    )
    detail_rows.sort(
        key=lambda row: (
            row["week_end"],
            str(row["location"]),
            str(row["display_name"]),
        )
    )

    category_counts = {
        "Consistently Strong": sum(
            row["overall_read"] == "Consistently Strong" for row in summary_rows
        ),
        "Consistent / Near Peer": sum(
            row["overall_read"] == "Consistent / Near Peer" for row in summary_rows
        ),
        "Variable / Inconsistent": sum(
            row["overall_read"] in {"Strong / Variable", "Mixed / Monitor", "Inconsistent"}
            for row in summary_rows
        ),
        "Below / Review": sum(
            row["overall_read"] in {"Consistently Below", "Below / Variable"}
            for row in summary_rows
        ),
        "Insufficient Data": sum(
            row["overall_read"] == "Insufficient Data" for row in summary_rows
        ),
    }
    store_consistency = []
    for location in sorted({str(row["location"]) for row in summary_rows}):
        location_rows = [row for row in summary_rows if row["location"] == location]
        store_consistency.append(
            {
                "location": location,
                "High": sum(row["consistency"] == "High" for row in location_rows),
                "Moderate": sum(
                    row["consistency"] == "Moderate" for row in location_rows
                ),
                "Low": sum(row["consistency"] == "Low" for row in location_rows),
                "Insufficient": sum(
                    row["consistency"] == "Insufficient" for row in location_rows
                ),
            }
        )
    return {
        "methodology_version": PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION,
        "week_ends": tuple(week_ends),
        "complete_week_count": len(global_full),
        "latest_complete_week_end": latest_complete_week_end,
        "summary_rows": summary_rows,
        "detail_rows": detail_rows,
        "category_counts": category_counts,
        "store_consistency": store_consistency,
        "overall_store": overall_store_performance(weekly_location_rows, week_ends),
    }


def metric_driver(field: str, change: float) -> str:
    if field == "check_average":
        return f"Check avg {format_change(change, 'currency')}"
    if field == "wine_pct":
        return f"Wine {format_change(change, 'pct_points')}"
    if field == "rate_of_sale_by_guest_count":
        direction = "improved" if change < 0 else "worsened"
        return f"Rate {direction} {abs(change):.3f}"
    minutes = abs(change) / 60
    direction = "faster" if change < 0 else "slower"
    return f"Ticket {minutes:.1f} min {direction}"


def merged_source_evidence(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    evidence: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        for item in row.get("source_evidence", []) or []:
            if not isinstance(item, dict):
                continue
            key = (
                str(item.get("source_file") or ""),
                str(item.get("sha256") or ""),
                str(item.get("report_date") or ""),
            )
            evidence[key] = {
                "source_file": excel_safe_text(key[0]),
                "sha256": key[1] or None,
                "report_date": key[2] or None,
                "format": item.get("format"),
                "parser_engine": item.get("parser_engine"),
                "report_date_source": item.get("report_date_source") or "Unknown",
            }
    return [evidence[key] for key in sorted(evidence)]


def recommended_server_follow_up(row: dict[str, Any]) -> str:
    action = str(row.get("action") or "Monitor")
    recurring = row.get("recurring_driver_fields", []) or []
    driver_text = ", ".join(
        field.replace("_", " ")
        for field in recurring
    )
    context_prompt = (
        "What changed in shift mix, section load, guest mix, check volume, staffing, "
        "or service conditions, and do manager observations corroborate the validated "
        "Sales / Guest and Wine % signal?"
    )
    if action == "Coaching Prompt":
        suffix = f" Recurring drivers: {driver_text}." if driver_text else ""
        return (
            "After reviewing comparable work context, what one measurable coaching "
            f"step is supported by the Sales / Guest and Wine % evidence?{suffix}"
        )
    if action == "Recognition Prompt":
        suffix = f" Recurring drivers: {driver_text}." if driver_text else ""
        return (
            "After reviewing comparable work context, what supported practice should "
            f"be recognized and shared?{suffix}"
        )
    if action == "Context Review":
        return context_prompt
    return (
        "What should be watched during the next complete, qualified week before "
        "opening any coaching or recognition prompt?"
    )


def management_metric_available(row: dict[str, Any], field: str) -> bool:
    if field == "check_count":
        return bool(row.get("check_count_available", False))
    if field in {"sales_per_check", "guests_per_check"}:
        return bool(
            row.get(
                "per_check_available",
                bool(row.get("check_count_available", False))
                and float(row.get("check_count", 0) or 0) > 0,
            )
        )
    if field == "rate_of_sale_by_guest_count":
        return bool(row.get("rate_available", True))
    if field == "average_ticket_time_seconds":
        return bool(row.get("ticket_time_available", True))
    value = row.get(field)
    return isinstance(value, (int, float)) and math.isfinite(float(value))


def management_metric_band(
    config: dict[str, Any],
    field: str,
    *,
    comparator: str = "movement",
) -> MetricBand:
    threshold = management_threshold(config, field, comparator=comparator)
    return MetricBand(
        neutral=float(threshold["neutral"]),
        strong=float(threshold["strong"]),
    )


def management_metric_score(
    change: float | None,
    config: dict[str, Any],
    field: str,
    *,
    comparator: str = "movement",
) -> int | None:
    threshold = management_threshold(config, field, comparator=comparator)
    return score_metric(
        change,
        management_metric_band(config, field, comparator=comparator),
        higher_is_better=not bool(threshold.get("lower_is_better")),
    )


def classification_from_metric_scores(
    scores: dict[str, int | None],
    *,
    positive_label: str,
    negative_label: str,
    unavailable_label: str,
) -> tuple[str, int]:
    if any(scores.get(field) is None for field in SERVER_TREND_FIELDS):
        return unavailable_label, 0
    return management_trend_classification(
        {field: int(scores[field] or 0) for field in SERVER_TREND_FIELDS},
        0,
        improving_label=positive_label,
        declining_label=negative_label,
    )


def weekly_row_from_daily_records(
    daily_records: Iterable[dict[str, Any]],
) -> dict[str, Any] | None:
    rows = list(daily_records)
    if not rows:
        return None
    gross_sales = sum(float(row.get("gross_sales", 0) or 0) for row in rows)
    guest_count = sum(float(row.get("guest_count", 0) or 0) for row in rows)
    wine_sales = sum(float(row.get("wine_sales", 0) or 0) for row in rows)
    active_dates = {
        as_date(row.get("report_date"))
        for row in rows
        if (
            float(row.get("guest_count", 0) or 0) > 0
            or float(row.get("gross_sales", 0) or 0) > 0
        )
    }
    context = operational_context_rollup(rows)
    if context["per_check_available"]:
        context["sales_per_check"] = gross_sales / context["check_count"]
        context["guests_per_check"] = guest_count / context["check_count"]
    return {
        "gross_sales": gross_sales,
        "guest_count": guest_count,
        "check_average": gross_sales / guest_count if guest_count else 0.0,
        "wine_sales": wine_sales,
        "wine_pct": wine_sales / gross_sales if gross_sales else 0.0,
        **context,
        "active_days": len({value for value in active_dates if value is not None}),
        "source_days": len({value for value in active_dates if value is not None}),
        "daily_records": rows,
    }


def evaluate_server_week_signal(
    current: dict[str, Any],
    weekly_server_rows: list[dict[str, Any]],
    full_by_location: dict[str, set[date]],
    config: dict[str, Any],
) -> dict[str, Any]:
    location = current["location"]
    current_week_end = current["week_end"]
    baseline_limit = int(config.get("dashboard_baseline_full_weeks", 4))
    min_prior_weeks = int(config.get("dashboard_min_prior_full_weeks", 2))
    min_prior_guests = float(config.get("dashboard_min_prior_guest_count", 50))
    peer_config = config.get("management_peer_reference", {})
    peer_prior_limit = int(peer_config.get("prior_full_weeks", 4))
    min_peer_weeks = int(peer_config.get("min_prior_full_weeks", 3))
    min_peers_per_week = int(
        peer_config.get("min_distinct_peers_per_week", 5)
    )
    min_peer_server_weeks = int(peer_config.get("min_peer_server_weeks", 20))
    eligible_week_ends = sorted(
        week_end
        for week_end in full_by_location.get(location, set())
        if week_end < current_week_end
    )
    self_week_ends = eligible_week_ends[-baseline_limit:]
    peer_week_ends = eligible_week_ends[-peer_prior_limit:]
    prior_rows = [
        row
        for row in weekly_server_rows
        if row["location"] == location
        and row["raw_user_name"] == current["raw_user_name"]
        and row["week_end"] in self_week_ends
    ]
    baseline = aggregate_weekly_rows(prior_rows)
    prior_guest_count = sum(
        float(row.get("guest_count", 0) or 0) for row in prior_rows
    )
    full_latest = current_week_end in full_by_location.get(location, set())
    current_sample_eligible = dashboard_trend_eligible(current, config)
    self_history_eligible = (
        len(prior_rows) >= min_prior_weeks
        and prior_guest_count >= min_prior_guests
    )

    changes: dict[str, float | None] = {}
    metric_scores: dict[str, int | None] = {}
    for field in SERVER_TREND_FIELDS:
        available = (
            baseline is not None
            and management_metric_available(current, field)
            and management_metric_available(baseline, field)
        )
        change = (
            float(current[field]) - float(baseline[field])
            if available and baseline is not None
            else None
        )
        changes[field] = change
        metric_scores[field] = management_metric_score(change, config, field)
    movement, composite_score = classification_from_metric_scores(
        metric_scores,
        positive_label=RecentMovement.UPWARD.value,
        negative_label=RecentMovement.DOWNWARD.value,
        unavailable_label=RecentMovement.NOT_EVALUATED.value,
    )

    peer_references: dict[str, Any] = {}
    peer_scores: dict[str, int | None] = {}
    peer_changes: dict[str, float | None] = {}
    for field in SERVER_TREND_FIELDS:
        observations = [
            PeerObservation(
                person_id=str(row.get("raw_user_name") or ""),
                location=str(row.get("location") or ""),
                week_end=row["week_end"],
                value=(
                    float(row[field])
                    if management_metric_available(row, field)
                    else None
                ),
                qualified=dashboard_trend_eligible(row, config),
                excluded=dashboard_excluded(row, config),
            )
            for row in weekly_server_rows
            if row["location"] == location and row["week_end"] in peer_week_ends
        ]
        reference = leave_one_out_same_store_peer_reference(
            observations,
            focal_person_id=str(current.get("raw_user_name") or ""),
            location=location,
            prior_week_ends=peer_week_ends,
            max_prior_weeks=peer_prior_limit,
            min_usable_weeks=min_peer_weeks,
            min_distinct_peers_per_week=min_peers_per_week,
            min_peer_weeks=min_peer_server_weeks,
        )
        peer_references[field] = reference
        peer_change = (
            float(current[field]) - float(reference.value)
            if reference.sufficient
            and reference.value is not None
            and management_metric_available(current, field)
            else None
        )
        peer_changes[field] = peer_change
        peer_scores[field] = management_metric_score(
            peer_change,
            config,
            field,
            comparator="peer",
        )
    peer_reference_available = all(
        reference.sufficient for reference in peer_references.values()
    )
    peer_score_label, peer_composite_score = classification_from_metric_scores(
        peer_scores,
        positive_label=PeerComparison.ABOVE.value,
        negative_label=PeerComparison.BELOW.value,
        unavailable_label=PeerComparison.UNAVAILABLE.value,
    )
    if (
        peer_score_label == "Stable"
        and peer_reference_available
    ):
        peer_score_label = PeerComparison.WITHIN.value

    combined_scores: dict[str, int | None] = {}
    for field in SERVER_TREND_FIELDS:
        movement_score = metric_scores.get(field)
        peer_score = peer_scores.get(field)
        if movement_score is None or peer_score is None:
            combined_scores[field] = None
        elif movement_score and peer_score and movement_score * peer_score > 0:
            direction = 1 if movement_score > 0 else -1
            combined_scores[field] = direction * min(
                abs(movement_score), abs(peer_score)
            )
        else:
            combined_scores[field] = 0
    candidate = classify_candidate(
        movement,
        peer_score_label,
        combined_scores,
        required_metrics=SERVER_TREND_FIELDS,
    )

    current_peer_rows = [
        row
        for row in weekly_server_rows
        if row["location"] == location
        and row["week_end"] == current_week_end
        and str(row.get("raw_user_name") or "").casefold()
        != str(current.get("raw_user_name") or "").casefold()
        and dashboard_trend_eligible(row, config)
        and not dashboard_excluded(row, config)
    ]
    current_peer_changes: dict[str, float | None] = {}
    for field in SERVER_TREND_FIELDS:
        valid_values = [
            float(row[field])
            for row in current_peer_rows
            if management_metric_available(row, field)
        ]
        reference = peer_references[field]
        current_peer_changes[field] = (
            float(median(valid_values)) - float(reference.value)
            if len(valid_values) >= min_peers_per_week
            and reference.sufficient
            and reference.value is not None
            else None
        )
    shock = assess_common_store_shock(
        candidate.polarity,
        peer_changes,
        current_peer_changes,
        {
            field: management_metric_band(config, field, comparator="peer")
            for field in SERVER_TREND_FIELDS
        },
        {
            field: not bool(
                management_threshold(
                    config,
                    field,
                    comparator="peer",
                ).get("lower_is_better")
            )
            for field in SERVER_TREND_FIELDS
        },
        candidate_drivers=candidate.agreeing_drivers,
    )
    base_qualified = (
        full_latest
        and current_sample_eligible
        and self_history_eligible
        and peer_reference_available
    )
    candidate_qualified = base_qualified and candidate.eligible
    full_candidate_polarity = (
        candidate.polarity
        if candidate_qualified and shock.guard_passed
        else CandidatePolarity.NONE
    )

    daily_records = list(current.get("daily_records", []) or [])
    leave_one_day_polarities: list[CandidatePolarity] = []
    if candidate_qualified and shock.guard_passed and daily_records:
        active_dates = sorted(
            {
                as_date(item.get("report_date"))
                for item in daily_records
                if (
                    float(item.get("guest_count", 0) or 0) > 0
                    or float(item.get("gross_sales", 0) or 0) > 0
                )
                and as_date(item.get("report_date")) is not None
            }
        )
        for removed_date in active_dates:
            reduced = weekly_row_from_daily_records(
                item
                for item in daily_records
                if as_date(item.get("report_date")) != removed_date
            )
            if reduced is None or not dashboard_trend_eligible(reduced, config):
                leave_one_day_polarities.append(CandidatePolarity.NONE)
                continue
            reduced_movement_scores: dict[str, int | None] = {}
            reduced_peer_scores: dict[str, int | None] = {}
            reduced_peer_changes: dict[str, float | None] = {}
            for field in SERVER_TREND_FIELDS:
                movement_change = (
                    float(reduced[field]) - float(baseline[field])
                    if baseline is not None
                    and management_metric_available(reduced, field)
                    and management_metric_available(baseline, field)
                    else None
                )
                peer_reference = peer_references[field]
                peer_change = (
                    float(reduced[field]) - float(peer_reference.value)
                    if peer_reference.sufficient
                    and peer_reference.value is not None
                    and management_metric_available(reduced, field)
                    else None
                )
                reduced_movement_scores[field] = management_metric_score(
                    movement_change, config, field
                )
                reduced_peer_scores[field] = management_metric_score(
                    peer_change,
                    config,
                    field,
                    comparator="peer",
                )
                reduced_peer_changes[field] = peer_change
            reduced_movement, _ = classification_from_metric_scores(
                reduced_movement_scores,
                positive_label=RecentMovement.UPWARD.value,
                negative_label=RecentMovement.DOWNWARD.value,
                unavailable_label=RecentMovement.NOT_EVALUATED.value,
            )
            reduced_peer_label, _ = classification_from_metric_scores(
                reduced_peer_scores,
                positive_label=PeerComparison.ABOVE.value,
                negative_label=PeerComparison.BELOW.value,
                unavailable_label=PeerComparison.UNAVAILABLE.value,
            )
            if reduced_peer_label == "Stable":
                reduced_peer_label = PeerComparison.WITHIN.value
            reduced_combined: dict[str, int | None] = {}
            for field in SERVER_TREND_FIELDS:
                left = reduced_movement_scores[field]
                right = reduced_peer_scores[field]
                if left is None or right is None:
                    reduced_combined[field] = None
                elif left and right and left * right > 0:
                    reduced_combined[field] = (
                        1 if left > 0 else -1
                    ) * min(abs(left), abs(right))
                else:
                    reduced_combined[field] = 0
            reduced_candidate = classify_candidate(
                reduced_movement,
                reduced_peer_label,
                reduced_combined,
                required_metrics=SERVER_TREND_FIELDS,
            )
            reduced_shock = assess_common_store_shock(
                reduced_candidate.polarity,
                reduced_peer_changes,
                current_peer_changes,
                {
                    field: management_metric_band(
                        config,
                        field,
                        comparator="peer",
                    )
                    for field in SERVER_TREND_FIELDS
                },
                {
                    field: not bool(
                        management_threshold(
                            config,
                            field,
                            comparator="peer",
                        ).get(
                            "lower_is_better"
                        )
                    )
                    for field in SERVER_TREND_FIELDS
                },
                candidate_drivers=reduced_candidate.agreeing_drivers,
            )
            leave_one_day_polarities.append(
                reduced_candidate.polarity
                if reduced_candidate.eligible and reduced_shock.guard_passed
                else CandidatePolarity.NONE
            )
    # When the store-shock guard alone blocks an otherwise qualified candidate,
    # day-removal stability is not the failing gate. Preserve a neutral/pass
    # value here so the persistence result reports the actual store-shock cause.
    day_stable = bool(candidate_qualified and not shock.guard_passed) or (
        leave_one_day_stability(full_candidate_polarity, leave_one_day_polarities)
    )

    if not full_latest:
        evidence_status = "Incomplete Week"
    elif not current_sample_eligible:
        evidence_status = "Limited Volume"
    elif not self_history_eligible:
        evidence_status = "Limited History"
    elif any(
        not management_metric_available(current, field)
        or baseline is None
        or not management_metric_available(baseline, field)
        for field in SERVER_TREND_FIELDS
    ):
        evidence_status = "Data Issue"
    elif not peer_reference_available:
        evidence_status = "Reference Unavailable"
    elif candidate_qualified and not shock.guard_passed:
        evidence_status = "Store Shock"
    elif candidate_qualified and not day_stable:
        evidence_status = "Sensitive"
    elif candidate_qualified and shock.guard_passed:
        evidence_status = "Stable"
    else:
        evidence_status = "Eligible"

    return {
        "prior_rows": prior_rows,
        "prior_week_ends": self_week_ends,
        "prior_guest_count": prior_guest_count,
        "baseline": baseline,
        "full_latest": full_latest,
        "current_sample_eligible": current_sample_eligible,
        "self_history_eligible": self_history_eligible,
        "changes": changes,
        "metric_scores": metric_scores,
        "composite_score": composite_score,
        "descriptive_movement": movement,
        "movement": movement if base_qualified else RecentMovement.NOT_EVALUATED.value,
        "peer_comparison": (
            peer_score_label
            if peer_reference_available
            else PeerComparison.UNAVAILABLE.value
        ),
        "peer_composite_score": peer_composite_score,
        "peer_changes": peer_changes,
        "peer_scores": peer_scores,
        "peer_references": peer_references,
        "peer_reference_available": peer_reference_available,
        "combined_scores": combined_scores,
        "candidate": candidate,
        "candidate_qualified": candidate_qualified,
        "store_shock": shock,
        "day_stable": day_stable,
        "leave_one_day_polarities": leave_one_day_polarities,
        "evidence_status": evidence_status,
    }


def management_server_rows(
    weekly_server_rows: list[dict[str, Any]],
    weekly_location_rows: list[dict[str, Any]],
    ranked_rows: list[dict[str, Any]],
    targets: dict[str, dict[str, float | None]],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    if not weekly_server_rows:
        return []
    latest_week_end = max(row["week_end"] for row in weekly_server_rows)
    long_term_limit = int(config.get("dashboard_long_term_full_weeks", 8))
    long_term_block = int(config.get("dashboard_long_term_block_weeks", 4))
    full_min_recent_guests = float(
        config.get("dashboard_long_term_full_min_recent_guests", 100)
    )
    full_min_earlier_guests = float(
        config.get("dashboard_long_term_full_min_earlier_guests", 100)
    )
    developing_min_total = int(
        config.get("dashboard_long_term_developing_min_total_weeks", 6)
    )
    developing_min_recent = int(
        config.get("dashboard_long_term_developing_min_recent_weeks", 3)
    )
    developing_min_earlier = int(
        config.get("dashboard_long_term_developing_min_earlier_weeks", 2)
    )
    developing_min_recent_guests = float(
        config.get("dashboard_long_term_developing_min_recent_guests", 75)
    )
    developing_min_earlier_guests = float(
        config.get("dashboard_long_term_developing_min_earlier_guests", 50)
    )
    full_by_location, _ = full_week_ends_by_location(weekly_location_rows)
    evaluations: dict[tuple[date, str, str], dict[str, Any]] = {}

    def evaluation_for(row: dict[str, Any]) -> dict[str, Any]:
        key = (row["week_end"], row["location"], row["raw_user_name"])
        if key not in evaluations:
            evaluations[key] = evaluate_server_week_signal(
                row, weekly_server_rows, full_by_location, config
            )
        return evaluations[key]

    output: list[dict[str, Any]] = []
    for current in weekly_server_rows:
        if current["week_end"] != latest_week_end or dashboard_excluded(current, config):
            continue
        location = current["location"]
        evaluated = evaluation_for(current)
        prior_rows = evaluated["prior_rows"]
        changes = evaluated["changes"]
        metric_scores = evaluated["metric_scores"]
        positive_drivers: list[str] = []
        negative_drivers: list[str] = []
        for field in SERVER_TREND_FIELDS:
            change = changes[field]
            score = metric_scores[field]
            if change is not None and score > 0:
                positive_drivers.append(metric_driver(field, change))
            elif change is not None and score < 0:
                negative_drivers.append(metric_driver(field, change))
        movement = evaluated["movement"]

        history_week_ends = sorted(
            week_end
            for week_end in full_by_location.get(location, set())
            if week_end <= latest_week_end
        )[-long_term_limit:]
        recent_history_ends = set(history_week_ends[-long_term_block:])
        earlier_history_ends = set(
            history_week_ends[-(long_term_block * 2) : -long_term_block]
        )
        server_history_rows = [
            row
            for row in weekly_server_rows
            if row["location"] == location
            and row["raw_user_name"] == current["raw_user_name"]
            and row["week_end"] in set(history_week_ends)
        ]
        recent_history_rows = [
            row for row in server_history_rows if row["week_end"] in recent_history_ends
        ]
        earlier_history_rows = [
            row for row in server_history_rows if row["week_end"] in earlier_history_ends
        ]
        recent_history_guests = sum(
            float(row.get("guest_count", 0) or 0) for row in recent_history_rows
        )
        earlier_history_guests = sum(
            float(row.get("guest_count", 0) or 0) for row in earlier_history_rows
        )
        history_total_weeks = len(server_history_rows)
        history_total_guests = recent_history_guests + earlier_history_guests
        full_history = (
            len(recent_history_rows) == long_term_block
            and len(earlier_history_rows) == long_term_block
            and history_total_weeks == long_term_block * 2
            and recent_history_guests >= full_min_recent_guests
            and earlier_history_guests >= full_min_earlier_guests
        )
        developing_history = (
            history_total_weeks >= developing_min_total
            and len(recent_history_rows) >= developing_min_recent
            and len(earlier_history_rows) >= developing_min_earlier
            and recent_history_guests >= developing_min_recent_guests
            and earlier_history_guests >= developing_min_earlier_guests
        )
        history_label = (
            "Full" if full_history else "Developing" if developing_history else "Building History"
        )
        history_through = max(
            (row["week_end"] for row in server_history_rows), default=None
        )
        history_used = (
            f"{history_total_weeks} weeks / {history_total_guests:,.0f} guests"
            + (f" through {history_through:%m/%d/%Y}" if history_through else "")
            + (
                f" (recent {len(recent_history_rows)}w / {recent_history_guests:,.0f}g; "
                f"earlier {len(earlier_history_rows)}w / {earlier_history_guests:,.0f}g)"
            )
        )
        recent_history = aggregate_weekly_rows(recent_history_rows)
        earlier_history = aggregate_weekly_rows(earlier_history_rows)
        long_term_changes: dict[str, float | None] = {}
        long_term_metric_scores: dict[str, int | None] = {}
        for field in SERVER_TREND_FIELDS:
            available = bool(
                recent_history
                and earlier_history
                and management_metric_available(recent_history, field)
                and management_metric_available(earlier_history, field)
            )
            change = (
                recent_history[field] - earlier_history[field]
                if available and recent_history and earlier_history
                else None
            )
            long_term_changes[field] = change
            long_term_metric_scores[field] = management_metric_score(
                change, config, field
            )
        scored_long_term_direction, long_term_composite_score = classification_from_metric_scores(
            long_term_metric_scores,
            positive_label=RecentMovement.UPWARD.value,
            negative_label=RecentMovement.DOWNWARD.value,
            unavailable_label=RecentMovement.NOT_EVALUATED.value,
        )
        long_term_direction = (
            scored_long_term_direction
            if evaluated["full_latest"] and history_label in {"Full", "Developing"}
            else RecentMovement.NOT_EVALUATED.value
        )

        benchmark_values = {
            field: evaluated["peer_references"][field].value
            for field in SERVER_TREND_FIELDS
        }
        benchmark_sources: dict[str, str] = {}
        level_statuses: dict[str, str] = {}
        for field in SERVER_TREND_FIELDS:
            target_value = targets.get(location, {}).get(field)
            benchmark_sources[field] = (
                "Same-store prior-four-week median"
                if evaluated["peer_references"][field].sufficient
                else "Reference unavailable"
            )
            score = evaluated["peer_scores"][field] if "peer_scores" in evaluated else None
            if score is None:
                level_statuses[field] = "Unavailable"
            else:
                level_statuses[field] = (
                    "Above"
                    if score > 0
                    else "Below"
                    if score < 0
                    else "Within"
                )
        peer_comparison = evaluated["peer_comparison"]

        previous_row = next(
            (
                row
                for row in weekly_server_rows
                if row["location"] == location
                and row["raw_user_name"] == current["raw_user_name"]
                and row["week_end"] == latest_week_end - timedelta(days=7)
            ),
            None,
        )
        week_over_week_changes = {
            field: (
                float(current[field]) - float(previous_row[field])
                if previous_row is not None
                and management_metric_available(current, field)
                and management_metric_available(previous_row, field)
                else None
            )
            for field in SERVER_TREND_FIELDS
        }
        week_over_week_movement = trend_note(
            week_over_week_changes["check_average"],
            week_over_week_changes["wine_pct"],
            None,
            None,
        )
        previous_evaluation = evaluation_for(previous_row) if previous_row else None
        current_candidate = evaluated["candidate"]
        current_week_signal = WeeklyCandidateSignal(
            week_end=latest_week_end,
            polarity=current_candidate.polarity,
            drivers=current_candidate.agreeing_drivers,
            qualified=evaluated["candidate_qualified"],
            leave_one_day_stable=evaluated["day_stable"],
            store_shock_guard_passed=evaluated["store_shock"].guard_passed,
        )
        previous_week_signal = None
        if previous_row is not None and previous_evaluation is not None:
            previous_candidate = previous_evaluation["candidate"]
            previous_week_signal = WeeklyCandidateSignal(
                week_end=previous_row["week_end"],
                polarity=previous_candidate.polarity,
                drivers=previous_candidate.agreeing_drivers,
                qualified=previous_evaluation["candidate_qualified"],
                leave_one_day_stable=previous_evaluation["day_stable"],
                store_shock_guard_passed=previous_evaluation[
                    "store_shock"
                ].guard_passed,
            )
        persistence = evaluate_two_week_persistence(
            current_week_signal, previous_week_signal
        )
        action = persistence.action.value
        if action == PromptAction.COACHING_PROMPT.value:
            priority = "Medium"
        elif action == PromptAction.RECOGNITION_PROMPT.value:
            priority = "Recognize"
        elif action == PromptAction.CONTEXT_REVIEW.value:
            priority = "Review"
        else:
            priority = "Monitor"
        evidence_status = evaluated["evidence_status"]
        if persistence.reason == "day_sensitive":
            evidence_status = "Sensitive"
        prominent = action != PromptAction.MONITOR.value

        why_parts = []
        if positive_drivers:
            why_parts.append("Improving: " + "; ".join(positive_drivers[:3]))
        if negative_drivers:
            why_parts.append("Watch: " + "; ".join(negative_drivers[:3]))
        if peer_comparison != PeerComparison.UNAVAILABLE.value:
            why_parts.append(f"Peer comparison {peer_comparison}")
        if long_term_direction != RecentMovement.NOT_EVALUATED.value:
            why_parts.append(f"8-week direction {long_term_direction} ({history_label})")
        if evaluated["store_shock"].common_store_shock:
            why_parts.append("Common-store movement guard applied")
        row = {
            **current,
            "prior_weeks": len(prior_rows),
            "prior_guest_count": evaluated["prior_guest_count"],
            "baseline": evaluated["baseline"],
            "changes": changes,
            "metric_scores": metric_scores,
            "rank_modifier": 0,
            "average_rank_movement": None,
            "composite_score": evaluated["composite_score"],
            "momentum": movement,
            "descriptive_recent_movement": evaluated["descriptive_movement"],
            "week_over_week_changes": week_over_week_changes,
            "week_over_week_movement": week_over_week_movement,
            "previous_week_end": (
                previous_row["week_end"] if previous_row is not None else None
            ),
            "long_term_direction": long_term_direction,
            "long_term_history_label": history_label,
            "history_used": history_used,
            "long_term_changes": long_term_changes,
            "long_term_metric_scores": long_term_metric_scores,
            "long_term_rank_modifier": 0,
            "long_term_average_rank_movement": None,
            "long_term_composite_score": long_term_composite_score,
            "performance_level": peer_comparison,
            "peer_composite_score": evaluated["peer_composite_score"],
            "benchmark_values": benchmark_values,
            "benchmark_sources": benchmark_sources,
            "level_statuses": level_statuses,
            "peer_changes": evaluated["peer_changes"],
            "peer_scores": evaluated["peer_scores"],
            "combined_scores": evaluated["combined_scores"],
            "target_values": {
                field: targets.get(location, {}).get(field)
                for field in SERVER_TREND_FIELDS
            },
            "context_metrics": {
                "check_count": current.get("check_count", 0.0),
                "check_count_available": current.get(
                    "check_count_available", False
                ),
                "per_check_available": current.get(
                    "per_check_available",
                    bool(current.get("check_count_available", False))
                    and float(current.get("check_count", 0) or 0) > 0,
                ),
                "sales_per_check": current.get("sales_per_check", 0.0),
                "guests_per_check": current.get("guests_per_check", 0.0),
                "rate_of_sale_by_guest_count": current.get(
                    "rate_of_sale_by_guest_count", 0.0
                ),
                "rate_available": current.get("rate_available", False),
                "average_ticket_time_seconds": current.get(
                    "average_ticket_time_seconds", 0.0
                ),
                "ticket_time_available": current.get(
                    "ticket_time_available", False
                ),
                "ticket_time_weight_basis": current.get(
                    "ticket_time_weight_basis",
                    "Unavailable (Check Count missing or incomplete)",
                ),
            },
            "positive_drivers": positive_drivers,
            "negative_drivers": negative_drivers,
            "priority": priority,
            "action": action,
            "confidence": evidence_status,
            "prominent": prominent,
            "why": " | ".join(why_parts) if why_parts else "No material movement",
            "candidate_polarity": current_candidate.polarity.value,
            "persistence_reason": persistence.reason,
            "recurring_driver_fields": list(persistence.recurring_drivers),
            "stability_result": (
                "Store-shock guard not passed; day-removal stability not applicable"
                if evaluated["candidate_qualified"]
                and not evaluated["store_shock"].guard_passed
                else
                "Stable under every active-day removal"
                if evaluated["day_stable"]
                else "Sensitive to at least one active-day removal"
                if evaluated["candidate_qualified"]
                else "Not applicable"
            ),
            "peer_cohort_size": min(
                reference.distinct_peer_count
                for reference in evaluated["peer_references"].values()
            ),
            "peer_cohort_weeks": min(
                len(reference.usable_weeks)
                for reference in evaluated["peer_references"].values()
            ),
            "peer_server_weeks": min(
                reference.peer_week_count
                for reference in evaluated["peer_references"].values()
            ),
            "threshold_version": config.get(
                "management_threshold_calibration", {}
            ).get("version", MANAGEMENT_METHODOLOGY_VERSION),
            "evidence_week_ends": [
                week_end.isoformat()
                for week_end in sorted(
                    {
                        latest_week_end,
                        *evaluated["prior_week_ends"],
                        *history_week_ends,
                    }
                )
            ],
            "source_evidence": merged_source_evidence(
                [current, *prior_rows, *server_history_rows]
            ),
        }
        row["recommended_next_step"] = recommended_server_follow_up(row)
        output.append(row)

    priority_order = {"Medium": 0, "Review": 1, "Recognize": 2, "Monitor": 3}
    output.sort(
        key=lambda row: (
            priority_order.get(row["priority"], 9),
            row["composite_score"] if row["priority"] in {"High", "Medium"} else -row["composite_score"],
            row["location"],
            row["display_name"],
        )
    )
    return output


def management_entity_rows(
    weekly_rows: list[dict[str, Any]],
    entity_field: str,
    targets: dict[str, dict[str, float | None]],
    config: dict[str, Any],
    allowed_baseline_weeks: dict[str, set[date]] | set[date],
) -> list[dict[str, Any]]:
    if not weekly_rows:
        return []
    baseline_limit = int(config.get("dashboard_baseline_full_weeks", 4))
    minimum_baseline_weeks = int(
        config.get("management_min_entity_baseline_weeks", 2)
    )
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in weekly_rows:
        grouped[row[entity_field]].append(row)
    output: list[dict[str, Any]] = []
    for entity, rows in grouped.items():
        rows.sort(key=lambda row: row["week_end"])
        latest = rows[-1]
        prior = rows[-2] if len(rows) > 1 else None
        allowed = allowed_baseline_weeks.get(entity, set()) if isinstance(allowed_baseline_weeks, dict) else allowed_baseline_weeks
        prior_full_ends = sorted(
            week_end for week_end in allowed if week_end < latest["week_end"]
        )[-baseline_limit:]
        baseline_rows = [row for row in rows if row["week_end"] in prior_full_ends]
        baseline = aggregate_weekly_rows(baseline_rows)
        if baseline and baseline_rows:
            baseline["gross_sales"] /= len(baseline_rows)
            baseline["guest_count"] /= len(baseline_rows)
            baseline["wine_sales"] /= len(baseline_rows)
        benchmark_values: dict[str, float | None] = {}
        benchmark_sources: dict[str, str] = {}
        for field, _, _ in MANAGEMENT_METRICS:
            target_value = targets.get(entity, {}).get(field)
            baseline_value = (
                baseline.get(field)
                if baseline
                and (
                    field not in SERVER_CONTEXT_FIELDS
                    or management_metric_available(baseline, field)
                )
                else None
            )
            benchmark_values[field] = (
                target_value if target_value is not None else baseline_value
            )
            benchmark_sources[field] = (
                "Target"
                if target_value is not None
                else f"{len(baseline_rows)}-week baseline"
                if baseline_value is not None
                else "Reference unavailable"
            )

        prior_changes = {
            field: (
                latest[field] - prior[field]
                if prior
                and management_metric_available(latest, field)
                and management_metric_available(prior, field)
                else None
            )
            for field, _, _ in MANAGEMENT_METRICS
        }
        benchmark_changes = {
            field: (
                latest[field] - benchmark_values[field]
                if benchmark_values[field] is not None
                and management_metric_available(latest, field)
                and (
                    field not in SERVER_CONTEXT_FIELDS
                    or (baseline is not None and management_metric_available(baseline, field))
                    or targets.get(entity, {}).get(field) is not None
                )
                else None
            )
            for field, _, _ in MANAGEMENT_METRICS
        }
        materiality = config.get("management_materiality", DEFAULT_CONFIG["management_materiality"])
        sales_pct = safe_pct_delta(latest["gross_sales"], benchmark_values["gross_sales"])
        guest_pct = safe_pct_delta(latest["guest_count"], benchmark_values["guest_count"])
        check_change = benchmark_changes["check_average"]
        wine_change = benchmark_changes["wine_pct"]
        negative_count = sum(
            [
                sales_pct is not None and sales_pct <= -float(materiality["sales_pct"]),
                guest_pct is not None and guest_pct <= -float(materiality["guest_pct"]),
                check_change is not None and check_change <= -float(materiality["check_average"]),
                wine_change is not None and wine_change <= -float(materiality["wine_pct"]),
            ]
        )
        if len(baseline_rows) < minimum_baseline_weeks:
            priority = "Monitor"
            status = "Limited History"
            focus = (
                f"Wait for at least {minimum_baseline_weeks} complete baseline weeks "
                "before opening an operational review."
            )
        elif sales_pct is not None and guest_pct is not None and sales_pct <= -float(materiality["sales_pct"]) and guest_pct <= -float(materiality["guest_pct"]):
            priority = "High"
            status = "Traffic Watch"
            focus = (
                "What changed in reservations, events, operating hours, staffing, "
                "or local demand?"
            )
        elif check_change is not None and wine_change is not None and check_change <= -float(materiality["check_average"]) and wine_change <= -float(materiality["wine_pct"]):
            priority = "Medium"
            status = "Upsell Watch"
            focus = (
                "Did guest mix, menu availability, training, or shift mix explain "
                "the Sales / Guest and Wine % change?"
            )
        elif negative_count >= 2:
            priority = "Medium"
            status = "Mixed Watch"
            focus = (
                "Which validated movement changed, what operating context explains it, "
                "and what one follow-up should be assigned?"
            )
        else:
            priority = "Monitor"
            status = "On Track / Mixed"
            focus = "What, if anything, needs follow-up before the next weekly review?"
        output.append(
            {
                "entity": entity,
                "latest": latest,
                "prior": prior,
                "baseline": baseline,
                "baseline_weeks": len(baseline_rows),
                "benchmark_values": benchmark_values,
                "benchmark_sources": benchmark_sources,
                "prior_changes": prior_changes,
                "benchmark_changes": benchmark_changes,
                "priority": priority,
                "status": status,
                "recommended_focus": focus,
                "evidence_week_ends": [
                    week_end.isoformat()
                    for week_end in sorted(
                        {latest["week_end"], *prior_full_ends}
                    )
                ],
                "source_evidence": merged_source_evidence(
                    [latest, *baseline_rows]
                ),
            }
        )
    output.sort(key=lambda row: (0 if row["entity"] == "All Stores" else 1, row["entity"]))
    return output


def find_sheet_header_row(ws, required_header: str) -> int | None:
    for row in range(1, min(ws.max_row, 12) + 1):
        if any(ws.cell(row=row, column=col).value == required_header for col in range(1, ws.max_column + 1)):
            return row
    return None


def records_from_sheet(ws, required_header: str) -> list[dict[str, Any]]:
    header_row = find_sheet_header_row(ws, required_header)
    if header_row is None:
        return []
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
    records: list[dict[str, Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        record = {
            str(header): excel_safe_cell_value(ws.cell(row=row, column=col))
            for col, header in enumerate(headers, start=1)
            if header
        }
        if record.get(required_header):
            records.append(record)
    return records


def records_from_table(ws, table_name: str) -> list[dict[str, Any]]:
    """Read records from one exact table without relying on its sheet position."""

    if table_name not in ws.tables:
        return []
    min_col, min_row, max_col, max_row = range_boundaries(
        ws.tables[table_name].ref
    )
    headers = [
        ws.cell(row=min_row, column=column).value
        for column in range(min_col, max_col + 1)
    ]
    records: list[dict[str, Any]] = []
    for row in range(min_row + 1, max_row + 1):
        record = {
            str(header): excel_safe_cell_value(ws.cell(row=row, column=column))
            for header, column in zip(
                headers, range(min_col, max_col + 1), strict=True
            )
            if header
        }
        if any(not is_blank(value) for value in record.values()):
            records.append(record)
    return records


def validate_action_board_records(
    records: list[dict[str, Any]],
    *,
    allowed_reviewers: Iterable[str] | None = None,
) -> list[dict[str, Any]]:
    """Normalize legacy headers and reject pasted workflow values that bypass Excel."""
    canonical_dispositions = {
        disposition.casefold(): disposition
        for disposition in REVIEW_DISPOSITION_CHOICES
    }
    allowed_reviewer_keys = (
        {
            str(reviewer).strip().casefold()
            for reviewer in allowed_reviewers
            if str(reviewer).strip()
        }
        if allowed_reviewers is not None
        else None
    )
    for row_number, record in enumerate(records, start=5):
        legacy_schema = "Manager Notes" in record and "Context Notes" not in record
        status_choices = (
            LEGACY_ACTION_STATUS_CHOICES if legacy_schema else ACTION_STATUS_CHOICES
        )
        canonical_statuses = {
            status.casefold(): status for status in status_choices
        }
        allowed_text = ", ".join(status_choices)
        for old, new in (
            ("Manager Notes", "Context Notes"),
            ("Performance Level", "Peer Comparison"),
            ("Momentum", "Recent Movement"),
            ("Confidence", "Evidence Status"),
        ):
            if is_blank(record.get(new)) and not is_blank(record.get(old)):
                record[new] = record.get(old)
        raw_status = record.get("Status")
        status_text = "" if is_blank(raw_status) else str(raw_status).strip()
        canonical_status = canonical_statuses.get(status_text.casefold())
        if canonical_status is None:
            action_id = str(record.get("Action ID") or f"row {row_number}")
            raise ValueError(
                f"Action Board Status values must be one of {allowed_text}; "
                f"received {raw_status!r} for action {action_id!r}. "
                "No workbooks were created and no source files were moved."
            )
        record["Status"] = canonical_status
        raw_disposition = record.get("Review Disposition")
        disposition_text = (
            "Pending Review"
            if is_blank(raw_disposition)
            else str(raw_disposition).strip()
        )
        canonical_disposition = canonical_dispositions.get(
            disposition_text.casefold()
        )
        if canonical_disposition is None:
            action_id = str(record.get("Action ID") or f"row {row_number}")
            allowed = ", ".join(REVIEW_DISPOSITION_CHOICES)
            raise ValueError(
                f"Action Board Review Disposition values must be one of {allowed}; "
                f"received {raw_disposition!r} for action {action_id!r}. "
                "No workbooks were created and no source files were moved."
            )
        record["Review Disposition"] = canonical_disposition
        if legacy_schema:
            record.setdefault("Reviewed By", "")
            record.setdefault("Review Date", None)
            continue
        review_completed = canonical_disposition != "Pending Review"
        status_requires_review = canonical_status.casefold() not in {
            "review needed",
            "open",
        }
        if review_completed or status_requires_review:
            reviewer = str(record.get("Reviewed By") or "").strip()
            review_date = as_date(record.get("Review Date"))
            if canonical_disposition == "Pending Review" or not reviewer or review_date is None:
                action_id = str(record.get("Action ID") or f"row {row_number}")
                raise ValueError(
                    f"Action {action_id!r} cannot move to {canonical_status} until "
                    "Review Disposition, Reviewed By, and Review Date are complete. "
                    "No workbooks were created and no source files were moved."
                )
            if (
                allowed_reviewer_keys is not None
                and reviewer.casefold() not in allowed_reviewer_keys
            ):
                action_id = str(record.get("Action ID") or f"row {row_number}")
                raise ValueError(
                    f"Action {action_id!r} Reviewed By must name an active person "
                    "from the Owner Roster. No workbooks were created and no "
                    "source files were moved."
                )
            record["Reviewed By"] = excel_safe_text(reviewer)
            record["Review Date"] = review_date
    return records


def owner_active_value(value: Any) -> bool:
    """Normalize the user-facing owner status without silently accepting typos."""
    if is_blank(value):
        return True
    if isinstance(value, bool):
        return value
    text = str(value).strip().casefold()
    if text in {"yes", "active"}:
        return True
    if text in {"no", "inactive"}:
        return False
    raise ValueError(
        f"Owner Roster Active values must be Yes or No; received {value!r}. "
        "No workbooks were created and no source files were moved."
    )


def normalize_owner_roster(entries: Iterable[Any]) -> list[dict[str, str]]:
    """Return a validated, display-safe owner roster in stable workbook order."""
    roster: list[dict[str, str]] = []
    seen: dict[str, str] = {}
    for entry in entries:
        if isinstance(entry, dict):
            raw_name = entry.get("Owner Name", entry.get("Name"))
            raw_active = entry.get("Active")
        else:
            raw_name = entry
            raw_active = True
        if is_blank(raw_name):
            continue
        name = str(excel_safe_text(str(raw_name).strip()))
        duplicate_key = name.casefold()
        if duplicate_key in seen:
            raise ValueError(
                f"Owner Roster contains the duplicate name {name!r}. "
                "Keep one row per owner; no workbooks were created and no source files were moved."
            )
        seen[duplicate_key] = name
        roster.append(
            {
                "Owner Name": name,
                "Active": "Yes" if owner_active_value(raw_active) else "No",
            }
        )
        if len(roster) > OWNER_ROSTER_MAX_ROWS:
            raise ValueError(
                f"Owner Roster supports at most {OWNER_ROSTER_MAX_ROWS} people. "
                "No workbooks were created and no source files were moved."
            )
    return roster


def active_owner_names(roster: Iterable[Any]) -> list[str]:
    return [
        row["Owner Name"]
        for row in normalize_owner_roster(roster)
        if row["Active"] == "Yes"
    ]


def owner_roster_from_sheet(ws) -> list[dict[str, str]]:
    """Read the current roster table, falling back to the legacy J6:J25 list."""
    entries: list[dict[str, Any]] = []
    if OWNER_ROSTER_TABLE_NAME in ws.tables:
        table = ws.tables[OWNER_ROSTER_TABLE_NAME]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        supported_anchor = (
            (min_col, min_row, max_col) == (1, 20, 2)
            or (
                ws.title == MANAGEMENT_CENTER_SHEET
                and (min_col, min_row, max_col) == (11, 12, 12)
            )
        )
        if not supported_anchor:
            raise ValueError(
                "Owner Roster must remain at its supported two-column anchor. "
                "No workbooks were created and no source files were moved."
            )
        if max_row - min_row > OWNER_ROSTER_MAX_ROWS:
            raise ValueError(
                f"Owner Roster supports at most {OWNER_ROSTER_MAX_ROWS} rows. "
                "No workbooks were created and no source files were moved."
            )
        actual_headers = tuple(
            str(ws.cell(row=min_row, column=col).value or "")
            for col in range(min_col, max_col + 1)
        )
        if actual_headers != OWNER_ROSTER_HEADERS:
            raise ValueError(
                "Owner Roster headers must be Owner Name and Active in that order. "
                "No workbooks were created and no source files were moved."
            )
        headers = {
            str(ws.cell(row=min_row, column=col).value): col
            for col in range(min_col, max_col + 1)
            if ws.cell(row=min_row, column=col).value
        }
        name_col = headers.get("Owner Name")
        active_col = headers.get("Active")
        if name_col is None or active_col is None:
            raise ValueError(
                "Owner Roster must contain Owner Name and Active columns. "
                "No workbooks were created and no source files were moved."
            )
        for row in range(min_row + 1, max_row + 1):
            name = excel_safe_cell_value(ws.cell(row=row, column=name_col))
            if is_blank(name):
                continue
            active_value = excel_safe_cell_value(ws.cell(row=row, column=active_col))
            if is_blank(active_value):
                raise ValueError(
                    f"Owner Roster row {row} has a name but no Active value. Choose Yes or No; "
                    "no workbooks were created and no source files were moved."
                )
            entries.append(
                {
                    "Owner Name": name,
                    "Active": active_value,
                }
            )
    else:
        for row in range(6, min(ws.max_row, 25) + 1):
            value = excel_safe_cell_value(ws.cell(row=row, column=10))
            if not is_blank(value):
                entries.append({"Owner Name": value, "Active": "Yes"})
    return normalize_owner_roster(entries)


def owner_roster_capacity_from_sheet(ws) -> int:
    if OWNER_ROSTER_TABLE_NAME not in ws.tables:
        return 20
    _, min_row, _, max_row = range_boundaries(ws.tables[OWNER_ROSTER_TABLE_NAME].ref)
    capacity = max(1, max_row - min_row)
    if capacity > OWNER_ROSTER_MAX_ROWS:
        raise ValueError(
            f"Owner Roster supports at most {OWNER_ROSTER_MAX_ROWS} rows. "
            "No workbooks were created and no source files were moved."
        )
    return capacity


def read_management_state(
    output_path: Path,
    *,
    allow_legacy_protection_upgrade: bool = False,
    expected_digest: str | None = None,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
) -> dict[str, Any]:
    state: dict[str, Any] = {
        "targets": {},
        "owners": [],
        "owner_roster": [],
        "owner_roster_capacity": OWNER_ROSTER_MIN_EDIT_ROWS,
        "active_actions": [],
        "action_history": [],
        "evidence_by_action_id": {},
    }
    if not output_path.exists():
        return state
    try:
        wb = load_workbook(output_path, data_only=False)
    except Exception as exc:
        raise RuntimeError(
            f"Could not read the existing master workbook at {output_path}. "
            "Close the workbook in Excel and rerun; no source files were moved."
        ) from exc
    try:
        verify_existing_management_workbook_integrity(
            output_path,
            expected_digest=expected_digest,
            expected_digest_scheme=expected_digest_scheme,
            legacy_reference_path=legacy_reference_path,
            allow_legacy_protection_upgrade=allow_legacy_protection_upgrade,
        )
    except Exception:
        wb.close()
        raise
    try:
        setup_sheet_name = (
            MANAGEMENT_CENTER_SHEET
            if MANAGEMENT_CENTER_SHEET in wb.sheetnames
            and "ManagementTargets" in wb[MANAGEMENT_CENTER_SHEET].tables
            and OWNER_ROSTER_TABLE_NAME in wb[MANAGEMENT_CENTER_SHEET].tables
            else "Management Setup"
            if "Management Setup" in wb.sheetnames
            else None
        )
        if setup_sheet_name is not None:
            ws = wb[setup_sheet_name]
            target_rows = records_from_table(ws, "ManagementTargets")
            if not target_rows:
                target_header_row = find_sheet_header_row(ws, "Entity")
                if target_header_row:
                    headers = {
                        ws.cell(row=target_header_row, column=col).value: col
                        for col in range(1, ws.max_column + 1)
                        if ws.cell(row=target_header_row, column=col).value
                    }
                    row = target_header_row + 1
                    while row <= ws.max_row:
                        entity = ws.cell(row=row, column=headers["Entity"]).value
                        if not entity:
                            break
                        target_rows.append(
                            {
                                str(header): excel_safe_cell_value(
                                    ws.cell(row=row, column=column)
                                )
                                for header, column in headers.items()
                            }
                        )
                        row += 1
            for target_row in target_rows:
                entity = target_row.get("Entity")
                if is_blank(entity):
                    continue
                values: dict[str, float | None] = {}
                for field, label in TARGET_FIELDS:
                    value = target_row.get(label)
                    if isinstance(value, (int, float)):
                        values[field] = (
                            float(value) * 60
                            if field == "average_ticket_time_seconds"
                            else float(value)
                        )
                    else:
                        values[field] = None
                state["targets"][str(entity)] = values
            state["owner_roster"] = owner_roster_from_sheet(ws)
            state["owners"] = active_owner_names(state["owner_roster"])
            state["owner_roster_capacity"] = owner_roster_capacity_from_sheet(ws)
        if (
            MANAGEMENT_CENTER_SHEET in wb.sheetnames
            and "ActionBoardTable" in wb[MANAGEMENT_CENTER_SHEET].tables
        ):
            state["active_actions"] = validate_action_board_records(
                records_from_table(wb[MANAGEMENT_CENTER_SHEET], "ActionBoardTable"),
                allowed_reviewers=state["owners"],
            )
        elif "Action Board" in wb.sheetnames:
            state["active_actions"] = validate_action_board_records(
                records_from_sheet(wb["Action Board"], "Action ID"),
                allowed_reviewers=state["owners"],
            )
        if (
            MANAGEMENT_CENTER_SHEET in wb.sheetnames
            and "ActionHistoryTable" in wb[MANAGEMENT_CENTER_SHEET].tables
        ):
            state["action_history"] = validate_action_board_records(
                records_from_table(wb[MANAGEMENT_CENTER_SHEET], "ActionHistoryTable")
            )
        elif "Action History" in wb.sheetnames:
            state["action_history"] = validate_action_board_records(
                records_from_sheet(wb["Action History"], "Action ID")
            )
        if "Evidence Detail" in wb.sheetnames:
            evidence_rows = records_from_sheet(wb["Evidence Detail"], "Evidence ID")
            state["evidence_by_action_id"] = {
                str(row.get("Action ID")): row
                for row in evidence_rows
                if row.get("Action ID")
            }
    finally:
        wb.close()
    return state


def workbook_digest_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return {"type": type(value).__name__, "value": value.isoformat()}
    if isinstance(value, float) and not math.isfinite(value):
        raise IntegrityError("Workbook integrity cannot hash NaN or infinite values.")
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, ArrayFormula):
        return {
            "type": "ArrayFormula",
            "ref": value.ref,
            "text": value.text,
        }
    if isinstance(value, DataTableFormula):
        return {
            "type": "DataTableFormula",
            **{
                key: item
                for key, item in sorted(value.__dict__.items())
            },
        }
    formula_text = getattr(value, "text", None)
    if isinstance(formula_text, str):
        return {"type": type(value).__name__, "value": formula_text}
    return {"type": type(value).__name__, "value": str(value)}


def workbook_digest_xml_element(node: Any) -> dict[str, Any]:
    """Convert an openpyxl XML element into canonical JSON-safe content."""
    return {
        "tag": str(node.tag),
        "attributes": {
            str(key): str(value)
            for key, value in sorted(node.attrib.items(), key=lambda item: str(item[0]))
        },
        "text": node.text,
        "children": [workbook_digest_xml_element(child) for child in list(node)],
    }


def workbook_digest_serialisable(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    to_tree = getattr(value, "to_tree", None)
    if not callable(to_tree):
        raise IntegrityError(
            f"Workbook integrity cannot serialize {type(value).__name__}."
        )
    try:
        return workbook_digest_xml_element(to_tree())
    except Exception as exc:
        raise IntegrityError(
            f"Workbook integrity could not serialize {type(value).__name__}: {exc}"
        ) from exc


def workbook_digest_style_payload(styleable: Any) -> dict[str, Any]:
    return {
        "font": workbook_digest_serialisable(styleable.font),
        "fill": workbook_digest_serialisable(styleable.fill),
        "border": workbook_digest_serialisable(styleable.border),
        "alignment": workbook_digest_serialisable(styleable.alignment),
        "number_format": styleable.number_format,
        "protection": workbook_digest_serialisable(styleable.protection),
    }


def workbook_digest_hyperlink_payload(
    hyperlink: Any,
    *,
    sheet_name: str,
    coordinate: str,
) -> dict[str, Any] | None:
    if hyperlink is None:
        return None
    target = hyperlink.target
    if target and not str(target).strip().startswith("#"):
        raise IntegrityError(
            "External hyperlinks are not allowed in the protected master workbook: "
            f"{sheet_name}!{coordinate}."
        )
    return {
        "target": target,
        "location": hyperlink.location,
        "tooltip": hyperlink.tooltip,
        "display": hyperlink.display,
    }


def workbook_digest_comment_payload(comment: Any) -> dict[str, Any] | None:
    if comment is None:
        return None
    return {
        "text": comment.text,
        "author": comment.author,
        "height": comment.height,
        "width": comment.width,
    }


def workbook_digest_editable_scalar(value: Any, *, sheet_name: str, coordinate: str) -> None:
    if isinstance(value, float) and not math.isfinite(value):
        raise IntegrityError(
            f"Editable workbook value is not finite at {sheet_name}!{coordinate}."
        )
    if value is None or isinstance(
        value,
        (str, int, float, bool, datetime, date, time, timedelta),
    ):
        return
    raise IntegrityError(
        "Only scalar values are allowed in editable workbook cells: "
        f"{sheet_name}!{coordinate}."
    )


def reject_external_formula_reference(
    value: Any,
    *,
    data_type: str,
    sheet_name: str,
    coordinate: str,
) -> None:
    if data_type != "f":
        return
    if isinstance(value, ArrayFormula):
        formula_fields = (value.ref, value.text)
    elif isinstance(value, DataTableFormula):
        formula_fields = tuple(value.__dict__.values())
    elif isinstance(value, str):
        formula_fields = (value,)
    else:
        raise IntegrityError(
            "Unsupported formula object in protected master workbook: "
            f"{sheet_name}!{coordinate} ({type(value).__name__})."
        )
    formula = "\n".join(str(item) for item in formula_fields if item is not None)
    if (
        re.search(r"\[(?:\d+|[^\]]+\.(?:xlsx?|xlsm|xlsb|csv))\]", formula, re.I)
        or re.search(r"(?:https?|file)://|\\\\|[A-Za-z]:\\", formula, re.I)
    ):
        raise IntegrityError(
            "External formula references are not allowed in the protected master "
            f"workbook: {sheet_name}!{coordinate}."
        )


def workbook_digest_dimension_payload(dimension: Any) -> dict[str, Any]:
    payload = {
        "hidden": bool(dimension.hidden),
        "outline_level": dimension.outlineLevel,
        "collapsed": bool(dimension.collapsed),
        "style": (
            workbook_digest_style_payload(dimension)
            if dimension.has_style
            else None
        ),
    }
    for attribute in (
        "min", "max", "width", "bestFit", "height", "thickTop", "thickBot",
    ):
        if hasattr(dimension, attribute):
            payload[attribute] = getattr(dimension, attribute)
    return payload


def workbook_digest_sheet_view_payload(ws: Any) -> dict[str, Any]:
    view = ws.sheet_view
    fields = (
        "windowProtection", "showFormulas", "showGridLines", "showRowColHeaders",
        "showZeros", "rightToLeft", "showRuler", "showOutlineSymbols",
        "defaultGridColor", "showWhiteSpace", "view", "colorId", "zoomScale",
        "zoomScaleNormal", "zoomScaleSheetLayoutView", "zoomScalePageLayoutView",
        "zoomToFit",
    )
    freeze_panes = ws.freeze_panes
    if hasattr(freeze_panes, "coordinate"):
        freeze_panes = freeze_panes.coordinate
    pane = view.pane
    return {
        "settings": {field: getattr(view, field) for field in fields},
        "freeze_panes": freeze_panes,
        "pane": (
            {
                "x_split": pane.xSplit,
                "y_split": pane.ySplit,
                "top_left_cell": pane.topLeftCell,
                "active_pane": pane.activePane,
                "state": pane.state,
            }
            if pane is not None
            else None
        ),
    }


def workbook_digest_chart_payload(chart: Any) -> dict[str, Any]:
    plots = getattr(chart, "_charts", [chart])
    axes = getattr(chart, "_axes", {})
    return {
        "type": type(chart).__name__,
        "plots": [
            {
                "type": type(plot).__name__,
                "definition": workbook_digest_serialisable(plot),
            }
            for plot in plots
        ],
        "title": workbook_digest_serialisable(getattr(chart, "title", None)),
        "legend": workbook_digest_serialisable(getattr(chart, "legend", None)),
        "layout": workbook_digest_serialisable(getattr(chart, "layout", None)),
        "graphical_properties": workbook_digest_serialisable(
            getattr(chart, "graphical_properties", None)
        ),
        "axes": [
            {
                "id": axis_id,
                "definition": workbook_digest_serialisable(axis),
            }
            for axis_id, axis in axes.items()
        ],
        "style": getattr(chart, "style", None),
        "rounded_corners": getattr(chart, "roundedCorners", None),
        "display_blanks": getattr(chart, "display_blanks", None),
        "visible_cells_only": getattr(chart, "visible_cells_only", None),
        "index_base": getattr(chart, "idx_base", None),
        "pivot_source": workbook_digest_serialisable(
            getattr(chart, "pivotSource", None)
        ),
        "pivot_formats": [
            workbook_digest_serialisable(pivot_format)
            for pivot_format in (getattr(chart, "pivotFormats", None) or [])
        ],
        "anchor": (
            workbook_digest_serialisable(chart.anchor)
            if hasattr(chart.anchor, "to_tree")
            else str(chart.anchor)
        ),
    }


def reject_unapproved_workbook_drawings(path: Path) -> None:
    """Allow generated charts while rejecting image, shape, and embedded-object overlays."""
    from xml.etree import ElementTree
    from zipfile import BadZipFile, ZipFile

    spreadsheet_drawing_ns = (
        "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    )
    chart_ns = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    office_relationship_ns = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    package_relationship_ns = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    anchor_tags = {"oneCellAnchor", "twoCellAnchor", "absoluteAnchor"}
    chart_anchor_children = {
        "from", "to", "pos", "ext", "graphicFrame", "clientData",
    }

    def local_name(tag: str) -> str:
        return tag.rsplit("}", 1)[-1]

    try:
        with ZipFile(path) as archive:
            part_names = set(archive.namelist())
            external_link_parts = sorted(
                name
                for name in part_names
                if name.lower().startswith("xl/externallinks/")
            )
            if external_link_parts:
                raise IntegrityError(
                    "External workbook links are not allowed in the protected master "
                    f"workbook ({external_link_parts[0]})."
                )
            unsupported_active_parts = sorted(
                name
                for name in part_names
                if name.lower() == "xl/connections.xml"
                or name.lower().startswith("xl/querytables/")
                or name.lower() == "xl/vbaproject.bin"
                or name.lower().startswith("xl/macrosheets/")
                or name.lower().startswith("xl/dialogsheets/")
            )
            if unsupported_active_parts:
                raise IntegrityError(
                    "Connections, queries, macros, and other active workbook parts are not "
                    "allowed in the protected master workbook "
                    f"({unsupported_active_parts[0]})."
                )
            for relationship_part in sorted(
                name
                for name in part_names
                if name.lower().endswith(".rels")
            ):
                relationships_root = ElementTree.fromstring(
                    archive.read(relationship_part)
                )
                for relationship in list(relationships_root):
                    if relationship.attrib.get("TargetMode", "").casefold() != "external":
                        continue
                    relationship_type = relationship.attrib.get("Type", "").casefold()
                    target = relationship.attrib.get("Target", "").strip()
                    # openpyxl historically serializes an in-workbook fragment as an
                    # external-mode hyperlink relationship. It is still local when the
                    # complete target is only a #fragment.
                    if relationship_type.endswith("/hyperlink") and target.startswith("#"):
                        continue
                    if relationship_type.endswith("/hyperlink"):
                        raise IntegrityError(
                            "External hyperlinks are not allowed in the protected master "
                            f"workbook ({relationship_part})."
                        )
                    raise IntegrityError(
                        "External relationships are not allowed in the protected master "
                        f"workbook ({relationship_part})."
                    )
            prohibited_binary_parts = sorted(
                name
                for name in part_names
                if name.lower().startswith(
                    ("xl/media/", "xl/embeddings/", "xl/activex/", "xl/ctrlprops/")
                )
            )
            if prohibited_binary_parts:
                raise IntegrityError(
                    "Images and embedded objects are not allowed in the protected master "
                    f"workbook ({prohibited_binary_parts[0]})."
                )

            unsupported_drawing_parts = sorted(
                name
                for name in part_names
                if name.lower().startswith("xl/drawings/")
                and "/_rels/" not in name.lower()
                and not name.lower().endswith(".xml")
            )
            if unsupported_drawing_parts:
                raise IntegrityError(
                    "Legacy VML and other non-chart drawings are not allowed in the "
                    f"protected master workbook ({unsupported_drawing_parts[0]})."
                )

            drawing_parts = sorted(
                name
                for name in part_names
                if name.lower().startswith("xl/drawings/")
                and "/_rels/" not in name.lower()
                and name.lower().endswith(".xml")
            )
            for drawing_part in drawing_parts:
                root = ElementTree.fromstring(archive.read(drawing_part))
                if root.tag != f"{{{spreadsheet_drawing_ns}}}wsDr":
                    raise IntegrityError(
                        "An unsupported drawing part was found in the protected master "
                        f"workbook ({drawing_part})."
                    )

                relationship_part = (
                    "xl/drawings/_rels/"
                    f"{drawing_part.rsplit('/', 1)[-1]}.rels"
                )
                if relationship_part not in part_names:
                    raise IntegrityError(
                        "A drawing relationship file is missing from the protected master "
                        f"workbook ({drawing_part})."
                    )
                relationships_root = ElementTree.fromstring(
                    archive.read(relationship_part)
                )
                relationships: dict[str, Any] = {}
                for relationship in list(relationships_root):
                    if relationship.tag != (
                        f"{{{package_relationship_ns}}}Relationship"
                    ):
                        raise IntegrityError(
                            "An unsupported drawing relationship was found in the protected "
                            f"master workbook ({relationship_part})."
                        )
                    relationship_id = relationship.attrib.get("Id")
                    if not relationship_id:
                        raise IntegrityError(
                            "A drawing relationship has no identifier in the protected master "
                            f"workbook ({relationship_part})."
                        )
                    relationships[relationship_id] = relationship.attrib

                used_relationships: set[str] = set()
                for anchor in list(root):
                    if local_name(anchor.tag) not in anchor_tags:
                        raise IntegrityError(
                            "Only generated chart drawings are allowed in the protected "
                            f"master workbook ({drawing_part})."
                        )
                    anchor_children = list(anchor)
                    if any(
                        local_name(child.tag) not in chart_anchor_children
                        for child in anchor_children
                    ):
                        raise IntegrityError(
                            "Images and non-chart drawings are not allowed in the protected "
                            f"master workbook ({drawing_part})."
                        )
                    frames = [
                        child
                        for child in anchor_children
                        if local_name(child.tag) == "graphicFrame"
                    ]
                    if len(frames) != 1:
                        raise IntegrityError(
                            "Images and non-chart drawings are not allowed in the protected "
                            f"master workbook ({drawing_part})."
                        )
                    chart_references = [
                        node
                        for node in frames[0].iter()
                        if node.tag == f"{{{chart_ns}}}chart"
                    ]
                    if len(chart_references) != 1:
                        raise IntegrityError(
                            "Only generated chart drawings are allowed in the protected "
                            f"master workbook ({drawing_part})."
                        )
                    relationship_id = chart_references[0].attrib.get(
                        f"{{{office_relationship_ns}}}id"
                    )
                    relationship = relationships.get(str(relationship_id))
                    if (
                        not relationship_id
                        or relationship is None
                        or not relationship.get("Type", "").endswith("/chart")
                        or relationship.get("TargetMode") == "External"
                    ):
                        raise IntegrityError(
                            "A chart drawing has an unsafe relationship in the protected "
                            f"master workbook ({drawing_part})."
                        )
                    used_relationships.add(str(relationship_id))

                if used_relationships != set(relationships):
                    raise IntegrityError(
                        "A non-chart drawing relationship was found in the protected master "
                        f"workbook ({relationship_part})."
                    )
    except IntegrityError:
        raise
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError) as exc:
        raise IntegrityError(
            f"Could not inspect workbook drawing integrity at {path}: {exc}"
        ) from exc


def workbook_digest_excluded_cells(wb: Workbook) -> set[tuple[str, str]]:
    """Exclude only the self-referential digest value, never a user input."""
    excluded: set[tuple[str, str]] = set()
    if "Run Notes" not in wb.sheetnames:
        return excluded
    ws = wb["Run Notes"]
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == RUN_NOTES_DIGEST_LABEL:
            excluded.add((ws.title, ws.cell(row=row, column=2).coordinate))
    return excluded


def approved_management_input_cells(wb: Workbook) -> set[tuple[str, str]]:
    approved: set[tuple[str, str]] = set()
    setup_sheet_name = (
        MANAGEMENT_CENTER_SHEET
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames
        and "ManagementTargets" in wb[MANAGEMENT_CENTER_SHEET].tables
        and OWNER_ROSTER_TABLE_NAME in wb[MANAGEMENT_CENTER_SHEET].tables
        else "Management Setup"
        if "Management Setup" in wb.sheetnames
        else None
    )
    if setup_sheet_name is not None:
        ws = wb[setup_sheet_name]
        if "ManagementTargets" in ws.tables:
            min_col, min_row, max_col, max_row = range_boundaries(
                ws.tables["ManagementTargets"].ref
            )
            headers = {
                ws.cell(row=min_row, column=col).value: col
                for col in range(min_col, max_col + 1)
                if ws.cell(row=min_row, column=col).value
            }
            target_columns = [
                headers[label] for _, label in TARGET_FIELDS if label in headers
            ]
            for row in range(min_row + 1, max_row + 1):
                for col in target_columns:
                    approved.add((ws.title, ws.cell(row=row, column=col).coordinate))
        else:
            target_header_row = find_sheet_header_row(ws, "Entity")
            if target_header_row:
                headers = {
                    ws.cell(row=target_header_row, column=col).value: col
                    for col in range(1, ws.max_column + 1)
                    if ws.cell(row=target_header_row, column=col).value
                }
                target_columns = [
                    headers[label] for _, label in TARGET_FIELDS if label in headers
                ]
                row = target_header_row + 1
                while row <= ws.max_row and not is_blank(
                    ws.cell(row=row, column=headers["Entity"]).value
                ):
                    for col in target_columns:
                        approved.add((ws.title, ws.cell(row=row, column=col).coordinate))
                    row += 1
        if OWNER_ROSTER_TABLE_NAME in ws.tables:
            min_col, min_row, max_col, max_row = range_boundaries(
                ws.tables[OWNER_ROSTER_TABLE_NAME].ref
            )
            for row in range(min_row + 1, max_row + 1):
                for col in range(min_col, max_col + 1):
                    approved.add((ws.title, ws.cell(row=row, column=col).coordinate))
        else:
            for row in range(6, 26):
                approved.add((ws.title, f"J{row}"))
    action_sheet_name = (
        MANAGEMENT_CENTER_SHEET
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames
        and "ActionBoardTable" in wb[MANAGEMENT_CENTER_SHEET].tables
        else "Action Board"
        if "Action Board" in wb.sheetnames
        else None
    )
    if action_sheet_name is not None:
        ws = wb[action_sheet_name]
        if "ActionBoardTable" in ws.tables:
            min_col, header_row, max_col, max_row = range_boundaries(
                ws.tables["ActionBoardTable"].ref
            )
        else:
            header_row = find_sheet_header_row(ws, "Action ID")
            min_col, max_col, max_row = 1, ws.max_column, ws.max_row
        if header_row:
            headers = {
                ws.cell(row=header_row, column=col).value: col
                for col in range(min_col, max_col + 1)
                if ws.cell(row=header_row, column=col).value
            }
            for header in (
                "Status",
                "Owner",
                "Due Date",
                "Context Notes",
                "Manager Notes",
                "Review Disposition",
                "Reviewed By",
                "Review Date",
            ):
                col = headers.get(header)
                if col is None:
                    continue
                for row in range(header_row + 1, max_row + 1):
                    approved.add((ws.title, ws.cell(row=row, column=col).coordinate))
    return approved


def workbook_metadata_payload_v2(path: Path) -> dict[str, Any]:
    """Reproduce the exact metadata-rich v2 digest contract for old manifests."""
    reject_unapproved_workbook_drawings(path)
    try:
        wb = load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        raise IntegrityError(f"Could not inspect master workbook integrity at {path}: {exc}") from exc
    try:
        approved = approved_management_input_cells(wb)
        digest_cells = workbook_digest_excluded_cells(wb)
        sheets: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            if ws._images:
                raise IntegrityError(
                    "Images and non-chart drawings are not allowed in the protected master "
                    f"workbook ({ws.title!r})."
                )
            cells: list[dict[str, Any]] = []
            for cell in sorted(ws._cells.values(), key=lambda item: (item.row, item.column)):
                key = (ws.title, cell.coordinate)
                is_approved = key in approved
                is_digest_cell = key in digest_cells
                hyperlink = workbook_digest_hyperlink_payload(
                    cell.hyperlink,
                    sheet_name=ws.title,
                    coordinate=cell.coordinate,
                )
                comment = workbook_digest_comment_payload(cell.comment)
                if (
                    cell.value is None
                    and not cell.has_style
                    and hyperlink is None
                    and comment is None
                    and not is_approved
                    and not is_digest_cell
                ):
                    continue
                if is_approved or is_digest_cell:
                    if cell.data_type == "f":
                        raise IntegrityError(
                            "Formulas are not allowed in editable or digest-excluded cells: "
                            f"{ws.title}!{cell.coordinate}."
                        )
                    workbook_digest_editable_scalar(
                        cell.value,
                        sheet_name=ws.title,
                        coordinate=cell.coordinate,
                    )
                    value_payload: Any = {
                        "excluded": (
                            "approved-management-scalar"
                            if is_approved
                            else "self-referential-digest"
                        )
                    }
                    data_type = "excluded-scalar"
                else:
                    value_payload = workbook_digest_value(cell.value)
                    data_type = cell.data_type
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "data_type": data_type,
                        "value": value_payload,
                        "style_name": getattr(cell, "style", None),
                        "style": workbook_digest_style_payload(cell),
                        "quote_prefix": bool(getattr(cell, "quotePrefix", False)),
                        "pivot_button": bool(getattr(cell, "pivotButton", False)),
                        "hyperlink": hyperlink,
                        "comment": comment,
                    }
                )
            tables: list[dict[str, Any]] = []
            for table in ws.tables.values():
                tables.append(
                    {
                        "name": table.name,
                        "definition": workbook_digest_serialisable(table),
                    }
                )
            validations = [
                workbook_digest_serialisable(validation)
                for validation in ws.data_validations.dataValidation
            ]
            conditional_formats = []
            for conditional_format in ws.conditional_formatting:
                conditional_formats.append(
                    {
                        "definition": workbook_digest_serialisable(conditional_format),
                        "differential_styles": [
                            workbook_digest_serialisable(rule.dxf)
                            for rule in conditional_format.rules
                        ],
                    }
                )
            charts = [workbook_digest_chart_payload(chart) for chart in ws._charts]
            unlocked_cells = sorted(
                cell.coordinate
                for cell in ws._cells.values()
                if cell.protection.locked is False
            )
            sheets.append(
                {
                    "name": ws.title,
                    "state": ws.sheet_state,
                    "cells": cells,
                    "tables": sorted(tables, key=lambda item: item["name"].casefold()),
                    "data_validations": validations,
                    "conditional_formats": conditional_formats,
                    "charts": charts,
                    "merged_cells": sorted(str(cell_range) for cell_range in ws.merged_cells.ranges),
                    "row_dimensions": [
                        {
                            "index": index,
                            **workbook_digest_dimension_payload(dimension),
                        }
                        for index, dimension in sorted(ws.row_dimensions.items())
                    ],
                    "column_dimensions": [
                        {
                            "index": index,
                            **workbook_digest_dimension_payload(dimension),
                        }
                        for index, dimension in sorted(ws.column_dimensions.items())
                    ],
                    "sheet_properties": workbook_digest_serialisable(ws.sheet_properties),
                    "sheet_format": workbook_digest_serialisable(ws.sheet_format),
                    "sheet_view": workbook_digest_sheet_view_payload(ws),
                    "auto_filter": workbook_digest_serialisable(ws.auto_filter),
                    "page_margins": workbook_digest_serialisable(ws.page_margins),
                    "page_setup": workbook_digest_serialisable(ws.page_setup),
                    "print_options": workbook_digest_serialisable(ws.print_options),
                    "header_footer": workbook_digest_serialisable(ws.HeaderFooter),
                    "row_breaks": workbook_digest_serialisable(ws.row_breaks),
                    "column_breaks": workbook_digest_serialisable(ws.col_breaks),
                    "unlocked_cells": unlocked_cells,
                    "protection": workbook_digest_serialisable(ws.protection),
                }
            )
        defined_names = [
            {
                "name": name,
                "definition": workbook_digest_serialisable(defined_name),
            }
            for name, defined_name in sorted(
                wb.defined_names.items(), key=lambda item: item[0].casefold()
            )
        ]
        return {
            "sheet_order": list(wb.sheetnames),
            "workbook_protection": workbook_digest_serialisable(wb.security),
            "calculation": workbook_digest_serialisable(wb.calculation),
            "defined_names": defined_names,
            "sheets": sheets,
        }
    finally:
        wb.close()


def workbook_metadata_sha256(path: Path) -> str:
    """Return the legacy v2 metadata-rich digest without reinterpreting it."""

    return canonical_json_sha256(workbook_metadata_payload_v2(path))


def workbook_internal_hyperlink_destination(
    hyperlink: Any,
    *,
    sheet_name: str,
    coordinate: str,
) -> str | None:
    """Return one canonical in-workbook destination across Excel serializations."""

    if hyperlink is None:
        return None
    target = str(hyperlink.target or "").strip()
    location = str(hyperlink.location or "").strip()
    if target and not target.startswith("#"):
        raise IntegrityError(
            "External hyperlinks are not allowed in the protected master workbook: "
            f"{sheet_name}!{coordinate}."
        )
    def canonical_destination(value: str) -> str:
        stripped = value.strip()
        if stripped.startswith("##"):
            raise IntegrityError(
                "Workbook hyperlink has more than one leading fragment marker: "
                f"{sheet_name}!{coordinate}."
            )
        return stripped[1:].strip() if stripped.startswith("#") else stripped

    destinations = [canonical_destination(value) for value in (target, location) if value]
    if any(
        re.search(r"(?:https?|file)://|\\\\|[A-Za-z]:[\\/]", value, re.I)
        or value.startswith(("/", "../", "..\\"))
        for value in destinations
    ):
        raise IntegrityError(
            "External hyperlinks are not allowed in the protected master workbook: "
            f"{sheet_name}!{coordinate}."
        )
    if not destinations:
        raise IntegrityError(
            f"Workbook hyperlink has no internal destination: {sheet_name}!{coordinate}."
        )
    if len(set(destinations)) != 1:
        raise IntegrityError(
            f"Workbook hyperlink has conflicting destinations: {sheet_name}!{coordinate}."
        )
    return destinations[0]


def workbook_semantic_hyperlink_payload(
    hyperlink: Any,
    *,
    sheet_name: str,
    coordinate: str,
    cell_value: Any = None,
) -> dict[str, Any] | None:
    if hyperlink is None:
        return None
    display = hyperlink.display
    if display is not None and cell_value is not None and str(display) == str(cell_value):
        # Excel materializes the visible cell text into the optional display
        # attribute when it saves an internal hyperlink.  The cell value is already
        # protected separately, so the two serializations are equivalent.
        display = None
    return {
        "destination": workbook_internal_hyperlink_destination(
            hyperlink,
            sheet_name=sheet_name,
            coordinate=coordinate,
        ),
        "tooltip": hyperlink.tooltip,
        "display": display,
    }


def workbook_hyperlink_matches(
    hyperlink: Any,
    expected: str,
    *,
    sheet_name: str,
    coordinate: str,
) -> bool:
    if hyperlink is None:
        return False
    return workbook_internal_hyperlink_destination(
        hyperlink,
        sheet_name=sheet_name,
        coordinate=coordinate,
    ) == (
        str(expected).strip()[1:].strip()
        if str(expected).strip().startswith("#")
        else str(expected).strip()
    )


def workbook_semantic_table_payload(table: Any) -> dict[str, Any]:
    def formula_payload(formula: Any) -> dict[str, Any] | None:
        if formula is None:
            return None
        return {
            "text": getattr(formula, "text", None),
            "array": bool(getattr(formula, "array", False)),
        }

    columns = []
    for column in table.tableColumns:
        calculated = getattr(column, "calculatedColumnFormula", None)
        totals = getattr(column, "totalsRowFormula", None)
        columns.append(
            {
                "id": column.id,
                "name": column.name,
                "totals_row_function": getattr(column, "totalsRowFunction", None),
                "totals_row_label": getattr(column, "totalsRowLabel", None),
                "unique_name": getattr(column, "uniqueName", None),
                "query_table_field_id": getattr(column, "queryTableFieldId", None),
                "header_row_dxf_id": getattr(column, "headerRowDxfId", None),
                "data_dxf_id": getattr(column, "dataDxfId", None),
                "totals_row_dxf_id": getattr(column, "totalsRowDxfId", None),
                "header_row_cell_style": getattr(column, "headerRowCellStyle", None),
                "data_cell_style": getattr(column, "dataCellStyle", None),
                "totals_row_cell_style": getattr(column, "totalsRowCellStyle", None),
                "calculated_formula": formula_payload(calculated),
                "totals_formula": formula_payload(totals),
                "xml_column_properties": workbook_digest_serialisable(
                    getattr(column, "xmlColumnPr", None)
                ),
                "extensions": workbook_digest_serialisable(
                    getattr(column, "extLst", None)
                ),
            }
        )
    return {
        "id": table.id,
        "name": table.name,
        "display_name": table.displayName,
        "comment": table.comment,
        "ref": table.ref,
        "table_type": table.tableType,
        "header_row_count": table.headerRowCount,
        "insert_row": bool(table.insertRow),
        "insert_row_shift": bool(table.insertRowShift),
        "totals_row_count": table.totalsRowCount,
        "totals_row_shown": table.totalsRowShown is not False,
        "published": bool(table.published),
        "header_row_dxf_id": table.headerRowDxfId,
        "data_dxf_id": table.dataDxfId,
        "totals_row_dxf_id": table.totalsRowDxfId,
        "header_row_border_dxf_id": table.headerRowBorderDxfId,
        "table_border_dxf_id": table.tableBorderDxfId,
        "totals_row_border_dxf_id": table.totalsRowBorderDxfId,
        "header_row_cell_style": table.headerRowCellStyle,
        "data_cell_style": table.dataCellStyle,
        "totals_row_cell_style": table.totalsRowCellStyle,
        "connection_id": table.connectionId,
        "auto_filter": workbook_digest_serialisable(table.autoFilter),
        "sort_state": workbook_digest_serialisable(table.sortState),
        "style": (
            {
                "name": table.tableStyleInfo.name,
                "show_first_column": bool(table.tableStyleInfo.showFirstColumn),
                "show_last_column": bool(table.tableStyleInfo.showLastColumn),
                "show_row_stripes": bool(table.tableStyleInfo.showRowStripes),
                "show_column_stripes": bool(table.tableStyleInfo.showColumnStripes),
            }
            if table.tableStyleInfo is not None
            else None
        ),
        "columns": columns,
    }


def workbook_semantic_validation_formula(value: Any) -> Any:
    if value is None:
        return None
    text = str(value)
    if re.fullmatch(r"=?[A-Za-z_\\][A-Za-z0-9_.]*", text):
        return text.lstrip("=")
    return value


def workbook_semantic_validation_payload(validation: Any) -> dict[str, Any]:
    return {
        "type": validation.type,
        "operator": validation.operator,
        "formula1": workbook_semantic_validation_formula(validation.formula1),
        "formula2": workbook_semantic_validation_formula(validation.formula2),
        "sqref": str(validation.sqref),
        "allow_blank": bool(validation.allowBlank),
        "show_drop_down": bool(validation.showDropDown),
        "show_input": bool(validation.showInputMessage),
        "show_error": bool(validation.showErrorMessage),
        "error_style": (
            validation.errorStyle
            or ("stop" if validation.showErrorMessage is not False else None)
        ),
        "error_title": validation.errorTitle,
        "error": validation.error,
        "prompt_title": validation.promptTitle,
        "prompt": validation.prompt,
        "ime_mode": validation.imeMode,
    }


def workbook_semantic_workbook_protection_payload(protection: Any) -> dict[str, Any] | None:
    if protection is None:
        return None
    payload = {
        "lock_structure": bool(protection.lockStructure),
        "lock_windows": bool(protection.lockWindows),
        "lock_revision": bool(protection.lockRevision),
        "workbook_password": protection.workbookPassword,
        "workbook_password_character_set": (
            protection.workbookPasswordCharacterSet
        ),
        "revisions_password": protection.revisionsPassword,
        "revisions_password_character_set": (
            protection.revisionsPasswordCharacterSet
        ),
        "workbook_algorithm": protection.workbookAlgorithmName,
        "workbook_hash": protection.workbookHashValue,
        "workbook_salt": protection.workbookSaltValue,
        "workbook_spin_count": protection.workbookSpinCount,
        "revisions_algorithm": protection.revisionsAlgorithmName,
        "revisions_hash": protection.revisionsHashValue,
        "revisions_salt": protection.revisionsSaltValue,
        "revisions_spin_count": protection.revisionsSpinCount,
    }
    if not any(value not in (None, False) for value in payload.values()):
        return None
    return payload


def workbook_semantic_number_format(number_format: str) -> str:
    """Normalize only Excel escape spelling that preserves literal display text."""

    # Backslashes inside a quoted literal are visible characters. Track quoted
    # runs explicitly because Excel permits an escaped quote (\") inside them.
    result: list[str] = []
    in_quote = False
    for index, character in enumerate(number_format):
        preceding_backslashes = 0
        cursor = index - 1
        while cursor >= 0 and number_format[cursor] == "\\":
            preceding_backslashes += 1
            cursor -= 1
        if character == '"' and preceding_backslashes % 2 == 0:
            in_quote = not in_quote
        if (
            character == "\\"
            and not in_quote
            and index + 1 < len(number_format)
            and number_format[index + 1] in "$+-"
            and preceding_backslashes % 2 == 0
        ):
            continue
        result.append(character)
    return "".join(result)


def workbook_semantic_color_payload(color: Any) -> dict[str, Any] | None:
    """Capture effective color semantics while ignoring OOXML alpha spelling."""

    if color is None:
        return None
    color_type = getattr(color, "type", None)
    payload: dict[str, Any] = {"type": color_type}
    if color_type == "rgb":
        rgb = getattr(color, "rgb", None)
        payload["value"] = str(rgb)[-6:].upper() if rgb else None
    elif color_type == "indexed":
        payload["value"] = getattr(color, "indexed", None)
    elif color_type == "theme":
        payload["value"] = getattr(color, "theme", None)
    elif color_type == "auto":
        payload["value"] = bool(getattr(color, "auto", False))
    else:
        payload.update(
            {
                "rgb": getattr(color, "rgb", None),
                "indexed": getattr(color, "indexed", None),
                "theme": getattr(color, "theme", None),
                "auto": getattr(color, "auto", None),
            }
        )
    tint = float(getattr(color, "tint", 0.0) or 0.0)
    payload["tint"] = round(tint, 8)
    return payload


def workbook_semantic_border_side_payload(side: Any) -> dict[str, Any] | None:
    if side is None or getattr(side, "style", None) is None:
        return None
    return {
        "style": side.style,
        "color": workbook_semantic_color_payload(side.color),
    }


def workbook_semantic_style_payload(styleable: Any) -> dict[str, Any]:
    """Capture visible style semantics without relying on workbook style IDs."""

    font = styleable.font
    fill = styleable.fill
    border = styleable.border
    alignment = styleable.alignment
    protection = styleable.protection
    fill_type = fill.fill_type
    fill_payload: dict[str, Any] = {"type": fill_type}
    if fill_type is not None:
        foreground = getattr(fill, "fgColor", None)
        background = getattr(fill, "bgColor", None)
        if foreground is not None:
            fill_payload["foreground"] = workbook_semantic_color_payload(foreground)
        if fill_type != "solid" and background is not None:
            fill_payload["background"] = workbook_semantic_color_payload(background)
        for attribute in ("degree", "left", "right", "top", "bottom"):
            value = getattr(fill, attribute, None)
            if value is not None:
                fill_payload[attribute] = value
        stops = getattr(fill, "stop", None)
        if stops:
            fill_payload["stops"] = [
                {
                    "position": stop.position,
                    "color": workbook_semantic_color_payload(stop.color),
                }
                for stop in stops
            ]
    return {
        "font": {
            # Excel writes the workbook default font name into otherwise explicit
            # font records.  In this workbook that default is Calibri.
            "name": font.name or "Calibri",
            "size": float(font.sz) if font.sz is not None else 11.0,
            "bold": bool(font.b),
            "italic": bool(font.i),
            "underline": font.u,
            "strike": bool(font.strike),
            "outline": bool(font.outline),
            "shadow": bool(font.shadow),
            "condense": bool(font.condense),
            "extend": bool(font.extend),
            "vertical_align": font.vertAlign,
            "charset": font.charset,
            "family": font.family,
            "scheme": font.scheme,
            "color": workbook_semantic_color_payload(font.color),
        },
        "fill": fill_payload,
        "border": {
            "left": workbook_semantic_border_side_payload(border.left),
            "right": workbook_semantic_border_side_payload(border.right),
            "top": workbook_semantic_border_side_payload(border.top),
            "bottom": workbook_semantic_border_side_payload(border.bottom),
            "diagonal": workbook_semantic_border_side_payload(border.diagonal),
            "vertical": workbook_semantic_border_side_payload(border.vertical),
            "horizontal": workbook_semantic_border_side_payload(border.horizontal),
            "start": workbook_semantic_border_side_payload(border.start),
            "end": workbook_semantic_border_side_payload(border.end),
            "diagonal_up": bool(border.diagonalUp),
            "diagonal_down": bool(border.diagonalDown),
            "outline": border.outline is not False,
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "text_rotation": int(alignment.textRotation or 0),
            "wrap_text": bool(alignment.wrapText),
            "shrink_to_fit": bool(alignment.shrinkToFit),
            "indent": float(alignment.indent or 0.0),
            "relative_indent": float(alignment.relativeIndent or 0.0),
            "justify_last_line": bool(alignment.justifyLastLine),
            "reading_order": float(alignment.readingOrder or 0.0),
        },
        "number_format": workbook_semantic_number_format(styleable.number_format),
        "protection": {
            "locked": protection.locked is not False,
            "hidden": bool(protection.hidden),
        },
    }


def workbook_semantic_calculation_payload(calculation: Any) -> dict[str, Any]:
    """Keep calculation behavior while normalizing omitted OOXML defaults."""

    return {
        # calcId identifies the Excel calculation-engine build, not behavior.
        "calc_mode": calculation.calcMode or "auto",
        "full_calc_on_load": calculation.fullCalcOnLoad is not False,
        "force_full_calc": calculation.forceFullCalc is not False,
        "reference_mode": calculation.refMode or "A1",
        "iterate": bool(calculation.iterate),
        "iterate_count": (
            calculation.iterateCount
            if calculation.iterateCount is not None
            else 100
        ),
        "iterate_delta": (
            calculation.iterateDelta
            if calculation.iterateDelta is not None
            else 0.001
        ),
        "full_precision": (
            True if calculation.fullPrecision is None else bool(calculation.fullPrecision)
        ),
        "calculation_completed": (
            True if calculation.calcCompleted is None else bool(calculation.calcCompleted)
        ),
        "calculate_on_save": (
            True if calculation.calcOnSave is None else bool(calculation.calcOnSave)
        ),
        "concurrent_calculation": (
            True if calculation.concurrentCalc is None else bool(calculation.concurrentCalc)
        ),
        "concurrent_manual_count": calculation.concurrentManualCount,
    }


def workbook_chart_semantic_xml(node: Any) -> dict[str, Any] | None:
    """Serialize chart semantics fail-closed, excluding only derived data caches."""

    local = str(node.tag).rsplit("}", 1)[-1]
    if local in {"numCache", "strCache", "multiLvlStrCache"}:
        # These values are derived from protected worksheet cells and Excel may add or
        # discard them on open/save. The source formula remains part of the digest.
        return None
    raw_attributes: dict[str, str] = {}
    for key, value in sorted(node.attrib.items(), key=lambda item: str(item[0])):
        expanded_key = str(key)
        text_value = str(value)
        if expanded_key.rsplit("}", 1)[-1] == "formatCode":
            text_value = workbook_semantic_number_format(text_value)
        raw_attributes[expanded_key] = text_value
    attributes_by_local = {
        key.rsplit("}", 1)[-1]: value for key, value in raw_attributes.items()
    }
    if local == "smooth" and attributes_by_local.get("val", "0").casefold() in {
        "0", "false",
    }:
        return None
    text = str(node.text) if node.text is not None else None
    xml_space = "{http://www.w3.org/XML/1998/namespace}space"
    if (
        local == "t"
        and raw_attributes.get(xml_space) == "preserve"
        and text is not None
        and text != text.strip()
    ):
        # Excel preserves the text while omitting this redundant marker on save.
        raw_attributes.pop(xml_space)
        attributes_by_local = {
            key.rsplit("}", 1)[-1]: value for key, value in raw_attributes.items()
        }
    if local == "f" and text is not None:
        # Excel may add/remove absolute-reference markers without changing the source.
        text = workbook_chart_formula(text)
    children = [
        payload
        for child in list(node)
        if (payload := workbook_chart_semantic_xml(child)) is not None
    ]
    if local == "dLbls" and not raw_attributes and text is None:
        false_label_tags = {
            "showLegendKey", "showVal", "showCatName", "showSerName",
            "showPercent", "showBubbleSize", "showLeaderLines", "showMarker",
        }
        if all(
            child["local_tag"] in false_label_tags
            and {
                key.rsplit("}", 1)[-1]: value
                for key, value in child["attributes"].items()
            }.get("val", "0").casefold()
            in {"0", "false"}
            and not child["children"]
            for child in children
        ):
            return None
    if (
        local == "numFmt"
        and attributes_by_local.get("formatCode") == "General"
        and attributes_by_local.get("sourceLinked", "1").casefold()
        in {"1", "true"}
        and set(attributes_by_local) <= {"formatCode", "sourceLinked"}
        and text is None
        and not children
    ):
        return None
    if local == "lstStyle" and not raw_attributes and text is None and not children:
        return None
    if local == "layout" and not raw_attributes and text is None and not children:
        return None
    if local == "spPr" and not raw_attributes and text is None and len(children) == 1:
        [line] = children
        if (
            line["tag"]
            == "{http://schemas.openxmlformats.org/drawingml/2006/main}ln"
            and not line["attributes"]
            and line["text"] is None
            and len(line["children"]) == 1
            and line["children"][0]["tag"]
            == "{http://schemas.openxmlformats.org/drawingml/2006/main}prstDash"
            and line["children"][0]["attributes"] == {"val": "solid"}
            and line["children"][0]["text"] is None
            and not line["children"][0]["children"]
        ):
            # openpyxl emits this no-op marker outline; Excel removes it.
            return None
    if (
        local == "rPr"
        and attributes_by_local == {"lang": "en-US"}
        and text is None
        and not children
    ):
        return None
    if (
        local == "autoTitleDeleted"
        and attributes_by_local.get("val", "0").casefold() in {"0", "false"}
        and text is None
        and not children
    ):
        return None
    if (
        local == "showDLblsOverMax"
        and attributes_by_local.get("val", "1").casefold() in {"1", "true"}
        and text is None
        and not children
    ):
        return None
    if (
        local == "date1904"
        and attributes_by_local.get("val", "0").casefold() in {"0", "false"}
        and text is None
        and not children
    ):
        return None
    if (
        local == "lang"
        and attributes_by_local.get("val", "").casefold() == "en-us"
        and text is None
        and not children
    ):
        return None
    if (
        local == "roundedCorners"
        and attributes_by_local.get("val", "1").casefold() in {"1", "true"}
        and text is None
        and not children
    ):
        # Desktop Excel materializes the schema/default rounded-corner setting
        # on first save even when the generated chart part omits it.
        return None
    if (
        local == "style"
        and attributes_by_local.get("val") == "2"
        and text is None
        and not children
    ):
        # Style 2 is Excel's first-save materialization for an omitted chart style.
        return None
    if local == "printSettings" and not raw_attributes and text is None:
        child_by_tag = {child["local_tag"]: child for child in children}
        header_footer = child_by_tag.get("headerFooter")
        page_margins = child_by_tag.get("pageMargins")
        page_setup = child_by_tag.get("pageSetup")
        if (
            set(child_by_tag) == {"headerFooter", "pageMargins", "pageSetup"}
            and header_footer is not None
            and not header_footer["attributes"]
            and header_footer["text"] is None
            and not header_footer["children"]
            and page_margins is not None
            and page_margins["attributes"]
            == {
                "b": "0.75",
                "footer": "0.3",
                "header": "0.3",
                "l": "0.7",
                "r": "0.7",
                "t": "0.75",
            }
            and page_margins["text"] is None
            and not page_margins["children"]
            and page_setup is not None
            and not page_setup["attributes"]
            and page_setup["text"] is None
            and not page_setup["children"]
        ):
            return None
    if local == "extLst" and not raw_attributes and text is None and len(children) == 1:
        [extension] = children
        extension_attributes = {
            key.rsplit("}", 1)[-1]: value
            for key, value in extension["attributes"].items()
        }
        if (
            extension["tag"]
            == "{http://schemas.openxmlformats.org/drawingml/2006/chart}ext"
            and extension_attributes.get("uri")
            == "{C3380CC4-5D6E-409C-BE32-E72D297353CC}"
            and extension["text"] is None
            and len(extension["children"]) == 1
            and extension["children"][0]["tag"]
            == "{http://schemas.microsoft.com/office/drawing/2014/chart}uniqueId"
        ):
            # Excel generates a series UUID on first save; it is identity metadata,
            # while the series index/order/formulas remain protected separately.
            return None
    return {
        "tag": str(node.tag),
        "local_tag": local,
        "attributes": raw_attributes,
        "text": text,
        "children": children,
    }


def workbook_chart_semantic_part(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    try:
        payload = workbook_chart_semantic_xml(value.to_tree())
        if (
            payload is not None
            and not payload["attributes"]
            and payload["text"] is None
            and not payload["children"]
        ):
            return None
        return payload
    except Exception as exc:
        raise IntegrityError(
            f"Workbook integrity could not inspect chart semantics: {exc}"
        ) from exc


def workbook_chart_anchor_payload(anchor: Any) -> dict[str, Any]:
    """Bind chart position and meaningful size while ignoring tiny EMU rounding."""

    if isinstance(anchor, str):
        return {"type": "cell", "coordinate": anchor}

    def emu(value: Any) -> int | None:
        if value is None:
            return None
        return int(round(int(value) / 10_000.0) * 10_000)

    def marker(value: Any) -> dict[str, Any] | None:
        if value is None:
            return None
        return {
            "column": value.col,
            "row": value.row,
            "column_offset": emu(value.colOff),
            "row_offset": emu(value.rowOff),
        }

    extent = getattr(anchor, "ext", None)
    position = getattr(anchor, "pos", None)
    client_data = getattr(anchor, "clientData", None)
    return {
        "type": type(anchor).__name__,
        "from": marker(getattr(anchor, "_from", None)),
        "to": marker(getattr(anchor, "to", None)),
        "position": (
            {"x": emu(position.x), "y": emu(position.y)}
            if position is not None
            else None
        ),
        "extent": (
            {"width": emu(extent.cx), "height": emu(extent.cy)}
            if extent is not None
            else None
        ),
        "edit_as": getattr(anchor, "editAs", None),
        "client_data": (
            {
                "locks_with_sheet": getattr(client_data, "fLocksWithSheet", None),
                "prints_with_sheet": getattr(client_data, "fPrintsWithSheet", None),
            }
            if client_data is not None
            else None
        ),
    }


def workbook_raw_chart_parts_payload(path: Path) -> list[dict[str, Any]]:
    """Hash chart-space XML that openpyxl does not fully restore when reading."""

    from xml.etree import ElementTree
    from zipfile import BadZipFile, ZipFile

    try:
        with ZipFile(path) as archive:
            names = sorted(
                name
                for name in archive.namelist()
                if name.startswith("xl/charts/chart") and name.endswith(".xml")
            )
            return [
                {
                    "part": name,
                    "definition": workbook_chart_semantic_xml(
                        ElementTree.fromstring(archive.read(name))
                    ),
                }
                for name in names
            ]
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError) as exc:
        raise IntegrityError(f"Workbook chart XML is invalid: {exc}") from exc


def workbook_chart_frame_payload(path: Path) -> list[dict[str, Any]]:
    """Hash visible and locking semantics from chart drawing frames."""

    from xml.etree import ElementTree
    from zipfile import BadZipFile, ZipFile

    def local_name(node: Any) -> str:
        return str(node.tag).rsplit("}", 1)[-1]

    def truthy(value: Any) -> bool:
        return str(value or "0").casefold() in {"1", "true"}

    try:
        payload: list[dict[str, Any]] = []
        with ZipFile(path) as archive:
            names = sorted(
                name
                for name in archive.namelist()
                if name.startswith("xl/drawings/drawing")
                and name.endswith(".xml")
            )
            for name in names:
                root = ElementTree.fromstring(archive.read(name))
                for frame_index, frame in enumerate(
                    node for node in root.iter() if local_name(node) == "graphicFrame"
                ):
                    nonvisual = next(
                        (
                            node
                            for node in frame.iter()
                            if local_name(node) == "cNvPr"
                        ),
                        None,
                    )
                    frame_properties = next(
                        (
                            node
                            for node in frame.iter()
                            if local_name(node) == "cNvGraphicFramePr"
                        ),
                        None,
                    )
                    locks = next(
                        (
                            node
                            for node in (frame_properties.iter() if frame_properties is not None else ())
                            if local_name(node) == "graphicFrameLocks"
                        ),
                        None,
                    )
                    lock_attributes = (
                        {
                            key: truthy(locks.attrib.get(key))
                            for key in (
                                "noChangeAspect",
                                "noDrilldown",
                                "noGrp",
                                "noMove",
                                "noResize",
                                "noSelect",
                            )
                        }
                        if locks is not None
                        else {
                            key: False
                            for key in (
                                "noChangeAspect",
                                "noDrilldown",
                                "noGrp",
                                "noMove",
                                "noResize",
                                "noSelect",
                            )
                        }
                    )
                    payload.append(
                        {
                            "part": name,
                            "frame_index": frame_index,
                            "macro": frame.attrib.get("macro") or None,
                            "name": (
                                nonvisual.attrib.get("name")
                                if nonvisual is not None
                                else None
                            ),
                            "description": (
                                nonvisual.attrib.get("descr")
                                if nonvisual is not None
                                else None
                            ),
                            "title": (
                                nonvisual.attrib.get("title")
                                if nonvisual is not None
                                else None
                            ),
                            "hidden": (
                                truthy(nonvisual.attrib.get("hidden"))
                                if nonvisual is not None
                                else False
                            ),
                            "locks": lock_attributes,
                        }
                    )
        return payload
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError) as exc:
        raise IntegrityError(f"Workbook chart drawing XML is invalid: {exc}") from exc


def workbook_chart_formula(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value)
    # Preserve literal dollar signs in quoted sheet names, string literals, and
    # external-workbook tokens. Only remove absolute markers inside A1 cell
    # references, which desktop Excel may add or discard on a no-edit save.
    protected_token = re.compile(
        r"('(?:[^']|'')*'|\"(?:[^\"]|\"\")*\"|\[[^\]]*\])"
    )
    cell_reference = re.compile(
        r"(?<![A-Za-z0-9_.])\$?([A-Za-z]{1,3})\$?(\d+)(?![A-Za-z0-9_])"
    )
    parts = protected_token.split(text)
    for index in range(0, len(parts), 2):
        parts[index] = cell_reference.sub(r"\1\2", parts[index])
    normalized = "".join(parts)
    return re.sub(
        r"^'([A-Za-z_][A-Za-z0-9_.]*)'!",
        r"\1!",
        normalized,
    )


def workbook_chart_text_payload(value: Any) -> dict[str, Any] | None:
    """Extract visible text and source formulas without empty rich-text defaults."""

    if value is None:
        return None
    to_tree = getattr(value, "to_tree", None)
    if not callable(to_tree):
        return {"text": str(value), "formulas": []}
    try:
        root = to_tree()
    except Exception as exc:
        raise IntegrityError(
            f"Workbook integrity could not inspect chart text: {exc}"
        ) from exc
    text: list[str] = []
    formulas: list[str] = []
    for node in root.iter():
        local = str(node.tag).rsplit("}", 1)[-1]
        if node.text is None:
            continue
        if local == "t":
            text.append(str(node.text))
        elif local == "f":
            formulas.append(workbook_chart_formula(node.text) or "")
    return {"text": text, "formulas": formulas}


def workbook_chart_color_choice_payload(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    for attribute in (
        "srgbClr", "schemeClr", "prstClr", "sysClr", "scrgbClr", "hslClr",
    ):
        color = getattr(value, attribute, None)
        if color is not None:
            return {"type": attribute, "value": str(color)}
    return workbook_chart_semantic_part(value)


def workbook_chart_graphical_properties_payload(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    line = getattr(value, "line", None)
    return {
        "no_fill": bool(getattr(value, "noFill", False)),
        "solid_fill": workbook_chart_color_choice_payload(
            getattr(value, "solidFill", None)
        ),
        "line": (
            {
                "no_fill": bool(getattr(line, "noFill", False)),
                "solid_fill": workbook_chart_color_choice_payload(
                    getattr(line, "solidFill", None)
                ),
                "width": getattr(line, "width", None),
                "dash": getattr(line, "prstDash", None) or "solid",
            }
            if line is not None
            else None
        ),
    }


def workbook_chart_data_source_payload(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    for attribute in ("strRef", "numRef", "multiLvlStrRef"):
        reference = getattr(value, attribute, None)
        if reference is not None and getattr(reference, "f", None) is not None:
            return {
                # Excel resolves a category source to strRef or numRef from the
                # protected source cells when it saves.  The referenced range,
                # not that derived cache type, is the substantive contract.
                "type": (
                    "multi_level_reference"
                    if attribute == "multiLvlStrRef"
                    else "reference"
                ),
                "formula": workbook_chart_formula(reference.f),
            }
    for attribute in ("strLit", "numLit"):
        literal = getattr(value, attribute, None)
        if literal is not None:
            points = [
                (point.idx, getattr(point, "v", None))
                for point in (getattr(literal, "pt", None) or [])
            ]
            return {"type": attribute, "points": points}
    return None


def workbook_chart_series_payload(
    series: Any,
    *,
    line_default: bool,
    bar_default: bool,
) -> dict[str, Any]:
    title = getattr(series, "tx", None)
    title_reference = getattr(title, "strRef", None)
    return {
        "index": getattr(series, "idx", None),
        "order": getattr(series, "order", None),
        "title": {
            "formula": workbook_chart_formula(
                getattr(title_reference, "f", None)
            ),
            "value": getattr(title, "v", None),
        },
        "categories": workbook_chart_data_source_payload(
            getattr(series, "cat", None) or getattr(series, "xVal", None)
        ),
        "values": workbook_chart_data_source_payload(
            getattr(series, "val", None) or getattr(series, "yVal", None)
        ),
        "bubble_size": workbook_chart_data_source_payload(
            getattr(series, "bubbleSize", None)
        ),
        "graphical_properties": workbook_chart_graphical_properties_payload(
            getattr(series, "graphicalProperties", None)
        ),
        "marker": {
            "symbol": getattr(getattr(series, "marker", None), "symbol", None),
            "size": getattr(getattr(series, "marker", None), "size", None),
        },
        "smooth": (
            getattr(series, "smooth", None) if line_default else None
        ),
        "invert_if_negative": (
            getattr(series, "invertIfNegative", None) if bar_default else None
        ),
    }


def workbook_chart_axis_payload(axis: Any) -> dict[str, Any]:
    scaling = getattr(axis, "scaling", None)
    number_format = getattr(axis, "numFmt", None)
    return {
        "id": getattr(axis, "axId", None),
        "deleted": getattr(axis, "delete", None),
        "position": getattr(axis, "axPos", None),
        "crosses": getattr(axis, "crosses", None),
        "title": workbook_chart_text_payload(getattr(axis, "title", None)),
        "number_format": (
            {
                "code": workbook_semantic_number_format(number_format.formatCode),
                "source_linked": number_format.sourceLinked,
            }
            if number_format is not None
            else None
        ),
        "tick_label_position": getattr(axis, "tickLblPos", None),
        "scaling": {
            "log_base": getattr(scaling, "logBase", None),
            "orientation": getattr(scaling, "orientation", None) or "minMax",
            "minimum": getattr(scaling, "min", None),
            "maximum": getattr(scaling, "max", None),
        },
        "major_unit": getattr(axis, "majorUnit", None),
        "minor_unit": getattr(axis, "minorUnit", None),
    }


def workbook_chart_semantic_payload(chart: Any) -> dict[str, Any]:
    """Keep complete chart semantics while normalizing enumerated Excel defaults."""

    plots = list(getattr(chart, "_charts", None) or [chart])
    all_axes: dict[str, Any] = {}
    for plot in plots:
        axes = getattr(plot, "_axes", None) or {}
        axis_items = axes.items() if hasattr(axes, "items") else enumerate(axes)
        for axis_id, axis in axis_items:
            all_axes[str(axis_id)] = axis
    return {
        "type": type(chart).__name__,
        "plots": [
            {
                "type": type(plot).__name__,
                "definition": workbook_chart_semantic_part(plot),
            }
            for plot in plots
        ],
        "title": workbook_chart_semantic_part(getattr(chart, "title", None)),
        "legend": workbook_chart_semantic_part(getattr(chart, "legend", None)),
        "layout": workbook_chart_semantic_part(getattr(chart, "layout", None)),
        "graphical_properties": workbook_chart_semantic_part(
            getattr(chart, "graphical_properties", None)
        ),
        "axes": [
            {
                "id": axis_id,
                "definition": workbook_chart_semantic_part(axis),
            }
            for axis_id, axis in sorted(all_axes.items(), key=lambda item: item[0])
        ],
        "style": getattr(chart, "style", None),
        "rounded_corners": getattr(chart, "roundedCorners", None),
        "display_blanks": getattr(chart, "display_blanks", None),
        "visible_cells_only": getattr(chart, "visible_cells_only", None),
        "index_base": getattr(chart, "idx_base", None),
        "pivot_source": workbook_chart_semantic_part(
            getattr(chart, "pivotSource", None)
        ),
        "pivot_formats": [
            workbook_chart_semantic_part(pivot_format)
            for pivot_format in (getattr(chart, "pivotFormats", None) or [])
        ],
        "anchor": workbook_chart_anchor_payload(chart.anchor),
    }


def workbook_semantic_conditional_formatting_payload(ws: Any) -> list[dict[str, Any]]:
    def normalized_xml(
        node: Any,
        *,
        ignored_attributes: frozenset[str] = frozenset(),
        normalize_argb: bool = False,
    ) -> dict[str, Any]:
        attributes: dict[str, str] = {}
        for key, value in sorted(node.attrib.items(), key=lambda item: str(item[0])):
            local_key = str(key).rsplit("}", 1)[-1]
            if local_key in ignored_attributes:
                continue
            text = str(value)
            if local_key == "formatCode":
                text = workbook_semantic_number_format(text)
            if normalize_argb and local_key in {"rgb", "fgColor", "bgColor"}:
                text = text[-6:].upper()
            attributes[local_key] = text
        return {
            "tag": str(node.tag).rsplit("}", 1)[-1],
            "attributes": attributes,
            "text": node.text,
            "children": [
                normalized_xml(
                    child,
                    ignored_attributes=ignored_attributes,
                    normalize_argb=normalize_argb,
                )
                for child in list(node)
            ],
        }

    payload: list[dict[str, Any]] = []
    for conditional_format in ws.conditional_formatting:
        priorities = [rule.priority for rule in conditional_format.rules]
        if len(priorities) != len(set(priorities)):
            raise IntegrityError(
                "Conditional-format rules must have unique priorities in the "
                f"protected master workbook ({ws.title}!{conditional_format.sqref})."
            )
        rules = [
            {
                "definition": normalized_xml(
                    rule.to_tree(),
                    ignored_attributes=frozenset({"dxfId"}),
                ),
                "differential_style": (
                    normalized_xml(rule.dxf.to_tree(), normalize_argb=True)
                    if rule.dxf is not None
                    else None
                ),
            }
            for rule in conditional_format.rules
        ]
        rules.sort(key=lambda item: json.dumps(item, sort_keys=True, default=str))
        payload.append(
            {
                "range": str(conditional_format.sqref),
                "pivot": bool(conditional_format.pivot),
                "rules": rules,
            }
        )
    return sorted(payload, key=lambda item: item["range"])


def workbook_semantic_differential_style_payload(style: Any) -> dict[str, Any]:
    def normalized_xml(node: Any) -> dict[str, Any]:
        attributes: dict[str, str] = {}
        for key, value in sorted(node.attrib.items(), key=lambda item: str(item[0])):
            expanded_key = str(key)
            local_key = expanded_key.rsplit("}", 1)[-1]
            text = str(value)
            if local_key == "formatCode":
                text = workbook_semantic_number_format(text)
            if local_key in {"rgb", "fgColor", "bgColor"}:
                text = text[-6:].upper()
            attributes[expanded_key] = text
        return {
            "tag": str(node.tag),
            "attributes": attributes,
            "text": node.text,
            "children": [normalized_xml(child) for child in list(node)],
        }

    try:
        return normalized_xml(style.to_tree())
    except Exception as exc:
        raise IntegrityError(
            f"Workbook integrity could not inspect a differential style: {exc}"
        ) from exc


def workbook_semantic_dimension_payload(dimension: Any) -> dict[str, Any]:
    payload = {
        "hidden": bool(dimension.hidden),
        "outline_level": dimension.outlineLevel,
        "collapsed": bool(dimension.collapsed),
        "style": (
            workbook_semantic_style_payload(dimension)
            if dimension.has_style
            else None
        ),
    }
    for attribute in ("width", "bestFit", "height", "thickTop", "thickBot"):
        if hasattr(dimension, attribute):
            value = getattr(dimension, attribute)
            payload[attribute] = value
    return payload


def workbook_semantic_column_dimensions_payload(ws: Any) -> list[dict[str, Any]]:
    """Canonicalize grouped/split column records to effective per-column state."""

    effective: dict[int, dict[str, Any]] = {}
    for index, dimension in ws.column_dimensions.items():
        default_index = column_index_from_string(str(index))
        start = dimension.min or default_index
        end = dimension.max or start
        payload = workbook_semantic_dimension_payload(dimension)
        for column in range(start, end + 1):
            effective[column] = payload
    return [
        {"index": get_column_letter(column), **effective[column]}
        for column in sorted(effective)
    ]


EXCEL_AUTOFIT_ROW_HEIGHTS: dict[str, dict[int, float]] = {
    "Methodology": {19: 30.0},
    "Weekly Server Metrics": {1: 19.5, 3: 30.0},
    "Weekly Server Rankings": {1: 19.5, 3: 30.0},
    "Server Week-over-Week Detail": {1: 19.5, 3: 30.0},
    "Weekly Location Metrics": {1: 19.5, 3: 45.0},
    "Store Week Trends": {1: 19.5, 3: 30.0},
    "Store Trend Summary": {1: 19.5, 3: 45.0},
    "Group Week Trends": {1: 19.5, 3: 30.0},
    "Group Trend Summary": {1: 19.5, 3: 45.0},
    "Daily Server Detail": {1: 19.5, 3: 30.0},
    "Daily Location Detail": {1: 19.5, 3: 45.0},
    "Server Trend Summary": {1: 19.5, 3: 30.0},
    "_Data Quality Detail": {1: 21.0, 3: 30.0},
    "_Consistency Calc": {3: 30.0},
    "Dashboard": {18: 30.0},
}

# Desktop Excel rewrites these explicit point sizes on its first save. Generated
# output uses the observed persisted value up front so every row height can remain
# exact in the substantive digest.
EXCEL_ROW_HEIGHT_SERIALIZATION: dict[float, float] = {
    20.0: 20.1,
    22.0: 21.95,
    26.0: 26.1,
    28.0: 27.95,
    32.0: 32.1,
    34.0: 33.95,
    38.0: 38.1,
    46.0: 45.95,
    50.0: 50.1,
}


def workbook_semantic_row_dimensions_payload(ws: Any) -> list[dict[str, Any]]:
    payload: list[dict[str, Any]] = []
    for index, dimension in sorted(ws.row_dimensions.items()):
        item = workbook_semantic_dimension_payload(dimension)
        payload.append({"index": index, **item})
    return payload


def workbook_semantic_sheet_format_payload(sheet_format: Any) -> dict[str, Any]:
    def size(value: Any) -> float | None:
        return None if value is None else round(float(value), 4)

    return {
        "base_column_width": sheet_format.baseColWidth,
        "default_column_width": size(sheet_format.defaultColWidth),
        "default_row_height": size(sheet_format.defaultRowHeight),
        "custom_height": bool(sheet_format.customHeight),
        "zero_height": bool(sheet_format.zeroHeight),
        "thick_top": bool(sheet_format.thickTop),
        "thick_bottom": bool(sheet_format.thickBottom),
        "outline_level_row": sheet_format.outlineLevelRow,
        "outline_level_column": sheet_format.outlineLevelCol,
    }


def workbook_semantic_page_setup_payload(ws: Any) -> dict[str, Any]:
    setup = ws.page_setup
    payload = {
        attribute: getattr(setup, attribute)
        for attribute in type(setup).__attrs__
    }
    page_setup_properties = ws.sheet_properties.pageSetUpPr
    if page_setup_properties is not None and bool(page_setup_properties.fitToPage):
        # Excel omits explicit 1-by-1 fit values because they are the defaults
        # while preserving sheetPr/fitToPage. Zero and non-default values remain.
        if payload["fitToWidth"] is None:
            payload["fitToWidth"] = 1
        if payload["fitToHeight"] is None:
            payload["fitToHeight"] = 1
    return payload


def workbook_semantic_sheet_properties_payload(properties: Any) -> dict[str, Any]:
    def attributes(value: Any) -> dict[str, Any] | None:
        if value is None:
            return None
        return {
            attribute: getattr(value, attribute)
            for attribute in type(value).__attrs__
        }

    return {
        **{
            attribute: getattr(properties, attribute)
            for attribute in type(properties).__attrs__
        },
        "tab_color": workbook_semantic_color_payload(properties.tabColor),
        "outline_properties": attributes(properties.outlinePr),
        "page_setup_properties": attributes(properties.pageSetUpPr),
    }


def workbook_semantic_sheet_view_payload(
    ws: Any,
    *,
    grouped_view_ids: frozenset[int],
) -> dict[str, Any]:
    freeze_panes = ws.freeze_panes
    if hasattr(freeze_panes, "coordinate"):
        freeze_panes = freeze_panes.coordinate
    views: list[dict[str, Any]] = []
    for view in ws.views.sheetView:
        pane = view.pane
        views.append(
            {
                "workbook_view_id": view.workbookViewId,
                # A single selected tab is transient active-sheet state. Two or
                # more selected tabs enable grouped edits, so that state is hashed.
                "tab_selected": (
                    bool(view.tabSelected)
                    if view.workbookViewId in grouped_view_ids
                    else False
                ),
                "show_formulas": bool(view.showFormulas),
                "show_grid_lines": (
                    True if view.showGridLines is None else bool(view.showGridLines)
                ),
                "show_row_col_headers": (
                    True
                    if view.showRowColHeaders is None
                    else bool(view.showRowColHeaders)
                ),
                "show_zeros": True if view.showZeros is None else bool(view.showZeros),
                "right_to_left": bool(view.rightToLeft),
                "show_ruler": True if view.showRuler is None else bool(view.showRuler),
                "show_outline_symbols": (
                    True
                    if view.showOutlineSymbols is None
                    else bool(view.showOutlineSymbols)
                ),
                "default_grid_color": (
                    True
                    if view.defaultGridColor is None
                    else bool(view.defaultGridColor)
                ),
                "grid_color_id": view.colorId,
                "window_protection": bool(view.windowProtection),
                "show_white_space": (
                    True if view.showWhiteSpace is None else bool(view.showWhiteSpace)
                ),
                "view": view.view or "normal",
                "zoom_scale": view.zoomScale or 100,
                "zoom_scale_normal": view.zoomScaleNormal,
                "zoom_scale_sheet_layout": view.zoomScaleSheetLayoutView,
                "zoom_scale_page_layout": view.zoomScalePageLayoutView,
                "zoom_to_fit": bool(view.zoomToFit),
                # Excel drops an explicit A1 because it is the default scroll origin.
                "top_left_cell": view.topLeftCell or "A1",
                "pane": (
                    {
                        "x_split": pane.xSplit,
                        "y_split": pane.ySplit,
                        "top_left_cell": pane.topLeftCell,
                        "active_pane": pane.activePane,
                        "state": pane.state,
                    }
                    if pane is not None
                    else None
                ),
            }
        )
    return {
        "freeze_panes": freeze_panes,
        "views": views,
    }


def workbook_semantic_defined_name_payload(
    name: str, defined_name: Any
) -> dict[str, Any]:
    return {
        "name": name,
        "text": defined_name.attr_text,
        "type": defined_name.type,
        "comment": defined_name.comment,
        "custom_menu": defined_name.customMenu,
        "description": defined_name.description,
        "help": defined_name.help,
        "status_bar": defined_name.statusBar,
        "local_sheet_id": defined_name.localSheetId,
        "hidden": bool(defined_name.hidden),
        "function": bool(defined_name.function),
        "vb_procedure": bool(defined_name.vbProcedure),
        "xlm": bool(defined_name.xlm),
        "function_group_id": defined_name.functionGroupId,
        "shortcut_key": defined_name.shortcutKey,
        "publish_to_server": bool(defined_name.publishToServer),
        "workbook_parameter": bool(defined_name.workbookParameter),
    }


def workbook_semantic_theme_payload(theme: bytes | str | None) -> dict[str, Any] | None:
    """Canonicalize the full theme because colors, fonts, and charts may reference it."""

    if theme is None:
        return None
    from xml.etree import ElementTree

    try:
        root = ElementTree.fromstring(theme)
    except (ElementTree.ParseError, TypeError, ValueError) as exc:
        raise IntegrityError(f"Workbook theme XML is invalid: {exc}") from exc

    def semantic_xml(node: Any) -> dict[str, Any]:
        text = (
            str(node.text).strip()
            if node.text is not None and str(node.text).strip()
            else None
        )
        return {
            "tag": str(node.tag),
            "attributes": {
                str(key): str(value)
                for key, value in sorted(
                    node.attrib.items(), key=lambda item: str(item[0])
                )
            },
            "text": text,
            "children": [semantic_xml(child) for child in list(node)],
        }

    return semantic_xml(root)


def reject_unsupported_workbook_extensions(path: Path) -> None:
    """Reject extension features that openpyxl would silently discard on load."""

    from xml.etree import ElementTree
    from zipfile import BadZipFile, ZipFile

    try:
        with ZipFile(path) as archive:
            spreadsheet_namespace = (
                "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            )
            allowed_nonstandard_elements = {
                "xl/workbook.xml": {
                    (
                        "http://schemas.openxmlformats.org/markup-compatibility/2006",
                        "AlternateContent",
                    ),
                    (
                        "http://schemas.openxmlformats.org/markup-compatibility/2006",
                        "Choice",
                    ),
                    (
                        "http://schemas.openxmlformats.org/markup-compatibility/2006",
                        "Fallback",
                    ),
                    (
                        "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac",
                        "absPath",
                    ),
                    (
                        "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
                        "revisionPtr",
                    ),
                    (
                        "http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures",
                        "calcFeatures",
                    ),
                    (
                        "http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures",
                        "feature",
                    ),
                },
                "xl/styles.xml": {
                    (
                        "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
                        "slicerStyles",
                    ),
                    (
                        "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main",
                        "timelineStyles",
                    ),
                },
            }
            relationship_namespace = (
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            )
            markup_compatibility_namespace = (
                "http://schemas.openxmlformats.org/markup-compatibility/2006"
            )
            xml_namespace = "http://www.w3.org/XML/1998/namespace"
            excel_2010_namespace = (
                "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
            )
            excel_revision3_namespace = (
                "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"
            )
            transient_revision_namespaces = {
                "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
                "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
                "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
                "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
            }
            excel_calc_features_namespace = (
                "http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures"
            )
            excel_calc_feature_names = (
                "microsoft.com:RD",
                "microsoft.com:Single",
                "microsoft.com:FV",
                "microsoft.com:CNMTM",
                "microsoft.com:LET_WF",
                "microsoft.com:LAMBDA_WF",
                "microsoft.com:ARRAYTEXT_WF",
            )

            def is_excel_default_calc_features(extension_list: Any) -> bool:
                if name != "xl/workbook.xml" or len(extension_list) != 1:
                    return False
                [extension] = list(extension_list)
                if (
                    str(extension.tag).rsplit("}", 1)[-1] != "ext"
                    or extension.attrib
                    != {"uri": "{B58B0392-4F1F-4190-BB64-5DF3571DCE5F}"}
                    or len(extension) != 1
                ):
                    return False
                [features] = list(extension)
                if features.tag != f"{{{excel_calc_features_namespace}}}calcFeatures":
                    return False
                feature_nodes = list(features)
                return (
                    not features.attrib
                    and tuple(
                        node.attrib.get("name")
                        for node in feature_nodes
                        if node.tag
                        == f"{{{excel_calc_features_namespace}}}feature"
                        and set(node.attrib) == {"name"}
                        and not list(node)
                    )
                    == excel_calc_feature_names
                    and len(feature_nodes) == len(excel_calc_feature_names)
                )
            names = [
                name
                for name in archive.namelist()
                if name == "xl/workbook.xml"
                or name == "xl/styles.xml"
                or name.startswith("xl/worksheets/sheet")
                or name.startswith("xl/tables/table")
            ]
            for name in names:
                if not name.endswith(".xml"):
                    continue
                root = ElementTree.fromstring(archive.read(name))
                for node in root.iter():
                    tag = str(node.tag)
                    namespace = (
                        tag.split("}", 1)[0].lstrip("{") if "}" in tag else ""
                    )
                    local_name = tag.rsplit("}", 1)[-1]
                    if (
                        namespace != spreadsheet_namespace
                        and (namespace, local_name)
                        not in allowed_nonstandard_elements.get(name, set())
                    ):
                        raise IntegrityError(
                            "Unsupported XML namespace is not allowed in the "
                            f"protected master workbook ({name})."
                        )
                    for attribute_name, attribute_value in node.attrib.items():
                        attribute_name = str(attribute_name)
                        if "}" not in attribute_name:
                            continue
                        attribute_namespace = (
                            attribute_name.split("}", 1)[0].lstrip("{")
                        )
                        attribute_local_name = attribute_name.rsplit("}", 1)[-1]
                        if attribute_namespace in {
                            relationship_namespace,
                            markup_compatibility_namespace,
                            xml_namespace,
                            *transient_revision_namespaces,
                        }:
                            continue
                        if (
                            attribute_namespace == excel_2010_namespace
                            and attribute_local_name == "knownFonts"
                            and str(attribute_value) == "1"
                            and name == "xl/styles.xml"
                        ):
                            continue
                        if (
                            attribute_namespace == excel_2010_namespace
                            and attribute_local_name == "dyDescent"
                            and str(attribute_value) in {"0.25", "0.3", "0.35"}
                            and name.startswith("xl/worksheets/sheet")
                        ):
                            continue
                        if (
                            attribute_namespace == excel_revision3_namespace
                            and attribute_local_name == "uid"
                            and name.startswith("xl/tables/table")
                            and re.fullmatch(
                                r"\{[0-9A-F]{8}(?:-[0-9A-F]{4}){3}-[0-9A-F]{12}\}",
                                str(attribute_value),
                            )
                        ):
                            continue
                        raise IntegrityError(
                            "Unsupported namespaced XML attribute is not allowed in "
                            f"the protected master workbook ({name})."
                        )
                unsupported_extension = next(
                    (
                        node
                        for node in root.iter()
                        if str(node.tag).rsplit("}", 1)[-1] == "extLst"
                        and list(node)
                        and name != "xl/styles.xml"
                        and not is_excel_default_calc_features(node)
                    ),
                    None,
                )
                if unsupported_extension is not None:
                    raise IntegrityError(
                        "Unsupported workbook extension content is not allowed in "
                        f"the protected master workbook ({name})."
                    )
                if name == "xl/workbook.xml":
                    unsupported_workbook_tags = {
                        "customWorkbookViews",
                        "fileRecoveryPr",
                        "fileSharing",
                        "functionGroups",
                        "smartTagPr",
                        "smartTagTypes",
                        "webPublishing",
                        "webPublishingObjects",
                    }
                    unsupported_workbook_tag = next(
                        (
                            str(node.tag).rsplit("}", 1)[-1]
                            for node in root.iter()
                            if str(node.tag).rsplit("}", 1)[-1]
                            in unsupported_workbook_tags
                        ),
                        None,
                    )
                    if unsupported_workbook_tag is not None:
                        raise IntegrityError(
                            f"Workbook {unsupported_workbook_tag} content is not allowed "
                            "in the protected master workbook."
                        )
                if name.startswith("xl/worksheets/sheet"):
                    if any(
                        str(node.tag).rsplit("}", 1)[-1] == "protectedRanges"
                        for node in root.iter()
                    ):
                        raise IntegrityError(
                            "Protected edit ranges are not allowed in the protected "
                            f"master workbook ({name})."
                        )
                    unsupported_sheet_tags = {
                        "ignoredErrors",
                        "customSheetViews",
                        "webPublishItems",
                    }
                    unsupported_sheet_tag = next(
                        (
                            str(node.tag).rsplit("}", 1)[-1]
                            for node in root.iter()
                            if str(node.tag).rsplit("}", 1)[-1]
                            in unsupported_sheet_tags
                        ),
                        None,
                    )
                    if unsupported_sheet_tag is not None:
                        raise IntegrityError(
                            f"Worksheet {unsupported_sheet_tag} content is not allowed "
                            f"in the protected master workbook ({name})."
                        )
                    for formula in (
                        node
                        for node in root.iter()
                        if str(node.tag).rsplit("}", 1)[-1] == "f"
                    ):
                        attributes = {
                            str(key).rsplit("}", 1)[-1]: value
                            for key, value in formula.attrib.items()
                        }
                        formula_type = attributes.get("t")
                        if formula_type is None:
                            allowed_attributes: set[str] = set()
                        elif formula_type == "array":
                            allowed_attributes = {"t", "ref"}
                        elif formula_type == "dataTable":
                            allowed_attributes = {
                                "t", "ref", "ca", "dt2D", "dtr", "r1", "r2",
                                "del1", "del2",
                            }
                        else:
                            allowed_attributes = set()
                        if (
                            formula_type not in (None, "array", "dataTable")
                            or set(attributes) - allowed_attributes
                        ):
                            raise IntegrityError(
                                "Unsupported worksheet formula attributes are not allowed "
                                f"in the protected master workbook ({name})."
                            )
    except IntegrityError:
        raise
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError) as exc:
        raise IntegrityError(f"Workbook extension XML is invalid: {exc}") from exc


def workbook_raw_properties_payload(path: Path) -> dict[str, Any]:
    from xml.etree import ElementTree
    from zipfile import BadZipFile, ZipFile

    try:
        with ZipFile(path) as archive:
            root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        properties = next(
            (
                node
                for node in root
                if str(node.tag).rsplit("}", 1)[-1] == "workbookPr"
            ),
            None,
        )
        return {
            str(key): str(value)
            for key, value in sorted(
                (properties.attrib.items() if properties is not None else ()),
                key=lambda item: str(item[0]),
            )
        }
    except (BadZipFile, ElementTree.ParseError, KeyError, OSError) as exc:
        raise IntegrityError(f"Workbook properties XML is invalid: {exc}") from exc


def workbook_generated_content_payload(path: Path) -> dict[str, Any]:
    """Build the current substantive view with only proven equivalents normalized."""

    reject_unapproved_workbook_drawings(path)
    reject_unsupported_workbook_extensions(path)
    try:
        wb = load_workbook(
            path,
            data_only=False,
            read_only=False,
            rich_text=True,
        )
    except Exception as exc:
        raise IntegrityError(f"Could not inspect master workbook integrity at {path}: {exc}") from exc
    try:
        if getattr(wb, "_external_links", []):
            raise IntegrityError(
                "External workbook links are not allowed in the protected master workbook."
            )
        approved = approved_management_input_cells(wb)
        digest_cells = workbook_digest_excluded_cells(wb)
        selected_view_count_by_id: dict[int, int] = defaultdict(int)
        for worksheet in wb.worksheets:
            for view in worksheet.views.sheetView:
                if view.tabSelected:
                    selected_view_count_by_id[int(view.workbookViewId)] += 1
        grouped_view_ids = frozenset(
            view_id
            for view_id, count in selected_view_count_by_id.items()
            if count > 1
        )
        sheets: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            scenarios = ws.scenarios
            if (
                getattr(scenarios, "scenario", [])
                or getattr(scenarios, "current", None) is not None
                or getattr(scenarios, "show", None) is not None
                or getattr(scenarios, "sqref", None) is not None
            ):
                raise IntegrityError(
                    "What-if scenarios are not allowed in the protected master "
                    f"workbook ({ws.title!r})."
                )
            if ws._images:
                raise IntegrityError(
                    "Images and non-chart drawings are not allowed in the protected master "
                    f"workbook ({ws.title!r})."
                )
            cells: list[dict[str, Any]] = []
            for cell in sorted(ws._cells.values(), key=lambda item: (item.row, item.column)):
                key = (ws.title, cell.coordinate)
                if isinstance(cell.value, CellRichText):
                    raise IntegrityError(
                        "Rich-text cells are not allowed in the protected master workbook: "
                        f"{ws.title}!{cell.coordinate}."
                    )
                reject_external_formula_reference(
                    cell.value,
                    data_type=cell.data_type,
                    sheet_name=ws.title,
                    coordinate=cell.coordinate,
                )
                is_approved = key in approved
                is_digest_cell = key in digest_cells
                hyperlink = workbook_semantic_hyperlink_payload(
                    cell.hyperlink,
                    sheet_name=ws.title,
                    coordinate=cell.coordinate,
                    cell_value=cell.value,
                )
                if (
                    cell.value is None
                    and hyperlink is None
                    and cell.comment is None
                    and not cell.has_style
                    and not is_approved
                    and not is_digest_cell
                ):
                    continue
                if is_approved or is_digest_cell:
                    if cell.data_type == "f":
                        raise IntegrityError(
                            "Formulas are not allowed in editable or digest-excluded cells: "
                            f"{ws.title}!{cell.coordinate}."
                        )
                    workbook_digest_editable_scalar(
                        cell.value,
                        sheet_name=ws.title,
                        coordinate=cell.coordinate,
                    )
                    value_payload: Any = {
                        "excluded": (
                            "approved-management-scalar"
                            if is_approved
                            else "self-referential-digest"
                        )
                    }
                    value_type = "excluded-scalar"
                else:
                    value_payload = workbook_digest_value(cell.value)
                    value_type = "formula" if cell.data_type == "f" else "scalar"
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "value_type": value_type,
                        "value": value_payload,
                        "data_type": (
                            None
                            if is_approved or is_digest_cell or cell.value is None
                            else cell.data_type
                        ),
                        "quote_prefix": bool(cell.quotePrefix),
                        "pivot_button": bool(cell.pivotButton),
                        # Preserve effective display semantics while excluding only
                        # workbook-internal style IDs/serialization ordering.
                        "style": workbook_semantic_style_payload(cell),
                        "hyperlink": hyperlink,
                        "comment": (
                            {
                                "text": cell.comment.text,
                                "author": cell.comment.author,
                            }
                            if cell.comment is not None
                            else None
                        ),
                        "protection": {
                            "locked": bool(cell.protection.locked),
                            "hidden": bool(cell.protection.hidden),
                        },
                    }
                )
            sheets.append(
                {
                    "name": ws.title,
                    "state": ws.sheet_state,
                    "cells": cells,
                    "tables": sorted(
                        (
                            workbook_semantic_table_payload(table)
                            for table in ws.tables.values()
                        ),
                        key=lambda item: item["name"].casefold(),
                    ),
                    "data_validations": {
                        "disable_prompts": bool(ws.data_validations.disablePrompts),
                        "prompt_x": ws.data_validations.xWindow,
                        "prompt_y": ws.data_validations.yWindow,
                        "items": [
                            workbook_semantic_validation_payload(validation)
                            for validation in ws.data_validations.dataValidation
                        ],
                    },
                    "conditional_formats": (
                        workbook_semantic_conditional_formatting_payload(ws)
                    ),
                    "charts": [
                        workbook_chart_semantic_payload(chart)
                        for chart in ws._charts
                    ],
                    "merged_cells": sorted(
                        str(cell_range) for cell_range in ws.merged_cells.ranges
                    ),
                    "print_area": str(ws.print_area) if ws.print_area else None,
                    "print_title_rows": ws.print_title_rows,
                    "print_title_columns": ws.print_title_cols,
                    "local_defined_names": [
                        workbook_semantic_defined_name_payload(name, defined_name)
                        for name, defined_name in sorted(
                            ws.defined_names.items(),
                            key=lambda item: item[0].casefold(),
                        )
                    ],
                    "row_dimensions": workbook_semantic_row_dimensions_payload(ws),
                    "column_dimensions": workbook_semantic_column_dimensions_payload(ws),
                    "sheet_format": workbook_semantic_sheet_format_payload(
                        ws.sheet_format
                    ),
                    "sheet_properties": workbook_semantic_sheet_properties_payload(
                        ws.sheet_properties
                    ),
                    "sheet_view": workbook_semantic_sheet_view_payload(
                        ws,
                        grouped_view_ids=grouped_view_ids,
                    ),
                    "auto_filter": workbook_digest_serialisable(ws.auto_filter),
                    "page_margins": workbook_digest_serialisable(ws.page_margins),
                    "page_setup": workbook_semantic_page_setup_payload(ws),
                    "print_options": workbook_digest_serialisable(ws.print_options),
                    "header_footer": workbook_digest_serialisable(ws.HeaderFooter),
                    "row_breaks": workbook_digest_serialisable(ws.row_breaks),
                    "column_breaks": workbook_digest_serialisable(ws.col_breaks),
                    "unlocked_cells": sorted(
                        cell.coordinate
                        for cell in ws._cells.values()
                        if cell.protection.locked is False
                    ),
                    "protection": workbook_digest_serialisable(ws.protection),
                }
            )
        return {
            "digest_scheme": WORKBOOK_DIGEST_SCHEME,
            "date_epoch": wb.epoch.isoformat(),
            "workbook_properties": workbook_raw_properties_payload(path),
            "workbook_views": [
                {
                    "visibility": view.visibility,
                    "minimized": bool(view.minimized),
                    "show_horizontal_scroll": view.showHorizontalScroll is not False,
                    "show_vertical_scroll": view.showVerticalScroll is not False,
                    "show_sheet_tabs": view.showSheetTabs is not False,
                    "tab_ratio": view.tabRatio,
                    "auto_filter_date_grouping": view.autoFilterDateGrouping is not False,
                }
                for view in wb.views
            ],
            "sheet_order": list(wb.sheetnames),
            "calculation": workbook_semantic_calculation_payload(wb.calculation),
            "theme": workbook_semantic_theme_payload(wb.loaded_theme),
            "indexed_colors": list(getattr(wb, "_colors", ()) or ()),
            "table_styles": workbook_digest_serialisable(
                getattr(wb, "_table_styles", None)
            ),
            "differential_styles": [
                json.loads(serialized)
                for serialized in sorted(
                    {
                        json.dumps(
                            workbook_semantic_differential_style_payload(style),
                            sort_keys=True,
                            separators=(",", ":"),
                        )
                        for style in getattr(
                            wb._differential_styles, "styles", []
                        )
                    }
                )
            ],
            "raw_chart_parts": workbook_raw_chart_parts_payload(path),
            "chart_frames": workbook_chart_frame_payload(path),
            "workbook_protection": workbook_semantic_workbook_protection_payload(
                wb.security
            ),
            "defined_names": [
                workbook_semantic_defined_name_payload(name, defined_name)
                for name, defined_name in sorted(
                    wb.defined_names.items(), key=lambda item: item[0].casefold()
                )
            ],
            "sheets": sheets,
        }
    finally:
        wb.close()


def workbook_generated_content_sha256(path: Path) -> str:
    return canonical_json_sha256(workbook_generated_content_payload(path))


def stamped_workbook_digest(wb: Workbook) -> str | None:
    if "Run Notes" not in wb.sheetnames:
        return None
    ws = wb["Run Notes"]
    rows = [
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=1).value == RUN_NOTES_DIGEST_LABEL
    ]
    if len(rows) > 1:
        raise IntegrityError("Run Notes contains a duplicated generated-content digest field.")
    if not rows:
        return None
    value = ws.cell(row=rows[0], column=2).value
    return str(value).strip().lower() if value else None


def stamped_workbook_digest_scheme(wb: Workbook) -> str | None:
    if "Run Notes" not in wb.sheetnames:
        return None
    ws = wb["Run Notes"]
    rows = [
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=1).value == "Digest Scheme"
    ]
    if len(rows) > 1:
        raise IntegrityError("Run Notes contains a duplicated digest scheme field.")
    if not rows:
        return None
    value = ws.cell(row=rows[0], column=2).value
    return str(value).strip() if value is not None else ""


def stamped_workbook_protection_contract(wb: Workbook) -> str | None:
    if "Run Notes" not in wb.sheetnames:
        return None
    ws = wb["Run Notes"]
    rows = [
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=1).value == WORKBOOK_PROTECTION_CONTRACT_LABEL
    ]
    if len(rows) > 1:
        raise IntegrityError("Run Notes contains a duplicated protection contract marker.")
    if not rows:
        return None
    value = ws.cell(row=rows[0], column=2).value
    return str(value).strip() if value is not None else ""


def stamp_generated_content_digest(path: Path) -> str:
    wb = load_workbook(path, data_only=False)
    try:
        if "Run Notes" not in wb.sheetnames:
            raise IntegrityError("Generated master workbook is missing Run Notes.")
        ws = wb["Run Notes"]
        scheme_rows = [
            row
            for row in range(1, ws.max_row + 1)
            if ws.cell(row=row, column=1).value == "Digest Scheme"
        ]
        if len(scheme_rows) > 1:
            raise IntegrityError("Run Notes contains a duplicated digest scheme field.")
        if scheme_rows:
            ws.cell(row=scheme_rows[0], column=2, value=WORKBOOK_DIGEST_SCHEME)
        else:
            ws.append(["Digest Scheme", WORKBOOK_DIGEST_SCHEME])
        wb.save(path)
    finally:
        wb.close()

    digest = workbook_generated_content_sha256(path)
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb["Run Notes"]
        digest_rows = [
            row
            for row in range(1, ws.max_row + 1)
            if ws.cell(row=row, column=1).value == RUN_NOTES_DIGEST_LABEL
        ]
        if not digest_rows:
            raise IntegrityError("Run Notes is missing the generated-content digest field.")
        if len(digest_rows) > 1:
            raise IntegrityError("Run Notes contains a duplicated generated-content digest field.")
        ws.cell(row=digest_rows[0], column=2, value=digest)
        wb.save(path)
    finally:
        wb.close()
    if workbook_generated_content_sha256(path) != digest:
        raise IntegrityError("Generated master workbook digest was not stable after save/reload.")
    return digest


def require_stop_style_list_validation(
    ws: Any,
    *,
    formula1: str,
    label: str,
    allow_blank: bool,
    sqref: str,
) -> None:
    matches = [
        validation
        for validation in ws.data_validations.dataValidation
        if validation.type == "list"
        and workbook_semantic_validation_formula(validation.formula1)
        == workbook_semantic_validation_formula(formula1)
    ]
    if len(matches) != 1:
        raise IntegrityError(f"The {label} list validation is missing or duplicated.")
    validation = matches[0]
    if (
        validation.showErrorMessage is not True
        or validation.errorStyle not in (None, "stop")
        or bool(validation.allowBlank) is not allow_blank
        or str(validation.sqref) != sqref
        or not validation.errorTitle
        or not validation.error
    ):
        raise IntegrityError(
            f"The {label} list validation must block invalid pasted or typed values."
        )


def expected_management_list_validations(
    wb: Workbook,
) -> list[tuple[str, str, bool, str, str]]:
    def column_range(column: str, first_row: int, last_row: int) -> str:
        first = f"{column}{first_row}"
        return first if first_row == last_row else f"{first}:{column}{last_row}"

    setup = (
        wb[MANAGEMENT_CENTER_SHEET]
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames
        and OWNER_ROSTER_TABLE_NAME in wb[MANAGEMENT_CENTER_SHEET].tables
        else wb["Management Setup"]
    )
    min_col, min_row, max_col, max_row = range_boundaries(
        setup.tables[OWNER_ROSTER_TABLE_NAME].ref
    )
    if max_col - min_col != 1:
        raise IntegrityError("The Owner Roster table has an unexpected shape.")
    roster_headers = {
        setup.cell(row=min_row, column=column).value: column
        for column in range(min_col, max_col + 1)
    }
    if tuple(roster_headers) != OWNER_ROSTER_HEADERS:
        raise IntegrityError("The Owner Roster table has unexpected headers.")
    active_column = get_column_letter(roster_headers["Active"])
    expected = [
        (
            setup.title,
            '"Yes,No"',
            True,
            f"{active_column}{min_row + 1}:{active_column}{max_row}",
            "Owner Roster Active",
        )
    ]
    action_board = (
        wb[MANAGEMENT_CENTER_SHEET]
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames
        else wb["Action Board"]
    )
    if "ActionBoardTable" in action_board.tables:
        action_min_col, action_min_row, action_max_col, action_max_row = range_boundaries(
            action_board.tables["ActionBoardTable"].ref
        )
        if action_min_col != 1 or action_max_col not in {
            len(ACTION_HEADERS),
            len(LEGACY_ACTION_HEADERS_V1),
        }:
            raise IntegrityError("The Action Board table has an unexpected shape.")
        expected_action_headers = (
            ACTION_HEADERS
            if action_max_col == len(ACTION_HEADERS)
            else LEGACY_ACTION_HEADERS_V1
        )
        actual_action_headers = [
            action_board.cell(row=action_min_row, column=column).value
            for column in range(action_min_col, action_max_col + 1)
        ]
        if actual_action_headers != expected_action_headers:
            raise IntegrityError(
                "The Action Board table headers do not match a supported schema."
            )
        action_status_choices = (
            ACTION_STATUS_CHOICES
            if action_max_col == len(ACTION_HEADERS)
            else LEGACY_ACTION_STATUS_CHOICES
        )
        action_columns = {
            header: get_column_letter(action_min_col + offset)
            for offset, header in enumerate(actual_action_headers)
        }
        expected.extend(
            [
                (
                    action_board.title,
                    f'"{",".join(action_status_choices)}"',
                    False,
                    column_range(
                        action_columns["Status"], action_min_row + 1, action_max_row
                    ),
                    "Action Board status",
                ),
            ]
        )
        if action_max_col == len(ACTION_HEADERS):
            expected.extend(
                [
                    (
                        action_board.title,
                        f"={OWNER_ROSTER_DEFINED_NAME}",
                        True,
                        (
                            column_range(
                                action_columns["Owner"],
                                action_min_row + 1,
                                action_max_row,
                            )
                            + " "
                            + column_range(
                                action_columns["Reviewed By"],
                                action_min_row + 1,
                                action_max_row,
                            )
                        ),
                        "Action Board owner and reviewer",
                    ),
                    (
                        action_board.title,
                        f'"{",".join(REVIEW_DISPOSITION_CHOICES)}"',
                        False,
                        column_range(
                            action_columns["Review Disposition"],
                            action_min_row + 1,
                            action_max_row,
                        ),
                        "Action Board review disposition",
                    ),
                ]
            )
        else:
            expected.append(
                (
                    action_board.title,
                    f"={OWNER_ROSTER_DEFINED_NAME}",
                    True,
                    column_range(
                        action_columns["Owner"], action_min_row + 1, action_max_row
                    ),
                    "Action Board owner",
                )
            )
    elif records_from_sheet(action_board, "Action ID"):
        raise IntegrityError("The Action Board table is missing.")
    return expected


def require_exact_validation_count(
    wb: Workbook,
    expected: list[tuple[str, str, bool, str, str]],
    *,
    contract_label: str,
) -> None:
    actual_count = sum(
        len(ws.data_validations.dataValidation) for ws in wb.worksheets
    )
    if actual_count != len(expected):
        raise IntegrityError(
            f"The {contract_label} workbook contains an unexpected number of data validations."
        )


def validate_management_workbook_controls(
    wb: Workbook,
    *,
    protect_objects_and_scenarios: bool,
    visible_sheets: Iterable[str] | None = None,
) -> tuple[Any, Any]:
    """Validate controls shared by strict and manifest-bound pre-contract workbooks."""

    visible_sheet_names = tuple(
        VISIBLE_MANAGEMENT_SHEETS if visible_sheets is None else visible_sheets
    )
    missing = [name for name in visible_sheet_names if name not in wb.sheetnames]
    if missing:
        raise IntegrityError(
            f"Generated master workbook is missing required sheets: {', '.join(missing)}"
        )
    if wb.security is None or not wb.security.lockStructure:
        raise IntegrityError("Generated master workbook structure protection is missing.")
    if not wb.security.workbookPassword:
        raise IntegrityError("Generated master workbook structure password is missing.")
    if (
        wb.calculation.calcMode not in (None, "auto")
        or wb.calculation.fullCalcOnLoad is False
        or wb.calculation.forceFullCalc is False
    ):
        raise IntegrityError(
            "Generated master workbook has explicit unsafe calculation settings."
        )
    if (
        wb.calculation.calcMode is None
        or wb.calculation.fullCalcOnLoad is None
        or wb.calculation.forceFullCalc is None
    ):
        warnings.warn(
            "Excel omitted default workbook calculation flags; they will be repaired on the "
            "next successful generated save.",
            UserWarning,
            stacklevel=2,
        )
    approved = approved_management_input_cells(wb)
    actual_unlocked = {
        (ws.title, cell.coordinate)
        for ws in wb.worksheets
        for cell in ws._cells.values()
        if cell.protection.locked is False
    }
    if actual_unlocked != approved:
        unexpected = sorted(actual_unlocked - approved)
        missing_unlocked = sorted(approved - actual_unlocked)
        detail = unexpected[:1] or missing_unlocked[:1]
        raise IntegrityError(
            "Generated master workbook editable-cell protection does not match the "
            f"approved allowlist ({detail[0] if detail else 'unknown mismatch'})."
        )
    for ws in wb.worksheets:
        expected_state = "visible" if ws.title in visible_sheet_names else "veryHidden"
        if ws.sheet_state != expected_state:
            raise IntegrityError(
                f"Worksheet {ws.title!r} has state {ws.sheet_state!r}; "
                f"expected {expected_state!r}."
            )
        if not ws.protection.sheet or not ws.protection.password:
            raise IntegrityError(f"Worksheet {ws.title!r} is not password-protected.")
        if (
            ws.protection.objects is not protect_objects_and_scenarios
            or ws.protection.scenarios is not protect_objects_and_scenarios
        ):
            expected = "protect" if protect_objects_and_scenarios else "leave unprotected"
            raise IntegrityError(
                f"Worksheet {ws.title!r} does not match the pre-approved protection "
                f"contract: it must {expected} drawing objects and scenarios."
            )
    setup = (
        wb[MANAGEMENT_CENTER_SHEET]
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames
        and "ManagementTargets" in wb[MANAGEMENT_CENTER_SHEET].tables
        and OWNER_ROSTER_TABLE_NAME in wb[MANAGEMENT_CENTER_SHEET].tables
        else wb["Management Setup"]
    )
    if "ManagementTargets" not in setup.tables or OWNER_ROSTER_TABLE_NAME not in setup.tables:
        raise IntegrityError("The management input surface is missing a required protected table.")
    if OWNER_ROSTER_DEFINED_NAME not in wb.defined_names:
        raise IntegrityError("The active-owner workbook name is missing.")
    owner_roster = owner_roster_from_sheet(setup)
    action_board = (
        wb[MANAGEMENT_CENTER_SHEET]
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames
        else wb["Action Board"]
    )
    action_records = (
        records_from_table(action_board, "ActionBoardTable")
        if "ActionBoardTable" in action_board.tables
        else records_from_sheet(action_board, "Action ID")
    )
    validate_action_board_records(
        action_records,
        allowed_reviewers=active_owner_names(owner_roster),
    )
    return setup, action_board


def require_pre_contract_list_validations(wb: Workbook) -> None:
    """Require the exact validation shape emitted immediately before PR #10."""

    expected = expected_management_list_validations(wb)
    actual = [
        (ws.title, validation)
        for ws in wb.worksheets
        for validation in ws.data_validations.dataValidation
    ]
    require_exact_validation_count(wb, expected, contract_label="pre-contract")
    for sheet_name, formula1, allow_blank, sqref, label in expected:
        matches = [
            validation
            for actual_sheet, validation in actual
            if actual_sheet == sheet_name
            and validation.type == "list"
            and validation.formula1 == formula1
        ]
        if len(matches) != 1:
            raise IntegrityError(f"The pre-contract {label} validation is missing or duplicated.")
        validation = matches[0]
        if (
            str(validation.sqref) != sqref
            or bool(validation.allowBlank) is not allow_blank
            or validation.showErrorMessage is not False
            or validation.errorStyle is not None
            or validation.errorTitle is not None
            or validation.error is not None
        ):
            raise IntegrityError(
                f"The pre-contract {label} validation does not match the approved legacy shape."
            )


def workbook_color_suffix(color: Any) -> str | None:
    value = getattr(color, "rgb", None)
    return str(value)[-6:].upper() if value else None


def require_current_workbook_usability_contract(
    wb: Workbook,
    *,
    previous_v2: bool = False,
    previous_tab_navigation: bool = False,
    previous_simplification: bool = False,
    visible_sheets: Iterable[str] | None = None,
    navigation_links: tuple[tuple[str, str], ...] | None = None,
) -> None:
    """Require the exact current or immediately preceding usability contract."""

    if not wb.sheetnames or wb.sheetnames[0] != "How to Use":
        raise IntegrityError("How to Use must be the first worksheet.")
    guide = wb["How to Use"]
    if guide.sheet_state != "visible":
        raise IntegrityError("How to Use must remain visible.")
    if any(cell.protection.locked is False for cell in guide._cells.values()):
        raise IntegrityError("How to Use must contain no editable cells.")
    if guide._charts or guide._images:
        raise IntegrityError("How to Use must not contain drawings.")
    if any(cell.comment is not None for cell in guide._cells.values()):
        raise IntegrityError("How to Use must not contain comments.")
    if (
        guide.freeze_panes != "A3"
        or guide.sheet_view.zoomScale != 85
        or any(
            guide.column_dimensions[get_column_letter(column)].width != 13
            for column in range(1, 13)
        )
    ):
        warnings.warn(
            "Excel rewrote How to Use view or sizing metadata; the generated workbook "
            "will repair it on the next successful run.",
            UserWarning,
            stacklevel=2,
        )
    guide_text = {
        str(cell.value)
        for cell in guide._cells.values()
        if cell.value not in (None, "")
    }
    required_headings = (
        PRE_SIMPLIFICATION_HOW_TO_USE_SECTION_HEADINGS
        if previous_simplification
        else HOW_TO_USE_SECTION_HEADINGS
    )
    missing_headings = [
        heading for heading in required_headings if heading not in guide_text
    ]
    if missing_headings:
        raise IntegrityError(
            "How to Use is missing required section headings: "
            + ", ".join(missing_headings)
        )
    methodology_version = (
        PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION
        if previous_v2
        else MANAGEMENT_METHODOLOGY_VERSION
    )
    rate_guidance = (
        PREVIOUS_RATE_OF_SALE_GUIDANCE
        if previous_v2
        else "Rate of Sale is opportunities divided by qualifying sales"
    )
    ticket_guidance = (
        PREVIOUS_TICKET_TIME_GUIDANCE
        if previous_v2
        else "Ticket Time is weighted only by complete Check Count coverage"
    )
    required_guide_phrases = (
        methodology_version,
        "never be the sole or determinative basis",
        "Do not infer protected characteristics from names.",
        "ExternalCheckRequired",
        "D — Status",
        "E — Owner",
        "F — Due Date",
        "N — Context Notes",
        "U — Review Disposition",
        "V — Reviewed By",
        "W — Review Date",
        rate_guidance,
        ticket_guidance,
    )
    flattened_guide_text = "\n".join(sorted(guide_text))
    missing_phrases = [
        phrase for phrase in required_guide_phrases if phrase not in flattened_guide_text
    ]
    missing_phrases.extend(
        choice
        for choice in (*ACTION_STATUS_CHOICES, *REVIEW_DISPOSITION_CHOICES)
        if choice not in flattened_guide_text
    )
    if missing_phrases:
        raise IntegrityError(
            "How to Use is missing required operating guidance: "
            + ", ".join(missing_phrases)
        )
    if "NotChecked" in flattened_guide_text:
        raise IntegrityError(
            "How to Use contains obsolete readiness terminology."
        )

    if previous_v2 or previous_tab_navigation:
        require_legacy_tab_navigation_contract(
            wb,
            previous_v2=previous_v2,
            visible_sheets=visible_sheets,
            navigation_links=navigation_links,
        )
    else:
        require_management_menu_contract(
            wb,
            visible_sheets=visible_sheets,
            navigation_links=navigation_links,
        )


def verify_recorded_workbook_digest(
    path: Path,
    *,
    stamped_digest: str | None,
    stamped_scheme: str | None,
    expected_digest: str | None,
    expected_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
) -> str:
    """Verify v4 directly or bridge a manifest-pinned v2/v3 workbook once."""

    required_digest = str(expected_digest or stamped_digest or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", required_digest):
        raise IntegrityError("The master workbook requires a valid recorded digest.")
    if stamped_digest != required_digest:
        raise IntegrityError(
            "Master workbook stamped digest does not match the manifest-recorded digest."
        )
    scheme = expected_scheme or stamped_scheme or LEGACY_WORKBOOK_DIGEST_SCHEME
    if (
        expected_scheme is not None
        and stamped_scheme is not None
        and expected_scheme != stamped_scheme
    ):
        raise IntegrityError(
            "Master workbook digest scheme does not match the manifest-recorded scheme."
        )
    if scheme == WORKBOOK_DIGEST_SCHEME:
        actual_digest = workbook_generated_content_sha256(path)
        if actual_digest != required_digest:
            raise IntegrityError(
                "Master workbook generated-content verification failed: "
                "substantive-content verification "
                f"expected {required_digest}; actual {actual_digest}. "
                "No outputs were replaced and no active source files were moved."
            )
        return actual_digest
    if scheme not in {
        LEGACY_WORKBOOK_DIGEST_SCHEME,
        PREVIOUS_WORKBOOK_DIGEST_SCHEME,
    }:
        raise IntegrityError(f"Unsupported master workbook digest scheme: {scheme!r}.")

    if scheme == PREVIOUS_WORKBOOK_DIGEST_SCHEME:
        if legacy_reference_path is None:
            raise IntegrityError(
                "The prior v3 master workbook requires its manifest-pinned archived "
                "reference before it can be upgraded to the v4 digest contract."
            )
        reference = load_workbook(
            legacy_reference_path,
            data_only=False,
            read_only=False,
        )
        try:
            reference_digest = stamped_workbook_digest(reference)
            reference_scheme = stamped_workbook_digest_scheme(reference)
        finally:
            reference.close()
        if (
            reference_digest != required_digest
            or reference_scheme != PREVIOUS_WORKBOOK_DIGEST_SCHEME
        ):
            raise IntegrityError(
                "The manifest-pinned archived master does not match its prior v3 "
                "digest contract."
            )
        actual_substantive = workbook_generated_content_sha256(path)
        reference_substantive = workbook_generated_content_sha256(
            legacy_reference_path
        )
        if actual_substantive != reference_substantive:
            raise IntegrityError(
                "Master workbook generated-content verification failed: the current "
                "v4 substantive view differs from the manifest-pinned prior v3 "
                "master. Restore the archived generated workbook before rerunning."
            )
        warnings.warn(
            "The manifest-pinned prior v3 master matches under the v4 substantive "
            "contract and will be regenerated on the next successful run.",
            UserWarning,
            stacklevel=2,
        )
        return required_digest

    actual_legacy_digest = workbook_metadata_sha256(path)
    if actual_legacy_digest == required_digest:
        return required_digest
    if legacy_reference_path is None:
        raise IntegrityError(
            "Master workbook generated-content verification failed under the legacy v2 "
            f"contract: expected {required_digest}; actual {actual_legacy_digest}."
        )
    reference_legacy_digest = workbook_metadata_sha256(legacy_reference_path)
    if reference_legacy_digest != required_digest:
        raise IntegrityError(
            "The manifest-pinned archived master does not match its legacy v2 digest."
        )
    actual_substantive = workbook_generated_content_sha256(path)
    reference_substantive = workbook_generated_content_sha256(legacy_reference_path)
    if actual_substantive != reference_substantive:
        raise IntegrityError(
            "Master workbook generated-content verification failed: "
            "substantive-content verification failed against the "
            "manifest-pinned archived master. Restore the archived generated workbook "
            "before rerunning."
        )
    warnings.warn(
        "Master workbook metadata differs under the legacy v2 digest; its substantive "
        "content matches the manifest-pinned archived master.",
        UserWarning,
        stacklevel=2,
    )
    return required_digest


def validate_management_workbook(
    path: Path,
    expected_digest: str | None = None,
    *,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
    previous_v2_usability: bool = False,
    previous_tab_navigation: bool = False,
    previous_simplification: bool = False,
    visible_sheets: Iterable[str] | None = None,
    navigation_links: tuple[tuple[str, str], ...] | None = None,
) -> str:
    """Validate protection, visibility, editable cells, tables, and digest."""
    reject_unapproved_workbook_drawings(path)
    wb = load_workbook(path, data_only=False)
    try:
        protection_contract = stamped_workbook_protection_contract(wb)
        if protection_contract != WORKBOOK_PROTECTION_CONTRACT:
            raise IntegrityError(
                "Generated master workbook has no supported protection contract marker."
            )
        setup, action_board = validate_management_workbook_controls(
            wb,
            protect_objects_and_scenarios=True,
            visible_sheets=visible_sheets,
        )
        visible_sheet_names = tuple(
            VISIBLE_MANAGEMENT_SHEETS
            if visible_sheets is None
            else visible_sheets
        )
        if "How to Use" in visible_sheet_names:
            require_current_workbook_usability_contract(
                wb,
                previous_v2=previous_v2_usability,
                previous_tab_navigation=previous_tab_navigation,
                previous_simplification=previous_simplification,
                visible_sheets=visible_sheet_names,
                navigation_links=navigation_links,
            )
        expected_validations = expected_management_list_validations(wb)
        require_exact_validation_count(
            wb, expected_validations, contract_label="strict protection-contract"
        )
        for sheet_name, formula1, allow_blank, sqref, label in expected_validations:
            require_stop_style_list_validation(
                wb[sheet_name],
                formula1=formula1,
                label=label,
                allow_blank=allow_blank,
                sqref=sqref,
            )
        stamped = stamped_workbook_digest(wb)
        stamped_scheme = stamped_workbook_digest_scheme(wb)
    finally:
        wb.close()
    return verify_recorded_workbook_digest(
        path,
        stamped_digest=stamped,
        stamped_scheme=stamped_scheme,
        expected_digest=expected_digest,
        expected_scheme=expected_digest_scheme,
        legacy_reference_path=legacy_reference_path,
    )


def validate_pre_guide_management_workbook(
    path: Path,
    expected_digest: str,
    *,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
) -> str:
    """Verify the exact protected v0.3.1 workbook before adding the guide."""

    required_digest = str(expected_digest).strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", required_digest):
        raise IntegrityError("The pre-guide workbook requires a valid recorded digest.")
    reject_unapproved_workbook_drawings(path)
    wb = load_workbook(path, data_only=False)
    try:
        if stamped_workbook_protection_contract(wb) != WORKBOOK_PROTECTION_CONTRACT:
            raise IntegrityError("The pre-guide workbook protection contract is unsupported.")
        validate_management_workbook_controls(
            wb,
            protect_objects_and_scenarios=True,
            visible_sheets=PRE_GUIDE_VISIBLE_MANAGEMENT_SHEETS,
        )
        expected_validations = expected_management_list_validations(wb)
        require_exact_validation_count(
            wb, expected_validations, contract_label="pre-guide v0.3.1 schema"
        )
        for sheet_name, formula1, allow_blank, sqref, label in expected_validations:
            require_stop_style_list_validation(
                wb[sheet_name],
                formula1=formula1,
                label=label,
                allow_blank=allow_blank,
                sqref=sqref,
            )
        stamped = stamped_workbook_digest(wb)
        stamped_scheme = stamped_workbook_digest_scheme(wb)
    finally:
        wb.close()
    return verify_recorded_workbook_digest(
        path,
        stamped_digest=stamped,
        stamped_scheme=stamped_scheme,
        expected_digest=required_digest,
        expected_scheme=expected_digest_scheme,
        legacy_reference_path=legacy_reference_path,
    )


def validate_v1_action_focus_workbook(
    path: Path,
    expected_digest: str,
    *,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
) -> str:
    """Verify the exact v0.2.x protected workbook before one-way v2 migration."""

    required_digest = str(expected_digest).strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", required_digest):
        raise IntegrityError("The v1 workbook requires a valid recorded digest.")
    reject_unapproved_workbook_drawings(path)
    wb = load_workbook(path, data_only=False)
    try:
        if stamped_workbook_protection_contract(wb) != WORKBOOK_PROTECTION_CONTRACT:
            raise IntegrityError("The v1 workbook protection contract is unsupported.")
        validate_management_workbook_controls(
            wb,
            protect_objects_and_scenarios=True,
            visible_sheets=LEGACY_V1_VISIBLE_MANAGEMENT_SHEETS,
        )
        expected_validations = expected_management_list_validations(wb)
        require_exact_validation_count(
            wb, expected_validations, contract_label="v1 action-focus schema"
        )
        for sheet_name, formula1, allow_blank, sqref, label in expected_validations:
            require_stop_style_list_validation(
                wb[sheet_name],
                formula1=formula1,
                label=label,
                allow_blank=allow_blank,
                sqref=sqref,
            )
        stamped = stamped_workbook_digest(wb)
        stamped_scheme = stamped_workbook_digest_scheme(wb)
    finally:
        wb.close()
    return verify_recorded_workbook_digest(
        path,
        stamped_digest=stamped,
        stamped_scheme=stamped_scheme,
        expected_digest=required_digest,
        expected_scheme=expected_digest_scheme,
        legacy_reference_path=legacy_reference_path,
    )


def validate_previous_action_schema_workbook(
    path: Path,
    expected_digest: str,
    *,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
) -> str:
    """Verify the exact manifest-bound workbook emitted before Action Focus."""

    required_digest = str(expected_digest).strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", required_digest):
        raise IntegrityError(
            "The previous workbook schema requires a valid recorded digest."
        )
    reject_unapproved_workbook_drawings(path)
    wb = load_workbook(path, data_only=False)
    try:
        if stamped_workbook_protection_contract(wb) != WORKBOOK_PROTECTION_CONTRACT:
            raise IntegrityError(
                "The previous workbook schema has an unsupported protection contract."
            )
        validate_management_workbook_controls(
            wb,
            protect_objects_and_scenarios=True,
            visible_sheets=PRE_ACTION_FOCUS_VISIBLE_MANAGEMENT_SHEETS,
        )
        expected_validations = expected_management_list_validations(wb)
        require_exact_validation_count(
            wb, expected_validations, contract_label="previous action schema"
        )
        for sheet_name, formula1, allow_blank, sqref, label in expected_validations:
            require_stop_style_list_validation(
                wb[sheet_name],
                formula1=formula1,
                label=label,
                allow_blank=allow_blank,
                sqref=sqref,
            )
        stamped = stamped_workbook_digest(wb)
        stamped_scheme = stamped_workbook_digest_scheme(wb)
    finally:
        wb.close()
    return verify_recorded_workbook_digest(
        path,
        stamped_digest=stamped,
        stamped_scheme=stamped_scheme,
        expected_digest=required_digest,
        expected_scheme=expected_digest_scheme,
        legacy_reference_path=legacy_reference_path,
    )


def validate_pre_contract_management_workbook(
    path: Path,
    expected_digest: str,
    *,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
) -> str:
    """Validate only the exact PR #9 workbook state pinned by an integrity manifest."""

    required_digest = str(expected_digest).strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", required_digest):
        raise IntegrityError("The pre-contract workbook requires a valid manifest digest.")
    reject_unapproved_workbook_drawings(path)
    wb = load_workbook(path, data_only=False)
    try:
        if stamped_workbook_protection_contract(wb) is not None:
            raise IntegrityError(
                "The compatibility verifier only accepts a markerless pre-contract workbook."
            )
        has_action_focus_schema = all(
            name in wb.sheetnames
            for name in ("Action Focus", "Evidence Detail")
        )
        legacy_visible_sheets = (
            LEGACY_V1_VISIBLE_MANAGEMENT_SHEETS
            if (
                has_action_focus_schema
                and "Rising & Falling Stars" in wb.sheetnames
                and "Recent Movement Signals" not in wb.sheetnames
            )
            else PRE_GUIDE_VISIBLE_MANAGEMENT_SHEETS
            if has_action_focus_schema
            else PRE_ACTION_FOCUS_VISIBLE_MANAGEMENT_SHEETS
        )
        validate_management_workbook_controls(
            wb,
            protect_objects_and_scenarios=False,
            visible_sheets=legacy_visible_sheets,
        )
        require_pre_contract_list_validations(wb)
        stamped = stamped_workbook_digest(wb)
        stamped_scheme = stamped_workbook_digest_scheme(wb)
    finally:
        wb.close()
    return verify_recorded_workbook_digest(
        path,
        stamped_digest=stamped,
        stamped_scheme=stamped_scheme,
        expected_digest=required_digest,
        expected_scheme=expected_digest_scheme,
        legacy_reference_path=legacy_reference_path,
    )


def pre_contract_management_workbook(path: Path) -> bool:
    wb = load_workbook(path, data_only=False)
    try:
        return (
            "Management Setup" in wb.sheetnames
            and OWNER_ROSTER_TABLE_NAME in wb["Management Setup"].tables
            and stamped_workbook_digest(wb) is not None
            and stamped_workbook_protection_contract(wb) is None
        )
    finally:
        wb.close()


def workbook_uses_previous_v2_usability(wb: Workbook) -> bool:
    if (
        stamped_workbook_protection_contract(wb)
        != WORKBOOK_PROTECTION_CONTRACT
        or "How to Use" not in wb.sheetnames
    ):
        return False
    guide_text = "\n".join(
        str(cell.value)
        for cell in wb["How to Use"]._cells.values()
        if cell.value not in (None, "")
    )
    return (
        PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION in guide_text
        and MANAGEMENT_METHODOLOGY_VERSION not in guide_text
    )


def workbook_uses_legacy_multi_link_navigation(wb: Workbook) -> bool:
    """Identify the exact manifest-bound v3 layout emitted before the menu bar."""

    if (
        stamped_workbook_protection_contract(wb)
        != WORKBOOK_PROTECTION_CONTRACT
        or "How to Use" not in wb.sheetnames
        or any(
            name not in wb.sheetnames
            for name in PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
        )
    ):
        return False
    guide_text = "\n".join(
        str(cell.value)
        for cell in wb["How to Use"]._cells.values()
        if cell.value not in (None, "")
    )
    if (
        MANAGEMENT_METHODOLOGY_VERSION not in guide_text
        or PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION in guide_text
    ):
        return False
    expected_labels = [
        label for label, _ in PRE_CONSOLIDATION_NAVIGATION_LINKS
    ]
    return all(
        [
            ws.cell(row=2, column=column).value
            for column in management_navigation_columns(
                ws, PRE_CONSOLIDATION_NAVIGATION_LINKS
            )
        ]
        == expected_labels
        for ws in (
            wb[name] for name in PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
        )
    )


def workbook_uses_pre_consolidation_layout(wb: Workbook) -> bool:
    """Identify the protected menu layout before action/trend tab consolidation."""

    if (
        stamped_workbook_protection_contract(wb)
        != WORKBOOK_PROTECTION_CONTRACT
        or "How to Use" not in wb.sheetnames
        or "Team Trends" in wb.sheetnames
        or any(
            name not in wb.sheetnames
            for name in PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
        )
    ):
        return False
    guide_text = "\n".join(
        str(cell.value)
        for cell in wb["How to Use"]._cells.values()
        if cell.value not in (None, "")
    )
    if (
        MANAGEMENT_METHODOLOGY_VERSION not in guide_text
        or PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION in guide_text
    ):
        return False
    for ws in (
        wb[name] for name in PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
    ):
        start_column, end_column = management_menu_bounds(ws)
        if (
            ws.sheet_state != "visible"
            or ws.cell(row=2, column=start_column).value != MANAGEMENT_MENU_LABEL
            or not workbook_hyperlink_matches(
                ws.cell(row=2, column=start_column).hyperlink,
                MANAGEMENT_MENU_TARGET,
                sheet_name=ws.title,
                coordinate=ws.cell(row=2, column=start_column).coordinate,
            )
            or str(
                next(
                    (
                        merged
                        for merged in ws.merged_cells.ranges
                        if merged.min_row <= 2 <= merged.max_row
                    ),
                    "",
                )
            )
            != (
                f"{get_column_letter(start_column)}2:"
                f"{get_column_letter(end_column)}2"
            )
        ):
            return False
    return True


def workbook_uses_pre_performance_layout(wb: Workbook) -> bool:
    """Identify the manifest-bound v0.4 layout before descriptive analytics."""

    if (
        stamped_workbook_protection_contract(wb)
        != WORKBOOK_PROTECTION_CONTRACT
        or "How to Use" not in wb.sheetnames
        or any(
            name in wb.sheetnames
            for name in (
                "Performance Dashboard",
                "Server Scorecards",
                "Weekly Performance",
                "Shared & Area Trends",
                "Methodology",
            )
        )
        or any(
            name not in wb.sheetnames
            for name in PRE_PERFORMANCE_VISIBLE_MANAGEMENT_SHEETS
        )
    ):
        return False
    guide_text = "\n".join(
        str(cell.value)
        for cell in wb["How to Use"]._cells.values()
        if cell.value not in (None, "")
    )
    if (
        MANAGEMENT_METHODOLOGY_VERSION not in guide_text
        or PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION in guide_text
    ):
        return False
    for ws in (
        wb[name] for name in PRE_PERFORMANCE_VISIBLE_MANAGEMENT_SHEETS
    ):
        start_column, end_column = management_menu_bounds(ws)
        if (
            ws.sheet_state != "visible"
            or ws.cell(row=2, column=start_column).value != MANAGEMENT_MENU_LABEL
            or not workbook_hyperlink_matches(
                ws.cell(row=2, column=start_column).hyperlink,
                MANAGEMENT_MENU_TARGET,
                sheet_name=ws.title,
                coordinate=ws.cell(row=2, column=start_column).coordinate,
            )
            or str(
                next(
                    (
                        merged
                        for merged in ws.merged_cells.ranges
                        if merged.min_row <= 2 <= merged.max_row
                    ),
                    "",
                )
            )
            != (
                f"{get_column_letter(start_column)}2:"
                f"{get_column_letter(end_column)}2"
            )
        ):
            return False
    return True


def workbook_uses_pre_simplification_layout(wb: Workbook) -> bool:
    """Identify the manifest-bound analytics layout before legacy-tab culling."""

    if (
        stamped_workbook_protection_contract(wb)
        != WORKBOOK_PROTECTION_CONTRACT
        or "How to Use" not in wb.sheetnames
        or any(
            name not in wb.sheetnames
            for name in PRE_SIMPLIFICATION_VISIBLE_MANAGEMENT_SHEETS
        )
        or any(
            wb[name].sheet_state != "visible"
            for name in PRE_SIMPLIFICATION_VISIBLE_MANAGEMENT_SHEETS
        )
    ):
        return False
    guide_text = "\n".join(
        str(cell.value)
        for cell in wb["How to Use"]._cells.values()
        if cell.value not in (None, "")
    )
    if (
        MANAGEMENT_METHODOLOGY_VERSION not in guide_text
        or PREVIOUS_MANAGEMENT_METHODOLOGY_VERSION in guide_text
    ):
        return False
    for ws in (
        wb[name] for name in PRE_SIMPLIFICATION_VISIBLE_MANAGEMENT_SHEETS
    ):
        start_column, end_column = management_menu_bounds(ws)
        if (
            ws.cell(row=2, column=start_column).value != MANAGEMENT_MENU_LABEL
            or not workbook_hyperlink_matches(
                ws.cell(row=2, column=start_column).hyperlink,
                MANAGEMENT_MENU_TARGET,
                sheet_name=ws.title,
                coordinate=ws.cell(row=2, column=start_column).coordinate,
            )
            or str(
                next(
                    (
                        merged
                        for merged in ws.merged_cells.ranges
                        if merged.min_row <= 2 <= merged.max_row
                    ),
                    "",
                )
            )
            != (
                f"{get_column_letter(start_column)}2:"
                f"{get_column_letter(end_column)}2"
            )
        ):
            return False
    return True


def management_workbook_upgrade_pending(path: Path) -> bool:
    if pre_contract_management_workbook(path):
        return True
    wb = load_workbook(path, data_only=False)
    try:
        return (
            workbook_uses_previous_v2_usability(wb)
            or workbook_uses_legacy_multi_link_navigation(wb)
            or workbook_uses_pre_consolidation_layout(wb)
            or workbook_uses_pre_performance_layout(wb)
            or workbook_uses_pre_simplification_layout(wb)
        )
    finally:
        wb.close()


def verify_existing_management_workbook_integrity(
    path: Path,
    *,
    expected_digest: str | None = None,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
    allow_legacy_protection_upgrade: bool = False,
) -> str | None:
    """Verify new-schema workbooks while allowing one-way migration from legacy files."""
    if not path.exists():
        return None
    reject_unapproved_workbook_drawings(path)
    wb = load_workbook(path, data_only=False)
    try:
        is_new_schema = (
            MANAGEMENT_CENTER_SHEET in wb.sheetnames
            and OWNER_ROSTER_TABLE_NAME in wb[MANAGEMENT_CENTER_SHEET].tables
        ) or (
            "Management Setup" in wb.sheetnames
            and OWNER_ROSTER_TABLE_NAME in wb["Management Setup"].tables
        )
        stamped = stamped_workbook_digest(wb)
        protection_contract = stamped_workbook_protection_contract(wb)
        has_action_focus_schema = all(
            name in wb.sheetnames for name in ("Action Focus", "Evidence Detail")
        )
        has_how_to_use_schema = "How to Use" in wb.sheetnames
        has_previous_v2_usability = workbook_uses_previous_v2_usability(wb)
        has_legacy_multi_link_navigation = (
            workbook_uses_legacy_multi_link_navigation(wb)
        )
        has_pre_consolidation_layout = workbook_uses_pre_consolidation_layout(
            wb
        )
        has_pre_performance_layout = workbook_uses_pre_performance_layout(wb)
        has_pre_simplification_layout = workbook_uses_pre_simplification_layout(
            wb
        )
        if (
            MANAGEMENT_CENTER_SHEET in wb.sheetnames
            and "ActionBoardTable" in wb[MANAGEMENT_CENTER_SHEET].tables
        ):
            action_min_col, action_min_row, action_max_col, _ = range_boundaries(
                wb[MANAGEMENT_CENTER_SHEET].tables["ActionBoardTable"].ref
            )
            action_headers = {
                str(wb[MANAGEMENT_CENTER_SHEET].cell(action_min_row, column).value)
                for column in range(action_min_col, action_max_col + 1)
            }
        else:
            action_headers = (
                {
                    str(cell.value)
                    for cell in wb["Action Board"][4]
                    if cell.value is not None
                }
                if "Action Board" in wb.sheetnames
                else set()
            )
        has_v2_action_schema = MANAGEMENT_CENTER_SHEET in wb.sheetnames or {
            "Context Notes",
            "Review Disposition",
            "Evidence Status",
        }.issubset(action_headers)
    finally:
        wb.close()
    if protection_contract == WORKBOOK_PROTECTION_CONTRACT:
        if has_v2_action_schema:
            if has_how_to_use_schema:
                if has_pre_simplification_layout:
                    if (
                        not allow_legacy_protection_upgrade
                        or expected_digest is None
                    ):
                        raise IntegrityError(
                            "The existing master workbook predates the current simplified "
                            "operator tab layout. It may be upgraded only when its exact "
                            "digest is pinned by the verified integrity manifest."
                        )
                    return validate_management_workbook(
                        path,
                        expected_digest,
                        expected_digest_scheme=expected_digest_scheme,
                        legacy_reference_path=legacy_reference_path,
                        previous_simplification=True,
                        visible_sheets=(
                            PRE_SIMPLIFICATION_VISIBLE_MANAGEMENT_SHEETS
                        ),
                        navigation_links=(
                            PRE_SIMPLIFICATION_NAVIGATION_LINKS
                        ),
                    )
                if has_pre_performance_layout:
                    if (
                        not allow_legacy_protection_upgrade
                        or expected_digest is None
                    ):
                        raise IntegrityError(
                            "The existing master workbook predates the current "
                            "performance, shared-number, and area-trend layout. It may be "
                            "upgraded only when its exact digest is pinned by the verified "
                            "integrity manifest."
                        )
                    return validate_management_workbook(
                        path,
                        expected_digest,
                        expected_digest_scheme=expected_digest_scheme,
                        legacy_reference_path=legacy_reference_path,
                        previous_simplification=True,
                        visible_sheets=PRE_PERFORMANCE_VISIBLE_MANAGEMENT_SHEETS,
                        navigation_links=PRE_PERFORMANCE_NAVIGATION_LINKS,
                    )
                if has_pre_consolidation_layout:
                    if (
                        not allow_legacy_protection_upgrade
                        or expected_digest is None
                    ):
                        raise IntegrityError(
                            "The existing master workbook predates the current "
                            "action and trend layout. It may be upgraded only when "
                            "its exact digest is pinned by the verified integrity "
                            "manifest."
                        )
                    return validate_management_workbook(
                        path,
                        expected_digest,
                        expected_digest_scheme=expected_digest_scheme,
                        legacy_reference_path=legacy_reference_path,
                        previous_simplification=True,
                        visible_sheets=(
                            PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
                        ),
                        navigation_links=(
                            PRE_CONSOLIDATION_NAVIGATION_LINKS
                        ),
                    )
                if has_previous_v2_usability:
                    if (
                        not allow_legacy_protection_upgrade
                        or expected_digest is None
                    ):
                        raise IntegrityError(
                            "The existing master workbook predates the current "
                            "usability contract. It may be upgraded only when its "
                            "exact digest is pinned by the verified integrity manifest."
                        )
                    return validate_management_workbook(
                        path,
                        expected_digest,
                        expected_digest_scheme=expected_digest_scheme,
                        legacy_reference_path=legacy_reference_path,
                        previous_v2_usability=True,
                        previous_simplification=True,
                        visible_sheets=(
                            PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
                        ),
                        navigation_links=(
                            PRE_CONSOLIDATION_NAVIGATION_LINKS
                        ),
                    )
                if has_legacy_multi_link_navigation:
                    if (
                        not allow_legacy_protection_upgrade
                        or expected_digest is None
                    ):
                        raise IntegrityError(
                            "The existing master workbook predates the current "
                            "navigation contract. It may be upgraded only when its "
                            "exact digest is pinned by the verified integrity manifest."
                        )
                    return validate_management_workbook(
                        path,
                        expected_digest,
                        expected_digest_scheme=expected_digest_scheme,
                        legacy_reference_path=legacy_reference_path,
                        previous_tab_navigation=True,
                        previous_simplification=True,
                        visible_sheets=(
                            PRE_CONSOLIDATION_VISIBLE_MANAGEMENT_SHEETS
                        ),
                        navigation_links=(
                            PRE_CONSOLIDATION_NAVIGATION_LINKS
                        ),
                    )
                return validate_management_workbook(
                    path,
                    expected_digest or stamped,
                    expected_digest_scheme=expected_digest_scheme,
                    legacy_reference_path=legacy_reference_path,
                )
            return validate_pre_guide_management_workbook(
                path,
                expected_digest or stamped or "",
                expected_digest_scheme=expected_digest_scheme,
                legacy_reference_path=legacy_reference_path,
            )
        if has_action_focus_schema:
            return validate_v1_action_focus_workbook(
                path,
                expected_digest or stamped or "",
                expected_digest_scheme=expected_digest_scheme,
                legacy_reference_path=legacy_reference_path,
            )
        return validate_previous_action_schema_workbook(
            path,
            expected_digest or stamped or "",
            expected_digest_scheme=expected_digest_scheme,
            legacy_reference_path=legacy_reference_path,
        )
    if protection_contract is not None:
        raise IntegrityError(
            "The existing master workbook declares an unsupported protection contract. "
            "Restore a verified generated-workbook version before rerunning."
        )
    if not is_new_schema and stamped is None:
        return None
    if stamped is None:
        raise IntegrityError(
            "The existing master workbook uses the protected owner-roster schema but its "
            "generated-content digest is missing. Restore an earlier Dropbox version or "
            "approved archive copy; no outputs were created and no source files were moved."
        )
    if not allow_legacy_protection_upgrade or expected_digest is None:
        raise IntegrityError(
            "The existing master workbook predates the current protection contract. "
            "It may be adopted only when its exact digest is pinned by the verified "
            "integrity manifest; the next ordinary weekly run will upgrade it."
        )
    return validate_pre_contract_management_workbook(
        path,
        expected_digest,
        expected_digest_scheme=expected_digest_scheme,
        legacy_reference_path=legacy_reference_path,
    )


def action_episode_id(entity_key: str, first_seen: date) -> str:
    digest = hashlib.sha256(f"{entity_key}|{first_seen.isoformat()}".encode("utf-8")).hexdigest()
    return digest[:12].upper()


def compact_server_evidence(row: dict[str, Any]) -> str:
    parts = []
    if row["positive_drivers"]:
        parts.append("Improving: " + "; ".join(row["positive_drivers"][:2]))
    if row["negative_drivers"]:
        parts.append("Watch: " + "; ".join(row["negative_drivers"][:2]))
    if row.get("long_term_direction") != RecentMovement.NOT_EVALUATED.value:
        parts.append(
            f"8-week {row['long_term_direction']} ({row['long_term_history_label']})"
        )
    parts.append(f"{row['guest_count']:,.0f} guests / {row['active_days']} days")
    return " | ".join(parts)


def stable_action_code(action: Any) -> str:
    codes = {
        "coaching prompt": "COACHING_PROMPT",
        "recognition prompt": "RECOGNITION_PROMPT",
        "context review": "CONTEXT_REVIEW",
        "coach now": "COACH_NOW",
        "protect performance": "PROTECT_PERFORMANCE",
        "reinforce improvement": "REINFORCE_IMPROVEMENT",
        "recognize & replicate": "RECOGNIZE_REPLICATE",
        "coach fundamentals": "COACH_FUNDAMENTALS",
        "store review": "STORE_REVIEW",
        "group review": "GROUP_REVIEW",
        "data quality": "DATA_QUALITY_REVIEW",
        "paused carryover": "PAUSED_CARRYOVER",
        "monitor": "MONITOR",
    }
    text = str(action or "").strip().casefold()
    return codes.get(text, re.sub(r"[^A-Z0-9]+", "_", text.upper()).strip("_") or "REVIEW")


def stable_reason_code(signal: dict[str, Any]) -> str:
    action = str(signal.get("Action") or "").strip().casefold()
    movement = str(
        signal.get("Recent Movement", signal.get("Momentum")) or ""
    ).strip().casefold()
    comparison = str(
        signal.get("Peer Comparison", signal.get("Performance Level")) or ""
    ).strip().casefold()
    entity_key = str(signal.get("Entity Key") or "").casefold()
    if action == "data quality":
        return "DQ_INCOMPLETE_LATEST_WEEK"
    if action == "paused carryover":
        return "LATEST_WEEK_INCOMPLETE_PRIOR_ACTION_RETAINED"
    if entity_key.startswith("server|"):
        if action == "coaching prompt":
            return "SERVER_TWO_WEEK_DOWNWARD_BELOW_PEER_STABLE"
        if action == "recognition prompt":
            return "SERVER_TWO_WEEK_UPWARD_ABOVE_PEER_STABLE"
        if action == "context review":
            if movement == "downward" and comparison == "below peer reference":
                return "SERVER_DOWNWARD_BELOW_PEER_CONTEXT_REVIEW"
            if movement == "upward" and comparison == "above peer reference":
                return "SERVER_UPWARD_ABOVE_PEER_CONTEXT_REVIEW"
            return "SERVER_CONTEXT_REVIEW"
        return "SERVER_MATERIAL_SIGNAL"
    if entity_key.startswith("store|"):
        return "STORE_" + stable_action_code(signal.get("Signal"))
    if entity_key.startswith("group|"):
        return "GROUP_" + stable_action_code(signal.get("Signal"))
    return stable_action_code(signal.get("Signal"))


def default_management_comparator_type(signal: dict[str, Any]) -> str:
    """Return a truthful comparator when older evidence omitted the field."""

    entity_key = str(signal.get("Entity Key") or "").strip().casefold()
    action = str(signal.get("Action") or "").strip().casefold()
    if entity_key.startswith("server|"):
        return "Same-store prior-four-week median"
    if entity_key.startswith(("store|", "group|")):
        return "Store/group rolling baseline"
    if entity_key.startswith("data-quality|") or action == "data quality":
        return "Not applicable"
    return "Not applicable"


def enrich_management_signal(signal: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(signal)
    week_ends = sorted(
        {
            str(value)
            for value in enriched.pop("_evidence_week_ends", []) or []
            if value
        }
    )
    sources = enriched.pop("_source_evidence", []) or []
    metric_evidence = enriched.pop("_metric_evidence", {}) or {}
    action_code = stable_action_code(enriched.get("Action"))
    reason_code = stable_reason_code(enriched)
    evidence_payload = {
        "entity_key": enriched.get("Entity Key"),
        "last_seen": (
            enriched.get("Last Seen").isoformat()
            if isinstance(enriched.get("Last Seen"), (date, datetime))
            else enriched.get("Last Seen")
        ),
        "action_code": action_code,
        "reason_code": reason_code,
        "week_ends": week_ends,
        "sources": sources,
        "metric_evidence": metric_evidence,
        "methodology_version": MANAGEMENT_METHODOLOGY_VERSION,
    }
    enriched.update(
        {
            "Action Code": action_code,
            "Reason Code": reason_code,
            "Evidence ID": canonical_json_sha256(evidence_payload)[:16].upper(),
            "Evidence Week Ends": ", ".join(week_ends),
            "Evidence Sources": json.dumps(
                sources, ensure_ascii=False, sort_keys=True, separators=(",", ":")
            ),
            "Metric Evidence": json.dumps(
                metric_evidence,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
            "Comparator Type": enriched.get("Comparator Type")
            or default_management_comparator_type(enriched),
            "Peer Cohort Size": enriched.get("Peer Cohort Size", 0),
            "Peer Cohort Weeks": enriched.get("Peer Cohort Weeks", 0),
            "Threshold Version": enriched.get(
                "Threshold Version", MANAGEMENT_METHODOLOGY_VERSION
            ),
            "Evidence Status": enriched.get(
                "Evidence Status",
                enriched.get("Confidence", ""),
            ),
            "Recurring Drivers": enriched.get("Recurring Drivers", ""),
            "Stability Result": enriched.get("Stability Result", ""),
            "Methodology Version": MANAGEMENT_METHODOLOGY_VERSION,
        }
    )
    return enriched


CONTEXT_ONLY_EVIDENCE_KEYS = frozenset(
    {
        *SERVER_CONTEXT_FIELDS,
        "check_count",
        "check_count_available",
        "per_check_available",
        "sales_per_check",
        "guests_per_check",
        "rate_available",
        "ticket_time_available",
        "ticket_time_weight_basis",
        "context_only_metrics",
        "target_values_context_only",
    }
)


def decision_only_evidence(value: Any) -> Any:
    """Remove operational context that must not reopen a people-review decision."""

    if isinstance(value, dict):
        return {
            str(key): decision_only_evidence(item)
            for key, item in value.items()
            if str(key) not in CONTEXT_ONLY_EVIDENCE_KEYS
        }
    if isinstance(value, list):
        return [decision_only_evidence(item) for item in value]
    return value


def review_reset_fingerprint(action: dict[str, Any]) -> str:
    """Hash only evidence allowed to reset a saved review disposition."""

    raw_metric_evidence = action.get(
        "Metric Evidence", action.get("_metric_evidence", {})
    )
    if isinstance(raw_metric_evidence, str):
        try:
            metric_evidence = json.loads(raw_metric_evidence)
        except json.JSONDecodeError:
            metric_evidence = {"unparsed_metric_evidence": raw_metric_evidence}
    elif isinstance(raw_metric_evidence, dict):
        metric_evidence = raw_metric_evidence
    else:
        metric_evidence = {}
    raw_sources = action.get("Evidence Sources", action.get("_source_evidence", []))
    if isinstance(raw_sources, str):
        try:
            parsed_sources = json.loads(raw_sources)
        except json.JSONDecodeError:
            parsed_sources = [{"unparsed_evidence_sources": raw_sources}]
    else:
        parsed_sources = raw_sources
    if not isinstance(parsed_sources, list):
        parsed_sources = [{"invalid_evidence_sources": parsed_sources}]
    evidence_sources = sorted(
        parsed_sources,
        key=lambda item: json.dumps(
            item, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ),
    )
    last_seen = action.get("Last Seen")
    if isinstance(last_seen, (date, datetime)):
        last_seen = last_seen.isoformat()
    payload = {
        "entity_key": action.get("Entity Key"),
        "last_seen": last_seen,
        "action_code": action.get("Action Code")
        or stable_action_code(action.get("Action")),
        "reason_code": action.get("Reason Code") or stable_reason_code(action),
        "evidence_week_ends": action.get("Evidence Week Ends")
        or action.get("_evidence_week_ends", []),
        "evidence_sources": evidence_sources,
        "metric_evidence": decision_only_evidence(metric_evidence),
        "comparator_type": action.get("Comparator Type"),
        "peer_cohort_size": action.get("Peer Cohort Size"),
        "peer_cohort_weeks": action.get("Peer Cohort Weeks"),
        "threshold_version": action.get("Threshold Version"),
        "evidence_status": action.get("Evidence Status"),
        "recurring_drivers": action.get("Recurring Drivers"),
        "stability_result": action.get("Stability Result"),
        "methodology_version": action.get(
            "Methodology Version", MANAGEMENT_METHODOLOGY_VERSION
        ),
    }
    return canonical_json_sha256(payload)


def refresh_management_evidence(
    signal: dict[str, Any],
    *,
    additional_week_ends: Iterable[date | str] = (),
    metric_evidence_updates: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Recompute evidence codes while preserving the prior evidence inputs."""

    refreshed = dict(signal)
    week_value = refreshed.get("Evidence Week Ends")
    if isinstance(week_value, str):
        week_ends = [
            item.strip() for item in week_value.split(",") if item.strip()
        ]
    elif isinstance(week_value, (list, tuple)):
        week_ends = [str(item) for item in week_value if item]
    else:
        last_seen = as_date(refreshed.get("Last Seen"))
        week_ends = [last_seen.isoformat()] if last_seen else []
    week_ends.extend(
        value.isoformat() if isinstance(value, date) else str(value)
        for value in additional_week_ends
        if value
    )

    def decoded_json(value: Any, expected_type: type, label: str) -> Any:
        if value in (None, ""):
            return expected_type()
        if isinstance(value, expected_type):
            return value
        try:
            parsed = json.loads(str(value))
        except json.JSONDecodeError as exc:
            raise IntegrityError(f"{label} is not valid JSON: {exc.msg}.") from exc
        if not isinstance(parsed, expected_type):
            raise IntegrityError(
                f"{label} must contain a JSON {expected_type.__name__}."
            )
        return parsed

    sources = decoded_json(
        refreshed.get("Evidence Sources"), list, "Evidence Sources"
    )
    metric_evidence = decoded_json(
        refreshed.get("Metric Evidence"), dict, "Metric Evidence"
    )
    if metric_evidence_updates:
        metric_evidence.update(metric_evidence_updates)
    if not sources and not metric_evidence:
        metric_evidence = {
            "signal": refreshed.get("Signal"),
            "why_it_matters": refreshed.get("Why It Matters"),
            "evidence_status": refreshed.get(
                "Evidence Status", refreshed.get("Confidence")
            ),
            "provenance_status": (
                "Legacy action carried forward before row-level source evidence "
                "was introduced."
            ),
        }
    for field in (
        "Evidence ID",
        "Action Code",
        "Reason Code",
        "Evidence Week Ends",
        "Evidence Sources",
        "Metric Evidence",
        "Methodology Version",
    ):
        refreshed.pop(field, None)
    refreshed.update(
        {
            "_evidence_week_ends": week_ends,
            "_source_evidence": sources,
            "_metric_evidence": metric_evidence,
        }
    )
    return enrich_management_signal(refreshed)


def build_management_action_signals(
    server_rows: list[dict[str, Any]],
    store_rows: list[dict[str, Any]],
    group_rows: list[dict[str, Any]],
    weekly_location_rows: list[dict[str, Any]],
    readiness: LatestWeekReadiness | None = None,
) -> list[dict[str, Any]]:
    signals: list[dict[str, Any]] = []
    analytical_ready = readiness is None or readiness.ready
    for row in server_rows if analytical_ready else []:
        if not row["prominent"] or row["action"] == "Monitor":
            continue
        family = (
            "recognition"
            if row.get("candidate_polarity") == CandidatePolarity.POSITIVE.value
            else "coaching"
        )
        entity_key = "|".join(
            ["server", row["location"], str(row["raw_user_name"]), family]
        ).casefold()
        signals.append(
            {
                "Entity Key": entity_key,
                "Priority": row["priority"],
                "Location": row["location"],
                "Person / Area": row["display_name"],
                "Action": row["action"],
                "Signal": (
                    f"{row['momentum']} / {row['performance_level']} / "
                    f"8-week {row['long_term_direction']}"
                ),
                "Why It Matters": compact_server_evidence(row),
                "Recommended Next Step": row["recommended_next_step"],
                "Peer Comparison": row["performance_level"],
                "Recent Movement": row["momentum"],
                "Evidence Status": row["confidence"],
                "Comparator Type": "Same-store prior-four-week median",
                "Peer Cohort Size": row.get("peer_cohort_size", 0),
                "Peer Cohort Weeks": row.get("peer_cohort_weeks", 0),
                "Threshold Version": row.get(
                    "threshold_version", MANAGEMENT_METHODOLOGY_VERSION
                ),
                "Recurring Drivers": ", ".join(
                    row.get("recurring_driver_fields", []) or []
                ),
                "Stability Result": row.get("stability_result", ""),
                "Last Seen": row["week_end"],
                "_evidence_week_ends": row.get("evidence_week_ends", []),
                "_source_evidence": row.get("source_evidence", []),
                "_metric_evidence": {
                    "decision_metric_fields": list(SERVER_PERSON_ACTION_FIELDS),
                    "current_sample": {
                        "guest_count": row.get("guest_count"),
                        "check_count": row.get("check_count"),
                        "check_count_available": row.get(
                            "check_count_available", False
                        ),
                        "active_days": row.get("active_days"),
                    },
                    "recent_changes": row.get("changes"),
                    "recent_metric_scores": row.get("metric_scores"),
                    "peer_changes": row.get("peer_changes"),
                    "peer_metric_scores": row.get("peer_scores"),
                    "combined_metric_scores": row.get("combined_scores"),
                    "long_term_changes": row.get("long_term_changes"),
                    "long_term_metric_scores": row.get("long_term_metric_scores"),
                    "benchmark_values": row.get("benchmark_values"),
                    "benchmark_sources": row.get("benchmark_sources"),
                    "target_values_context_only": row.get("target_values"),
                    "peer_cohort_size": row.get("peer_cohort_size"),
                    "peer_cohort_weeks": row.get("peer_cohort_weeks"),
                    "peer_server_weeks": row.get("peer_server_weeks"),
                    "persistence_reason": row.get("persistence_reason"),
                    "stability_result": row.get("stability_result"),
                    "history_used": row.get("history_used"),
                    "context_only_metrics": row.get("context_metrics"),
                },
            }
        )
    analytical_rows = (("store", store_rows), ("group", group_rows)) if analytical_ready else ()
    for action_type, rows in analytical_rows:
        for row in rows:
            if row["priority"] == "Monitor":
                continue
            entity = row["entity"]
            latest = row["latest"]
            sales_pct = safe_pct_delta(latest["gross_sales"], row["benchmark_values"]["gross_sales"])
            guest_pct = safe_pct_delta(latest["guest_count"], row["benchmark_values"]["guest_count"])
            entity_key = f"{action_type}|{entity}|{row['status']}".casefold()
            signals.append(
                {
                    "Entity Key": entity_key,
                    "Priority": row["priority"],
                    "Location": entity,
                    "Person / Area": entity,
                    "Action": "Store Review" if action_type == "store" else "Group Review",
                    "Signal": row["status"],
                    "Why It Matters": (
                        f"Sales {sales_pct:+.1%}; guests {guest_pct:+.1%} vs benchmark"
                        if sales_pct is not None and guest_pct is not None
                        else "Material movement requires management review"
                    ),
                    "Recommended Next Step": row["recommended_focus"],
                    "Peer Comparison": row["status"],
                    "Recent Movement": "Operational Watch",
                    "Evidence Status": (
                        "Eligible" if row["baseline_weeks"] >= 2 else "Limited History"
                    ),
                    "Comparator Type": "Store/group rolling baseline",
                    "Peer Cohort Size": 0,
                    "Peer Cohort Weeks": row["baseline_weeks"],
                    "Threshold Version": MANAGEMENT_METHODOLOGY_VERSION,
                    "Recurring Drivers": "",
                    "Stability Result": "Not applicable",
                    "Last Seen": latest["week_end"],
                    "_evidence_week_ends": row.get("evidence_week_ends", []),
                    "_source_evidence": row.get("source_evidence", []),
                    "_metric_evidence": {
                        "decision_metrics": {
                            field: latest.get(field)
                            for field in (
                                "gross_sales",
                                "guest_count",
                                *SERVER_PERSON_ACTION_FIELDS,
                            )
                        },
                        "context_only_metrics": {
                            field: latest.get(field)
                            for field in (
                                "check_count",
                                "check_count_available",
                                "per_check_available",
                                "sales_per_check",
                                "guests_per_check",
                                *SERVER_CONTEXT_FIELDS,
                                "rate_available",
                                "ticket_time_available",
                                "ticket_time_weight_basis",
                            )
                        },
                        "benchmark_values": row.get("benchmark_values"),
                        "benchmark_sources": row.get("benchmark_sources"),
                        "prior_changes": row.get("prior_changes"),
                        "benchmark_changes": row.get("benchmark_changes"),
                        "baseline_weeks": row.get("baseline_weeks"),
                    },
                }
            )
    if readiness is not None and readiness.latest_week_end is not None:
        latest_by_location = {
            row["location"]: row for row in readiness.latest_location_rows
        }
        for location in readiness.location_gaps:
            row = latest_by_location.get(location)
            source_days = int(row.get("source_days", 0)) if row else 0
            detail = (
                f"{source_days} of {OPERATING_WEEK_DAYS} source days"
                if row
                else "No current-week store data"
            )
            entity_key = f"data-quality|{location}|latest-week".casefold()
            signals.append(
                {
                    "Entity Key": entity_key,
                    "Priority": "Review",
                    "Location": location,
                    "Person / Area": location,
                    "Action": "Data Quality",
                    "Signal": "Incomplete Latest Week",
                    "Why It Matters": detail,
                    "Recommended Next Step": (
                        "Which source reports are missing or incomplete, and who will "
                        "confirm them before any people review?"
                    ),
                    "Peer Comparison": "Preliminary",
                    "Recent Movement": "Not Evaluated",
                    "Evidence Status": "Incomplete Week",
                    "Last Seen": readiness.latest_week_end,
                    "_evidence_week_ends": [readiness.latest_week_end.isoformat()],
                    "_source_evidence": (
                        row.get("source_evidence", []) if row else []
                    ),
                    "_metric_evidence": {
                        "source_days": source_days,
                        "expected_source_days": OPERATING_WEEK_DAYS,
                        "missing_dates": [
                            item.isoformat() for item in readiness.missing_dates
                        ],
                    },
                }
            )
        if readiness.missing_dates and not readiness.location_gaps:
            signals.append(
                {
                    "Entity Key": "data-quality|all-stores|missing-daily-reports",
                    "Priority": "Review",
                    "Location": "All Stores",
                    "Person / Area": "All Stores",
                    "Action": "Data Quality",
                    "Signal": "Incomplete Latest Week",
                    "Why It Matters": f"Missing {readiness.missing_text}",
                    "Recommended Next Step": (
                        "Which source reports are missing or incomplete, and who will "
                        "confirm them before any people review?"
                    ),
                    "Peer Comparison": "Preliminary",
                    "Recent Movement": "Not Evaluated",
                    "Evidence Status": "Incomplete Week",
                    "Last Seen": readiness.latest_week_end,
                    "_evidence_week_ends": [readiness.latest_week_end.isoformat()],
                    "_source_evidence": merged_source_evidence(
                        readiness.latest_location_rows
                    ),
                    "_metric_evidence": {
                        "missing_dates": [
                            item.isoformat() for item in readiness.missing_dates
                        ],
                        "expected_dates": [
                            item.isoformat() for item in readiness.expected_dates
                        ],
                    },
                }
            )
    elif weekly_location_rows:
        latest_week_end = max(row["week_end"] for row in weekly_location_rows)
        for row in weekly_location_rows:
            if row["week_end"] != latest_week_end or row.get("source_days", 0) >= OPERATING_WEEK_DAYS:
                continue
            entity_key = f"data-quality|{row['location']}|short-week".casefold()
            signals.append(
                {
                    "Entity Key": entity_key,
                    "Priority": "Review",
                    "Location": row["location"],
                    "Person / Area": row["location"],
                    "Action": "Data Quality",
                    "Signal": "Incomplete Latest Week",
                    "Why It Matters": f"{row['source_days']} of {OPERATING_WEEK_DAYS} source days",
                    "Recommended Next Step": (
                        "Which source reports are missing or incomplete, and who will "
                        "confirm them before any people review?"
                    ),
                    "Peer Comparison": "Preliminary",
                    "Recent Movement": "Not Evaluated",
                    "Evidence Status": "Incomplete Week",
                    "Last Seen": latest_week_end,
                    "_evidence_week_ends": [latest_week_end.isoformat()],
                    "_source_evidence": row.get("source_evidence", []),
                    "_metric_evidence": {
                        "source_days": row.get("source_days"),
                        "expected_source_days": OPERATING_WEEK_DAYS,
                    },
                }
            )
    return [enrich_management_signal(signal) for signal in signals]


def merge_management_actions(
    signals: list[dict[str, Any]],
    state: dict[str, Any],
    readiness: LatestWeekReadiness | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    evidence_by_action_id = state.get("evidence_by_action_id", {})
    prior_rows = []
    for row in state.get("active_actions", []):
        enriched_prior = dict(row)
        evidence = evidence_by_action_id.get(str(row.get("Action ID")), {})
        enriched_prior.update(
            {
                field: evidence[field]
                for field in ACTION_EVIDENCE_FIELDS
                if field in evidence
            }
        )
        prior_rows.append(enriched_prior)
    prior_active = {
        str(row.get("Entity Key", "")).casefold(): row
        for row in prior_rows
        if row.get("Entity Key")
    }
    history = []
    for row in state.get("action_history", []):
        enriched_history = dict(row)
        evidence = evidence_by_action_id.get(str(row.get("Action ID")), {})
        enriched_history.update(
            {
                field: evidence[field]
                for field in ACTION_EVIDENCE_FIELDS
                if field in evidence
            }
        )
        history.append(enriched_history)
    current: list[dict[str, Any]] = []
    matched_keys: set[str] = set()
    completed_statuses = {"complete", "dismissed"}
    for signal in signals:
        key = str(signal["Entity Key"]).casefold()
        matched_keys.add(key)
        prior = prior_active.get(key)
        prior_status = str(prior.get("Status", "")).casefold() if prior else ""
        last_seen = as_date(signal.get("Last Seen")) or date.today()
        if prior and prior_status not in completed_statuses:
            first_seen = as_date(prior.get("First Seen")) or last_seen
            action_id = str(prior.get("Action ID") or action_episode_id(key, first_seen))
            status = prior.get("Status") or "Review Needed"
            if (
                "Review Disposition" not in prior
                and str(status).casefold() in {"open", "in progress", "blocked"}
            ):
                status = "Review Needed"
            owner = prior.get("Owner") or ""
            due_date = as_date(prior.get("Due Date"))
            notes = prior.get("Context Notes", prior.get("Manager Notes")) or ""
            same_evidence = (
                str(prior.get("Evidence ID") or "")
                == str(signal.get("Evidence ID") or "")
                or review_reset_fingerprint(prior)
                == review_reset_fingerprint(signal)
            )
            if not same_evidence:
                status = "Review Needed"
            review_disposition = (
                prior.get("Review Disposition") or "Pending Review"
                if same_evidence
                else "Pending Review"
            )
            reviewed_by = prior.get("Reviewed By") or "" if same_evidence else ""
            review_date = (
                as_date(prior.get("Review Date")) if same_evidence else None
            )
        else:
            if prior:
                completed = dict(prior)
                completed["Signal State"] = completed.get("Signal State") or "Completed"
                history.append(completed)
            first_seen = last_seen
            action_id = action_episode_id(key, first_seen)
            status, owner, due_date, notes = "Review Needed", "", None, ""
            review_disposition, reviewed_by, review_date = (
                "Pending Review",
                "",
                None,
            )
        weeks_open = max(1, ((last_seen - first_seen).days // 7) + 1)
        current.append(
            {
                "Action ID": action_id,
                **signal,
                "Status": status,
                "Owner": owner,
                "Due Date": due_date,
                "First Seen": first_seen,
                "Weeks Open": weeks_open,
                "Context Notes": notes,
                "Signal State": "Current",
                "Review Disposition": review_disposition,
                "Reviewed By": reviewed_by,
                "Review Date": review_date,
            }
        )
    for key, prior in prior_active.items():
        if key in matched_keys:
            continue
        prior_status = str(prior.get("Status", "")).casefold()
        is_data_quality = (
            key.startswith("data-quality|")
            or str(prior.get("Action") or "").casefold() == "data quality"
        )
        if (
            readiness is not None
            and not readiness.ready
            and prior_status not in completed_statuses
            and not is_data_quality
        ):
            paused = dict(prior)
            if "Review Disposition" not in paused:
                paused["Status"] = "Review Needed"
                paused["Context Notes"] = (
                    paused.get("Context Notes")
                    or paused.get("Manager Notes")
                    or ""
                )
                paused["Review Disposition"] = "Pending Review"
                paused["Reviewed By"] = ""
                paused["Review Date"] = None
            prior_last_seen = as_date(paused.get("Last Seen"))
            first_seen = (
                as_date(paused.get("First Seen"))
                or prior_last_seen
                or readiness.latest_week_end
                or date.today()
            )
            last_seen = max(
                candidate
                for candidate in (readiness.latest_week_end, prior_last_seen, first_seen)
                if candidate is not None
            )
            paused["Last Seen"] = last_seen
            paused["First Seen"] = first_seen
            paused["Weeks Open"] = max(1, ((last_seen - first_seen).days // 7) + 1)
            prior_signal = str(paused.get("Signal") or "Prior action").strip()
            if not prior_signal.casefold().startswith("paused / carryover"):
                paused["Signal"] = f"PAUSED / CARRYOVER - {prior_signal}"
            missing_text = readiness.missing_text or "latest-week data"
            paused["Priority"] = "Paused"
            paused["Action"] = "Paused Carryover"
            paused["Why It Matters"] = (
                f"Prior action retained; the latest week is incomplete ({missing_text}) "
                "and was not used to confirm or clear it."
            )
            paused["Recommended Next Step"] = (
                "Keep the manual assignment on hold until a complete week can confirm, "
                "update, or clear the prior signal."
            )
            paused["Peer Comparison"] = "Preliminary"
            paused["Recent Movement"] = "Not Evaluated"
            paused["Evidence Status"] = "Paused"
            paused["Signal State"] = "Paused / Carryover"
            paused_context = {
                "latest_week_end": (
                    readiness.latest_week_end.isoformat()
                    if readiness.latest_week_end
                    else None
                ),
                "expected_dates": [
                    value.isoformat() for value in readiness.expected_dates
                ],
                "received_dates": sorted(
                    value.isoformat() for value in readiness.received_dates
                ),
                "missing_dates": [
                    value.isoformat() for value in readiness.missing_dates
                ],
                "location_gaps": list(readiness.location_gaps),
                "missing_text": missing_text,
            }
            current.append(
                refresh_management_evidence(
                    paused,
                    additional_week_ends=(
                        [readiness.latest_week_end]
                        if readiness.latest_week_end
                        else []
                    ),
                    metric_evidence_updates={
                        "paused_carryover": paused_context
                    },
                )
            )
            continue
        cleared = dict(prior)
        cleared["Signal State"] = "Cleared"
        history.append(cleared)

    history_by_id: dict[str, dict[str, Any]] = {}
    current_ids = {str(row["Action ID"]) for row in current}
    for row in history:
        action_id = str(row.get("Action ID") or "")
        if action_id and action_id not in current_ids:
            history_by_id[action_id] = row
    priority_order = {
        "High": 0,
        "Medium": 1,
        "Recognize": 2,
        "Share": 3,
        "Review": 4,
        "Monitor": 5,
        "Paused": 6,
    }
    current.sort(
        key=lambda row: (
            priority_order.get(str(row.get("Priority")), 9),
            str(row.get("Location", "")),
            str(row.get("Person / Area", "")),
        )
    )
    history_rows = sorted(
        history_by_id.values(),
        key=lambda row: (
            as_date(row.get("Last Seen")) or date.min,
            str(row.get("Location", "")),
        ),
        reverse=True,
    )
    return current, history_rows


def inactive_owner_assignments(
    current_actions: Iterable[dict[str, Any]],
    roster: Iterable[Any],
) -> list[dict[str, Any]]:
    """Keep prior assignments visible while flagging owners unavailable for new work."""
    normalized = normalize_owner_roster(roster)
    active = {
        row["Owner Name"].casefold()
        for row in normalized
        if row["Active"] == "Yes"
    }
    known = {row["Owner Name"].casefold(): row["Active"] for row in normalized}
    warnings: list[dict[str, Any]] = []
    for action in current_actions:
        status = str(action.get("Status") or "Open").strip().casefold()
        owner = str(action.get("Owner") or "").strip()
        if not owner or status in {"complete", "dismissed"} or owner.casefold() in active:
            continue
        warnings.append(
            {
                "Owner": owner,
                "Action ID": action.get("Action ID"),
                "Location": action.get("Location"),
                "Person / Area": action.get("Person / Area"),
                "Status": action.get("Status") or "Open",
                "Reason": (
                    "Inactive"
                    if known.get(owner.casefold()) == "No"
                    else "Not on roster"
                ),
            }
        )
    return sorted(
        warnings,
        key=lambda row: (
            str(row["Reason"]), str(row["Owner"]), str(row.get("Action ID") or "")
        ),
    )


def remove_sheet_if_present(wb: Workbook, name: str) -> None:
    if name in wb.sheetnames:
        wb.remove(wb[name])


def worksheet_column_is_hidden(ws: Any, column: int) -> bool:
    """Return effective hidden state without materializing split column records."""

    hidden = False
    for index, dimension in ws.column_dimensions.items():
        default_index = column_index_from_string(str(index))
        start = dimension.min or default_index
        end = dimension.max or start
        if start <= column <= end:
            hidden = dimension.hidden is True
    return hidden


def management_navigation_columns(
    ws,
    navigation_links: tuple[tuple[str, str], ...] | None = None,
) -> tuple[int, ...]:
    """Return the visible columns used by a legacy multi-link navigation bar."""

    links = (
        MANAGEMENT_NAVIGATION_LINKS
        if navigation_links is None
        else navigation_links
    )
    columns: list[int] = []
    column = 1
    while len(columns) < len(links):
        if not worksheet_column_is_hidden(ws, column):
            columns.append(column)
        column += 1
        if column > 256:
            raise IntegrityError(
                f"Worksheet {ws.title!r} has too few visible columns for navigation."
            )
    return tuple(columns)


def management_menu_bounds(ws) -> tuple[int, int]:
    """Return the visible start and end columns for the uniform workbook menu."""

    columns: list[int] = []
    column = 1
    while len(columns) < MANAGEMENT_MENU_VISIBLE_COLUMN_COUNT:
        if not worksheet_column_is_hidden(ws, column):
            columns.append(column)
        column += 1
        if column > 256:
            raise IntegrityError(
                f"Worksheet {ws.title!r} has too few visible columns for the workbook menu."
            )
    return columns[0], columns[-1]


def require_management_title_and_freeze_contract(ws) -> None:
    title_merges = [
        merged
        for merged in ws.merged_cells.ranges
        if merged.min_row == merged.max_row == 1
        and merged.min_col == 1
        and merged.max_col >= 12
    ]
    if len(title_merges) != 1:
        raise IntegrityError(
            f"Worksheet {ws.title!r} title band must extend through column L."
        )
    frozen_at = ws.freeze_panes
    if hasattr(frozen_at, "coordinate"):
        frozen_at = frozen_at.coordinate
    frozen_row_match = re.search(r"\d+", str(frozen_at or ""))
    if not frozen_row_match or int(frozen_row_match.group()) < 3:
        warnings.warn(
            f"Excel rewrote the frozen-pane view on worksheet {ws.title!r}; the next "
            "successful generated save will repair it.",
            UserWarning,
            stacklevel=2,
        )


def require_legacy_tab_navigation_contract(
    wb: Workbook,
    *,
    previous_v2: bool,
    visible_sheets: Iterable[str] | None = None,
    navigation_links: tuple[tuple[str, str], ...] | None = None,
) -> None:
    """Validate the exact protected multi-link layout used before v0.4.0."""

    visible_sheet_names = tuple(
        VISIBLE_MANAGEMENT_SHEETS if visible_sheets is None else visible_sheets
    )
    links = (
        MANAGEMENT_NAVIGATION_LINKS
        if navigation_links is None
        else navigation_links
    )
    expected_labels = [label for label, _ in links]
    for ws in (wb[name] for name in visible_sheet_names):
        navigation_columns = management_navigation_columns(ws, links)
        actual_labels = [
            ws.cell(row=2, column=column).value
            for column in navigation_columns
        ]
        if actual_labels != expected_labels:
            raise IntegrityError(
                f"Worksheet {ws.title!r} does not contain the approved legacy "
                "navigation labels."
            )
        if any(
            worksheet_column_is_hidden(ws, column)
            for column in navigation_columns
        ):
            raise IntegrityError(
                f"Worksheet {ws.title!r} legacy navigation must remain fully visible."
            )
        if ws.row_dimensions[2].height != 24:
            warnings.warn(
                f"Excel rewrote the navigation-row height on worksheet {ws.title!r}; "
                "the next successful generated save will repair it.",
                UserWarning,
                stacklevel=2,
            )
        require_management_title_and_freeze_contract(ws)
        for column, (_, target) in zip(
            navigation_columns, links, strict=True
        ):
            cell = ws.cell(row=2, column=column)
            expected_target = f"#'{target}'!A1"
            if not workbook_hyperlink_matches(
                cell.hyperlink,
                expected_target,
                sheet_name=ws.title,
                coordinate=cell.coordinate,
            ):
                raise IntegrityError(
                    f"Worksheet {ws.title!r} has an invalid legacy navigation target."
                )
            is_current = target == ws.title
            expected_fill = "7A1E1E" if previous_v2 and is_current else "F2F2F2"
            expected_font_color = (
                "FFFFFF" if previous_v2 and is_current else "7A1E1E"
            )
            expected_underline = (
                None if previous_v2 else ("single" if is_current else None)
            )
            expected_side_style = "thin" if previous_v2 else None
            if (
                workbook_color_suffix(cell.fill.fgColor) != expected_fill
                or workbook_color_suffix(cell.font.color) != expected_font_color
                or cell.font.bold is not True
                or cell.font.size != 9
                or cell.font.underline != expected_underline
                or cell.alignment.horizontal != "center"
                or cell.alignment.vertical != "center"
                or cell.alignment.shrink_to_fit is not True
                or cell.border.top.style != "thin"
                or cell.border.bottom.style != "thin"
                or getattr(cell.border.left, "style", None) != expected_side_style
                or getattr(cell.border.right, "style", None) != expected_side_style
            ):
                raise IntegrityError(
                    f"Worksheet {ws.title!r} navigation style does not match the "
                    "contract for the legacy layout."
                )


def require_management_menu_contract(
    wb: Workbook,
    *,
    visible_sheets: Iterable[str] | None = None,
    navigation_links: tuple[tuple[str, str], ...] | None = None,
) -> None:
    """Validate the locked, width-independent current workbook menu."""

    guide = wb["How to Use"]
    if guide["A44"].value != HOW_TO_USE_SECTION_HEADINGS[6]:
        raise IntegrityError(
            "How to Use does not place the workbook map at the approved menu target."
        )
    if {
        str(merged)
        for merged in guide.merged_cells.ranges
        if merged.min_row <= 45 <= merged.max_row
    } != {"A45:B45", "C45:J45", "K45:L45"}:
        raise IntegrityError(
            "How to Use contains an invalid workbook-map heading layout."
        )
    if guide["A45"].value != "Sheet - click to open":
        raise IntegrityError(
            "How to Use is missing the approved workbook-map heading."
        )
    visible_sheet_names = tuple(
        VISIBLE_MANAGEMENT_SHEETS if visible_sheets is None else visible_sheets
    )
    links = (
        MANAGEMENT_NAVIGATION_LINKS
        if navigation_links is None
        else navigation_links
    )
    for row, (_, target) in enumerate(links, start=46):
        cell = guide.cell(row=row, column=1)
        if (
            cell.value != target
            or not workbook_hyperlink_matches(
                cell.hyperlink,
                f"#'{target}'!A1",
                sheet_name=guide.title,
                coordinate=cell.coordinate,
            )
            or cell.font.underline != "single"
            or cell.protection.locked is not True
        ):
            raise IntegrityError(
                "How to Use contains an invalid workbook-map navigation link."
            )
        expected_merges = {
            f"A{row}:B{row}",
            f"C{row}:J{row}",
            f"K{row}:L{row}",
        }
        actual_merges = {
            str(merged)
            for merged in guide.merged_cells.ranges
            if merged.min_row <= row <= merged.max_row
        }
        if actual_merges != expected_merges:
            raise IntegrityError(
                "How to Use contains an invalid workbook-map row layout."
            )

    for ws in (wb[name] for name in visible_sheet_names):
        start_column, end_column = management_menu_bounds(ws)
        expected_range = (
            f"{get_column_letter(start_column)}2:"
            f"{get_column_letter(end_column)}2"
        )
        menu_merges = {
            str(merged)
            for merged in ws.merged_cells.ranges
            if merged.min_row <= 2 <= merged.max_row
        }
        if menu_merges != {expected_range}:
            raise IntegrityError(
                f"Worksheet {ws.title!r} does not use the approved workbook-menu merge."
            )
        if (
            worksheet_column_is_hidden(ws, start_column)
            or worksheet_column_is_hidden(ws, end_column)
        ):
            raise IntegrityError(
                f"Worksheet {ws.title!r} workbook menu must start and end visibly."
            )
        if ws.row_dimensions[2].height != 24:
            warnings.warn(
                f"Excel rewrote the menu-row height on worksheet {ws.title!r}; the next "
                "successful generated save will repair it.",
                UserWarning,
                stacklevel=2,
            )
        require_management_title_and_freeze_contract(ws)
        cell = ws.cell(row=2, column=start_column)
        if (
            cell.value != MANAGEMENT_MENU_LABEL
            or not workbook_hyperlink_matches(
                cell.hyperlink,
                MANAGEMENT_MENU_TARGET,
                sheet_name=ws.title,
                coordinate=cell.coordinate,
            )
        ):
            raise IntegrityError(
                f"Worksheet {ws.title!r} has an invalid navigation target."
            )
        if (
            workbook_color_suffix(cell.fill.fgColor) != "F2F2F2"
            or workbook_color_suffix(cell.font.color) != "7A1E1E"
            or cell.font.bold is not True
            or cell.font.size != 9
            or cell.font.underline != "single"
            or cell.protection.locked is not True
            or cell.alignment.horizontal != "left"
            or cell.alignment.vertical != "center"
            or cell.alignment.shrink_to_fit is not True
            or cell.alignment.indent != 1
            or cell.border.top.style != "thin"
            or cell.border.bottom.style != "thin"
            or getattr(cell.border.left, "style", None) is not None
            or getattr(cell.border.right, "style", None) is not None
        ):
            raise IntegrityError(
                f"Worksheet {ws.title!r} navigation style does not match the contract."
            )
        if cell.protection.locked is not True:
            raise IntegrityError(
                f"Worksheet {ws.title!r} workbook menu must remain locked."
            )


def add_management_navigation(ws) -> None:
    """Add one consistent menu link without changing data-column widths."""

    neutral_fill = PatternFill("solid", fgColor="F2F2F2")
    navigation_border = Border(
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )
    start_column, end_column = management_menu_bounds(ws)
    # Set the anchor border before merging so openpyxl materializes the same
    # top/bottom edge cells on the first save and every subsequent reload.
    # Assigning it only after merge makes the protected semantic digest drift
    # when openpyxl later formats the merged range.
    ws.cell(row=2, column=start_column).border = navigation_border
    ws.merge_cells(
        start_row=2,
        start_column=start_column,
        end_row=2,
        end_column=end_column,
    )
    cell = ws.cell(row=2, column=start_column, value=MANAGEMENT_MENU_LABEL)
    cell.hyperlink = MANAGEMENT_MENU_TARGET
    cell.fill = neutral_fill
    cell.font = Font(
        color="7A1E1E",
        bold=True,
        underline="single",
        size=9,
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        shrink_to_fit=True,
        indent=1,
    )
    cell.border = navigation_border
    ws.row_dimensions[2].height = 24


def configure_management_print_layout(ws) -> None:
    """Use a readable, sheet-specific printable width."""

    last_column = max(ws.max_column, MANAGEMENT_MENU_VISIBLE_COLUMN_COUNT)
    last_row = max(ws.max_row, 2)
    if not ws.print_area:
        ws.print_area = (
            f"A1:{get_column_letter(last_column)}{last_row}"
        )
    if not ws.print_title_rows:
        ws.print_title_rows = "1:2"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = MANAGEMENT_PRINT_WIDTHS.get(ws.title, 1)
    if ws.title in {
        "Data Quality",
        "Server Scorecards",
        "Weekly Performance",
        "Shared & Area Trends",
    }:
        # Wide analytical tables are more useful when each row stays intact.
        # Tabloid landscape keeps the complete view legible on one page wide.
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    if ws.page_setup.fitToHeight is None:
        ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True


def style_management_title(ws, title: str, end_col: int) -> None:
    ws.sheet_view.showGridLines = False
    end_col = max(end_col, MANAGEMENT_MENU_VISIBLE_COLUMN_COUNT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(
        row=1,
        column=1,
        value=title,
    )
    cell.fill = PatternFill("solid", fgColor="7A1E1E")
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32


def style_performance_consistency_value(cell: Any, label: str) -> None:
    fill_color, font_color = PERFORMANCE_CONSISTENCY_COLORS.get(
        label, ("FFFFFF", "262626")
    )
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(color=font_color, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_performance_consistency_calc_sheet(
    wb: Workbook, model: dict[str, Any]
) -> None:
    remove_sheet_if_present(wb, "_Consistency Calc")
    ws = wb.create_sheet("_Consistency Calc")
    style_management_title(ws, "Performance & Consistency Calculation Layer", 21)
    headers = [
        "Week End",
        "Location",
        "Server",
        "Gross Sales",
        "Guests",
        "Sales / Guest",
        "Wine Sales",
        "Wine %",
        "Active Days",
        "Base Sample",
        "Qualified Peers",
        "Peer Sales / Guest",
        "Peer Wine %",
        "Sales / Guest Gap",
        "Wine % Gap",
        "Guest Weight",
        "Qualified Gross Sales",
        "Qualified Guests",
        "Qualified Wine Sales",
        "Period",
        "Weekly Band",
    ]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, item in enumerate(model["detail_rows"], start=4):
        values = [
            item["week_end"],
            excel_safe_text(item["location"]),
            excel_safe_text(item["display_name"]),
            item["gross_sales"],
            item["guest_count"],
            item["check_average"],
            item["wine_sales"],
            item["wine_pct"],
            item["active_days"],
            "Qualified" if item["base_sample"] else "Limited",
            item["distinct_peer_count"],
            item["peer_spg"],
            item["peer_wine_pct"],
            item["spg_gap"],
            item["wine_gap"],
            item["guest_weight"] if item["qualified"] else None,
            item["qualified_gross_sales"],
            item["qualified_guests"],
            item["qualified_wine_sales"],
            item["period"],
            item["weekly_band"],
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=column, value=value)
        ws.cell(row=row_index, column=1).number_format = "mm/dd/yyyy"
        for column in (4, 6, 7, 12, 14, 17, 19):
            ws.cell(row=row_index, column=column).number_format = "$#,##0.00"
        for column in (8, 13, 15):
            ws.cell(row=row_index, column=column).number_format = "0.0%"
        for column in range(1, len(headers) + 1):
            ws.cell(row=row_index, column=column).alignment = Alignment(
                vertical="center", wrap_text=column in {2, 3, 9, 20, 21}
            )

    store_rows = model["store_consistency"]
    for column, header in enumerate(
        ("Location", "High", "Moderate", "Low", "Insufficient"), start=23
    ):
        cell = ws.cell(row=1, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
    for row_index, item in enumerate(store_rows, start=2):
        ws.cell(row=row_index, column=23, value=excel_safe_text(item["location"]))
        for column, label in enumerate(
            ("High", "Moderate", "Low", "Insufficient"), start=24
        ):
            ws.cell(row=row_index, column=column, value=item[label])
    widths = {
        "A": 15,
        "B": 24,
        "C": 24,
        "D": 15,
        "E": 11,
        "F": 15,
        "G": 15,
        "H": 12,
        "I": 12,
        "J": 15,
        "K": 15,
        "L": 17,
        "M": 14,
        "N": 18,
        "O": 14,
        "P": 13,
        "Q": 20,
        "R": 16,
        "S": 19,
        "T": 12,
        "U": 15,
        "W": 24,
        "X": 12,
        "Y": 12,
        "Z": 12,
        "AA": 14,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A4"
    ws.sheet_view.topLeftCell = "A1"


def write_server_scorecards_sheet(
    wb: Workbook, model: dict[str, Any], config: dict[str, Any]
) -> None:
    remove_sheet_if_present(wb, "Server Scorecards")
    ws = wb.create_sheet("Server Scorecards")
    style_management_title(ws, "Server Performance & Consistency Scorecards", 16)
    latest = model["latest_complete_week_end"]
    subtitle = (
        f"Validated snapshot of up to eight complete weeks through {latest:%m/%d/%Y} | "
        "performance and consistency are deliberately shown as separate dimensions"
        if latest
        else "No complete shared week is available for performance and consistency review."
    )
    ws.merge_cells("A3:P3")
    ws["A3"] = subtitle
    ws["A3"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws["A3"].font = Font(color="5C5C5C", size=11)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("A4:P4")
    ws["A4"] = (
        "SPG means Sales / Guest. Performance uses capped guest-weighted average peer gaps; "
        "higher positive gaps are stronger. Consistency uses sample SD of weekly peer gaps; "
        "lower SD is steadier. These are selling-outcome review aids, not total job performance."
    )
    ws["A4"].fill = PatternFill("solid", fgColor="E8F0FA")
    ws["A4"].font = Font(color="3E6FA8", size=11)
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 42

    for start_col, end_col, label, fill, font in (
        (1, 6, "IDENTITY, CLASSIFICATION & CONFIDENCE", "E9E9E9", "641318"),
        (7, 10, "PERFORMANCE | average peer gap up | higher = stronger", "E8F3E8", "1F5B23"),
        (11, 12, "CONSISTENCY | sample SD down | steadier", "E8F0FA", "3E6FA8"),
        (13, 16, "SAMPLE & RECENT FOUR-WEEK MOVEMENT", "E9E9E9", "5C5C5C"),
    ):
        ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        cell = ws.cell(row=5, column=start_col, value=label)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=font, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Location",
        "Server",
        "Overall Read",
        "Performance Level",
        "Consistency",
        "Confidence",
        "8W Sales / Guest",
        "Avg SPG Gap vs Peer",
        "8W Wine %",
        "Avg Wine Gap vs Peer",
        "Weekly SPG Gap SD",
        "Weekly Wine Gap SD",
        "Qualified Weeks",
        "Qualified Guests",
        "Recent SPG Delta",
        "Recent Wine Delta",
    ]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="8B1E23")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 42

    rows = model["summary_rows"]
    for row_index, item in enumerate(rows, start=7):
        values = [
            excel_safe_text(item["location"]),
            excel_safe_text(item["display_name"]),
            item["overall_read"],
            item["performance_level"],
            item["consistency"],
            item["confidence"],
            item["eight_week_spg"],
            item["average_spg_gap"],
            item["eight_week_wine_pct"],
            item["average_wine_gap"],
            item["weekly_spg_gap_sd"],
            item["weekly_wine_gap_sd"],
            item["qualified_weeks"],
            item["qualified_guests"],
            item["recent_spg_delta"],
            item["recent_wine_delta"],
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=column <= 6)
        style_performance_consistency_value(
            ws.cell(row=row_index, column=3), str(item["overall_read"])
        )
        style_performance_consistency_value(
            ws.cell(row=row_index, column=5), str(item["consistency"])
        )
        style_performance_consistency_value(
            ws.cell(row=row_index, column=6), str(item["confidence"])
        )
        for column in (8, 10):
            value = ws.cell(row=row_index, column=column).value
            if value is not None:
                ws.cell(row=row_index, column=column).font = Font(
                    color="1F5B23" if float(value) >= 0 else "8B1E23"
                )
        for column, threshold in (
            (15, PERFORMANCE_CONSISTENCY_RECENT_SPG_THRESHOLD),
            (16, PERFORMANCE_CONSISTENCY_RECENT_WINE_THRESHOLD),
        ):
            value = ws.cell(row=row_index, column=column).value
            if value is not None and abs(float(value)) >= threshold:
                fill, font = (
                    ("E8F3E8", "1F5B23")
                    if float(value) > 0
                    else ("F8E8E9", "8B1E23")
                )
                ws.cell(row=row_index, column=column).fill = PatternFill(
                    "solid", fgColor=fill
                )
                ws.cell(row=row_index, column=column).font = Font(
                    color=font, bold=True
                )
        ws.cell(row=row_index, column=7).number_format = "$0.00"
        ws.cell(row=row_index, column=8).number_format = "+$0.00;-$0.00;$0.00"
        ws.cell(row=row_index, column=9).number_format = "0.0%"
        ws.cell(row=row_index, column=10).number_format = "+0.0%;-0.0%;0.0%"
        ws.cell(row=row_index, column=11).number_format = "$0.00"
        ws.cell(row=row_index, column=12).number_format = "0.0%"
        ws.cell(row=row_index, column=13).number_format = "0"
        ws.cell(row=row_index, column=14).number_format = "#,##0"
        ws.cell(row=row_index, column=15).number_format = "+$0.00;-$0.00;$0.00"
        ws.cell(row=row_index, column=16).number_format = "+0.0%;-0.0%;0.0%"

    if rows:
        table = Table(displayName="ServerScorecardsTable", ref=f"A6:P{6 + len(rows)}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    widths = {
        "A": 22,
        "B": 24,
        "C": 25,
        "D": 19,
        "E": 15,
        "F": 14,
        "G": 17,
        "H": 18,
        "I": 13,
        "J": 18,
        "K": 18,
        "L": 18,
        "M": 15,
        "N": 16,
        "O": 16,
        "P": 16,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "G7"
    ws.print_title_cols = "A:B"
    ws.page_setup.fitToHeight = 1
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.zoomScale = 75


def write_weekly_performance_sheet(wb: Workbook, model: dict[str, Any]) -> None:
    remove_sheet_if_present(wb, "Weekly Performance")
    week_ends = list(model["week_ends"])
    end_column = 5 + len(week_ends) + 2
    end_letter = get_column_letter(max(end_column, MANAGEMENT_MENU_VISIBLE_COLUMN_COUNT))
    ws = wb.create_sheet("Weekly Performance")
    style_management_title(ws, "Weekly Performance Pattern", end_column)
    ws.merge_cells(f"A3:{end_letter}3")
    ws["A3"] = (
        "Repeated colors show consistency; alternating colors show variability. Gray weeks "
        "did not meet the person sample or same-store peer gate. A dash means no row exists."
    )
    ws["A3"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws["A3"].font = Font(color="5C5C5C", size=11)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    for column, label in enumerate(
        ("Strong", "Near Peer", "Mixed", "Below", "Not Qualified"), start=1
    ):
        cell = ws.cell(row=4, column=column, value=label)
        style_performance_consistency_value(cell, label)
    if max(end_column, MANAGEMENT_MENU_VISIBLE_COLUMN_COUNT) >= 6:
        ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=max(end_column, 12))
        ws.cell(row=4, column=6, value="Read dated weeks left to right: oldest to newest")
        ws.cell(row=4, column=6).fill = PatternFill("solid", fgColor="F5F5F5")
        ws.cell(row=4, column=6).font = Font(color="5C5C5C")
        ws.cell(row=4, column=6).alignment = Alignment(vertical="center")
    headers = [
        "Location",
        "Server",
        "Overall Read",
        "Consistency",
        "Confidence",
        *[week_end.strftime("%m/%d/%Y") for week_end in week_ends],
        "Qualified Weeks",
        "Guests",
    ]
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="8B1E23")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[6].height = 38
    rows = model["summary_rows"]
    for row_index, item in enumerate(rows, start=7):
        values = [
            excel_safe_text(item["location"]),
            excel_safe_text(item["display_name"]),
            item["overall_read"],
            item["consistency"],
            item["confidence"],
            *[item["weekly_bands"].get(week_end, "—") for week_end in week_ends],
            item["qualified_weeks"],
            item["qualified_guests"],
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(
                horizontal="center" if column >= 3 else None,
                vertical="center",
                wrap_text=column <= 5,
            )
        for column in range(6, 6 + len(week_ends)):
            style_performance_consistency_value(
                ws.cell(row=row_index, column=column),
                str(ws.cell(row=row_index, column=column).value),
            )
        ws.cell(row=row_index, column=end_column).number_format = "#,##0"
    if rows:
        table = Table(
            displayName="WeeklyPerformanceTable",
            ref=f"A6:{get_column_letter(end_column)}{6 + len(rows)}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    widths = {"A": 22, "B": 24, "C": 25, "D": 16, "E": 14}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for column in range(6, 6 + len(week_ends)):
        ws.column_dimensions[get_column_letter(column)].width = 14
    ws.column_dimensions[get_column_letter(end_column - 1)].width = 16
    ws.column_dimensions[get_column_letter(end_column)].width = 12
    ws.freeze_panes = "F7"
    ws.print_title_cols = "A:B"
    ws.page_setup.fitToHeight = 1
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.zoomScale = 75


def write_performance_methodology_sheet(
    wb: Workbook, model: dict[str, Any], config: dict[str, Any]
) -> None:
    remove_sheet_if_present(wb, "Methodology")
    ws = wb.create_sheet("Methodology")
    style_management_title(ws, "How Performance & Consistency Are Measured", 12)
    latest = model["latest_complete_week_end"]
    ws.merge_cells("A3:L3")
    ws["A3"] = (
        f"Methodology {PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION} | up to eight complete "
        f"weeks through {latest:%m/%d/%Y}: same-store peer comparison and week-to-week variation."
        if latest
        else f"Methodology {PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION} | no complete shared week is available."
    )
    ws["A3"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws["A3"].font = Font(color="5C5C5C", size=11)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")

    for column, label in enumerate(("Overall Read", "Meaning", "Management Use"), start=1):
        cell = ws.cell(row=5, column=column, value=label)
        cell.fill = PatternFill("solid", fgColor="8B1E23")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("E5:H5")
    ws["E5"] = "MODEL NOTES"
    ws["E5"].fill = PatternFill("solid", fgColor="8B1E23")
    ws["E5"].font = Font(color="FFFFFF", bold=True)
    ws["E5"].alignment = Alignment(horizontal="center", vertical="center")
    interpretation_rows = (
        ("Consistently Strong", "Above the performance band with low volatility", "Recognize and identify repeatable practices"),
        ("Consistent / Near Peer", "Near peer norms with low volatility", "Reliable pattern; monitor normal movement"),
        ("Strong / Variable", "Strong average result with uneven weeks", "Diagnose what changes by shift or assignment"),
        ("Mixed / Monitor", "Near peer norms with moderate volatility", "Watch for a clearer pattern"),
        ("Inconsistent", "Low week-to-week consistency", "Review context before coaching the process"),
        ("Below / Variable or Consistently Below", "Below the peer band; variability is shown separately", "Validate context before any action"),
        ("Insufficient Data", "Too few qualified weeks or guests", "Do not draw a conclusion"),
    )
    for row_index, values in enumerate(interpretation_rows, start=6):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row_index, column=1).font = Font(bold=True, color="641318")
        # The Meaning column wraps to three lines for several classifications.
        # Keep enough vertical room in both Excel and the two-page print layout.
        ws.row_dimensions[row_index].height = 50
    ws.merge_cells("E6:H12")
    ws["E6"] = (
        "• SPG means Sales / Guest (Gross Sales divided by Guest Count). SPG and Wine % are "
        "the only person-level outcome metrics in this descriptive layer.\n"
        "• Each qualified current-roster server-week is compared with other qualified current-roster "
        "servers in the same store and week.\n"
        "• Consistency is the sample standard deviation of weekly peer gaps.\n"
        "• Gross sales, Rate of Sale, Ticket Time, and existing Management Center Current Actions signals do not change "
        "these labels.\n"
        "• Schedule, section, party mix, events, tenure, and staffing can explain outcomes. Use this "
        "as a review aid, never as a standalone employment decision."
    )
    ws["E6"].fill = PatternFill("solid", fgColor="E8F0FA")
    ws["E6"].font = Font(color="262626")
    ws["E6"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A14:L14")
    ws["A14"] = "CALCULATION FLOW"
    ws["A14"].fill = PatternFill("solid", fgColor="E9E9E9")
    ws["A14"].font = Font(color="641318", bold=True)
    flow_rows = (
        ("1", "Qualify the week", "Minimum guests and active days", "4", "Measure consistency", "Sample SD of weekly peer gaps"),
        ("2", "Build the peer reference", "Median of other qualified current-roster servers", "5", "Set confidence", "Qualified weeks and guests"),
        ("3", "Measure performance", "Capped guest-weighted average peer gaps", "6", "Combine the two axes", "Overall Read shown on the dashboard"),
    )
    for row_index, values in enumerate(flow_rows, start=15):
        positions = (1, 2, 3, 5, 6, 7)
        for column, value in zip(positions, values, strict=True):
            ws.cell(row=row_index, column=column, value=value)
            ws.cell(row=row_index, column=column).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
        ws.merge_cells(start_row=row_index, start_column=7, end_row=row_index, end_column=8)
        for column in (1, 5):
            ws.cell(row=row_index, column=column).fill = PatternFill(
                "solid", fgColor="F8E8E9"
            )
            ws.cell(row=row_index, column=column).font = Font(
                color="641318", bold=True
            )
            ws.cell(row=row_index, column=column).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        ws.row_dimensions[row_index].height = 34

    ws["A19"] = "Current source"
    ws["B19"] = "Weekly Server Metrics"
    ws["C19"] = f"{model['complete_week_count']} complete shared weeks; latest eight used"
    ws["E19"] = "Model version"
    ws["F19"] = PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION
    ws.merge_cells("F19:H19")
    for column in (1, 5):
        ws.cell(row=19, column=column).font = Font(bold=True, color="641318")
    for column in range(1, 9):
        ws.cell(row=19, column=column).alignment = Alignment(wrap_text=True, vertical="center")

    for column, label in enumerate(("Setting", "Value", "Decision Rule / Purpose"), start=1):
        cell = ws.cell(row=21, column=column, value=label)
        cell.fill = PatternFill("solid", fgColor="8B1E23")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("E21:H21")
    ws["E21"] = "WHY THE SAME THRESHOLD APPEARS TWICE"
    ws["E21"].fill = PatternFill("solid", fgColor="8B1E23")
    ws["E21"].font = Font(color="FFFFFF", bold=True)
    ws["E21"].alignment = Alignment(horizontal="center", vertical="center")

    spg_threshold = management_threshold(config, "check_average")
    wine_threshold = management_threshold(config, "wine_pct")
    min_peers = int(
        config.get("management_peer_reference", {}).get(
            "min_distinct_peers_per_week",
            DEFAULT_CONFIG["management_peer_reference"]["min_distinct_peers_per_week"],
        )
    )
    settings = (
        ("Latest complete week", latest, "Source-driven end date"),
        ("Analysis window (weeks)", PERFORMANCE_CONSISTENCY_WINDOW_WEEKS, "Rolling complete-week view"),
        ("Minimum guests in a week", int(config.get("dashboard_min_guest_count_for_trends", 25)), "Base sample gate"),
        ("Minimum active days in a week", int(config.get("dashboard_min_active_days_for_trends", 3)), "Base sample gate"),
        ("Minimum qualified peers", min_peers, "Same store/week, excluding the server"),
        ("Weekly guest-weight cap", PERFORMANCE_CONSISTENCY_GUEST_WEIGHT_CAP, "Prevents one large week from dominating"),
        ("High confidence: qualified weeks", PERFORMANCE_CONSISTENCY_HIGH_MIN_WEEKS, "High-confidence gate"),
        ("High confidence: qualified guests", PERFORMANCE_CONSISTENCY_HIGH_MIN_GUESTS, "High-confidence gate"),
        ("Provisional: qualified weeks", PERFORMANCE_CONSISTENCY_PROVISIONAL_MIN_WEEKS, "Provisional gate"),
        ("Provisional: qualified guests", PERFORMANCE_CONSISTENCY_PROVISIONAL_MIN_GUESTS, "Provisional gate"),
        ("High consistency maximum: weekly SPG gap SD", float(spg_threshold["neutral"]), "Lower is steadier; both SD gates must pass"),
        ("High consistency maximum: weekly Wine gap SD", float(wine_threshold["neutral"]), "Lower is steadier; both SD gates must pass"),
        ("Moderate consistency maximum: weekly SPG gap SD", float(spg_threshold["strong"]), "Applies only after the High gate is missed"),
        ("Moderate consistency maximum: weekly Wine gap SD", float(wine_threshold["strong"]), "Applies only after the High gate is missed"),
        ("Performance boundary (+/-): average SPG gap", float(spg_threshold["neutral"]), "Positive can support Strong; negative can support Below"),
        ("Performance boundary (+/-): average Wine gap", float(wine_threshold["neutral"]), "Positive can support Strong; negative can support Below"),
        ("Recent movement display threshold: SPG", PERFORMANCE_CONSISTENCY_RECENT_SPG_THRESHOLD, "Highlights material recent-four minus prior-four movement"),
        ("Recent movement display threshold: Wine", PERFORMANCE_CONSISTENCY_RECENT_WINE_THRESHOLD, "Highlights material recent-four minus prior-four movement"),
    )
    for row_index, values in enumerate(settings, start=22):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row_index].height = 30
    ws["B22"].number_format = "mm/dd/yyyy"
    for row_index in (32, 34, 36, 38):
        ws.cell(row=row_index, column=2).number_format = "$0.00"
    for row_index in (33, 35, 37, 39):
        ws.cell(row=row_index, column=2).number_format = "0.0%"

    example = next(
        (
            row
            for row in model["summary_rows"]
            if row["average_spg_gap"] is not None
            and row["weekly_spg_gap_sd"] is not None
        ),
        None,
    )
    ws.merge_cells("E22:F22")
    ws.merge_cells("G22:H22")
    for coordinate, label, fill, font in (
        ("E22", "PERFORMANCE", "2E7D32", "FFFFFF"),
        ("G22", "CONSISTENCY", "3E6FA8", "FFFFFF"),
    ):
        ws[coordinate] = label
        ws[coordinate].fill = PatternFill("solid", fgColor=fill)
        ws[coordinate].font = Font(color=font, bold=True)
        ws[coordinate].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E23:F25")
    ws.merge_cells("G23:H25")
    if example:
        ws["E23"] = (
            "Average SPG gap vs peer\nHigher positive = stronger\n\n"
            f"{excel_safe_text(example['display_name'])}: {example['average_spg_gap']:+.2f} vs "
            f"+{float(spg_threshold['neutral']):.2f} minimum\nResult: {example['performance_level']}"
        )
        ws["G23"] = (
            "Weekly SPG gap sample SD\nLower = steadier\n\n"
            f"{excel_safe_text(example['display_name'])}: {example['weekly_spg_gap_sd']:.2f} vs "
            f"{float(spg_threshold['neutral']):.2f} maximum\nResult: {example['consistency']} consistency"
        )
    else:
        ws["E23"] = "Average peer gap measures performance; higher positive is stronger."
        ws["G23"] = "Sample SD of weekly peer gaps measures consistency; lower is steadier."
    for coordinate, fill, font in (
        ("E23", "E8F3E8", "1F5B23"),
        ("G23", "E8F0FA", "3E6FA8"),
    ):
        ws[coordinate].fill = PatternFill("solid", fgColor=fill)
        ws[coordinate].font = Font(color=font)
        ws[coordinate].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.merge_cells("E26:H31")
    ws["E26"] = (
        "Same number, opposite decision rule.\n\n"
        "Performance: a capped guest-weighted average peer gap must meet a positive or "
        "negative boundary, with the other metric on the same side of zero.\n\n"
        "Consistency: BOTH weekly-gap sample SDs must be at or below their maximums.\n\n"
        "Never compare an average peer gap directly with an SD result."
    )
    ws["E26"].fill = PatternFill("solid", fgColor="FFF4CC")
    ws["E26"].alignment = Alignment(wrap_text=True, vertical="center")

    ws.merge_cells("A41:L41")
    ws["A41"] = (
        "Protected generated view: values are recomputed from verified weekly source data on "
        "each successful run. Technical audit sheets remain veryHidden and protected. The "
        "existing Management Center Current Actions readiness, evidence, persistence, and human-review gates remain separate."
    )
    ws["A41"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws["A41"].font = Font(color="5C5C5C")
    ws["A41"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[41].height = 42
    widths = {
        "A": 38,
        "B": 18,
        "C": 42,
        "D": 3,
        "E": 19,
        "F": 26,
        "G": 37,
        "H": 3,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for column in ("I", "J", "K", "L"):
        ws.column_dimensions[column].width = 3
    ws.freeze_panes = "A5"
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.zoomScale = 80
    ws.print_area = "A1:L41"


def write_shared_area_trends_sheet(
    wb: Workbook,
    model: dict[str, Any],
) -> None:
    remove_sheet_if_present(wb, "Shared & Area Trends")
    ws = wb.create_sheet("Shared & Area Trends")
    style_management_title(ws, "Shared POS & Area Trends", 17)
    week_ends = tuple(model.get("week_ends", ()))
    latest = model.get("latest_complete_week_end")
    ws.merge_cells("A3:Q3")
    ws["A3"] = (
        f"Latest {len(week_ends)} complete shared weeks through {latest:%m/%d/%Y}. "
        "Check Average is Gross Sales / Guests; shared identities remain outside person-level scoring."
        if latest
        else "No complete shared week is available for shared-number or area trends."
    )
    ws["A3"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws["A3"].font = Font(color="5C5C5C", size=11)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 32

    def write_trend_table(
        *,
        section_row: int,
        title: str,
        rows: list[dict[str, Any]],
        identity_header: str,
        identity_key: str,
        context_header: str,
        context_key: str,
        table_name: str,
    ) -> int:
        last_column = 3 + len(week_ends) + 5
        ws.merge_cells(
            start_row=section_row,
            start_column=1,
            end_row=section_row,
            end_column=last_column,
        )
        section = ws.cell(row=section_row, column=1, value=title)
        section.fill = PatternFill("solid", fgColor="8B1E23")
        section.font = Font(color="FFFFFF", bold=True)
        section.alignment = Alignment(vertical="center")
        header_row = section_row + 1
        headers = [
            "Location",
            identity_header,
            context_header,
            *(f"{week_end:%m/%d} CA" for week_end in week_ends),
            "Latest Sales",
            "Latest Guests",
            "Latest Check Avg",
            "Latest Wine %",
            "WoW Check Avg",
        ]
        for column, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=column, value=header)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(bold=True, color="1F1F1F")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 30

        data_start = header_row + 1
        for row_index, item in enumerate(rows, start=data_start):
            values = [
                excel_safe_text(item["location"]),
                excel_safe_text(item[identity_key]),
                excel_safe_text(item[context_key]),
                *(
                    item["weekly_metrics"].get(week_end, {}).get("check_average")
                    for week_end in week_ends
                ),
                item["latest_gross_sales"],
                item["latest_guest_count"],
                item["latest_check_average"],
                item["latest_wine_pct"],
                item["wow_check_average"],
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=column, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=column <= 3)
                if row_index % 2:
                    cell.fill = PatternFill("solid", fgColor="F8F8F8")
            week_start_column = 4
            for column in range(week_start_column, week_start_column + len(week_ends)):
                ws.cell(row=row_index, column=column).number_format = "$#,##0.00"
            metric_start = week_start_column + len(week_ends)
            ws.cell(row=row_index, column=metric_start).number_format = "$#,##0"
            ws.cell(row=row_index, column=metric_start + 1).number_format = "#,##0"
            ws.cell(row=row_index, column=metric_start + 2).number_format = "$#,##0.00"
            ws.cell(row=row_index, column=metric_start + 3).number_format = "0.0%"
            delta_cell = ws.cell(row=row_index, column=metric_start + 4)
            delta_cell.number_format = "+$0.00;-$0.00;$0.00"
            if delta_cell.value is not None:
                delta_cell.font = Font(
                    color="1F5B23" if float(delta_cell.value) >= 0 else "8B1E23",
                    bold=True,
                )
        if rows:
            last_row = data_start + len(rows) - 1
            table = Table(
                displayName=table_name,
                ref=(
                    f"A{header_row}:{get_column_letter(last_column)}{last_row}"
                ),
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
            return last_row
        placeholder_row = data_start
        ws.merge_cells(
            start_row=placeholder_row,
            start_column=1,
            end_row=placeholder_row,
            end_column=last_column,
        )
        placeholder = ws.cell(
            row=placeholder_row,
            column=1,
            value="No matching activity is available in the selected complete-week window.",
        )
        placeholder.fill = PatternFill("solid", fgColor="FFF2CC")
        placeholder.font = Font(color="7F6000")
        placeholder.alignment = Alignment(wrap_text=True, vertical="center")
        return placeholder_row

    shared_end = write_trend_table(
        section_row=5,
        title="SHARED POS NUMBERS | weekly operating results, never person-level scoring",
        rows=list(model.get("shared_rows", [])),
        identity_header="Shared #",
        identity_key="shared_number",
        context_header="Source POS Label",
        context_key="pos_label",
        table_name="SharedPosTrendTable",
    )
    area_section_row = shared_end + 3
    area_end = write_trend_table(
        section_row=area_section_row,
        title=(
            "AREA CHECK AVERAGE | Bar, Patio, Dining Room, Banquets, and Wine Dinners"
        ),
        rows=list(model.get("area_rows", [])),
        identity_header="Area",
        identity_key="area",
        context_header="Mapping",
        context_key="mapping_status",
        table_name="AreaCheckAverageTrendTable",
    )
    note_row = area_end + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=17)
    ws.cell(row=note_row, column=1).value = (
        "Wine Dinners stays unavailable until a real POS label or shared number is configured. "
        "A maintainer can update the weekly area mapping. Dining Room is the residual set of "
        "eligible named servers after shared IDs and explicitly mapped areas are removed."
    )
    ws.cell(row=note_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=note_row, column=1).font = Font(color="7F6000")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[note_row].height = 42
    for column, width in SHARED_AREA_TRENDS_COLUMN_WIDTHS.items():
        # Preserve the operator's Excel AutoFit result as a deterministic fixed
        # width; the bestFit flag itself is Excel save metadata, not layout.
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "D7"
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.zoomScale = 72
    ws.page_setup.fitToHeight = 1
    ws.print_area = f"A1:Q{note_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"


def configure_generated_chart_excel_defaults(
    chart: Any,
    *,
    x_axis_position: str,
    y_axis_position: str,
    line_series: bool,
) -> None:
    """Write the defaults desktop Excel materializes on its first save."""

    chart.roundedCorners = False
    if chart.style is None:
        chart.style = 2
    chart.varyColors = False
    if hasattr(chart, "smooth"):
        chart.smooth = False
    if chart.title is not None:
        chart.title.overlay = False
    if chart.legend is not None:
        chart.legend.overlay = False
    for axis, position in (
        (chart.x_axis, x_axis_position),
        (chart.y_axis, y_axis_position),
    ):
        axis.delete = False
        axis.axPos = position
        if axis.title is not None:
            axis.title.overlay = False
        if axis.tickLblPos is None:
            axis.tickLblPos = "nextTo"
        axis.crosses = "autoZero"
        if axis.numFmt is None:
            axis.numFmt = "General"
            axis.numFmt.sourceLinked = True
    if hasattr(chart.x_axis, "auto"):
        chart.x_axis.auto = False
    if hasattr(chart.x_axis, "lblAlgn"):
        chart.x_axis.lblAlgn = "ctr"
    if hasattr(chart.x_axis, "noMultiLvlLbl"):
        chart.x_axis.noMultiLvlLbl = True
    if hasattr(chart.y_axis, "crossBetween"):
        chart.y_axis.crossBetween = "between"
    for series in chart.series:
        if line_series:
            series.smooth = False
        else:
            series.invertIfNegative = False


def write_performance_dashboard_sheet(
    wb: Workbook,
    model: dict[str, Any],
    config: dict[str, Any],
    readiness: LatestWeekReadiness,
) -> None:
    remove_sheet_if_present(wb, "Performance Dashboard")
    ws = wb.create_sheet("Performance Dashboard")
    style_management_title(ws, "Red Onion Performance Dashboard", 15)
    latest = model["latest_complete_week_end"]
    rows = model["summary_rows"]
    ws.merge_cells("A3:O3")
    ws["A3"] = (
        f"Generated eight-week snapshot through {latest:%m/%d/%Y} | {len(rows)} current "
        "servers | current-roster same-store peer comparison"
        if latest
        else "No complete shared week is available for the performance snapshot."
    )
    ws["A3"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws["A3"].font = Font(color="5C5C5C", size=11)
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    evaluated = [row for row in rows if row["confidence"] != "Insufficient"]
    high_count = sum(row["consistency"] == "High" for row in evaluated)
    moderate_count = sum(row["consistency"] == "Moderate" for row in evaluated)
    low_count = sum(row["consistency"] == "Low" for row in evaluated)
    insufficient_count = len(rows) - len(evaluated)
    summary_parts = [
        f"{high_count} of {len(evaluated)} evaluated servers show high consistency"
        if evaluated
        else "No servers meet the evaluated-data gate"
    ]
    if moderate_count:
        summary_parts.append(f"{moderate_count} show moderate consistency")
    if low_count:
        summary_parts.append(f"{low_count} show low consistency")
    summary_parts.append(f"{insufficient_count} need more qualified data")
    ws.merge_cells("A4:O4")
    ws["A4"] = "; ".join(summary_parts) + "."
    ws["A4"].fill = PatternFill("solid", fgColor="E8F0FA")
    ws["A4"].font = Font(color="3E6FA8", bold=True, size=12)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 30

    overall = model.get("overall_store")
    ws.merge_cells("A6:O6")
    ws["A6"] = (
        f"OVERALL STORE PERFORMANCE | latest complete week {latest:%m/%d/%Y}"
        if latest
        else "OVERALL STORE PERFORMANCE"
    )
    ws["A6"].fill = PatternFill("solid", fgColor="641318")
    ws["A6"].font = Font(color="FFFFFF", bold=True, size=12)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")

    def overall_delta_text(value: float | None, kind: str, baseline_weeks: int) -> str:
        if value is None or baseline_weeks == 0:
            return "No prior complete-week baseline"
        if kind == "pct":
            change = f"{value:+.1%}"
        elif kind == "currency":
            change = f"{'+' if value >= 0 else '-'}${abs(value):.2f}"
        else:
            change = f"{value * 100:+.1f} pts"
        return f"{change} vs prior {baseline_weeks}-week baseline"

    baseline_week_count = int(overall.get("baseline_week_count", 0)) if overall else 0
    overall_cards = (
        (
            "Gross Sales",
            overall.get("gross_sales") if overall else None,
            "$#,##0",
            overall_delta_text(
                overall.get("gross_sales_change_pct") if overall else None,
                "pct",
                baseline_week_count,
            ),
            overall.get("gross_sales_change_pct") if overall else None,
        ),
        (
            "Guests",
            overall.get("guest_count") if overall else None,
            "#,##0",
            overall_delta_text(
                overall.get("guest_count_change_pct") if overall else None,
                "pct",
                baseline_week_count,
            ),
            overall.get("guest_count_change_pct") if overall else None,
        ),
        (
            "Check Average",
            overall.get("check_average") if overall else None,
            "$#,##0.00",
            overall_delta_text(
                overall.get("check_average_change") if overall else None,
                "currency",
                baseline_week_count,
            ),
            overall.get("check_average_change") if overall else None,
        ),
        (
            "Wine Mix",
            overall.get("wine_pct") if overall else None,
            "0.0%",
            overall_delta_text(
                overall.get("wine_pct_change") if overall else None,
                "points",
                baseline_week_count,
            ),
            overall.get("wine_pct_change") if overall else None,
        ),
        (
            "Complete Stores",
            f"{int(overall.get('complete_location_count', 0))} of {len(config.get('locations', {}))}"
            if overall
            else "—",
            "General",
            "Source location totals; not server-row sums",
            None,
        ),
    )
    for start_column, (label, value, number_format, note, delta) in zip(
        (1, 4, 7, 10, 13), overall_cards, strict=True
    ):
        end_column = start_column + 2
        ws.merge_cells(
            start_row=7,
            start_column=start_column,
            end_row=7,
            end_column=end_column,
        )
        ws.merge_cells(
            start_row=8,
            start_column=start_column,
            end_row=9,
            end_column=end_column,
        )
        ws.merge_cells(
            start_row=10,
            start_column=start_column,
            end_row=10,
            end_column=end_column,
        )
        header = ws.cell(row=7, column=start_column, value=label)
        header.fill = PatternFill("solid", fgColor="3E6FA8")
        header.font = Font(color="FFFFFF", bold=True)
        header.alignment = Alignment(horizontal="center", vertical="center")
        value_cell = ws.cell(row=8, column=start_column, value=value)
        value_cell.fill = PatternFill("solid", fgColor="E8F0FA")
        value_cell.font = Font(color="1F4E78", bold=True, size=21)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format
        note_cell = ws.cell(row=10, column=start_column, value=note)
        note_cell.fill = PatternFill("solid", fgColor="E8F0FA")
        note_cell.font = Font(
            color=("1F5B23" if float(delta) >= 0 else "8B1E23")
            if delta is not None
            else "5C5C5C",
            bold=delta is not None,
            size=9,
        )
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[6].height = 24
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 28
    ws.row_dimensions[9].height = 28
    ws.row_dimensions[10].height = 34

    card_data = (
        ("Consistently Strong", "Strong results + low volatility", "2E7D32", "E8F3E8"),
        ("Consistent / Near Peer", "Near peer norms + low volatility", "2F6F6D", "E3F0EF"),
        ("Variable / Inconsistent", "Results change materially week to week", "B77900", "FFF4CC"),
        ("Below / Review", "Below peer band after sample gates", "8B1E23", "F8E8E9"),
        ("Insufficient Data", "No conclusion until volume improves", "5C5C5C", "E9E9E9"),
    )
    counts = model["category_counts"]
    for start_column, (label, note, accent, body) in zip(
        (1, 4, 7, 10, 13), card_data, strict=True
    ):
        end_column = start_column + 2
        ws.merge_cells(start_row=12, start_column=start_column, end_row=12, end_column=end_column)
        ws.merge_cells(start_row=13, start_column=start_column, end_row=14, end_column=end_column)
        ws.merge_cells(start_row=15, start_column=start_column, end_row=15, end_column=end_column)
        header = ws.cell(row=12, column=start_column, value=label)
        header.fill = PatternFill("solid", fgColor=accent)
        header.font = Font(color="FFFFFF", bold=True)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value = ws.cell(row=13, column=start_column, value=counts[label])
        value.fill = PatternFill("solid", fgColor=body)
        value.font = Font(color=accent, bold=True, size=24)
        value.alignment = Alignment(horizontal="center", vertical="center")
        note_cell = ws.cell(row=15, column=start_column, value=note)
        note_cell.fill = PatternFill("solid", fgColor=body)
        note_cell.font = Font(color=accent, size=9)
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[12].height = 24
    ws.row_dimensions[13].height = 28
    ws.row_dimensions[14].height = 28
    ws.row_dimensions[15].height = 34

    consistent_rows = [
        row
        for row in rows
        if row["overall_read"] in {"Consistently Strong", "Consistent / Near Peer"}
    ][:6]
    review_rows = [
        row
        for row in rows
        if row["overall_read"]
        in {
            "Strong / Variable",
            "Mixed / Monitor",
            "Inconsistent",
            "Consistently Below",
            "Below / Variable",
        }
    ][:6]
    ws.merge_cells("A17:G17")
    ws["A17"] = "CONSISTENT PERFORMANCE | strongest current peer-adjusted results"
    ws["A17"].fill = PatternFill("solid", fgColor="E8F3E8")
    ws["A17"].font = Font(color="1F5B23", bold=True)
    ws.merge_cells("I17:O17")
    ws["I17"] = "VARIABILITY / REVIEW | context check before coaching"
    ws["I17"].fill = PatternFill("solid", fgColor="FFF4CC")
    ws["I17"].font = Font(color="B77900", bold=True)
    list_headers = [
        "Server",
        "Location",
        "Overall Read",
        "Confidence",
        "Avg SPG Gap",
        "Avg Wine Gap",
        "Weeks",
    ]
    for start_column in (1, 9):
        for offset, header in enumerate(list_headers):
            cell = ws.cell(row=18, column=start_column + offset, value=header)
            cell.fill = PatternFill("solid", fgColor="8B1E23")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def write_dashboard_list(start_column: int, selected: list[dict[str, Any]]) -> None:
        for offset in range(6):
            row_index = 19 + offset
            if offset < len(selected):
                item = selected[offset]
                values = (
                    excel_safe_text(item["display_name"]),
                    excel_safe_text(item["location"]),
                    item["overall_read"],
                    item["confidence"],
                    item["average_spg_gap"],
                    item["average_wine_gap"],
                    item["qualified_weeks"],
                )
            else:
                values = ("—", "No additional", None, None, None, None, None)
            for column_offset, value in enumerate(values):
                cell = ws.cell(
                    row=row_index, column=start_column + column_offset, value=value
                )
                cell.alignment = Alignment(vertical="center", wrap_text=column_offset < 4)
            ws.cell(row=row_index, column=start_column + 4).number_format = "+$0.00;-$0.00;$0.00"
            ws.cell(row=row_index, column=start_column + 5).number_format = "+0.0%;-0.0%;0.0%"

    write_dashboard_list(1, consistent_rows)
    write_dashboard_list(9, review_rows)

    calc_ws = wb["_Consistency Calc"]
    if model["store_consistency"]:
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.style = 10
        chart.title = "Consistency by Store"
        chart.height = 7.0
        chart.width = 13.0
        max_row = 1 + len(model["store_consistency"])
        data = Reference(calc_ws, min_col=24, max_col=27, min_row=1, max_row=max_row)
        cats = Reference(calc_ws, min_col=23, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        category_formula = f"'_Consistency Calc'!$W$2:$W${max_row}"
        for series in chart.series:
            series.cat = AxDataSource(strRef=StrRef(f=category_formula))
        for series, color in zip(
            chart.series, ("2E7D32", "B77900", "8B1E23", "7F7F7F"), strict=True
        ):
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color
        chart.legend.position = "b"
        configure_generated_chart_excel_defaults(
            chart,
            x_axis_position="l",
            y_axis_position="b",
            line_series=False,
        )
        ws.add_chart(chart, "A27")

    ws.merge_cells("I27:O27")
    ws["I27"] = "READ THE WORKBOOK IN THIS ORDER"
    ws["I27"].fill = PatternFill("solid", fgColor="8B1E23")
    ws["I27"].font = Font(color="FFFFFF", bold=True)
    ws.merge_cells("I28:O41")
    ws["I28"] = (
        "1. Confidence: Is there enough qualified data?\n\n"
        "2. Performance Level: Average SPG and Wine peer gaps. Higher positive is stronger.\n\n"
        "3. Consistency: Sample SD of weekly peer gaps. Lower is steadier.\n\n"
        "4. Weekly Performance: Inspect the dated pattern, then verify shift and assignment context.\n\n"
        "5. Management Center Current Actions: Treat its readiness, evidence, persistence, and human-review gates as separate.\n\n"
        "Important: this measures selling outcomes, not total job performance."
    )
    ws["I28"].fill = PatternFill("solid", fgColor="E8F0FA")
    ws["I28"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A44:O44")
    latest_status = (
        f"latest source week complete ({len(readiness.received_dates)} of {len(readiness.expected_dates)} daily reports)"
        if readiness.ready
        else f"latest source week preliminary ({readiness.missing_text or 'coverage incomplete'})"
    )
    ws["A44"] = (
        f"Data quality: {latest_status}. This view uses only complete shared weeks through "
        f"{latest:%m/%d/%Y}. See Management Center readiness before review; detailed source "
        "reconciliation remains in the protected Data Quality audit layer."
        if latest
        else f"Data quality: {latest_status}. See Management Center readiness before review."
    )
    ws["A44"].fill = PatternFill("solid", fgColor="F5F5F5")
    ws["A44"].font = Font(color="5C5C5C")
    ws["A44"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[44].height = 42
    for column in range(1, 16):
        ws.column_dimensions[get_column_letter(column)].width = 14
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 22
    ws.freeze_panes = "A18"
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.zoomScale = 80
    ws.print_area = "A1:O44"


def write_performance_consistency_sheets(
    wb: Workbook,
    model: dict[str, Any],
    config: dict[str, Any],
    readiness: LatestWeekReadiness,
) -> None:
    write_performance_consistency_calc_sheet(wb, model)
    write_server_scorecards_sheet(wb, model, config)
    write_weekly_performance_sheet(wb, model)
    write_performance_methodology_sheet(wb, model, config)
    write_performance_dashboard_sheet(wb, model, config, readiness)


def write_guide_text_row(
    ws,
    row: int,
    text: str,
    *,
    height: float = 36,
    fill_color: str = "FFFFFF",
    bold: bool = False,
    font_color: str = "1F1F1F",
) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(bold=bold, color=font_color, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = height


def write_guide_section_header(ws, row: int, label: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = PatternFill("solid", fgColor="E7E6E6")
    cell.font = Font(bold=True, color="7A1E1E", size=11)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24


def write_how_to_use_sheet(wb: Workbook) -> None:
    """Write the locked start-here guide for the protected management workbook."""

    remove_sheet_if_present(wb, "How to Use")
    ws = wb.create_sheet("How to Use")
    style_management_title(ws, "How to Use This Workbook", 12)
    for column in range(1, 13):
        ws.column_dimensions[get_column_letter(column)].width = 13
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 85

    write_guide_section_header(ws, 3, HOW_TO_USE_SECTION_HEADINGS[0])
    write_guide_text_row(
        ws,
        4,
        (
            "This workbook is a deterministic, rule-based observational coaching aid. "
            "It does not determine performance and is not statistically confident, "
            "predictive, causal, or validated for demographic fairness. A signal may "
            "never be the sole or determinative basis for discipline, termination, pay, "
            "promotion, demotion, scheduling, reduced opportunity, training, or a formal "
            "performance rating."
        ),
        height=54,
        fill_color="FFF2CC",
        bold=True,
        font_color="7F6000",
    )
    write_guide_text_row(
        ws,
        5,
        (
            f"Methodology {MANAGEMENT_METHODOLOGY_VERSION}. "
            f"{MANAGEMENT_SIGNAL_DISCLAIMER}"
        ),
        height=30,
        fill_color="F3F4F6",
    )

    write_guide_section_header(ws, 6, HOW_TO_USE_SECTION_HEADINGS[1])
    workflow_steps = (
        (
            "1. Read LAST RUN STATUS.txt. Ready means verified within that run's stated "
            "local scope. Attention means stop. NotEvaluated makes no verification claim. "
            "Use Outcome and Stage to confirm whether the run completed. "
            "ExternalCheckRequired means the local run does not evaluate that external "
            "assurance; independent recovery still requires a separate check."
        ),
        (
            "2. Open Management Center and confirm the latest Tuesday–Sunday week is complete, "
            "reconciled, and free of unresolved source, identity, or owner problems."
        ),
        (
            "3. Read Performance Dashboard, Server Scorecards, and Weekly Performance in "
            "that order: confidence first, then peer-adjusted performance, then consistency "
            "and the dated weekly pattern. Methodology defines every gate. These views are "
            "descriptive selling-outcome review aids, not total job performance."
        ),
        (
            "4. Use the overall store card on Performance Dashboard and Shared & Area "
            "Trends for the operations layer: sales, guests, Check Average, Wine Mix, shared "
            "POS results, and guest-weighted area Check Average. Use Methodology for the "
            "definitions and limits. Do not make a person-level decision from a card."
        ),
        (
            "5. Use Management Center as the single current queue. Its readiness, peer-reference, "
            "persistence, stability, evidence, and human-review gates remain separate from the "
            "descriptive performance/consistency labels. Review the evidence fields, then "
            "complete only the blue targets, owner-roster, and action fields."
        ),
        (
            "6. Track follow-up in Management Center. Completed or dismissed items move from "
            "Current Actions to Action History on the next successful run."
        ),
    )
    workflow_step_heights = (48, 36, 42, 42, 48, 36)
    for row, (step, height) in enumerate(
        zip(workflow_steps, workflow_step_heights, strict=True), start=7
    ):
        write_guide_text_row(ws, row, step, height=height)

    write_guide_section_header(ws, 13, HOW_TO_USE_SECTION_HEADINGS[2])
    required_review_rows = (
        "Before dispositioning a generated row, verify:",
        "• Source date, identity, location, completeness, and reconciliation",
        "• Latest sample and peer-reference sufficiency",
        (
            "• Shift, daypart, section, party mix, staffing, event, training, tenure, "
            "or menu context that could explain the result"
        ),
        "• Recurring metric drivers and leave-one-active-day stability",
        "• Independent manager observations supporting any coaching or recognition",
        "Do not infer protected characteristics from names.",
    )
    for row, text in enumerate(required_review_rows, start=14):
        write_guide_text_row(
            ws,
            row,
            text,
            height=42 if row == 17 else 30,
            fill_color="FCE8E6" if row == 20 else "FFFFFF",
            bold=row in {14, 20},
            font_color="9C0006" if row == 20 else "1F1F1F",
        )

    write_guide_section_header(ws, 21, HOW_TO_USE_SECTION_HEADINGS[3])
    signal_rows = (
        "Not Evaluated or Reference Unavailable — gates were not met; no prompt.",
        "Context Review — investigate context only; this is not a coaching conclusion.",
        (
            "Coaching Prompt — second consecutive aligned qualified signal with a "
            "recurring driver and stability; human corroboration is still required."
        ),
        (
            "Recognition Prompt — the same persistence and stability requirements apply; "
            "human corroboration is still required."
        ),
        (
            "Monitor — observe another qualified period; no action quota exists. "
            "Management targets and displayed rank are context only and cannot create "
            "or change a person-level prompt."
        ),
    )
    for row, text in enumerate(signal_rows, start=22):
        write_guide_text_row(ws, row, text, height=42 if row >= 24 else 36)

    write_guide_section_header(ws, 27, HOW_TO_USE_SECTION_HEADINGS[4])
    write_guide_text_row(
        ws,
        28,
        (
            "The normal tab strip keeps How to Use, five analytical tabs, and one Management Center. "
            "It consolidates data readiness, targets, owners, current actions, and action history. "
            "Superseded summaries and audit-only sheets remain preserved but hidden. Blue cells "
            "are the only fields carried forward. "
            f"Review > Unprotect Sheet uses the lowercase password '{WORKBOOK_OPERATOR_PASSWORD}'. "
            "Work on a copy: saved generated or structural changes are detected and can stop the "
            "next protected run."
        ),
        height=66,
        fill_color="D9EAF7",
        bold=True,
        font_color="1F4E78",
    )
    ws.merge_cells("A29:B29")
    ws.merge_cells("C29:L29")
    ws["A29"] = "Field"
    ws["C29"] = "Rule"
    for cell in (ws["A29"], ws["C29"]):
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[29].height = 24
    editable_fields = (
        (
            "D — Status",
            (
                "Generated rows start Review Needed. Open may remain pending. In Progress, "
                "Blocked, Complete, and Dismissed require completed review fields."
            ),
        ),
        ("E — Owner", "Select an active roster member or leave blank."),
        ("F — Due Date", "Enter a real date or leave blank."),
        (
            "N — Context Notes",
            "Record only the minimum information needed to explain the decision.",
        ),
        (
            "U — Review Disposition",
            "Use one of the exact approved choices in Disposition Meanings below.",
        ),
        (
            "V — Reviewed By",
            "Required for a completed disposition and must name an active roster member.",
        ),
        ("W — Review Date", "Required for a completed disposition."),
    )
    for row, (field, rule) in enumerate(editable_fields, start=30):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        ws.cell(row=row, column=1, value=field).font = Font(bold=True, color="1F4E78")
        ws.cell(row=row, column=3, value=rule)
        for col in (1, 3):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor="D9EAF7" if col == 1 else "FFFFFF")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 42 if row == 30 else 36

    write_guide_section_header(ws, 37, HOW_TO_USE_SECTION_HEADINGS[5])
    disposition_descriptions = {
        "Pending Review": "review is incomplete",
        "Coaching Accepted": (
            "independent manager review supports a coaching conversation"
        ),
        "Recognition Accepted": "independent manager review supports recognition",
        "Context Explains": "operating or assignment context explains the signal",
        "Data Issue": (
            "source, identity, metric, or lineage problem; stop and escalate for correction"
        ),
        "Monitor": "no action now; watch a later qualified week",
    }
    for row, disposition in enumerate(REVIEW_DISPOSITION_CHOICES, start=38):
        write_guide_text_row(
            ws,
            row,
            f"{disposition} — {disposition_descriptions[disposition]}.",
            height=36,
        )

    write_guide_section_header(ws, 44, HOW_TO_USE_SECTION_HEADINGS[6])
    for start_col, end_col, value in (
        (1, 2, "Sheet - click to open"),
        (3, 10, "Purpose"),
        (11, 12, "Editable?"),
    ):
        ws.merge_cells(
            start_row=45,
            start_column=start_col,
            end_row=45,
            end_column=end_col,
        )
        cell = ws.cell(row=45, column=start_col, value=value)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[45].height = 24
    workbook_map = (
        (
            "How to Use",
            "start-here workflow, interpretation, controls, and limitations",
            "No — locked",
        ),
        (
            "Performance Dashboard",
            "eight-week confidence, peer-adjusted performance, and consistency overview",
            "No — locked",
        ),
        (
            "Server Scorecards",
            "separate performance, consistency, confidence, sample, and recent movement fields",
            "No — locked",
        ),
        (
            "Weekly Performance",
            "dated Strong, Near Peer, Mixed, Below, Not Qualified, and missing-row pattern",
            "No — locked",
        ),
        (
            "Shared & Area Trends",
            "eight-week shared POS results and guest-weighted Check Average by operating area",
            "No — locked by default",
        ),
        (
            "Methodology",
            "definitions, calculation flow, gates, thresholds, and limitations",
            "No — locked",
        ),
        (
            "Management Center",
            (
                "data readiness, editable targets and owners, current action queue, and "
                "preserved action history"
            ),
            "Yes — blue cells only",
        ),
    )
    workbook_map_start_row = 46
    for row, (sheet_name, purpose, editable) in enumerate(
        workbook_map, start=workbook_map_start_row
    ):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
        sheet_cell = ws.cell(row=row, column=1, value=sheet_name)
        sheet_cell.hyperlink = f"#'{sheet_name}'!A1"
        sheet_cell.font = Font(
            bold=True,
            color="7A1E1E",
            underline="single",
        )
        ws.cell(row=row, column=3, value=purpose)
        ws.cell(row=row, column=11, value=editable)
        for col in (1, 3, 11):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = (
            42 if sheet_name in {
                "How to Use",
                "Performance Dashboard",
                "Server Scorecards",
                "Weekly Performance",
                "Shared & Area Trends",
                "Methodology",
                "Management Center",
            }
            else 36
        )

    limits_header_row = workbook_map_start_row + len(workbook_map)
    write_guide_section_header(ws, limits_header_row, HOW_TO_USE_SECTION_HEADINGS[7])
    limit_rows = (
        (
            "Rate of Sale is opportunities divided by qualifying sales, so lower is "
            "better; multi-row totals use opportunities divided by inferred qualifying "
            "sales. Ticket Time is weighted only by complete Check Count coverage. "
            "Legacy weeks without Check Count show Ticket Time as unavailable. Both "
            "metrics are context only and cannot create person-level prompts."
        ),
        (
            "The source does not observe assignment and equal-opportunity context. Correct "
            "bad source data through the protected run; do not rewrite prior evidence."
        ),
        (
            "Record context or methodology disagreements without deleting the original "
            "signal."
        ),
        (
            "Evidence export or sharing remains a separate manual, exact-fingerprint "
            "approval workflow."
        ),
        (
            "Restricted employee-performance information must remain limited to authorized "
            "users."
        ),
    )
    limits_start_row = limits_header_row + 1
    for row, text in enumerate(limit_rows, start=limits_start_row):
        write_guide_text_row(
            ws,
            row,
            text,
            height=72 if row == limits_start_row else 42 if row == limits_start_row + 1 else 36,
            fill_color="FFF2CC" if row == limits_start_row else "FFFFFF",
        )

    ws.print_area = f"A1:L{limits_start_row + len(limit_rows) - 1}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True


def write_management_setup_sheet(
    wb: Workbook,
    targets: dict[str, dict[str, float | None]],
    owners: list[Any],
    config: dict[str, Any],
    roster_capacity: int | None = None,
) -> None:
    roster = normalize_owner_roster(owners)
    active_names = active_owner_names(roster)
    remove_sheet_if_present(wb, "Management Setup")
    ws = wb.create_sheet("Management Setup")
    style_management_title(ws, "Management Setup", 7)
    ws.merge_cells("A3:G3")
    ws["A3"] = (
        "Blue cells are the only management inputs. The designated custodian maintains targets "
        "and the owner roster; managers update task fields on the Action Board. Changes are "
        "preserved on the next weekly run."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A3"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws.row_dimensions[3].height = 46
    ws.merge_cells("A4:G4")
    ws["A4"] = (
        f"{len(active_names)} active owner{'s' if len(active_names) != 1 else ''} available "
        "for new assignments."
        if active_names
        else "ATTENTION: No active owners are configured. Add names in the Owner Roster below."
    )
    ws["A4"].fill = PatternFill(
        "solid", fgColor="D9EAD3" if active_names else "FFF2CC"
    )
    ws["A4"].font = Font(bold=True, color="274E13" if active_names else "7F6000")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")

    headers = ["Entity", *[label for _, label in TARGET_FIELDS]]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[5].height = 30
    entities = ["All Stores", *config.get("locations", {}).keys()]
    for row_index, entity in enumerate(entities, start=6):
        ws.cell(row=row_index, column=1, value=entity).font = Font(bold=True)
        for col, (field, _) in enumerate(TARGET_FIELDS, start=2):
            value = targets.get(entity, {}).get(field)
            if field == "average_ticket_time_seconds" and value is not None:
                value = value / 60
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="right")
            cell.protection = Protection(locked=False)
    for row in range(6, 6 + len(entities)):
        ws.cell(row=row, column=2).number_format = "$#,##0"
        ws.cell(row=row, column=3).number_format = "#,##0"
        ws.cell(row=row, column=4).number_format = "$0.00"
        ws.cell(row=row, column=5).number_format = "0.0%"
        ws.cell(row=row, column=6).number_format = "0.000"
        ws.cell(row=row, column=7).number_format = "0.0"
    table = Table(displayName="ManagementTargets", ref=f"A5:G{5 + len(entities)}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    threshold_headers = ["Metric", "Neutral Band", "Strong Band", "Better Direction", "Use"]
    for col, header in enumerate(threshold_headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="E7E6E6")
        cell.font = Font(bold=True)
    metric_labels = {
        "check_average": "Check Average",
        "wine_pct": "Wine %",
        "rate_of_sale_by_guest_count": "Rate of Sale",
        "average_ticket_time_seconds": "Ticket Time",
    }
    for row, field in enumerate(metric_labels, start=13):
        threshold = management_threshold(config, field)
        neutral = float(threshold["neutral"])
        strong = float(threshold["strong"])
        if field == "average_ticket_time_seconds":
            neutral, strong = neutral / 60, strong / 60
        ws.cell(row=row, column=1, value=metric_labels[field])
        ws.cell(row=row, column=2, value=neutral)
        ws.cell(row=row, column=3, value=strong)
        ws.cell(row=row, column=4, value="Lower" if threshold.get("lower_is_better") else "Higher")
        ws.cell(row=row, column=5, value="Recent and 8-week trend scoring; benchmark status")
        if field == "wine_pct":
            ws.cell(row=row, column=2).number_format = "0.0%"
            ws.cell(row=row, column=3).number_format = "0.0%"
        elif field == "check_average":
            ws.cell(row=row, column=2).number_format = "$0.00"
            ws.cell(row=row, column=3).number_format = "$0.00"
        else:
            ws.cell(row=row, column=2).number_format = "0.000" if field != "average_ticket_time_seconds" else "0.0"
            ws.cell(row=row, column=3).number_format = "0.000" if field != "average_ticket_time_seconds" else "0.0"

    roster_title_row = 18
    roster_header_row = 20
    style_section_header(ws, roster_title_row, 1, 2, "Owner Roster")
    ws.merge_cells(start_row=19, start_column=1, end_row=19, end_column=7)
    ws.cell(
        row=19,
        column=1,
        value=(
            "Add one person per row and use Yes/No to control new assignments. Mark departing "
            "people No instead of deleting them so existing and historical assignments remain clear."
        ),
    )
    ws.cell(row=19, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(row=19, column=1).fill = PatternFill("solid", fgColor="F3F4F6")
    ws.row_dimensions[19].height = 34
    for col, header in enumerate(OWNER_ROSTER_HEADERS, start=1):
        cell = ws.cell(row=roster_header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    roster_data_rows = max(
        OWNER_ROSTER_MIN_EDIT_ROWS,
        len(roster) + OWNER_ROSTER_SPARE_ROWS,
        int(roster_capacity or 0),
    )
    roster_data_rows = min(OWNER_ROSTER_MAX_ROWS, roster_data_rows)
    roster_last_row = roster_header_row + roster_data_rows
    for offset in range(roster_data_rows):
        row_index = roster_header_row + 1 + offset
        entry = roster[offset] if offset < len(roster) else None
        name_cell = ws.cell(
            row=row_index,
            column=1,
            value=excel_safe_text(entry["Owner Name"]) if entry else None,
        )
        active_cell = ws.cell(
            row=row_index,
            column=2,
            value=entry["Active"] if entry else None,
        )
        for cell in (name_cell, active_cell):
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.protection = Protection(locked=False)
        active_cell.alignment = Alignment(horizontal="center")

    roster_table = Table(
        displayName=OWNER_ROSTER_TABLE_NAME,
        ref=f"A{roster_header_row}:B{roster_last_row}",
    )
    roster_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(roster_table)
    active_validation = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Choose Yes or No",
        error="Select Yes or No from the Owner Roster Active list.",
    )
    ws.add_data_validation(active_validation)
    active_validation.add(f"B{roster_header_row + 1}:B{roster_last_row}")

    widths = {"A": 30, "B": 18, "C": 18, "D": 25, "E": 48, "F": 16, "G": 22}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"


def write_owner_validation_sheet(
    wb: Workbook,
    owners: list[Any],
    roster_capacity: int | None = None,
) -> None:
    roster = normalize_owner_roster(owners)
    setup = (
        wb[MANAGEMENT_CENTER_SHEET]
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames
        and OWNER_ROSTER_TABLE_NAME in wb[MANAGEMENT_CENTER_SHEET].tables
        else wb["Management Setup"]
    )
    min_col, min_row, max_col, max_row = range_boundaries(
        setup.tables[OWNER_ROSTER_TABLE_NAME].ref
    )
    headers = {
        setup.cell(row=min_row, column=column).value: column
        for column in range(min_col, max_col + 1)
    }
    if tuple(headers) != OWNER_ROSTER_HEADERS:
        raise ValueError("Owner Roster headers are not supported.")
    capacity = max_row - min_row
    name_column = get_column_letter(headers["Owner Name"])
    active_column = get_column_letter(headers["Active"])
    remove_sheet_if_present(wb, OWNER_VALIDATION_SHEET)
    ws = wb.create_sheet(OWNER_VALIDATION_SHEET)
    ws["A1"] = "Active Owners"
    for offset in range(capacity):
        setup_row = min_row + 1 + offset
        ws.cell(
            row=2 + offset,
            column=1,
            value=(
                f'=IF(\'{setup.title}\'!${active_column}${setup_row}="Yes",'
                f'\'{setup.title}\'!${name_column}${setup_row},"")'
            ),
        )
    end_row = capacity + 1
    wb.defined_names.pop(OWNER_ROSTER_DEFINED_NAME, None)
    wb.defined_names.add(
        DefinedName(
            OWNER_ROSTER_DEFINED_NAME,
            attr_text=f"'{OWNER_VALIDATION_SHEET}'!$A$2:$A${end_row}",
        )
    )
    ws.sheet_state = "veryHidden"


def action_row_values(row: dict[str, Any]) -> list[Any]:
    return [excel_safe_text(row.get(header)) for header in ACTION_HEADERS]


def compact_action_row_height(values: list[Any]) -> int:
    """Estimate a safe Excel height for the visible wrapped action fields."""

    visible_wrapped_capacity = {
        11: 39,  # Why It Matters, width 46
        12: 41,  # Recommended Next Step, width 48
        14: 29,  # Context Notes, width 34
        19: 15,  # Evidence Status, width 18
        21: 19,  # Review Disposition, width 22
    }
    estimated_lines = 1
    for column, line_capacity in visible_wrapped_capacity.items():
        text = str(values[column - 1] or "")
        cell_lines = sum(
            max(1, math.ceil(len(line) / line_capacity))
            for line in (text.splitlines() or [""])
        )
        estimated_lines = max(estimated_lines, cell_lines)
    return min(300, max(30, estimated_lines * 15))


def write_action_tracking_region(
    ws,
    rows: list[dict[str, Any]],
    *,
    header_row: int,
    editable: bool,
    table_name: str,
    compact_rows: bool = False,
) -> int:
    """Write one action table at an explicit row on a standalone or combined sheet."""

    for col, header in enumerate(ACTION_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[header_row].height = 30
    for row_index, row in enumerate(rows, start=header_row + 1):
        values = action_row_values(row)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.alignment = Alignment(
                vertical="top", wrap_text=col in {10, 11, 12, 14, 15, 16, 19, 21}
            )
        priority_style = priority_fill(row.get("Priority"))
        if priority_style:
            ws.cell(row=row_index, column=3).fill = priority_style
        ws.cell(row=row_index, column=6).number_format = EXCEL_DATE_NUMBER_FORMAT
        ws.cell(row=row_index, column=13).number_format = EXCEL_DATE_NUMBER_FORMAT
        ws.cell(row=row_index, column=17).number_format = EXCEL_DATE_NUMBER_FORMAT
        ws.cell(row=row_index, column=23).number_format = EXCEL_DATE_NUMBER_FORMAT
        if compact_rows:
            ws.row_dimensions[row_index].height = compact_action_row_height(values)
        else:
            ws.row_dimensions[row_index].height = 60
    if rows:
        table = Table(
            displayName=table_name,
            ref=(
                f"A{header_row}:{get_column_letter(len(ACTION_HEADERS))}"
                f"{header_row + len(rows)}"
            ),
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        first, last = header_row + 1, header_row + len(rows)
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        amber_fill = PatternFill("solid", fgColor="FFF2CC")
        green_fill = PatternFill("solid", fgColor="D9EAD3")
        blue_fill = PatternFill("solid", fgColor="D9EAF7")
        ws.conditional_formatting.add(
            f"C{first}:C{last}",
            FormulaRule(formula=[f'$C{first}="High"'], fill=red_fill),
        )
        ws.conditional_formatting.add(
            f"C{first}:C{last}",
            FormulaRule(formula=[f'$C{first}="Medium"'], fill=amber_fill),
        )
        ws.conditional_formatting.add(
            f"C{first}:C{last}",
            FormulaRule(formula=[f'$C{first}="Recognize"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            f"D{first}:D{last}",
            FormulaRule(formula=[f'$D{first}="Complete"'], fill=green_fill),
        )
        ws.conditional_formatting.add(
            f"D{first}:D{last}",
            FormulaRule(formula=[f'$D{first}="Blocked"'], fill=red_fill),
        )
        ws.conditional_formatting.add(
            f"F{first}:F{last}",
            FormulaRule(
                formula=[
                    f'AND($F{first}<TODAY(),$F{first}<>"",'
                    f'$D{first}<>"Complete",$D{first}<>"Dismissed")'
                ],
                fill=red_fill,
            ),
        )
        if editable:
            for row in range(first, last + 1):
                for col in (4, 5, 6, 14, 21, 22, 23):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = blue_fill
                    cell.protection = Protection(locked=False)
            status_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(ACTION_STATUS_CHOICES)}"',
                allow_blank=False,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Choose a valid status",
                error="Select a status from the action list.",
            )
            owner_validation = DataValidation(
                type="list",
                formula1=f"={OWNER_ROSTER_DEFINED_NAME}",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Choose an active owner",
                error="Select an active owner from the Owner Roster or leave the cell blank.",
            )
            disposition_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(REVIEW_DISPOSITION_CHOICES)}"',
                allow_blank=False,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Choose a review disposition",
                error="Select a review disposition from the list.",
            )
            ws.add_data_validation(status_validation)
            ws.add_data_validation(owner_validation)
            ws.add_data_validation(disposition_validation)
            status_validation.add(f"D{first}:D{last}")
            owner_validation.add(f"E{first}:E{last}")
            owner_validation.add(f"V{first}:V{last}")
            disposition_validation.add(f"U{first}:U{last}")
            ws.conditional_formatting.add(
                f"E{first}:E{last}",
                FormulaRule(
                    formula=[
                        f'AND($E{first}<>"",COUNTIF({OWNER_ROSTER_DEFINED_NAME},'
                        f'$E{first})=0,$D{first}<>"Complete",$D{first}<>"Dismissed")'
                    ],
                    fill=amber_fill,
                ),
            )
            ws.conditional_formatting.add(
                f"U{first}:W{last}",
                FormulaRule(
                    formula=[
                        f'AND(OR($D{first}="Complete",$D{first}="Dismissed"),'
                        f'OR($U{first}="Pending Review",$U{first}="",'
                        f'$V{first}="",$W{first}=""))'
                    ],
                    fill=red_fill,
                ),
            )
    return header_row + max(1, len(rows))


def configure_action_tracking_columns(ws, *, hide_internal: bool) -> None:
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].hidden = True
    if hide_internal:
        for column in ("J", "M", "O", "P", "Q", "R"):
            ws.column_dimensions[column].hidden = True
    ws.column_dimensions["T"].hidden = True
    widths = {
        "C": 12,
        "D": 14,
        "E": 18,
        "F": 13,
        "G": 20,
        "H": 24,
        "I": 22,
        "J": 22,
        "K": 46,
        "L": 48,
        "M": 14,
        "N": 34,
        "O": 20,
        "P": 14,
        "Q": 13,
        "R": 12,
        "S": 18,
        "T": 14,
        "U": 22,
        "V": 18,
        "W": 13,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def write_action_tracking_sheet(
    wb: Workbook,
    name: str,
    rows: list[dict[str, Any]],
    *,
    editable: bool,
) -> None:
    remove_sheet_if_present(wb, name)
    ws = wb.create_sheet(name)
    style_management_title(ws, name, len(ACTION_HEADERS))
    if editable:
        ws.merge_cells("C3:W3")
        note = ws["C3"]
        note.value = (
            "Single action queue: start with the highest-priority row, review the "
            "generated evidence, and complete only the blue management fields."
        )
        note.fill = PatternFill("solid", fgColor="FFF2CC")
        note.font = Font(bold=True, color="7F6000")
        note.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[3].height = 36
    header_row = 4
    write_action_tracking_region(
        ws,
        rows,
        header_row=header_row,
        editable=editable,
        table_name="ActionBoardTable" if editable else "ActionHistoryTable",
    )
    configure_action_tracking_columns(ws, hide_internal=editable)
    ws.freeze_panes = "G5"
    ws.sheet_view.zoomScale = 80
    ws.print_title_rows = "1:4"


def write_management_center_sheet(
    wb: Workbook,
    targets: dict[str, dict[str, float | None]],
    owners: list[Any],
    current_actions: list[dict[str, Any]],
    action_history: list[dict[str, Any]],
    config: dict[str, Any],
    readiness: LatestWeekReadiness,
    owner_warnings: list[dict[str, Any]],
    roster_capacity: int | None = None,
) -> None:
    """Consolidate readiness, setup, current actions, and history in one operator tab."""

    roster = normalize_owner_roster(owners)
    active_names = active_owner_names(roster)
    remove_sheet_if_present(wb, MANAGEMENT_CENTER_SHEET)
    ws = wb.create_sheet(MANAGEMENT_CENTER_SHEET)
    style_management_title(ws, "Management Center", len(ACTION_HEADERS))
    ws.merge_cells("C3:W3")
    ws["C3"] = (
        "Weekly control center: confirm data readiness, maintain targets and owners, then "
        "update only the blue cells in Current Actions. Completed and dismissed items move "
        "to the read-only history below. The default printout covers the current management "
        "snapshot; history remains available below for on-screen lookup."
    )
    ws["C3"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["C3"].font = Font(bold=True, color="7F6000")
    ws["C3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 42

    style_section_header(ws, 5, 3, 9, "DATA READINESS")
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=9)
    ws.row_dimensions[5].height = 24
    coverage_text = (
        f"{len(readiness.received_dates)} of {len(readiness.expected_dates)} expected "
        f"{OPERATING_WEEK_LABEL} business dates received"
        if readiness.expected_dates
        else "No complete reporting window is available"
    )
    readiness_rows = [
        (
            "Weekly status",
            "Ready for management review"
            if readiness.ready
            else "Attention required; person-level actions are suppressed",
        ),
        (
            "Latest week",
            readiness.latest_week_end.strftime("%m/%d/%Y")
            if readiness.latest_week_end
            else "Unavailable",
        ),
        ("Source coverage", coverage_text),
        ("Missing inputs", readiness.missing_text or "None"),
        (
            "Owner review",
            (
                f"{len(owner_warnings)} open action assignment(s) need owner review"
                if owner_warnings
                else "No open actions are assigned to inactive or unlisted owners"
            ),
        ),
    ]
    for row_index, (label, value) in enumerate(readiness_rows, start=6):
        ws.merge_cells(
            start_row=row_index, start_column=3, end_row=row_index, end_column=5
        )
        ws.merge_cells(
            start_row=row_index, start_column=6, end_row=row_index, end_column=9
        )
        label_cell = ws.cell(row=row_index, column=3, value=label)
        value_cell = ws.cell(row=row_index, column=6, value=value)
        label_cell.fill = PatternFill("solid", fgColor="F3F4F6")
        label_cell.font = Font(bold=True)
        value_cell.fill = PatternFill(
            "solid",
            fgColor=(
                "D9EAD3"
                if row_index == 6 and readiness.ready
                else "F4CCCC"
                if row_index == 6
                else "FFF2CC"
                if row_index == 10 and owner_warnings
                else "FFFFFF"
            ),
        )
        value_cell.font = Font(
            bold=row_index in {6, 10},
            color=(
                "274E13"
                if row_index == 6 and readiness.ready
                else "9C0006"
                if row_index == 6
                else "7F6000"
                if row_index == 10 and owner_warnings
                else "262626"
            ),
        )
        for cell in (label_cell, value_cell):
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row_index].height = 30 if row_index != 9 else 42

    target_header_row = 12
    style_section_header(ws, 11, 3, 9, "MANAGEMENT TARGETS | blue cells are editable")
    ws.merge_cells(start_row=11, start_column=3, end_row=11, end_column=9)
    target_headers = ["Entity", *[label for _, label in TARGET_FIELDS]]
    for col, header in enumerate(target_headers, start=3):
        cell = ws.cell(row=target_header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[target_header_row].height = 45
    entities = ["All Stores", *config.get("locations", {}).keys()]
    for row_index, entity in enumerate(entities, start=target_header_row + 1):
        ws.cell(row=row_index, column=3, value=entity).font = Font(bold=True)
        for col, (field, _) in enumerate(TARGET_FIELDS, start=4):
            value = targets.get(entity, {}).get(field)
            if field == "average_ticket_time_seconds" and value is not None:
                value = value / 60
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="right")
            cell.protection = Protection(locked=False)
        ws.cell(row=row_index, column=4).number_format = "$#,##0"
        ws.cell(row=row_index, column=5).number_format = "#,##0"
        ws.cell(row=row_index, column=6).number_format = "$0.00"
        ws.cell(row=row_index, column=7).number_format = "0.0%"
        ws.cell(row=row_index, column=8).number_format = "0.000"
        ws.cell(row=row_index, column=9).number_format = "0.0"
        ws.row_dimensions[row_index].height = 24
    targets_last_row = target_header_row + len(entities)
    target_table = Table(
        displayName="ManagementTargets",
        ref=f"C{target_header_row}:I{targets_last_row}",
    )
    target_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(target_table)

    roster_header_row = 12
    style_section_header(
        ws,
        11,
        11,
        12,
        (
            f"OWNER ROSTER | {len(active_names)} active"
            if active_names
            else "OWNER ROSTER | add an active owner"
        ),
    )
    ws.merge_cells(start_row=11, start_column=11, end_row=11, end_column=12)
    ws.row_dimensions[11].height = 30
    for col, header in enumerate(OWNER_ROSTER_HEADERS, start=11):
        cell = ws.cell(row=roster_header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    roster_data_rows = min(
        OWNER_ROSTER_MAX_ROWS,
        max(
            OWNER_ROSTER_MIN_EDIT_ROWS,
            len(roster) + OWNER_ROSTER_SPARE_ROWS,
            int(roster_capacity or 0),
        ),
    )
    roster_last_row = roster_header_row + roster_data_rows
    visible_roster_rows = min(
        roster_data_rows,
        max(OWNER_ROSTER_VISIBLE_EDIT_ROWS, len(roster) + OWNER_ROSTER_SPARE_ROWS),
    )
    for offset in range(roster_data_rows):
        row_index = roster_header_row + 1 + offset
        entry = roster[offset] if offset < len(roster) else None
        name_cell = ws.cell(
            row=row_index,
            column=11,
            value=excel_safe_text(entry["Owner Name"]) if entry else None,
        )
        active_cell = ws.cell(
            row=row_index,
            column=12,
            value=entry["Active"] if entry else None,
        )
        for cell in (name_cell, active_cell):
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.protection = Protection(locked=False)
        active_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row_index].height = max(
            ws.row_dimensions[row_index].height or 15, 21
        )
        ws.row_dimensions[row_index].hidden = (
            offset >= visible_roster_rows and row_index > targets_last_row
        )
    roster_table = Table(
        displayName=OWNER_ROSTER_TABLE_NAME,
        ref=f"K{roster_header_row}:L{roster_last_row}",
    )
    roster_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(roster_table)
    active_validation = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Choose Yes or No",
        error="Select Yes or No from the Owner Roster Active list.",
    )
    ws.add_data_validation(active_validation)
    active_validation.add(f"L{roster_header_row + 1}:L{roster_last_row}")

    current_title_row = max(roster_last_row, targets_last_row) + 2
    style_section_header(
        ws,
        current_title_row,
        3,
        len(ACTION_HEADERS),
        f"CURRENT ACTIONS | {len(current_actions)} open review item(s)",
    )
    ws.merge_cells(
        start_row=current_title_row,
        start_column=3,
        end_row=current_title_row,
        end_column=len(ACTION_HEADERS),
    )
    ws.merge_cells(
        start_row=current_title_row + 1,
        start_column=3,
        end_row=current_title_row + 1,
        end_column=len(ACTION_HEADERS),
    )
    ws.cell(
        row=current_title_row + 1,
        column=3,
        value=(
            "Review the evidence and operational context, then update Status, Owner, Due "
            "Date, Context Notes, Review Disposition, Reviewed By, and Review Date in blue."
        ),
    )
    ws.cell(row=current_title_row + 1, column=3).fill = PatternFill(
        "solid", fgColor="D9EAF7"
    )
    ws.cell(row=current_title_row + 1, column=3).alignment = Alignment(
        wrap_text=True, vertical="center"
    )
    ws.row_dimensions[current_title_row + 1].height = 36
    current_end_row = write_action_tracking_region(
        ws,
        current_actions,
        header_row=current_title_row + 2,
        editable=True,
        table_name="ActionBoardTable",
        compact_rows=True,
    )

    history_title_row = current_end_row + 2
    style_section_header(
        ws,
        history_title_row,
        3,
        len(ACTION_HEADERS),
        f"ACTION HISTORY | {len(action_history)} completed or dismissed item(s)",
    )
    ws.merge_cells(
        start_row=history_title_row,
        start_column=3,
        end_row=history_title_row,
        end_column=len(ACTION_HEADERS),
    )
    history_end_row = write_action_tracking_region(
        ws,
        action_history,
        header_row=history_title_row + 1,
        editable=False,
        table_name="ActionHistoryTable",
        compact_rows=True,
    )

    configure_action_tracking_columns(ws, hide_internal=True)
    ws.freeze_panes = "C6"
    ws.sheet_view.topLeftCell = "C1"
    ws.sheet_view.zoomScale = 75
    ws.print_title_rows = "1:5"
    ws.print_area = f"A1:W{current_end_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.pageOrder = "overThenDown"
    ws.print_options.horizontalCentered = True


def write_action_focus_sheet(
    wb: Workbook, current_actions: list[dict[str, Any]]
) -> None:
    remove_sheet_if_present(wb, "Action Focus")
    ws = wb.create_sheet("Action Focus")
    headers = [
        "Priority",
        "Status",
        "Owner",
        "Due Date",
        "Location",
        "Person / Area",
        "Action",
        "Evidence Status",
        "Review Disposition",
        "Manager Question",
        "Why It Matters",
        "Weeks Open",
        "Open Action",
    ]
    style_management_title(ws, "Action Focus", len(headers))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws.cell(
        row=3,
        column=1,
        value=(
            "Start with the highest-priority row. Use the Manager Question to review "
            "context, then open Action Board to document the owner, due date, and "
            "review disposition."
        ),
    )
    ws.cell(row=3, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=3, column=1).alignment = Alignment(
        wrap_text=True, vertical="center"
    )
    ws.row_dimensions[3].height = 48
    actionable = [
        row
        for row in current_actions
        if str(row.get("Status") or "").strip().casefold()
        not in {"complete", "dismissed"}
    ]
    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    board_rows = {
        str(row.get("Action ID")): index
        for index, row in enumerate(current_actions, start=5)
    }
    for row_index, action in enumerate(actionable, start=header_row + 1):
        action_id = str(action.get("Action ID") or "")
        values = [
            action.get("Priority"),
            action.get("Status"),
            action.get("Owner"),
            as_date(action.get("Due Date")),
            action.get("Location"),
            action.get("Person / Area"),
            action.get("Action"),
            action.get("Evidence Status"),
            action.get("Review Disposition"),
            action.get("Recommended Next Step"),
            action.get("Why It Matters"),
            action.get("Weeks Open"),
            "Open in Action Board",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=excel_safe_text(value))
            cell.alignment = Alignment(
                vertical="top", wrap_text=col in {6, 7, 8, 9, 10, 11}
            )
        ws.cell(row=row_index, column=4).number_format = EXCEL_DATE_NUMBER_FORMAT
        target_row = board_rows.get(action_id)
        if target_row:
            link = ws.cell(row=row_index, column=13)
            link.hyperlink = f"#'Action Board'!C{target_row}"
            link.style = "Hyperlink"
        fill = priority_fill(action.get("Priority"))
        if fill:
            ws.cell(row=row_index, column=1).fill = fill
        ws.row_dimensions[row_index].height = 54
    if actionable:
        table = Table(
            displayName="ActionFocusTable",
            ref=f"A{header_row}:M{header_row + len(actionable)}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    for column, width in {
        "A": 12,
        "B": 15,
        "C": 18,
        "D": 13,
        "E": 20,
        "F": 24,
        "G": 23,
        "H": 18,
        "I": 22,
        "J": 50,
        "K": 44,
        "L": 12,
        "M": 21,
    }.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "E6"
    ws.sheet_view.zoomScale = 85
    ws.print_title_rows = "1:5"


def action_with_evidence_fallback(action: dict[str, Any]) -> dict[str, Any]:
    if (
        action.get("Evidence ID")
        and action.get("Action Code") == stable_action_code(action.get("Action"))
        and action.get("Reason Code") == stable_reason_code(action)
    ):
        return dict(action)
    return refresh_management_evidence(action)


def write_evidence_detail_sheet(
    wb: Workbook,
    current_actions: list[dict[str, Any]],
    action_history: list[dict[str, Any]],
) -> None:
    remove_sheet_if_present(wb, "Evidence Detail")
    ws = wb.create_sheet("Evidence Detail")
    style_management_title(ws, "Evidence Detail", len(EVIDENCE_DETAIL_HEADERS))
    ws.merge_cells(
        start_row=3,
        start_column=1,
        end_row=3,
        end_column=len(EVIDENCE_DETAIL_HEADERS),
    )
    ws.cell(
        row=3,
        column=1,
        value=(
            "Person-level actions use Sales / Guest and Wine % only; Check Count, "
            "Rate of Sale, and Ticket Time are context only. "
            "Operator view: exact raw Evidence Sources and Metric Evidence columns are "
            "hidden. Their protected values remain unchanged for carry-forward, approved "
            "evidence export, and audit."
        ),
    )
    ws.cell(row=3, column=1).fill = PatternFill("solid", fgColor="D9EAF7")
    ws.cell(row=3, column=1).font = Font(bold=True, color="1F4E78")
    ws.cell(row=3, column=1).alignment = Alignment(
        wrap_text=True, vertical="center"
    )
    ws.row_dimensions[3].height = 54
    evidence_by_action: dict[str, dict[str, Any]] = {}
    for action in [*action_history, *current_actions]:
        action_id = str(action.get("Action ID") or "")
        if action_id:
            evidence_by_action[action_id] = action_with_evidence_fallback(action)
    rows = sorted(
        evidence_by_action.values(),
        key=lambda row: (
            str(row.get("Signal State") or "") != "Current",
            as_date(row.get("Last Seen")) or date.min,
            str(row.get("Action ID") or ""),
        ),
    )
    header_row = 4
    for col, header in enumerate(EVIDENCE_DETAIL_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 42
    action_sheet_name = (
        MANAGEMENT_CENTER_SHEET
        if (
            MANAGEMENT_CENTER_SHEET in wb.sheetnames
            and "ActionBoardTable" in wb[MANAGEMENT_CENTER_SHEET].tables
        )
        else "Action Board"
        if (
            "Action Board" in wb.sheetnames
            and "ActionBoardTable" in wb["Action Board"].tables
        )
        else None
    )
    action_table_rows: dict[str, int] = {}
    if action_sheet_name is not None:
        action_sheet = wb[action_sheet_name]
        min_col, action_header_row, max_col, action_last_row = range_boundaries(
            action_sheet.tables["ActionBoardTable"].ref
        )
        action_id_column = next(
            (
                column
                for column in range(min_col, max_col + 1)
                if action_sheet.cell(action_header_row, column).value == "Action ID"
            ),
            None,
        )
        if action_id_column is not None:
            action_table_rows = {
                str(action_sheet.cell(row, action_id_column).value): row
                for row in range(action_header_row + 1, action_last_row + 1)
                if action_sheet.cell(row, action_id_column).value not in (None, "")
            }
    for row_index, action in enumerate(rows, start=header_row + 1):
        for col, header in enumerate(EVIDENCE_DETAIL_HEADERS, start=1):
            value = action.get(header)
            cell = ws.cell(row=row_index, column=col, value=excel_safe_text(value))
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=header
                in {
                    "Evidence Week Ends",
                    "Comparator Type",
                    "Evidence Status",
                    "Recurring Drivers",
                    "Stability Result",
                    "Review Disposition",
                },
            )
            if header in {"Due Date", "Last Seen", "Review Date"}:
                cell.number_format = EXCEL_DATE_NUMBER_FORMAT
        action_id = str(action.get("Action ID") or "")
        target_row = action_table_rows.get(action_id)
        if action_sheet_name is not None and target_row:
            evidence_cell = ws.cell(row=row_index, column=1)
            evidence_cell.hyperlink = f"#'{action_sheet_name}'!C{target_row}"
            evidence_cell.style = "Hyperlink"
        visible_lengths = [
            len(str(action.get(header) or ""))
            for header in (
                "Evidence Week Ends",
                "Comparator Type",
                "Evidence Status",
                "Recurring Drivers",
                "Stability Result",
                "Review Disposition",
            )
        ]
        ws.row_dimensions[row_index].height = (
            90
            if max(visible_lengths, default=0) >= 100
            or sum(visible_lengths) >= 260
            else 75
        )
    if rows:
        table = Table(
            displayName="ManagementEvidenceDetail",
            ref=(
                f"A{header_row}:{get_column_letter(len(EVIDENCE_DETAIL_HEADERS))}"
                f"{header_row + len(rows)}"
            ),
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    widths = {
        "A": 20,
        "B": 16,
        "C": 25,
        "D": 46,
        "E": 12,
        "F": 14,
        "G": 18,
        "H": 13,
        "I": 20,
        "J": 24,
        "K": 42,
        "L": 68,
        "M": 68,
        "N": 20,
        "O": 13,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for header in ("Evidence Sources", "Metric Evidence"):
        ws.column_dimensions[
            get_column_letter(EVIDENCE_DETAIL_HEADERS.index(header) + 1)
        ].hidden = True
    for header, width in {
        "Comparator Type": 34,
        "Peer Cohort Size": 16,
        "Peer Cohort Weeks": 16,
        "Threshold Version": 18,
        "Evidence Status": 18,
        "Recurring Drivers": 32,
        "Stability Result": 42,
        "Review Disposition": 22,
        "Reviewed By": 18,
        "Review Date": 13,
    }.items():
        ws.column_dimensions[
            get_column_letter(EVIDENCE_DETAIL_HEADERS.index(header) + 1)
        ].width = width
    ws.freeze_panes = "E5"
    ws.sheet_view.zoomScale = 70
    ws.print_title_rows = "1:4"


def operational_sample_text(row: dict[str, Any]) -> str:
    check_text = (
        f"{float(row.get('check_count', 0) or 0):,.0f} checks"
        if row.get("check_count_available")
        else "checks unavailable"
    )
    return (
        f"{float(row.get('guest_count', 0) or 0):,.0f} guests / "
        f"{check_text} / {int(row.get('active_days', 0) or 0)} days"
    )


def write_team_trends_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    remove_sheet_if_present(wb, "Server Scorecard")
    remove_sheet_if_present(wb, "Team Trends")
    headers = [
        "Location",
        "Server",
        "Week End",
        "Prior Week End",
        "Current Sample",
        "Sales / Guest",
        "WoW Sales / Guest Δ",
        "Wine %",
        "WoW Wine % Δ",
        "WoW Movement",
        "4-Week Movement",
        "8-Week Direction",
        "Peer Comparison",
        "Evidence Status",
        "Action Gate",
        "Trend Drivers",
        "History Used",
    ]
    data = []
    for row in sorted(
        rows,
        key=lambda item: (
            str(item.get("location") or ""),
            str(item.get("display_name") or ""),
        ),
    ):
        driver_parts = []
        if row["positive_drivers"]:
            driver_parts.append("Up: " + "; ".join(row["positive_drivers"]))
        if row["negative_drivers"]:
            driver_parts.append("Watch: " + "; ".join(row["negative_drivers"]))
        if not driver_parts:
            driver_parts.append("No material 4-week driver")
        wow_changes = row.get("week_over_week_changes", {})
        data.append(
            [
                row["location"],
                excel_safe_text(row["display_name"]),
                row["week_end"],
                row.get("previous_week_end"),
                operational_sample_text(row),
                row["check_average"],
                wow_changes.get("check_average"),
                row["wine_pct"],
                wow_changes.get("wine_pct"),
                row.get("week_over_week_movement", "No prior week"),
                row.get("descriptive_recent_movement", row["momentum"]),
                row["long_term_direction"],
                row["performance_level"],
                row["confidence"],
                row["action"],
                " | ".join(driver_parts),
                row["history_used"],
            ]
        )
    ws = write_table_sheet(
        wb, "Team Trends", headers, data, "TeamTrends",
        widths={
            "A": 20, "B": 24, "C": 15, "D": 20, "E": 24, "F": 20,
            "G": 18, "H": 12, "I": 20, "J": 20, "K": 23, "L": 22,
            "M": 23, "N": 20, "O": 18, "P": 42, "Q": 52,
        },
    )
    style_management_title(
        ws,
        "Team Trends — All Current Servers | Descriptive WoW + Gated Actions",
        len(headers),
    )
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "E4"
    ws.sheet_view.zoomScale = 80
    ws.print_title_rows = "1:3"
    ws.row_dimensions[3].height = 30
    for row in range(4, 4 + len(data)):
        ws.cell(row=row, column=3).number_format = EXCEL_DATE_NUMBER_FORMAT
        ws.cell(row=row, column=4).number_format = EXCEL_DATE_NUMBER_FORMAT
        ws.cell(row=row, column=6).number_format = "$0.00"
        ws.cell(row=row, column=7).number_format = "+$0.00;[Red]-$0.00;$0.00"
        ws.cell(row=row, column=8).number_format = "0.0%"
        ws.cell(row=row, column=9).number_format = "+0.0%;[Red]-0.0%;0.0%"
        for col in (10, 11, 12, 13, 14, 15):
            fill = priority_fill(ws.cell(row=row, column=col).value)
            if fill:
                ws.cell(row=row, column=col).fill = fill
        for col in (5, 16, 17):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 54


def write_rising_falling_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    remove_sheet_if_present(wb, "Rising & Falling Stars")
    remove_sheet_if_present(wb, "Recent Movement Signals")
    star_rows = [
        row
        for row in rows
        if row["prominent"]
        and row["momentum"]
        in {RecentMovement.UPWARD.value, RecentMovement.DOWNWARD.value}
    ]
    headers = [
        "Category",
        "Action",
        "Location",
        "Server",
        "Current Sample",
        "Peer Comparison",
        "Recent Movement",
        "8-Week Direction",
        "History Used",
        "Positive Drivers",
        "Watch Drivers",
        "Manager Question",
    ]
    data = [
        [
            f"{row['momentum']} Movement Signal",
            row["action"],
            row["location"],
            excel_safe_text(row["display_name"]),
            operational_sample_text(row),
            row["performance_level"],
            row["momentum"],
            row["long_term_direction"],
            row["history_used"],
            "; ".join(row["positive_drivers"]),
            "; ".join(row["negative_drivers"]),
            row["recommended_next_step"],
        ]
        for row in star_rows
    ]
    ws = write_table_sheet(
        wb,
        "Recent Movement Signals",
        headers,
        data,
        "RecentMovementSignalsV3",
        widths={
            "A": 26, "B": 24, "C": 20, "D": 24, "E": 22, "F": 23,
            "G": 18, "H": 18, "I": 54, "J": 40, "K": 40, "L": 48,
        },
    )
    style_management_title(ws, "Recent Movement Signals", len(headers))
    ws.freeze_panes = "E4"
    ws.sheet_view.zoomScale = 80
    ws.print_title_rows = "1:3"
    ws.row_dimensions[3].height = 30
    for row in range(4, 4 + len(data)):
        for col in (1, 2, 6, 7, 8):
            fill = priority_fill(ws.cell(row=row, column=col).value)
            if fill:
                ws.cell(row=row, column=col).fill = fill
        for col in (5, 9, 10, 11, 12):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 66


def format_management_value(field: str, value: float | None) -> tuple[Any, str]:
    if value is None:
        return None, "General"
    if field == "gross_sales":
        return value, "$#,##0"
    if field in {"guest_count", "check_count"}:
        return value, "#,##0"
    if field in {"check_average", "sales_per_check"}:
        return value, "$0.00"
    if field == "guests_per_check":
        return value, "0.00"
    if field == "wine_pct":
        return value, "0.0%"
    if field == "rate_of_sale_by_guest_count":
        return value, "0.000"
    if field == "average_ticket_time_seconds":
        return duration_fraction(value), "[h]:mm"
    return value, "0.00"


def management_metric_status(field: str, item: dict[str, Any], config: dict[str, Any] | None = None) -> str:
    config = config or DEFAULT_CONFIG
    if field in SERVER_CONTEXT_FIELDS:
        return (
            "Context Only"
            if management_metric_available(item["latest"], field)
            else "Unavailable"
        )
    current = item["latest"][field]
    benchmark = item["benchmark_values"].get(field)
    if benchmark is None:
        return "No Baseline"
    if field in {"gross_sales", "guest_count"}:
        delta = safe_pct_delta(current, benchmark)
        threshold_key = "sales_pct" if field == "gross_sales" else "guest_pct"
        threshold = float(config.get("management_materiality", DEFAULT_CONFIG["management_materiality"])[threshold_key])
        return "Above" if delta is not None and delta >= threshold else "Watch" if delta is not None and delta <= -threshold else "On Track"
    threshold = management_threshold(config, field)
    variance = directional_variance(current, benchmark, threshold)
    neutral = float(threshold["neutral"])
    return "Above" if variance >= neutral else "Watch" if variance <= -neutral else "On Track"


def write_store_group_scorecards_sheet(
    wb: Workbook,
    rows: list[dict[str, Any]],
    config: dict[str, Any],
    weekly_location_rows: list[dict[str, Any]],
    weekly_group_rows: list[dict[str, Any]],
    readiness: LatestWeekReadiness | None = None,
) -> None:
    remove_sheet_if_present(wb, "Store & Group Scorecards")
    ws = wb.create_sheet("Store & Group Scorecards")
    style_management_title(ws, "Store & Group Scorecards", 7)
    ws.freeze_panes = "A4"
    current_row = 4
    scorecard_start_rows: list[int] = []
    print_section_start_rows: list[int] = []
    guest_chart_row: int | None = None
    display_items: list[tuple[str, dict[str, Any] | None, str]] = []
    if readiness is None:
        display_items = [(item["entity"], item, "current") for item in rows]
    else:
        by_entity = {item["entity"]: item for item in rows}
        configured = set(readiness.configured_locations)
        for item in rows:
            if item["entity"] not in configured:
                state = "current" if readiness.ready else "preliminary"
                display_items.append((item["entity"], item, state))
        for location in readiness.configured_locations:
            item = by_entity.get(location)
            item_week_end = as_date(item.get("latest", {}).get("week_end")) if item else None
            if item is None or (
                item_week_end is not None and item_week_end != readiness.latest_week_end
            ):
                display_items.append((location, None, "missing"))
            else:
                state = "current" if readiness.ready else "preliminary"
                display_items.append((location, item, state))

    for entity, item, readiness_state in display_items:
        scorecard_start_rows.append(current_row)
        print_section_start_rows.append(current_row)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        if readiness_state == "missing":
            title_text = f"{entity} | Missing | No current-week data; comparisons are paused."
            title_priority = "Review"
        elif readiness_state == "preliminary":
            missing_text = readiness.missing_text if readiness else "latest-week data"
            title_text = f"{entity} | Preliminary | Comparisons paused; missing {missing_text or 'latest-week data'}."
            title_priority = "Review"
        else:
            title_text = f"{entity} | {item['status']} | {item['recommended_focus']}"
            title_priority = item["priority"]
        title = ws.cell(row=current_row, column=1, value=title_text)
        title.fill = priority_fill(title_priority) or PatternFill("solid", fgColor="E7E6E6")
        title.font = Font(bold=True)
        title.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[current_row].height = 54
        for col, header in enumerate(["Metric", "Current", "vs Prior", "vs Benchmark", "Benchmark", "Source", "Status"], start=1):
            cell = ws.cell(row=current_row + 1, column=col, value=header)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        ws.row_dimensions[current_row + 1].height = 30
        for offset, (field, label, _) in enumerate(
            STORE_GROUP_DISPLAY_METRICS, start=2
        ):
            row = current_row + offset
            ws.cell(row=row, column=1, value=label)
            if readiness_state == "missing":
                ws.cell(row=row, column=6, value="No current-week data")
                ws.cell(row=row, column=7, value="Missing")
                fill = priority_fill("Review")
                if fill:
                    ws.cell(row=row, column=7).fill = fill
                continue

            is_workload_context = field in {
                "check_count",
                "sales_per_check",
                "guests_per_check",
            }
            context_available = management_metric_available(item["latest"], field)
            current_raw = item["latest"].get(field) if context_available else None
            current_value, number_format = format_management_value(
                field, current_raw
            )
            if is_workload_context:
                prior_context_available = bool(
                    item.get("prior")
                    and management_metric_available(item["prior"], field)
                )
                prior_change = (
                    float(item["latest"].get(field, 0) or 0)
                    - float(item["prior"].get(field, 0) or 0)
                    if context_available and prior_context_available
                    else None
                )
                benchmark_change = None
                benchmark_value, benchmark_format = None, "General"
            else:
                prior_change = item["prior_changes"][field]
                benchmark_change = item["benchmark_changes"][field]
                benchmark_value, benchmark_format = format_management_value(
                    field, item["benchmark_values"][field]
                )
            ws.cell(row=row, column=2, value=current_value).number_format = number_format
            if readiness_state == "preliminary":
                ws.cell(row=row, column=5, value=benchmark_value).number_format = benchmark_format
                source = (
                    "Preliminary workload context"
                    if is_workload_context
                    else f"Preliminary; {item['benchmark_sources'][field]}"
                )
                ws.cell(row=row, column=6, value=source)
                status = "Preliminary"
            elif is_workload_context:
                ws.cell(row=row, column=3, value=prior_change).number_format = number_format
                ws.cell(row=row, column=6, value="Current workload context")
                status = "Context Only" if context_available else "Unavailable"
            elif field in {"gross_sales", "guest_count"}:
                prior_value = safe_pct_delta(item["latest"][field], item["prior"][field] if item["prior"] else None)
                benchmark_delta = safe_pct_delta(item["latest"][field], item["benchmark_values"][field])
                ws.cell(row=row, column=3, value=prior_value).number_format = "0.0%"
                ws.cell(row=row, column=4, value=benchmark_delta).number_format = "0.0%"
            elif field == "average_ticket_time_seconds":
                ws.cell(row=row, column=3, value=prior_change / 60 if prior_change is not None else None).number_format = "0.0"
                ws.cell(row=row, column=4, value=benchmark_change / 60 if benchmark_change is not None else None).number_format = "0.0"
            else:
                ws.cell(row=row, column=3, value=prior_change).number_format = number_format
                ws.cell(row=row, column=4, value=benchmark_change).number_format = number_format
            if readiness_state == "current" and not is_workload_context:
                ws.cell(row=row, column=5, value=benchmark_value).number_format = benchmark_format
                ws.cell(row=row, column=6, value=item["benchmark_sources"][field])
                status = management_metric_status(field, item, config)
            ws.cell(row=row, column=7, value=status)
            fill = priority_fill(
                "Review" if status == "Preliminary" else
                "Review" if status == "Unavailable" else
                "Medium" if status == "Watch" else
                "Recognize" if status == "Above" else "Monitor"
            )
            if fill:
                ws.cell(row=row, column=7).fill = fill
            for col in range(1, 8):
                ws.cell(row=row, column=col).alignment = Alignment(
                    wrap_text=col in {1, 6, 7},
                    vertical="center",
                )
            ws.row_dimensions[row].height = 30
        current_row += len(STORE_GROUP_DISPLAY_METRICS) + 4
        if entity == "All Stores" and weekly_group_rows:
            guest_chart_row = current_row
            print_section_start_rows.append(current_row)
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=7,
            )
            ws.cell(
                row=current_row,
                column=1,
                value="All Stores | Complete-Week Guest Trend",
            )
            ws.cell(row=current_row, column=1).fill = PatternFill(
                "solid", fgColor="D9EAF7"
            )
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
            ws.row_dimensions[current_row].height = 54
            ws.merge_cells(
                start_row=current_row + 1,
                start_column=1,
                end_row=current_row + 6,
                end_column=7,
            )
            ws.cell(
                row=current_row + 1,
                column=1,
                value=(
                    "Group-level operational context only. Use this trend to ask "
                    "what changed in demand, events, hours, or capacity; do not "
                    "attribute the movement to an individual."
                ),
            )
            ws.cell(row=current_row + 1, column=1).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
            for row_index in range(current_row + 1, current_row + 13):
                ws.row_dimensions[row_index].height = 30
            current_row += len(STORE_GROUP_DISPLAY_METRICS) + 4
    for column, width in {"A": 24, "B": 16, "C": 16, "D": 18, "E": 16, "F": 20, "G": 14}.items():
        ws.column_dimensions[column].width = width
    ws.column_dimensions["H"].width = 3
    ws.sheet_view.zoomScale = 80
    for section_start in print_section_start_rows[1:]:
        ws.row_breaks.append(Break(id=section_start - 1))
    sales_chart_row = scorecard_start_rows[0] if scorecard_start_rows else 4
    if guest_chart_row is None:
        guest_chart_row = sales_chart_row + len(STORE_GROUP_DISPLAY_METRICS) + 4
    add_management_scorecard_charts(
        wb,
        ws,
        weekly_location_rows,
        weekly_group_rows,
        sales_anchor=f"I{sales_chart_row}",
        guest_anchor=f"I{guest_chart_row}",
    )


def latest_location_completeness(
    weekly_location_rows: list[dict[str, Any]],
    latest_week_end: date | None,
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[str], list[str], bool]:
    latest_rows = [
        row for row in weekly_location_rows if row["week_end"] == latest_week_end
    ]
    latest_by_location = {row["location"]: row for row in latest_rows}
    configured_locations = list(config.get("locations", {}))
    if not configured_locations:
        configured_locations = sorted(latest_by_location)
    location_gaps = [
        location
        for location in configured_locations
        if location not in latest_by_location
        or latest_by_location[location].get("source_days", 0) < OPERATING_WEEK_DAYS
    ]
    complete = bool(configured_locations) and not location_gaps
    return latest_rows, configured_locations, location_gaps, complete


def recent_completeness_week_ends(
    weekly_location_rows: list[dict[str, Any]],
    max_weeks: int = DATA_QUALITY_COMPLETENESS_WEEKS,
) -> list[date]:
    """Return a latest-first, continuous completeness window without prehistory."""
    observed_week_ends = {
        parsed
        for row in weekly_location_rows
        if (parsed := as_date(row.get("week_end"))) is not None
    }
    if not observed_week_ends or max_weeks < 1:
        return []
    earliest_week_end = min(observed_week_ends)
    latest_week_end = max(observed_week_ends)
    available_span = ((latest_week_end - earliest_week_end).days // 7) + 1
    return [
        latest_week_end - timedelta(days=7 * offset)
        for offset in range(min(max_weeks, available_span))
    ]


def write_data_quality_completeness_heatmap(
    ws,
    weekly_location_rows: list[dict[str, Any]],
    configured_locations: list[str],
    *,
    start_row: int,
) -> int:
    """Write a compact categorical matrix and return its final occupied row."""
    week_ends = recent_completeness_week_ends(weekly_location_rows)
    locations = list(configured_locations) or sorted(
        {
            str(row.get("location") or "").strip()
            for row in weekly_location_rows
            if str(row.get("location") or "").strip()
        }
    )
    title_row = start_row
    note_row = start_row + 1
    header_row = start_row + 2
    data_start_row = start_row + 3
    heatmap_end_column = max(12, 1 + len(locations))

    ws.merge_cells(
        start_row=title_row,
        start_column=1,
        end_row=title_row,
        end_column=heatmap_end_column,
    )
    week_label = (
        f"Latest {len(week_ends)} Week{'s' if len(week_ends) != 1 else ''}"
        if week_ends
        else "No Weekly History"
    )
    ws.cell(
        row=title_row,
        column=1,
        value=f"Source Completeness - {week_label}",
    )
    ws.cell(row=title_row, column=1).fill = PatternFill(
        "solid", fgColor="D9EAF7"
    )
    ws.cell(row=title_row, column=1).font = Font(bold=True, color="7A1E1E")

    ws.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=heatmap_end_column,
    )
    ws.cell(
        row=note_row,
        column=1,
        value=(
            f"Each cell represents one {OPERATING_WEEK_LABEL} location-week. "
            f"Complete = {OPERATING_WEEK_DAYS} source days; Partial = 1-"
            f"{OPERATING_WEEK_DAYS - 1}; Missing = no usable location-week row. "
            "Latest week appears first."
        ),
    )
    ws.cell(row=note_row, column=1).fill = PatternFill(
        "solid", fgColor="F7F7F7"
    )
    ws.cell(row=note_row, column=1).font = Font(color="595959", size=9)
    ws.cell(row=note_row, column=1).alignment = Alignment(
        wrap_text=True, vertical="center"
    )
    ws.row_dimensions[note_row].height = 30

    headers = ["Week Ending", *locations]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[header_row].height = 30

    if not week_ends or not locations:
        ws.merge_cells(
            start_row=data_start_row,
            start_column=1,
            end_row=data_start_row,
            end_column=heatmap_end_column,
        )
        ws.cell(
            row=data_start_row,
            column=1,
            value="No weekly location history is available for the completeness view.",
        )
        ws.cell(row=data_start_row, column=1).alignment = Alignment(
            wrap_text=True, vertical="center"
        )
        return data_start_row

    row_lookup = {
        (as_date(row.get("week_end")), str(row.get("location") or "")): row
        for row in weekly_location_rows
    }
    status_styles = {
        "Complete": ("D9EAD3", "274E13"),
        "Partial": ("FFF2CC", "7F6000"),
        "Missing": ("F4CCCC", "9C0006"),
    }
    matrix_border = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )
    for row_index, week_end in enumerate(week_ends, start=data_start_row):
        date_cell = ws.cell(row=row_index, column=1, value=week_end)
        date_cell.number_format = EXCEL_DATE_NUMBER_FORMAT
        date_cell.fill = PatternFill("solid", fgColor="F3F4F6")
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.border = matrix_border
        for col, location in enumerate(locations, start=2):
            item = row_lookup.get((week_end, location))
            source_days = int(item.get("source_days", 0) or 0) if item else 0
            status = (
                "Complete"
                if source_days >= OPERATING_WEEK_DAYS
                else "Partial"
                if source_days > 0
                else "Missing"
            )
            fill_color, font_color = status_styles[status]
            cell = ws.cell(row=row_index, column=col, value=status)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(bold=True, color=font_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = matrix_border
        ws.row_dimensions[row_index].height = 24
    return data_start_row + len(week_ends) - 1


def write_management_data_quality_sheet(
    wb: Workbook,
    weekly_location_rows: list[dict[str, Any]],
    config: dict[str, Any],
    readiness: LatestWeekReadiness | None = None,
    owner_warnings: list[dict[str, Any]] | None = None,
    records: list[MetricRecord] | None = None,
) -> None:
    remove_sheet_if_present(wb, "Data Quality")
    ws = wb.create_sheet("Data Quality")
    style_management_title(ws, "Data Quality", 8)
    if readiness is None:
        latest_week_end = max((row["week_end"] for row in weekly_location_rows), default=None)
        latest_rows, configured_locations, _location_gaps, latest_complete = latest_location_completeness(
            weekly_location_rows, latest_week_end, config
        )
    else:
        latest_week_end = readiness.latest_week_end
        latest_rows = list(readiness.latest_location_rows)
        configured_locations = list(readiness.configured_locations)
        latest_complete = readiness.ready
    latest_by_location = {row["location"]: row for row in latest_rows}
    ws.merge_cells("A3:L3")
    ws["A3"] = (
        f"Latest week ending {latest_week_end:%m/%d/%Y} is complete and suitable for management trends."
        if latest_complete and latest_week_end
        else "Latest week is incomplete. Management trends are preliminary and server actions are suppressed."
    )
    report_audit = config.get("_report_audit", {})
    if report_audit:
        ws["A3"] = (
            f"{ws['A3'].value} Input audit: "
            f"{report_audit.get('unique_business_days', 0)} unique business days used; "
            f"{report_audit.get('duplicate_files_ignored', 0)} semantic duplicate files ignored; "
            "no report conflicts."
        )
    ws["A3"].fill = PatternFill("solid", fgColor="D9EAD3" if latest_complete else "F4CCCC")
    ws["A3"].font = Font(bold=True)
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 48
    for col, header in enumerate(
        [
            "Latest Week",
            "Location",
            "Active Days",
            "Source Days",
            "Check Count",
            "Ticket Time Basis",
            "Status",
            "Management Use",
        ],
        start=1,
    ):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
    for row_index, location in enumerate(configured_locations, start=6):
        row = latest_by_location.get(location)
        status = (
            "Missing" if row is None
            else "Complete" if row.get("source_days", 0) >= OPERATING_WEEK_DAYS
            else "Short Week"
        )
        values = [
            latest_week_end,
            location,
            row.get("active_days", 0) if row else 0,
            row.get("source_days", 0) if row else 0,
            (
                row.get("check_count", 0)
                if row and row.get("check_count_available")
                else "Unavailable"
            ),
            (
                row.get(
                    "ticket_time_weight_basis",
                    "Unavailable (Check Count missing or incomplete)",
                )
                if row
                else "Unavailable"
            ),
            status,
            (
                "People review eligible; operational context complete"
                if status == "Complete"
                and row
                and row.get("check_count_available")
                and row.get("ticket_time_available")
                and row.get("rate_available")
                else "People review eligible; Rate of Sale unavailable"
                if status == "Complete"
                and row
                and row.get("check_count_available")
                and row.get("ticket_time_available")
                else "People review eligible; Ticket Time unavailable"
                if status == "Complete"
                and row
                and row.get("check_count_available")
                else "People review eligible; check-based context unavailable"
                if status == "Complete"
                else "Preliminary only"
            ),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)
            ws.cell(row=row_index, column=col).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
        ws.cell(row=row_index, column=1).number_format = EXCEL_DATE_NUMBER_FORMAT
        if row and row.get("check_count_available"):
            ws.cell(row=row_index, column=5).number_format = "#,##0"
        fill = priority_fill("Recognize" if status == "Complete" else "High")
        if fill:
            ws.cell(row=row_index, column=7).fill = fill
        ws.row_dimensions[row_index].height = 72

    heatmap_start_row = max(10, 8 + len(configured_locations))
    heatmap_end = write_data_quality_completeness_heatmap(
        ws,
        weekly_location_rows,
        configured_locations,
        start_row=heatmap_start_row,
    )
    historical_title_row = heatmap_end + 2
    historical_header_row = historical_title_row + 1
    historical_data_start = historical_title_row + 2
    ws.cell(row=historical_title_row, column=1, value="Historical Exceptions")
    ws.cell(row=historical_title_row, column=1).font = Font(bold=True)
    ws.cell(row=historical_title_row, column=1).fill = PatternFill(
        "solid", fgColor="E7E6E6"
    )
    historical = [
        row for row in weekly_location_rows if row.get("source_days", 0) < OPERATING_WEEK_DAYS
    ]
    for col, header in enumerate(["Week End", "Location", "Active Days", "Source Days", "Status", "Benchmark Treatment"], start=1):
        cell = ws.cell(row=historical_header_row, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.font = Font(bold=True)
    for row_index, row in enumerate(
        sorted(historical, key=lambda item: (item["week_end"], item["location"])),
        start=historical_data_start,
    ):
        values = [row["week_end"], row["location"], row["active_days"], row["source_days"], "Short Week", "Excluded"]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)
        ws.cell(row=row_index, column=1).number_format = EXCEL_DATE_NUMBER_FORMAT
    warning_start = max(
        historical_title_row + 4,
        historical_title_row + 3 + len(historical),
    )
    ws.merge_cells(
        start_row=warning_start,
        start_column=1,
        end_row=warning_start,
        end_column=6,
    )
    ws.cell(row=warning_start, column=1, value="Owner Assignment Review")
    ws.cell(row=warning_start, column=1).font = Font(bold=True)
    ws.cell(row=warning_start, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
    warning_headers = [
        "Owner", "Action ID", "Location", "Person / Area", "Status", "Reason"
    ]
    for col, header in enumerate(warning_headers, start=1):
        cell = ws.cell(row=warning_start + 1, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.font = Font(bold=True)
    warnings = owner_warnings or []
    if not warnings:
        ws.merge_cells(
            start_row=warning_start + 2,
            start_column=1,
            end_row=warning_start + 2,
            end_column=6,
        )
        ws.cell(
            row=warning_start + 2,
            column=1,
            value="No open actions are assigned to inactive or unlisted owners.",
        )
    else:
        for row_index, warning in enumerate(warnings, start=warning_start + 2):
            for col, header in enumerate(warning_headers, start=1):
                ws.cell(
                    row=row_index,
                    column=col,
                    value=excel_safe_text(warning.get(header)),
                )
            ws.cell(row=row_index, column=6).fill = PatternFill(
                "solid", fgColor="FFF2CC"
            )

    source_start = warning_start + max(4, len(warnings) + 3)
    print_note_row = source_start - 1
    ws.merge_cells(
        start_row=print_note_row,
        start_column=1,
        end_row=print_note_row,
        end_column=12,
    )
    ws.cell(
        row=print_note_row,
        column=1,
        value=(
            "Compact print view ends here. Source-by-source parsing and provenance "
            "remain available below in the protected workbook."
        ),
    )
    ws.cell(row=print_note_row, column=1).fill = PatternFill(
        "solid", fgColor="D9EAF7"
    )
    ws.cell(row=print_note_row, column=1).alignment = Alignment(
        wrap_text=True, vertical="center"
    )
    ws.row_dimensions[print_note_row].height = 30
    ws.merge_cells(
        start_row=source_start,
        start_column=1,
        end_row=source_start,
        end_column=8,
    )
    ws.cell(
        row=source_start,
        column=1,
        value="Source Parsing and Business-Date Provenance",
    )
    ws.cell(row=source_start, column=1).font = Font(bold=True)
    ws.cell(row=source_start, column=1).fill = PatternFill(
        "solid", fgColor="D9E1F2"
    )
    source_headers = [
        "Source File",
        "Report Date",
        "Date Source",
        "Format",
        "Parser",
        "SHA-256",
        "Locations",
        "Status",
    ]
    for col, header in enumerate(source_headers, start=1):
        cell = ws.cell(row=source_start + 1, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="F3F4F6")
        cell.font = Font(bold=True)
    source_groups: dict[str, list[MetricRecord]] = defaultdict(list)
    for record in records or []:
        source_groups[record.source_file].append(record)
    for row_index, (source_file, source_records) in enumerate(
        sorted(source_groups.items()), start=source_start + 2
    ):
        first_record = source_records[0]
        locations = sorted({record.location for record in source_records})
        hashes = {record.source_sha256 for record in source_records if record.source_sha256}
        date_sources = {
            record.report_date_source for record in source_records
        }
        parsers = {record.parser_engine for record in source_records if record.parser_engine}
        formats = {record.source_format for record in source_records if record.source_format}
        unavailable_metrics = sorted(
            {
                metric
                for record in source_records
                if (
                    record.guest_count > 0
                    or record.gross_sales > 0
                    or (record.check_count_available and record.check_count > 0)
                )
                for metric, available in (
                    ("Rate of Sale", record.rate_available),
                    ("Check Count", record.check_count_available),
                    (
                        "Ticket Time",
                        record.ticket_time_available
                        and record.check_count_available,
                    ),
                )
                if not available
            }
        )
        provenance_verified = (
            len(hashes) == 1
            and len(date_sources) == 1
            and len(parsers) == 1
            and len(formats) == 1
        )
        status = (
            "Metric unavailable: " + ", ".join(unavailable_metrics)
            if unavailable_metrics
            else "Verified"
            if provenance_verified
            else "Review"
        )
        values = [
            excel_safe_text(source_file),
            first_record.report_date,
            "; ".join(sorted(date_sources)),
            ", ".join(sorted(formats)) or "Unknown",
            ", ".join(sorted(parsers)) or "Unknown",
            next(iter(hashes)) if len(hashes) == 1 else "Unavailable",
            ", ".join(locations),
            status,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)
            ws.cell(row=row_index, column=col).alignment = Alignment(
                wrap_text=col in {1, 3, 5, 7, 8},
                vertical="center",
            )
        ws.cell(row=row_index, column=2).number_format = EXCEL_DATE_NUMBER_FORMAT
        if status != "Verified":
            ws.cell(row=row_index, column=8).fill = PatternFill(
                "solid", fgColor="FFF2CC"
            )
        ws.row_dimensions[row_index].height = 42

    for column, width in {
        "A": 44,
        "B": 16,
        "C": 34,
        "D": 13,
        "E": 18,
        "F": 69,
        "G": 22,
        "H": 42,
    }.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"
    print_end_column = get_column_letter(
        max(12, 1 + len(configured_locations))
    )
    ws.print_area = f"A1:{print_end_column}{print_note_row}"


def write_management_chart_data(
    wb: Workbook,
    weekly_location_rows: list[dict[str, Any]],
    weekly_group_rows: list[dict[str, Any]],
) -> tuple[int, int]:
    remove_sheet_if_present(wb, "_Dashboard Chart Data")
    ws = wb.create_sheet("_Dashboard Chart Data")
    locations = sorted({row["location"] for row in weekly_location_rows})
    _, global_full_week_ends = full_week_ends_by_location(weekly_location_rows)
    week_ends = sorted(global_full_week_ends)[-8:]
    ws.cell(row=1, column=1, value="Week End")
    for col, location in enumerate(locations, start=2):
        ws.cell(row=1, column=col, value=location.replace("RC ", ""))
    location_lookup = {
        (row["week_end"], row["location"]): row for row in weekly_location_rows
    }
    for row_index, week_end in enumerate(week_ends, start=2):
        ws.cell(row=row_index, column=1, value=week_end.strftime("%m/%d"))
        for col, location in enumerate(locations, start=2):
            row = location_lookup.get((week_end, location))
            ws.cell(row=row_index, column=col, value=row["gross_sales"] if row else None)

    group_lookup = {row["week_end"]: row for row in weekly_group_rows}
    group_rows = [group_lookup[week_end] for week_end in week_ends if week_end in group_lookup]
    ws.cell(row=1, column=6, value="Week End")
    ws.cell(row=1, column=7, value="Guest Count")
    for row_index, row in enumerate(group_rows, start=2):
        ws.cell(row=row_index, column=6, value=row["week_end"].strftime("%m/%d"))
        ws.cell(row=row_index, column=7, value=row["guest_count"])
    ws.sheet_state = "veryHidden"
    return len(week_ends), len(group_rows)


def add_management_scorecard_charts(
    wb: Workbook,
    ws,
    weekly_location_rows: list[dict[str, Any]],
    weekly_group_rows: list[dict[str, Any]],
    *,
    sales_anchor: str = "I4",
    guest_anchor: str = "I20",
) -> None:
    location_points, group_points = write_management_chart_data(
        wb, weekly_location_rows, weekly_group_rows
    )
    chart_ws = wb["_Dashboard Chart Data"]
    if location_points:
        locations = sorted({row["location"] for row in weekly_location_rows})
        chart = LineChart()
        chart.title = "Complete-Week Sales Trend by Store"
        chart.y_axis.title = "Gross Sales"
        chart.y_axis.numFmt = "$#,##0"
        chart.x_axis.tickLblPos = "low"
        chart.height = 7
        chart.width = 12.5
        data = Reference(
            chart_ws,
            min_col=2,
            max_col=1 + len(locations),
            min_row=1,
            max_row=1 + location_points,
        )
        cats = Reference(chart_ws, min_col=1, min_row=2, max_row=1 + location_points)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        category_formula = f"'_Dashboard Chart Data'!$A$2:$A${1 + location_points}"
        for series in chart.series:
            series.cat = AxDataSource(strRef=StrRef(f=category_formula))
        for series, location, color in zip(
            chart.series, locations, ("4472C4", "C0504D")
        ):
            series.tx = SeriesLabel(v=location.replace("RC ", ""))
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 28575
        # Bottom placement keeps both store names inside the chart frame when
        # the scorecard is exported or printed at one page wide.
        chart.legend.position = "b"
        configure_generated_chart_excel_defaults(
            chart,
            x_axis_position="b",
            y_axis_position="l",
            line_series=True,
        )
        ws.add_chart(chart, sales_anchor)
    if group_points:
        chart = LineChart()
        chart.title = "Complete-Week All-Stores Guest Trend"
        chart.y_axis.title = "Guests"
        chart.y_axis.numFmt = "#,##0"
        chart.x_axis.tickLblPos = "low"
        chart.height = 7
        chart.width = 12.5
        data = Reference(chart_ws, min_col=7, min_row=1, max_row=1 + group_points)
        cats = Reference(chart_ws, min_col=6, min_row=2, max_row=1 + group_points)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        category_formula = f"'_Dashboard Chart Data'!$F$2:$F${1 + group_points}"
        for series in chart.series:
            series.cat = AxDataSource(strRef=StrRef(f=category_formula))
        if chart.series:
            chart.series[0].graphicalProperties.line.solidFill = "7A1E1E"
            chart.series[0].graphicalProperties.line.width = 28575
        chart.legend = None
        configure_generated_chart_excel_defaults(
            chart,
            x_axis_position="b",
            y_axis_position="l",
            line_series=True,
        )
        ws.add_chart(chart, guest_anchor)


def signed_management_delta(field: str, value: float | None) -> str:
    if value is None:
        return "n/a"
    if field == "gross_sales":
        return compact_money(value, signed=True)
    if field == "guest_count":
        return compact_count(value, signed=True)
    if field == "check_average":
        return compact_money(value, signed=True)
    if field == "wine_pct":
        return compact_pct_points(value)
    if field == "rate_of_sale_by_guest_count":
        return f"{value:+.3f}"
    return compact_minutes(value / 60)


def latest_week_report_coverage(
    records: list[MetricRecord], latest_week_end: date | None
) -> tuple[list[date], set[date], list[date]]:
    if latest_week_end is None:
        return [], set(), []
    week_start = latest_week_end - timedelta(days=OPERATING_WEEK_DAYS - 1)
    expected = [week_start + timedelta(days=offset) for offset in range(OPERATING_WEEK_DAYS)]
    received = {
        record.report_date
        for record in records
        if week_start <= record.report_date <= latest_week_end
        and is_operating_day(record.report_date)
    }
    missing = [report_date for report_date in expected if report_date not in received]
    return expected, received, missing


def latest_week_readiness(
    records: list[MetricRecord],
    weekly_location_rows: list[dict[str, Any]],
    config: dict[str, Any],
) -> LatestWeekReadiness:
    latest_week_end = max((row["week_end"] for row in weekly_location_rows), default=None)
    latest_rows, configured_locations, location_gaps, locations_complete = (
        latest_location_completeness(weekly_location_rows, latest_week_end, config)
    )
    expected_dates, received_dates, missing_dates = latest_week_report_coverage(
        records, latest_week_end
    )
    return LatestWeekReadiness(
        latest_week_end=latest_week_end,
        latest_location_rows=tuple(latest_rows),
        configured_locations=tuple(configured_locations),
        location_gaps=tuple(location_gaps),
        expected_dates=tuple(expected_dates),
        received_dates=frozenset(received_dates),
        missing_dates=tuple(missing_dates),
        ready=locations_complete and not missing_dates,
    )


def deduplicated_dashboard_actions(
    action_rows: list[dict[str, Any]],
    priorities: set[str],
    limit: int | None = None,
) -> list[dict[str, Any]]:
    priority_order = {"High": 0, "Medium": 1, "Review": 2, "Recognize": 3, "Share": 4}
    candidates = [
        row
        for row in action_rows
        if str(row.get("Priority") or "") in priorities
        and str(row.get("Status") or "Open").casefold() not in {"complete", "dismissed"}
    ]
    candidates.sort(
        key=lambda row: (
            priority_order.get(str(row.get("Priority") or ""), 9),
            1
            if str(row.get("Entity Key") or "").casefold().startswith("group|")
            else 0,
            str(row.get("Location") or ""),
            str(row.get("Person / Area") or ""),
        )
    )
    selected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in candidates:
        identity = str(row.get("Entity Key") or "").strip().casefold()
        if not identity:
            identity = "|".join(
                str(row.get(field) or "").strip().casefold()
                for field in ("Location", "Person / Area", "Action", "Signal")
            )
        if identity in seen:
            continue
        seen.add(identity)
        selected.append(row)
        if limit is not None and len(selected) >= limit:
            break
    return selected


def dashboard_management_move(item: dict[str, Any]) -> str:
    next_step = (
        str(item.get("Recommended Next Step") or "").strip().rstrip(".?")
    )
    if next_step:
        return f"{next_step}?"
    return "What evidence and operating context need review, and who owns the follow-up?"


def write_dashboard_card(
    ws,
    start_col: int,
    title: str,
    value: str | int,
    note: str,
    accent_color: str,
) -> None:
    end_col = start_col + 3
    ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=end_col)
    ws.merge_cells(start_row=7, start_column=start_col, end_row=8, end_column=end_col)
    ws.merge_cells(start_row=9, start_column=start_col, end_row=9, end_column=end_col)
    title_cell = ws.cell(row=6, column=start_col, value=title)
    title_cell.fill = PatternFill("solid", fgColor=accent_color)
    title_cell.font = Font(bold=True, color="FFFFFF", size=10)
    title_cell.alignment = Alignment(
        horizontal="center", vertical="center", shrink_to_fit=True
    )
    value_cell = ws.cell(row=7, column=start_col, value=value)
    value_cell.fill = PatternFill("solid", fgColor="F7F7F7")
    value_cell.font = Font(bold=True, color="7A1E1E", size=20)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    note_cell = ws.cell(row=9, column=start_col, value=note)
    note_cell.fill = PatternFill("solid", fgColor="F7F7F7")
    note_cell.font = Font(color="595959", size=9)
    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, minimum_height in ((6, 20), (7, 24), (8, 24), (9, 32)):
        current_height = ws.row_dimensions[row_index].height or 0
        ws.row_dimensions[row_index].height = max(current_height, minimum_height)


def write_management_dashboard_sheet(
    wb: Workbook,
    records: list[MetricRecord],
    weekly_location_rows: list[dict[str, Any]],
    weekly_group_rows: list[dict[str, Any]],
    server_rows: list[dict[str, Any]],
    store_rows: list[dict[str, Any]],
    group_rows: list[dict[str, Any]],
    action_rows: list[dict[str, Any]],
    config: dict[str, Any] | None = None,
    readiness: LatestWeekReadiness | None = None,
) -> None:
    config = config or DEFAULT_CONFIG
    remove_sheet_if_present(wb, "Dashboard")
    ws = wb.create_sheet("Dashboard")
    style_management_title(ws, "Red Onion Weekly Brief", 12)
    widths = {
        "A": 11, "B": 11, "C": 12, "D": 12, "E": 12, "F": 12,
        "G": 12, "H": 12, "I": 12, "J": 16, "K": 14, "L": 14,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A5"
    readiness = readiness or latest_week_readiness(records, weekly_location_rows, config)
    latest_week_end = readiness.latest_week_end
    latest_location_rows = list(readiness.latest_location_rows)
    configured_locations = list(readiness.configured_locations)
    expected_dates = readiness.expected_dates
    received_dates = readiness.received_dates
    latest_complete = readiness.ready
    missing_text = readiness.missing_text
    current_location_names = {row["location"] for row in latest_location_rows}
    current_store_rows = {
        item["entity"]: item
        for item in store_rows
        if item["entity"] in current_location_names
    }
    _, global_full = full_week_ends_by_location(weekly_location_rows)

    ws.merge_cells("A3:L3")
    ws["A3"] = (
        f"Week ending {latest_week_end:%m/%d/%Y} | {len(received_dates)} of {len(expected_dates)} daily reports | "
        f"{len(global_full)} complete weeks available"
        if latest_week_end
        else "No management data available"
    )
    ws["A3"].font = Font(bold=True, color="595959")
    ws["A3"].alignment = Alignment(vertical="center")
    ws.merge_cells("A4:L4")
    ws["A4"] = (
        "READY FOR HUMAN REVIEW - rule-based prompts are available for context checking"
        if latest_complete
        else f"PRELIMINARY - missing {missing_text or 'daily reports'}; comparisons, actions, and recognition are paused"
    )
    ws["A4"].fill = PatternFill("solid", fgColor="D9EAD3" if latest_complete else "F4CCCC")
    ws["A4"].font = Font(bold=True)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 28

    group = group_rows[0] if group_rows else None
    all_action_items = deduplicated_dashboard_actions(
        action_rows, {"High", "Medium", "Review"}
    )
    selected_actions = all_action_items[:3]
    sales_delta = None
    sales_source = "benchmark"
    if group:
        sales_delta = safe_pct_delta(
            group["latest"]["gross_sales"],
            group["benchmark_values"].get("gross_sales"),
        )
        sales_source = str(
            group.get("benchmark_sources", {}).get("gross_sales") or "benchmark"
        )
    if not latest_complete:
        sales_value, sales_note, sales_color = (
            "PAUSED",
            f"Missing {missing_text or 'daily reports'}",
            "7F7F7F",
        )
    elif sales_delta is None:
        sales_value, sales_note, sales_color = (
            "NO BASELINE",
            "Sales benchmark is not available",
            "7F7F7F",
        )
    else:
        sales_value = f"{sales_delta:+.1%}"
        sales_note = f"Gross sales vs {sales_source.lower()}"
        sales_color = (
            "548235"
            if sales_delta >= 0
            else "C00000"
            if sales_delta <= -0.05
            else "BF9000"
        )
    write_dashboard_card(
        ws, 1, "Sales vs Benchmark", sales_value, sales_note, sales_color
    )

    traffic_delta = None
    traffic_source = "benchmark"
    if group:
        traffic_delta = safe_pct_delta(
            group["latest"]["guest_count"], group["benchmark_values"].get("guest_count")
        )
        traffic_source = str(group.get("benchmark_sources", {}).get("guest_count") or "benchmark")
    if not latest_complete:
        traffic_value, traffic_note, traffic_color = "PAUSED", f"Missing {missing_text or 'daily reports'}", "7F7F7F"
    elif traffic_delta is None:
        traffic_value, traffic_note, traffic_color = "NO BASELINE", "Guest traffic benchmark is not available", "7F7F7F"
    else:
        traffic_value = f"{traffic_delta:+.1%}"
        traffic_note = f"Guest traffic vs {traffic_source.lower()}"
        traffic_color = "548235" if traffic_delta >= 0 else "C00000" if traffic_delta <= -0.05 else "BF9000"
    write_dashboard_card(
        ws, 5, "Guests vs Benchmark", traffic_value, traffic_note, traffic_color
    )

    if latest_complete:
        action_value: str | int = len(all_action_items)
        action_note = "Open coaching-context review items"
        action_color = "C00000" if all_action_items else "548235"
    else:
        action_value, action_note, action_color = "PAUSED", f"Missing {missing_text or 'daily reports'}", "7F7F7F"
    write_dashboard_card(
        ws, 9, "Review Items", action_value, action_note, action_color
    )

    ws.merge_cells("A11:L11")
    ws["A11"] = "TOP THREE REVIEW ITEMS"
    ws["A11"].fill = PatternFill("solid", fgColor="E7E6E6")
    ws["A11"].font = Font(bold=True, color="7A1E1E")
    action_headers = [
        (1, 2, "Priority"), (3, 5, "Person / Location"), (6, 10, "Manager Question"),
        (11, 11, "Owner"), (12, 12, "Due"),
    ]
    for start_col, end_col, header in action_headers:
        if end_col > start_col:
            ws.merge_cells(start_row=12, start_column=start_col, end_row=12, end_column=end_col)
        cell = ws.cell(row=12, column=start_col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if latest_complete:
        for row_index in range(13, 16):
            item = selected_actions[row_index - 13] if row_index - 13 < len(selected_actions) else None
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
            ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=5)
            ws.merge_cells(start_row=row_index, start_column=6, end_row=row_index, end_column=10)
            if item:
                priority = str(item.get("Priority") or "Review")
                person = str(item.get("Person / Area") or "").strip()
                location = str(item.get("Location") or "").strip()
                person_location = person if person.casefold() == location.casefold() else " | ".join(
                    part for part in (person, location) if part
                )
                owner = str(item.get("Owner") or "Unassigned")
                due_date = as_date(item.get("Due Date"))
                ws.cell(row=row_index, column=1, value=excel_safe_text(priority))
                ws.cell(row=row_index, column=3, value=excel_safe_text(person_location))
                ws.cell(
                    row=row_index,
                    column=6,
                    value=excel_safe_text(dashboard_management_move(item)),
                )
                ws.cell(row=row_index, column=11, value=excel_safe_text(owner))
                if due_date:
                    ws.cell(
                        row=row_index, column=12, value=due_date
                    ).number_format = EXCEL_DATE_NUMBER_FORMAT
                else:
                    ws.cell(row=row_index, column=12, value="Not set")
                fill = priority_fill(priority)
                if fill:
                    ws.cell(row=row_index, column=1).fill = fill
                if owner == "Unassigned":
                    ws.cell(row=row_index, column=11).fill = PatternFill("solid", fgColor="FFF2CC")
                if not due_date:
                    ws.cell(row=row_index, column=12).fill = PatternFill("solid", fgColor="FFF2CC")
            else:
                ws.cell(row=row_index, column=1, value="—")
                ws.cell(row=row_index, column=3, value="No additional priority item")
            for col in (1, 3, 6, 11, 12):
                ws.cell(row=row_index, column=col).alignment = Alignment(
                    wrap_text=True, vertical="center"
                )
            ws.row_dimensions[row_index].height = 66
    else:
        ws.merge_cells("A13:L15")
        ws["A13"] = (
            f"Actions are paused until the latest week is complete. Missing: {missing_text or 'daily reports'}."
        )
        ws["A13"].fill = PatternFill("solid", fgColor="FCE8E6")
        ws["A13"].font = Font(bold=True, color="9C0006")
        ws["A13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A17:L17")
    ws["A17"] = "STORE SNAPSHOT"
    ws["A17"].fill = PatternFill("solid", fgColor="E7E6E6")
    ws["A17"].font = Font(bold=True, color="7A1E1E")
    store_headers = [
        (1, 2, "Location"), (3, 4, "Status"), (5, 5, "Sales"), (6, 6, "Guests"),
        (7, 7, "Checks"), (8, 8, "Sales / Guest"), (9, 9, "Sales / Check"),
        (10, 10, "Guests / Check"), (11, 12, "Manager Question"),
    ]
    for start_col, end_col, header in store_headers:
        if end_col > start_col:
            ws.merge_cells(start_row=18, start_column=start_col, end_row=18, end_column=end_col)
        cell = ws.cell(row=18, column=start_col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index in range(19, 21):
        location_index = row_index - 19
        location = (
            configured_locations[location_index]
            if location_index < len(configured_locations)
            else None
        )
        item = current_store_rows.get(location) if location else None
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
        ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=4)
        ws.merge_cells(start_row=row_index, start_column=11, end_row=row_index, end_column=12)
        if item:
            latest = item["latest"]
            status = item["status"] if latest_complete else "Preliminary"
            focus = item["recommended_focus"] if latest_complete else (
                f"Comparisons paused; missing {missing_text or 'daily reports'}."
            )
            ws.cell(row=row_index, column=1, value=item["entity"])
            ws.cell(row=row_index, column=3, value=status)
            ws.cell(row=row_index, column=5, value=latest["gross_sales"]).number_format = "$#,##0"
            ws.cell(row=row_index, column=6, value=latest["guest_count"]).number_format = "#,##0"
            if latest.get("check_count_available"):
                ws.cell(
                    row=row_index, column=7, value=latest.get("check_count", 0)
                ).number_format = "#,##0"
            else:
                ws.cell(row=row_index, column=7, value="n/a")
            if management_metric_available(latest, "sales_per_check"):
                ws.cell(
                    row=row_index, column=9, value=latest.get("sales_per_check", 0)
                ).number_format = "$0.00"
                ws.cell(
                    row=row_index,
                    column=10,
                    value=latest.get("guests_per_check", 0),
                ).number_format = "0.00"
            else:
                for col in (9, 10):
                    ws.cell(row=row_index, column=col, value="n/a")
            ws.cell(
                row=row_index, column=8, value=latest["check_average"]
            ).number_format = "$0.00"
            ws.cell(row=row_index, column=11, value=focus)
            fill = priority_fill(item["priority"] if latest_complete else "Review")
            if fill:
                ws.cell(row=row_index, column=3).fill = fill
        else:
            ws.cell(row=row_index, column=1, value=location or "No store data")
            ws.cell(row=row_index, column=3, value="Missing")
            ws.cell(
                row=row_index,
                column=11,
                value="What source data is missing, and who will confirm it?",
            )
            fill = priority_fill("Review")
            if fill:
                ws.cell(row=row_index, column=3).fill = fill
        for col in (1, 3, 5, 6, 7, 8, 9, 10, 11):
            ws.cell(row=row_index, column=col).alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="center" if col in {5, 6, 7, 8, 9, 10} else None,
            )
        for col in range(5, 11):
            ws.cell(row=row_index, column=col).border = Border(
                right=Side(style="thin", color="D9E1F2")
            )
        ws.row_dimensions[row_index].height = 60

    ws.merge_cells("A22:L22")
    ws["A22"] = "RECOGNITION REVIEW"
    ws["A22"].fill = PatternFill("solid", fgColor="E7E6E6")
    ws["A22"].font = Font(bold=True, color="7A1E1E")
    ws.merge_cells("A23:L23")
    recognition = deduplicated_dashboard_actions(
        action_rows, {"Recognize", "Share"}, limit=1
    )
    if not latest_complete:
        recognition_text = (
            f"Recognition is paused until the latest week is complete. Missing: {missing_text or 'daily reports'}."
        )
        recognition_fill = "FCE8E6"
    elif recognition:
        item = recognition[0]
        person = str(item.get("Person / Area") or item.get("Location") or "Team")
        location = str(item.get("Location") or "")
        identity = person if person.casefold() == location.casefold() else " | ".join(
            part for part in (person, location) if part
        )
        recognition_text = f"{identity} — {dashboard_management_move(item)}"
        recognition_fill = "D9EAD3"
    else:
        recognition_text = "No recognition item met the current complete-week thresholds."
        recognition_fill = "F3F4F6"
    ws["A23"] = excel_safe_text(recognition_text)
    ws["A23"].fill = PatternFill("solid", fgColor=recognition_fill)
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[23].height = 34

    ws.merge_cells("A24:L24")
    check_covered_locations = sum(
        bool(row.get("check_count_available")) for row in latest_location_rows
    )
    ws["A24"] = (
        f"Data quality: {len(received_dates)} of {len(expected_dates)} reports received; "
        f"Check Count complete for {check_covered_locations} of "
        f"{len(latest_location_rows)} current store rows; {len(global_full)} complete "
        "all-store weeks available."
        if latest_complete
        else f"Data quality: {len(received_dates)} of {len(expected_dates)} reports received; missing {missing_text or 'daily reports'}."
    )
    ws["A24"].fill = PatternFill("solid", fgColor="D9EAF7" if latest_complete else "F4CCCC")
    ws["A24"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[24].height = 30

    ws.print_area = "A1:L24"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.sheet_view.zoomScale = 90


def write_management_run_notes(
    wb: Workbook,
    records: list[MetricRecord],
    source_dir: Path,
    public_start: date,
    public_end: date,
    config: dict[str, Any],
) -> None:
    remove_sheet_if_present(wb, "Run Notes")
    ws = wb.create_sheet("Run Notes")
    style_management_title(ws, "Red Onion Server Master", 2)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 96
    ws.freeze_panes = "A3"
    ws.sheet_view.zoomScale = 85
    report_audit = config.get("_report_audit", {})
    integrity = config.get("_integrity", {})
    provenance = integrity.get("provenance", {})
    git_provenance = provenance.get("git", {})
    config_provenance = provenance.get("config", {})
    requirements_provenance = provenance.get("requirements", {})
    business_dates = {record.report_date for record in records}
    check_count_dates = {
        record.report_date
        for record in records
        if record.check_count_available
    }
    note_rows = [
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Run ID", integrity.get("run_id", "Standalone workbook generation")),
        ("Generator Commit", git_provenance.get("commit", "Not available")),
        ("Config SHA-256", config_provenance.get("sha256", integrity.get("effective_config_sha256", "Not available"))),
        ("Requirements SHA-256", requirements_provenance.get("sha256", "Not available")),
        ("Previous Manifest SHA-256", integrity.get("previous_manifest_sha256", "Integrity baseline or standalone generation")),
        ("Digest Scheme", WORKBOOK_DIGEST_SCHEME),
        (WORKBOOK_PROTECTION_CONTRACT_LABEL, WORKBOOK_PROTECTION_CONTRACT),
        (
            WORKBOOK_OPERATOR_UNLOCK_LABEL,
            f"Review > Unprotect Sheet or Workbook uses {WORKBOOK_OPERATOR_PASSWORD}. "
            "This is an accidental-edit guard only; saved generated-content or structure "
            "changes are detected by the digest and can stop the next protected run.",
        ),
        (RUN_NOTES_DIGEST_LABEL, "Pending save/reload validation"),
        ("Source Folder", str(source_dir)),
        ("Operating Week", f"{OPERATING_WEEK_LABEL}; Mondays are closed."),
        ("Raw Reports Read", len({record.source_file for record in records})),
        ("Unique Business Days Used", report_audit.get("unique_business_days", len(business_dates))),
        (
            "Check Count Coverage",
            (
                f"{len(check_count_dates)} of {len(business_dates)} business days. "
                "Ticket Time totals are available only where every active row in the "
                "rollup has Check Count."
            ),
        ),
        ("Semantic Duplicates Ignored", report_audit.get("duplicate_files_ignored", 0)),
        ("Canonical History Folder", report_audit.get("canonical_archive", "Not recorded")),
        ("Report Conflicts", "None; conflicting same-date reports stop the run before output or archiving."),
        ("Date Coverage", format_date_range(min(r.report_date for r in records), max(r.report_date for r in records))),
        ("Public Snapshot Dates", format_date_range(public_start, public_end)),
        ("Recent Movement", f"Latest complete Tuesday-Sunday week versus up to {config.get('dashboard_baseline_full_weeks', 4)} prior complete person-weeks. Partial weeks, limited samples, missing metrics, and insufficient references are Not Evaluated."),
        ("Evidence Status", f"Eligibility requires latest guests >= {config.get('dashboard_min_guest_count_for_trends', 25)} AND active days >= {config.get('dashboard_min_active_days_for_trends', 3)}, plus at least {config.get('dashboard_min_prior_full_weeks', 2)} prior full weeks and {config.get('dashboard_min_prior_guest_count', 50)} prior guests. This is not statistical confidence."),
        ("8-Week Direction", f"Compares the most recent {config.get('dashboard_long_term_block_weeks', 4)} complete weeks with the preceding {config.get('dashboard_long_term_block_weeks', 4)}. Full uses eight server weeks with at least {config.get('dashboard_long_term_full_min_recent_guests', 100)} recent / {config.get('dashboard_long_term_full_min_earlier_guests', 100)} earlier guests; Developing requires at least {config.get('dashboard_long_term_developing_min_total_weeks', 6)} usable weeks, {config.get('dashboard_long_term_developing_min_recent_weeks', 3)} recent / {config.get('dashboard_long_term_developing_min_earlier_weeks', 2)} earlier weeks, and {config.get('dashboard_long_term_developing_min_recent_guests', 75)} recent / {config.get('dashboard_long_term_developing_min_earlier_guests', 50)} earlier guests."),
        ("Peer Comparison", "Person-level comparison uses a leave-one-person-out same-store median pooled across the prior four complete weeks. It fails closed when cohort sufficiency gates are not met. Store totals are never a person fallback."),
        ("Targets", "Management Setup targets apply only to store/group operational context. They do not create person-level coaching prompts."),
        ("Signal Scoring", "Person-level movement and peer comparison use separately calibrated Sales / Guest and Wine % bands only. Check Count, Rate of Sale, and Ticket Time are context only. Rank is display-only and contributes no points. A prompt requires aligned movement and peer context, recurring drivers in two consecutive qualified weeks, a common-store-shock guard, and leave-one-active-day stability."),
        ("Review Gate", "Generated rows start at Review Needed. A later workflow state requires Review Disposition, Reviewed By, and Review Date; the signal cannot be the sole basis for an adverse employment decision."),
        (
            "Team Trends",
            "Shows every current non-excluded server with exact consecutive-week "
            "Sales / Guest and Wine % deltas plus descriptive 4-week and 8-week "
            "direction. These visible movements do not bypass the action gates.",
        ),
        (
            "Performance & Consistency",
            f"Descriptive methodology {PERFORMANCE_CONSISTENCY_METHODOLOGY_VERSION} uses the latest "
            f"{PERFORMANCE_CONSISTENCY_WINDOW_WEEKS} complete shared weeks, current-roster same-store "
            f"weekly medians with at least {config.get('management_peer_reference', {}).get('min_distinct_peers_per_week', 5)} peers, "
            f"a {PERFORMANCE_CONSISTENCY_GUEST_WEIGHT_CAP:,.0f}-guest weekly weight cap, and sample SD. "
            "It measures selling outcomes only and does not bypass readiness, evidence, persistence, "
            "stability, review, or Action Board controls.",
        ),
        (
            "Shared POS & Area Trends",
            "Shared four-digit POS identities are summarized separately from people. Bar, "
            "Patio, Dining Room, Banquets, and Wine Dinners use guest-weighted Gross Sales / "
            "Guests across complete shared weeks. Wine Dinners remains unavailable until an "
            "actual source name or shared number is configured.",
        ),
        ("Technical Trend Detail", "Server Week-over-Week Detail remains protected descriptive audit context. Generated coaching prompts use qualified Recent Movement, peer comparison, evidence stability, and two-week persistence."),
        ("Action Tracking", "Owner, due date, status, context notes, and review fields carry forward between weekly runs. Cleared signals move to Action History."),
        (
            "Evidence Contract",
            "Action and reason codes, exact evidence weeks, source hashes/parser "
            f"provenance, comparator/cohort metadata, review disposition, and metric inputs use methodology {MANAGEMENT_METHODOLOGY_VERSION}.",
        ),
        ("Metric Rule", "Sales / Guest and Wine % are recalculated from rolled-up sales, guests, and wine sales. Sales / Check and Guests / Check are shown only with complete Check Count coverage."),
        ("Metric Caveat", "Red Onion defines Rate of Sale as opportunities divided by qualifying sales, where lower is better; totals use opportunities divided by inferred qualifying sales. Ticket Time is Check Count-weighted only and is unavailable for legacy periods without Check Count. Both are descriptive context and cannot create person-level prompts."),
    ]
    for row, (label, value) in enumerate(note_rows, start=4):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = (
            60
            if label
            in {
                "8-Week Direction",
                "Peer Comparison",
                "Signal Scoring",
                "Metric Caveat",
            }
            else 42
            if label in {"Check Count Coverage", "Metric Rule", "Team Trends", "Performance & Consistency"}
            else 30
        )


def protect_worksheet(ws, password: str) -> None:
    """Enable an accidental-edit boundary while keeping review controls usable."""
    ws.protection.set_password(password)
    ws.protection.sheet = True
    ws.protection.objects = True
    ws.protection.scenarios = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def finalize_management_workbook(wb: Workbook) -> None:
    # This documented convenience password is an accidental-edit guard, not a
    # security boundary. The substantive workbook digest and manifest chain
    # remain the authoritative controls for generated content and structure.
    protection_password = WORKBOOK_OPERATOR_PASSWORD
    visible_management_sheets = tuple(
        name for name in VISIBLE_MANAGEMENT_SHEETS if name in wb.sheetnames
    )
    if not visible_management_sheets:
        # Compatibility for narrow fixture and migration workbooks assembled from
        # the former standalone management surfaces. Strict validation still
        # requires the declared contract for any published workbook.
        visible_management_sheets = tuple(
            name
            for name in ("Management Setup", "Action Board", "Run Notes")
            if name in wb.sheetnames
        )
    for sheet in wb.worksheets:
        if sheet.title in visible_management_sheets:
            sheet.sheet_state = "visible"
            add_management_navigation(sheet)
            configure_management_print_layout(sheet)
        else:
            sheet.sheet_state = "veryHidden"
        protect_worksheet(sheet, protection_password)
    ordered = [wb[name] for name in visible_management_sheets]
    ordered.extend(
        sheet for sheet in wb.worksheets if sheet.title not in visible_management_sheets
    )
    wb._sheets = ordered
    wb.active = 0
    tab_colors = {
        "How to Use": "FFD966",
        "Performance Dashboard": "8B1E23",
        "Server Scorecards": "2F6F6D",
        "Weekly Performance": "3E6FA8",
        "Shared & Area Trends": "8064A2",
        "Methodology": "B77900",
        "Management Center": "4472C4",
        "Dashboard": "7A1E1E",
        "Action Board": "C00000", "Team Trends": "5B9BD5",
        "Store & Group Scorecards": "70AD47",
        "Evidence Detail": "8064A2", "Action History": "A5A5A5",
        "Data Quality": "5B9BD5", "Management Setup": "4472C4",
        "Run Notes": "7F7F7F",
    }
    for name, color in tab_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color
    # Desktop Excel materializes these AutoFit results on its first no-edit save.
    # Write them into generated output up front so the visible row sizes remain
    # part of the substantive digest instead of being treated as serializer noise.
    for sheet_name, row_heights in EXCEL_AUTOFIT_ROW_HEIGHTS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_index, height in row_heights.items():
            if ws.row_dimensions[row_index].height is None:
                ws.row_dimensions[row_index].height = height
    for ws in wb.worksheets:
        for dimension in ws.row_dimensions.values():
            if dimension.height is None:
                continue
            dimension.height = EXCEL_ROW_HEIGHT_SERIALIZATION.get(
                float(dimension.height),
                float(dimension.height),
            )
    wb.security = WorkbookProtection(lockStructure=True, lockWindows=False)
    wb.security.set_workbook_password(protection_password)


def write_master_workbook(
    records: list[MetricRecord],
    output_path: Path,
    config: dict[str, Any],
    source_dir: Path,
    public_start: date,
    public_end: date,
) -> Path:
    integrity_context = config.get("_integrity", {})
    state = read_management_state(
        output_path,
        allow_legacy_protection_upgrade=bool(
            integrity_context.get("allow_legacy_master_upgrade", False)
        ),
        expected_digest=integrity_context.get(
            "expected_master_generated_content_sha256"
        ),
        expected_digest_scheme=integrity_context.get(
            "expected_master_generated_content_digest_scheme"
        ),
        legacy_reference_path=(
            Path(integrity_context["legacy_master_reference_path"])
            if integrity_context.get("legacy_master_reference_path")
            else None
        ),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(f".{output_path.stem}.{os.getpid()}.tmp.xlsx")
    if temp_path.exists():
        temp_path.unlink()
    wb: Workbook | None = None
    try:
        _write_master_workbook_base(records, temp_path, config, source_dir, public_start, public_end)
        wb = load_workbook(temp_path)
        weekly_server_rows, weekly_location_rows = weekly_rollups(records)
        rank_min_guest_count = int(
            config.get("master_min_guest_count_for_rankings", config.get("public_min_guest_count", 1))
        )
        ranked_rows = weekly_server_rank_rows(weekly_server_rows, rank_min_guest_count)
        weekly_group_rows = group_weekly_rows(records)
        readiness = latest_week_readiness(records, weekly_location_rows, config)
        full_by_location, global_full = full_week_ends_by_location(weekly_location_rows)
        server_rows = management_server_rows(
            weekly_server_rows, weekly_location_rows, ranked_rows, state["targets"], config
        )
        performance_consistency_model = build_performance_consistency_model(
            weekly_server_rows, weekly_location_rows, config
        )
        shared_area_trends_model = build_shared_area_trends_model(
            weekly_server_rows, weekly_location_rows, config
        )
        store_rows = management_entity_rows(
            weekly_location_rows, "location", state["targets"], config, full_by_location
        )
        group_rows = management_entity_rows(
            weekly_group_rows, "group", state["targets"], config, global_full
        )
        signals = build_management_action_signals(
            server_rows, store_rows, group_rows, weekly_location_rows, readiness
        )
        current_actions, action_history = merge_management_actions(signals, state, readiness)
        owner_warnings = inactive_owner_assignments(
            current_actions, state["owner_roster"]
        )
        visible_server_rows = server_rows if readiness.ready else []

        if "Data Quality" in wb.sheetnames:
            remove_sheet_if_present(wb, "_Data Quality Detail")
            wb["Data Quality"].title = "_Data Quality Detail"
        for name in (
            "How to Use", "Performance Dashboard", "Server Scorecards", "Weekly Performance", "Shared & Area Trends",
            "Methodology", "Management Center", "_Consistency Calc", "Dashboard", "Action Focus", "Action Board", "Rising & Falling Stars",
            "Recent Movement Signals", "Run Notes",
            "Server Scorecard", "Team Trends", "Store & Group Scorecards", "Action History", "Management Setup",
            "Evidence Detail",
        ):
            remove_sheet_if_present(wb, name)
        write_team_trends_sheet(wb, visible_server_rows)
        write_store_group_scorecards_sheet(
            wb,
            [*group_rows, *store_rows],
            config,
            weekly_location_rows,
            weekly_group_rows,
            readiness,
        )
        write_management_data_quality_sheet(
            wb,
            weekly_location_rows,
            config,
            readiness,
            owner_warnings,
            records=records,
        )
        write_management_center_sheet(
            wb,
            state["targets"],
            state["owner_roster"],
            current_actions,
            action_history,
            config,
            readiness,
            owner_warnings,
            state["owner_roster_capacity"],
        )
        write_evidence_detail_sheet(wb, current_actions, action_history)
        write_owner_validation_sheet(
            wb,
            state["owner_roster"],
            state["owner_roster_capacity"],
        )
        write_performance_consistency_sheets(
            wb,
            performance_consistency_model,
            config,
            readiness,
        )
        write_shared_area_trends_sheet(wb, shared_area_trends_model)
        write_management_run_notes(wb, records, source_dir, public_start, public_end, config)
        write_management_dashboard_sheet(
            wb, records, weekly_location_rows, weekly_group_rows, visible_server_rows,
            store_rows, group_rows, current_actions, config, readiness,
        )
        write_how_to_use_sheet(wb)
        finalize_management_workbook(wb)
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
        wb.save(temp_path)
        wb.close()
        wb = None
        digest = stamp_generated_content_digest(temp_path)
        validate_management_workbook(temp_path, digest)
        os.replace(temp_path, output_path)
    except PermissionError as exc:
        raise RuntimeError(
            f"Could not replace {output_path.name}. Close the master workbook in Excel and rerun; "
            "no source files were moved."
        ) from exc
    finally:
        if wb is not None:
            wb.close()
        if temp_path.exists():
            temp_path.unlink()
    return output_path


def default_integrity_anchor_dir() -> Path:
    """Return machine-local trusted state kept outside the Dropbox operator tree."""

    configured = os.environ.get(INTEGRITY_ANCHOR_ENVIRONMENT_VARIABLE, "").strip()
    if configured:
        return Path(configured).expanduser().absolute()
    if os.name == "nt":
        local_app_data = os.environ.get("LOCALAPPDATA", "").strip()
        base = (
            Path(local_app_data)
            if local_app_data
            else Path.home() / "AppData" / "Local"
        )
        return (base / "RedOnionMetrics" / "integrity-anchors").absolute()
    xdg_state_home = os.environ.get("XDG_STATE_HOME", "").strip()
    base = Path(xdg_state_home) if xdg_state_home else Path.home() / ".local" / "state"
    return (base / "red-onion-weekly-metrics" / "integrity-anchors").absolute()


def integrity_archive_identity(archive_dir: Path) -> str:
    """Return a stable local identity without disclosing the archive path in state."""

    resolved = os.path.normpath(str(archive_dir.resolve()))
    if os.name == "nt":
        resolved = os.path.normcase(resolved)
    return canonical_json_sha256({"archive_directory": resolved})


def integrity_anchor_path(
    archive_dir: Path,
    anchor_dir: Path | None = None,
) -> Path:
    """Resolve the per-archive trusted-head file outside the operator workspace."""

    archive_dir = archive_dir.resolve()
    operations_root = archive_dir.parent
    supplied_root = Path(anchor_dir or default_integrity_anchor_dir()).expanduser().absolute()
    if os.path.lexists(supplied_root) and path_is_link_or_reparse(supplied_root):
        raise IntegrityError(
            "The trusted integrity-anchor folder is a link or reparse point. Restore "
            "the protected machine-local folder before rerunning."
        )
    root = supplied_root.resolve()
    try:
        root.relative_to(operations_root)
    except ValueError:
        pass
    else:
        raise IntegrityError(
            "The trusted integrity-anchor folder must be outside the Dropbox/operator "
            f"workspace: {operations_root}."
        )
    if os.path.lexists(root) and not root.is_dir():
        raise IntegrityError(
            f"The trusted integrity-anchor location is not a folder: {root}."
        )
    anchor = root / f"{integrity_archive_identity(archive_dir)}.json"
    if os.path.lexists(anchor) and path_is_link_or_reparse(anchor):
        raise IntegrityError(
            "The trusted integrity-anchor file is a link or reparse point. Restore the "
            "protected machine-local anchor before rerunning."
        )
    return anchor


def integrity_anchor_exists(
    archive_dir: Path,
    anchor_dir: Path | None = None,
) -> bool:
    """Return whether this archive has ever been pinned on the selected machine."""

    return os.path.lexists(integrity_anchor_path(archive_dir, anchor_dir))


def reject_legacy_raw_mutation_if_protected(
    archive_dir: Path,
    operation: str,
    *,
    anchor_dir: Path | None = None,
) -> None:
    """Keep legacy helpers from mutating an archive governed by manifests."""

    if (
        latest_integrity_manifest_path(archive_dir) is not None
        or integrity_anchor_exists(archive_dir, anchor_dir)
    ):
        raise IntegrityError(
            f"Refusing {operation} outside the locked integrity transaction. "
            "Use the supported --migrate-history-only or weekly-run CLI workflow so "
            "the raw archive and trusted manifest head advance together."
        )


def _read_integrity_anchor(
    archive_dir: Path,
    anchor_dir: Path | None = None,
) -> tuple[Path, str]:
    anchor = integrity_anchor_path(archive_dir, anchor_dir)
    if not os.path.lexists(anchor):
        raise IntegrityError(
            "The machine-local trusted integrity anchor is missing. Do not create a new "
            "baseline until a technical maintainer has established whether the prior "
            "anchor or manifest history must be restored."
        )
    if not anchor.is_file():
        raise IntegrityError(f"The trusted integrity anchor is not a file: {anchor}.")
    try:
        payload = read_json_manifest(anchor, root=anchor.parent)
    except IntegrityError as exc:
        raise IntegrityError(f"The trusted integrity anchor is invalid: {exc}") from exc
    if payload.get("schema_version") != INTEGRITY_ANCHOR_SCHEMA_VERSION:
        raise IntegrityError(
            "The trusted integrity anchor has an unsupported schema version."
        )
    expected_identity = integrity_archive_identity(archive_dir)
    if not secrets.compare_digest(
        str(payload.get("archive_identity_sha256", "")), expected_identity
    ):
        raise IntegrityError(
            "The trusted integrity anchor belongs to a different archive directory."
        )
    manifest_name = payload.get("manifest_path")
    if (
        not isinstance(manifest_name, str)
        or not manifest_name
        or Path(manifest_name).name != manifest_name
        or "/" in manifest_name
        or "\\" in manifest_name
    ):
        raise IntegrityError(
            "The trusted integrity anchor contains an invalid manifest path."
        )
    manifest_sha256 = payload.get("manifest_sha256")
    if not isinstance(manifest_sha256, str) or not re.fullmatch(
        r"[0-9a-f]{64}", manifest_sha256
    ):
        raise IntegrityError(
            "The trusted integrity anchor contains an invalid manifest SHA-256."
        )
    manifest = integrity_manifest_dir(archive_dir) / manifest_name
    if not os.path.lexists(manifest):
        raise IntegrityError(
            "The manifest pinned by the machine-local integrity anchor is missing. "
            "Restore the recorded manifest history before rerunning."
        )
    if path_is_link_or_reparse(manifest) or not manifest.is_file():
        raise IntegrityError(
            "The manifest pinned by the machine-local integrity anchor is not a regular file."
        )
    actual_payload = read_json_manifest(manifest, root=integrity_manifest_dir(archive_dir))
    actual_sha256 = canonical_json_sha256(actual_payload)
    if not secrets.compare_digest(actual_sha256, manifest_sha256):
        raise IntegrityError(
            "The Dropbox integrity manifest head does not match the machine-local trusted "
            "anchor. Restore the recorded raw data and manifest history before rerunning."
        )
    return manifest.resolve(), manifest_sha256


def verify_integrity_anchor(
    archive_dir: Path,
    anchor_dir: Path | None = None,
    *,
    expected_manifest: Path | None = None,
    expected_sha256: str | None = None,
) -> tuple[Path, str]:
    """Require the current Dropbox head to match independently stored trusted state."""

    manifest, manifest_sha256 = _read_integrity_anchor(archive_dir, anchor_dir)
    latest = latest_integrity_manifest_path(archive_dir)
    if latest is None or latest.resolve() != manifest:
        raise IntegrityError(
            "The Dropbox integrity manifest head differs from the machine-local trusted "
            "anchor. Remove no files; ask the technical maintainer to reconcile the history."
        )
    if expected_manifest is not None and manifest != expected_manifest.resolve():
        raise IntegrityError(
            "The trusted integrity manifest head changed during this run."
        )
    if expected_sha256 is not None and not secrets.compare_digest(
        manifest_sha256, expected_sha256
    ):
        raise IntegrityError(
            "The trusted integrity manifest hash changed during this run."
        )
    return manifest, manifest_sha256


def _write_integrity_anchor(
    archive_dir: Path,
    manifest: Path,
    manifest_sha256: str,
    anchor_dir: Path | None = None,
) -> Path:
    anchor = integrity_anchor_path(archive_dir, anchor_dir)
    anchor.parent.mkdir(parents=True, exist_ok=True)
    # Recheck after directory creation so a substituted leaf never becomes trusted.
    if path_is_link_or_reparse(anchor.parent):
        raise IntegrityError(
            "The trusted integrity-anchor folder became a link or reparse point."
        )
    manifest_root = integrity_manifest_dir(archive_dir).resolve()
    try:
        manifest_name = manifest.resolve().relative_to(manifest_root).as_posix()
    except ValueError as exc:
        raise IntegrityError(
            "Cannot pin an integrity manifest outside the managed manifest folder."
        ) from exc
    if "/" in manifest_name:
        raise IntegrityError("Trusted integrity manifests must be direct files in the manifest folder.")
    payload = {
        "schema_version": INTEGRITY_ANCHOR_SCHEMA_VERSION,
        "archive_identity_sha256": integrity_archive_identity(archive_dir),
        "manifest_path": manifest_name,
        "manifest_sha256": manifest_sha256,
        "updated_at_utc": datetime.now(timezone.utc).isoformat(),
    }
    if os.name != "nt":
        os.chmod(anchor.parent, 0o700)
    # This atomic replace is deliberately the final operation. If it raises, the
    # prior anchor remains; once it succeeds, callers must treat the new manifest
    # as committed and must never roll it back beneath the advanced anchor.
    return write_json_manifest_atomic(anchor, payload, root=anchor.parent)


def initialize_integrity_anchor(
    archive_dir: Path,
    manifest: Path,
    manifest_sha256: str,
    anchor_dir: Path | None = None,
) -> Path:
    """Pin a virgin or explicitly adopted verified chain without any reset path."""

    if integrity_anchor_exists(archive_dir, anchor_dir):
        raise IntegrityError(
            "A trusted integrity anchor already exists; refusing to replace it."
        )
    latest = latest_integrity_manifest_path(archive_dir)
    if latest is None or latest.resolve() != manifest.resolve():
        raise IntegrityError("Only the current integrity manifest head can be trusted.")
    chain = verify_manifest_chain(manifest, integrity_manifest_dir(archive_dir))
    if not chain or not secrets.compare_digest(chain[0].sha256, manifest_sha256):
        raise IntegrityError("The integrity manifest changed before it could be trusted.")
    written = _write_integrity_anchor(
        archive_dir, manifest, manifest_sha256, anchor_dir
    )
    return written


def rebind_restored_integrity_anchor(
    archive_dir: Path,
    output_dir: Path,
    source_anchor: Path,
    anchor_dir: Path | None = None,
) -> tuple[Path, Path]:
    """Rebind a verified restored head to a replacement machine/path identity."""

    archive_dir = archive_dir.resolve()
    output_dir = output_dir.resolve()
    source_anchor = regular_file_without_reparse_ancestors(
        source_anchor, purpose="restored integrity anchor"
    )
    source_payload = read_json_manifest(source_anchor, root=source_anchor.parent)
    if source_payload.get("schema_version") != INTEGRITY_ANCHOR_SCHEMA_VERSION:
        raise IntegrityError(
            "The restored integrity anchor has an unsupported schema version."
        )
    source_identity = str(source_payload.get("archive_identity_sha256") or "")
    if not re.fullmatch(r"[0-9a-f]{64}", source_identity):
        raise IntegrityError(
            "The restored integrity anchor has an invalid archive identity."
        )
    manifest_name = source_payload.get("manifest_path")
    if (
        not isinstance(manifest_name, str)
        or not manifest_name
        or Path(manifest_name).name != manifest_name
        or "/" in manifest_name
        or "\\" in manifest_name
    ):
        raise IntegrityError(
            "The restored integrity anchor contains an invalid manifest path."
        )
    source_manifest_sha256 = str(source_payload.get("manifest_sha256") or "")
    if not re.fullmatch(r"[0-9a-f]{64}", source_manifest_sha256):
        raise IntegrityError(
            "The restored integrity anchor contains an invalid manifest SHA-256."
        )
    latest = latest_integrity_manifest_path(archive_dir)
    if latest is None or latest.name != manifest_name:
        raise IntegrityError(
            "The restored archive head does not match the manifest named by the "
            "backed-up integrity anchor."
        )
    _, verified_sha256 = verify_integrity_state(
        archive_dir, output_dir, latest
    )
    if not secrets.compare_digest(verified_sha256, source_manifest_sha256):
        raise IntegrityError(
            "The restored archive and managed outputs do not match the backed-up "
            "integrity anchor."
        )
    destination = integrity_anchor_path(archive_dir, anchor_dir)
    if os.path.lexists(destination):
        raise IntegrityError(
            "A trusted integrity anchor already exists for the restored archive; "
            "refusing to replace it."
        )
    written = initialize_integrity_anchor(
        archive_dir, latest, verified_sha256, anchor_dir
    )
    receipt = written.parent / (
        "restore-rebind-"
        f"{integrity_archive_identity(archive_dir)[:16]}-"
        f"{verified_sha256[:16]}.receipt.json"
    )
    receipt_payload = {
        "schema_version": 1,
        "contract": "IntegrityAnchorRestoreRebindReceiptV1",
        "rebound_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_anchor_file": source_anchor.name,
        "source_anchor_sha256": sha256_file(source_anchor),
        "source_archive_identity_sha256": source_identity,
        "restored_archive_identity_sha256": integrity_archive_identity(
            archive_dir
        ),
        "manifest_path": latest.name,
        "manifest_sha256": verified_sha256,
        "target_anchor_file": written.name,
        "verification": (
            "Full manifest chain, raw inventory, generated archive, published "
            "outputs, and management workbook digest verified before rebind."
        ),
    }
    try:
        write_json_manifest_atomic(
            receipt, receipt_payload, root=written.parent
        )
    except Exception as exc:
        raise IntegrityError(
            f"The restored anchor was rebound at {written}, but its audit receipt "
            "could not be written. Preserve the anchor and investigate before "
            "continuing."
        ) from exc
    return written, receipt


def advance_integrity_anchor(
    archive_dir: Path,
    previous_manifest: Path,
    previous_sha256: str,
    new_manifest: Path,
    new_sha256: str,
    anchor_dir: Path | None = None,
) -> Path:
    """Compare-and-swap the trusted head after a verified linear advancement."""

    anchored_manifest, anchored_sha256 = _read_integrity_anchor(archive_dir, anchor_dir)
    if anchored_manifest != previous_manifest.resolve() or not secrets.compare_digest(
        anchored_sha256, previous_sha256
    ):
        raise IntegrityError(
            "The trusted integrity head changed before the new manifest could be committed."
        )
    latest = latest_integrity_manifest_path(archive_dir)
    if latest is None or latest.resolve() != new_manifest.resolve():
        raise IntegrityError("The new manifest is not the sole current integrity head.")
    chain = verify_manifest_chain(new_manifest, integrity_manifest_dir(archive_dir))
    if (
        len(chain) < 2
        or chain[0].path != new_manifest.resolve()
        or not secrets.compare_digest(chain[0].sha256, new_sha256)
        or chain[1].path != previous_manifest.resolve()
        or not secrets.compare_digest(chain[1].sha256, previous_sha256)
    ):
        raise IntegrityError(
            "The new manifest is not a verified direct successor of the trusted head."
        )
    written = _write_integrity_anchor(
        archive_dir, new_manifest, new_sha256, anchor_dir
    )
    return written


def integrity_manifest_dir(archive_dir: Path) -> Path:
    return managed_subdirectory(
        archive_dir,
        INTEGRITY_MANIFEST_FOLDER,
        purpose="integrity manifest",
        create=False,
    )


@contextmanager
def workflow_run_lock(archive_dir: Path) -> Iterable[Path]:
    """Hold one OS-enforced, nonblocking lock for the complete workflow run."""

    archive_dir = archive_dir.resolve()
    archive_dir.mkdir(parents=True, exist_ok=True)
    root = managed_subdirectory(
        archive_dir,
        INTEGRITY_MANIFEST_FOLDER,
        purpose="integrity manifest",
        create=True,
    )
    lock_path = root / ".weekly-snapshot.lock"
    if os.path.lexists(lock_path) and path_is_link_or_reparse(lock_path):
        raise IntegrityError(
            "The workflow lock is a link or reparse point. Restore the normal lock file "
            "inside 03 Archive\\run-manifests before rerunning."
        )
    handle = lock_path.open("a+b")
    acquired = False
    try:
        handle.seek(0, os.SEEK_END)
        if handle.tell() == 0:
            handle.write(b"\0")
            handle.flush()
        handle.seek(0)
        try:
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl

                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except OSError as exc:
            raise IntegrityError(
                "Another weekly snapshot or integrity operation is already running. "
                "Let it finish before launching again."
            ) from exc
        acquired = True
        yield lock_path
    finally:
        if acquired:
            handle.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        handle.close()


def generated_workbook_archive_dir(archive_dir: Path) -> Path:
    return managed_subdirectory(
        archive_dir,
        GENERATED_WORKBOOK_ARCHIVE_FOLDER,
        purpose="generated-workbook archive",
        create=False,
    )


def inventory_dicts(items: Iterable[FileFingerprint]) -> list[dict[str, Any]]:
    return [item.to_dict() for item in items]


def manifest_inventory(payload: dict[str, Any], field: str) -> tuple[FileFingerprint, ...]:
    values = payload.get(field)
    if not isinstance(values, list):
        raise IntegrityError(f"Integrity manifest is missing the {field!r} inventory.")
    try:
        return tuple(FileFingerprint.from_mapping(value) for value in values)
    except (TypeError, ValueError) as exc:
        raise IntegrityError(f"Integrity manifest has an invalid {field!r} inventory: {exc}") from exc


def manifest_pinned_master_snapshot(
    archive_dir: Path,
    payload: dict[str, Any],
) -> Path:
    """Resolve exactly one archived master fingerprint owned by this manifest run."""

    run_id = payload.get("run_id")
    if not isinstance(run_id, str) or not run_id.strip():
        raise IntegrityError(
            "The prior manifest cannot identify its archived master without a run ID."
        )
    suffix = (
        f"/{run_id}/published/Red_Onion_Server_Master.xlsx"
    ).casefold()
    inventory = manifest_inventory(payload, "derived_archive_inventory")
    matches = [
        item
        for item in inventory
        if f"/{item.path.lstrip('/')}".casefold().endswith(suffix)
    ]
    if len(matches) > 1:
        raise IntegrityError(
            "The prior manifest must pin exactly one archived published master for its run."
        )
    root = generated_workbook_archive_dir(archive_dir)

    def verified_candidate(item: FileFingerprint) -> Path:
        candidate = managed_recursive_file(
            [root],
            root / Path(item.path),
            purpose="manifest-pinned archived master workbook",
        )
        if sha256_file(candidate) != item.sha256:
            raise IntegrityError(
                "The manifest-pinned archived master workbook fingerprint does not match."
            )
        return candidate

    if matches:
        return verified_candidate(matches[0])

    expected_v2 = payload.get("master_generated_content_sha256")
    if not isinstance(expected_v2, str) or not re.fullmatch(r"[0-9a-f]{64}", expected_v2):
        raise IntegrityError(
            "The legacy manifest cannot resolve an archived master without its v2 digest."
        )
    archived_items = [
        item
        for item in inventory
        if f"/{item.path.lstrip('/')}".casefold().endswith(
            "/published/red_onion_server_master.xlsx"
        )
    ]
    digest_matches: list[Path] = []
    for item in archived_items:
        candidate = verified_candidate(item)
        if workbook_metadata_sha256(candidate) == expected_v2:
            digest_matches.append(candidate)
    if not digest_matches:
        raise IntegrityError(
            "The legacy manifest inventory contains no archived published master matching "
            "its frozen v2 digest."
        )
    substantive_digests = {
        workbook_generated_content_sha256(candidate) for candidate in digest_matches
    }
    if len(substantive_digests) != 1:
        raise IntegrityError(
            "The legacy manifest inventory contains semantically conflicting archived "
            "masters for its frozen v2 digest."
        )
    return sorted(digest_matches, key=lambda path: str(path).casefold())[0]


def latest_integrity_manifest_path(archive_dir: Path) -> Path | None:
    root = integrity_manifest_dir(archive_dir)
    if not root.exists():
        return None
    manifests = sorted(
        (
            path
            for path in root.glob("*.json")
            if path.is_file() and not path.name.startswith(".")
        ),
        key=lambda path: path.name.casefold(),
    )
    return manifests[-1] if manifests else None


def managed_published_output_paths(output_dir: Path) -> list[Path]:
    if not output_dir.exists():
        return []
    paths: list[Path] = []
    for path in output_dir.glob("*.xlsx"):
        if path.name == "Red_Onion_Server_Master.xlsx" or path.name.startswith((".", "~$")):
            continue
        paths.append(
            managed_direct_child(
                output_dir,
                path,
                purpose="published workbook",
                require_file=True,
            )
        )
    return sorted(paths, key=lambda path: path.name.casefold())


def effective_config_sha256(config: dict[str, Any]) -> str:
    return canonical_json_sha256(
        {key: value for key, value in config.items() if not str(key).startswith("_")}
    )


def workflow_provenance(config_path: Path, config: dict[str, Any]) -> dict[str, Any]:
    repository_root = REPOSITORY_ROOT.resolve()
    requirements_path = PROGRAM_DIR / "requirements.txt"
    config_for_provenance: Path | None = None
    try:
        resolved_config = config_path.resolve()
        resolved_config.relative_to(repository_root)
        if resolved_config.is_file():
            config_for_provenance = resolved_config
    except ValueError:
        pass
    provenance = collect_provenance(
        repository_root,
        config_path=config_for_provenance,
        requirements_path=requirements_path,
    )
    provenance["effective_config_sha256"] = effective_config_sha256(config)
    if config_for_provenance is None:
        provenance["config"] = {
            "path": config_path.name,
            "exists": config_path.is_file(),
            "sha256": sha256_file(config_path) if config_path.is_file() else None,
            "effective_sha256": provenance["effective_config_sha256"],
        }
    return provenance


def timestamped_manifest_path(root: Path, run_id: str, kind: str) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S.%fZ")
    safe_kind = re.sub(r"[^a-z0-9-]+", "-", kind.casefold()).strip("-")
    return root / f"{timestamp}-{safe_kind}-{run_id}.json"


def current_master_digest_state(output_dir: Path) -> tuple[str | None, str | None]:
    master_path = managed_master_workbook_path(output_dir)
    if not os.path.lexists(master_path):
        return None, None
    wb = load_workbook(master_path, data_only=False, read_only=False)
    try:
        scheme = stamped_workbook_digest_scheme(wb) or LEGACY_WORKBOOK_DIGEST_SCHEME
        stamped = stamped_workbook_digest(wb)
    finally:
        wb.close()
    if scheme in {
        LEGACY_WORKBOOK_DIGEST_SCHEME,
        PREVIOUS_WORKBOOK_DIGEST_SCHEME,
    }:
        if not stamped or not re.fullmatch(r"[0-9a-f]{64}", stamped):
            raise IntegrityError(
                "Prior master workbook is missing its frozen stamped digest."
            )
        return stamped, scheme
    if scheme == WORKBOOK_DIGEST_SCHEME:
        return workbook_generated_content_sha256(master_path), scheme
    raise IntegrityError(f"Unsupported master workbook digest scheme: {scheme!r}.")


def current_master_digest(output_dir: Path) -> str | None:
    return current_master_digest_state(output_dir)[0]


def build_integrity_state(
    archive_dir: Path,
    output_dir: Path,
) -> dict[str, Any]:
    raw_root = canonical_daily_archive_dir(archive_dir)
    derived_root = generated_workbook_archive_dir(archive_dir)
    validate_managed_tree_no_reparse(raw_root, purpose="canonical raw archive")
    validate_managed_tree_no_reparse(
        derived_root, purpose="generated-workbook archive"
    )
    raw_inventory = build_raw_inventory(raw_root)
    derived_inventory = build_raw_inventory(derived_root)
    public_paths = managed_published_output_paths(output_dir)
    public_inventory = build_raw_inventory(output_dir, public_paths)
    master_digest, master_scheme = current_master_digest_state(output_dir)
    state = {
        "raw_inventory": inventory_dicts(raw_inventory),
        "derived_archive_inventory": inventory_dicts(derived_inventory),
        "published_output_inventory": inventory_dicts(public_inventory),
        "master_generated_content_sha256": master_digest,
    }
    if state["master_generated_content_sha256"] is not None:
        state["master_generated_content_digest_scheme"] = master_scheme
        state["master_metadata_sha256"] = workbook_metadata_sha256(
            managed_master_workbook_path(output_dir)
        )
    return state


def merge_inventory(
    base: Iterable[FileFingerprint], updates: Iterable[FileFingerprint]
) -> tuple[FileFingerprint, ...]:
    """Apply explicit fingerprint additions/replacements to a recorded inventory."""

    merged = {item.path.casefold(): item for item in base}
    for item in updates:
        merged[item.path.casefold()] = item
    return tuple(sorted(merged.values(), key=lambda item: item.path.casefold()))


def fingerprint_matching(
    root: Path,
    path: Path,
    *,
    expected_sha256: str,
    expected_size: int | None = None,
) -> FileFingerprint:
    """Fingerprint a planned file and require its content to match the plan."""

    fingerprint = fingerprint_file(root, path)
    if fingerprint.sha256 != expected_sha256 or (
        expected_size is not None and fingerprint.size != expected_size
    ):
        raise IntegrityError(
            f"Managed file changed during the transaction: {fingerprint.path}."
        )
    return fingerprint


def expected_integrity_state(
    previous_payload: dict[str, Any],
    *,
    raw_updates: Iterable[FileFingerprint] = (),
    derived_updates: Iterable[FileFingerprint] = (),
    published_updates: Iterable[FileFingerprint] = (),
    master_generated_content_sha256: str | None | object = Ellipsis,
    master_generated_content_digest_scheme: str | None | object = Ellipsis,
    master_metadata_sha256: str | None | object = Ellipsis,
) -> dict[str, Any]:
    """Build prior-state-plus-explicit-delta without inventorying unrelated changes."""

    master_digest = (
        previous_payload.get("master_generated_content_sha256")
        if master_generated_content_sha256 is Ellipsis
        else master_generated_content_sha256
    )
    state = {
        "raw_inventory": inventory_dicts(
            merge_inventory(manifest_inventory(previous_payload, "raw_inventory"), raw_updates)
        ),
        "derived_archive_inventory": inventory_dicts(
            merge_inventory(
                manifest_inventory(previous_payload, "derived_archive_inventory"),
                derived_updates,
            )
        ),
        "published_output_inventory": inventory_dicts(
            merge_inventory(
                manifest_inventory(previous_payload, "published_output_inventory"),
                published_updates,
            )
        ),
        "master_generated_content_sha256": master_digest,
    }
    scheme = (
        previous_payload.get(
            "master_generated_content_digest_scheme",
            previous_payload.get("_master_generated_content_digest_scheme"),
        )
        if master_generated_content_digest_scheme is Ellipsis
        else master_generated_content_digest_scheme
    )
    metadata_digest = (
        previous_payload.get(
            "master_metadata_sha256",
            previous_payload.get("_master_metadata_sha256"),
        )
        if master_metadata_sha256 is Ellipsis
        else master_metadata_sha256
    )
    if master_digest is not None and scheme is not None:
        state["master_generated_content_digest_scheme"] = scheme
    if master_digest is not None and metadata_digest is not None:
        state["master_metadata_sha256"] = metadata_digest
    return state


def verify_expected_integrity_state(
    archive_dir: Path,
    output_dir: Path,
    expected_state: dict[str, Any],
) -> None:
    """Reject every post-state change not represented by the explicit run delta."""

    actual_state = build_integrity_state(archive_dir, output_dir)
    if canonical_json_sha256(actual_state) != canonical_json_sha256(expected_state):
        raise IntegrityError(
            "Managed raw data or derivative state changed outside this run's planned "
            "transaction. No new integrity manifest was committed."
        )


def assert_manifest_head(
    archive_dir: Path,
    expected_manifest: Path,
    expected_sha256: str,
    anchor_dir: Path | None = None,
) -> None:
    """Compare-and-swap guard for a single linear manifest history."""

    verify_integrity_anchor(
        archive_dir,
        anchor_dir,
        expected_manifest=expected_manifest,
        expected_sha256=expected_sha256,
    )
    latest = latest_integrity_manifest_path(archive_dir)
    if latest is None or latest.resolve() != expected_manifest.resolve():
        raise IntegrityError(
            "The integrity manifest head changed during this run; refusing to create a "
            "forked history."
        )
    payload = read_json_manifest(latest, root=integrity_manifest_dir(archive_dir))
    if canonical_json_sha256(payload) != expected_sha256:
        raise IntegrityError(
            "The integrity manifest head changed after preflight; refusing to extend it."
        )


def write_integrity_manifest(
    *,
    archive_dir: Path,
    output_dir: Path,
    config_path: Path,
    config: dict[str, Any],
    kind: str,
    run_id: str,
    previous_manifest: Path | None,
    expected_previous_sha256: str | None = None,
    integrity_state: dict[str, Any] | None = None,
    details: dict[str, Any] | None = None,
) -> Path:
    root = integrity_manifest_dir(archive_dir)
    payload: dict[str, Any] = {
        "schema_version": 1,
        "kind": kind,
        "run_id": run_id,
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        **(
            build_integrity_state(archive_dir, output_dir)
            if integrity_state is None
            else integrity_state
        ),
        "provenance": workflow_provenance(config_path, config),
    }
    if details:
        payload["details"] = details
    manifest_path = timestamped_manifest_path(root, run_id, kind)
    return write_chained_manifest_atomic(
        manifest_path,
        payload,
        root,
        previous_manifest_path=previous_manifest,
        expected_previous_sha256=expected_previous_sha256,
    )


def verify_integrity_state(
    archive_dir: Path,
    output_dir: Path,
    latest_manifest: Path,
    *,
    allow_legacy_master_upgrade: bool = False,
) -> tuple[dict[str, Any], str]:
    root = integrity_manifest_dir(archive_dir)
    chain = verify_manifest_chain(latest_manifest, root)
    if not chain:
        raise IntegrityError("The integrity manifest chain is empty.")
    payload = dict(chain[0].payload)
    if payload.get("schema_version") != 1:
        raise IntegrityError(
            f"Unsupported integrity manifest schema: {payload.get('schema_version')!r}."
        )

    raw_root = canonical_daily_archive_dir(archive_dir)
    derived_root = generated_workbook_archive_dir(archive_dir)
    validate_managed_tree_no_reparse(raw_root, purpose="canonical raw archive")
    validate_managed_tree_no_reparse(
        derived_root, purpose="generated-workbook archive"
    )
    try:
        verify_raw_inventory(raw_root, manifest_inventory(payload, "raw_inventory"))
    except IntegrityError as exc:
        raise IntegrityError(f"Canonical raw archive verification failed. {exc}") from exc
    try:
        verify_raw_inventory(
            derived_root,
            manifest_inventory(payload, "derived_archive_inventory"),
        )
    except IntegrityError as exc:
        raise IntegrityError(f"Generated-workbook archive verification failed. {exc}") from exc
    try:
        verify_raw_inventory(
            output_dir,
            manifest_inventory(payload, "published_output_inventory"),
            managed_published_output_paths(output_dir),
        )
    except IntegrityError as exc:
        raise IntegrityError(f"Published workbook verification failed. {exc}") from exc

    expected_master = payload.get("master_generated_content_sha256")
    master_path = managed_master_workbook_path(output_dir)
    if expected_master is not None:
        if not isinstance(expected_master, str) or not re.fullmatch(r"[0-9a-f]{64}", expected_master):
            raise IntegrityError("Integrity manifest contains an invalid master workbook digest.")
        if not os.path.lexists(master_path):
            raise IntegrityError(
                "Master workbook verification failed: the recorded master workbook is missing."
            )
        recorded_scheme = payload.get("master_generated_content_digest_scheme")
        if recorded_scheme is None:
            # Schema-v1 manifests emitted before v3 always recorded the exact v2
            # metadata-rich digest. Absence is legacy meaning, never v3 inference.
            recorded_scheme = LEGACY_WORKBOOK_DIGEST_SCHEME
        if recorded_scheme not in {
            LEGACY_WORKBOOK_DIGEST_SCHEME,
            PREVIOUS_WORKBOOK_DIGEST_SCHEME,
            WORKBOOK_DIGEST_SCHEME,
        }:
            raise IntegrityError(
                f"Integrity manifest contains an unsupported workbook digest scheme: "
                f"{recorded_scheme!r}."
            )
        legacy_reference: Path | None = None
        current_metadata = workbook_metadata_sha256(master_path)
        if recorded_scheme == PREVIOUS_WORKBOOK_DIGEST_SCHEME:
            legacy_reference = manifest_pinned_master_snapshot(archive_dir, payload)
        elif (
            recorded_scheme == LEGACY_WORKBOOK_DIGEST_SCHEME
            and current_metadata != expected_master
        ):
            legacy_reference = manifest_pinned_master_snapshot(archive_dir, payload)
        expected_metadata = payload.get("master_metadata_sha256")
        metadata_drift = (
            recorded_scheme == LEGACY_WORKBOOK_DIGEST_SCHEME
            and legacy_reference is not None
        )
        if (
            recorded_scheme in {
                PREVIOUS_WORKBOOK_DIGEST_SCHEME,
                WORKBOOK_DIGEST_SCHEME,
            }
            and expected_metadata is None
        ):
            raise IntegrityError(
                "Substantive-content integrity manifest is missing the required "
                "master metadata digest."
            )
        if expected_metadata is not None:
            if not isinstance(expected_metadata, str) or not re.fullmatch(
                r"[0-9a-f]{64}", expected_metadata
            ):
                raise IntegrityError(
                    "Integrity manifest contains an invalid master metadata digest."
                )
            if current_metadata != expected_metadata:
                metadata_drift = True
                warnings.warn(
                    "Master workbook metadata differs from the recorded generated version; "
                    "substantive verification continues.",
                    UserWarning,
                    stacklevel=2,
                )
        # Compatibility is permitted only when the manifest-recorded digest is
        # independently trusted by explicit adoption or an existing local anchor.
        verify_existing_management_workbook_integrity(
            master_path,
            expected_digest=expected_master,
            expected_digest_scheme=recorded_scheme,
            legacy_reference_path=legacy_reference,
            allow_legacy_protection_upgrade=allow_legacy_master_upgrade,
        )
        payload["_master_generated_content_digest_scheme"] = recorded_scheme
        payload["_master_metadata_drift"] = metadata_drift
        payload["_master_metadata_sha256"] = current_metadata
        if legacy_reference is not None:
            payload["_legacy_master_reference_path"] = str(legacy_reference)
        payload["_legacy_master_upgrade_pending"] = (
            allow_legacy_master_upgrade
            and management_workbook_upgrade_pending(master_path)
        )
    elif os.path.lexists(master_path):
        raise IntegrityError(
            "An unrecorded master workbook appeared after the integrity baseline. "
            "Reconcile it before running the automation."
        )
    return payload, chain[0].sha256


def ensure_integrity_preflight(
    archive_dir: Path,
    output_dir: Path,
    config_path: Path,
    config: dict[str, Any],
    *,
    allow_initialize: bool = False,
    anchor_dir: Path | None = None,
) -> tuple[Path, dict[str, Any], str]:
    latest = latest_integrity_manifest_path(archive_dir)
    anchor_exists = integrity_anchor_exists(archive_dir, anchor_dir)
    created_baseline = False
    if anchor_exists:
        anchored_manifest, anchored_sha256 = verify_integrity_anchor(
            archive_dir, anchor_dir
        )
        if latest is None or latest.resolve() != anchored_manifest:
            raise IntegrityError(
                "The established manifest history is missing or no longer matches the "
                "machine-local trusted anchor. Explicit baseline initialization cannot "
                "replace established integrity history."
            )
        latest = anchored_manifest
    elif latest is not None and not allow_initialize:
        raise IntegrityError(
            "Integrity manifests exist, but this runner has no machine-local trusted-head "
            "anchor. A technical maintainer must explicitly inspect and adopt the existing "
            "chain with --initialize-integrity-baseline before ordinary runs can continue."
        )
    if latest is None:
        if not allow_initialize:
            raise IntegrityError(
                "No integrity baseline exists. Run the explicit "
                "--initialize-integrity-baseline operation after confirming the current "
                "raw archive and finished reports are the intended starting state."
            )
        run_id = str(uuid.uuid4())
        latest = write_integrity_manifest(
            archive_dir=archive_dir,
            output_dir=output_dir,
            config_path=config_path,
            config=config,
            kind="integrity-baseline",
            run_id=run_id,
            previous_manifest=None,
            details={
                "purpose": (
                    "Initial clearly labeled inventory of existing raw history and derivatives."
                )
            },
        )
        created_baseline = True
    try:
        payload, manifest_hash = verify_integrity_state(
            archive_dir,
            output_dir,
            latest,
            allow_legacy_master_upgrade=allow_initialize or anchor_exists,
        )
        if anchor_exists:
            if not secrets.compare_digest(manifest_hash, anchored_sha256):
                raise IntegrityError(
                    "The verified manifest hash differs from the machine-local trusted anchor."
                )
        else:
            initialize_integrity_anchor(
                archive_dir,
                latest,
                manifest_hash,
                anchor_dir,
            )
    except Exception:
        if (
            created_baseline
            and latest.exists()
            and not integrity_anchor_exists(archive_dir, anchor_dir)
        ):
            latest.unlink()
        raise
    return latest, payload, manifest_hash


def rollback_created_files(
    paths: Iterable[Path],
    managed_root: Path,
    *,
    expected_hashes: dict[Path, str],
) -> list[Path]:
    """Quarantine first, then delete only this run's exact file version."""

    root = managed_root.resolve()
    normalized_hashes = {
        path.absolute(): digest for path, digest in expected_hashes.items()
    }
    conflicts: list[Path] = []
    for path in reversed(list(paths)):
        candidate = path.absolute()
        try:
            candidate.relative_to(root)
        except ValueError:
            conflicts.append(candidate)
            continue
        expected_hash = normalized_hashes.get(candidate)
        if not os.path.lexists(candidate):
            continue
        if expected_hash is None:
            conflicts.append(candidate)
            continue
        try:
            managed_recursive_file([root], candidate, purpose="rollback target")
        except IntegrityError:
            conflicts.append(candidate)
            continue

        quarantine = candidate.parent / (
            f".{candidate.name}.{uuid.uuid4().hex}.rollback"
        )
        if os.path.lexists(quarantine):
            conflicts.append(candidate)
            continue
        os.replace(candidate, quarantine)
        quarantine_matches = (
            not path_is_link_or_reparse(quarantine)
            and quarantine.is_file()
            and sha256_file(quarantine) == expected_hash
        )
        if quarantine_matches:
            quarantine.unlink()
            continue

        if not os.path.lexists(candidate):
            os.replace(quarantine, candidate)
            conflicts.append(candidate)
        else:
            # A new entry appeared after quarantine. Preserve both versions and
            # report the hidden recovery path rather than overwriting either one.
            conflicts.append(quarantine)
    return conflicts


def remove_snapshot_run(snapshot_run_dir: Path, generated_root: Path) -> None:
    root = generated_root.resolve()
    candidate = snapshot_run_dir.resolve()
    try:
        relative = candidate.relative_to(root)
    except ValueError as exc:
        raise IntegrityError(f"Refusing snapshot rollback outside {root}: {candidate}") from exc
    if len(relative.parts) < 2:
        raise IntegrityError(f"Refusing broad snapshot rollback target: {candidate}")
    if candidate.exists():
        shutil.rmtree(candidate)


def validate_staged_outputs(stage_dir: Path, paths: Iterable[Path]) -> list[Path]:
    root = stage_dir.resolve()
    validated: list[Path] = []
    seen_names: set[str] = set()
    for supplied in paths:
        path = supplied.resolve()
        try:
            relative = path.relative_to(root)
        except ValueError as exc:
            raise IntegrityError(f"Generated output escaped the staging folder: {path}") from exc
        if len(relative.parts) != 1 or not path.is_file():
            raise IntegrityError(f"Generated output is missing or unexpectedly nested: {path}")
        key = path.name.casefold()
        if key in seen_names:
            raise IntegrityError(f"Generated output name was duplicated: {path.name}")
        seen_names.add(key)
        sha256_file(path)
        validated.append(path)
    if not validated:
        raise IntegrityError("The run did not create any staged workbooks.")
    master = root / "Red_Onion_Server_Master.xlsx"
    if master not in validated:
        raise IntegrityError("The staged master workbook is missing.")
    validate_management_workbook(master)
    return validated


def snapshot_and_publish_outputs(
    *,
    staged_paths: Iterable[Path],
    output_dir: Path,
    archive_dir: Path,
    week_end: date,
    run_id: str,
    expected_existing_hashes: dict[str, str | None],
) -> tuple[
    list[Path],
    Path,
    dict[Path, OutputRollback],
    dict[Path, str],
    list[FileFingerprint],
    list[FileFingerprint],
]:
    staged_paths = list(staged_paths)
    generated_root = managed_subdirectory(
        archive_dir,
        GENERATED_WORKBOOK_ARCHIVE_FOLDER,
        purpose="generated-workbook archive",
        create=True,
    )
    generated_week = managed_subdirectory(
        generated_root,
        f"week-ending-{week_end.isoformat()}",
        purpose="generated-workbook archive week",
        create=True,
    )
    snapshot_run = managed_subdirectory(
        generated_week,
        run_id,
        purpose="generated-workbook run snapshot",
        create=True,
    )
    published_snapshot = managed_subdirectory(
        snapshot_run,
        "published",
        purpose="published workbook snapshot",
        create=True,
    )
    replaced_snapshot = managed_subdirectory(
        snapshot_run,
        "replaced",
        purpose="replaced workbook snapshot",
        create=True,
    )
    final_paths: list[Path] = []
    rollback_backups: dict[Path, OutputRollback] = {}
    staged_hashes: dict[Path, str] = {}
    derived_updates: list[FileFingerprint] = []
    published_updates: list[FileFingerprint] = []
    try:
        for staged in staged_paths:
            final = managed_direct_child(
                output_dir,
                output_dir / staged.name,
                purpose="published output",
                require_file=False,
            )
            final_paths.append(final)
            staged_hashes[final] = sha256_file(staged)
            staged_size = staged.stat().st_size
            published_backup = published_snapshot / staged.name
            verified_copy_file(staged, published_backup)
            derived_updates.append(
                fingerprint_matching(
                    generated_root,
                    published_backup,
                    expected_sha256=staged_hashes[final],
                    expected_size=staged_size,
                )
            )
            expected_original = expected_existing_hashes.get(staged.name.casefold(), Ellipsis)
            if final.exists():
                if expected_original is Ellipsis or expected_original is None:
                    raise IntegrityError(
                        f"An unrecorded output appeared during the run: {final.name}."
                    )
                actual_original = sha256_file(final)
                if actual_original != expected_original:
                    raise IntegrityError(
                        f"{final.name} changed after it was imported for this run. "
                        "The newer edit was preserved; rerun to carry it forward."
                    )
                backup = replaced_snapshot / staged.name
                backup_hash = verified_copy_file(final, backup)
                if backup_hash != actual_original:
                    raise IntegrityError(f"Output backup verification failed for {final.name}.")
                derived_updates.append(
                    fingerprint_matching(
                        generated_root,
                        backup,
                        expected_sha256=actual_original,
                        expected_size=final.stat().st_size,
                    )
                )
                rollback_backups[final] = OutputRollback(
                    backup, actual_original, backup_hash
                )
            else:
                if expected_original is not Ellipsis and expected_original is not None:
                    raise IntegrityError(f"Recorded output disappeared during the run: {final.name}.")
                rollback_backups[final] = OutputRollback(None, None, None)
        for staged, final in zip(staged_paths, final_paths):
            rollback = rollback_backups[final]
            if rollback.original_sha256 is not None:
                displaced = output_dir / f".{final.name}.{run_id}.replacing"
                rollback = OutputRollback(
                    rollback.backup,
                    rollback.original_sha256,
                    rollback.backup_sha256,
                    displaced,
                )
                rollback_backups[final] = rollback
                if os.path.lexists(displaced):
                    raise IntegrityError(f"Unexpected output replacement collision: {displaced.name}")
                os.replace(final, displaced)
                if (
                    path_is_link_or_reparse(displaced)
                    or not displaced.is_file()
                    or sha256_file(displaced) != rollback.original_sha256
                ):
                    if not os.path.lexists(final):
                        os.replace(displaced, final)
                    raise IntegrityError(
                        f"{final.name} changed at publication time. The newer entry was "
                        "preserved and this run was stopped."
                    )
            copied_hash = verified_copy_file(staged, final, replace=False)
            if copied_hash != staged_hashes[final]:
                raise IntegrityError(f"Published workbook verification failed for {final.name}.")
            if final.name != "Red_Onion_Server_Master.xlsx":
                published_updates.append(
                    fingerprint_matching(
                        output_dir,
                        final,
                        expected_sha256=staged_hashes[final],
                        expected_size=staged.stat().st_size,
                    )
                )
        for rollback in rollback_backups.values():
            if rollback.displaced is not None and rollback.displaced.exists():
                rollback.displaced.unlink()
        return (
            final_paths,
            snapshot_run,
            rollback_backups,
            staged_hashes,
            derived_updates,
            published_updates,
        )
    except Exception as exc:
        conflicts = rollback_published_outputs(rollback_backups, staged_hashes)
        if conflicts:
            raise IntegrityError(
                "Output rollback preserved a newer conflicting file and retained the "
                f"recovery snapshot at {snapshot_run}. Resolve {conflicts[0].name} before "
                "rerunning."
            ) from exc
        remove_snapshot_run(snapshot_run, generated_root)
        raise


def rollback_published_outputs(
    rollback_backups: dict[Path, OutputRollback],
    staged_hashes: dict[Path, str],
) -> list[Path]:
    """Restore only this run's output, never a newer manager/Dropbox edit."""

    conflicts: list[Path] = []
    for final, rollback in reversed(list(rollback_backups.items())):
        staged_hash = staged_hashes.get(final)
        recovery_source: Path | None = None
        displaced_valid = False
        if rollback.backup is not None:
            if (
                not rollback.backup.is_file()
                or path_is_link_or_reparse(rollback.backup)
                or sha256_file(rollback.backup) != rollback.backup_sha256
            ):
                conflicts.append(final)
                continue
            displaced_valid = (
                rollback.displaced is not None
                and rollback.displaced.is_file()
                and not path_is_link_or_reparse(rollback.displaced)
                and sha256_file(rollback.displaced) == rollback.original_sha256
            )
            recovery_source = rollback.displaced if displaced_valid else rollback.backup

        if os.path.lexists(final) and (
            path_is_link_or_reparse(final) or not final.is_file()
        ):
            conflicts.append(final)
            continue

        quarantine: Path | None = None
        current_hash: str | None = None
        if os.path.lexists(final):
            quarantine = final.parent / (
                f".{final.name}.{uuid.uuid4().hex}.rollback"
            )
            if os.path.lexists(quarantine):
                conflicts.append(final)
                continue
            os.replace(final, quarantine)
            if path_is_link_or_reparse(quarantine) or not quarantine.is_file():
                if not os.path.lexists(final):
                    os.replace(quarantine, final)
                conflicts.append(final)
                continue
            current_hash = sha256_file(quarantine)

        def publish_without_overwrite(source: Path, expected_hash: str | None) -> bool:
            if expected_hash is None or os.path.lexists(final):
                return False
            try:
                return verified_copy_file(source, final, replace=False) == expected_hash
            except (FileExistsError, IntegrityError, OSError):
                return False

        if rollback.backup is None:
            if quarantine is None:
                continue
            if current_hash == staged_hash:
                quarantine.unlink()
            else:
                restored = publish_without_overwrite(quarantine, current_hash)
                if restored:
                    quarantine.unlink()
                    conflicts.append(final)
                else:
                    conflicts.append(quarantine)
            continue

        assert rollback.original_sha256 is not None
        assert recovery_source is not None
        if quarantine is None or current_hash == staged_hash:
            restored = publish_without_overwrite(
                recovery_source, rollback.original_sha256
            )
            if quarantine is not None and current_hash == staged_hash:
                quarantine.unlink()
            if not restored:
                conflicts.append(final)
                continue
        elif current_hash == rollback.original_sha256:
            restored = publish_without_overwrite(
                quarantine, rollback.original_sha256
            )
            if restored:
                quarantine.unlink()
            else:
                conflicts.append(quarantine)
                continue
        else:
            restored = publish_without_overwrite(quarantine, current_hash)
            if restored:
                quarantine.unlink()
                conflicts.append(final)
            else:
                conflicts.append(quarantine)
            continue

        if displaced_valid and rollback.displaced is not None and rollback.displaced.exists():
            rollback.displaced.unlink()
    return conflicts


def record_run_stage(
    args: argparse.Namespace,
    stage: RunStage,
    message: str,
    *,
    readiness: dict[str, RunReadiness | str] | None = None,
    details: dict[str, Any] | None = None,
) -> None:
    recorder = getattr(args, "_run_attempt_recorder", None)
    if isinstance(recorder, RunAttemptRecorder):
        recorder.update(
            stage,
            message,
            readiness=readiness,
            details=details,
        )


def nearest_existing_directory(path: Path) -> Path:
    candidate = path.absolute()
    while not candidate.exists() and candidate.parent != candidate:
        candidate = candidate.parent
    if not candidate.exists():
        raise OSError(f"No existing parent directory was found for {path}.")
    return candidate


def available_disk_bytes(path: Path) -> int:
    return int(shutil.disk_usage(nearest_existing_directory(path)).free)


def assert_staging_capacity(
    output_dir: Path,
    archive_dir: Path,
    captured_inputs: Iterable[CapturedActiveInput],
) -> int:
    input_bytes = sum(item.fingerprint.size for item in captured_inputs)
    required = max(256 * 1024 * 1024, input_bytes * 5)
    for label, path in (("finished reports", output_dir), ("archive", archive_dir)):
        free = available_disk_bytes(path)
        if free < required:
            raise OSError(
                f"Insufficient disk space for {label}: {free / (1024 ** 2):.0f} MiB "
                f"free; at least {required / (1024 ** 2):.0f} MiB required."
            )
    return required


def run(args: argparse.Namespace) -> list[Path]:
    config_path = Path(args.config).resolve()
    # Validate before a folder, run lock, environment, or report artifact is created.
    args._validated_config = load_config(config_path)
    archive_dir = Path(args.archive_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    operation = (
        "integrity-baseline"
        if bool(getattr(args, "initialize_integrity_baseline", False))
        else "history-migration"
        if bool(getattr(args, "migrate_history_only", False))
        else "history-rebuild"
        if bool(getattr(args, "rebuild_from_history", False))
        else "weekly-run"
    )
    run_id = str(uuid.uuid4())
    args._attempt_run_id = run_id
    with workflow_run_lock(archive_dir):
        attempt_root = managed_subdirectory(
            archive_dir,
            RUN_ATTEMPT_FOLDER,
            purpose="run-attempt log",
            create=True,
        )
        status_path = (
            managed_direct_child(
                output_dir,
                output_dir / LAST_RUN_STATUS_FILE,
                purpose="last-run status",
                require_file=False,
            )
            if output_dir.is_dir()
            else None
        )
        recorder = RunAttemptRecorder(
            run_id=run_id,
            operation=operation,
            attempt_path=timestamped_manifest_path(
                attempt_root, run_id, "attempt"
            ),
            status_path=status_path,
        )
        verified_release_commit = os.environ.get(
            "RED_ONION_VERIFIED_RELEASE_COMMIT", ""
        )
        if re.fullmatch(r"[0-9a-fA-F]{40}", verified_release_commit):
            recorder.readiness["release"] = RunReadiness.READY.value
            recorder.details["verified_release_commit"] = (
                verified_release_commit.lower()
            )
        args._run_attempt_recorder = recorder
        recorder.update(
            RunStage.WAITING_FOR_LOCK,
            "Configuration validated and exclusive workflow lock acquired.",
        )
        try:
            generated = _run_with_lock_held(args)
        except Exception as exc:
            try:
                recorder.fail(exc)
            except Exception:
                # Attempt logging must never replace the original operational error.
                pass
            raise
        if recorder.status_path is None and output_dir.is_dir():
            recorder.status_path = managed_direct_child(
                output_dir,
                output_dir / LAST_RUN_STATUS_FILE,
                purpose="last-run status",
                require_file=False,
            )
        recorder.succeed(
            "Run completed successfully.",
            details={"generated_files": [path.name for path in generated]},
        )
        return generated


def _run_with_lock_held(args: argparse.Namespace) -> list[Path]:
    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    archive_dir = Path(args.archive_dir).resolve()
    config_path = Path(args.config).resolve()
    config = getattr(args, "_validated_config", None)
    if config is None:
        config = load_config(config_path)
    configured_anchor_dir = getattr(args, "integrity_anchor_dir", None)
    anchor_dir = (
        Path(configured_anchor_dir).expanduser().absolute()
        if configured_anchor_dir
        else default_integrity_anchor_dir()
    )

    migration_sources = [
        Path(path).resolve() for path in (getattr(args, "migrate_history_from", None) or [])
    ]
    migration_only = bool(getattr(args, "migrate_history_only", False))
    rebuild_from_history = bool(getattr(args, "rebuild_from_history", False))
    initialize_baseline = bool(
        getattr(args, "initialize_integrity_baseline", False)
    )
    if migration_only and not migration_sources:
        raise ValueError("--migrate-history-only requires at least one --migrate-history-from folder.")
    if initialize_baseline and (migration_only or migration_sources):
        raise ValueError(
            "--initialize-integrity-baseline cannot be combined with history migration options."
        )
    if rebuild_from_history and (migration_only or initialize_baseline):
        raise ValueError(
            "--rebuild-from-history cannot be combined with "
            "--migrate-history-only or --initialize-integrity-baseline."
        )
    if rebuild_from_history and (args.week_start or args.week_end):
        raise ValueError(
            "--rebuild-from-history selects the latest complete historical week "
            "and cannot be combined with --week-start or --week-end."
        )

    record_run_stage(
        args,
        RunStage.INTEGRITY_PREFLIGHT,
        "Verifying the trusted manifest head and managed report history.",
        readiness={"integrity": RunReadiness.RUNNING},
    )
    if initialize_baseline:
        latest, _, _ = ensure_integrity_preflight(
            archive_dir,
            output_dir,
            config_path,
            config,
            allow_initialize=True,
            anchor_dir=anchor_dir,
        )
        return [latest]

    active_paths: list[Path] = []
    if not migration_only and not rebuild_from_history:
        active_paths = daily_report_paths(input_dir)
        if not active_paths:
            raise FileNotFoundError(
                f"No active daily reports ({DAILY_REPORT_FORMAT_LABEL}) found in {input_dir}. "
                "Drop current Red Onion daily reports into 01 Daily Reports - Drop Here and rerun."
            )

    previous_manifest, previous_payload, previous_manifest_hash = ensure_integrity_preflight(
        archive_dir,
        output_dir,
        config_path,
        config,
        anchor_dir=anchor_dir,
    )
    legacy_master_upgrade_pending = bool(
        previous_payload.get("_legacy_master_upgrade_pending", False)
    )
    if migration_only and legacy_master_upgrade_pending:
        raise IntegrityError(
            "History-only migration is blocked while the manifest-pinned master workbook "
            "awaits its one-way workbook-contract upgrade. Run the next ordinary weekly "
            "snapshot or an approved combined history rebuild; no files were changed."
        )

    migration_plan = (
        build_history_migration_plan(
            migration_sources,
            archive_dir,
            config,
            expected_raw_inventory=manifest_inventory(
                previous_payload, "raw_inventory"
            ),
        )
        if migration_sources
        else None
    )
    if migration_only:
        assert migration_plan is not None
        copied_paths: list[Path] = []
        copied_hashes: dict[Path, str] = {}
        migration_manifest: Path | None = None
        try:
            assert_manifest_head(
                archive_dir,
                previous_manifest,
                previous_manifest_hash,
                anchor_dir,
            )
            verify_integrity_state(archive_dir, output_dir, previous_manifest)
            verify_captured_migration_inputs(migration_plan.captured_sources)
            copied_paths = list(apply_history_migration_plan(migration_plan))
            copied_hashes = history_migration_expected_hashes(
                migration_plan, copied_paths
            )
            if copied_paths:
                raw_root = canonical_daily_archive_dir(archive_dir)
                raw_updates = [fingerprint_file(raw_root, path) for path in copied_paths]
                expected_state = expected_integrity_state(
                    previous_payload, raw_updates=raw_updates
                )
                verify_expected_integrity_state(archive_dir, output_dir, expected_state)
                assert_manifest_head(
                    archive_dir,
                    previous_manifest,
                    previous_manifest_hash,
                    anchor_dir,
                )
                migration_manifest = write_integrity_manifest(
                    archive_dir=archive_dir,
                    output_dir=output_dir,
                    config_path=config_path,
                    config=config,
                    kind="history-migration",
                    run_id=str(uuid.uuid4()),
                    previous_manifest=previous_manifest,
                    expected_previous_sha256=previous_manifest_hash,
                    integrity_state=expected_state,
                    details={
                        "copied_raw_files": [path.name for path in copied_paths],
                        "source_manifest_sha256": previous_manifest_hash,
                    },
                )
                _, migration_manifest_hash = verify_integrity_state(
                    archive_dir, output_dir, migration_manifest
                )
                advance_integrity_anchor(
                    archive_dir,
                    previous_manifest,
                    previous_manifest_hash,
                    migration_manifest,
                    migration_manifest_hash,
                    anchor_dir,
                )
        except Exception:
            if migration_manifest is not None and migration_manifest.exists():
                migration_manifest.unlink()
            conflicts = rollback_created_files(
                copied_paths,
                canonical_daily_archive_dir(archive_dir),
                expected_hashes=copied_hashes,
            )
            if conflicts:
                raise IntegrityError(
                    "History migration rollback preserved a changed replacement: "
                    f"{conflicts[0]}"
                )
            raise
        return copied_paths

    active_captures: tuple[CapturedActiveInput, ...] = ()
    active_resolution = ReportResolution({}, (), ())
    active_records_by_path: dict[Path, list[MetricRecord]] = {}
    active_records: list[MetricRecord] = []
    active_week_end: date | None = None
    if rebuild_from_history:
        record_run_stage(
            args,
            RunStage.READING_INPUTS,
            "Reading only manifest-pinned history for a protected rebuild.",
        )
    else:
        record_run_stage(
            args,
            RunStage.READING_INPUTS,
            "Capturing and validating the active weekly input files.",
        )
        active_captures = tuple(capture_active_inputs(active_paths, input_dir))
        with tempfile.TemporaryDirectory(
            prefix=".weekly-input-", dir=str(integrity_manifest_dir(archive_dir))
        ) as input_stage_name:
            input_stage = Path(input_stage_name)
            captured_parse_paths: list[Path] = []
            for capture in active_captures:
                staged_input = input_stage / capture.source.name
                written_hash = verified_write_bytes(capture.content, staged_input)
                if written_hash != capture.fingerprint.sha256:
                    raise IntegrityError(
                        f"Input staging verification failed for {capture.source.name}."
                    )
                captured_parse_paths.append(staged_input)
            raw_active_records_by_path = read_reports_by_path(
                captured_parse_paths, config
            )
        active_resolution = resolve_report_duplicates(raw_active_records_by_path)
        active_records_by_path = active_resolution.records_by_path
        _, active_week_end = active_week_for_paths(active_records_by_path)
        active_records = flatten_report_records(active_records_by_path)

    if migration_plan is not None:
        historical_records_by_path = migration_plan.effective_records_by_path
        historical_duplicate_paths = migration_plan.duplicate_paths
    else:
        archived_captures = capture_archived_report_inputs(
            archive_dir,
            expected_inventory=manifest_inventory(
                previous_payload, "raw_inventory"
            ),
        )
        archived_records_by_path = read_captured_reports_by_path(
            archived_captures, config
        )
        historical_resolution = resolve_report_duplicates(archived_records_by_path)
        historical_records_by_path = historical_resolution.records_by_path
        historical_duplicate_paths = historical_resolution.duplicate_paths

    combined_records_by_path = dict(historical_records_by_path)
    combined_records_by_path.update(active_records_by_path)
    combined_resolution = resolve_report_duplicates(combined_records_by_path)
    records = flatten_report_records(combined_resolution.records_by_path)
    validate_daily_location_reconciliation(records, config)

    duplicate_paths = sorted(
        {
            *active_resolution.duplicate_paths,
            *historical_duplicate_paths,
            *combined_resolution.duplicate_paths,
        },
        key=lambda path: str(path).casefold(),
    )
    config["_report_audit"] = {
        "canonical_archive": str(canonical_daily_archive_dir(archive_dir)),
        "unique_business_days": len(combined_resolution.business_dates),
        "duplicate_files_ignored": len(duplicate_paths),
        "duplicate_file_names": [path.name for path in duplicate_paths],
        "conflicts": 0,
    }

    if rebuild_from_history:
        _, weekly_location_rows = weekly_rollups(records)
        full_by_location, _ = full_week_ends_by_location(weekly_location_rows)
        configured_locations = set(config["locations"])
        missing_locations = sorted(
            configured_locations - set(full_by_location),
            key=str.casefold,
        )
        complete_week_ends = (
            set.intersection(
                *(full_by_location[location] for location in configured_locations)
            )
            if not missing_locations
            else set()
        )
        if not complete_week_ends:
            missing_detail = (
                " Missing all history for: " + ", ".join(missing_locations) + "."
                if missing_locations
                else ""
            )
            raise ValueError(
                "History rebuild requires at least one complete Tuesday-Sunday "
                "week for every configured location."
                f"{missing_detail} No files were changed."
            )
        active_week_end = max(complete_week_ends)
        public_start, public_end = week_period_for(active_week_end)
    else:
        public_start, public_end = selected_public_dates(
            active_records, args.week_start, args.week_end
        )
    output_history_records = [
        record
        for record in records
        if not rebuild_from_history or record.report_date <= public_end
    ]
    validated_data_quality_coverage(
        output_history_records, public_start, public_end
    )
    selected_records = [
        record
        for record in output_history_records
        if public_start <= record.report_date <= public_end and is_operating_day(record.report_date)
    ]
    if not selected_records:
        raise ValueError(f"No records found between {public_start} and {public_end}.")

    assert active_week_end is not None
    required_staging_bytes = assert_staging_capacity(
        output_dir, archive_dir, active_captures
    )
    print("Preflight summary:")
    print(f"  Snapshot dates: {public_start:%Y-%m-%d} through {public_end:%Y-%m-%d}")
    print(
        "  Active input files: "
        + (
            "0 (protected history rebuild)"
            if rebuild_from_history
            else str(len(active_captures))
        )
    )
    print(f"  Unique business days in history: {len(combined_resolution.business_dates)}")
    print(f"  Semantic duplicate files ignored: {len(duplicate_paths)}")
    print(
        "  Selected-week readiness: "
        + (
            "Ready"
            if latest_week_readiness(
                [
                    record
                    for record in records
                    if record.report_date <= public_end
                ],
                [
                    row
                    for row in weekly_rollups(records)[1]
                    if row["week_end"] <= public_end
                ],
                config,
            ).ready
            else "Preliminary"
        )
    )
    print(
        f"  Staging capacity reserved: {required_staging_bytes / (1024 ** 2):.0f} MiB minimum"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    run_id = str(getattr(args, "_attempt_run_id", None) or uuid.uuid4())
    provenance = workflow_provenance(config_path, config)
    config["_integrity"] = {
        "run_id": run_id,
        "previous_manifest_sha256": previous_manifest_hash,
        "effective_config_sha256": provenance["effective_config_sha256"],
        "provenance": provenance,
        "allow_legacy_master_upgrade": legacy_master_upgrade_pending,
        "expected_master_generated_content_sha256": previous_payload.get(
            "master_generated_content_sha256"
        ),
        "expected_master_generated_content_digest_scheme": previous_payload.get(
            "_master_generated_content_digest_scheme",
            previous_payload.get("master_generated_content_digest_scheme"),
        ),
        "legacy_master_reference_path": previous_payload.get(
            "_legacy_master_reference_path"
        ),
    }

    migration_copied: list[Path] = []
    migration_copied_hashes: dict[Path, str] = {}
    active_copies: list[VerifiedArchiveCopy] = []
    final_paths: list[Path] = []
    snapshot_run: Path | None = None
    rollback_backups: dict[Path, OutputRollback] = {}
    staged_hashes: dict[Path, str] = {}
    derived_updates: list[FileFingerprint] = []
    published_updates: list[FileFingerprint] = []
    manifest_committed = False
    new_manifest_path: Path | None = None
    with tempfile.TemporaryDirectory(prefix=".weekly-run-", dir=str(output_dir)) as stage_name:
        stage_dir = Path(stage_name)
        try:
            record_run_stage(
                args,
                RunStage.BUILDING_WORKBOOKS,
                "Building exact staged workbook artifacts.",
                readiness={"workbook": RunReadiness.RUNNING},
            )
            staged_paths: list[Path] = []
            for location in config["locations"]:
                location_records = [
                    record for record in selected_records if record.location == location
                ]
                if not location_records:
                    continue
                staged_paths.append(
                    write_public_workbook(
                        location,
                        selected_records,
                        stage_dir,
                        config,
                        public_start,
                        public_end,
                    )
                )

            staged_master = stage_dir / "Red_Onion_Server_Master.xlsx"
            current_master = managed_master_workbook_path(output_dir)
            expected_existing_hashes = {
                Path(item.path).name.casefold(): item.sha256
                for item in manifest_inventory(
                    previous_payload, "published_output_inventory"
                )
            }
            if os.path.lexists(current_master):
                expected_existing_hashes[current_master.name.casefold()] = (
                    verified_copy_file(current_master, staged_master)
                )
            else:
                expected_existing_hashes[current_master.name.casefold()] = None
            staged_paths.append(
                write_master_workbook(
                    output_history_records,
                    staged_master,
                    config,
                    (
                        canonical_daily_archive_dir(archive_dir)
                        if rebuild_from_history
                        else input_dir
                    ),
                    public_start,
                    public_end,
                )
            )
            staged_paths = validate_staged_outputs(stage_dir, staged_paths)
            staged_master_digest = workbook_generated_content_sha256(staged_master)
            staged_master_metadata_digest = workbook_metadata_sha256(staged_master)

            record_run_stage(
                args,
                RunStage.PUBLISHING,
                "Staged workbooks validated; rechecking inputs and publishing exact bytes.",
                readiness={
                    "workbook": RunReadiness.READY,
                    "distribution": RunReadiness.RUNNING,
                },
            )
            if not rebuild_from_history:
                verify_captured_active_inputs(active_captures, input_dir)
            assert_manifest_head(
                archive_dir,
                previous_manifest,
                previous_manifest_hash,
                anchor_dir,
            )
            verify_integrity_state(
                archive_dir,
                output_dir,
                previous_manifest,
                allow_legacy_master_upgrade=legacy_master_upgrade_pending,
            )
            if migration_plan is not None:
                verify_captured_migration_inputs(migration_plan.captured_sources)
                migration_copied = list(apply_history_migration_plan(migration_plan))
                migration_copied_hashes = history_migration_expected_hashes(
                    migration_plan, migration_copied
                )
            if not rebuild_from_history:
                active_copies = copy_captured_active_files_verified(
                    active_captures, archive_dir, active_week_end
                )
            (
                final_paths,
                snapshot_run,
                rollback_backups,
                staged_hashes,
                derived_updates,
                published_updates,
            ) = snapshot_and_publish_outputs(
                staged_paths=staged_paths,
                output_dir=output_dir,
                archive_dir=archive_dir,
                week_end=active_week_end,
                run_id=run_id,
                expected_existing_hashes=expected_existing_hashes,
            )

            if not rebuild_from_history:
                verify_captured_active_inputs(active_captures, input_dir)
            input_inventory = tuple(capture.fingerprint for capture in active_captures)
            raw_root = canonical_daily_archive_dir(archive_dir).resolve()
            archived_destinations = [
                copy.destination.resolve().relative_to(raw_root).as_posix()
                for copy in active_copies
            ]
            capture_by_name = {
                capture.source.name.casefold(): capture for capture in active_captures
            }
            raw_updates = [fingerprint_file(raw_root, path) for path in migration_copied]
            raw_updates.extend(
                fingerprint_matching(
                    raw_root,
                    copy.destination,
                    expected_sha256=copy.sha256,
                    expected_size=capture_by_name[
                        copy.source.name.casefold()
                    ].fingerprint.size,
                )
                for copy in active_copies
            )
            expected_state = expected_integrity_state(
                previous_payload,
                raw_updates=raw_updates,
                derived_updates=derived_updates,
                published_updates=published_updates,
                master_generated_content_sha256=staged_master_digest,
                master_generated_content_digest_scheme=WORKBOOK_DIGEST_SCHEME,
                master_metadata_sha256=staged_master_metadata_digest,
            )
            verify_expected_integrity_state(archive_dir, output_dir, expected_state)
            assert_manifest_head(
                archive_dir,
                previous_manifest,
                previous_manifest_hash,
                anchor_dir,
            )
            record_run_stage(
                args,
                RunStage.COMMITTING_MANIFEST,
                "Published workbooks verified; committing the new integrity manifest.",
            )
            new_manifest_path = write_integrity_manifest(
                archive_dir=archive_dir,
                output_dir=output_dir,
                config_path=config_path,
                config=config,
                kind=(
                    "history-rebuild"
                    if rebuild_from_history
                    else "weekly-run"
                ),
                run_id=run_id,
                previous_manifest=previous_manifest,
                expected_previous_sha256=previous_manifest_hash,
                integrity_state=expected_state,
                details={
                    "source_manifest_sha256": previous_manifest_hash,
                    "rebuild_from_history": rebuild_from_history,
                    "active_input_inventory": inventory_dicts(input_inventory),
                    "archived_destinations": archived_destinations,
                    "history_migration_files": [path.name for path in migration_copied],
                    "public_snapshot_start": public_start.isoformat(),
                    "public_snapshot_end": public_end.isoformat(),
                    (
                        "history_week_end"
                        if rebuild_from_history
                        else "active_week_end"
                    ): active_week_end.isoformat(),
                    "published_workbooks": [path.name for path in final_paths],
                },
            )
            _, new_manifest_hash = verify_integrity_state(
                archive_dir, output_dir, new_manifest_path
            )
            advance_integrity_anchor(
                archive_dir,
                previous_manifest,
                previous_manifest_hash,
                new_manifest_path,
                new_manifest_hash,
                anchor_dir,
            )
            # The manifest and trusted anchor now form the committed state. Any later
            # status-reporting or intake-cleanup failure must not roll back outputs while
            # leaving the independently stored anchor pointed at a removed manifest.
            manifest_committed = True
            record_run_stage(
                args,
                RunStage.COMMITTING_MANIFEST,
                (
                    "Integrity manifest and trusted anchor verified; exact local "
                    "publication is complete."
                ),
                readiness={"distribution": RunReadiness.READY},
            )
            if not rebuild_from_history:
                quarantine_and_delete_captured_inputs(
                    active_captures, active_copies, input_dir, run_id
                )
            return final_paths
        except Exception as exc:
            if not manifest_committed:
                if new_manifest_path is not None and new_manifest_path.exists():
                    new_manifest_path.unlink()
                output_conflicts: list[Path] = []
                if rollback_backups:
                    output_conflicts = rollback_published_outputs(
                        rollback_backups, staged_hashes
                    )
                active_created = [copy for copy in active_copies if copy.created]
                raw_conflicts = rollback_created_files(
                    [copy.destination for copy in active_copies if copy.created],
                    canonical_daily_archive_dir(archive_dir),
                    expected_hashes={
                        copy.destination: copy.sha256 for copy in active_created
                    },
                )
                raw_conflicts.extend(
                    rollback_created_files(
                        migration_copied,
                        canonical_daily_archive_dir(archive_dir),
                        expected_hashes=migration_copied_hashes,
                    )
                )
                if snapshot_run is not None and not output_conflicts:
                    remove_snapshot_run(
                        snapshot_run, generated_workbook_archive_dir(archive_dir)
                    )
                if output_conflicts:
                    raise IntegrityError(
                        "The failed run preserved a newer output and retained its recovery "
                        f"snapshot at {snapshot_run}. Resolve {output_conflicts[0].name} "
                        "before rerunning."
                    ) from exc
                if raw_conflicts:
                    raise IntegrityError(
                        "The failed run preserved a changed raw-archive replacement instead "
                        f"of deleting it: {raw_conflicts[0]}. Reconcile it before rerunning."
                    ) from exc
            raise


def evidence_fingerprint_path(candidate_path: Path) -> Path:
    return candidate_path.with_name(candidate_path.name + ".fingerprint.json")


def evidence_approval_template_path(candidate_path: Path) -> Path:
    return candidate_path.with_name(candidate_path.name + ".approval-template.json")


def parse_json_cell(value: Any, *, label: str, expected_type: type) -> Any:
    if value in (None, ""):
        return expected_type()
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError as exc:
        raise IntegrityError(f"{label} is not valid JSON: {exc.msg}.") from exc
    if not isinstance(parsed, expected_type):
        raise IntegrityError(f"{label} must contain a JSON {expected_type.__name__}.")
    return parsed


def validate_v2_management_evidence_workbook(
    workbook_path: Path,
    expected_digest: str,
    *,
    expected_digest_scheme: str | None = None,
    legacy_reference_path: Path | None = None,
) -> str:
    """Accept only current or protected pre-guide workbooks with the V2 action schema."""

    wb = load_workbook(workbook_path, data_only=False)
    try:
        protection_contract = stamped_workbook_protection_contract(wb)
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames:
            management_center = wb[MANAGEMENT_CENTER_SHEET]
            has_current_evidence_surfaces = (
                "ActionBoardTable" in management_center.tables
                and OWNER_ROSTER_TABLE_NAME in management_center.tables
            )
            if has_current_evidence_surfaces:
                action_min_col, action_min_row, action_max_col, _ = range_boundaries(
                    management_center.tables["ActionBoardTable"].ref
                )
                action_headers = {
                    str(management_center.cell(action_min_row, column).value)
                    for column in range(action_min_col, action_max_col + 1)
                    if management_center.cell(action_min_row, column).value is not None
                }
            else:
                action_headers = set()
        else:
            # Exact fallback for protected V2 workbooks created before the
            # management surfaces were consolidated into Management Center.
            action_headers = (
                {
                    str(cell.value)
                    for cell in wb["Action Board"][4]
                    if cell.value is not None
                }
                if "Action Board" in wb.sheetnames
                else set()
            )
        has_v2_action_schema = {
            "Context Notes",
            "Review Disposition",
            "Evidence Status",
        }.issubset(action_headers)
        has_guide = "How to Use" in wb.sheetnames
    finally:
        wb.close()

    if (
        protection_contract != WORKBOOK_PROTECTION_CONTRACT
        or not has_v2_action_schema
    ):
        raise IntegrityError(
            "Management Evidence Package V2 requires the protected V2 action schema; "
            "legacy workbook evidence cannot be relabeled or promoted as V2."
        )
    if has_guide:
        return validate_management_workbook(
            workbook_path,
            expected_digest,
            expected_digest_scheme=expected_digest_scheme,
            legacy_reference_path=legacy_reference_path,
        )
    return validate_pre_guide_management_workbook(
        workbook_path,
        expected_digest,
        expected_digest_scheme=expected_digest_scheme,
        legacy_reference_path=legacy_reference_path,
    )


def verified_evidence_source(
    args: argparse.Namespace,
) -> tuple[dict[str, Any], list[dict[str, Any]], Path]:
    output_dir = Path(args.output_dir).resolve()
    archive_dir = Path(args.archive_dir).resolve()
    configured_anchor_dir = getattr(args, "integrity_anchor_dir", None)
    anchor_dir = (
        Path(configured_anchor_dir).expanduser().absolute()
        if configured_anchor_dir
        else default_integrity_anchor_dir()
    )
    latest_manifest = latest_integrity_manifest_path(archive_dir)
    if latest_manifest is None:
        raise IntegrityError(
            "Management evidence cannot be exported without an integrity manifest."
        )
    anchored_manifest, anchored_sha256 = verify_integrity_anchor(
        archive_dir, anchor_dir
    )
    if anchored_manifest != latest_manifest.resolve():
        raise IntegrityError(
            "Management evidence export requires the current trusted manifest head."
        )
    manifest_payload, manifest_sha256 = verify_integrity_state(
        archive_dir, output_dir, latest_manifest
    )
    if not secrets.compare_digest(manifest_sha256, anchored_sha256):
        raise IntegrityError(
            "Management evidence export manifest differs from the trusted anchor."
        )
    workbook_path = managed_master_workbook_path(output_dir)
    expected_digest = manifest_payload.get("master_generated_content_sha256")
    if not isinstance(expected_digest, str):
        raise IntegrityError(
            "The current manifest does not record a management workbook digest."
        )
    validate_v2_management_evidence_workbook(
        workbook_path,
        expected_digest,
        expected_digest_scheme=manifest_payload.get(
            "_master_generated_content_digest_scheme"
        ),
        legacy_reference_path=(
            Path(manifest_payload["_legacy_master_reference_path"])
            if manifest_payload.get("_legacy_master_reference_path")
            else None
        ),
    )
    wb = load_workbook(workbook_path, data_only=False)
    try:
        if MANAGEMENT_CENTER_SHEET in wb.sheetnames:
            management_center = wb[MANAGEMENT_CENTER_SHEET]
            allowed_reviewers = active_owner_names(
                owner_roster_from_sheet(management_center)
            )
            action_records = records_from_table(
                management_center, "ActionBoardTable"
            )
        else:
            # Exact fallback for protected V2 workbooks created before the
            # management surfaces were consolidated into Management Center.
            allowed_reviewers = (
                active_owner_names(
                    owner_roster_from_sheet(wb["Management Setup"])
                )
                if "Management Setup" in wb.sheetnames
                else None
            )
            action_records = records_from_sheet(wb["Action Board"], "Action ID")
        action_rows = validate_action_board_records(
            action_records,
            allowed_reviewers=allowed_reviewers,
        )
        evidence_rows = records_from_sheet(wb["Evidence Detail"], "Evidence ID")
    finally:
        wb.close()
    evidence_by_action = {
        str(row.get("Action ID")): row
        for row in evidence_rows
        if row.get("Action ID")
    }
    export_rows: list[dict[str, Any]] = []
    for action in action_rows:
        status = str(action.get("Status") or "").strip().casefold()
        if status in {"complete", "dismissed"}:
            continue
        action_id = str(action.get("Action ID") or "")
        evidence = evidence_by_action.get(action_id)
        if evidence is None:
            raise IntegrityError(
                f"Action {action_id!r} is missing its Evidence Detail record."
            )
        # Evidence Detail is generated and locked, while Status/Owner/Due Date on
        # Action Board remain deliberately editable after the run. Let the live
        # action row win on overlapping fields so approved exports reflect the
        # manager's current decisions without weakening the evidence payload.
        export_rows.append({**evidence, **action})
    source = {
        "manifest_path": latest_manifest.name,
        "manifest_sha256": manifest_sha256,
        "manifest_run_id": manifest_payload.get("run_id"),
        "manifest_created_at_utc": manifest_payload.get("created_at_utc"),
        "workbook_file": workbook_path.name,
        "workbook_generated_content_sha256": expected_digest,
        "workbook_generated_content_digest_scheme": manifest_payload.get(
            "_master_generated_content_digest_scheme"
        ),
        "generator_commit": (
            manifest_payload.get("provenance", {}).get("git", {}).get("commit")
        ),
        "effective_config_sha256": (
            manifest_payload.get("provenance", {}).get("effective_config_sha256")
        ),
        "methodology_version": MANAGEMENT_METHODOLOGY_VERSION,
    }
    return source, export_rows, workbook_path


def build_management_evidence_package(
    args: argparse.Namespace,
) -> ManagementEvidencePackageV2:
    source, rows, _ = verified_evidence_source(args)
    created_text = str(source.get("manifest_created_at_utc") or "")
    try:
        created = datetime.fromisoformat(created_text.replace("Z", "+00:00"))
    except ValueError as exc:
        raise IntegrityError(
            "The trusted manifest has an invalid creation timestamp."
        ) from exc
    records: list[EvidenceRecordV2] = []
    for row in rows:
        due_date = as_date(row.get("Due Date"))
        records.append(
            EvidenceRecordV2(
                action_id=str(row.get("Action ID") or ""),
                evidence_id=str(row.get("Evidence ID") or ""),
                action_code=str(row.get("Action Code") or ""),
                reason_code=str(row.get("Reason Code") or ""),
                location=str(row.get("Location") or ""),
                person_or_area=str(row.get("Person / Area") or ""),
                priority=str(row.get("Priority") or ""),
                status=str(row.get("Status") or ""),
                owner=str(row.get("Owner") or ""),
                due_date=due_date.isoformat() if due_date else None,
                recommended_next_step=str(row.get("Recommended Next Step") or ""),
                why_it_matters=str(row.get("Why It Matters") or ""),
                evidence_week_ends=str(row.get("Evidence Week Ends") or ""),
                evidence_sources=tuple(
                    parse_json_cell(
                        row.get("Evidence Sources"),
                        label=f"Evidence Sources for {row.get('Action ID')}",
                        expected_type=list,
                    )
                ),
                metric_evidence=parse_json_cell(
                    row.get("Metric Evidence"),
                    label=f"Metric Evidence for {row.get('Action ID')}",
                    expected_type=dict,
                ),
                methodology_version=str(
                    row.get("Methodology Version")
                    or MANAGEMENT_METHODOLOGY_VERSION
                ),
                comparator_type=str(
                    row.get("Comparator Type")
                    or default_management_comparator_type(row)
                ),
                peer_cohort_size=int(row.get("Peer Cohort Size") or 0),
                peer_cohort_weeks=int(row.get("Peer Cohort Weeks") or 0),
                threshold_version=str(
                    row.get("Threshold Version")
                    or MANAGEMENT_METHODOLOGY_VERSION
                ),
                evidence_status=str(row.get("Evidence Status") or ""),
                recurring_drivers=str(row.get("Recurring Drivers") or ""),
                stability_result=str(row.get("Stability Result") or ""),
                review_disposition=str(
                    row.get("Review Disposition") or "Pending Review"
                ),
                reviewed_by=str(row.get("Reviewed By") or ""),
                review_date=(
                    as_date(row.get("Review Date")).isoformat()
                    if as_date(row.get("Review Date"))
                    else None
                ),
            )
        )
    records.sort(key=lambda item: (item.priority, item.location, item.action_id))
    return ManagementEvidencePackageV2(
        source=source,
        records=tuple(records),
        retention_delete_after=(created + timedelta(days=365)).date().isoformat(),
    )


def stage_management_evidence(
    args: argparse.Namespace, candidate_path: Path
) -> list[Path]:
    candidate_path = candidate_path.absolute()
    candidate_path.parent.mkdir(parents=True, exist_ok=True)
    candidate_path = managed_direct_child(
        candidate_path.parent,
        candidate_path,
        purpose="management evidence candidate",
        require_file=False,
    )
    package = build_management_evidence_package(args).to_dict()
    write_json_atomic(candidate_path, package)
    candidate_sha256 = sha256_file(candidate_path)
    fingerprint_path = evidence_fingerprint_path(candidate_path)
    fingerprint = {
        "schema_version": 2,
        "contract": "ManagementEvidenceCandidateFingerprintV2",
        "candidate_file": candidate_path.name,
        "candidate_sha256": candidate_sha256,
        "candidate_size": candidate_path.stat().st_size,
        "source": package["source"],
        "record_count": len(package["records"]),
        "retention_delete_after": package["retention"]["delete_after"],
        "approval_required": True,
        "automatic_upload": False,
        "automatic_send": False,
    }
    write_json_atomic(fingerprint_path, fingerprint)
    fingerprint_sha256 = sha256_file(fingerprint_path)
    template_path = evidence_approval_template_path(candidate_path)
    write_json_atomic(
        template_path,
        {
            "schema_version": 2,
            "decision": "APPROVE",
            "candidate_sha256": candidate_sha256,
            "fingerprint_sha256": fingerprint_sha256,
            "approved_by": "",
            "approved_at_utc": "",
            "purpose": "",
        },
    )
    return [candidate_path, fingerprint_path, template_path]


def promote_approved_management_evidence(
    args: argparse.Namespace,
    candidate_path: Path,
    approval_path: Path,
) -> list[Path]:
    candidate_path = regular_file_without_reparse_ancestors(
        candidate_path, purpose="management evidence candidate"
    )
    fingerprint_path = regular_file_without_reparse_ancestors(
        evidence_fingerprint_path(candidate_path),
        purpose="management evidence fingerprint",
    )
    approval_path = regular_file_without_reparse_ancestors(
        approval_path, purpose="management evidence approval"
    )
    candidate_sha256 = sha256_file(candidate_path)
    fingerprint_sha256 = sha256_file(fingerprint_path)
    approval_sha256 = sha256_file(approval_path)
    fingerprint = read_json_manifest(fingerprint_path)
    approval = read_json_manifest(approval_path)
    if approval.get("schema_version") != 2 or approval.get("decision") != "APPROVE":
        raise IntegrityError(
            "Approval must use schema_version 2 and the exact decision APPROVE."
        )
    for field, actual in (
        ("candidate_sha256", candidate_sha256),
        ("fingerprint_sha256", fingerprint_sha256),
    ):
        if not secrets.compare_digest(str(approval.get(field) or ""), actual):
            raise IntegrityError(
                f"Approval {field} does not match the reviewed evidence artifact."
            )
    for field in ("approved_by", "approved_at_utc", "purpose"):
        if not str(approval.get(field) or "").strip():
            raise IntegrityError(f"Approval field {field} is required.")
    try:
        approved_at = datetime.fromisoformat(
            str(approval["approved_at_utc"]).replace("Z", "+00:00")
        )
    except ValueError as exc:
        raise IntegrityError("approved_at_utc must be an ISO-8601 timestamp.") from exc
    if approved_at.tzinfo is None:
        raise IntegrityError("approved_at_utc must include a timezone.")
    if not secrets.compare_digest(
        str(fingerprint.get("candidate_sha256") or ""), candidate_sha256
    ):
        raise IntegrityError("The fingerprint does not bind the candidate bytes.")
    if (
        fingerprint.get("schema_version") != 2
        or fingerprint.get("contract")
        != "ManagementEvidenceCandidateFingerprintV2"
    ):
        raise IntegrityError(
            "The fingerprint must use ManagementEvidenceCandidateFingerprintV2."
        )
    package = read_json_manifest(candidate_path)
    package_contract = (package.get("contract"), package.get("schema_version"))
    if (
        package_contract != ("ManagementEvidencePackageV2", 2)
        or package.get("retention", {}).get("days") != 365
        or package.get("distribution", {}).get("automatic_upload") is not False
        or package.get("distribution", {}).get("automatic_send") is not False
    ):
        raise IntegrityError("The candidate does not satisfy the approved evidence contract.")
    current_source, _, _ = verified_evidence_source(args)
    if canonical_json_sha256(package.get("source", {})) != canonical_json_sha256(
        current_source
    ):
        raise IntegrityError(
            "The reviewed evidence candidate is stale relative to the trusted current run."
        )
    archive_dir = Path(args.archive_dir).resolve()
    approved_root = managed_subdirectory(
        archive_dir,
        "approved-management-evidence",
        purpose="approved management evidence",
        create=True,
    )
    destination = managed_direct_child(
        approved_root,
        approved_root / f"approved-{candidate_sha256[:16]}.json",
        purpose="approved management evidence",
        require_file=False,
    )
    content = candidate_path.read_bytes()
    if destination.exists():
        if sha256_file(destination) != candidate_sha256:
            raise IntegrityError(
                "An approved evidence file with the same identifier has different bytes."
            )
    else:
        written_sha256 = verified_write_bytes(content, destination)
        if written_sha256 != candidate_sha256:
            raise IntegrityError("Approved evidence exact-byte verification failed.")
    receipt_path = approved_root / (
        f"approved-{candidate_sha256[:16]}-{approval_sha256[:16]}.receipt.json"
    )
    write_json_atomic(
        receipt_path,
        {
            "schema_version": 2,
            "contract": "ManagementEvidenceApprovalReceiptV2",
            "candidate_sha256": candidate_sha256,
            "fingerprint_sha256": fingerprint_sha256,
            "approval_sha256": approval_sha256,
            "approved_by": approval["approved_by"],
            "approved_at_utc": approval["approved_at_utc"],
            "purpose": approval["purpose"],
            "approved_file": destination.name,
            "retention_delete_after": package["retention"]["delete_after"],
            "automatic_upload": False,
            "automatic_send": False,
        },
    )
    return [destination, receipt_path]


def build_health_check(args: argparse.Namespace) -> dict[str, Any]:
    """Inspect runtime readiness without creating folders, locks, reports, or state."""

    checked_at = datetime.now(timezone.utc).isoformat()
    config_path = Path(args.config).resolve()
    input_dir = Path(args.input_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    archive_dir = Path(args.archive_dir).resolve()
    configured_anchor_dir = getattr(args, "integrity_anchor_dir", None)
    anchor_dir = (
        Path(configured_anchor_dir).expanduser().absolute()
        if configured_anchor_dir
        else default_integrity_anchor_dir()
    )
    checks: list[dict[str, Any]] = []

    def add(
        name: str,
        status: str,
        detail: str,
        **structured_detail: Any,
    ) -> None:
        checks.append(
            {
                "name": name,
                "status": status,
                "detail": safe_message(detail),
                **structured_detail,
            }
        )

    try:
        config = load_config(config_path)
        add(
            "Configuration",
            "Ready",
            f"Valid schema with {len(config['locations'])} configured locations.",
        )
    except Exception as exc:
        config = None
        add("Configuration", "Attention", str(exc))

    python_ready = (3, 10) <= sys.version_info[:2] < (3, 13)
    add(
        "Python",
        "Ready" if python_ready else "Attention",
        f"{sys.version.split()[0]} (supported: 3.10-3.12).",
    )
    for label, path in (
        ("Input folder", input_dir),
        ("Finished reports folder", output_dir),
        ("Archive folder", archive_dir),
    ):
        add(
            label,
            "Ready" if path.is_dir() else "Attention",
            str(path) if path.is_dir() else f"Missing directory: {path}",
        )

    try:
        free = available_disk_bytes(output_dir)
        minimum = 256 * 1024 * 1024
        add(
            "Disk capacity",
            "Ready" if free >= minimum else "Attention",
            f"{free / (1024 ** 3):.2f} GiB free; minimum {minimum / (1024 ** 2):.0f} MiB.",
        )
    except Exception as exc:
        add("Disk capacity", "Attention", str(exc))

    integrity_status = "Attention"
    integrity_detail = "Archive and finished-report folders must exist."
    latest_manifest: Path | None = None
    verified_manifest_payload: dict[str, Any] = {}
    if archive_dir.is_dir() and output_dir.is_dir() and config is not None:
        try:
            latest_manifest = latest_integrity_manifest_path(archive_dir)
            if latest_manifest is None:
                raise IntegrityError(
                    "No integrity baseline exists; initialize it before a weekly run."
                )
            if not integrity_anchor_exists(archive_dir, anchor_dir):
                raise IntegrityError(
                    "The machine-local trusted-head anchor is missing."
                )
            anchored_manifest, anchored_sha256 = verify_integrity_anchor(
                archive_dir, anchor_dir
            )
            if anchored_manifest != latest_manifest.resolve():
                raise IntegrityError(
                    "The latest manifest does not match the machine-local trusted head."
                )
            verified_manifest_payload, verified_sha256 = verify_integrity_state(
                archive_dir,
                output_dir,
                latest_manifest,
                # The trusted anchor already pins this exact manifest head. A
                # read-only health check must therefore recognize a supported
                # one-way workbook-layout upgrade just as ordinary preflight does.
                allow_legacy_master_upgrade=True,
            )
            if not secrets.compare_digest(verified_sha256, anchored_sha256):
                raise IntegrityError(
                    "The verified manifest hash differs from the trusted anchor."
                )
            integrity_status = "Ready"
            integrity_detail = (
                f"Manifest chain and managed outputs verified at {latest_manifest.name}."
            )
        except Exception as exc:
            integrity_detail = str(exc)
    add("Integrity", integrity_status, integrity_detail)

    recorded_master_digest = verified_manifest_payload.get(
        "master_generated_content_sha256"
    )
    workbook_status = (
        "Ready"
        if integrity_status == "Ready"
        and isinstance(recorded_master_digest, str)
        and re.fullmatch(r"[0-9a-fA-F]{64}", recorded_master_digest)
        else "Attention"
    )
    workbook_detail = (
        (
            "Protected substantive master content matches the manifest; workbook metadata "
            "differs from the recorded generated version and will be refreshed on the next "
            "successful run."
            if verified_manifest_payload.get("_master_metadata_drift")
            else "Protected substantive master content matches the manifest-recorded digest."
        )
        if workbook_status == "Ready"
        else (
            "The verified manifest does not record a protected management workbook."
            if integrity_status == "Ready"
            else integrity_detail
        )
    )
    add(
        "Workbook",
        workbook_status,
        workbook_detail,
        metadata_drift=bool(verified_manifest_payload.get("_master_metadata_drift")),
    )

    published_inventory = verified_manifest_payload.get(
        "published_output_inventory"
    )
    publication_status = (
        "Ready"
        if workbook_status == "Ready"
        and isinstance(published_inventory, list)
        and bool(published_inventory)
        else "Attention"
    )
    publication_detail = (
        "Exact managed per-location workbook bytes and protected generated master "
        f"content are verified by {latest_manifest.name}."
        if publication_status == "Ready" and latest_manifest is not None
        else (
            "The verified manifest does not record managed per-location published workbooks."
            if integrity_status == "Ready"
            else integrity_detail
        )
    )
    add(
        "Local publication",
        publication_status,
        publication_detail,
    )

    overall = (
        "Ready"
        if checks and all(item["status"] == "Ready" for item in checks)
        else "Attention"
    )
    return {
        "schema_version": 1,
        "checked_at_utc": checked_at,
        "overall": overall,
        "readiness": {
            "release": (
                RunReadiness.READY.value
                if re.fullmatch(
                    r"[0-9a-fA-F]{40}",
                    os.environ.get("RED_ONION_VERIFIED_RELEASE_COMMIT", ""),
                )
                else RunReadiness.NOT_EVALUATED.value
            ),
            "integrity": (
                RunReadiness.READY.value
                if integrity_status == "Ready"
                else RunReadiness.ATTENTION.value
            ),
            "workbook": (
                RunReadiness.READY.value
                if workbook_status == "Ready"
                else RunReadiness.ATTENTION.value
            ),
            "distribution": (
                RunReadiness.READY.value
                if publication_status == "Ready"
                else RunReadiness.ATTENTION.value
            ),
            "recovery": RunReadiness.EXTERNAL_CHECK_REQUIRED.value,
        },
        "latest_manifest": latest_manifest.name if latest_manifest else None,
        "checks": checks,
        "note": (
            "Independent Google Drive backup freshness and restore-test evidence require "
            "a separate external verification; this local command does not access Drive."
        ),
    }


def print_health_check(payload: dict[str, Any]) -> None:
    print(f"Red Onion health: {payload['overall']}")
    for check in payload["checks"]:
        print(f"  [{check['status']}] {check['name']}: {check['detail']}")
    print(
        "  "
        f"[{payload['readiness']['recovery']}] Independent recovery: {payload['note']}"
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build Red Onion weekly metric workbooks.")
    parser.add_argument(
        "--input-dir",
        default=str(DEFAULT_INPUT_DIR),
        help=f"Folder containing active raw daily {DAILY_REPORT_FORMAT_LABEL} reports.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Folder for generated workbooks.",
    )
    parser.add_argument(
        "--archive-dir",
        default=str(DEFAULT_ARCHIVE_DIR),
        help="Folder for archived source files.",
    )
    parser.add_argument(
        "--integrity-anchor-dir",
        default=str(default_integrity_anchor_dir()),
        help=(
            "Machine-local trusted-head folder outside the Dropbox/operator workspace. "
            "Keep it restricted to the automation runner account."
        ),
    )
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH), help="Path to config JSON.")
    parser.add_argument("--week-start", help="Optional public snapshot start date, YYYY-MM-DD.")
    parser.add_argument("--week-end", help="Optional public snapshot end date, YYYY-MM-DD.")
    parser.add_argument(
        "--migrate-history-from",
        action="append",
        default=[],
        metavar="FOLDER",
        help=(
            "Copy validated daily reports from a legacy folder into the canonical processed "
            "archive. Repeat this option for additional source folders."
        ),
    )
    parser.add_argument(
        "--migrate-history-only",
        action="store_true",
        help="Perform the copy-only history migration without generating weekly workbooks.",
    )
    parser.add_argument(
        "--rebuild-from-history",
        action="store_true",
        help=(
            "Maintainer-only: rebuild and publish the latest complete week from "
            "manifest-pinned history without reading, moving, or deleting active "
            "drop-folder files. May be combined with --migrate-history-from."
        ),
    )
    parser.add_argument(
        "--initialize-integrity-baseline",
        action="store_true",
        help=(
            "Create or verify the clearly labeled raw/derivative integrity baseline without "
            "generating workbooks or moving active reports."
        ),
    )
    parser.add_argument(
        "--rebind-restored-integrity-anchor",
        metavar="SOURCE_ANCHOR_JSON",
        help=(
            "Verify a backed-up trusted anchor against a restored manifest chain "
            "and managed outputs, then bind that exact head to this machine/path. "
            "This is a recovery-only operation and never processes weekly reports."
        ),
    )
    parser.add_argument(
        "--validate-config",
        action="store_true",
        help="Validate configuration and exit without creating or changing runtime state.",
    )
    parser.add_argument(
        "--health-check",
        action="store_true",
        help="Run a read-only local readiness check and exit.",
    )
    parser.add_argument(
        "--health-check-json",
        action="store_true",
        help="Run the read-only readiness check and print versioned JSON.",
    )
    parser.add_argument(
        "--export-management-evidence",
        metavar="CANDIDATE_JSON",
        help=(
            "Stage a local ManagementEvidencePackageV2 candidate and exact fingerprint. "
            "With --approval-file, promote the already-reviewed exact bytes into the "
            "local approved-evidence archive."
        ),
    )
    parser.add_argument(
        "--approval-file",
        help=(
            "Explicit schema-v2 approval bound to the staged candidate and fingerprint. "
            "This never uploads or sends the evidence."
        ),
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.rebuild_from_history:
        incompatible_rebuild = (
            args.validate_config
            or args.health_check
            or args.health_check_json
            or args.initialize_integrity_baseline
            or args.migrate_history_only
            or bool(args.rebind_restored_integrity_anchor)
            or bool(args.export_management_evidence)
            or bool(args.approval_file)
            or bool(args.week_start)
            or bool(args.week_end)
        )
        if incompatible_rebuild:
            raise SystemExit(
                "ERROR: --rebuild-from-history may be combined only with "
                "--migrate-history-from and normal path/config options."
            )
    if args.rebind_restored_integrity_anchor:
        incompatible = (
            args.validate_config
            or args.health_check
            or args.health_check_json
            or args.initialize_integrity_baseline
            or args.migrate_history_only
            or args.rebuild_from_history
            or bool(args.migrate_history_from)
            or bool(args.export_management_evidence)
            or bool(args.approval_file)
        )
        if incompatible:
            raise SystemExit(
                "ERROR: --rebind-restored-integrity-anchor must be run as a "
                "standalone recovery operation."
            )
        try:
            rebound, receipt = rebind_restored_integrity_anchor(
                Path(args.archive_dir),
                Path(args.output_dir),
                Path(args.rebind_restored_integrity_anchor),
                Path(args.integrity_anchor_dir),
            )
        except Exception as exc:
            raise SystemExit(f"ERROR: {exc}") from exc
        print("Restored integrity head verified and rebound:")
        print(f"  Anchor: {rebound}")
        print(f"  Audit receipt: {receipt}")
        return
    if args.validate_config:
        try:
            config = load_config(Path(args.config).resolve())
        except Exception as exc:
            raise SystemExit(f"ERROR: {exc}") from exc
        print(
            "Configuration valid: "
            f"{len(config['locations'])} location(s), "
            f"{len(config)} supported top-level fields."
        )
        return
    if args.health_check or args.health_check_json:
        payload = build_health_check(args)
        if args.health_check_json:
            print(json.dumps(payload, sort_keys=True, indent=2))
        else:
            print_health_check(payload)
        if payload["overall"] != "Ready":
            raise SystemExit(2)
        return
    if args.approval_file and not args.export_management_evidence:
        raise SystemExit(
            "ERROR: --approval-file requires --export-management-evidence CANDIDATE_JSON."
        )
    if args.export_management_evidence:
        try:
            candidate_path = Path(args.export_management_evidence).absolute()
            if args.approval_file:
                paths = promote_approved_management_evidence(
                    args, candidate_path, Path(args.approval_file).absolute()
                )
                print("Approved evidence retained locally (no upload or send):")
            else:
                paths = stage_management_evidence(args, candidate_path)
                print("Evidence candidate staged; review and complete the approval template:")
            for path in paths:
                print(f"  {path}")
        except Exception as exc:
            raise SystemExit(f"ERROR: {exc}") from exc
        return
    try:
        generated = run(args)
    except Exception as exc:
        raise SystemExit(f"ERROR: {exc}") from exc
    if args.initialize_integrity_baseline:
        print("Integrity baseline verified:")
        for path in generated:
            print(f"  {path}")
    elif args.migrate_history_only:
        if generated:
            print("History copied to the canonical archive:")
            for path in generated:
                print(f"  {path}")
        else:
            print("History migration complete: the canonical archive is already up to date.")
    elif args.rebuild_from_history:
        print("History rebuilt and published:")
        for path in generated:
            print(f"  {path}")
    else:
        print("Generated:")
        for path in generated:
            print(f"  {path}")


if __name__ == "__main__":
    main()
